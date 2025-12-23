"""
Этап 5: Валидация результатов заполнения
Проверка корректности заполнения шаблонов
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell import MergedCell


class FillValidator:
    """Класс для валидации результатов заполнения шаблонов."""
    
    def __init__(self, filled_template_path: Path,
                 original_template_path: Optional[Path] = None,
                 semantic_mapping_path: Optional[Path] = None):
        """
        Инициализация валидатора.
        
        Args:
            filled_template_path: Путь к заполненному шаблону
            original_template_path: Путь к оригинальному шаблону (опционально)
            semantic_mapping_path: Путь к семантическому маппингу (опционально)
        """
        self.filled_template_path = Path(filled_template_path)
        self.original_template_path = Path(original_template_path) if original_template_path else None
        self.semantic_mapping_path = Path(semantic_mapping_path) if semantic_mapping_path else None
        
        self.filled_workbook = None
        self.original_workbook = None
        self.semantic_mapping = {}
        
        self.validation_report = {
            "validation_date": datetime.now().isoformat(),
            "filled_template": str(self.filled_template_path),
            "status": "unknown",
            "score": 0.0,
            "validations": {},
            "issues": [],
            "warnings": [],
            "statistics": {}
        }
    
    def load_data(self) -> None:
        """Загрузка данных для валидации."""
        if not self.filled_template_path.exists():
            raise FileNotFoundError(f"Заполненный шаблон не найден: {self.filled_template_path}")
        
        self.filled_workbook = load_workbook(self.filled_template_path, data_only=True)
        
        if self.original_template_path and self.original_template_path.exists():
            self.original_workbook = load_workbook(self.original_template_path, data_only=True)
        
        if self.semantic_mapping_path and self.semantic_mapping_path.exists():
            self.semantic_mapping = json.loads(
                self.semantic_mapping_path.read_text(encoding="utf-8")
            )
    
    def validate(self) -> Dict[str, Any]:
        """
        Валидация заполнения.
        
        Returns:
            Словарь с отчетом о валидации
        """
        # Валидация структуры
        structural_validation = self._validate_structure()
        self.validation_report["validations"]["structural"] = structural_validation
        
        # Валидация заполнения
        filling_validation = self._validate_filling()
        self.validation_report["validations"]["filling"] = filling_validation
        
        # Валидация форматов
        format_validation = self._validate_formats()
        self.validation_report["validations"]["formats"] = format_validation
        
        # Валидация единиц измерения
        units_validation = self._validate_units()
        self.validation_report["validations"]["units"] = units_validation
        
        # Валидация по семантическому маппингу
        if self.semantic_mapping:
            semantic_validation = self._validate_semantic_mapping()
            self.validation_report["validations"]["semantic"] = semantic_validation
        
        # Расчет общего статуса
        self._calculate_overall_status()
        
        return self.validation_report
    
    def _validate_structure(self) -> Dict[str, Any]:
        """Валидация структуры шаблона."""
        validation = {
            "status": "passed",
            "issues": [],
            "warnings": [],
            "details": {}
        }
        
        if not self.filled_workbook:
            validation["status"] = "failed"
            validation["issues"].append("Не удалось загрузить заполненный шаблон")
            return validation
        
        # Проверка наличия листов
        if self.original_workbook:
            original_sheets = set(self.original_workbook.sheetnames)
            filled_sheets = set(self.filled_workbook.sheetnames)
            
            missing_sheets = original_sheets - filled_sheets
            extra_sheets = filled_sheets - original_sheets
            
            if missing_sheets:
                validation["warnings"].append(
                    f"Отсутствуют листы: {', '.join(missing_sheets)}"
                )
            
            if extra_sheets:
                validation["warnings"].append(
                    f"Дополнительные листы: {', '.join(extra_sheets)}"
                )
            
            validation["details"]["sheet_comparison"] = {
                "original_sheets_count": len(original_sheets),
                "filled_sheets_count": len(filled_sheets),
                "missing_sheets": list(missing_sheets),
                "extra_sheets": list(extra_sheets)
            }
        else:
            validation["details"]["filled_sheets_count"] = len(self.filled_workbook.sheetnames)
        
        return validation
    
    def _validate_filling(self) -> Dict[str, Any]:
        """Валидация заполнения ячеек."""
        validation = {
            "status": "passed",
            "filled_cells": 0,
            "empty_cells": 0,
            "issues": [],
            "warnings": [],
            "details": {}
        }
        
        if not self.filled_workbook:
            return validation
        
        total_cells = 0
        filled_cells = 0
        empty_cells = 0
        
        for sheet_name in self.filled_workbook.sheetnames:
            ws = self.filled_workbook[sheet_name]
            sheet_filled = 0
            sheet_empty = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    total_cells += 1
                    if cell.value is not None and str(cell.value).strip() != "":
                        filled_cells += 1
                        sheet_filled += 1
                    else:
                        empty_cells += 1
                        sheet_empty += 1
            
            validation["details"][sheet_name] = {
                "filled_cells": sheet_filled,
                "empty_cells": sheet_empty
            }
        
        validation["filled_cells"] = filled_cells
        validation["empty_cells"] = empty_cells
        
        # Оценка заполнения
        if total_cells > 0:
            fill_rate = (filled_cells / total_cells) * 100
            
            if fill_rate < 30:
                validation["status"] = "poor"
                validation["warnings"].append(f"Низкий процент заполнения: {fill_rate:.1f}%")
            elif fill_rate < 60:
                validation["status"] = "partial"
                validation["warnings"].append(f"Умеренный процент заполнения: {fill_rate:.1f}%")
            else:
                validation["status"] = "good"
            
            validation["details"]["fill_rate"] = fill_rate
            validation["details"]["total_cells"] = total_cells
        
        return validation
    
    def _validate_formats(self) -> Dict[str, Any]:
        """Валидация форматов данных."""
        validation = {
            "status": "passed",
            "issues": [],
            "warnings": [],
            "details": {}
        }
        
        if not self.filled_workbook:
            return validation
        
        format_errors = 0
        format_warnings = 0
        
        for sheet_name in self.filled_workbook.sheetnames:
            ws = self.filled_workbook[sheet_name]
            sheet_errors = []
            sheet_warnings = []
            
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.value is None:
                        continue
                    
                    # Проверка формата чисел
                    if cell.data_type == "n":
                        value = cell.value
                        if isinstance(value, (int, float)):
                            # Проверка на разумные значения
                            if abs(value) > 1e10:
                                format_warnings += 1
                                sheet_warnings.append(
                                    f"{cell.coordinate}: Подозрительно большое значение: {value}"
                                )
                    # Проверка формата текста
                    elif cell.data_type == "s":
                        value = str(cell.value)
                        # Проверка на пустые строки
                        if value.strip() == "":
                            format_warnings += 1
                            sheet_warnings.append(
                                f"{cell.coordinate}: Пустая строка"
                            )
            
            if sheet_errors:
                validation["details"][sheet_name] = {
                    "errors": sheet_errors,
                    "warnings": sheet_warnings
                }
        
        if format_errors > 0:
            validation["status"] = "failed"
        
        if format_warnings > 0:
            validation["warnings"].append(f"Найдено {format_warnings} предупреждений о форматах")
        
        validation["details"]["format_errors"] = format_errors
        validation["details"]["format_warnings"] = format_warnings
        
        return validation
    
    def _validate_units(self) -> Dict[str, Any]:
        """Валидация единиц измерения."""
        validation = {
            "status": "passed",
            "issues": [],
            "warnings": [],
            "details": {}
        }
        
        # Здесь можно добавить проверку единиц измерения
        # Например, проверка соответствия единиц в заголовках и значениях
        
        return validation
    
    def _validate_semantic_mapping(self) -> Dict[str, Any]:
        """Валидация по семантическому маппингу."""
        validation = {
            "status": "passed",
            "mapped_cells": 0,
            "unmapped_cells": 0,
            "issues": [],
            "warnings": [],
            "details": {}
        }
        
        if not self.semantic_mapping:
            return validation
        
        mappings = self.semantic_mapping.get("mappings", [])
        mapped_count = 0
        unmapped_count = 0
        
        for mapping in mappings:
            cell_address = mapping.get("cell_address")
            sheet_name = mapping.get("sheet")
            confidence = mapping.get("confidence", 0.0)
            
            if not cell_address or not sheet_name:
                continue
            
            if sheet_name not in self.filled_workbook.sheetnames:
                unmapped_count += 1
                continue
            
            try:
                ws = self.filled_workbook[sheet_name]
                cell = ws[cell_address]
                
                # Проверка на объединенную ячейку
                if isinstance(cell, MergedCell):
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            cell = ws[merged_range.min_col][merged_range.min_row]
                            break
                
                if cell.value is not None:
                    mapped_count += 1
                else:
                    unmapped_count += 1
                    if confidence > 0.5:
                        validation["warnings"].append(
                            f"Ячейка {cell_address} в листе {sheet_name} не заполнена "
                            f"(уверенность маппинга: {confidence:.2f})"
                        )
            except Exception as e:
                unmapped_count += 1
                validation["issues"].append(
                    f"Ошибка проверки {cell_address} в листе {sheet_name}: {str(e)}"
                )
        
        validation["mapped_cells"] = mapped_count
        validation["unmapped_cells"] = unmapped_count
        
        # Оценка статуса
        total_checked = mapped_count + unmapped_count
        if total_checked > 0:
            mapping_rate = (mapped_count / total_checked) * 100
            
            if mapping_rate < 50:
                validation["status"] = "poor"
            elif mapping_rate < 80:
                validation["status"] = "partial"
            else:
                validation["status"] = "good"
            
            validation["details"]["mapping_rate"] = mapping_rate
        
        return validation
    
    def _calculate_overall_status(self) -> None:
        """Расчет общего статуса валидации."""
        validations = self.validation_report["validations"]
        
        # Подсчет очков
        scores = []
        
        # Структурная валидация
        structural = validations.get("structural", {})
        if structural.get("status") == "passed":
            scores.append(1.0)
        else:
            scores.append(0.5)
        
        # Валидация заполнения
        filling = validations.get("filling", {})
        fill_rate = filling.get("details", {}).get("fill_rate", 0.0)
        scores.append(fill_rate / 100.0)
        
        # Валидация форматов
        formats = validations.get("formats", {})
        if formats.get("status") == "passed":
            scores.append(1.0)
        else:
            scores.append(0.5)
        
        # Семантическая валидация
        if "semantic" in validations:
            semantic = validations["semantic"]
            mapping_rate = semantic.get("details", {}).get("mapping_rate", 0.0)
            scores.append(mapping_rate / 100.0)
        
        # Среднее значение
        if scores:
            overall_score = sum(scores) / len(scores)
            self.validation_report["score"] = overall_score
            
            # Определение статуса
            if overall_score >= 0.8:
                status = "excellent"
            elif overall_score >= 0.6:
                status = "good"
            elif overall_score >= 0.4:
                status = "acceptable"
            else:
                status = "poor"
            
            self.validation_report["status"] = status
        
        # Сбор всех проблем и предупреждений
        all_issues = []
        all_warnings = []
        
        for validation_name, validation_result in validations.items():
            all_issues.extend(validation_result.get("issues", []))
            all_warnings.extend(validation_result.get("warnings", []))
        
        self.validation_report["issues"] = all_issues
        self.validation_report["warnings"] = all_warnings
        
        # Статистика
        self.validation_report["statistics"] = {
            "total_issues": len(all_issues),
            "total_warnings": len(all_warnings),
            "validations_count": len(validations)
        }
    
    def save(self, output_path: Path) -> None:
        """Сохранение отчета о валидации."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.validation_report, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def validate_filling(filled_template_path: Path,
                    output_path: Path,
                    original_template_path: Optional[Path] = None,
                    semantic_mapping_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Валидация заполнения шаблона.
    
    Args:
        filled_template_path: Путь к заполненному шаблону
        output_path: Путь для сохранения отчета
        original_template_path: Путь к оригинальному шаблону (опционально)
        semantic_mapping_path: Путь к семантическому маппингу (опционально)
    
    Returns:
        Словарь с отчетом о валидации
    """
    validator = FillValidator(filled_template_path, original_template_path, semantic_mapping_path)
    validator.load_data()
    report = validator.validate()
    validator.save(output_path)
    
    return report


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Валидация заполнения шаблонов")
    parser.add_argument("--filled", required=True, help="Путь к заполненному шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения отчета")
    parser.add_argument("--original", help="Путь к оригинальному шаблону")
    parser.add_argument("--semantic", help="Путь к семантическому маппингу")
    
    args = parser.parse_args()
    
    filled_path = Path(args.filled)
    output_path = Path(args.output)
    original_path = Path(args.original) if args.original else None
    semantic_path = Path(args.semantic) if args.semantic else None
    
    if not filled_path.exists():
        raise FileNotFoundError(f"Файл не найден: {filled_path}")
    
    print("=" * 80)
    print("ВАЛИДАЦИЯ ЗАПОЛНЕНИЯ ШАБЛОНА")
    print("=" * 80)
    
    report = validate_filling(filled_path, output_path, original_path, semantic_path)
    
    print("\n✅ Валидация завершена")
    print(f"  Статус: {report['status']}")
    print(f"  Оценка: {report['score']:.2%}")
    print(f"  Проблем: {len(report['issues'])}")
    print(f"  Предупреждений: {len(report['warnings'])}")
    print(f"\n📁 Отчет сохранен: {output_path}")

