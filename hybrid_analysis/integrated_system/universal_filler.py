"""
Этап 5: Универсальная система заполнения шаблонов
Интеграция всех предыдущих анализов для интеллектуального заполнения
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell import MergedCell


class UniversalFiller:
    """
    Универсальная система заполнения шаблонов.
    Интегрирует технический, семантический, ML и сравнительный анализы.
    """
    
    def __init__(self,
                 template_path: Path,
                 data_path: Path,
                 structural_analysis_path: Optional[Path] = None,
                 semantic_mapping_path: Optional[Path] = None,
                 ml_patterns_path: Optional[Path] = None,
                 adapter_path: Optional[Path] = None):
        """
        Инициализация универсального заполнителя.
        
        Args:
            template_path: Путь к шаблону Excel
            data_path: Путь к JSON файлу с данными
            structural_analysis_path: Путь к техническому анализу (опционально)
            semantic_mapping_path: Путь к семантическому маппингу (опционально)
            ml_patterns_path: Путь к ML паттернам (опционально)
            adapter_path: Путь к адаптерам (опционально)
        """
        self.template_path = Path(template_path)
        self.data_path = Path(data_path)
        self.structural_analysis_path = Path(structural_analysis_path) if structural_analysis_path else None
        self.semantic_mapping_path = Path(semantic_mapping_path) if semantic_mapping_path else None
        self.ml_patterns_path = Path(ml_patterns_path) if ml_patterns_path else None
        self.adapter_path = Path(adapter_path) if adapter_path else None
        
        # Загруженные данные
        self.workbook = None
        self.data = {}
        self.structural_analysis = {}
        self.semantic_mapping = {}
        self.ml_patterns = {}
        self.adapters = {}
        
        # Результаты заполнения
        self.fill_results = {
            "filled_cells": 0,
            "skipped_cells": 0,
            "errors": [],
            "warnings": [],
            "filled_addresses": []
        }
    
    def load_all_data(self) -> None:
        """Загрузка всех необходимых данных."""
        # Загрузка шаблона
        if not self.template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
        
        self.workbook = load_workbook(self.template_path, data_only=False)
        
        # Загрузка данных
        if not self.data_path.exists():
            raise FileNotFoundError(f"Файл данных не найден: {self.data_path}")
        
        self.data = json.loads(self.data_path.read_text(encoding="utf-8"))
        
        # Загрузка технического анализа
        if self.structural_analysis_path and self.structural_analysis_path.exists():
            self.structural_analysis = json.loads(
                self.structural_analysis_path.read_text(encoding="utf-8")
            )
        
        # Загрузка семантического маппинга
        if self.semantic_mapping_path and self.semantic_mapping_path.exists():
            self.semantic_mapping = json.loads(
                self.semantic_mapping_path.read_text(encoding="utf-8")
            )
        
        # Загрузка ML паттернов
        if self.ml_patterns_path and self.ml_patterns_path.exists():
            self.ml_patterns = json.loads(
                self.ml_patterns_path.read_text(encoding="utf-8")
            )
        
        # Загрузка адаптеров
        if self.adapter_path and self.adapter_path.exists():
            self.adapters = json.loads(
                self.adapter_path.read_text(encoding="utf-8")
            )
    
    def fill(self) -> Dict[str, Any]:
        """
        Заполнение шаблона данными.
        
        Returns:
            Словарь с результатами заполнения
        """
        print(f"Заполнение шаблона: {self.template_path.name}")
        print(f"Данные: {self.data_path.name}")
        
        # Нормализация данных
        normalized_data = self._normalize_data(self.data)
        
        # Заполнение на основе семантического маппинга (приоритет)
        if self.semantic_mapping:
            print("  Использование семантического маппинга...")
            self._fill_by_semantic_mapping(normalized_data)
        
        # Заполнение на основе структурного анализа
        elif self.structural_analysis:
            print("  Использование структурного анализа...")
            self._fill_by_structural_analysis(normalized_data)
        
        # Заполнение на основе адаптеров
        if self.adapters:
            print("  Применение адаптеров...")
            self._apply_adapters(normalized_data)
        
        # Применение ML паттернов для форматирования
        if self.ml_patterns:
            print("  Применение ML паттернов форматирования...")
            self._apply_ml_patterns()
        
        return self.fill_results
    
    def _normalize_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Нормализация структуры данных."""
        normalized = {"resources": {}}
        
        # Если данные в формате file-based, преобразуем в resource-based
        if isinstance(data, dict):
            # Проверяем наличие resources напрямую
            if "resources" in data:
                normalized["resources"] = data["resources"]
            else:
                # Преобразуем из file-based формата
                for file_data in data.values():
                    if isinstance(file_data, dict) and "resources" in file_data:
                        for resource_type, resource_data in file_data["resources"].items():
                            if resource_type not in normalized["resources"]:
                                normalized["resources"][resource_type] = {}
                            if isinstance(resource_data, dict):
                                for quarter, quarter_data in resource_data.items():
                                    if quarter not in normalized["resources"][resource_type]:
                                        normalized["resources"][resource_type][quarter] = quarter_data
        
        return normalized
    
    def _fill_by_semantic_mapping(self, normalized_data: Dict[str, Any]) -> None:
        """Заполнение на основе семантического маппинга."""
        mappings = self.semantic_mapping.get("mappings", [])
        
        for mapping in mappings:
            cell_address = mapping.get("cell_address")
            sheet_name = mapping.get("sheet")
            data_path = mapping.get("data_path")
            confidence = mapping.get("confidence", 0.0)
            
            if not cell_address or not sheet_name or not data_path:
                continue
            
            # Пропускаем ячейки с очень низкой уверенностью (можно настроить)
            # Но для тестирования заполняем все доступные маппинги
            if confidence < 0.1:
                self.fill_results["skipped_cells"] += 1
                continue
            
            # Получение значения из данных
            value = self._get_value_from_path(normalized_data, data_path)
            
            if value is None:
                self.fill_results["warnings"].append(
                    f"Данные не найдены для {cell_address} ({data_path})"
                )
                continue
            
            # Заполнение ячейки
            try:
                if sheet_name in self.workbook.sheetnames:
                    ws = self.workbook[sheet_name]
                    cell = ws[cell_address]
                    
                    # Проверка на объединенную ячейку
                    if isinstance(cell, MergedCell):
                        # Получаем главную ячейку
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                cell = ws[merged_range.min_col][merged_range.min_row]
                                break
                    
                    # Форматирование значения
                    formatted_value = self._format_value(value, cell, mapping)
                    cell.value = formatted_value
                    
                    self.fill_results["filled_cells"] += 1
                    self.fill_results["filled_addresses"].append({
                        "sheet": sheet_name,
                        "address": cell_address,
                        "value": formatted_value,
                        "data_path": data_path
                    })
            except Exception as e:
                self.fill_results["errors"].append(
                    f"Ошибка заполнения {cell_address} в листе {sheet_name}: {str(e)}"
                )
    
    def _fill_by_structural_analysis(self, normalized_data: Dict[str, Any]) -> None:
        """Заполнение на основе структурного анализа."""
        sheets = self.structural_analysis.get("sheets", {})
        
        for sheet_name, sheet_info in sheets.items():
            if sheet_name not in self.workbook.sheetnames:
                continue
            
            ws = self.workbook[sheet_name]
            cells = sheet_info.get("cells", [])
            
            for row_data in cells:
                for cell_info in row_data.get("cells", []):
                    cell_address = cell_info.get("address")
                    value = cell_info.get("value")
                    
                    # Простая эвристика: заполняем пустые ячейки
                    if not value or value.strip() == "":
                        # Попытка определить тип данных по позиции
                        semantic_type = self._guess_semantic_type(cell_address, sheet_name)
                        if semantic_type:
                            data_value = self._get_value_by_semantic_type(
                                normalized_data, semantic_type
                            )
                            if data_value is not None:
                                try:
                                    ws[cell_address].value = data_value
                                    self.fill_results["filled_cells"] += 1
                                except Exception as e:
                                    self.fill_results["errors"].append(
                                        f"Ошибка заполнения {cell_address}: {str(e)}"
                                    )
    
    def _apply_adapters(self, normalized_data: Dict[str, Any]) -> None:
        """Применение адаптеров для заполнения."""
        adapters = self.adapters.get("adapters", {})
        
        for sheet_name, adapter in adapters.items():
            if adapter.get("type") != "map":
                continue
            
            if sheet_name not in self.workbook.sheetnames:
                continue
            
            ws = self.workbook[sheet_name]
            cell_mappings = adapter.get("cell_mappings", [])
            
            for mapping in cell_mappings:
                source_cell = mapping.get("source_cell")
                target_cell = mapping.get("target_cell")
                
                if not source_cell or not target_cell:
                    continue
                
                # Получение значения из source
                try:
                    source_ws = ws  # Может быть другой лист
                    source_value = source_ws[source_cell].value
                    
                    if source_value is not None:
                        # Применение трансформаций
                        transformations = adapter.get("transformations", [])
                        transformed_value = self._apply_transformations(
                            source_value, transformations, target_cell
                        )
                        
                        ws[target_cell].value = transformed_value
                        self.fill_results["filled_cells"] += 1
                except Exception as e:
                    self.fill_results["warnings"].append(
                        f"Не удалось применить адаптер для {target_cell}: {str(e)}"
                    )
    
    def _apply_ml_patterns(self) -> None:
        """Применение ML паттернов для форматирования."""
        patterns = self.ml_patterns.get("patterns", {})
        
        for pattern_name, pattern_info in patterns.items():
            # Применение паттернов форматирования
            # Здесь можно добавить логику применения ML-предсказаний
            pass
    
    def _get_value_from_path(self, data: Dict[str, Any], path: str) -> Optional[Any]:
        """Получение значения из данных по пути."""
        try:
            parts = path.split(".")
            current = data
            
            for part in parts:
                if isinstance(current, dict):
                    current = current.get(part)
                elif isinstance(current, list):
                    try:
                        current = current[int(part)]
                    except (ValueError, IndexError):
                        return None
                else:
                    return None
                
                if current is None:
                    return None
            
            return current
        except Exception:
            return None
    
    def _get_value_by_semantic_type(self, data: Dict[str, Any], semantic_type: str) -> Optional[Any]:
        """Получение значения по семантическому типу."""
        # Простая эвристика
        if semantic_type == "electricity_active":
            electricity = data.get("resources", {}).get("electricity", {})
            if electricity:
                # Берем первый доступный квартал
                quarters = list(electricity.keys())
                if quarters:
                    quarter_data = electricity[quarters[0]]
                    return quarter_data.get("quarter_totals", {}).get("active_kwh")
        
        elif semantic_type == "electricity_reactive":
            electricity = data.get("resources", {}).get("electricity", {})
            if electricity:
                quarters = list(electricity.keys())
                if quarters:
                    quarter_data = electricity[quarters[0]]
                    return quarter_data.get("quarter_totals", {}).get("reactive_kvarh")
        
        elif semantic_type == "gas_volume":
            gas = data.get("resources", {}).get("gas", {})
            if gas:
                quarters = list(gas.keys())
                if quarters:
                    quarter_data = gas[quarters[0]]
                    return quarter_data.get("quarter_totals", {}).get("volume_m3")
        
        elif semantic_type == "water_volume":
            water = data.get("resources", {}).get("water", {})
            if water:
                quarters = list(water.keys())
                if quarters:
                    quarter_data = water[quarters[0]]
                    return quarter_data.get("quarter_totals", {}).get("volume_m3")
        
        return None
    
    def _guess_semantic_type(self, cell_address: str, sheet_name: str) -> Optional[str]:
        """Попытка определить семантический тип ячейки."""
        # Простая эвристика на основе имени листа и позиции
        sheet_lower = sheet_name.lower()
        
        if "электри" in sheet_lower or "electricity" in sheet_lower:
            return "electricity_active"
        elif "газ" in sheet_lower or "gas" in sheet_lower:
            return "gas_volume"
        elif "вод" in sheet_lower or "water" in sheet_lower:
            return "water_volume"
        
        return None
    
    def _format_value(self, value: Any, cell, mapping: Dict[str, Any]) -> Any:
        """Форматирование значения для ячейки."""
        # Применение ML паттернов если есть
        semantic_type = mapping.get("semantic_type")
        
        if isinstance(value, (int, float)):
            # Округление
            return round(float(value), 4) if value else None
        
        return value
    
    def _apply_transformations(self, value: Any, transformations: List[Dict[str, Any]],
                              target_cell: str) -> Any:
        """Применение трансформаций к значению."""
        result = value
        
        for transformation in transformations:
            if transformation.get("target_cell") != target_cell:
                continue
            
            trans_type = transformation.get("transformation_type")
            
            if trans_type == "number_to_string":
                result = str(result) if result is not None else None
            elif trans_type == "string_to_number":
                try:
                    result = float(result) if result is not None else None
                except (TypeError, ValueError):
                    pass
            elif trans_type == "formula":
                # Формулы обрабатываются отдельно
                pass
        
        return result
    
    def save(self, output_path: Path) -> None:
        """Сохранение заполненного шаблона."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        
        # Сохранение отчета о заполнении
        report_path = output_path.with_suffix(".fill_report.json")
        report = {
            "template_path": str(self.template_path),
            "data_path": str(self.data_path),
            "output_path": str(output_path),
            "fill_date": datetime.now().isoformat(),
            "results": self.fill_results
        }
        
        report_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    
    def get_fill_statistics(self) -> Dict[str, Any]:
        """Получение статистики заполнения."""
        total_attempted = self.fill_results["filled_cells"] + self.fill_results["skipped_cells"]
        success_rate = (self.fill_results["filled_cells"] / total_attempted * 100) if total_attempted > 0 else 0
        
        return {
            "filled_cells": self.fill_results["filled_cells"],
            "skipped_cells": self.fill_results["skipped_cells"],
            "total_attempted": total_attempted,
            "success_rate": round(success_rate, 2),
            "errors_count": len(self.fill_results["errors"]),
            "warnings_count": len(self.fill_results["warnings"])
        }


def fill_template(template_path: Path,
                 data_path: Path,
                 output_path: Path,
                 structural_analysis_path: Optional[Path] = None,
                 semantic_mapping_path: Optional[Path] = None,
                 ml_patterns_path: Optional[Path] = None,
                 adapter_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Заполнение шаблона данными.
    
    Args:
        template_path: Путь к шаблону
        data_path: Путь к данным
        output_path: Путь для сохранения заполненного шаблона
        structural_analysis_path: Путь к техническому анализу (опционально)
        semantic_mapping_path: Путь к семантическому маппингу (опционально)
        ml_patterns_path: Путь к ML паттернам (опционально)
        adapter_path: Путь к адаптерам (опционально)
    
    Returns:
        Словарь с результатами заполнения
    """
    filler = UniversalFiller(
        template_path,
        data_path,
        structural_analysis_path,
        semantic_mapping_path,
        ml_patterns_path,
        adapter_path
    )
    
    filler.load_all_data()
    results = filler.fill()
    filler.save(output_path)
    
    return results


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Универсальное заполнение шаблонов")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--data", required=True, help="Путь к данным JSON")
    parser.add_argument("--output", required=True, help="Путь для сохранения результата")
    parser.add_argument("--structural", help="Путь к техническому анализу")
    parser.add_argument("--semantic", help="Путь к семантическому маппингу")
    parser.add_argument("--ml", help="Путь к ML паттернам")
    parser.add_argument("--adapter", help="Путь к адаптерам")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    data_path = Path(args.data)
    output_path = Path(args.output)
    structural_path = Path(args.structural) if args.structural else None
    semantic_path = Path(args.semantic) if args.semantic else None
    ml_path = Path(args.ml) if args.ml else None
    adapter_path = Path(args.adapter) if args.adapter else None
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    if not data_path.exists():
        raise FileNotFoundError(f"Файл данных не найден: {data_path}")
    
    print("=" * 80)
    print("УНИВЕРСАЛЬНОЕ ЗАПОЛНЕНИЕ ШАБЛОНА")
    print("=" * 80)
    
    results = fill_template(
        template_path,
        data_path,
        output_path,
        structural_path,
        semantic_path,
        ml_path,
        adapter_path
    )
    
    print("\n✅ Результаты:")
    print(f"  Заполнено ячеек: {results['filled_cells']}")
    print(f"  Пропущено ячеек: {results['skipped_cells']}")
    print(f"  Ошибок: {len(results['errors'])}")
    print(f"  Предупреждений: {len(results['warnings'])}")
    print(f"\n📁 Результат сохранен: {output_path}")

