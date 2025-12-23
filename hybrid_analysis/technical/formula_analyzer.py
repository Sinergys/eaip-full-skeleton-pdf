"""
Этап 1.2: Анализ формул и ссылок
Обнаружение всех формул Excel и их зависимостей
"""

import json
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import re


class FormulaAnalyzer:
    """Анализатор формул и ссылок в Excel шаблонах."""
    
    def __init__(self, template_path: Path):
        """
        Инициализация анализатора.
        
        Args:
            template_path: Путь к Excel шаблону
        """
        self.template_path = template_path
        self.workbook = None
        self.formulas = {}
        self.dependencies = {}
    
    def analyze(self) -> Dict[str, Any]:
        """
        Анализ формул и ссылок.
        
        Returns:
            Словарь с информацией о формулах
        """
        self.workbook = load_workbook(self.template_path, data_only=False)
        
        result = {
            "template_path": str(self.template_path),
            "template_name": self.template_path.stem,
            "sheets": {}
        }
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            sheet_formulas = self._analyze_sheet(ws)
            result["sheets"][sheet_name] = sheet_formulas
        
        # Общая статистика
        total_formulas = sum(len(s["formulas"]) for s in result["sheets"].values())
        result["total_formulas"] = total_formulas
        
        return result
    
    def _analyze_sheet(self, ws) -> Dict[str, Any]:
        """Анализ формул в одном листе."""
        formulas = []
        dependencies = []
        named_ranges = []
        
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                
                if cell.data_type == 'f' and cell.value:
                    formula_str = str(cell.value)
                    
                    # Извлечение зависимостей (ссылок на другие ячейки)
                    cell_refs = self._extract_cell_references(formula_str)
                    range_refs = self._extract_range_references(formula_str)
                    
                    formula_info = {
                        "address": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "formula": formula_str,
                        "dependencies": {
                            "cells": cell_refs,
                            "ranges": range_refs
                        }
                    }
                    
                    formulas.append(formula_info)
                    
                    # Добавление зависимостей
                    for ref in cell_refs:
                        dependencies.append({
                            "from": cell.coordinate,
                            "to": ref,
                            "type": "cell"
                        })
                    
                    for ref_range in range_refs:
                        dependencies.append({
                            "from": cell.coordinate,
                            "to": ref_range,
                            "type": "range"
                        })
        
        # Поиск именованных диапазонов
        if hasattr(self.workbook, 'defined_names'):
            for name, definition in self.workbook.defined_names.items():
                named_ranges.append({
                    "name": name,
                    "definition": str(definition)
                })
        
        return {
            "formulas": formulas,
            "dependencies": dependencies,
            "named_ranges": named_ranges,
            "formulas_count": len(formulas)
        }
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """Извлечение ссылок на отдельные ячейки из формулы."""
        # Паттерн для ссылок на ячейки: A1, B2, Sheet1!A1 и т.д.
        pattern = r'([A-Z]+[0-9]+)'
        matches = re.findall(pattern, formula)
        return list(set(matches))
    
    def _extract_range_references(self, formula: str) -> List[str]:
        """Извлечение ссылок на диапазоны из формулы."""
        # Паттерн для диапазонов: A1:B2, Sheet1!A1:B2 и т.д.
        pattern = r'([A-Z]+[0-9]+:[A-Z]+[0-9]+)'
        matches = re.findall(pattern, formula)
        return list(set(matches))
    
    def save(self, output_path: Path) -> None:
        """
        Сохранение результатов анализа.
        
        Args:
            output_path: Путь для сохранения JSON
        """
        result = self.analyze()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def analyze_formulas(template_path: Path, output_path: Path) -> Dict[str, Any]:
    """
    Анализ формул и сохранение результатов.
    
    Args:
        template_path: Путь к шаблону
        output_path: Путь для сохранения результатов
    
    Returns:
        Словарь с информацией о формулах
    """
    analyzer = FormulaAnalyzer(template_path)
    result = analyzer.analyze()
    analyzer.save(output_path)
    return result


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Анализ формул Excel")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    output_path = Path(args.output)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    print(f"Анализ формул: {template_path}")
    result = analyze_formulas(template_path, output_path)
    
    print(f"\n✅ Результаты сохранены в: {output_path}")
    print("📊 Статистика:")
    print(f"  Всего формул: {result['total_formulas']}")
    for sheet_name, sheet_info in result["sheets"].items():
        print(f"  {sheet_name}: {sheet_info['formulas_count']} формул")

