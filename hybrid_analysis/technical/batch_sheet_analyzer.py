"""
Скрипт для массового анализа всех листов шаблона
Проводит полный структурный и семантический анализ каждого листа
"""

import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from hybrid_analysis.technical.structural_parser import StructuralParser
from hybrid_analysis.technical.ai_table_structure_analyzer import AITableStructureAnalyzer
from openpyxl import load_workbook


def analyze_all_sheets(
    template_path: Path,
    output_dir: Path,
    max_rows: Optional[int] = None,
    sheets_to_analyze: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    Анализ всех листов шаблона.
    
    Args:
        template_path: Путь к шаблону
        output_dir: Директория для сохранения результатов
        max_rows: Максимальное количество строк для анализа
        sheets_to_analyze: Список листов для анализа (None = все)
        
    Returns:
        Словарь с результатами анализа всех листов
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Получаем список листов
    workbook = load_workbook(template_path, read_only=True)
    all_sheets = workbook.sheetnames
    workbook.close()
    
    sheets_to_process = sheets_to_analyze if sheets_to_analyze else all_sheets
    
    print("=" * 80)
    print("🔬 ПОЛНЫЙ АНАЛИЗ ВСЕХ ЛИСТОВ ШАБЛОНА")
    print("=" * 80)
    print(f"Шаблон: {template_path}")
    print(f"Выходная директория: {output_dir}")
    print(f"Листов для анализа: {len(sheets_to_process)}")
    print()
    
    results = {
        "template_path": str(template_path),
        "template_name": template_path.stem,
        "total_sheets": len(all_sheets),
        "analyzed_sheets": len(sheets_to_process),
        "sheets_analysis": {}
    }
    
    # Анализ каждого листа
    for idx, sheet_name in enumerate(sheets_to_process, 1):
        print(f"[{idx}/{len(sheets_to_process)}] Анализ листа: '{sheet_name}'")
        
        try:
            # 1. Структурный анализ
            print("  📋 Структурный анализ...")
            structural_parser = StructuralParser(template_path)
            structural_result = structural_parser.parse(max_rows=max_rows, sheet_name=sheet_name)
            structural_parser.close()
            
            sheet_structure = structural_result["sheets"].get(sheet_name, {})
            
            # 2. Анализ структуры таблиц
            print("  🔍 Анализ структуры таблиц...")
            table_analyzer = AITableStructureAnalyzer(template_path, use_ai=False)
            table_structure = table_analyzer.analyze_sheet(sheet_name, max_rows=max_rows)
            table_analyzer.close()
            
            # Объединяем результаты
            sheet_analysis = {
                "sheet_name": sheet_name,
                "structural_analysis": {
                    "max_row": sheet_structure.get("max_row"),
                    "max_column": sheet_structure.get("max_column"),
                    "merged_ranges": sheet_structure.get("merged_ranges", []),
                    "data_region": sheet_structure.get("data_region"),
                    "cells_count": sheet_structure.get("cells_count", 0)
                },
                "table_structure_analysis": table_structure,
                "summary": {
                    "total_tables": table_structure["summary"]["total_tables"],
                    "total_rows": table_structure["summary"]["total_rows"],
                    "total_columns": table_structure["summary"]["total_columns"],
                    "sheet_purpose": table_structure["semantic_analysis"]["sheet_purpose"],
                    "resource_types": table_structure["semantic_analysis"]["resource_types"],
                    "time_periods": table_structure["semantic_analysis"]["time_periods"]
                }
            }
            
            results["sheets_analysis"][sheet_name] = sheet_analysis
            
            # Сохраняем индивидуальный файл для каждого листа
            sheet_output_file = output_dir / f"sheet_{sheet_name.replace(' ', '_').replace('/', '_')}_analysis.json"
            sheet_output_file.write_text(
                json.dumps(sheet_analysis, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )
            
            print("  ✅ Завершено")
            print(f"     - Таблиц найдено: {sheet_analysis['summary']['total_tables']}")
            print(f"     - Назначение: {sheet_analysis['summary']['sheet_purpose']}")
            print(f"     - Ресурсы: {', '.join(sheet_analysis['summary']['resource_types']) or 'нет'}")
            print(f"     - Сохранено: {sheet_output_file.name}")
            print()
            
        except Exception as e:
            print(f"  ❌ Ошибка при анализе листа '{sheet_name}': {e}")
            results["sheets_analysis"][sheet_name] = {
                "error": str(e),
                "sheet_name": sheet_name
            }
            print()
    
    # Сохраняем сводный файл
    summary_file = output_dir / "all_sheets_analysis_summary.json"
    summary_file.write_text(
        json.dumps(results, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    
    print("=" * 80)
    print("✅ АНАЛИЗ ВСЕХ ЛИСТОВ ЗАВЕРШЕН")
    print("=" * 80)
    print("\n📊 Итоговая статистика:")
    print(f"  Всего листов в шаблоне: {results['total_sheets']}")
    print(f"  Проанализировано: {results['analyzed_sheets']}")
    print(f"  Успешно: {len([s for s in results['sheets_analysis'].values() if 'error' not in s])}")
    print(f"  С ошибками: {len([s for s in results['sheets_analysis'].values() if 'error' in s])}")
    print("\n💾 Результаты сохранены в:")
    print(f"  - Сводный файл: {summary_file}")
    print(f"  - Индивидуальные файлы: {output_dir}")
    
    return results


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Массовый анализ всех листов шаблона")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--output-dir", required=True, help="Директория для сохранения результатов")
    parser.add_argument("--max-rows", type=int, help="Максимальное количество строк для анализа")
    parser.add_argument("--sheets", nargs="+", help="Конкретные листы для анализа (если не указано - все)")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    analyze_all_sheets(
        template_path=template_path,
        output_dir=output_dir,
        max_rows=args.max_rows,
        sheets_to_analyze=args.sheets
    )

