"""
Этап 1.3: Определение типов данных
Классификация содержимого ячеек
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import re
from datetime import datetime


class DataTypeClassifier:
    """Классификатор типов данных в ячейках Excel."""
    
    # Категории типов данных
    CATEGORIES = {
        "text": "Текст",
        "number": "Число",
        "date": "Дата",
        "formula": "Формула",
        "placeholder": "Placeholder",
        "header": "Заголовок",
        "label": "Метка",
        "empty": "Пусто"
    }
    
    def __init__(self, template_path: Path):
        """
        Инициализация классификатора.
        
        Args:
            template_path: Путь к Excel шаблону
        """
        self.template_path = template_path
        self.workbook = None
        self.classification = {}
    
    def classify(self, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """
        Классификация типов данных.
        
        Args:
            max_rows: Максимальное количество строк для анализа
        
        Returns:
            Словарь с классификацией
        """
        self.workbook = load_workbook(self.template_path, data_only=False)
        
        self.classification = {
            "template_path": str(self.template_path),
            "template_name": self.template_path.stem,
            "sheets": {},
            "summary": {
                "categories": {cat: 0 for cat in self.CATEGORIES.keys()}
            }
        }
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            sheet_classification = self._classify_sheet(ws, max_rows)
            self.classification["sheets"][sheet_name] = sheet_classification
            
            # Обновление сводки
            for cat, count in sheet_classification["summary"].items():
                self.classification["summary"]["categories"][cat] += count
        
        return self.classification
    
    def _classify_sheet(self, ws, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """Классификация одного листа."""
        cells_by_type = {cat: [] for cat in self.CATEGORIES.keys()}
        
        rows_to_parse = min(max_rows or ws.max_row, ws.max_row)
        
        for row_idx in range(1, rows_to_parse + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if isinstance(cell, MergedCell):
                    continue
                
                cell_type = self._classify_cell(cell)
                cells_by_type[cell_type].append({
                    "address": cell.coordinate,
                    "row": row_idx,
                    "column": col_idx,
                    "value": str(cell.value)[:200] if cell.value is not None else None
                })
        
        summary = {cat: len(cells) for cat, cells in cells_by_type.items()}
        
        return {
            "cells_by_type": cells_by_type,
            "summary": summary
        }
    
    def _classify_cell(self, cell) -> str:
        """
        Классификация одной ячейки.
        
        Args:
            cell: Ячейка Excel
        
        Returns:
            Категория типа данных
        """
        if cell.value is None:
            return "empty"
        
        if cell.data_type == 'f':
            return "formula"
        
        if isinstance(cell.value, (int, float)):
            return "number"
        
        if isinstance(cell.value, bool):
            return "number"  # Булевы значения как числа
        
        if isinstance(cell.value, datetime):
            return "date"
        
        cell_str = str(cell.value)
        
        # Проверка на placeholder
        if self._is_placeholder(cell_str):
            return "placeholder"
        
        # Проверка на заголовок (обычно в первых строках)
        if cell.row <= 5 and self._is_header(cell_str):
            return "header"
        
        # Проверка на метку
        if self._is_label(cell_str):
            return "label"
        
        # Проверка на дату в текстовом формате
        if self._is_date_like(cell_str):
            return "date"
        
        return "text"
    
    def _is_placeholder(self, value: str) -> bool:
        """Проверка на placeholder."""
        # {{key}} формат
        if re.search(r'\{\{[^}]+\}\}', value):
            return True
        
        # Подчеркивания с контекстом
        if re.search(r'_+', value):
            context = value.lower()
            if any(keyword in context for keyword in ['год', 'year', 'квартал', 'quarter', 'месяц', 'month']):
                return True
        
        # 20___ формат
        if re.search(r'20_+', value):
            return True
        
        return False
    
    def _is_header(self, value: str) -> bool:
        """Проверка на заголовок."""
        # Заголовки обычно короткие, начинаются с заглавной буквы
        if len(value) < 100 and value and value[0].isupper():
            # Проверка на ключевые слова заголовков
            header_keywords = ['таблица', 'структура', 'баланс', 'динамика', 'энергия', 'ресурс']
            if any(keyword in value.lower() for keyword in header_keywords):
                return True
        
        return False
    
    def _is_label(self, value: str) -> bool:
        """Проверка на метку."""
        # Метки обычно короткие, описывают категории
        if len(value) < 50:
            label_keywords = ['потребление', 'нужды', 'производство', 'технологические', 'собственные']
            if any(keyword in value.lower() for keyword in label_keywords):
                return True
        
        return False
    
    def _is_date_like(self, value: str) -> bool:
        """Проверка на дату в текстовом формате."""
        # Паттерны дат: DD.MM.YYYY, DD/MM/YYYY и т.д.
        date_patterns = [
            r'\d{1,2}[./]\d{1,2}[./]\d{2,4}',
            r'\d{4}-\d{2}-\d{2}'
        ]
        
        for pattern in date_patterns:
            if re.search(pattern, value):
                return True
        
        return False
    
    def save(self, output_path: Path) -> None:
        """
        Сохранение результатов классификации.
        
        Args:
            output_path: Путь для сохранения JSON
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.classification, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def classify_data_types(template_path: Path, output_path: Path, max_rows: Optional[int] = None) -> Dict[str, Any]:
    """
    Классификация типов данных и сохранение результатов.
    
    Args:
        template_path: Путь к шаблону
        output_path: Путь для сохранения результатов
        max_rows: Максимальное количество строк для анализа
    
    Returns:
        Словарь с классификацией
    """
    classifier = DataTypeClassifier(template_path)
    result = classifier.classify(max_rows=max_rows)
    classifier.save(output_path)
    return result


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Классификация типов данных")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    parser.add_argument("--max-rows", type=int, help="Максимальное количество строк")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    output_path = Path(args.output)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    print(f"Классификация типов данных: {template_path}")
    result = classify_data_types(template_path, output_path, max_rows=args.max_rows)
    
    print(f"\n✅ Результаты сохранены в: {output_path}")
    print("📊 Сводка по категориям:")
    for cat, count in result["summary"]["categories"].items():
        if count > 0:
            print(f"  {DataTypeClassifier.CATEGORIES[cat]}: {count}")

