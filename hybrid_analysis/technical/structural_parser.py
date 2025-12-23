"""
Этап 1.1: Структурный парсинг шаблонов
Извлечение координат всех ячеек из Excel шаблонов
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter


class StructuralParser:
    """Парсер для извлечения структурной информации из Excel шаблонов."""
    
    def __init__(self, template_path: Path):
        """
        Инициализация парсера.
        
        Args:
            template_path: Путь к Excel шаблону
        """
        self.template_path = template_path
        self.workbook = None
        self.structure = {}
    
    def parse(self, max_rows: Optional[int] = None, max_cols: Optional[int] = None, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Парсинг структуры шаблона.
        
        Args:
            max_rows: Максимальное количество строк для анализа (None = все)
            max_cols: Максимальное количество столбцов для анализа (None = все)
            sheet_name: Имя листа для анализа (None = все листы)
        
        Returns:
            Словарь со структурной информацией
        """
        self.workbook = load_workbook(self.template_path, data_only=False)
        
        self.structure = {
            "template_path": str(self.template_path),
            "template_name": self.template_path.stem,
            "total_sheets": len(self.workbook.sheetnames),
            "sheets": {}
        }
        
        sheets_to_parse = [sheet_name] if sheet_name else self.workbook.sheetnames
        
        for sheet_name_item in sheets_to_parse:
            if sheet_name_item not in self.workbook.sheetnames:
                raise ValueError(f"Лист '{sheet_name_item}' не найден в шаблоне")
            ws = self.workbook[sheet_name_item]
            sheet_info = self._parse_sheet(ws, max_rows, max_cols)
            self.structure["sheets"][sheet_name_item] = sheet_info
        
        return self.structure
    
    def _parse_sheet(self, ws, max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> Dict[str, Any]:
        """Парсинг одного листа."""
        sheet_info = {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "cells": [],
            "data_region": None
        }
        
        # Определение области данных
        data_start_row = None
        data_end_row = None
        data_start_col = None
        data_end_col = None
        
        # Парсинг ячеек
        rows_to_parse = min(max_rows or ws.max_row, ws.max_row)
        cols_to_parse = min(max_cols or ws.max_column, ws.max_column)
        
        for row_idx in range(1, rows_to_parse + 1):
            row_data = []
            for col_idx in range(1, cols_to_parse + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if isinstance(cell, MergedCell):
                    continue
                
                cell_info = {
                    "address": cell.coordinate,
                    "row": row_idx,
                    "column": col_idx,
                    "column_letter": get_column_letter(col_idx),
                    "value": str(cell.value) if cell.value is not None else None,
                    "data_type": cell.data_type,
                    "is_merged": False,
                    "merged_range": None
                }
                
                # Проверка на объединенные ячейки
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        cell_info["is_merged"] = True
                        cell_info["merged_range"] = str(merged_range)
                        break
                
                if cell.value is not None:
                    # Определение области данных
                    if data_start_row is None:
                        data_start_row = row_idx
                        data_start_col = col_idx
                    data_end_row = row_idx
                    data_end_col = max(data_end_col or col_idx, col_idx)
                    
                    row_data.append(cell_info)
            
            if row_data:
                sheet_info["cells"].append({
                    "row": row_idx,
                    "cells": row_data
                })
        
        # Область данных
        if data_start_row:
            sheet_info["data_region"] = {
                "start_row": data_start_row,
                "end_row": data_end_row,
                "start_column": data_start_col,
                "end_column": data_end_col,
                "start_address": f"{get_column_letter(data_start_col)}{data_start_row}",
                "end_address": f"{get_column_letter(data_end_col)}{data_end_row}"
            }
        
        sheet_info["cells_count"] = sum(len(row["cells"]) for row in sheet_info["cells"])
        
        return sheet_info
    
    def save(self, output_path: Path) -> None:
        """
        Сохранение результатов парсинга.
        
        Args:
            output_path: Путь для сохранения JSON
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.structure, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    
    def close(self) -> None:
        """Закрытие рабочей книги."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None


def parse_template(template_path: Path, output_path: Path, max_rows: Optional[int] = None) -> Dict[str, Any]:
    """
    Парсинг шаблона и сохранение результатов.
    
    Args:
        template_path: Путь к шаблону
        output_path: Путь для сохранения результатов
        max_rows: Максимальное количество строк для анализа
    
    Returns:
        Словарь со структурной информацией
    """
    parser = StructuralParser(template_path)
    structure = parser.parse(max_rows=max_rows)
    parser.save(output_path)
    return structure


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Структурный парсинг Excel шаблонов")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    parser.add_argument("--sheet", help="Имя листа для анализа (если не указано - все листы)")
    parser.add_argument("--max-rows", type=int, help="Максимальное количество строк")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    output_path = Path(args.output)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    print(f"Парсинг шаблона: {template_path}")
    if args.sheet:
        print(f"Анализ листа: {args.sheet}")
        parser_obj = StructuralParser(template_path)
        structure = parser_obj.parse(max_rows=args.max_rows, sheet_name=args.sheet)
        parser_obj.save(output_path)
    else:
        structure = parse_template(template_path, output_path, max_rows=args.max_rows)
    
    print(f"\n✅ Результаты сохранены в: {output_path}")
    print("📊 Статистика:")
    print(f"  Листов: {structure['total_sheets']}")
    for sheet_name, sheet_info in structure["sheets"].items():
        print(f"  {sheet_name}: {sheet_info['cells_count']} ячеек")

