"""
Инструмент для ИИ-анализа структуры таблиц в листах Excel
Глубокий анализ заголовков, подзаголовков, данных, связей
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class AITableStructureAnalyzer:
    """Анализатор структуры таблиц с использованием ИИ для понимания семантики."""
    
    def __init__(self, template_path: Path, use_ai: bool = True):
        """
        Инициализация анализатора.
        
        Args:
            template_path: Путь к Excel шаблону
            use_ai: Использовать ли ИИ для анализа (если False - только структурный анализ)
        """
        self.template_path = template_path
        self.use_ai = use_ai
        self.workbook = None
        
    def analyze_sheet(self, sheet_name: str, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """
        Анализ структуры таблиц в одном листе.
        
        Args:
            sheet_name: Имя листа для анализа
            max_rows: Максимальное количество строк для анализа
            
        Returns:
            Словарь с анализом структуры таблиц
        """
        if self.workbook is None:
            self.workbook = load_workbook(self.template_path, data_only=False, read_only=True)
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
        
        ws = self.workbook[sheet_name]
        
        # Структурный анализ
        structure = self._structural_analysis(ws, max_rows)
        
        # Определение таблиц в листе
        tables = self._detect_tables(ws, structure, max_rows)
        
        # Анализ каждой таблицы
        analyzed_tables = []
        for table in tables:
            table_analysis = self._analyze_table_structure(ws, table, structure)
            analyzed_tables.append(table_analysis)
        
        # Семантический анализ листа (без ИИ - правила)
        semantic_analysis = self._semantic_analysis(ws, structure, analyzed_tables, max_rows)
        
        result = {
            "sheet_name": sheet_name,
            "structure": structure,
            "tables": analyzed_tables,
            "semantic_analysis": semantic_analysis,
            "summary": {
                "total_tables": len(analyzed_tables),
                "total_rows": structure.get("max_row", 0),
                "total_columns": structure.get("max_column", 0),
                "data_regions": len([t for t in analyzed_tables if t.get("data_rows", [])])
            }
        }
        
        return result
    
    def _structural_analysis(self, ws, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """Структурный анализ листа."""
        rows_to_analyze = min(max_rows or ws.max_row, ws.max_row) if max_rows else ws.max_row
        
        # Определение области данных
        data_start_row = None
        data_end_row = None
        data_start_col = None
        data_end_col = None
        
        # Сбор информации о ячейках
        cell_info = {}
        for row_idx in range(1, min(rows_to_analyze + 1, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    coord = cell.coordinate
                    cell_info[coord] = {
                        "address": coord,
                        "row": row_idx,
                        "column": col_idx,
                        "value": str(cell.value),
                        "data_type": cell.data_type,
                        "has_formula": cell.data_type == 'f'
                    }
                    
                    if data_start_row is None:
                        data_start_row = row_idx
                        data_start_col = col_idx
                    data_end_row = row_idx
                    data_end_col = max(data_end_col or col_idx, col_idx)
        
        # Анализ заголовков (первые несколько строк обычно содержат заголовки)
        headers = self._extract_headers(ws, min(10, rows_to_analyze))
        
        return {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "data_region": {
                "start_row": data_start_row,
                "end_row": data_end_row,
                "start_column": data_start_col,
                "end_column": data_end_col
            } if data_start_row else None,
            "headers": headers,
            "cells_info": cell_info
        }
    
    def _extract_headers(self, ws, max_header_rows: int = 10) -> List[Dict[str, Any]]:
        """Извлечение заголовков из первых строк."""
        headers = []
        
        for row_idx in range(1, min(max_header_rows + 1, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 50)):  # Ограничиваем столбцы
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_values.append({
                        "column": col_idx,
                        "column_letter": get_column_letter(col_idx),
                        "value": str(cell.value).strip(),
                        "is_merged": False
                    })
            
            if row_values:
                headers.append({
                    "row": row_idx,
                    "cells": row_values
                })
        
        return headers
    
    def _detect_tables(self, ws, structure: Dict, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
        """Определение таблиц в листе."""
        tables = []
        data_region = structure.get("data_region")
        
        if not data_region:
            return tables
        
        start_row = data_region["start_row"]
        end_row = min(data_region["end_row"], max_rows or data_region["end_row"])
        start_col = data_region["start_column"]
        end_col = data_region["end_column"]
        
        # Ищем начало таблиц (обычно это строки с заголовками)
        current_table_start = None
        
        for row_idx in range(start_row, end_row + 1):
            # Проверяем, является ли строка заголовком таблицы
            is_header_row = self._is_header_row(ws, row_idx, start_col, end_col)
            
            if is_header_row and current_table_start is None:
                # Начало новой таблицы
                current_table_start = row_idx
            elif current_table_start is not None:
                # Проверяем, не закончилась ли таблица (пустая строка или новый заголовок)
                is_empty = self._is_empty_row(ws, row_idx, start_col, end_col)
                is_new_header = is_header_row and (row_idx - current_table_start) > 2
                
                if is_empty or is_new_header:
                    # Завершаем текущую таблицу
                    tables.append({
                        "start_row": current_table_start,
                        "end_row": row_idx - 1,
                        "start_column": start_col,
                        "end_column": end_col
                    })
                    current_table_start = row_idx if is_new_header else None
        
        # Добавляем последнюю таблицу
        if current_table_start is not None:
            tables.append({
                "start_row": current_table_start,
                "end_row": end_row,
                "start_column": start_col,
                "end_column": end_col
            })
        
        # Если не нашли таблицы по заголовкам, создаем одну большую
        if not tables:
            tables.append({
                "start_row": start_row,
                "end_row": end_row,
                "start_column": start_col,
                "end_column": end_col
            })
        
        return tables
    
    def _is_header_row(self, ws, row_idx: int, start_col: int, end_col: int) -> bool:
        """Проверка, является ли строка заголовком таблицы."""
        header_indicators = ["№", "наименование", "показатель", "год", "квартал", "месяц", 
                           "электроэнергия", "газ", "вода", "мазут", "уголь", "итого", "всего",
                           "пункты", "учет", "учёт", "вид", "место", "коэффициент", "дата",
                           "мероприятия", "мероприятие", "название", "описание", "эффект"]
        
        text_values = []
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                text_values.append(str(cell.value).lower().strip())
        
        # Проверяем наличие индикаторов заголовков
        text_joined = " ".join(text_values)
        has_header_indicators = any(indicator in text_joined for indicator in header_indicators)
        
        # Проверяем, что в строке есть несколько непустых ячеек (обычно заголовки имеют несколько столбцов)
        non_empty_count = len([v for v in text_values if v])
        
        # Дополнительная проверка: заголовки обычно не содержат только числа
        numeric_only = all(
            any(c.isdigit() for c in v.replace(".", "").replace(",", "").replace(" ", "")) 
            for v in text_values if v
        ) and len(text_values) > 0
        
        return has_header_indicators and non_empty_count >= 2 and not numeric_only
    
    def _is_empty_row(self, ws, row_idx: int, start_col: int, end_col: int) -> bool:
        """Проверка, является ли строка пустой."""
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                return False
        return True
    
    def _analyze_table_structure(self, ws, table: Dict, structure: Dict) -> Dict[str, Any]:
        """Детальный анализ структуры одной таблицы."""
        start_row = table["start_row"]
        end_row = table["end_row"]
        start_col = table["start_column"]
        end_col = table["end_column"]
        
        # Извлечение заголовков таблицы
        header_rows = []
        data_rows = []
        
        # Ищем заголовки (первые 1-3 строки таблицы)
        header_end_row = min(start_row + 3, end_row)
        for row_idx in range(start_row, header_end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_data.append({
                        "column": col_idx,
                        "column_letter": get_column_letter(col_idx),
                        "value": str(cell.value).strip(),
                        "data_type": cell.data_type,
                        "has_formula": cell.data_type == 'f'
                    })
            if row_data:
                header_rows.append({
                    "row": row_idx,
                    "cells": row_data
                })
        
        # Извлечение данных
        data_start = header_end_row + 1 if header_rows else start_row
        for row_idx in range(data_start, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_data.append({
                        "column": col_idx,
                        "column_letter": get_column_letter(col_idx),
                        "value": str(cell.value).strip(),
                        "data_type": cell.data_type,
                        "has_formula": cell.data_type == 'f'
                    })
            if row_data:
                data_rows.append({
                    "row": row_idx,
                    "cells": row_data
                })
        
        # Определение типов столбцов
        column_types = self._detect_column_types(header_rows, data_rows)
        
        return {
            "table_range": {
                "start_row": start_row,
                "end_row": end_row,
                "start_column": start_col,
                "end_column": end_col,
                "start_address": f"{get_column_letter(start_col)}{start_row}",
                "end_address": f"{get_column_letter(end_col)}{end_row}"
            },
            "header_rows": header_rows,
            "data_rows": data_rows,
            "column_types": column_types,
            "row_count": len(data_rows),
            "column_count": end_col - start_col + 1
        }
    
    def _detect_column_types(self, header_rows: List[Dict], data_rows: List[Dict]) -> List[Dict[str, Any]]:
        """Определение типов столбцов на основе заголовков и данных."""
        if not header_rows or not data_rows:
            return []
        
        # Берем первую строку заголовков
        first_header = header_rows[0]["cells"] if header_rows else []
        column_types = []
        
        for col_info in first_header:
            col_idx = col_info["column"]
            header_text = col_info["value"].lower()
            
            # Определяем тип столбца
            col_type = "text"
            semantic_type = "unknown"
            
            # Проверяем по заголовку
            if any(word in header_text for word in ["№", "номер", "no"]):
                col_type = "number"
                semantic_type = "index"
            elif any(word in header_text for word in ["год", "year"]):
                col_type = "number"
                semantic_type = "year"
            elif any(word in header_text for word in ["квартал", "quarter"]):
                col_type = "text"
                semantic_type = "quarter"
            elif any(word in header_text for word in ["месяц", "month"]):
                col_type = "text"
                semantic_type = "month"
            elif any(word in header_text for word in ["электроэнергия", "электричество", "квт", "квар"]):
                col_type = "number"
                semantic_type = "electricity"
            elif any(word in header_text for word in ["газ", "м³", "м3"]):
                col_type = "number"
                semantic_type = "gas"
            elif any(word in header_text for word in ["вода", "water"]):
                col_type = "number"
                semantic_type = "water"
            elif any(word in header_text for word in ["мазут", "уголь", "fuel"]):
                col_type = "number"
                semantic_type = "fuel"
            elif any(word in header_text for word in ["стоимость", "цена", "сум", "cost"]):
                col_type = "number"
                semantic_type = "cost"
            elif any(word in header_text for word in ["итого", "всего", "total", "sum"]):
                col_type = "number"
                semantic_type = "total"
            else:
                # Анализируем данные в столбце
                sample_values = [row["cells"] for row in data_rows[:5]]
                numeric_count = 0
                for row_cells in sample_values:
                    for cell in row_cells:
                        if cell["column"] == col_idx:
                            try:
                                float(cell["value"].replace(",", ".").replace(" ", ""))
                                numeric_count += 1
                            except:
                                pass
                
                if numeric_count > len(sample_values) * 0.7:
                    col_type = "number"
            
            column_types.append({
                "column": col_idx,
                "column_letter": col_info["column_letter"],
                "header": col_info["value"],
                "type": col_type,
                "semantic_type": semantic_type
            })
        
        return column_types
    
    def _semantic_analysis(self, ws, structure: Dict, tables: List[Dict], max_rows: Optional[int] = None) -> Dict[str, Any]:
        """Семантический анализ листа (без ИИ, правила)."""
        semantic_info = {
            "sheet_purpose": "unknown",
            "resource_types": [],
            "time_periods": [],
            "data_categories": [],
            "key_indicators": []
        }
        
        # Анализируем заголовки для определения назначения листа
        all_text = []
        for header_row in structure.get("headers", []):
            for cell in header_row.get("cells", []):
                text = cell["value"].lower()
                all_text.append(text)
        
        text_joined = " ".join(all_text)
        
        # Определяем назначение листа (приоритет по специфичности)
        # Сначала проверяем специфичные термины
        if any(word in text_joined for word in ["dastur", "дастур", "hujjat", "хужат", "qarori", "карори", "qonun", "қонун"]):
            semantic_info["sheet_purpose"] = "regulatory_documentation"
        elif any(word in text_joined for word in ["узел учета", "узел учёта", "учет", "учёт", "metering", "счетчик", "счётчик"]):
            semantic_info["sheet_purpose"] = "metering_nodes"
        elif any(word in text_joined for word in ["мероприятия", "мероприятие", "measures", "events", "мероприят"]):
            semantic_info["sheet_purpose"] = "measures"
        elif any(word in text_joined for word in ["баланс", "balance"]):
            semantic_info["sheet_purpose"] = "balance"
        elif any(word in text_joined for word in ["динамика", "dynamic", "динамика сред", "динамика ср"]):
            semantic_info["sheet_purpose"] = "dynamics"
        elif any(word in text_joined for word in ["мазут", "уголь", "fuel", "coal"]):
            semantic_info["sheet_purpose"] = "fuel_consumption"
        elif any(word in text_joined for word in ["расход на единиц", "расход  на ед.п", "consumption per unit", "расход на ед"]):
            semantic_info["sheet_purpose"] = "consumption_per_unit"
        elif any(word in text_joined for word in ["структура", "structure", "распределение"]):
            semantic_info["sheet_purpose"] = "structure"
        elif any(word in text_joined for word in ["сводн", "summary", "итог", "сводные"]):
            semantic_info["sheet_purpose"] = "summary"
        
        # Определяем типы энергоресурсов
        if any(word in text_joined for word in ["электроэнергия", "электричество", "квт"]):
            semantic_info["resource_types"].append("electricity")
        if any(word in text_joined for word in ["газ", "gas"]):
            semantic_info["resource_types"].append("gas")
        if any(word in text_joined for word in ["вода", "water"]):
            semantic_info["resource_types"].append("water")
        if any(word in text_joined for word in ["мазут", "fuel"]):
            semantic_info["resource_types"].append("fuel")
        if any(word in text_joined for word in ["уголь", "coal"]):
            semantic_info["resource_types"].append("coal")
        
        # Определяем временные периоды
        if any(word in text_joined for word in ["год", "year"]):
            semantic_info["time_periods"].append("year")
        if any(word in text_joined for word in ["квартал", "quarter"]):
            semantic_info["time_periods"].append("quarter")
        if any(word in text_joined for word in ["месяц", "month"]):
            semantic_info["time_periods"].append("month")
        
        # Определяем категории данных
        if any(word in text_joined for word in ["технологич", "technological"]):
            semantic_info["data_categories"].append("technological")
        if any(word in text_joined for word in ["собственн", "own"]):
            semantic_info["data_categories"].append("own_needs")
        if any(word in text_joined for word in ["производственн", "production"]):
            semantic_info["data_categories"].append("production")
        if any(word in text_joined for word in ["хозяйственн", "бытов", "household"]):
            semantic_info["data_categories"].append("household")
        
        return semantic_info
    
    def close(self):
        """Закрытие рабочей книги."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def analyze_sheet_structure(
    template_path: Path,
    sheet_name: str,
    output_path: Optional[Path] = None,
    max_rows: Optional[int] = None
) -> Dict[str, Any]:
    """
    Анализ структуры таблиц в листе.
    
    Args:
        template_path: Путь к шаблону
        sheet_name: Имя листа
        output_path: Путь для сохранения результатов (опционально)
        max_rows: Максимальное количество строк для анализа
        
    Returns:
        Словарь с результатами анализа
    """
    analyzer = AITableStructureAnalyzer(template_path)
    
    try:
        result = analyzer.analyze_sheet(sheet_name, max_rows)
        
        if output_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(
                json.dumps(result, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )
        
        return result
    finally:
        analyzer.close()


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="ИИ-анализ структуры таблиц в листе Excel")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--sheet", required=True, help="Имя листа для анализа")
    parser.add_argument("--output", help="Путь для сохранения JSON")
    parser.add_argument("--max-rows", type=int, help="Максимальное количество строк")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    output_path = Path(args.output) if args.output else None
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    print(f"Анализ листа '{args.sheet}' в шаблоне: {template_path}")
    result = analyze_sheet_structure(template_path, args.sheet, output_path, args.max_rows)
    
    print("\n✅ Анализ завершен")
    print("📊 Результаты:")
    print(f"  Найдено таблиц: {result['summary']['total_tables']}")
    print(f"  Всего строк: {result['summary']['total_rows']}")
    print(f"  Всего столбцов: {result['summary']['total_columns']}")
    print(f"  Назначение листа: {result['semantic_analysis']['sheet_purpose']}")
    print(f"  Типы ресурсов: {', '.join(result['semantic_analysis']['resource_types']) or 'нет'}")
    
    if output_path:
        print(f"\n💾 Результаты сохранены в: {output_path}")

