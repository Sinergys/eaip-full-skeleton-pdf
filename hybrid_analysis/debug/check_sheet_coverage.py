"""
Скрипт для проверки покрытия листов шаблона
Показывает какие листы обрабатываются и какие пропускаются
"""

import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))


class SheetCoverageChecker:
    """Проверка покрытия листов шаблона."""
    
    def __init__(self, template_path: Path, mapping_path: Optional[Path] = None):
        """
        Инициализация проверки.
        
        Args:
            template_path: Путь к шаблону Excel
            mapping_path: Путь к файлу маппинга (опционально)
        """
        self.template_path = template_path
        self.mapping_path = mapping_path
        self.workbook = None
        self.sheet_names = []
        self.mapped_sheets = set()
        
    def load_template(self) -> None:
        """Загрузка шаблона."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
        
        self.workbook = load_workbook(self.template_path, read_only=True, data_only=True)
        self.sheet_names = self.workbook.sheetnames
        
    def load_mapping(self) -> Dict[str, Any]:
        """Загрузка маппинга."""
        if not self.mapping_path or not self.mapping_path.exists():
            return {"mappings": []}
        
        with open(self.mapping_path, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def check_function_coverage(self) -> Dict[str, List[str]]:
        """Проверка покрытия листов функциями заполнения."""
        # Имена функций заполнения из tools/fill_energy_passport.py
        function_sheets = {
            "fill_struktura_pr2": [
                "Структура пр 2 ", "Struktura pr2", "Структура пр 2"
            ],
            "fill_balans_sheet": [
                "Баланс", "04_Баланс", "Balans", "04_Balans"
            ],
            "fill_dinamika_sheet": [
                "Динамика ср", "05_Динамика", "Dinamika sr", "05_Dinamika", "Динамика"
            ],
            "fill_meropriyatiya_sheet": [
                "Мериаприятия 1 ", "06_Мероприятия", "Мероприятия", "Meropriyatiya", 
                "06_Meropriyatiya", "Мероприятия 1"
            ],
            "fill_nodes_sheet": [
                "Узел учета ", "01_Узлы учета", "Узлы учета", "Uzel ucheta", 
                "01_Uzly_ucheta", "Узел учета"
            ],
            "fill_fuel_dynamics_sheet": [
                "мазут,уголь 5 ", "мазут,уголь 5", "05_Мазут_Уголь", "Fuel Dynamics",
                "Мазут,уголь 5", "мазут уголь 5"
            ],
            "fill_specific_consumption_sheet": [
                "Расход  на ед.п", "Расход на ед.п", "06_Расход_на_ед", "Specific Consumption",
                "Расход на единиц", "Расход на ед"
            ]
        }
        
        covered = {}
        uncovered = []
        
        for sheet_name in self.sheet_names:
            found = False
            for func_name, variants in function_sheets.items():
                if sheet_name in variants or sheet_name.strip() in [v.strip() for v in variants]:
                    if func_name not in covered:
                        covered[func_name] = []
                    covered[func_name].append(sheet_name)
                    self.mapped_sheets.add(sheet_name)
                    found = True
                    break
            
            if not found:
                uncovered.append(sheet_name)
        
        return {
            "covered": covered,
            "uncovered": uncovered
        }
    
    def check_mapping_coverage(self, mapping_data: Dict[str, Any]) -> Dict[str, Any]:
        """Проверка покрытия листов в маппинге."""
        mappings = mapping_data.get("mappings", [])
        
        mapped_sheets = set()
        for mapping in mappings:
            sheet_name = mapping.get("sheet")
            if sheet_name:
                mapped_sheets.add(sheet_name)
                self.mapped_sheets.add(sheet_name)
        
        all_sheets = set(self.sheet_names)
        unmapped_sheets = all_sheets - mapped_sheets
        
        return {
            "mapped_sheets": sorted(list(mapped_sheets)),
            "unmapped_sheets": sorted(list(unmapped_sheets)),
            "total_mappings": len(mappings),
            "mappings_by_sheet": self._count_mappings_by_sheet(mappings)
        }
    
    def _count_mappings_by_sheet(self, mappings: List[Dict]) -> Dict[str, int]:
        """Подсчет маппингов по листам."""
        counts = {}
        for mapping in mappings:
            sheet_name = mapping.get("sheet")
            if sheet_name:
                counts[sheet_name] = counts.get(sheet_name, 0) + 1
        return counts
    
    def check_data_coverage(self) -> Dict[str, Any]:
        """Проверка заполненности листов данными."""
        if not self.workbook:
            self.load_template()
        
        coverage = {}
        
        for sheet_name in self.sheet_names:
            ws = self.workbook[sheet_name]
            
            # Подсчет заполненных и пустых ячеек
            total_cells = 0
            filled_cells = 0
            empty_cells = 0
            
            # Проверяем область с данными (первые 100 строк и 50 столбцов)
            for row in ws.iter_rows(max_row=min(100, ws.max_row), max_col=min(50, ws.max_column)):
                for cell in row:
                    total_cells += 1
                    if cell.value is not None:
                        value_str = str(cell.value).strip()
                        # Не считаем пустыми: формулы, пробелы, нули в текстовых контекстах
                        if value_str and value_str not in ["", "None", "0"]:
                            filled_cells += 1
                        else:
                            empty_cells += 1
                    else:
                        empty_cells += 1
            
            fill_percentage = (filled_cells / total_cells * 100) if total_cells > 0 else 0
            
            coverage[sheet_name] = {
                "total_cells": total_cells,
                "filled_cells": filled_cells,
                "empty_cells": empty_cells,
                "fill_percentage": round(fill_percentage, 2),
                "max_row": ws.max_row,
                "max_column": ws.max_column
            }
        
        return coverage
    
    def generate_report(self) -> Dict[str, Any]:
        """Генерация полного отчета."""
        self.load_template()
        
        # Проверка покрытия функциями
        function_coverage = self.check_function_coverage()
        
        # Проверка маппинга
        mapping_data = {}
        if self.mapping_path:
            mapping_data = self.load_mapping()
            mapping_coverage = self.check_mapping_coverage(mapping_data)
        else:
            mapping_coverage = {
                "mapped_sheets": [],
                "unmapped_sheets": self.sheet_names,
                "total_mappings": 0,
                "mappings_by_sheet": {}
            }
        
        # Проверка заполненности данными
        data_coverage = self.check_data_coverage()
        
        report = {
            "template_path": str(self.template_path),
            "total_sheets": len(self.sheet_names),
            "sheet_names": self.sheet_names,
            "function_coverage": function_coverage,
            "mapping_coverage": mapping_coverage,
            "data_coverage": data_coverage,
            "summary": {
                "covered_by_functions": len(function_coverage.get("covered", {})),
                "uncovered_by_functions": len(function_coverage.get("uncovered", [])),
                "covered_by_mapping": len(mapping_coverage.get("mapped_sheets", [])),
                "uncovered_by_mapping": len(mapping_coverage.get("unmapped_sheets", [])),
                "sheets_requiring_attention": self._identify_sheets_requiring_attention(
                    function_coverage, mapping_coverage, data_coverage
                )
            }
        }
        
        return report
    
    def _identify_sheets_requiring_attention(self, function_coverage, mapping_coverage, data_coverage) -> List[str]:
        """Определение листов, требующих внимания."""
        requiring_attention = []
        
        uncovered_by_functions = function_coverage.get("uncovered", [])
        uncovered_by_mapping = mapping_coverage.get("unmapped_sheets", [])
        
        # Листы, которые не покрыты ни функциями, ни маппингом
        uncovered = set(uncovered_by_functions) & set(uncovered_by_mapping)
        requiring_attention.extend(list(uncovered))
        
        # Листы с низким процентом заполнения (< 10%)
        for sheet_name, coverage_info in data_coverage.items():
            if coverage_info["fill_percentage"] < 10 and sheet_name not in requiring_attention:
                requiring_attention.append(sheet_name)
        
        return sorted(requiring_attention)
    
    def close(self):
        """Закрытие рабочей книги."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def print_report(report: Dict[str, Any]) -> None:
    """Вывод отчета в консоль."""
    print("=" * 80)
    print("📋 ОТЧЕТ О ПОКРЫТИИ ЛИСТОВ ШАБЛОНА")
    print("=" * 80)
    print(f"Шаблон: {report['template_path']}")
    print(f"Всего листов: {report['total_sheets']}")
    print()
    
    # Функциональное покрытие
    print("🔧 ПОКРЫТИЕ ФУНКЦИЯМИ ЗАПОЛНЕНИЯ:")
    print("-" * 80)
    func_coverage = report["function_coverage"]
    for func_name, sheets in func_coverage.get("covered", {}).items():
        print(f"  ✅ {func_name}:")
        for sheet in sheets:
            print(f"     - {sheet}")
    
    uncovered = func_coverage.get("uncovered", [])
    if uncovered:
        print(f"\n  ❌ Листы без функций заполнения ({len(uncovered)}):")
        for sheet in uncovered:
            print(f"     - {sheet}")
    print()
    
    # Покрытие маппингом
    print("🗺️  ПОКРЫТИЕ МАППИНГОМ:")
    print("-" * 80)
    mapping_coverage = report["mapping_coverage"]
    mapped = mapping_coverage.get("mapped_sheets", [])
    print(f"  ✅ Листов с маппингом: {len(mapped)}")
    for sheet in mapped:
        count = mapping_coverage["mappings_by_sheet"].get(sheet, 0)
        print(f"     - {sheet}: {count} маппингов")
    
    unmapped = mapping_coverage.get("unmapped_sheets", [])
    if unmapped:
        print(f"\n  ❌ Листы без маппинга ({len(unmapped)}):")
        for sheet in unmapped:
            print(f"     - {sheet}")
    print()
    
    # Заполненность данными
    print("📊 ЗАПОЛНЕННОСТЬ ДАННЫМИ:")
    print("-" * 80)
    data_coverage = report["data_coverage"]
    for sheet_name, coverage_info in sorted(data_coverage.items()):
        percentage = coverage_info["fill_percentage"]
        status = "✅" if percentage > 50 else "⚠️" if percentage > 10 else "❌"
        print(f"  {status} {sheet_name}: {percentage}% "
              f"({coverage_info['filled_cells']}/{coverage_info['total_cells']} ячеек)")
    print()
    
    # Сводка
    print("📋 СВОДКА:")
    print("-" * 80)
    summary = report["summary"]
    print(f"  Покрыто функциями: {summary['covered_by_functions']}/{report['total_sheets']}")
    print(f"  Покрыто маппингом: {summary['covered_by_mapping']}/{report['total_sheets']}")
    
    requiring_attention = summary["sheets_requiring_attention"]
    if requiring_attention:
        print(f"\n  ⚠️  Листы, требующие внимания ({len(requiring_attention)}):")
        for sheet in requiring_attention:
            print(f"     - {sheet}")
    else:
        print("\n  ✅ Все листы покрыты!")
    
    print("=" * 80)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Проверка покрытия листов шаблона")
    parser.add_argument("--template", required=True, help="Путь к шаблону Excel")
    parser.add_argument("--mapping", help="Путь к файлу маппинга JSON")
    parser.add_argument("--output", help="Путь для сохранения отчета JSON")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    mapping_path = Path(args.mapping) if args.mapping else None
    output_path = Path(args.output) if args.output else None
    
    with SheetCoverageChecker(template_path, mapping_path) as checker:
        report = checker.generate_report()
        
        print_report(report)
        
        if output_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            print(f"\n💾 Отчет сохранен в: {output_path}")

