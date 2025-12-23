"""
Этап 3.1: Анализ паттернов заполнения
Выявление паттернов форматирования и заполнения
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from collections import defaultdict
import re


class PatternAnalyzer:
    """Анализатор паттернов заполнения шаблонов."""
    
    def __init__(self, template_path: Path, filled_template_path: Optional[Path] = None):
        """
        Инициализация анализатора.
        
        Args:
            template_path: Путь к исходному шаблону
            filled_template_path: Путь к заполненному шаблону (опционально)
        """
        self.template_path = template_path
        self.filled_template_path = filled_template_path
        self.patterns = {}
    
    def analyze(self) -> Dict[str, Any]:
        """
        Анализ паттернов заполнения.
        
        Returns:
            Словарь с паттернами
        """
        self.patterns = {
            "template_name": self.template_path.stem,
            "number_formats": {},
            "text_patterns": {},
            "filling_patterns": {},
            "statistics": {}
        }
        
        # Анализ исходного шаблона
        template_patterns = self._analyze_template()
        self.patterns["template_patterns"] = template_patterns
        
        # Анализ заполненного шаблона (если есть)
        if self.filled_template_path and self.filled_template_path.exists():
            filled_patterns = self._analyze_filled_template()
            self.patterns["filled_patterns"] = filled_patterns
            self.patterns["filling_patterns"] = self._compare_patterns(template_patterns, filled_patterns)
        
        # Статистика
        self.patterns["statistics"] = self._calculate_statistics()
        
        return self.patterns
    
    def _analyze_template(self) -> Dict[str, Any]:
        """Анализ паттернов исходного шаблона."""
        workbook = load_workbook(self.template_path, data_only=False)
        
        patterns = {
            "sheets": {},
            "number_formats": defaultdict(list),
            "text_patterns": defaultdict(int)
        }
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            sheet_patterns = self._analyze_sheet(ws)
            patterns["sheets"][sheet_name] = sheet_patterns
            
            # Агрегация форматов
            for format_type, cells in sheet_patterns.get("number_formats", {}).items():
                patterns["number_formats"][format_type].extend(cells)
        
        return patterns
    
    def _analyze_sheet(self, ws) -> Dict[str, Any]:
        """Анализ паттернов одного листа."""
        patterns = {
            "number_formats": defaultdict(list),
            "text_patterns": defaultdict(int),
            "data_regions": []
        }
        
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                
                if cell.value is None:
                    continue
                
                # Анализ форматов чисел
                if isinstance(cell.value, (int, float)):
                    format_info = self._analyze_number_format(cell)
                    format_type = format_info["type"]
                    patterns["number_formats"][format_type].append({
                        "address": cell.coordinate,
                        "value": cell.value,
                        "format": format_info
                    })
                
                # Анализ текстовых паттернов
                elif isinstance(cell.value, str):
                    text_pattern = self._analyze_text_pattern(cell.value)
                    patterns["text_patterns"][text_pattern] += 1
        
        return patterns
    
    def _analyze_number_format(self, cell) -> Dict[str, Any]:
        """Анализ формата числа."""
        value = cell.value
        number_format = cell.number_format if hasattr(cell, 'number_format') else "General"
        
        # Определение типа формата
        if abs(value) >= 1000:
            format_type = "large_number"
            precision = self._detect_precision(value)
        elif abs(value) < 1 and value != 0:
            format_type = "decimal"
            precision = self._detect_precision(value)
        else:
            format_type = "integer" if isinstance(value, int) or value.is_integer() else "decimal"
            precision = self._detect_precision(value)
        
        return {
            "type": format_type,
            "precision": precision,
            "number_format": number_format,
            "magnitude": self._get_magnitude(value)
        }
    
    def _detect_precision(self, value: float) -> int:
        """Определение точности числа."""
        if isinstance(value, int) or value.is_integer():
            return 0
        
        # Подсчет значащих цифр после запятой
        str_value = str(value)
        if '.' in str_value:
            decimal_part = str_value.split('.')[1]
            # Убираем завершающие нули
            decimal_part = decimal_part.rstrip('0')
            return len(decimal_part)
        
        return 0
    
    def _get_magnitude(self, value: float) -> str:
        """Определение порядка величины."""
        abs_value = abs(value)
        if abs_value >= 1000000:
            return "millions"
        elif abs_value >= 1000:
            return "thousands"
        elif abs_value >= 1:
            return "units"
        else:
            return "fractions"
    
    def _analyze_text_pattern(self, text: str) -> str:
        """Анализ текстового паттерна."""
        text_lower = text.lower()
        
        if re.search(r'\{\{[^}]+\}\}', text):
            return "placeholder_curly"
        elif re.search(r'_+', text) and any(kw in text_lower for kw in ['год', 'квартал', 'месяц']):
            return "placeholder_underscore"
        elif any(kw in text_lower for kw in ['таблица', 'структура', 'баланс']):
            return "header"
        elif any(kw in text_lower for kw in ['потребление', 'нужды', 'производство']):
            return "category_label"
        elif re.search(r'\d+[./]\d+[./]\d+', text):
            return "date_like"
        else:
            return "text"
    
    def _analyze_filled_template(self) -> Dict[str, Any]:
        """Анализ паттернов заполненного шаблона."""
        workbook = load_workbook(self.filled_template_path, data_only=False)
        
        patterns = {
            "sheets": {},
            "filling_statistics": {}
        }
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            filled_count = 0
            empty_count = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.value is None:
                        empty_count += 1
                    else:
                        filled_count += 1
            
            patterns["sheets"][sheet_name] = {
                "filled_cells": filled_count,
                "empty_cells": empty_count,
                "fill_rate": filled_count / (filled_count + empty_count) if (filled_count + empty_count) > 0 else 0
            }
        
        return patterns
    
    def _compare_patterns(self, template_patterns: Dict, filled_patterns: Dict) -> Dict[str, Any]:
        """Сравнение паттернов шаблона и заполненного файла."""
        comparison = {
            "format_changes": [],
            "filling_statistics": filled_patterns.get("filling_statistics", {})
        }
        
        return comparison
    
    def _calculate_statistics(self) -> Dict[str, Any]:
        """Расчет статистики паттернов."""
        stats = {
            "total_number_formats": len(self.patterns.get("number_formats", {})),
            "format_distribution": {},
            "text_pattern_distribution": {}
        }
        
        # Распределение форматов чисел
        for format_type, cells in self.patterns.get("number_formats", {}).items():
            stats["format_distribution"][format_type] = len(cells)
        
        # Распределение текстовых паттернов
        template_patterns = self.patterns.get("template_patterns", {})
        for sheet_name, sheet_data in template_patterns.get("sheets", {}).items():
            for pattern, count in sheet_data.get("text_patterns", {}).items():
                if pattern not in stats["text_pattern_distribution"]:
                    stats["text_pattern_distribution"][pattern] = 0
                stats["text_pattern_distribution"][pattern] += count
        
        return stats
    
    def save(self, output_path: Path) -> None:
        """
        Сохранение результатов анализа.
        
        Args:
            output_path: Путь для сохранения JSON
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.patterns, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8"
        )


def analyze_patterns(
    template_path: Path,
    output_path: Path,
    filled_template_path: Optional[Path] = None
) -> Dict[str, Any]:
    """
    Анализ паттернов заполнения и сохранение результатов.
    
    Args:
        template_path: Путь к шаблону
        output_path: Путь для сохранения результатов
        filled_template_path: Путь к заполненному шаблону (опционально)
    
    Returns:
        Словарь с паттернами
    """
    analyzer = PatternAnalyzer(template_path, filled_template_path)
    patterns = analyzer.analyze()
    analyzer.save(output_path)
    return patterns


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Анализ паттернов заполнения")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    parser.add_argument("--filled-template", help="Путь к заполненному шаблону")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    output_path = Path(args.output)
    filled_path = Path(args.filled_template) if args.filled_template else None
    
    print("Анализ паттернов заполнения...")
    patterns = analyze_patterns(template_path, output_path, filled_path)
    
    print(f"\n✅ Результаты сохранены в: {output_path}")
    print("📊 Статистика:")
    print(f"  Форматов чисел: {patterns['statistics']['total_number_formats']}")
    print("  Распределение форматов:")
    for format_type, count in patterns['statistics']['format_distribution'].items():
        print(f"    - {format_type}: {count}")

