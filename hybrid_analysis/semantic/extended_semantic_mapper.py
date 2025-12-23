"""
Расширенный семантический маппер
Использование полного анализа структуры для создания маппинга всех листов
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
import re


class ExtendedSemanticMapper:
    """Расширенный маппер для создания полного семантического маппинга."""
    
    def __init__(self,
                 template_path: Path,
                 full_analysis_path: Path,
                 aggregated_data_path: Path,
                 ontology_path: Optional[Path] = None):
        """
        Инициализация расширенного маппера.
        
        Args:
            template_path: Путь к шаблону Excel
            full_analysis_path: Путь к полному анализу листов
            aggregated_data_path: Путь к агрегированным данным
            ontology_path: Путь к онтологии (опционально)
        """
        self.template_path = Path(template_path)
        self.full_analysis_path = Path(full_analysis_path)
        self.aggregated_data_path = Path(aggregated_data_path)
        self.ontology_path = Path(ontology_path) if ontology_path else None
        
        self.workbook = None
        self.full_analysis = {}
        self.aggregated_data = {}
        self.ontology = {}
        
        self.extended_mapping = {
            "template_name": self.template_path.stem,
            "mappings": [],
            "unmapped_cells": [],
            "statistics": {
                "total_cells": 0,
                "mapped_cells": 0,
                "unmapped_cells": 0,
                "coverage_percentage": 0.0
            }
        }
    
    def load_data(self) -> None:
        """Загрузка всех необходимых данных."""
        # Загрузка шаблона
        if not self.template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
        
        self.workbook = load_workbook(self.template_path, data_only=True)
        
        # Загрузка полного анализа
        if not self.full_analysis_path.exists():
            raise FileNotFoundError(f"Анализ не найден: {self.full_analysis_path}")
        
        self.full_analysis = json.loads(
            self.full_analysis_path.read_text(encoding="utf-8")
        )
        
        # Загрузка агрегированных данных
        if not self.aggregated_data_path.exists():
            raise FileNotFoundError(f"Данные не найдены: {self.aggregated_data_path}")
        
        self.aggregated_data = json.loads(
            self.aggregated_data_path.read_text(encoding="utf-8")
        )
        
        # Загрузка онтологии
        if self.ontology_path and self.ontology_path.exists():
            self.ontology = json.loads(
                self.ontology_path.read_text(encoding="utf-8")
            )
    
    def create_extended_mapping(self) -> Dict[str, Any]:
        """
        Создание расширенного маппинга для всех листов.
        
        Returns:
            Словарь с расширенным маппингом
        """
        print("Создание расширенного семантического маппинга...")
        print(f"Шаблон: {self.template_path.name}")
        print(f"Анализ листов: {self.full_analysis_path.name}")
        
        # Нормализация данных
        normalized_data = self._normalize_data(self.aggregated_data)
        
        # Получение классификации листов
        sheet_classifications = self._get_sheet_classifications()
        
        # Маппинг для каждого листа
        for sheet_name in self.workbook.sheetnames:
            if sheet_name not in sheet_classifications:
                continue
            
            sheet_info = sheet_classifications[sheet_name]
            sheet_type = sheet_info.get("type", "unknown")
            
            print(f"\nОбработка листа: {sheet_name} (тип: {sheet_type})")
            
            # Маппинг в зависимости от типа листа
            if sheet_type == "balance":
                self._map_balance_sheet(sheet_name, normalized_data, sheet_info)
            elif sheet_type == "dynamics":
                self._map_dynamics_sheet(sheet_name, normalized_data, sheet_info)
            elif sheet_type == "consumption_per_unit":
                self._map_consumption_sheet(sheet_name, normalized_data, sheet_info)
            elif sheet_type == "metering_nodes":
                self._map_metering_nodes_sheet(sheet_name, normalized_data, sheet_info)
            elif sheet_type == "measures":
                self._map_measures_sheet(sheet_name, normalized_data, sheet_info)
            elif sheet_type == "regulatory_documentation":
                self._map_regulatory_sheet(sheet_name, normalized_data, sheet_info)
            else:
                # Общий маппинг для неизвестных типов
                self._map_generic_sheet(sheet_name, normalized_data, sheet_info)
        
        # Расчет статистики
        self._calculate_statistics()
        
        return self.extended_mapping
    
    def _get_sheet_classifications(self) -> Dict[str, Dict[str, Any]]:
        """Получение классификации листов из анализа."""
        classifications = {}
        
        # Попытка получить из сводки (sheets_analysis или sheets)
        if "sheets_analysis" in self.full_analysis:
            for sheet_name, sheet_data in self.full_analysis["sheets_analysis"].items():
                # Определение типа листа по имени и структуре
                sheet_type = self._determine_sheet_type(sheet_name, sheet_data)
                classifications[sheet_name] = {
                    "type": sheet_type,
                    "description": "",
                    "table_structure": sheet_data.get("table_structure_analysis", {}),
                    "semantic_profile": {},
                    "structural_analysis": sheet_data.get("structural_analysis", {})
                }
        elif "sheets" in self.full_analysis:
            for sheet_name, sheet_data in self.full_analysis["sheets"].items():
                classifications[sheet_name] = {
                    "type": sheet_data.get("type", "unknown"),
                    "description": sheet_data.get("description", ""),
                    "table_structure": sheet_data.get("table_structure", {}),
                    "semantic_profile": sheet_data.get("semantic_profile", {})
                }
        
        return classifications
    
    def _determine_sheet_type(self, sheet_name: str, sheet_data: Dict[str, Any]) -> str:
        """Определение типа листа по имени и структуре."""
        sheet_lower = sheet_name.lower()
        
        # Классификация по имени
        if "динамик" in sheet_lower or "dynamics" in sheet_lower:
            return "dynamics"
        elif "баланс" in sheet_lower or "balance" in sheet_lower:
            return "balance"
        elif "расход" in sheet_lower or "consumption" in sheet_lower:
            return "consumption_per_unit"
        elif "узел" in sheet_lower or "учет" in sheet_lower or "metering" in sheet_lower:
            return "metering_nodes"
        elif "мероприя" in sheet_lower or "measures" in sheet_lower:
            return "measures"
        elif "структур" in sheet_lower or "structure" in sheet_lower:
            return "balance"  # Структура потребления - это тоже баланс
        elif sheet_name == "Sheet1" or "dastur" in sheet_lower:
            return "regulatory_documentation"
        elif "мазут" in sheet_lower or "уголь" in sheet_lower or "fuel" in sheet_lower:
            return "balance"  # Мазут/уголь обычно в балансе
        
        return "unknown"
    
    def _normalize_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Нормализация структуры данных."""
        normalized = {"resources": {}}
        
        if isinstance(data, dict):
            if "resources" in data:
                normalized["resources"] = data["resources"]
            else:
                # Преобразуем из file-based формата
                for file_data in data.values():
                    if isinstance(file_data, dict) and "resources" in file_data:
                        for resource_type, resource_data in file_data["resources"].items():
                            if resource_type not in normalized["resources"]:
                                normalized["resources"][resource_type] = {}
                            if isinstance(resource_data, dict):
                                for quarter, quarter_data in resource_data.items():
                                    if quarter not in normalized["resources"][resource_type]:
                                        normalized["resources"][resource_type][quarter] = quarter_data
        
        return normalized
    
    def _map_balance_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                          sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа баланса на основе детального анализа."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        ws = self.workbook[sheet_name]
        table_structure = sheet_info.get("table_structure", {})
        
        # Для листа "Структура пр 2" используем известную структуру
        if "Структура пр 2" in sheet_name:
            self._map_struktura_sheet(sheet_name, normalized_data, sheet_info)
            return
        
        # Поиск таблиц с данными
        if "tables" in table_structure:
            for table in table_structure["tables"]:
                self._map_table_cells(ws, table, normalized_data, sheet_name)
    
    def _map_struktura_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                            sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа 'Структура пр 2'."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        ws = self.workbook[sheet_name]
        
        # Структура листа "Структура пр 2":
        # Строка 6: заголовки с единицами измерения
        # Колонка C: Электроэнергия актив (кВт·ч)
        # Колонка D: Электроэнергия реактив (кВар·ч)
        # Колонка F: Газ (тыс. м³)
        # Колонка N: Вода (тыс. м³)
        
        # Строки с данными по кварталам (известная структура):
        # 2022-Q1: строка 9, колонки C, D, F, N
        # 2022-Q2: строка 9, колонки C+4, D+4, F+4, N+4 и т.д.
        
        quarter_row_map = {
            "2022-Q1": {"row": 9, "col_active": 3, "col_reactive": 4, "col_gas": 6, "col_water": 14},
            "2022-Q2": {"row": 9, "col_active": 19, "col_reactive": 20, "col_gas": 22, "col_water": 30},
            "2022-Q3": {"row": 9, "col_active": 35, "col_reactive": 36, "col_gas": 38, "col_water": 46},
            "2022-Q4": {"row": 9, "col_active": 51, "col_reactive": 52, "col_gas": 54, "col_water": 62}
        }
        
        # Маппинг для каждого квартала
        for quarter, positions in quarter_row_map.items():
            # Проверка наличия данных
            if quarter not in normalized_data.get("resources", {}).get("electricity", {}):
                continue
            
            # Маппинг активной электроэнергии
            self._add_mapping(
                sheet_name,
                ws.cell(row=positions["row"], column=positions["col_active"]).coordinate,
                "electricity_active",
                f"resources.electricity.{quarter}.quarter_totals.active_kwh",
                confidence=0.95
            )
            
            # Маппинг реактивной электроэнергии
            self._add_mapping(
                sheet_name,
                ws.cell(row=positions["row"], column=positions["col_reactive"]).coordinate,
                "electricity_reactive",
                f"resources.electricity.{quarter}.quarter_totals.reactive_kvarh",
                confidence=0.95
            )
            
            # Маппинг газа
            if quarter in normalized_data.get("resources", {}).get("gas", {}):
                self._add_mapping(
                    sheet_name,
                    ws.cell(row=positions["row"], column=positions["col_gas"]).coordinate,
                    "gas_volume",
                    f"resources.gas.{quarter}.quarter_totals.volume_m3",
                    confidence=0.95
                )
            
            # Маппинг воды
            if quarter in normalized_data.get("resources", {}).get("water", {}):
                self._add_mapping(
                    sheet_name,
                    ws.cell(row=positions["row"], column=positions["col_water"]).coordinate,
                    "water_volume",
                    f"resources.water.{quarter}.quarter_totals.volume_m3",
                    confidence=0.95
                )
    
    def _map_dynamics_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                           sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа динамики на основе детального анализа структуры."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        ws = self.workbook[sheet_name]
        table_structure = sheet_info.get("table_structure", {})
        
        # Определение колонок для ресурсов на основе заголовков
        # Электроэнергия: колонка C (3), кварталы I-IV в колонках C-F (3-6)
        # Тепловая энергия: колонка H (8), кварталы I-IV в колонках H-K (8-11)
        # Природный газ: колонка M (13), кварталы I-IV в колонках M-P (13-16)
        
        resource_columns = {
            "electricity": {
                "start_col": 3,  # C
                "quarters": {"I": 3, "II": 4, "III": 5, "IV": 6},  # C, D, E, F
                "semantic_type": "electricity_active"
            },
            "gas": {
                "start_col": 13,  # M
                "quarters": {"I": 13, "II": 14, "III": 15, "IV": 16},  # M, N, O, P
                "semantic_type": "gas_volume"
            },
            "heat": {
                "start_col": 8,  # H
                "quarters": {"I": 8, "II": 9, "III": 10, "IV": 11},  # H, I, J, K
                "semantic_type": "heat_energy"
            }
        }
        
        # Определение строк с данными
        # Строка 13: "Хозяйственно-бытовое потребление"
        # Строка 15: "Общее потребление энергоресурсов"
        data_rows = [13, 15]
        
        # Маппинг для каждого ресурса и каждого квартала
        quarters_map = {"I": "Q1", "II": "Q2", "III": "Q3", "IV": "Q4"}
        
        # Получение доступных годов из данных
        years = set()
        if "resources" in normalized_data:
            for resource_type, resource_data in normalized_data["resources"].items():
                for quarter_key in resource_data.keys():
                    if "-Q" in quarter_key:
                        year = quarter_key.split("-")[0]
                        years.add(year)
        
        if not years:
            # Используем год по умолчанию
            years = {"2022"}
        
        # Маппинг для строки 15 (Общее потребление)
        target_row = 15
        
        for year in sorted(years):
            for quarter_roman, quarter_num in quarters_map.items():
                quarter = f"{year}-{quarter_num}"
                
                # Маппинг электроэнергии
                if "electricity" in normalized_data.get("resources", {}) and \
                   quarter in normalized_data["resources"]["electricity"]:
                    col = resource_columns["electricity"]["quarters"][quarter_roman]
                    cell_address = ws.cell(row=target_row, column=col).coordinate
                    self._add_mapping(
                        sheet_name,
                        cell_address,
                        "electricity_active",
                        f"resources.electricity.{quarter}.quarter_totals.active_kwh",
                        confidence=0.9
                    )
                
                # Маппинг газа
                if "gas" in normalized_data.get("resources", {}) and \
                   quarter in normalized_data["resources"]["gas"]:
                    col = resource_columns["gas"]["quarters"][quarter_roman]
                    cell_address = ws.cell(row=target_row, column=col).coordinate
                    self._add_mapping(
                        sheet_name,
                        cell_address,
                        "gas_volume",
                        f"resources.gas.{quarter}.quarter_totals.volume_m3",
                        confidence=0.9
                    )
                
                # Маппинг тепловой энергии (если есть в данных)
                if "heat" in normalized_data.get("resources", {}) and \
                   quarter in normalized_data["resources"]["heat"]:
                    col = resource_columns["heat"]["quarters"][quarter_roman]
                    cell_address = ws.cell(row=target_row, column=col).coordinate
                    self._add_mapping(
                        sheet_name,
                        cell_address,
                        "heat_energy",
                        f"resources.heat.{quarter}.quarter_totals.energy_gcal",
                        confidence=0.9
                    )
    
    def _map_consumption_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                              sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа расхода на единицу продукции."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        ws = self.workbook[sheet_name]
        # Поиск ячеек с формулами расхода
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Поиск формул вида =C{row}/G{row}
                    if "/" in cell.value and "=" in cell.value:
                        # Это формула расхода, не заполняем напрямую
                        pass
    
    def _map_metering_nodes_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                                 sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа узлов учета."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        ws = self.workbook[sheet_name]
        # Узлы учета обычно заполняются вручную, но можно добавить заголовки
    
    def _map_measures_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                           sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа энергосберегающих мероприятий."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        # Мероприятия обычно заполняются вручную
        pass
    
    def _map_regulatory_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                             sheet_info: Dict[str, Any]) -> None:
        """Маппинг для листа нормативной документации."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        # Нормативная документация обычно статична
        pass
    
    def _map_generic_sheet(self, sheet_name: str, normalized_data: Dict[str, Any],
                          sheet_info: Dict[str, Any]) -> None:
        """Общий маппинг для неизвестных типов листов."""
        if sheet_name not in self.workbook.sheetnames:
            return
        
        ws = self.workbook[sheet_name]
        semantic_profile = sheet_info.get("semantic_profile", {})
        
        # Поиск семантических типов в профиле
        if "cells" in semantic_profile:
            for cell_address, cell_semantic in semantic_profile["cells"].items():
                semantic_type = cell_semantic.get("semantic_type")
                resource_type = cell_semantic.get("resource_type")
                
                if semantic_type and resource_type:
                    # Попытка найти соответствующие данные
                    data_path = self._find_data_path(semantic_type, resource_type)
                    if data_path:
                        self._add_mapping(
                            sheet_name,
                            cell_address,
                            semantic_type,
                            data_path,
                            confidence=0.6
                        )
    
    def _map_table_cells(self, ws, table: Dict[str, Any],
                        normalized_data: Dict[str, Any], sheet_name: str) -> None:
        """Маппинг ячеек таблицы."""
        start_row = table.get("start_row", 1)
        end_row = table.get("end_row", ws.max_row)
        start_col = table.get("start_col", 1)
        end_col = table.get("end_col", ws.max_column)
        
        # Поиск заголовков для определения колонок
        headers = {}
        for col in range(start_col, end_col + 1):
            header_cell = ws.cell(row=start_row, column=col)
            if header_cell.value:
                header_text = str(header_cell.value).lower()
                # Определение типа данных по заголовку
                if "электро" in header_text or "electricity" in header_text:
                    if "активн" in header_text or "active" in header_text:
                        headers[col] = "electricity_active"
                    elif "реактивн" in header_text or "reactive" in header_text:
                        headers[col] = "electricity_reactive"
                elif "газ" in header_text or "gas" in header_text:
                    headers[col] = "gas_volume"
                elif "вод" in header_text or "water" in header_text:
                    headers[col] = "water_volume"
        
        # Маппинг данных в таблице
        # Здесь можно добавить логику для заполнения данных по кварталам
    
    def _extract_quarters_from_table(self, table: Dict[str, Any], ws) -> List[Dict[str, Any]]:
        """Извлечение информации о кварталах из таблицы."""
        quarters = []
        
        # Простой поиск кварталов в первой колонке
        start_row = table.get("start_row", 1)
        end_row = table.get("end_row", ws.max_row)
        
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value:
                value_str = str(cell.value)
                # Поиск паттернов кварталов (2022-Q1, Q1, Q1 2022)
                quarter_match = re.search(r'(\d{4})-Q(\d)|Q(\d).*(\d{4})|Q(\d)', value_str)
                if quarter_match:
                    if quarter_match.group(1) and quarter_match.group(2):
                        year = quarter_match.group(1)
                        quarter_num = quarter_match.group(2)
                    elif quarter_match.group(4) and quarter_match.group(3):
                        year = quarter_match.group(4)
                        quarter_num = quarter_match.group(3)
                    else:
                        # Используем текущий год по умолчанию
                        year = "2022"
                        quarter_num = quarter_match.group(5)
                    
                    quarter = f"{year}-Q{quarter_num}"
                    quarters.append({
                        "quarter": quarter,
                        "row": row,
                        "col": 3  # По умолчанию колонка C
                    })
        
        return quarters
    
    def _find_data_path(self, semantic_type: str, resource_type: str) -> Optional[str]:
        """Поиск пути к данным по семантическому типу."""
        # Нормализация данных для поиска
        normalized = self._normalize_data(self.aggregated_data)
        # Получение первого доступного квартала
        resources = normalized.get("resources", {})
        if resource_type in resources:
            quarters = list(resources[resource_type].keys())
            if quarters:
                quarter = sorted(quarters)[0]
                
                if semantic_type == "electricity_active":
                    return f"resources.electricity.{quarter}.quarter_totals.active_kwh"
                elif semantic_type == "electricity_reactive":
                    return f"resources.electricity.{quarter}.quarter_totals.reactive_kvarh"
                elif semantic_type == "gas_volume":
                    return f"resources.gas.{quarter}.quarter_totals.volume_m3"
                elif semantic_type == "water_volume":
                    return f"resources.water.{quarter}.quarter_totals.volume_m3"
        
        return None
    
    def _add_mapping(self, sheet_name: str, cell_address: str, semantic_type: str,
                    data_path: str, confidence: float = 0.5) -> None:
        """Добавление маппинга."""
        mapping = {
            "cell_address": cell_address,
            "sheet": sheet_name,
            "semantic_type": semantic_type,
            "data_path": data_path,
            "confidence": confidence
        }
        self.extended_mapping["mappings"].append(mapping)
        self.extended_mapping["statistics"]["mapped_cells"] += 1
    
    def _calculate_statistics(self) -> None:
        """Расчет статистики маппинга."""
        mapped = self.extended_mapping["statistics"]["mapped_cells"]
        unmapped = self.extended_mapping["statistics"]["unmapped_cells"]
        total = mapped + unmapped
        
        if total > 0:
            coverage = (mapped / total) * 100
            self.extended_mapping["statistics"]["coverage_percentage"] = round(coverage, 2)
    
    def save(self, output_path: Path) -> None:
        """Сохранение расширенного маппинга."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.extended_mapping, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def create_extended_mapping(template_path: Path,
                           full_analysis_path: Path,
                           aggregated_data_path: Path,
                           output_path: Path,
                           ontology_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Создание расширенного семантического маппинга.
    
    Args:
        template_path: Путь к шаблону
        full_analysis_path: Путь к полному анализу листов
        aggregated_data_path: Путь к агрегированным данным
        output_path: Путь для сохранения результата
        ontology_path: Путь к онтологии (опционально)
    
    Returns:
        Словарь с расширенным маппингом
    """
    mapper = ExtendedSemanticMapper(
        template_path,
        full_analysis_path,
        aggregated_data_path,
        ontology_path
    )
    
    mapper.load_data()
    mapping = mapper.create_extended_mapping()
    mapper.save(output_path)
    
    return mapping


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Расширенный семантический маппинг")
    parser.add_argument("--template", required=True, help="Путь к шаблону")
    parser.add_argument("--analysis", required=True, help="Путь к полному анализу листов")
    parser.add_argument("--data", required=True, help="Путь к агрегированным данным")
    parser.add_argument("--output", required=True, help="Путь для сохранения результата")
    parser.add_argument("--ontology", help="Путь к онтологии")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    analysis_path = Path(args.analysis)
    data_path = Path(args.data)
    output_path = Path(args.output)
    ontology_path = Path(args.ontology) if args.ontology else None
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    if not analysis_path.exists():
        raise FileNotFoundError(f"Анализ не найден: {analysis_path}")
    if not data_path.exists():
        raise FileNotFoundError(f"Данные не найдены: {data_path}")
    
    print("=" * 80)
    print("СОЗДАНИЕ РАСШИРЕННОГО СЕМАНТИЧЕСКОГО МАППИНГА")
    print("=" * 80)
    
    mapping = create_extended_mapping(
        template_path,
        analysis_path,
        data_path,
        output_path,
        ontology_path
    )
    
    print(f"\n✅ Маппинг создан и сохранен в: {output_path}")
    print("\n📊 Статистика:")
    print(f"  Замаплено ячеек: {mapping['statistics']['mapped_cells']}")
    print(f"  Незамаплено ячеек: {mapping['statistics']['unmapped_cells']}")
    print(f"  Покрытие: {mapping['statistics']['coverage_percentage']:.2f}%")

