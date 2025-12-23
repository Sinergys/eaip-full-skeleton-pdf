"""
Этап 4.1: Структурное сопоставление шаблонов
Сравнение структуры листов и ячеек между шаблонами
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook


class StructuralComparator:
    """Класс для сравнения структур двух шаблонов."""
    
    def __init__(self, template1_path: Path, template2_path: Path):
        """
        Инициализация компаратора.
        
        Args:
            template1_path: Путь к первому шаблону (new_energy_passport.xlsx)
            template2_path: Путь ко второму шаблону (template_metin.xlsx)
        """
        self.template1_path = template1_path
        self.template2_path = template2_path
        self.template1_structure = None
        self.template2_structure = None
        self.comparison = {}
    
    def load_structures(self, template1_structure_path: Optional[Path] = None,
                       template2_structure_path: Optional[Path] = None) -> None:
        """
        Загрузка структур шаблонов.
        
        Args:
            template1_structure_path: Путь к JSON структуры первого шаблона (опционально)
            template2_structure_path: Путь к JSON структуры второго шаблона (опционально)
        """
        # Загрузка структуры первого шаблона
        if template1_structure_path and template1_structure_path.exists():
            self.template1_structure = json.loads(
                template1_structure_path.read_text(encoding="utf-8")
            )
        else:
            # Парсинг из Excel
            self.template1_structure = self._parse_template_structure(self.template1_path)
        
        # Загрузка структуры второго шаблона
        if template2_structure_path and template2_structure_path.exists():
            self.template2_structure = json.loads(
                template2_structure_path.read_text(encoding="utf-8")
            )
        else:
            # Парсинг из Excel
            self.template2_structure = self._parse_template_structure(self.template2_path)
    
    def _parse_template_structure(self, template_path: Path) -> Dict[str, Any]:
        """Быстрый парсинг структуры шаблона."""
        workbook = load_workbook(template_path, data_only=True)
        
        structure = {
            "template_path": str(template_path),
            "template_name": template_path.stem,
            "total_sheets": len(workbook.sheetnames),
            "sheets": {}
        }
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            structure["sheets"][sheet_name] = {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges_count": len(ws.merged_cells.ranges),
                "has_formulas": False,
                "cells_with_data": 0
            }
            
            # Подсчет ячеек с данными
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        structure["sheets"][sheet_name]["cells_with_data"] += 1
                        if cell.data_type == 'f':
                            structure["sheets"][sheet_name]["has_formulas"] = True
        
        return structure
    
    def compare(self) -> Dict[str, Any]:
        """
        Сравнение структур шаблонов.
        
        Returns:
            Словарь с результатами сравнения
        """
        if not self.template1_structure or not self.template2_structure:
            raise ValueError("Структуры шаблонов не загружены. Вызовите load_structures()")
        
        self.comparison = {
            "template1": {
                "name": self.template1_structure["template_name"],
                "path": self.template1_structure["template_path"],
                "total_sheets": self.template1_structure["total_sheets"]
            },
            "template2": {
                "name": self.template2_structure["template_name"],
                "path": self.template2_structure["template_path"],
                "total_sheets": self.template2_structure["total_sheets"]
            },
            "sheet_comparison": self._compare_sheets(),
            "structural_differences": self._find_structural_differences(),
            "similarity_metrics": self._calculate_similarity_metrics(),
            "conversion_feasibility": self._assess_conversion_feasibility()
        }
        
        return self.comparison
    
    def _compare_sheets(self) -> List[Dict[str, Any]]:
        """Сравнение листов между шаблонами."""
        sheets1 = set(self.template1_structure["sheets"].keys())
        sheets2 = set(self.template2_structure["sheets"].keys())
        
        common_sheets = sheets1.intersection(sheets2)
        unique_to_1 = sheets1 - sheets2
        unique_to_2 = sheets2 - sheets1
        
        comparison = []
        
        # Общие листы
        for sheet_name in common_sheets:
            sheet1 = self.template1_structure["sheets"][sheet_name]
            sheet2 = self.template2_structure["sheets"][sheet_name]
            
            comparison.append({
                "sheet_name": sheet_name,
                "status": "common",
                "template1": {
                    "max_row": sheet1["max_row"],
                    "max_column": sheet1["max_column"],
                    "cells_with_data": sheet1.get("cells_with_data", 0),
                    "merged_ranges": sheet1.get("merged_ranges_count", 0)
                },
                "template2": {
                    "max_row": sheet2["max_row"],
                    "max_column": sheet2["max_column"],
                    "cells_with_data": sheet2.get("cells_with_data", 0),
                    "merged_ranges": sheet2.get("merged_ranges_count", 0)
                },
                "differences": {
                    "row_diff": sheet1["max_row"] - sheet2["max_row"],
                    "col_diff": sheet1["max_column"] - sheet2["max_column"],
                    "data_diff": sheet1.get("cells_with_data", 0) - sheet2.get("cells_with_data", 0)
                }
            })
        
        # Уникальные листы для template1
        for sheet_name in unique_to_1:
            sheet1 = self.template1_structure["sheets"][sheet_name]
            comparison.append({
                "sheet_name": sheet_name,
                "status": "unique_to_template1",
                "template1": {
                    "max_row": sheet1["max_row"],
                    "max_column": sheet1["max_column"],
                    "cells_with_data": sheet1.get("cells_with_data", 0)
                }
            })
        
        # Уникальные листы для template2
        for sheet_name in unique_to_2:
            sheet2 = self.template2_structure["sheets"][sheet_name]
            comparison.append({
                "sheet_name": sheet_name,
                "status": "unique_to_template2",
                "template2": {
                    "max_row": sheet2["max_row"],
                    "max_column": sheet2["max_column"],
                    "cells_with_data": sheet2.get("cells_with_data", 0)
                }
            })
        
        return comparison
    
    def _find_structural_differences(self) -> Dict[str, Any]:
        """Поиск структурных различий."""
        differences = {
            "sheet_count_diff": self.template1_structure["total_sheets"] - self.template2_structure["total_sheets"],
            "total_cells_diff": 0,
            "complexity_diff": {}
        }
        
        # Подсчет общей сложности
        total_cells1 = sum(s.get("cells_with_data", 0) for s in self.template1_structure["sheets"].values())
        total_cells2 = sum(s.get("cells_with_data", 0) for s in self.template2_structure["sheets"].values())
        
        differences["total_cells_diff"] = total_cells1 - total_cells2
        
        differences["complexity_diff"] = {
            "template1_total_cells": total_cells1,
            "template2_total_cells": total_cells2,
            "relative_complexity": total_cells1 / total_cells2 if total_cells2 > 0 else 0
        }
        
        return differences
    
    def _calculate_similarity_metrics(self) -> Dict[str, float]:
        """Вычисление метрик схожести."""
        # Сравнение имен листов
        sheets1 = set(self.template1_structure["sheets"].keys())
        sheets2 = set(self.template2_structure["sheets"].keys())
        
        common_sheets = sheets1.intersection(sheets2)
        all_sheets = sheets1.union(sheets2)
        
        sheet_similarity = len(common_sheets) / len(all_sheets) if all_sheets else 0.0
        
        # Сравнение структуры листов
        structural_similarity = 0.0
        if common_sheets:
            similarities = []
            for sheet_name in common_sheets:
                sheet1 = self.template1_structure["sheets"][sheet_name]
                sheet2 = self.template2_structure["sheets"][sheet_name]
                
                # Сравнение размеров
                row_sim = min(sheet1["max_row"], sheet2["max_row"]) / max(sheet1["max_row"], sheet2["max_row"]) if max(sheet1["max_row"], sheet2["max_row"]) > 0 else 0
                col_sim = min(sheet1["max_column"], sheet2["max_column"]) / max(sheet1["max_column"], sheet2["max_column"]) if max(sheet1["max_column"], sheet2["max_column"]) > 0 else 0
                
                similarities.append((row_sim + col_sim) / 2)
            
            structural_similarity = sum(similarities) / len(similarities) if similarities else 0.0
        
        return {
            "sheet_name_similarity": sheet_similarity,
            "structural_similarity": structural_similarity,
            "overall_similarity": (sheet_similarity + structural_similarity) / 2
        }
    
    def _assess_conversion_feasibility(self) -> Dict[str, Any]:
        """Оценка возможности конвертации."""
        feasibility = {
            "is_feasible": True,
            "confidence": 0.0,
            "challenges": [],
            "recommendations": []
        }
        
        similarity = self.comparison.get("similarity_metrics", {}).get("overall_similarity", 0.0)
        
        if similarity > 0.7:
            feasibility["confidence"] = similarity
            feasibility["recommendations"].append("Высокая схожесть структур - конвертация возможна напрямую")
        elif similarity > 0.5:
            feasibility["confidence"] = similarity
            feasibility["challenges"].append("Умеренные структурные различия требуют адаптации")
            feasibility["recommendations"].append("Рекомендуется создание адаптеров для специфичных листов")
        else:
            feasibility["confidence"] = similarity
            feasibility["challenges"].append("Значительные структурные различия")
            feasibility["recommendations"].append("Требуется детальный анализ и создание специальных адаптеров")
        
        return feasibility
    
    def save(self, output_path: Path) -> None:
        """Сохранение результатов сравнения."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.comparison, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def compare_templates(template1_path: Path, template2_path: Path,
                     output_path: Path,
                     template1_structure_path: Optional[Path] = None,
                     template2_structure_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Сравнение двух шаблонов.
    
    Args:
        template1_path: Путь к первому шаблону
        template2_path: Путь ко второму шаблону
        output_path: Путь для сохранения результатов
        template1_structure_path: Путь к JSON структуры первого шаблона (опционально)
        template2_structure_path: Путь к JSON структуры второго шаблона (опционально)
    
    Returns:
        Словарь с результатами сравнения
    """
    comparator = StructuralComparator(template1_path, template2_path)
    comparator.load_structures(template1_structure_path, template2_structure_path)
    comparison = comparator.compare()
    comparator.save(output_path)
    
    return comparison


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Структурное сравнение шаблонов")
    parser.add_argument("--template1", required=True, help="Путь к первому шаблону")
    parser.add_argument("--template2", required=True, help="Путь ко второму шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    parser.add_argument("--structure1", help="Путь к JSON структуры первого шаблона")
    parser.add_argument("--structure2", help="Путь к JSON структуры второго шаблона")
    
    args = parser.parse_args()
    
    template1_path = Path(args.template1)
    template2_path = Path(args.template2)
    output_path = Path(args.output)
    structure1_path = Path(args.structure1) if args.structure1 else None
    structure2_path = Path(args.structure2) if args.structure2 else None
    
    if not template1_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template1_path}")
    if not template2_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template2_path}")
    
    print("Сравнение шаблонов:")
    print(f"  Template 1: {template1_path}")
    print(f"  Template 2: {template2_path}")
    
    comparison = compare_templates(template1_path, template2_path, output_path,
                                  structure1_path, structure2_path)
    
    print(f"\n✅ Результаты сохранены в: {output_path}")
    print("\n📊 Результаты сравнения:")
    print(f"  Схожесть структур: {comparison['similarity_metrics']['overall_similarity']:.2%}")
    print(f"  Совместимость: {comparison['conversion_feasibility']['confidence']:.2%}")
    print(f"  Общие листы: {len([s for s in comparison['sheet_comparison'] if s['status'] == 'common'])}")
    print(f"  Уникальные листы (Template 1): {len([s for s in comparison['sheet_comparison'] if s['status'] == 'unique_to_template1'])}")
    print(f"  Уникальные листы (Template 2): {len([s for s in comparison['sheet_comparison'] if s['status'] == 'unique_to_template2'])}")

