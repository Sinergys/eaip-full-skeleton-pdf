"""
Этап 4.3: Генерация адаптеров
Создание мостов между форматами шаблонов
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook


class AdapterGenerator:
    """Класс для генерации адаптеров между шаблонами."""
    
    def __init__(self, structural_comparison_path: Path, 
                 semantic_comparison_path: Optional[Path] = None):
        """
        Инициализация генератора адаптеров.
        
        Args:
            structural_comparison_path: Путь к результатам структурного сравнения
            semantic_comparison_path: Путь к результатам семантического сравнения (опционально)
        """
        self.structural_comparison_path = structural_comparison_path
        self.semantic_comparison_path = semantic_comparison_path
        self.structural_comparison = {}
        self.semantic_comparison = {}
        self.adapters = {}
    
    def load_comparisons(self) -> None:
        """Загрузка результатов сравнений."""
        if self.structural_comparison_path.exists():
            self.structural_comparison = json.loads(
                self.structural_comparison_path.read_text(encoding="utf-8")
            )
        
        if self.semantic_comparison_path and self.semantic_comparison_path.exists():
            self.semantic_comparison = json.loads(
                self.semantic_comparison_path.read_text(encoding="utf-8")
            )
    
    def generate(self, template1_path: Path, template2_path: Path) -> Dict[str, Any]:
        """
        Генерация адаптеров.
        
        Args:
            template1_path: Путь к первому шаблону
            template2_path: Путь ко второму шаблону
        
        Returns:
            Словарь с адаптерами
        """
        self.adapters = {
            "source_template": self.structural_comparison.get("template1", {}).get("name", ""),
            "target_template": self.structural_comparison.get("template2", {}).get("name", ""),
            "adapters": {},
            "conversion_rules": {},
            "metadata": {
                "generation_date": None,
                "confidence": 0.0
            }
        }
        
        # Генерация адаптеров для листов
        sheet_adapters = self._generate_sheet_adapters(template1_path, template2_path)
        self.adapters["adapters"] = sheet_adapters
        
        # Генерация правил конвертации
        conversion_rules = self._generate_conversion_rules()
        self.adapters["conversion_rules"] = conversion_rules
        
        # Расчет метаданных
        self.adapters["metadata"]["confidence"] = self._calculate_confidence()
        
        return self.adapters
    
    def _generate_sheet_adapters(self, template1_path: Path, template2_path: Path) -> Dict[str, Any]:
        """Генерация адаптеров для листов."""
        adapters = {}
        
        sheet_comparison = self.structural_comparison.get("sheet_comparison", [])
        
        for sheet_info in sheet_comparison:
            sheet_name = sheet_info.get("sheet_name")
            status = sheet_info.get("status")
            
            if status == "common":
                # Адаптер для общего листа
                adapter = self._create_sheet_adapter(
                    template1_path, template2_path, sheet_name, sheet_info
                )
                adapters[sheet_name] = adapter
            
            elif status == "unique_to_template1":
                # Адаптер для уникального листа (пропуск или создание пустого)
                adapters[sheet_name] = {
                    "type": "skip",
                    "reason": "Лист существует только в source template",
                    "target_sheet": None
                }
            
            elif status == "unique_to_template2":
                # Адаптер для нового листа в target
                adapters[sheet_name] = {
                    "type": "create",
                    "reason": "Лист существует только в target template",
                    "source_sheet": None,
                    "target_sheet": sheet_name
                }
        
        return adapters
    
    def _create_sheet_adapter(self, template1_path: Path, template2_path: Path,
                             sheet_name: str, sheet_info: Dict[str, Any]) -> Dict[str, Any]:
        """Создание адаптера для одного листа."""
        adapter = {
            "type": "map",
            "source_sheet": sheet_name,
            "target_sheet": sheet_name,
            "cell_mappings": [],
            "transformations": [],
            "notes": []
        }
        
        # Загрузка шаблонов для анализа ячеек
        try:
            wb1 = load_workbook(template1_path, data_only=True)
            wb2 = load_workbook(template2_path, data_only=True)
            
            if sheet_name in wb1.sheetnames and sheet_name in wb2.sheetnames:
                ws1 = wb1[sheet_name]
                ws2 = wb2[sheet_name]
                
                # Анализ ячеек с данными
                cell_mappings = self._map_cells(ws1, ws2, sheet_name)
                adapter["cell_mappings"] = cell_mappings
                
                # Определение трансформаций
                transformations = self._determine_transformations(ws1, ws2, cell_mappings)
                adapter["transformations"] = transformations
                
        except Exception as e:
            adapter["notes"].append(f"Ошибка при анализе листа: {str(e)}")
        
        return adapter
    
    def _map_cells(self, ws1, ws2, sheet_name: str) -> List[Dict[str, Any]]:
        """Маппинг ячеек между листами."""
        mappings = []
        
        # Простой маппинг: ячейки с одинаковыми координатами
        max_row = min(ws1.max_row, ws2.max_row)
        max_col = min(ws1.max_column, ws2.max_column)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell1 = ws1.cell(row=row, column=col)
                cell2 = ws2.cell(row=row, column=col)
                
                # Если обе ячейки имеют значения
                if cell1.value is not None or cell2.value is not None:
                    mapping = {
                        "source_cell": cell1.coordinate,
                        "target_cell": cell2.coordinate,
                        "source_value": str(cell1.value) if cell1.value is not None else None,
                        "target_value": str(cell2.value) if cell2.value is not None else None,
                        "source_type": cell1.data_type,
                        "target_type": cell2.data_type,
                        "mapping_type": "direct"  # или "semantic", "calculated"
                    }
                    
                    # Определение типа маппинга на основе семантики
                    if self.semantic_comparison:
                        semantic_mapping = self._find_semantic_mapping(
                            sheet_name, cell1.coordinate, cell2.coordinate
                        )
                        if semantic_mapping:
                            mapping["mapping_type"] = "semantic"
                            mapping["semantic_type"] = semantic_mapping.get("semantic_type")
                            mapping["confidence"] = semantic_mapping.get("confidence", 0.0)
                    
                    mappings.append(mapping)
        
        return mappings
    
    def _find_semantic_mapping(self, sheet_name: str, source_cell: str, target_cell: str) -> Optional[Dict[str, Any]]:
        """Поиск семантического маппинга для ячейки."""
        # Поиск в семантическом сравнении
        if not self.semantic_comparison:
            return None
        
        # Поиск в semantic_mapping
        semantic_mapping = self.semantic_comparison.get("semantic_mapping", {})
        
        # Здесь можно добавить логику поиска семантического типа ячейки
        # Пока возвращаем None
        
        return None
    
    def _determine_transformations(self, ws1, ws2, cell_mappings: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Определение необходимых трансформаций."""
        transformations = []
        
        for mapping in cell_mappings:
            source_type = mapping.get("source_type")
            target_type = mapping.get("target_type")
            
            transformation = {
                "source_cell": mapping["source_cell"],
                "target_cell": mapping["target_cell"],
                "transformation_type": "none",
                "formula": None,
                "description": None
            }
            
            # Определение типа трансформации
            if source_type != target_type:
                if source_type == "n" and target_type == "s":
                    transformation["transformation_type"] = "number_to_string"
                    transformation["formula"] = '=TEXT({source_cell}, "0")'
                elif source_type == "s" and target_type == "n":
                    transformation["transformation_type"] = "string_to_number"
                    transformation["formula"] = '=VALUE({source_cell})'
                else:
                    transformation["transformation_type"] = "type_conversion"
            
            # Проверка на формулы
            if mapping.get("source_type") == "f":
                transformation["transformation_type"] = "formula"
                transformation["description"] = "Требуется копирование формулы с адаптацией ссылок"
            
            if transformation["transformation_type"] != "none":
                transformations.append(transformation)
        
        return transformations
    
    def _generate_conversion_rules(self) -> Dict[str, Any]:
        """Генерация правил конвертации."""
        rules = {
            "common_sheets": [],
            "unique_sheets": [],
            "cell_rules": [],
            "semantic_rules": []
        }
        
        # Правила для листов
        sheet_comparison = self.structural_comparison.get("sheet_comparison", [])
        for sheet_info in sheet_comparison:
            if sheet_info.get("status") == "common":
                rules["common_sheets"].append({
                    "sheet_name": sheet_info.get("sheet_name"),
                    "rule": "direct_copy",
                    "confidence": 1.0
                })
            elif sheet_info.get("status") == "unique_to_template1":
                rules["unique_sheets"].append({
                    "sheet_name": sheet_info.get("sheet_name"),
                    "rule": "skip",
                    "reason": "Лист отсутствует в target template"
                })
            elif sheet_info.get("status") == "unique_to_template2":
                rules["unique_sheets"].append({
                    "sheet_name": sheet_info.get("sheet_name"),
                    "rule": "create_empty",
                    "reason": "Новый лист в target template"
                })
        
        # Семантические правила
        if self.semantic_comparison:
            semantic_rules = self.semantic_comparison.get("conversion_rules", [])
            rules["semantic_rules"] = semantic_rules
        
        return rules
    
    def _calculate_confidence(self) -> float:
        """Расчет уверенности в адаптерах."""
        # Базовая уверенность на основе структурного сходства
        similarity = self.structural_comparison.get("similarity_metrics", {}).get("overall_similarity", 0.0)
        
        # Увеличение уверенности при наличии семантического анализа
        if self.semantic_comparison:
            semantic_confidence = len(self.semantic_comparison.get("common_concepts", [])) / max(
                len(self.semantic_comparison.get("common_concepts", [])) + 
                len(self.semantic_comparison.get("unique_to_template1", [])) +
                len(self.semantic_comparison.get("unique_to_template2", [])), 1
            )
            similarity = (similarity + semantic_confidence) / 2
        
        return similarity
    
    def save(self, output_path: Path) -> None:
        """Сохранение адаптеров."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(self.adapters, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def generate_adapters(structural_comparison_path: Path,
                     output_path: Path,
                     template1_path: Path,
                     template2_path: Path,
                     semantic_comparison_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Генерация адаптеров между шаблонами.
    
    Args:
        structural_comparison_path: Путь к результатам структурного сравнения
        output_path: Путь для сохранения адаптеров
        template1_path: Путь к первому шаблону
        template2_path: Путь ко второму шаблону
        semantic_comparison_path: Путь к результатам семантического сравнения (опционально)
    
    Returns:
        Словарь с адаптерами
    """
    generator = AdapterGenerator(structural_comparison_path, semantic_comparison_path)
    generator.load_comparisons()
    adapters = generator.generate(template1_path, template2_path)
    generator.save(output_path)
    
    return adapters


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Генерация адаптеров между шаблонами")
    parser.add_argument("--structural", required=True, help="Путь к структурному сравнению")
    parser.add_argument("--template1", required=True, help="Путь к первому шаблону")
    parser.add_argument("--template2", required=True, help="Путь ко второму шаблону")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    parser.add_argument("--semantic", help="Путь к семантическому сравнению")
    
    args = parser.parse_args()
    
    structural_path = Path(args.structural)
    template1_path = Path(args.template1)
    template2_path = Path(args.template2)
    output_path = Path(args.output)
    semantic_path = Path(args.semantic) if args.semantic else None
    
    if not structural_path.exists():
        raise FileNotFoundError(f"Файл не найден: {structural_path}")
    if not template1_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template1_path}")
    if not template2_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template2_path}")
    
    print("Генерация адаптеров:")
    print(f"  Template 1: {template1_path}")
    print(f"  Template 2: {template2_path}")
    
    adapters = generate_adapters(structural_path, output_path, template1_path, template2_path, semantic_path)
    
    print(f"\n✅ Адаптеры сохранены в: {output_path}")
    print("\n📊 Результаты генерации:")
    print(f"  Уверенность: {adapters['metadata']['confidence']:.2%}")
    print(f"  Адаптеров листов: {len(adapters['adapters'])}")
    print(f"  Правил конвертации: {len(adapters['conversion_rules'].get('semantic_rules', []))}")

