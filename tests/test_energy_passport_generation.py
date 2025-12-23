"""
Тесты для генерации энергопаспорта.
Проверяют корректность заполнения данных, расчетов и соответствие требованиям.
"""
import pytest
from pathlib import Path
from openpyxl import load_workbook
import tempfile
import shutil

# Импортируем функции генерации
import sys
from pathlib import Path as PathLib

_project_root = PathLib(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from services.reports.energy_passport.generator import generate_energy_passport
from services.reports.energy_passport.data_collector import collect_energy_passport_data


@pytest.fixture
def template_path():
    """Путь к шаблону энергопаспорта"""
    template = Path("data/source_files/audit_sinergys/Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx")
    if not template.exists():
        pytest.skip(f"Шаблон не найден: {template}")
    return template


@pytest.fixture
def test_aggregated_data():
    """Тестовые агрегированные данные для Метин Ирода"""
    return {
        "resources": {
            "gas": {
                "2022-Q1": {
                    "year": 2022,
                    "quarter": 1,
                    "months": [
                        {"month": "Январь", "values": {"volume_m3": 2500.0}},
                        {"month": "Февраль", "values": {"volume_m3": 2400.0}},
                        {"month": "Март", "values": {"volume_m3": 2600.0}},
                    ],
                    "quarter_totals": {"volume_m3": 7500.0},
                },
                "2022-Q2": {
                    "year": 2022,
                    "quarter": 2,
                    "months": [
                        {"month": "Апрель", "values": {"volume_m3": 2200.0}},
                        {"month": "Май", "values": {"volume_m3": 2100.0}},
                        {"month": "Июнь", "values": {"volume_m3": 2300.0}},
                    ],
                    "quarter_totals": {"volume_m3": 6600.0},
                },
                "2022-Q3": {
                    "year": 2022,
                    "quarter": 3,
                    "months": [
                        {"month": "Июль", "values": {"volume_m3": 2000.0}},
                        {"month": "Август", "values": {"volume_m3": 1900.0}},
                        {"month": "Сентябрь", "values": {"volume_m3": 2100.0}},
                    ],
                    "quarter_totals": {"volume_m3": 6000.0},
                },
                "2022-Q4": {
                    "year": 2022,
                    "quarter": 4,
                    "months": [
                        {"month": "Октябрь", "values": {"volume_m3": 2400.0}},
                        {"month": "Ноябрь", "values": {"volume_m3": 2500.0}},
                        {"month": "Декабрь", "values": {"volume_m3": 2600.0}},
                    ],
                    "quarter_totals": {"volume_m3": 7500.0},
                },
                "2023-Q1": {
                    "year": 2023,
                    "quarter": 1,
                    "months": [
                        {"month": "Январь", "values": {"volume_m3": 4800.0}},
                        {"month": "Февраль", "values": {"volume_m3": 5100.0}},
                        {"month": "Март", "values": {"volume_m3": 4919.0}},
                    ],
                    "quarter_totals": {"volume_m3": 14819.0},  # Должно быть 14.819 тыс. м³
                },
            },
            "electricity": {
                "2022-Q1": {
                    "year": 2022,
                    "quarter": 1,
                    "quarter_totals": {"active_kwh": 100000.0},
                },
            },
        }
    }


@pytest.fixture
def test_enterprise_data():
    """Тестовые данные предприятия"""
    return {
        "id": "metin-iroda",
        "name": "Метин Ирода",
        "inn": "123456789",
    }


@pytest.fixture
def test_building_data():
    """Тестовые данные о здании"""
    return {
        "area_m2": 5000.0,  # 5000 м²
    }


def test_gas_calculation_from_monthly_data(template_path, test_aggregated_data, test_enterprise_data, test_building_data):
    """Тест: проверка правильного расчета газа из помесячных данных"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "test_passport.xlsx"
        
        # Генерируем паспорт
        generate_energy_passport(
            enterprise_id=test_enterprise_data["id"],
            year=2023,
            template_path=template_path,
            output_path=output_path,
            aggregated_data=test_aggregated_data,
            enterprise_data=test_enterprise_data,
            building_data=test_building_data,
        )
        
        # Проверяем результат
        assert output_path.exists(), "Паспорт должен быть создан"
        
        wb = load_workbook(output_path, data_only=True)
        ws = wb["Структура пр 2"]
        
        # Проверяем E32 (2023 Q1 газ) - должно быть 14.819 тыс. м³
        # Сумма месяцев: 4800 + 5100 + 4919 = 14819 м³ = 14.819 тыс. м³
        e32_value = ws["E32"].value
        assert e32_value is not None, "E32 должна быть заполнена"
        assert abs(e32_value - 14.819) < 0.001, f"E32 должна быть 14.819, получено {e32_value}"
        
        wb.close()


def test_gas_yearly_totals(template_path, test_aggregated_data, test_enterprise_data, test_building_data):
    """Тест: проверка годовых итогов по газу"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "test_passport.xlsx"
        
        # Добавляем данные за все кварталы 2022 года
        # Итого: 7500 + 6600 + 6000 + 7500 = 27600 м³ = 27.6 тыс. м³
        
        generate_energy_passport(
            enterprise_id=test_enterprise_data["id"],
            year=2022,
            template_path=template_path,
            output_path=output_path,
            aggregated_data=test_aggregated_data,
            enterprise_data=test_enterprise_data,
            building_data=test_building_data,
        )
        
        assert output_path.exists()
        
        wb = load_workbook(output_path, data_only=True)
        ws = wb["Структура пр 2"]
        
        # Проверяем, что квартальные суммы правильные
        # Q1: 7500 м³ = 7.5 тыс. м³
        # Q2: 6600 м³ = 6.6 тыс. м³
        # Q3: 6000 м³ = 6.0 тыс. м³
        # Q4: 7500 м³ = 7.5 тыс. м³
        
        # TODO: Найти ячейки для кварталов и проверить значения
        
        wb.close()


def test_electricity_by_product(template_path, test_aggregated_data, test_enterprise_data, test_building_data):
    """Тест: проверка заполнения электроэнергии по видам продукции"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "test_passport.xlsx"
        
        generate_energy_passport(
            enterprise_id=test_enterprise_data["id"],
            year=2023,
            template_path=template_path,
            output_path=output_path,
            aggregated_data=test_aggregated_data,
            enterprise_data=test_enterprise_data,
            building_data=test_building_data,
        )
        
        assert output_path.exists()
        
        wb = load_workbook(output_path, data_only=True)
        ws = wb["Структура пр 2"]
        
        # Проверяем заполнение данных по продукции
        # Строка 17: Трубы ХВС, норма=630, факт 2022=657
        # TODO: Проверить конкретные ячейки
        
        wb.close()


def test_gas_specific_consumption(template_path, test_aggregated_data, test_enterprise_data, test_building_data):
    """Тест: проверка расчета удельного расхода газа"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "test_passport.xlsx"
        
        generate_energy_passport(
            enterprise_id=test_enterprise_data["id"],
            year=2022,
            template_path=template_path,
            output_path=output_path,
            aggregated_data=test_aggregated_data,
            enterprise_data=test_enterprise_data,
            building_data=test_building_data,
        )
        
        assert output_path.exists()
        
        wb = load_workbook(output_path, data_only=True)
        
        # Проверяем лист "Удельный расход газа"
        if "Удельный расход газа" in wb.sheetnames:
            ws = wb["Удельный расход газа"]
            
            # Годовой газ 2022: 27600 м³
            # Площадь: 5000 м²
            # Удельный расход: 27600 / 5000 = 5.52 м³/(м²·год)
            
            fact_per_m2 = ws["C5"].value
            assert fact_per_m2 is not None, "Удельный расход на м² должен быть рассчитан"
            expected = 27600.0 / 5000.0  # 5.52
            assert abs(fact_per_m2 - expected) < 0.01, f"Ожидалось {expected}, получено {fact_per_m2}"
        
        wb.close()


def test_collect_energy_passport_data(test_aggregated_data):
    """Тест: проверка сбора данных для паспорта"""
    passport_data = collect_energy_passport_data(
        enterprise_id="test-enterprise",
        year=2023,
        aggregated_data=test_aggregated_data,
    )
    
    assert passport_data.year == 2023
    assert passport_data.enterprise_id == "test-enterprise"
    assert passport_data.gas_data is not None
    assert len(passport_data.electricity_by_product) == 5
    
    # Проверяем данные по газу
    assert 2023 in passport_data.gas_data.yearly
    assert 2023 in passport_data.gas_data.quarterly
    
    # Проверяем квартальные данные 2023-Q1
    assert passport_data.gas_data.quarterly[2023][1] == 14819.0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

