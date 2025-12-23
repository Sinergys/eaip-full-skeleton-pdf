"""Простой тест заполнения электроэнергии"""
import sys
from pathlib import Path

# Добавляем путь к tools
sys.path.insert(0, str(Path(__file__).parent / "tools"))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Проверяем сгенерированный файл
output_path = Path("test_output_metin/generated_passport_fixed.xlsx")

if not output_path.exists():
    print(f"Файл не найден: {output_path}")
    sys.exit(1)

wb = load_workbook(output_path, data_only=True)
ws = wb["Структура пр 2"]

print("Проверка Q1 (строка 17):")
print(f"  B17: {ws['B17'].value}")
print(f"  C17: {ws['C17'].value}")

print("\nПроверка Q2 (строка 39):")
print(f"  Q39: {ws['Q39'].value}")
print(f"  R39: {ws['R39'].value}")

print("\nПроверка Q3 (строка 61):")
print(f"  AF61: {ws['AF61'].value}")
print(f"  AG61: {ws['AG61'].value}")

print("\nПроверка Q4 (строка 83):")
print(f"  AU83: {ws['AU83'].value}")
print(f"  AV83: {ws['AV83'].value}")

wb.close()

