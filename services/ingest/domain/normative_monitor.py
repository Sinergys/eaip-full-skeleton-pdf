"""
Модуль мониторинга критических полей энергопаспорта
Автоматическая проверка всех критических полей и генерация отчета
"""
import logging
from typing import Dict, Any, Optional, List
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Импорт функций проверки
try:
    from domain.normative_validator import validate_against_normative
    from domain.normative_integration import validate_and_log_critical_field, NormativeValidationResult

    HAS_VALIDATOR = True
except ImportError:
    HAS_VALIDATOR = False
    logger.warning("normative_validator модуль не найден. Мониторинг недоступен.")

# Импорт database
try:
    import database

    HAS_DATABASE = True
except ImportError:
    HAS_DATABASE = False
    logger.warning("database модуль не найден.")


# Список критических полей для мониторинга
CRITICAL_FIELDS = [
    {
        "field_name": "Удельный расход",
        "sheet_name": "Динамика ср",
        "column": 7,  # Колонка G
        "row_start": 2,  # Начало данных
        "tolerance_percent": 10.0,
        "description": "Удельный расход электроэнергии на единицу продукции",
    },
    {
        "field_name": "Удельный расход по кварталам",
        "sheet_name": "Расход на ед.п",
        "column": None,  # Динамические колонки (C, H, M для разных лет)
        "row_start": 5,
        "tolerance_percent": 10.0,
        "description": "Удельный расход по кварталам",
    },
    # Можно добавить больше полей
]


def read_field_value_from_passport(
    passport_path: str,
    field_name: str,
    sheet_name: str,
    column: Optional[int] = None,
    row: Optional[int] = None,
) -> Optional[float]:
    """
    Прочитать значение поля из энергопаспорта

    Args:
        passport_path: Путь к файлу паспорта
        field_name: Название поля
        sheet_name: Имя листа
        column: Номер колонки (если известен)
        row: Номер строки (если известен)

    Returns:
        Значение поля или None
    """
    try:
        workbook = load_workbook(passport_path, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Лист '{sheet_name}' не найден в паспорте")
            return None

        sheet = workbook[sheet_name]

        # Если указаны колонка и строка - читаем напрямую
        if column and row:
            cell = sheet.cell(row=row, column=column)
            value = cell.value
            if isinstance(value, (int, float)):
                return float(value)
            return None

        # Иначе ищем по названию поля в заголовках
        # Простой поиск по первой строке
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col_idx).value
            if header and field_name.lower() in str(header).lower():
                # Ищем первое числовое значение в этой колонке
                for row_idx in range(2, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, (int, float)):
                        workbook.close()
                        return float(cell_value)

        workbook.close()
        return None

    except Exception as e:
        logger.error(f"Ошибка чтения поля {field_name} из паспорта: {e}")
        return None


def monitor_critical_fields_from_passport(
    passport_path: str,
    enterprise_id: Optional[int] = None,
    batch_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Мониторинг всех критических полей из энергопаспорта

    Args:
        passport_path: Путь к файлу энергопаспорта
        enterprise_id: ID предприятия
        batch_id: ID загрузки

    Returns:
        Словарь с результатами мониторинга
    """
    if not HAS_VALIDATOR:
        return {
            "error": "Модуль проверки нормативов недоступен",
            "fields_checked": 0,
            "violations": [],
            "compliant": [],
            "unknown": [],
        }

    violations = []
    compliant = []
    unknown = []

    passport_file = Path(passport_path)
    if not passport_file.exists():
        return {
            "error": f"Файл паспорта не найден: {passport_path}",
            "fields_checked": 0,
            "violations": [],
            "compliant": [],
            "unknown": [],
        }

    # Проверяем каждое критическое поле
    for field_config in CRITICAL_FIELDS:
        field_name = field_config["field_name"]
        sheet_name = field_config["sheet_name"]
        column = field_config.get("column")
        row_start = field_config.get("row_start", 2)
        tolerance = field_config.get("tolerance_percent", 10.0)

        # Пытаемся прочитать значение из паспорта
        # Для упрощения проверяем первую строку данных
        actual_value = read_field_value_from_passport(
            passport_path,
            field_name,
            sheet_name,
            column=column,
            row=row_start,
        )

        if actual_value is None:
            unknown.append({
                "field_name": field_name,
                "sheet_name": sheet_name,
                "reason": "Значение не найдено в паспорте",
            })
            continue

        # Проверяем соответствие нормативу
        try:
            validation = validate_against_normative(
                actual_value=actual_value,
                field_name=field_name,
                sheet_name=sheet_name,
                tolerance_percent=tolerance,
            )

            result = {
                "field_name": field_name,
                "sheet_name": sheet_name,
                "actual_value": actual_value,
                "status": validation.get("status", "unknown"),
                "normative_value": validation.get("normative"),
                "deviation_percent": validation.get("deviation_percent", 0.0),
                "message": validation.get("message", ""),
            }

            if validation.get("status") == "violation":
                violations.append(result)
            elif validation.get("status") == "compliant":
                compliant.append(result)
            else:
                unknown.append(result)

        except Exception as e:
            logger.error(f"Ошибка проверки поля {field_name}: {e}")
            unknown.append({
                "field_name": field_name,
                "sheet_name": sheet_name,
                "reason": f"Ошибка проверки: {e}",
            })

    return {
        "passport_path": str(passport_path),
        "enterprise_id": enterprise_id,
        "batch_id": batch_id,
        "fields_checked": len(CRITICAL_FIELDS),
        "violations_count": len(violations),
        "compliant_count": len(compliant),
        "unknown_count": len(unknown),
        "violations": violations,
        "compliant": compliant,
        "unknown": unknown,
        "status": "has_violations" if len(violations) > 0 else "compliant",
    }


def get_monitoring_summary(enterprise_id: Optional[int] = None) -> Dict[str, Any]:
    """
    Получить сводку мониторинга для предприятия

    Args:
        enterprise_id: ID предприятия

    Returns:
        Сводка мониторинга
    """
    if not HAS_DATABASE:
        return {
            "error": "База данных недоступна",
            "total_violations": 0,
            "recent_violations": [],
        }

    try:
        # Получаем последние нарушения
        violations = database.get_normative_violations(
            enterprise_id=enterprise_id,
            status="violation",
            limit=20,
        )

        # Группируем по полям
        fields_summary = {}
        for violation in violations:
            field_key = f"{violation.get('field_name')}_{violation.get('sheet_name')}"
            if field_key not in fields_summary:
                fields_summary[field_key] = {
                    "field_name": violation.get("field_name"),
                    "sheet_name": violation.get("sheet_name"),
                    "count": 0,
                    "max_deviation": 0.0,
                    "latest": None,
                }
            
            fields_summary[field_key]["count"] += 1
            deviation = violation.get("deviation_percent", 0.0)
            if deviation > fields_summary[field_key]["max_deviation"]:
                fields_summary[field_key]["max_deviation"] = deviation
            
            if not fields_summary[field_key]["latest"]:
                fields_summary[field_key]["latest"] = violation

        return {
            "total_violations": len(violations),
            "fields_with_violations": len(fields_summary),
            "fields_summary": list(fields_summary.values()),
            "recent_violations": violations[:10],
        }

    except Exception as e:
        logger.error(f"Ошибка получения сводки мониторинга: {e}")
        return {
            "error": str(e),
            "total_violations": 0,
            "recent_violations": [],
        }

