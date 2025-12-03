"""
Парсер энергетических паспортов Excel
Согласно Постановлению КМ РУз №690

Автор: Energy Audit System
Дата: 26 октября 2025
"""

from openpyxl import load_workbook
from typing import Dict, List, Optional, Any
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelPassportParser:
    """Парсер энергетических паспортов из Excel файлов"""

    def __init__(self, file_path: str):
        """
        Инициализация парсера

        Args:
            file_path: Путь к файлу энергопаспорта
        """
        self.file_path = file_path
        self.workbook = None
        self.data = {}

    def parse(self) -> Dict[str, Any]:
        """
        Главная функция парсинга

        Returns:
            Словарь с извлеченными данными
        """
        try:
            logger.info(f"Начало парсинга: {self.file_path}")

            # Загрузка файла
            self.workbook = load_workbook(self.file_path, data_only=False)

            # Извлечение данных из разных листов
            self.data = {
                "file_path": self.file_path,
                "parsed_at": datetime.now().isoformat(),
                "enterprise": self._parse_enterprise_info(),
                "energy_resources": self._parse_energy_resources(),
                "electricity": self._parse_electricity(),
                "gas": self._parse_gas(),
                "water": self._parse_water(),
                "fuel": self._parse_fuel(),
                "equipment": self._parse_equipment(),
                "buildings": self._parse_buildings(),
                "calculations": self._parse_calculations(),
                "measures": self._parse_measures(),
                "formulas": self._extract_formulas(),
            }

            logger.info(f"Парсинг завершен успешно: {len(self.data)} разделов")
            return self.data

        except Exception as e:
            logger.error(f"Ошибка парсинга файла {self.file_path}: {e}")
            raise

        finally:
            if self.workbook:
                self.workbook.close()

    def _parse_enterprise_info(self) -> Dict[str, Any]:
        """Извлечение общей информации о предприятии"""
        try:
            # Обычно на первом листе или листе "Паспорт"
            sheet = self._get_sheet(["Паспорт", "Титульный", "Sheet1", 0])

            info = {
                "name": self._find_cell_value(
                    sheet, ["Название", "Наименование", "предприятие"]
                ),
                "inn": self._find_cell_value(sheet, ["ИНН", "INN"]),
                "address": self._find_cell_value(sheet, ["Адрес", "адрес"]),
                "industry": self._find_cell_value(sheet, ["Отрасль", "отрасль"]),
                "director": self._find_cell_value(
                    sheet, ["Директор", "директор", "руководитель"]
                ),
                "auditor": self._find_cell_value(
                    sheet, ["Аудитор", "аудитор", "энергоаудитор"]
                ),
                "reporting_year": self._find_cell_value(
                    sheet, ["Год", "отчетный год", "период"]
                ),
            }

            logger.info(f"Предприятие: {info.get('name', 'Не найдено')}")
            return info

        except Exception as e:
            logger.error(f"Ошибка извлечения информации о предприятии: {e}")
            return {}

    def _parse_energy_resources(self) -> Dict[str, Any]:
        """Извлечение данных по энергоресурсам (сводная таблица)"""
        try:
            sheet = self._get_sheet(["Энергоресурсы", "Ресурсы", "Баланс"])

            resources = {
                "electricity": {
                    "consumption_kwh": self._find_numeric_value(
                        sheet, ["Электроэнергия", "кВт"]
                    ),
                    "cost_uzs": self._find_numeric_value(
                        sheet, ["Электроэнергия", "сум"]
                    ),
                },
                "gas": {
                    "consumption_m3": self._find_numeric_value(
                        sheet, ["Газ", "м3", "м³"]
                    ),
                    "cost_uzs": self._find_numeric_value(sheet, ["Газ", "сум"]),
                },
                "water": {
                    "consumption_m3": self._find_numeric_value(
                        sheet, ["Вода", "м3", "м³"]
                    ),
                    "cost_uzs": self._find_numeric_value(sheet, ["Вода", "сум"]),
                },
                "heat": {
                    "consumption_gcal": self._find_numeric_value(
                        sheet, ["Тепло", "Гкал"]
                    ),
                    "cost_uzs": self._find_numeric_value(sheet, ["Тепло", "сум"]),
                },
            }

            return resources

        except Exception as e:
            logger.warning(f"Ошибка извлечения энергоресурсов: {e}")
            return {}

    def _parse_electricity(self) -> Dict[str, Any]:
        """Детальные данные по электроэнергии"""
        try:
            sheet = self._get_sheet(["Электроэнергия", "Электричество", "ТП"])

            electricity = {
                "total_consumption": self._find_numeric_value(
                    sheet, ["Общее потребление", "Всего"]
                ),
                "transformers": self._parse_transformers(sheet),
                "monthly_data": self._parse_monthly_data(sheet, "кВт"),
                "power_factor": self._find_numeric_value(
                    sheet, ["cos", "коэффициент мощности"]
                ),
                "losses": self._find_numeric_value(sheet, ["потери", "Потери"]),
            }

            return electricity

        except Exception as e:
            logger.warning(f"Ошибка извлечения данных по электроэнергии: {e}")
            return {}

    def _parse_gas(self) -> Dict[str, Any]:
        """Детальные данные по газу"""
        try:
            sheet = self._get_sheet(["Газ", "Газоснабжение"])

            gas = {
                "natural_gas_m3": self._find_numeric_value(
                    sheet, ["Природный газ", "м3"]
                ),
                "liquified_gas_kg": self._find_numeric_value(
                    sheet, ["Сжиженный газ", "кг"]
                ),
                "heating_m3": self._find_numeric_value(sheet, ["Отопление", "м3"]),
                "technology_m3": self._find_numeric_value(
                    sheet, ["Технологические нужды"]
                ),
                "monthly_data": self._parse_monthly_data(sheet, "м3"),
            }

            return gas

        except Exception as e:
            logger.warning(f"Ошибка извлечения данных по газу: {e}")
            return {}

    def _parse_water(self) -> Dict[str, Any]:
        """Детальные данные по воде"""
        try:
            sheet = self._get_sheet(["Вода", "Водоснабжение"])

            water = {
                "cold_water_m3": self._find_numeric_value(
                    sheet, ["Холодная вода", "холодн"]
                ),
                "hot_water_m3": self._find_numeric_value(
                    sheet, ["Горячая вода", "горяч"]
                ),
                "wastewater_m3": self._find_numeric_value(
                    sheet, ["Сточные воды", "стоки"]
                ),
                "monthly_data": self._parse_monthly_data(sheet, "м3"),
            }

            return water

        except Exception as e:
            logger.warning(f"Ошибка извлечения данных по воде: {e}")
            return {}

    def _parse_fuel(self) -> Dict[str, Any]:
        """Данные по ГСМ"""
        try:
            sheet = self._get_sheet(["ГСМ", "Топливо", "Fuel"])

            fuel = {
                "petrol_liters": self._find_numeric_value(sheet, ["Бензин", "л"]),
                "diesel_liters": self._find_numeric_value(sheet, ["Дизель", "ДТ"]),
                "total_cost": self._find_numeric_value(sheet, ["Стоимость", "сум"]),
            }

            return fuel

        except Exception as e:
            logger.warning(f"Ошибка извлечения данных по ГСМ: {e}")
            return {}

    def _parse_equipment(self) -> List[Dict[str, Any]]:
        """Список оборудования"""
        try:
            sheet = self._get_sheet(["Оборудование", "Equipment"])

            equipment_list = []

            # Поиск таблицы с оборудованием
            for row in sheet.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    if cell.value and any(
                        keyword in str(cell.value).lower()
                        for keyword in ["наименование", "название", "оборудование"]
                    ):
                        # Найдена шапка таблицы
                        start_row = cell.row + 1
                        equipment_list = self._extract_table(
                            sheet, start_row, ["название", "мощность", "год", "статус"]
                        )
                        break
                if equipment_list:
                    break

            return equipment_list

        except Exception as e:
            logger.warning(f"Ошибка извлечения оборудования: {e}")
            return []

    def _parse_buildings(self) -> List[Dict[str, Any]]:
        """Данные по зданиям"""
        try:
            sheet = self._get_sheet(["Здания", "Buildings", "Объекты"])

            buildings = []

            # Поиск таблицы со зданиями
            for row in sheet.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    if cell.value and any(
                        keyword in str(cell.value).lower()
                        for keyword in ["здание", "объект", "площадь"]
                    ):
                        start_row = cell.row + 1
                        buildings = self._extract_table(
                            sheet, start_row, ["название", "площадь", "этажность"]
                        )
                        break
                if buildings:
                    break

            return buildings

        except Exception as e:
            logger.warning(f"Ошибка извлечения данных по зданиям: {e}")
            return []

    def _parse_calculations(self) -> Dict[str, Any]:
        """Расчетные показатели"""
        try:
            sheet = self._get_sheet(["Расчеты", "Показатели", "Calculations"])

            calculations = {
                "specific_consumption": self._find_numeric_value(
                    sheet, ["Удельный расход"]
                ),
                "energy_class": self._find_cell_value(
                    sheet, ["Энергетический класс", "класс"]
                ),
                "power_factor": self._find_numeric_value(
                    sheet, ["cos φ", "коэффициент мощности"]
                ),
                "losses_percent": self._find_numeric_value(sheet, ["Потери", "%"]),
                "overconsumption": self._find_numeric_value(sheet, ["Перерасход"]),
            }

            return calculations

        except Exception as e:
            logger.warning(f"Ошибка извлечения расчетов: {e}")
            return {}

    def _parse_measures(self) -> List[Dict[str, Any]]:
        """Мероприятия по энергосбережению"""
        try:
            sheet = self._get_sheet(["Мероприятия", "Меры", "Measures"])

            measures = []

            # Поиск таблицы с мероприятиями
            for row in sheet.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    if cell.value and any(
                        keyword in str(cell.value).lower()
                        for keyword in ["мероприятие", "название", "экономия"]
                    ):
                        start_row = cell.row + 1
                        measures = self._extract_table(
                            sheet,
                            start_row,
                            ["название", "экономия", "затраты", "срок"],
                        )
                        break
                if measures:
                    break

            return measures

        except Exception as e:
            logger.warning(f"Ошибка извлечения мероприятий: {e}")
            return []

    def _extract_formulas(self) -> List[Dict[str, str]]:
        """Извлечение формул из ячеек"""
        formulas = []

        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == "f":  # Formula
                            formulas.append(
                                {
                                    "sheet": sheet_name,
                                    "cell": cell.coordinate,
                                    "formula": cell.value,
                                    "result": cell.internal_value
                                    if hasattr(cell, "internal_value")
                                    else None,
                                }
                            )

            logger.info(f"Извлечено формул: {len(formulas)}")

        except Exception as e:
            logger.warning(f"Ошибка извлечения формул: {e}")

        return formulas

    # ========== Вспомогательные методы ==========

    def _get_sheet(self, names: List):
        """Получить лист по списку возможных имен"""
        for name in names:
            if isinstance(name, int):
                return self.workbook.worksheets[name]
            if name in self.workbook.sheetnames:
                return self.workbook[name]

        # Вернуть первый лист если не нашли
        return self.workbook.worksheets[0]

    def _find_cell_value(self, sheet, keywords: List[str]) -> Optional[str]:
        """Найти значение ячейки по ключевым словам"""
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value).lower()
                    for keyword in keywords:
                        if keyword.lower() in cell_text:
                            # Значение обычно в следующей ячейке справа или снизу
                            if cell.offset(column=1).value:
                                return str(cell.offset(column=1).value)
                            elif cell.offset(row=1).value:
                                return str(cell.offset(row=1).value)
        return None

    def _find_numeric_value(self, sheet, keywords: List[str]) -> Optional[float]:
        """Найти числовое значение"""
        value = self._find_cell_value(sheet, keywords)
        if value:
            try:
                # Убрать пробелы и заменить запятую на точку
                cleaned = value.replace(" ", "").replace(",", ".")
                return float(cleaned)
            except (ValueError, AttributeError, TypeError):
                # Если не удалось преобразовать в число, игнорируем
                pass
        return None

    def _parse_monthly_data(self, sheet, unit: str) -> List[Dict[str, Any]]:
        """Извлечь помесячные данные"""
        monthly = []
        months = [
            "январь",
            "февраль",
            "март",
            "апрель",
            "май",
            "июнь",
            "июль",
            "август",
            "сентябрь",
            "октябрь",
            "ноябрь",
            "декабрь",
        ]

        for month in months:
            value = self._find_numeric_value(sheet, [month])
            if value:
                monthly.append(
                    {"month": month.capitalize(), "value": value, "unit": unit}
                )

        return monthly

    def _parse_transformers(self, sheet) -> List[Dict[str, Any]]:
        """Извлечь данные по трансформаторным подстанциям"""
        transformers = []

        for tp_num in [1, 2, 3]:
            tp_name = f"ТП-{tp_num}"
            consumption = self._find_numeric_value(sheet, [tp_name, f"ТП {tp_num}"])
            if consumption:
                transformers.append({"name": tp_name, "consumption_kwh": consumption})

        return transformers

    def _extract_table(
        self, sheet, start_row: int, column_keywords: List[str]
    ) -> List[Dict[str, Any]]:
        """Извлечь данные из таблицы"""
        table_data = []

        # Найти колонки по ключевым словам
        header_row = sheet[start_row - 1]
        columns = {}

        for idx, cell in enumerate(header_row):
            if cell.value:
                cell_text = str(cell.value).lower()
                for keyword in column_keywords:
                    if keyword.lower() in cell_text:
                        columns[keyword] = idx

        # Извлечь строки
        for row in sheet.iter_rows(min_row=start_row, max_row=start_row + 100):
            row_data = {}
            has_data = False

            for keyword, col_idx in columns.items():
                if col_idx < len(row) and row[col_idx].value:
                    row_data[keyword] = row[col_idx].value
                    has_data = True

            if has_data:
                table_data.append(row_data)
            elif len(table_data) > 0:
                # Пустая строка после данных - конец таблицы
                break

        return table_data


# Функция для быстрого использования
def parse_passport_file(file_path: str) -> Dict[str, Any]:
    """
    Быстрая функция парсинга энергопаспорта

    Args:
        file_path: Путь к файлу

    Returns:
        Словарь с данными
    """
    parser = ExcelPassportParser(file_path)
    return parser.parse()


if __name__ == "__main__":
    # Тест парсера
    import sys

    logging.basicConfig(level=logging.INFO)

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"Парсинг файла: {file_path}")

        try:
            data = parse_passport_file(file_path)
            print("\n✅ Успешно извлечено:")
            print(f"   - Предприятие: {data.get('enterprise', {}).get('name')}")
            print(f"   - Формул: {len(data.get('formulas', []))}")
            print(f"   - Оборудование: {len(data.get('equipment', []))}")
            print(f"   - Здания: {len(data.get('buildings', []))}")
            print(f"   - Мероприятия: {len(data.get('measures', []))}")
        except Exception as e:
            print(f"\n❌ Ошибка: {e}")
    else:
        print("Использование: python excel_passport_parser.py <путь_к_файлу>")
