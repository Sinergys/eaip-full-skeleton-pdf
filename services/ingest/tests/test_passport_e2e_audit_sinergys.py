"""
E2E тест полного потока генерации энергопаспорта для audit_sinergys.

Тест выполняет:
1. Загрузку всех необходимых файлов из data/source_files/audit_sinergys
2. Создание нового batch через загрузку файлов
3. Проверку readiness
4. Генерацию паспорта с шаблоном new_energy_passport
5. Проверку структуры, формул и данных в итоговом файле

Используемые файлы:
- pererashod.xlsx - электроэнергия
- gaz.xlsx - газ
- voda.xlsx - вода
- otoplenie.xlsx - тепло
- oborudovanie.xlsx - оборудование
- schetchiki.xlsx - узлы учета
- teploprovodnost.xlsx - ограждающие конструкции

Сценарий: METIN IRODA (полный набор данных)
"""

import sys
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional

import pytest
from openpyxl import load_workbook

# Добавляем путь к модулям
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "eaip_full_skeleton" / "services" / "ingest"))

# Импортируем app после добавления пути
from main import app

# Используем starlette.testclient напрямую
# Проблема совместимости версий httpx/starlette решается использованием правильного импорта
from starlette.testclient import TestClient

# Путь к тестовым файлам
# Файлы находятся в корне проекта, а не в eaip_full_skeleton
# Пробуем несколько вариантов путей
if (PROJECT_ROOT / "data" / "source_files" / "audit_sinergys").exists():
    TEST_FILES_DIR = PROJECT_ROOT / "data" / "source_files" / "audit_sinergys"
elif (PROJECT_ROOT.parent / "data" / "source_files" / "audit_sinergys").exists():
    TEST_FILES_DIR = PROJECT_ROOT.parent / "data" / "source_files" / "audit_sinergys"
else:
    # Последний вариант: абсолютный путь
    TEST_FILES_DIR = Path("c:/eaip/data/source_files/audit_sinergys")

# Маппинг файлов к типам ресурсов
# Примечание: resource_type используется для указания типа ресурса при загрузке
# Если файл не соответствует стандартным типам, можно оставить None
TEST_FILES = {
    "electro act react.xlsx": {
        "resource_type": "electricity",
        "enterprise_name": "METIN IRODA",
    },  # Месячные данные по электроэнергии
    "pererashod.xlsx": {
        "resource_type": "electricity",
        "enterprise_name": "METIN IRODA",
    },  # Категории использования (by_usage)
    "gaz.xlsx": {"resource_type": "gas", "enterprise_name": "METIN IRODA"},
    "voda.xlsx": {"resource_type": "water", "enterprise_name": "METIN IRODA"},
    "otoplenie.xlsx": {"resource_type": "heat", "enterprise_name": "METIN IRODA"},
    "oborudovanie.xlsx": {
        "resource_type": None,
        "enterprise_name": "METIN IRODA",
    },  # Оборудование определяется автоматически
    "schetchiki.xlsx": {
        "resource_type": None,
        "enterprise_name": "METIN IRODA",
    },  # Узлы учета определяются автоматически
    "teploprovodnost.xlsx": {
        "resource_type": None,
        "enterprise_name": "METIN IRODA",
    },  # Ограждающие конструкции определяются автоматически
}


def get_test_files() -> List[Path]:
    """Собирает список файлов для теста из TEST_FILES_DIR"""
    files = []
    missing_files = []
    for filename in TEST_FILES.keys():
        file_path = TEST_FILES_DIR / filename
        if file_path.exists():
            files.append(file_path)
        else:
            missing_files.append(str(file_path))

    if missing_files:
        pytest.skip(f"Тестовые файлы не найдены: {', '.join(missing_files)}")

    return files


def upload_file(
    client: TestClient,
    file_path: Path,
    enterprise_name: str,
    resource_type: Optional[str] = None,
) -> Dict:
    """Загружает файл через API и возвращает ответ"""
    with open(file_path, "rb") as f:
        files = {
            "file": (
                file_path.name,
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        data = {
            "enterprise_name": enterprise_name,
        }
        if resource_type:
            data["resource_type"] = resource_type

        response = client.post("/web/upload", files=files, data=data)
        assert response.status_code in [200, 201], (
            f"Ошибка загрузки файла {file_path.name}: {response.status_code} - {response.text}"
        )

        result = response.json()
        # Если файл уже был загружен (дубликат), возвращается существующий batch_id
        # Это нормально для теста, продолжаем с этим batch_id
        return result


def wait_for_parsing(client: TestClient, batch_id: str, timeout: int = 300) -> bool:
    """Ожидает завершения парсинга файла"""
    start_time = time.time()
    while time.time() - start_time < timeout:
        response = client.get(f"/api/progress/{batch_id}")
        if response.status_code == 200:
            data = response.json()
            if data.get("is_completed"):
                if data.get("has_error"):
                    pytest.fail(
                        f"Ошибка при парсинге batch {batch_id}: {data.get('error')}"
                    )
                return True
        time.sleep(2)
    pytest.fail(f"Таймаут ожидания парсинга для batch {batch_id}")


def get_enterprise_id_from_batch(client: TestClient, batch_id: str) -> int:
    """Получает enterprise_id из batch"""
    response = client.get(f"/api/uploads/{batch_id}")
    assert response.status_code == 200, f"Ошибка получения batch: {response.text}"
    data = response.json()
    return data["enterprise_id"]


@pytest.fixture(scope="function")
def client():
    """Создает TestClient для теста"""
    return TestClient(app)


@pytest.fixture(scope="function")
def uploaded_batches(client: TestClient) -> List[str]:
    """Загружает все файлы и возвращает список batch_id"""
    batches = []
    enterprise_name = "METIN IRODA"

    test_files = get_test_files()
    if len(test_files) == 0:
        pytest.skip("Тестовые файлы не найдены в data/source_files/audit_sinergys")

    print(f"\n{'=' * 80}")
    print("ЗАГРУЗКА ФАЙЛОВ ДЛЯ E2E ТЕСТА")
    print(f"{'=' * 80}")

    for file_path in test_files:
        filename = file_path.name
        file_info = TEST_FILES.get(filename, {})
        resource_type = file_info.get("resource_type")

        print(f"\n📤 Загрузка: {filename}")
        if resource_type:
            print(f"   Тип ресурса: {resource_type}")

        response_data = upload_file(client, file_path, enterprise_name, resource_type)
        batch_id = response_data.get("batch_id")

        assert batch_id, f"Batch ID не получен для файла {filename}"
        print(f"   ✅ Batch ID: {batch_id}")

        # Ждем завершения парсинга
        print("   ⏳ Ожидание парсинга...")
        wait_for_parsing(client, batch_id)
        print("   ✅ Парсинг завершен")

        batches.append(batch_id)

    print(f"\n✅ Всего загружено файлов: {len(batches)}")
    return batches


def test_passport_generation_e2e(client: TestClient, uploaded_batches: List[str]):
    """
    Полный E2E тест генерации энергопаспорта.

    Шаги:
    1. Проверка readiness для предприятия
    2. Генерация паспорта с шаблоном new_energy_passport
    3. Проверка структуры файла
    4. Проверка формул
    5. Проверка данных
    """
    # Получаем enterprise_id из первого batch
    enterprise_id = get_enterprise_id_from_batch(client, uploaded_batches[0])

    print(f"\n{'=' * 80}")
    print("E2E ТЕСТ ГЕНЕРАЦИИ ЭНЕРГОПАСПОРТА")
    print(f"{'=' * 80}")
    print(f"Enterprise ID: {enterprise_id}")
    print(f"Загружено batches: {len(uploaded_batches)}")

    # Шаг 1: Проверка readiness
    print("\n📋 Шаг 1: Проверка readiness...")
    response = client.get(f"/api/enterprises/{enterprise_id}/generation-readiness")
    assert response.status_code == 200, f"Ошибка проверки readiness: {response.text}"

    readiness_data = response.json()
    overall_status = readiness_data.get("overall_status", "unknown")
    print(f"   Overall status: {overall_status}")
    print(f"   Completeness: {readiness_data.get('completeness_score', 0):.2%}")

    # Проверяем, что статус не "blocked"
    assert overall_status != "blocked", f"Readiness blocked: {readiness_data}"

    # Шаг 2: Генерация паспорта
    print("\n📋 Шаг 2: Генерация паспорта...")
    # Используем последний batch_id для генерации
    batch_id = uploaded_batches[-1]

    # Генерируем с шаблоном new_energy_passport
    response = client.post(
        f"/api/generate-passport/{batch_id}",
        params={"template_name": "new_energy_passport"},
    )

    assert response.status_code == 200, f"Ошибка генерации паспорта: {response.text}"

    # Сохраняем файл во временный файл
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        tmp_file.write(response.content)
        passport_path = Path(tmp_file.name)

    print(f"   ✅ Паспорт сохранен: {passport_path}")
    print(f"   Размер: {passport_path.stat().st_size} байт")

    # Шаг 3: Проверка структуры
    print("\n📋 Шаг 3: Проверка структуры файла...")
    wb = load_workbook(passport_path, data_only=False)

    print(f"   Всего листов: {len(wb.sheetnames)}")
    print(f"   Листы: {', '.join(wb.sheetnames)}")

    # Проверяем наличие ключевых листов
    key_sheets = {
        "Struktura pr2": [
            "Структура пр 2",
            "Struktura pr2",
            "02_Структура",
            "Структура пр 2 ",
            "Struktura pr2 ",
        ],
        "Balans": ["04_Баланс", "Баланс", "Balance", "Balans", "Баланс ", "04_Баланс "],
        "Equipment": [
            "03_Оборудование",
            "Equipment",
            "Оборудование",
            "Sheet1",
            "03_Оборудование ",
            "Equipment ",
        ],
        "Nodes": [
            "01_Узлы учета",
            "Узел учета",
            "Узлы учета",
            "Nodes",
            "Uzel ucheta",
            "Uzel ucheta ",
        ],
    }

    found_sheets = {}
    for category, names in key_sheets.items():
        found = None
        for name in names:
            for sheet_name in wb.sheetnames:
                if sheet_name.strip() == name.strip() or sheet_name == name:
                    found = sheet_name
                    break
            if found:
                break
        found_sheets[category] = found
        status = "✅" if found else "❌"
        print(f"   {status} {category}: {found if found else 'НЕ НАЙДЕН'}")

    # Проверяем, что все ключевые листы найдены
    missing_sheets = [cat for cat, sheet in found_sheets.items() if not sheet]
    assert not missing_sheets, f"Не найдены листы: {missing_sheets}"

    # Шаг 4: Проверка формул
    print("\n📋 Шаг 4: Проверка формул...")

    # Проверяем формулы на листе Balans
    balans_sheet_name = found_sheets["Balans"]
    balans_ws = wb[balans_sheet_name]
    balans_formulas = []
    for row in balans_ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value:
                balans_formulas.append((cell.coordinate, str(cell.value)))

    print(f"   Balans: найдено формул: {len(balans_formulas)}")
    if balans_formulas:
        print("   Примеры формул на Balans:")
        for coord, formula in balans_formulas[:5]:
            print(f"     {coord}: {formula[:60]}...")

    assert len(balans_formulas) > 0, "На листе Balans не найдено формул!"

    # Проверяем формулы на листе Equipment
    equipment_sheet_name = found_sheets["Equipment"]
    equipment_ws = wb[equipment_sheet_name]
    equipment_formulas = []
    for row in equipment_ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value:
                equipment_formulas.append((cell.coordinate, str(cell.value)))

    print(f"   Equipment: найдено формул: {len(equipment_formulas)}")
    if equipment_formulas:
        print("   Примеры формул на Equipment:")
        for coord, formula in equipment_formulas[:5]:
            print(f"     {coord}: {formula[:60]}...")

    assert len(equipment_formulas) > 0, "На листе Equipment не найдено формул!"

    # Шаг 5: Проверка данных
    print("\n📋 Шаг 5: Проверка данных...")

    # Проверяем заполнение Struktura pr2
    struktura_sheet_name = found_sheets["Struktura pr2"]
    struktura_ws = wb[struktura_sheet_name]

    # Проверяем заполнение ресурсов по кварталам
    # Согласно build_quarter_mapping, данные в строке 9 для каждого квартала
    resources_found = {
        "electricity": False,
        "gas": False,
        "water": False,
        "heat": False,
    }

    # Проверяем колонки для кварталов 2022-Q1 (колонки 3, 4, 6, 14)
    quarter_cols = [
        3,
        19,
        35,
        51,
        67,
        83,
        99,
        115,
        131,
        147,
        163,
        179,
    ]  # Все кварталы 2022-2024

    quarters_with_data = 0
    for quarter_col in quarter_cols:
        row = 9
        active_cell = struktura_ws.cell(row=row, column=quarter_col)
        reactive_cell = struktura_ws.cell(row=row, column=quarter_col + 1)
        gas_cell = struktura_ws.cell(row=row, column=quarter_col + 3)
        water_cell = struktura_ws.cell(row=row, column=quarter_col + 11)

        has_data = False
        if (
            active_cell.value
            and isinstance(active_cell.value, (int, float))
            and active_cell.value > 0
        ):
            resources_found["electricity"] = True
            has_data = True
        if (
            reactive_cell.value
            and isinstance(reactive_cell.value, (int, float))
            and reactive_cell.value > 0
        ):
            has_data = True
        if (
            gas_cell.value
            and isinstance(gas_cell.value, (int, float))
            and gas_cell.value > 0
        ):
            resources_found["gas"] = True
            has_data = True
        if (
            water_cell.value
            and isinstance(water_cell.value, (int, float))
            and water_cell.value > 0
        ):
            resources_found["water"] = True
            has_data = True

        if has_data:
            quarters_with_data += 1

    print("   Struktura pr2:")
    print(f"     - Кварталов с данными: {quarters_with_data}/12")
    print("     - Ресурсы найдены:")
    for resource, found in resources_found.items():
        status = "✅" if found else "❌"
        print(f"       {status} {resource}")

    # Проверяем, что хотя бы некоторые ресурсы заполнены (кроме угля и мазута)
    required_resources = ["electricity", "gas", "water"]
    found_required = sum(1 for r in required_resources if resources_found[r])
    assert found_required > 0, (
        f"Не найдены данные по обязательным ресурсам. Найдено: {resources_found}"
    )

    # Проверяем данные на листе Balans
    balans_data_cells = []
    for row in balans_ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                if isinstance(cell.value, (int, float)) and cell.value > 0:
                    balans_data_cells.append((cell.coordinate, cell.value))

    print(f"   Balans: найдено ячеек с данными: {len(balans_data_cells)}")
    assert len(balans_data_cells) > 0, "На листе Balans не найдено данных!"

    wb.close()

    print(f"\n{'=' * 80}")
    print("✅ ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ!")
    print(f"{'=' * 80}")
    print(f"Временный файл паспорта: {passport_path}")
    print("Для ручной проверки откройте файл в Excel")

    # Не удаляем файл, чтобы можно было проверить вручную
    # passport_path.unlink()

    # Возвращаем путь к файлу для возможной дальнейшей проверки
    return passport_path
