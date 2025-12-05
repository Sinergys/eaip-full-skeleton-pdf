from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from pydantic import BaseModel
from uuid import uuid4
import json
import os
import httpx
import hashlib
import shutil
from pathlib import Path
from typing import Optional, Dict, Any, List
import logging
import sys

# –ó–∞–≥—Ä—É–∑–∫–∞ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã—Ö –æ–∫—Ä—É–∂–µ–Ω–∏—è –∏–∑ .env —Ñ–∞–π–ª–∞
try:
    from dotenv import load_dotenv
    # –ó–∞–≥—Ä—É–∂–∞–µ–º .env –∏–∑ –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏–∏ —Å–µ—Ä–≤–∏—Å–∞
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        load_dotenv(env_path)
        logging.getLogger(__name__).info(f"‚úÖ –ó–∞–≥—Ä—É–∂–µ–Ω .env —Ñ–∞–π–ª: {env_path}")
    else:
        # –ü—Ä–æ–±—É–µ–º –∑–∞–≥—Ä—É–∑–∏—Ç—å –∏–∑ –∫–æ—Ä–Ω—è –ø—Ä–æ–µ–∫—Ç–∞
        project_env = Path(__file__).resolve().parent.parent.parent.parent / ".env"
        if project_env.exists():
            load_dotenv(project_env)
            logging.getLogger(__name__).info(f"‚úÖ –ó–∞–≥—Ä—É–∂–µ–Ω .env —Ñ–∞–π–ª: {project_env}")
except ImportError:
    # python-dotenv –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω, –∏—Å–ø–æ–ª—å–∑—É–µ–º —Ç–æ–ª—å–∫–æ —Å–∏—Å—Ç–µ–º–Ω—ã–µ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã–µ
    pass
from file_parser import parse_file
import database`r`nimport postgres_db
from openpyxl import load_workbook
from settings.excel_semantic_settings import get_excel_semantic_mode
from ai.ai_excel_semantic_parser import CanonicalSourceData
from utils.canonical_collector import collect_canonical_from_workbook
from utils.canonical_collector import analyze_workbook_sheets
from utils.canonical_to_passport import canonical_to_passport_payload
from domain.passport_requirements import (
    evaluate_generation_readiness,
    GenerationReadinessResult,
)
from utils.energy_aggregator import (
    aggregate_energy_data,
    should_aggregate_file,
    write_aggregation_json,
    aggregate_from_db_json,
    aggregate_usage_categories,
    distribute_categories_by_quarter,
)
from utils.equipment_parser import (
    is_equipment_file,
    parse_equipment_workbook,
    write_equipment_json,
)
from utils.building_envelope_parser import (
    is_envelope_file,
    parse_building_envelope,
    write_envelope_json,
)
from utils.nodes_parser import (
    is_nodes_file,
    parse_nodes_workbook,
    write_nodes_json,
    load_nodes_from_json,
)
from utils.balance_sheet_detector import is_balance_sheet_file
from utils.balance_sheet_node_extractor import extract_node_consumption_from_balance_sheet
from utils.aggregation_log import log_aggregation_event
from utils.intelligent_router import IntelligentRouter
from utils.progress_tracker import (
    FileType,
    ProcessingStage,
    create_progress_tracker,
    get_progress_tracker,
    remove_progress_tracker,
)
from utils.readiness_validator import (
    validate_generation_readiness,
    get_upload_checklist,
)
from utils.data_validator import (
    validate_data_for_template,
)

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –î–û –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è logger
# –£—Ä–æ–≤–µ–Ω—å –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –º–æ–∂–Ω–æ –∏–∑–º–µ–Ω–∏—Ç—å —á–µ—Ä–µ–∑ –ø–µ—Ä–µ–º–µ–Ω–Ω—É—é –æ–∫—Ä—É–∂–µ–Ω–∏—è LOG_LEVEL
log_level = os.getenv("LOG_LEVEL", "DEBUG").upper()
logging.basicConfig(
    level=getattr(logging, log_level, logging.DEBUG),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)
logger.info(f"üîß –£—Ä–æ–≤–µ–Ω—å –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω: {log_level}")

# –ü–µ—Ä–µ–∫–ª—é—á–∞—Ç–µ–ª—å —Ä–µ–∂–∏–º–æ–≤ —Ä–∞–±–æ—Ç—ã —Å–∏—Å—Ç–µ–º—ã
SYSTEM_MODE = os.getenv("SYSTEM_MODE", "debug").lower()  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é debug –¥–ª—è —Ä–∞–∑—Ä–∞–±–æ—Ç–∫–∏
if SYSTEM_MODE not in ["debug", "production"]:
    SYSTEM_MODE = "debug"
    logger.warning(f"‚ö†Ô∏è –ù–µ–≤–µ—Ä–Ω—ã–π SYSTEM_MODE, —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω —Ä–µ–∂–∏–º 'debug'")
logger.info(f"üîß –†–µ–∂–∏–º —Ä–∞–±–æ—Ç—ã —Å–∏—Å—Ç–µ–º—ã: {SYSTEM_MODE.upper()} (–¥–ª—è –∏–∑–º–µ–Ω–µ–Ω–∏—è —É—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ SYSTEM_MODE=debug –∏–ª–∏ SYSTEM_MODE=production, –∏–ª–∏ –∏—Å–ø–æ–ª—å–∑—É–π—Ç–µ –ø–µ—Ä–µ–∫–ª—é—á–∞—Ç–µ–ª—å –≤ –≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–µ)")

# –î–æ–±–∞–≤–ª—è–µ–º tools –≤ –ø—É—Ç—å –¥–ª—è –∏–º–ø–æ—Ä—Ç–∞ –≥–µ–Ω–µ—Ä–∞—Ç–æ—Ä–∞
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent / "tools"))
try:
    from pkm690_excel_generator import PKM690ExcelGenerator

    HAS_GENERATOR = True
except ImportError as e:
    logger.warning(f"PKM690 –≥–µ–Ω–µ—Ä–∞—Ç–æ—Ä –Ω–µ –¥–æ—Å—Ç—É–ø–µ–Ω: {e}")
    HAS_GENERATOR = False

try:
    from fill_energy_passport import (
        fill_struktura_pr2,
        fill_nodes_sheet,
        load_default_nodes,
        fill_building_envelope_sheet,
        fill_equipment_sheet,
        fill_balans_sheet,
        fill_dinamika_sheet,
        fill_fuel_dynamics_sheet,
        fill_specific_consumption_sheet,
        fill_meropriyatiya_sheet,
        fill_monthly_sheet,
    )

    HAS_FILLER = True
except ImportError as e:
    logger.warning(f"fill_energy_passport –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω: {e}")
    HAS_FILLER = False

# –ò–º–ø–æ—Ä—Ç AI –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ–ª—è —Ñ–æ—Ä–º—É–ª
try:
    from utils.ai_formula_restorer import AIFormulaRestorer

    HAS_FORMULA_RESTORER = True
except ImportError as e:
    logger.warning(f"ai_formula_restorer –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω: {e}")
    HAS_FORMULA_RESTORER = False

try:
    from utils.word_report_generator import WordReportGenerator

    HAS_WORD_GENERATOR = True
except ImportError as e:
    logger.warning(f"WordReportGenerator –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω: {e}")
    HAS_WORD_GENERATOR = False

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent

WEB_DIR = Path(__file__).resolve().parent / "web"
WEB_DIR.mkdir(parents=True, exist_ok=True)

database.init_db()

app = FastAPI(title="EAIP ingest", version="0.1.0")


class ValidateRequest(BaseModel):
    batchId: str


class EnterpriseCreate(BaseModel):
    name: str


class EditablePayload(BaseModel):
    text: str


# Inbox directory for uploaded files
INBOX_DIR = os.getenv("INBOX_DIR", "/data/inbox")
os.makedirs(INBOX_DIR, exist_ok=True)

AGGREGATED_DIR = Path(
    os.getenv("AGGREGATED_DIR", os.path.join(INBOX_DIR, "aggregated"))
)
AGGREGATED_DIR.mkdir(parents=True, exist_ok=True)

# –ö–æ–Ω—Å—Ç–∞–Ω—Ç—ã —Å–æ–≥–ª–∞—Å–Ω–æ –¢–ó (—Ä–∞–∑–¥–µ–ª 4.1, 4.2)
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".docx", ".pdf", ".jpg", ".jpeg", ".png"}
ALLOWED_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # XLSX
    "application/vnd.ms-excel.sheet.macroEnabled.12",  # XLSM
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # DOCX
    "application/pdf",
    "image/jpeg",
    "image/jpg",
    "image/png",
}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 –ú–ë —Å–æ–≥–ª–∞—Å–Ω–æ –¢–ó (—Ä–∞–∑–¥–µ–ª 4.2)
FILE_TYPE_LABELS = {
    ".xlsx": "Excel (XLSX)",
    ".xlsm": "Excel —Å –º–∞–∫—Ä–æ—Å–∞–º–∏ (XLSM)",
    ".docx": "Word (DOCX)",
    ".pdf": "PDF",
    ".jpg": "–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ (JPG)",
    ".jpeg": "–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ (JPEG)",
    ".png": "–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ (PNG)",
}


@app.get("/api/batches/{batch_id}/generation-readiness")
def api_generation_readiness(batch_id: str):
    """
    –û—Ü–µ–Ω–∫–∞ –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –∫ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –ø–∞—Å–ø–æ—Ä—Ç–∞ –Ω–∞ –æ—Å–Ω–æ–≤–µ CanonicalSourceData.
    """
    upload = database.get_upload_by_batch(batch_id)
    if not upload:
        raise HTTPException(status_code=404, detail=f"–ó–∞–≥—Ä—É–∑–∫–∞ {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")

    raw_json = upload.get("raw_json") or {}
    canonical = None
    if isinstance(raw_json, dict):
        canonical_dict = raw_json.get("canonical_source")
        if isinstance(canonical_dict, dict):
            try:
                canonical = CanonicalSourceData.parse_obj(canonical_dict)
            except Exception:
                canonical = None
    if canonical is None:
        # Try to reconstruct from the original file if name is known
        try:
            filename = upload.get("filename")
            if filename:
                canonical = collect_canonical_from_workbook(filename)
        except Exception:
            canonical = None

    result = evaluate_generation_readiness(canonical)
    return {
        "batch_id": batch_id,
        "overall_status": result.overall_status,
        "missing_required": [rf.__dict__ for rf in result.missing_required],
        "missing_optional": [rf.__dict__ for rf in result.missing_optional],
        "notes": result.notes,
        "mode": get_excel_semantic_mode(),
    }


@app.get("/api/batches/{batch_id}/canonical-debug")
def api_canonical_debug(batch_id: str):
    """
    –í–Ω—É—Ç—Ä–µ–Ω–Ω–∏–π –æ—Ç–ª–∞–¥–æ—á–Ω—ã–π —ç–Ω–¥–ø–æ–∏–Ω—Ç –¥–ª—è –ø—Ä–æ—Å–º–æ—Ç—Ä–∞ CanonicalSourceData –∏ –≤–∫–ª–∞–¥–∞ AI.
    –ù–µ –≤—ã–ø–æ–ª–Ω—è–µ—Ç –≥–µ–Ω–µ—Ä–∞—Ü–∏—é –∏ –Ω–µ –º–µ–Ω—è–µ—Ç —Å–æ—Å—Ç–æ—è–Ω–∏–µ.
    """
    mode = get_excel_semantic_mode()
    logger.info("Canonical debug requested for batch_id=%s mode=%s", batch_id, mode)

    upload = database.get_upload_by_batch(batch_id)
    if not upload:
        raise HTTPException(status_code=404, detail=f"–ó–∞–≥—Ä—É–∑–∫–∞ {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")

    # 1) –ü—ã—Ç–∞–µ–º—Å—è –∏–∑–≤–ª–µ—á—å canonical_source –∏–∑ —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã—Ö raw_json
    raw_json = upload.get("raw_json") or {}
    canonical = None
    if isinstance(raw_json, dict):
        canonical_dict = raw_json.get("canonical_source")
        if isinstance(canonical_dict, dict):
            try:
                canonical = CanonicalSourceData.parse_obj(canonical_dict)
            except Exception:
                canonical = None

    # 2) –ï—Å–ª–∏ –Ω–µ—Ç ‚Äî –ø—ã—Ç–∞–µ–º—Å—è —Ä–µ–∫–æ–Ω—Å—Ç—Ä—É–∏—Ä–æ–≤–∞—Ç—å –∏–∑ –∏—Å—Ö–æ–¥–Ω–æ–≥–æ —Ñ–∞–π–ª–∞ (best-effort)
    if canonical is None:
        try:
            filename = upload.get("filename")
            if filename:
                canonical = collect_canonical_from_workbook(filename)
        except Exception as exc:
            logger.warning(
                "Failed to reconstruct canonical for batch_id=%s: %s", batch_id, exc
            )
            canonical = None

    # –§–æ—Ä–º–∏—Ä—É–µ–º –æ—Ç–≤–µ—Ç
    response = {
        "batch_id": batch_id,
        "mode": mode,
        "canonical_source": canonical.dict() if canonical else None,
        "provenance": (canonical.provenance if canonical else {}) or {},
        "sheets": [],
    }
    # –ü–æ–ø—ã—Ç–∫–∞ –¥–æ–±–∞–≤–∏—Ç—å sheet-level –¥–µ—Ç–∞–ª–∏ (on-the-fly –∞–Ω–∞–ª–∏–∑ –∏—Å—Ö–æ–¥–Ω–æ–≥–æ —Ñ–∞–π–ª–∞)
    try:
        filename = upload.get("filename")
        if filename:
            response["sheets"] = analyze_workbook_sheets(filename)
    except Exception as exc:
        logger.warning("Sheet-level debug analysis failed for %s: %s", batch_id, exc)
    return response


# Mapping of resource type codes to human-friendly labels
RESOURCE_LABELS: Dict[str, str] = {
    "electricity": "–≠–ª–µ–∫—Ç—Ä–æ—ç–Ω–µ—Ä–≥–∏—è",
    "gas": "–ì–∞–∑",
    "heat": "–¢–µ–ø–ª–æ–≤–∞—è —ç–Ω–µ—Ä–≥–∏—è",
    "water": "–í–æ–¥–∞",
    "fuel": "–¢–æ–ø–ª–∏–≤–æ –∏ –ì–°–ú",
    "equipment": "–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
    "envelope": "–†–∞—Å—á–µ—Ç —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º",
    "nodes": "–£–∑–ª—ã —É—á–µ—Ç–∞",
    "other": "–ü—Ä–æ—á–µ–µ",
}

# –•—Ä–∞–Ω–∏–ª–∏—â–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –ø–∞—Ä—Å–∏–Ω–≥–∞ (–≤ –ø—Ä–æ–¥–∞–∫—à–µ–Ω–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å Redis/PostgreSQL)
# –î–ª—è —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏ –æ—Å—Ç–∞–≤–∏–ª–∏ –∫–µ—à, –Ω–æ –ø–µ—Ä–≤–∏—á–Ω–æ–µ —Ö—Ä–∞–Ω–∏–ª–∏—â–µ ‚Äî SQLite.
parsing_results_cache: Dict[str, Dict[str, Any]] = {}

# Enable CORS for web interface
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify exact origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def render_html_page(filename: str) -> HTMLResponse:
    file_path = WEB_DIR / filename
    if not file_path.exists():
        logger.error("HTML page %s not found in %s", filename, WEB_DIR)
        raise HTTPException(status_code=500, detail="–°—Ç—Ä–∞–Ω–∏—Ü–∞ –≤—Ä–µ–º–µ–Ω–Ω–æ –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–∞")
    return HTMLResponse(content=file_path.read_text(encoding="utf-8"))


def normalize_file_type(parsing_result: Optional[Dict[str, Any]], file_ext: str) -> str:
    if parsing_result and parsing_result.get("file_type"):
        return str(parsing_result.get("file_type")).lower()
    ext = file_ext.lower()
    if ext in {".xlsx", ".xlsm"}:
        return "excel"
    if ext in {".docx"}:
        return "docx"
    if ext in {".pdf"}:
        return "pdf"
    return "unknown"


def get_file_type_enum(file_ext: str) -> FileType:
    """–ü—Ä–µ–æ–±—Ä–∞–∑—É–µ—Ç —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ —Ñ–∞–π–ª–∞ –≤ FileType enum"""
    ext = file_ext.lower()
    if ext in {".xlsx", ".xlsm", ".xls"}:
        return FileType.EXCEL
    elif ext in {".docx"}:
        return FileType.WORD
    elif ext in {".pdf"}:
        return FileType.PDF
    elif ext in {".jpg", ".jpeg", ".png"}:
        return FileType.IMAGE
    else:
        return FileType.UNKNOWN


def build_parsing_summary(
    parsing_result: Optional[Dict[str, Any]], file_ext: str
) -> Optional[Dict[str, Any]]:
    if not parsing_result or not parsing_result.get("parsed"):
        return None

    data = parsing_result.get("data") or {}
    file_type = normalize_file_type(parsing_result, file_ext)

    if file_type == "pdf":
        metadata = data.get("metadata", {})
        return {
            "pages": metadata.get("num_pages", 0),
            "characters": data.get("total_characters", 0),
            "tables": data.get("total_tables", 0),
        }
    if file_type == "excel":
        sheets = data.get("sheets", [])
        return {
            "sheets": len(sheets),
            "total_rows": sum(sheet.get("max_row", 0) for sheet in sheets),
        }
    if file_type == "docx":
        return {
            "paragraphs": len(data.get("paragraphs", [])),
            "tables": len(data.get("tables", [])),
        }
    return None


def build_editable_text(parsing_result: Optional[Dict[str, Any]]) -> str:
    if not parsing_result or not parsing_result.get("parsed"):
        return ""

    data = parsing_result.get("data") or {}
    file_type = normalize_file_type(parsing_result, "")

    if file_type == "pdf":
        text = data.get("text")
        return text if isinstance(text, str) else ""

    if file_type == "excel":
        lines: List[str] = []
        sheets = data.get("sheets", [])
        for sheet in sheets:
            name = sheet.get("name") or "–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è"
            lines.append(f"=== –õ–∏—Å—Ç: {name} ===")
            rows = sheet.get("rows") or []
            if not rows:
                lines.append("(–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö)")
            else:
                for row in rows[:100]:
                    if isinstance(row, list):
                        values = ["" if value is None else str(value) for value in row]
                    elif isinstance(row, dict):
                        values = [
                            "" if value is None else str(value)
                            for value in row.values()
                        ]
                    else:
                        values = ["" if row is None else str(row)]
                    lines.append("\t".join(values))
            lines.append("")
        return "\n".join(lines).strip()

    if file_type == "docx":
        paragraphs = data.get("paragraphs") or []
        return "\n".join(
            paragraph.get("text", "")
            for paragraph in paragraphs
            if paragraph.get("text")
        )

    return json.dumps(data, ensure_ascii=False, indent=2)


def ensure_parsing_cached(batch_id: str) -> Optional[Dict[str, Any]]:
    """
    –û–±–µ—Å–ø–µ—á–∏–≤–∞–µ—Ç –Ω–∞–ª–∏—á–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –ø–∞—Ä—Å–∏–Ω–≥–∞ –≤ –∫—ç—à–µ.
    –ï—Å–ª–∏ –¥–∞–Ω–Ω—ã—Ö –Ω–µ—Ç –≤ –∫—ç—à–µ, –∑–∞–≥—Ä—É–∂–∞–µ—Ç –∏—Ö –∏–∑ –ë–î.
    """
    if batch_id not in parsing_results_cache:
        record = database.get_upload_by_batch(batch_id)
        if record:
            # –§–æ—Ä–º–∏—Ä—É–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –∫—ç—à–∞ –∏–∑ –∑–∞–ø–∏—Å–∏ –ë–î
            raw_json = record.get("raw_json")
            if raw_json:
                # –ï—Å–ª–∏ raw_json —É–∂–µ —è–≤–ª—è–µ—Ç—Å—è –ø–æ–ª–Ω–æ–π —Å—Ç—Ä—É–∫—Ç—É—Ä–æ–π (—Å batch_id, parsing –∏ —Ç.–¥.)
                if isinstance(raw_json, dict) and "batch_id" in raw_json:
                    parsing_results_cache[batch_id] = raw_json
                else:
                    # –ï—Å–ª–∏ raw_json - —ç—Ç–æ —Ç–æ–ª—å–∫–æ –¥–∞–Ω–Ω—ã–µ –ø–∞—Ä—Å–∏–Ω–≥–∞, —Ñ–æ—Ä–º–∏—Ä—É–µ–º –ø–æ–ª–Ω—É—é —Å—Ç—Ä—É–∫—Ç—É—Ä—É
                    parsing_results_cache[batch_id] = {
                        "batch_id": batch_id,
                        "filename": record.get("filename"),
                        "file_path": None,  # –ü—É—Ç—å –Ω–µ —Å–æ—Ö—Ä–∞–Ω—è–µ—Ç—Å—è –≤ –ë–î
                        "file_type": record.get("file_type"),
                        "file_size": record.get("file_size"),
                        "parsing": raw_json if isinstance(raw_json, dict) else {},
                        "status": record.get("status", "unknown"),
                        "resource_type": record.get("parsing_summary", {}).get("resource_type") if record.get("parsing_summary") else None,
                    }
            else:
                # –ï—Å–ª–∏ raw_json –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç, —Å–æ–∑–¥–∞–µ–º –º–∏–Ω–∏–º–∞–ª—å–Ω—É—é —Å—Ç—Ä—É–∫—Ç—É—Ä—É
                logger.warning(f"raw_json –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –¥–ª—è batch_id={batch_id}, —Å–æ–∑–¥–∞–Ω–∞ –º–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Å—Ç—Ä—É–∫—Ç—É—Ä–∞")
                parsing_results_cache[batch_id] = {
                    "batch_id": batch_id,
                    "filename": record.get("filename"),
                    "file_path": None,
                    "file_type": record.get("file_type"),
                    "file_size": record.get("file_size"),
                    "parsing": None,
                    "status": record.get("status", "error"),
                    "error": "–î–∞–Ω–Ω—ã–µ –ø–∞—Ä—Å–∏–Ω–≥–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã –≤ –ë–î",
                }
    return parsing_results_cache.get(batch_id)


@app.get("/health")
def health():
    return {"service": "ingest", "status": "ok"}


@app.get("/test-xlsm")
def test_xlsm():
    """–ü—Ä–æ—Å—Ç–æ–π —Ç–µ—Å—Ç–æ–≤—ã–π endpoint –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ .xlsm –ø–æ–¥–¥–µ—Ä–∂–∫–∏"""
    return {
        "xlsm_in_allowed": ".xlsm" in ALLOWED_EXTENSIONS,
        "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
        "message": "–¢–µ—Å—Ç –ø–æ–¥–¥–µ—Ä–∂–∫–∏ .xlsm",
    }


@app.get("/api/debug/extensions")
@app.get("/debug/extensions")
def debug_extensions():
    """–î–∏–∞–≥–Ω–æ—Å—Ç–∏—á–µ—Å–∫–∏–π endpoint –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã—Ö —Ñ–æ—Ä–º–∞—Ç–æ–≤"""
    from datetime import datetime

    try:
        return {
            "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
            "xlsm_supported": ".xlsm" in ALLOWED_EXTENSIONS,
            "allowed_mime_types": list(ALLOWED_MIME_TYPES),
            "file_type_labels": FILE_TYPE_LABELS,
            "timestamp": datetime.now().isoformat(),
            "code_version": "2025-01-16-xlsm-support",
            "server_status": "running",
        }
    except Exception as e:
        return {
            "error": str(e),
            "allowed_extensions": list(ALLOWED_EXTENSIONS)
            if "ALLOWED_EXTENSIONS" in globals()
            else "not found",
        }


@app.get("/web/upload")
def upload_page():
    return render_html_page("upload.html")


@app.get("/web/files")
def files_page():
    """–í–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –¥–ª—è –ø—Ä–æ—Å–º–æ—Ç—Ä–∞ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤"""
    return render_html_page("files.html")


@app.get("/web/results")
def results_page():
    return render_html_page("results.html")


@app.get("/web/normative")
def normative_upload_page():
    """–í–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
    return render_html_page("normative_upload.html")


@app.get("/web/normative/upload")
def normative_upload_page_alt():
    """–í–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ (–∞–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω—ã–π URL)"""
    return render_html_page("normative_upload.html")


@app.get("/web/normative/dashboard")
def web_normative_dashboard():
    """–î–∞—à–±–æ—Ä–¥ —Å–æ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–æ–π –Ω–æ—Ä–º–∞—Ç–∏–≤–æ–≤"""
    return render_html_page("normative_dashboard.html")


@app.get("/api/normative/test")
def test_normative_endpoint():
    """–¢–µ—Å—Ç–æ–≤—ã–π endpoint –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ API –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
    return {
        "status": "ok",
        "message": "API –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –¥–æ—Å—Ç—É–ø–µ–Ω",
        "endpoints": {
            "upload": "/api/normative/upload (POST)",
            "documents": "/api/normative/documents (GET)",
            "rules": "/api/normative/rules/{rule_type} (GET)",
            "ai_status": "/api/normative/ai-status (GET)",
        },
    }


@app.get("/api/normative/ai-status")
def get_ai_status():
    """–ü–æ–ª—É—á–∏—Ç—å —Å—Ç–∞—Ç—É—Å AI –¥–ª—è –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
    try:
        from settings.ai_settings import get_ai_status, get_ai_settings
        from domain.normative_importer import get_normative_importer

        # –ü–æ–ª—É—á–∞–µ–º —Å—Ç–∞—Ç—É—Å AI –∏–∑ –µ–¥–∏–Ω–æ–≥–æ –º–æ–¥—É–ª—è –Ω–∞—Å—Ç—Ä–æ–µ–∫
        ai_status = get_ai_status()
        get_ai_settings()

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å –∏–º–ø–æ—Ä—Ç–µ—Ä–∞
        importer = get_normative_importer()
        importer_available = importer is not None
        ai_parser_available = (
            importer and importer.ai_parser is not None and importer.ai_parser.enabled
            if importer
            else False
        )

        return {
            **ai_status,
            "importer_available": importer_available,
            "ai_parser_available": ai_parser_available,
        }
    except Exception as e:
        logger.exception("–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å—Ç–∞—Ç—É—Å–∞ AI")
        return {
            "ai_enabled": False,
            "ai_provider": "unknown",
            "has_api_key": False,
            "has_valid_config": False,
            "importer_available": False,
            "ai_parser_available": False,
            "error": str(e),
            "message": f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å—Ç–∞—Ç—É—Å–∞ AI: {e}",
        }


@app.get("/api/debug/extensions")
def api_debug_extensions():
    """–î–∏–∞–≥–Ω–æ—Å—Ç–∏—á–µ—Å–∫–∏–π endpoint –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã—Ö —Ñ–æ—Ä–º–∞—Ç–æ–≤ (API –≤–µ—Ä—Å–∏—è)"""
    from datetime import datetime

    try:
        return {
            "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
            "xlsm_supported": ".xlsm" in ALLOWED_EXTENSIONS,
            "allowed_mime_types": list(ALLOWED_MIME_TYPES),
            "file_type_labels": FILE_TYPE_LABELS,
            "timestamp": datetime.now().isoformat(),
            "code_version": "2025-01-16-xlsm-support",
            "server_status": "running",
        }
    except Exception as e:
        return {
            "error": str(e),
            "allowed_extensions": list(ALLOWED_EXTENSIONS)
            if "ALLOWED_EXTENSIONS" in globals()
            else "not found",
        }


@app.get("/api/enterprises")
def api_list_enterprises():
    return {"items": database.list_enterprises()}


@app.post("/api/enterprises")
def api_create_enterprise(payload: EnterpriseCreate):
    enterprise = database.get_or_create_enterprise(payload.name)
    return enterprise


@app.get("/api/enterprises/{enterprise_id}/uploads")
def api_enterprise_history(enterprise_id: int):
    enterprise = database.get_enterprise_by_id(enterprise_id)
    if not enterprise:
        raise HTTPException(status_code=404, detail="–ü—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")
    history = database.list_uploads_for_enterprise(enterprise_id)
    return {"enterprise": enterprise, "uploads": history}


@app.get("/api/enterprises/{enterprise_id}/upload-checklist")
def api_get_upload_checklist(enterprise_id: int):
    """
    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —á–µ–∫-–ª–∏—Å—Ç —Ç—Ä–µ–±—É–µ–º—ã—Ö —Ñ–∞–π–ª–æ–≤ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è.

    Returns:
        –°–ª–æ–≤–∞—Ä—å —Å —á–µ–∫-–ª–∏—Å—Ç–æ–º —Ç—Ä–µ–±—É–µ–º—ã—Ö –∏ –æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤.
    """
    enterprise = database.get_enterprise_by_id(enterprise_id)
    if not enterprise:
        raise HTTPException(status_code=404, detail="–ü—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")

    checklist = get_upload_checklist(enterprise_id)
    return {
        "enterprise_id": enterprise_id,
        "enterprise_name": enterprise.get("name"),
        **checklist,
    }


@app.get("/api/enterprises/{enterprise_id}/generation-readiness")
def api_get_generation_readiness(enterprise_id: int):
    """
    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å—Ç–∞—Ç—É—Å –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ —ç–Ω–µ—Ä–≥–µ—Ç–∏—á–µ—Å–∫–æ–≥–æ –ø–∞—Å–ø–æ—Ä—Ç–∞.

    Returns:
        –°–ª–æ–≤–∞—Ä—å —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ –ø—Ä–æ–≤–µ—Ä–∫–∏ –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏:
        - ready: –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç—å –∫ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏
        - completeness_score: –ø–æ–∫–∞–∑–∞—Ç–µ–ª—å –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ (0.0-1.0)
        - missing_resources: —Å–ø–∏—Å–æ–∫ –Ω–µ–¥–æ—Å—Ç–∞—é—â–∏—Ö —Ä–µ—Å—É—Ä—Å–æ–≤
        - missing_files: —Å–ø–∏—Å–æ–∫ –Ω–µ–¥–æ—Å—Ç–∞—é—â–∏—Ö —Ñ–∞–π–ª–æ–≤
        - available_resources: —Å–ø–∏—Å–æ–∫ –¥–æ—Å—Ç—É–ø–Ω—ã—Ö —Ä–µ—Å—É—Ä—Å–æ–≤
        - warnings: —Å–ø–∏—Å–æ–∫ –ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–µ–Ω–∏–π
        - progress_percentage: –ø—Ä–æ—Ü–µ–Ω—Ç –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏
    """
    enterprise = database.get_enterprise_by_id(enterprise_id)
    if not enterprise:
        raise HTTPException(status_code=404, detail="–ü—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")

    readiness = validate_generation_readiness(enterprise_id)
    return {
        "enterprise_id": enterprise_id,
        "enterprise_name": enterprise.get("name"),
        **readiness,
    }


@app.get("/api/uploads/{batch_id}")
def api_get_upload(batch_id: str):
    record = database.get_upload_by_batch(batch_id)
    if not record:
        raise HTTPException(status_code=404, detail="–ó–∞–≥—Ä—É–∑–∫–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")
    return record


@app.get("/api/uploads/{batch_id}/editable")
def api_get_editable(batch_id: str):
    record = database.get_upload_by_batch(batch_id)
    if not record:
        raise HTTPException(status_code=404, detail="–ó–∞–≥—Ä—É–∑–∫–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")
    return {
        "batch_id": batch_id,
        "editable_text": record.get("editable_text") or "",
        "updated_at": record.get("parsed_updated_at"),
    }


@app.post("/api/uploads/{batch_id}/editable")
def api_update_editable(batch_id: str, payload: EditablePayload):
    ensure_parsing_cached(batch_id)
    database.update_editable_text(batch_id, payload.text)
    return {"batch_id": batch_id, "status": "saved"}


@app.get("/ingest/parse/{batch_id}")
async def get_parsing_results(batch_id: str):
    data = ensure_parsing_cached(batch_id)
    if not data:
        raise HTTPException(
            status_code=404,
            detail=f"–†–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–∞—Ä—Å–∏–Ω–≥–∞ –¥–ª—è batch_id {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω—ã",
        )
    return JSONResponse(content=data)


@app.get("/api/progress/{batch_id}")
async def get_progress(batch_id: str):
    """
    –ü–æ–ª—É—á–µ–Ω–∏–µ —Ç–µ–∫—É—â–µ–≥–æ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞

    Returns:
        –°—Ç–∞—Ç—É—Å –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ —Å –¥–µ—Ç–∞–ª—å–Ω–æ–π –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–µ–π –ø–æ —ç—Ç–∞–ø–∞–º
    """
    tracker = get_progress_tracker(batch_id)
    if not tracker:
        # –ï—Å–ª–∏ —Ç—Ä–µ–∫–µ—Ä –Ω–µ –Ω–∞–π–¥–µ–Ω, –≤–æ–∑–º–æ–∂–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∫–∞ –µ—â–µ –Ω–µ –Ω–∞—á–∞–ª–∞—Å—å –∏–ª–∏ —É–∂–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∞
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –ë–î
        upload = database.get_upload_by_batch(batch_id)
        if upload:
            # –í–æ–∑–≤—Ä–∞—â–∞–µ–º —Ñ–∏–Ω–∞–ª—å–Ω—ã–π —Å—Ç–∞—Ç—É—Å –∏–∑ –ë–î
            return {
                "batch_id": batch_id,
                "file_type": "unknown",
                "overall_progress": 100,
                "current_stage": "completed",
                "stages": {},
                "is_completed": True,
                "has_error": upload.get("status") == "error",
                "error": None
                if upload.get("status") != "error"
                else "–û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞ —Å –æ—à–∏–±–∫–æ–π",
                "message": "–û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞. –¢—Ä–µ–∫–µ—Ä —É–¥–∞–ª–µ–Ω –∏–∑ –ø–∞–º—è—Ç–∏.",
            }
        raise HTTPException(
            status_code=404,
            detail=f"–¢—Ä–µ–∫–µ—Ä –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –¥–ª—è batch_id {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω",
        )

    return tracker.get_status()


@app.post("/web/upload/{batch_id}/cancel")
async def cancel_upload(batch_id: str):
    """
    –û—Ç–º–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω–æ–≥–æ —Ñ–∞–π–ª–∞

    Args:
        batch_id: –£–Ω–∏–∫–∞–ª—å–Ω—ã–π –∏–¥–µ–Ω—Ç–∏—Ñ–∏–∫–∞—Ç–æ—Ä –∑–∞–≥—Ä—É–∑–∫–∏

    Returns:
        –°—Ç–∞—Ç—É—Å –æ—Ç–º–µ–Ω—ã
    """
    tracker = get_progress_tracker(batch_id)
    if not tracker:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –ë–î, –≤–æ–∑–º–æ–∂–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∫–∞ —É–∂–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∞
        upload = database.get_upload_by_batch(batch_id)
        if upload:
            return {
                "batch_id": batch_id,
                "cancelled": False,
                "message": "–û–±—Ä–∞–±–æ—Ç–∫–∞ —É–∂–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∞, –æ—Ç–º–µ–Ω–∞ –Ω–µ–≤–æ–∑–º–æ–∂–Ω–∞",
            }
        raise HTTPException(
            status_code=404,
            detail=f"–¢—Ä–µ–∫–µ—Ä –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –¥–ª—è batch_id {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω",
        )

    if tracker.is_cancelled():
        return {
            "batch_id": batch_id,
            "cancelled": True,
            "message": "–û–±—Ä–∞–±–æ—Ç–∫–∞ —É–∂–µ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ —Ä–∞–Ω–µ–µ",
        }

    if tracker.completed_at:
        return {
            "batch_id": batch_id,
            "cancelled": False,
            "message": "–û–±—Ä–∞–±–æ—Ç–∫–∞ —É–∂–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∞, –æ—Ç–º–µ–Ω–∞ –Ω–µ–≤–æ–∑–º–æ–∂–Ω–∞",
        }

    tracker.cancel()
    logger.info(f"–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º: batch_id={batch_id}")

    return {
        "batch_id": batch_id,
        "cancelled": True,
        "message": "–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞",
        "status": tracker.get_status(),
    }


@app.get("/api/diagnose/pdf")
async def diagnose_pdf_endpoint(
    file_path: Optional[str] = None, batch_id: Optional[str] = None
):
    """
    –î–∏–∞–≥–Ω–æ—Å—Ç–∏–∫–∞ PDF —Ñ–∞–π–ª–∞

    Args:
        file_path: –ü—Ä—è–º–æ–π –ø—É—Ç—å –∫ PDF —Ñ–∞–π–ª—É (–¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è)
        batch_id: ID –∑–∞–≥—Ä—É–∑–∫–∏ –¥–ª—è –¥–∏–∞–≥–Ω–æ—Å—Ç–∏–∫–∏ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω–æ–≥–æ —Ñ–∞–π–ª–∞

    Returns:
        –î–∏–∞–≥–Ω–æ—Å—Ç–∏—á–µ—Å–∫–∏–π –æ—Ç—á–µ—Ç
    """
    from utils.pdf_diagnostics import diagnose_pdf

    if batch_id:
        # –ü–æ–ª—É—á–∞–µ–º –ø—É—Ç—å –∫ —Ñ–∞–π–ª—É –∏–∑ –∑–∞–≥—Ä—É–∑–∫–∏
        upload = database.get_upload_by_batch(batch_id)
        if not upload:
            raise HTTPException(
                status_code=404, detail=f"–ó–∞–≥—Ä—É–∑–∫–∞ {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞"
            )

        file_path = upload.get("file_path")
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(
                status_code=404, detail=f"–§–∞–π–ª –¥–ª—è batch_id {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω"
            )

    if not file_path:
        raise HTTPException(status_code=400, detail="–£–∫–∞–∂–∏—Ç–µ file_path –∏–ª–∏ batch_id")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"–§–∞–π–ª {file_path} –Ω–µ –Ω–∞–π–¥–µ–Ω")

    try:
        report = diagnose_pdf(file_path)
        return report
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –¥–∏–∞–≥–Ω–æ—Å—Ç–∏–∫–∏ PDF {file_path}: {e}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –¥–∏–∞–≥–Ω–æ—Å—Ç–∏–∫–∏: {e}")


@app.get("/ingest/parse/{batch_id}/summary")
async def get_parsing_summary(batch_id: str):
    data = ensure_parsing_cached(batch_id)
    if not data:
        raise HTTPException(
            status_code=404,
            detail=f"–†–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–∞—Ä—Å–∏–Ω–≥–∞ –¥–ª—è batch_id {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω—ã",
        )

    parsing_data = data.get("parsing") or {}
    if not parsing_data or not parsing_data.get("parsed"):
        return {
            "batch_id": batch_id,
            "status": data.get("status", "error"),
            "message": "–§–∞–π–ª –Ω–µ –±—ã–ª —Ä–∞—Å–ø–æ–∑–Ω–∞–Ω –∏–ª–∏ –ø—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞",
        }

    file_path = data.get("file_path")
    file_ext = Path(file_path).suffix.lower() if file_path else ""
    summary = build_parsing_summary(parsing_data, file_ext) or {}
    summary.update(
        {
            "batch_id": batch_id,
            "filename": data.get("filename"),
            "file_type": data.get("file_type"),
            "status": data.get("status", "success"),
            "parsed": True,
        }
    )
    return summary


@app.post("/ingest/validate")
async def proxy_validate(req: ValidateRequest):
    """–ü—Ä–æ–∫—Å–∏ –¥–ª—è –≤—ã–∑–æ–≤–∞ validate —Å–µ—Ä–≤–∏—Å–∞ –∏–∑ –≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                "http://validate:8002/validate/run", json={"batchId": req.batchId}
            )
            resp.raise_for_status()
            return resp.json()
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Validate service timeout")
    except httpx.HTTPStatusError as exc:
        raise HTTPException(
            status_code=exc.response.status_code,
            detail=f"Validate service error: {exc.response.text}",
        )
    except httpx.RequestError as exc:
        raise HTTPException(
            status_code=503, detail=f"Validate service unavailable: {exc}"
        )


def validate_file(file: UploadFile):
    """–í–∞–ª–∏–¥–∞—Ü–∏—è —Ñ–∞–π–ª–∞ —Å–æ–≥–ª–∞—Å–Ω–æ –¢–ó (—Ä–∞–∑–¥–µ–ª 4.1) —Å —É–ª—É—á—à–µ–Ω–Ω—ã–º –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ–º"""
    logger.info("=" * 70)
    logger.info(f"üîç [VALIDATE] –ù–ê–ß–ê–õ–û –í–ê–õ–ò–î–ê–¶–ò–ò —Ñ–∞–π–ª–∞: {file.filename}")
    logger.info("=" * 70)
    
    # –ü—Ä–æ–≤–µ—Ä–∫–∞ 1: –ò–º—è —Ñ–∞–π–ª–∞
    if not file.filename:
        logger.error("‚ùå [VALIDATE] –û–®–ò–ë–ö–ê: –ò–º—è —Ñ–∞–π–ª–∞ –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç")
        return False, "–ò–º—è —Ñ–∞–π–ª–∞ –æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ"

    file_ext = Path(file.filename).suffix.lower()
    logger.info(f"üìã [VALIDATE] –†–∞—Å—à–∏—Ä–µ–Ω–∏–µ —Ñ–∞–π–ª–∞: {file_ext}")
    logger.info(f"üìã [VALIDATE] –†–∞–∑—Ä–µ—à–µ–Ω–Ω—ã–µ —Ä–∞—Å—à–∏—Ä–µ–Ω–∏—è: {sorted(ALLOWED_EXTENSIONS)}")

    # –ü—Ä–æ–≤–µ—Ä–∫–∞ 2: –†–∞—Å—à–∏—Ä–µ–Ω–∏–µ —Ñ–∞–π–ª–∞
    if file_ext not in ALLOWED_EXTENSIONS:
        allowed_str = ", ".join(sorted(ALLOWED_EXTENSIONS))
        error_msg = f"–ù–µ–ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã–π —Ñ–æ—Ä–º–∞—Ç —Ñ–∞–π–ª–∞: {file_ext}. –†–∞–∑—Ä–µ—à–µ–Ω—ã: {allowed_str}"
        logger.error("=" * 70)
        logger.error(f"‚ùå [VALIDATE] –û–®–ò–ë–ö–ê –í–ê–õ–ò–î–ê–¶–ò–ò: {error_msg}")
        logger.error(f"   –§–∞–π–ª: {file.filename}")
        logger.error(f"   –†–∞—Å—à–∏—Ä–µ–Ω–∏–µ: {file_ext}")
        logger.error(f"   –†–∞–∑—Ä–µ—à–µ–Ω–Ω—ã–µ: {allowed_str}")
        logger.error("=" * 70)
        return False, error_msg

    logger.info(f"‚úÖ [VALIDATE] –†–∞—Å—à–∏—Ä–µ–Ω–∏–µ {file_ext} —Ä–∞–∑—Ä–µ—à–µ–Ω–æ")

    # –ü—Ä–æ–≤–µ—Ä–∫–∞ 3: MIME type
    content_type = getattr(file, "content_type", None)
    logger.info(f"üìã [VALIDATE] MIME type —Ñ–∞–π–ª–∞: {content_type or '–Ω–µ —É–∫–∞–∑–∞–Ω'}")

    if content_type:
        # –°–ø–µ—Ü–∏–∞–ª—å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –¥–ª—è Word —Ñ–∞–π–ª–æ–≤ (–∞–Ω–∞–ª–æ–≥–∏—á–Ω–æ .xlsm)
        if file_ext == ".docx":
            word_mime_types = [
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "application/msword",  # –°—Ç–∞—Ä—ã–µ –≤–µ—Ä—Å–∏–∏ Word
                "application/octet-stream",  # –ù–µ–∫–æ—Ç–æ—Ä—ã–µ –±—Ä–∞—É–∑–µ—Ä—ã
            ]
            if content_type in word_mime_types or content_type in ALLOWED_MIME_TYPES:
                logger.info(f"‚úÖ [VALIDATE] MIME type –¥–ª—è Word —Ñ–∞–π–ª–∞ –ø—Ä–∏–Ω—è—Ç: {content_type}")
            else:
                logger.warning(
                    f"‚ö†Ô∏è [VALIDATE] –ù–µ–æ–∂–∏–¥–∞–Ω–Ω—ã–π MIME type –¥–ª—è .docx —Ñ–∞–π–ª–∞: {content_type}, "
                    f"–Ω–æ —Ä–∞–∑—Ä–µ—à–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É (—Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ)"
                )
        # –î–ª—è Excel —Ñ–∞–π–ª–æ–≤ —Å –º–∞–∫—Ä–æ—Å–∞–º–∏ –º–æ–∂–µ—Ç –±—ã—Ç—å —Ä–∞–∑–Ω—ã–π MIME type
        elif file_ext == ".xlsm":
            # –ü—Ä–∏–Ω–∏–º–∞–µ–º –∫–∞–∫ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã–π XLSX MIME type, —Ç–∞–∫ –∏ —Å–ø–µ—Ü–∏–∞–ª—å–Ω—ã–π –¥–ª—è –º–∞–∫—Ä–æ—Å–æ–≤
            xlsm_mime_types = [
                "application/vnd.ms-excel.sheet.macroEnabled.12",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/octet-stream",  # –ù–µ–∫–æ—Ç–æ—Ä—ã–µ –±—Ä–∞—É–∑–µ—Ä—ã –æ—Ç–ø—Ä–∞–≤–ª—è—é—Ç —Ç–∞–∫
            ]
            if (
                content_type not in ALLOWED_MIME_TYPES
                and content_type not in xlsm_mime_types
            ):
                logger.warning(
                    f"‚ö†Ô∏è [VALIDATE] –ù–µ–æ–∂–∏–¥–∞–Ω–Ω—ã–π MIME type –¥–ª—è .xlsm —Ñ–∞–π–ª–∞: {content_type}, –Ω–æ —Ä–∞–∑—Ä–µ—à–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É"
                )
            else:
                logger.info(
                    f"‚úÖ [VALIDATE] MIME type –¥–ª—è .xlsm —Ñ–∞–π–ª–∞ –ø—Ä–∏–Ω—è—Ç: {content_type}"
                )
        elif content_type not in ALLOWED_MIME_TYPES:
            # –ö–∞—Ä—Ç–∞ –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã—Ö —Ä–∞–∑—Ä–µ—à–µ–Ω–Ω—ã—Ö MIME-—Ç–∏–ø–æ–≤ –¥–ª—è –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã—Ö —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–π
            EXTRA_MIME_MAP = {
                ".docx": [
                    "application/msword",  # –°—Ç–∞—Ä—ã–µ –≤–µ—Ä—Å–∏–∏ Word
                    "application/octet-stream",
                ],
                ".xlsm": [
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/octet-stream",
                ],
            }

    if not content_type:
        logger.warning(
            f"‚ö†Ô∏è [VALIDATE] MIME type –Ω–µ —É–∫–∞–∑–∞–Ω –¥–ª—è —Ñ–∞–π–ª–∞ {file.filename}, "
            f"–Ω–æ —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ {file_ext} –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ - —Ä–∞–∑—Ä–µ—à–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É"
        )
    else:
        is_allowed = content_type in ALLOWED_MIME_TYPES
        if not is_allowed and file_ext in EXTRA_MIME_MAP:
            is_allowed = content_type in EXTRA_MIME_MAP[file_ext]

        if is_allowed:
            logger.info(f"‚úÖ [VALIDATE] MIME type {content_type} —Ä–∞–∑—Ä–µ—à–µ–Ω –¥–ª—è {file_ext}")
        else:
            error_msg = f"–ù–µ–ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º—ã–π —Ç–∏–ø —Ñ–∞–π–ª–∞: {content_type}. –†–∞–∑—Ä–µ—à–µ–Ω—ã: {', '.join(sorted(ALLOWED_MIME_TYPES))}"
            logger.error("=" * 70)
            logger.error(f"‚ùå [VALIDATE] –û–®–ò–ë–ö–ê –í–ê–õ–ò–î–ê–¶–ò–ò: {error_msg}")
            logger.error(f"   –§–∞–π–ª: {file.filename}")
            logger.error(f"   –†–∞—Å—à–∏—Ä–µ–Ω–∏–µ: {file_ext}")
            logger.error(f"   MIME type: {content_type}")
            logger.error(f"   –†–∞–∑—Ä–µ—à–µ–Ω–Ω—ã–µ MIME types: {sorted(ALLOWED_MIME_TYPES)}")
            logger.error("=" * 70)
            return False, error_msg
        else:
            logger.info(f"‚úÖ [VALIDATE] MIME type {content_type} —Ä–∞–∑—Ä–µ—à–µ–Ω")
    else:
        logger.warning(
            f"‚ö†Ô∏è [VALIDATE] MIME type –Ω–µ —É–∫–∞–∑–∞–Ω –¥–ª—è —Ñ–∞–π–ª–∞ {file.filename}, "
            f"–Ω–æ —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ {file_ext} –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ - —Ä–∞–∑—Ä–µ—à–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É"
        )

    logger.info("=" * 70)
    logger.info(f"‚úÖ [VALIDATE] –í–ê–õ–ò–î–ê–¶–ò–Ø –ü–†–û–ô–î–ï–ù–ê: {file.filename}")
    logger.info("=" * 70)
    return True, None


async def validate_file_size(file: UploadFile):
    content = await file.read()
    await file.seek(0)

    if len(content) > MAX_FILE_SIZE:
        max_mb = MAX_FILE_SIZE / (1024 * 1024)
        file_mb = len(content) / (1024 * 1024)
        return (
            False,
            f"–†–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞ ({file_mb:.2f} –ú–ë) –ø—Ä–µ–≤—ã—à–∞–µ—Ç –º–∞–∫—Å–∏–º–∞–ª—å–Ω–æ –¥–æ–ø—É—Å—Ç–∏–º—ã–π ({max_mb} –ú–ë)",
        )

    if len(content) == 0:
        return False, "–§–∞–π–ª –ø—É—Å—Ç–æ–π"

    return True, None


@app.post("/web/upload")
async def upload_file(
    file: UploadFile = File(...),
    enterprise_id: Optional[int] = Form(None),
    enterprise_name: Optional[str] = Form(None),
    resource_type: Optional[str] = Form(None),
    system_mode: Optional[str] = Form(None),
):
    """–ü—Ä–∏—ë–º —Ñ–∞–π–ª–∞ —á–µ—Ä–µ–∑ –≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å —Å –ø—Ä–∏–≤—è–∑–∫–æ–π –∫ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—é"""
    if file.filename:
        file_ext = Path(file.filename).suffix.lower()
        logger.info(f"üì§ [UPLOAD] –ù–∞—á–∞–ª–æ –∑–∞–≥—Ä—É–∑–∫–∏ —Ñ–∞–π–ª–∞: {file.filename}")
        logger.info(f"üì§ [UPLOAD] –†–∞—Å—à–∏—Ä–µ–Ω–∏–µ —Ñ–∞–π–ª–∞: {file_ext}")
        logger.info(f"üì§ [UPLOAD] –†–∞–∑—Ä–µ—à–µ–Ω–Ω—ã–µ —Ä–∞—Å—à–∏—Ä–µ–Ω–∏—è: {sorted(ALLOWED_EXTENSIONS)}")
        logger.info(f"üì§ [UPLOAD] MIME type: {getattr(file, 'content_type', '–Ω–µ —É–∫–∞–∑–∞–Ω')}")

    if enterprise_name and enterprise_name.strip():
        enterprise = database.get_or_create_enterprise(enterprise_name)
    elif enterprise_id is not None:
        enterprise = database.get_enterprise_by_id(int(enterprise_id))
        if not enterprise:
            raise HTTPException(status_code=400, detail="–ü—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")
    else:
        raise HTTPException(
            status_code=400, detail="–£–∫–∞–∂–∏—Ç–µ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ –∏–ª–∏ –≤–≤–µ–¥–∏—Ç–µ –Ω–æ–≤–æ–µ"
        )

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ç–∏–ø, —É–∫–∞–∑–∞–Ω–Ω—ã–π –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º (–µ—Å–ª–∏ –µ—Å—Ç—å)
    user_provided_type = (
        resource_type if resource_type and resource_type in RESOURCE_LABELS else None
    )

    # –í—Ä–µ–º–µ–Ω–Ω–∞—è –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è resource_type –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –≤ –∫–µ—à
    # –ë—É–¥–µ—Ç –ø–µ—Ä–µ–æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∞ –ø–æ—Å–ª–µ –ø–∞—Ä—Å–∏–Ω–≥–∞ –Ω–∞ –æ—Å–Ω–æ–≤–µ —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ
    if not resource_type or resource_type not in RESOURCE_LABELS:
        resource_type = (
            "other"  # –í—Ä–µ–º–µ–Ω–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ, –±—É–¥–µ—Ç –ø–µ—Ä–µ–æ–ø—Ä–µ–¥–µ–ª–µ–Ω–æ –ø–æ—Å–ª–µ –ø–∞—Ä—Å–∏–Ω–≥–∞
        )

    is_valid, error_msg = validate_file(file)
    if not is_valid:
        logger.error(
            f"‚ùå [UPLOAD] –í–∞–ª–∏–¥–∞—Ü–∏—è –Ω–µ –ø—Ä–æ–π–¥–µ–Ω–∞ –¥–ª—è {file.filename}: {error_msg}"
        )
        raise HTTPException(status_code=400, detail=error_msg)
    logger.info(f"‚úÖ [UPLOAD] –í–∞–ª–∏–¥–∞—Ü–∏—è –ø—Ä–æ–π–¥–µ–Ω–∞ –¥–ª—è {file.filename}")

    # –ü—Ä–æ–≤–µ—Ä–∫–∞ —Ä–∞–∑–º–µ—Ä–∞ —Ñ–∞–π–ª–∞ –¥–ª—è –≤—Å–µ—Ö —Ç–∏–ø–æ–≤
    size_valid, size_error = await validate_file_size(file)
    if not size_valid:
        raise HTTPException(status_code=400, detail=size_error)

    os.makedirs(INBOX_DIR, exist_ok=True)
    batch_id = str(uuid4())
    safe_filename = os.path.basename(file.filename)
    file_ext = Path(file.filename).suffix.lower()
    dst = os.path.join(INBOX_DIR, f"{batch_id}__{safe_filename}")

    # –°–æ–∑–¥–∞–µ–º —Ç—Ä–µ–∫–µ—Ä –ø—Ä–æ–≥—Ä–µ—Å—Å–∞
    file_type_enum = get_file_type_enum(file_ext)
    tracker = create_progress_tracker(batch_id, file_type_enum)
    tracker.update_stage(
        ProcessingStage.UPLOAD, progress=0, message="–ù–∞—á–∞–ª–æ –∑–∞–≥—Ä—É–∑–∫–∏ —Ñ–∞–π–ª–∞"
    )

    try:
        file_hash = hashlib.sha1()
        file_size_total = 0
        # –ü–æ–ª—É—á–∞–µ–º —Ä–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞ –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –∑–∞–≥—Ä—É–∑–∫–∏
        file.file.seek(0, 2)  # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –≤ –∫–æ–Ω–µ—Ü
        total_size = file.file.tell()
        file.file.seek(0)  # –í–æ–∑–≤—Ä–∞—â–∞–µ–º—Å—è –≤ –Ω–∞—á–∞–ª–æ

        with open(dst, "wb") as output_file:
            while True:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞
                if tracker.is_cancelled():
                    os.remove(dst)  # –£–¥–∞–ª—è–µ–º —á–∞—Å—Ç–∏—á–Ω–æ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–π —Ñ–∞–π–ª
                    raise HTTPException(
                        status_code=499, detail="–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
                    )

                chunk = file.file.read(8192)
                if not chunk:
                    break
                output_file.write(chunk)
                file_hash.update(chunk)
                file_size_total += len(chunk)
                # –û–±–Ω–æ–≤–ª—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∑–∞–≥—Ä—É–∑–∫–∏
                if total_size > 0:
                    upload_progress = int((file_size_total / total_size) * 100)
                    tracker.update_stage(
                        ProcessingStage.UPLOAD,
                        progress=upload_progress,
                        message=f"–ó–∞–≥—Ä—É–∑–∫–∞ —Ñ–∞–π–ª–∞: {file_size_total / 1024:.1f} –ö–ë –∏–∑ {total_size / 1024:.1f} –ö–ë",
                    )

        file_digest = file_hash.hexdigest()
        file_size = os.path.getsize(dst)
        file_mtime = os.path.getmtime(dst)
        tracker.complete_stage(ProcessingStage.UPLOAD, "–§–∞–π–ª –∑–∞–≥—Ä—É–∂–µ–Ω —É—Å–ø–µ—à–Ω–æ")
    except Exception as exc:  # pragma: no cover - unexpected IO failure
        logger.exception("Failed to write uploaded file")
        tracker.set_error(ProcessingStage.UPLOAD, f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏: {exc}")
        raise HTTPException(
            status_code=500, detail=f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å —Ñ–∞–π–ª: {exc}"
        ) from exc

    file_type_label = FILE_TYPE_LABELS.get(file_ext, "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π —Ç–∏–ø")

    # –í–∞–ª–∏–¥–∞—Ü–∏—è
    tracker.update_stage(
        ProcessingStage.VALIDATION, progress=50, message="–ü—Ä–æ–≤–µ—Ä–∫–∞ —Ñ–∞–π–ª–∞..."
    )

    existing_upload = database.find_duplicate_upload(
        enterprise_id=enterprise["id"],
        filename=safe_filename,
        file_size=file_size,
        file_hash=file_digest,
    )

    if existing_upload:
        existing_batch_id = existing_upload["batch_id"]
        
        # üîß –ü–ï–†–ï–ö–õ–Æ–ß–ê–¢–ï–õ–¨ –†–ï–ñ–ò–ú–û–í: debug / production
        # –ü—Ä–∏–æ—Ä–∏—Ç–µ—Ç: –ø–∞—Ä–∞–º–µ—Ç—Ä –∏–∑ —Ñ–æ—Ä–º—ã > –ø–µ—Ä–µ–º–µ–Ω–Ω–∞—è –æ–∫—Ä—É–∂–µ–Ω–∏—è > production
        current_mode = system_mode.lower() if system_mode else SYSTEM_MODE
        if current_mode not in ["debug", "production"]:
            current_mode = SYSTEM_MODE
        
        if current_mode == "debug":
            # –†–ï–ñ–ò–ú –û–¢–õ–ê–î–ö–ò: –≤—Å–µ–≥–¥–∞ –ø–µ—Ä–µ–æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—Ç—å
            logger.info(
                f"üîß [DEBUG MODE] –î—É–±–ª–∏–∫–∞—Ç –Ω–∞–π–¥–µ–Ω (batch_id={existing_batch_id}), "
                f"–ø–µ—Ä–µ–æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª (—Ä–µ–∂–∏–º –æ—Ç–ª–∞–¥–∫–∏, —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω —á–µ—Ä–µ–∑ {'–≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å' if system_mode else '–ø–µ—Ä–µ–º–µ–Ω–Ω—É—é –æ–∫—Ä—É–∂–µ–Ω–∏—è'})"
            )
            deleted = database.delete_upload_by_batch_id(existing_batch_id)
            if deleted:
                logger.info(f"‚úÖ –°—Ç–∞—Ä–∞—è –∑–∞–ø–∏—Å—å —É–¥–∞–ª–µ–Ω–∞, –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª –∑–∞–Ω–æ–≤–æ")
            # –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É
        else:
            # –†–ï–ñ–ò–ú –†–ê–ë–û–¢–´: –ø—Ä–æ–≤–µ—Ä—è–µ–º –∏–∑–º–µ–Ω–µ–Ω–∏—è –ø–æ hash
            existing_hash = existing_upload.get("file_hash", "")
            
            if existing_hash == file_digest:
                # Hash —Å–æ–≤–ø–∞–¥–∞–µ—Ç - —Ñ–∞–π–ª –Ω–µ –∏–∑–º–µ–Ω–∏–ª—Å—è, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º
                logger.info(
                    f"‚úÖ –ù–∞–π–¥–µ–Ω –¥—É–±–ª–∏–∫–∞—Ç {safe_filename} –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise['name']} "
                    f"(batch_id={existing_batch_id}). –§–∞–π–ª –Ω–µ –∏–∑–º–µ–Ω–∏–ª—Å—è (hash —Å–æ–≤–ø–∞–¥–∞–µ—Ç), –ø—Ä–æ–ø—É—Å–∫–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É"
                )
                os.remove(dst)
                tracker.complete_stage(
                    ProcessingStage.VALIDATION, "–§–∞–π–ª —É–∂–µ –±—ã–ª –∑–∞–≥—Ä—É–∂–µ–Ω —Ä–∞–Ω–µ–µ (–±–µ–∑ –∏–∑–º–µ–Ω–µ–Ω–∏–π)"
                )
                tracker.complete()
                
                # –û—á–∏—â–∞–µ–º —Ç—Ä–µ–∫–µ—Ä –¥–ª—è –¥—É–±–ª–∏–∫–∞—Ç–∞
                import threading
                def cleanup_tracker():
                    import time
                    time.sleep(60)
                    remove_progress_tracker(batch_id)
                threading.Thread(target=cleanup_tracker, daemon=True).start()
                
                return {
                    "batch_id": existing_batch_id,
                    "saved": existing_upload["filename"],
                    "file_type": existing_upload["file_type"],
                    "file_size": existing_upload["file_size"],
                    "parsing_status": existing_upload["status"],
                    "enterprise": {"id": enterprise["id"], "name": enterprise["name"]},
                    "parsing_summary": existing_upload.get("parsing_summary"),
                    "duplicate": True,
                    "skipped": True,
                    "reason": "–§–∞–π–ª –Ω–µ –∏–∑–º–µ–Ω–∏–ª—Å—è (hash —Å–æ–≤–ø–∞–¥–∞–µ—Ç)"
                }
            else:
                # Hash –æ—Ç–ª–∏—á–∞–µ—Ç—Å—è - —Ñ–∞–π–ª –æ–±–Ω–æ–≤–ª–µ–Ω, –ø–µ—Ä–µ–æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º
                logger.info(
                    f"üîÑ –î—É–±–ª–∏–∫–∞—Ç –Ω–∞–π–¥–µ–Ω, –Ω–æ —Ñ–∞–π–ª –æ–±–Ω–æ–≤–ª–µ–Ω (hash –∏–∑–º–µ–Ω–∏–ª—Å—è). "
                    f"–ü–µ—Ä–µ–æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª (batch_id={existing_batch_id})"
                )
                deleted = database.delete_upload_by_batch_id(existing_batch_id)
                if deleted:
                    logger.info(f"‚úÖ –°—Ç–∞—Ä–∞—è –∑–∞–ø–∏—Å—å —É–¥–∞–ª–µ–Ω–∞, –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –æ–±–Ω–æ–≤–ª–µ–Ω–Ω—ã–π —Ñ–∞–π–ª")
                # –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É

    tracker.complete_stage(ProcessingStage.VALIDATION, "–í–∞–ª–∏–¥–∞—Ü–∏—è –ø—Ä–æ–π–¥–µ–Ω–∞")

    parsing_result: Optional[Dict[str, Any]] = None
    parsing_error: Optional[str] = None
    try:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞ –ø–µ—Ä–µ–¥ –Ω–∞—á–∞–ª–æ–º –ø–∞—Ä—Å–∏–Ω–≥–∞
        if tracker.is_cancelled():
            os.remove(dst)  # –£–¥–∞–ª—è–µ–º —Ñ–∞–π–ª –ø—Ä–∏ –æ—Ç–º–µ–Ω–µ
            raise HTTPException(
                status_code=499, detail="–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
            )

        logger.info("–ù–∞—á–∏–Ω–∞—é –ø–∞—Ä—Å–∏–Ω–≥ —Ñ–∞–π–ª–∞: %s", dst)
        tracker.update_stage(
            ProcessingStage.PARSING, progress=10, message="–ù–∞—á–∞–ª–æ –ø–∞—Ä—Å–∏–Ω–≥–∞ —Ñ–∞–π–ª–∞..."
        )

        # –î–ª—è –±–æ–ª—å—à–∏—Ö —Ñ–∞–π–ª–æ–≤ –æ–±–Ω–æ–≤–ª—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –ø–∞—Ä—Å–∏–Ω–≥–∞
        if file_type_enum == FileType.EXCEL:
            tracker.update_stage(
                ProcessingStage.PARSING, progress=30, message="–ß—Ç–µ–Ω–∏–µ –ª–∏—Å—Ç–æ–≤ Excel..."
            )
        elif file_type_enum == FileType.PDF:
            tracker.update_stage(
                ProcessingStage.PARSING,
                progress=30,
                message="–ò–∑–≤–ª–µ—á–µ–Ω–∏–µ —Ç–µ–∫—Å—Ç–∞ –∏–∑ PDF...",
            )
            # –î–ª—è PDF –º–æ–∂–µ—Ç –ø–æ—Ç—Ä–µ–±–æ–≤–∞—Ç—å—Å—è OCR, –ø–æ—ç—Ç–æ–º—É –∑–∞—Ä–∞–Ω–µ–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º —ç—Ç–∞–ø
            tracker.update_stage(
                ProcessingStage.OCR, progress=0, message="–ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ—Å—Ç–∏ OCR..."
            )
        elif file_type_enum == FileType.IMAGE:
            tracker.update_stage(
                ProcessingStage.PARSING, progress=50, message="–ó–∞–≥—Ä—É–∑–∫–∞ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è..."
            )

        # –ü–µ—Ä–µ–¥–∞–µ–º batch_id –≤ parse_file –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –æ—Ç–º–µ–Ω—ã
        try:
            parsing_result = parse_file(dst, batch_id=batch_id)
        except InterruptedError as cancel_exc:
            # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º
            logger.info(f"–ü–∞—Ä—Å–∏–Ω–≥ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω: {cancel_exc}")
            tracker.set_error(
                ProcessingStage.PARSING, "–û–±—Ä–∞–±–æ—Ç–∫–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
            )
            os.remove(dst)  # –£–¥–∞–ª—è–µ–º —Ñ–∞–π–ª –ø—Ä–∏ –æ—Ç–º–µ–Ω–µ
            raise HTTPException(
                status_code=499, detail="–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
            )

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª—Å—è –ª–∏ OCR –¥–ª—è PDF
        pdf_data = parsing_result.get("data", {})
        if file_type_enum == FileType.PDF:
            is_scanned = pdf_data.get("is_scanned", False)
            ocr_attempted = pdf_data.get("ocr_attempted", False)
            ocr_success = pdf_data.get("ocr_success", False)
            ocr_error = pdf_data.get("ocr_error")

            if is_scanned:
                if ocr_attempted:
                    if ocr_success:
                        # OCR —É—Å–ø–µ—à–Ω–æ –ø—Ä–∏–º–µ–Ω–µ–Ω
                        char_count = pdf_data.get("total_characters", 0)
                        tracker.update_stage(
                            ProcessingStage.OCR,
                            progress=100,
                            message=f"OCR —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏–µ –∑–∞–≤–µ—Ä—à–µ–Ω–æ ({char_count} —Å–∏–º–≤–æ–ª–æ–≤)",
                            metadata={
                                "char_count": char_count,
                                "pages": pdf_data.get("metadata", {}).get(
                                    "num_pages", 0
                                ),
                            },
                        )
                        tracker.complete_stage(
                            ProcessingStage.OCR, "OCR —É—Å–ø–µ—à–Ω–æ –ø—Ä–∏–º–µ–Ω–µ–Ω"
                        )
                    else:
                        # OCR –Ω–µ —É–¥–∞–ª—Å—è (poppler –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω –∏–ª–∏ –¥—Ä—É–≥–∞—è –æ—à–∏–±–∫–∞)
                        if ocr_error == "poppler_not_installed":
                            tracker.update_stage(
                                ProcessingStage.OCR,
                                progress=0,
                                message="OCR –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω: poppler –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω. –§–∞–π–ª –æ–±—Ä–∞–±–æ—Ç–∞–Ω –±–µ–∑ OCR.",
                                metadata={
                                    "error": "poppler_not_installed",
                                    "is_scanned": True,
                                },
                            )
                        else:
                            tracker.update_stage(
                                ProcessingStage.OCR,
                                progress=0,
                                message="OCR –Ω–µ –∏–∑–≤–ª–µ–∫ —Ç–µ–∫—Å—Ç. –§–∞–π–ª –æ–±—Ä–∞–±–æ—Ç–∞–Ω –±–µ–∑ OCR.",
                                metadata={"error": "ocr_failed", "is_scanned": True},
                            )
                else:
                    # –°–∫–∞–Ω–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –¥–æ–∫—É–º–µ–Ω—Ç, –Ω–æ OCR –Ω–µ –±—ã–ª –ø–æ–ø—ã—Ç–∫–∞ (HAS_OCR = False)
                    tracker.update_stage(
                        ProcessingStage.OCR,
                        progress=0,
                        message="–û–±–Ω–∞—Ä—É–∂–µ–Ω —Å–∫–∞–Ω–∏—Ä–æ–≤–∞–Ω–Ω—ã–π PDF, –Ω–æ OCR –±–∏–±–ª–∏–æ—Ç–µ–∫–∏ –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω—ã.",
                        metadata={"is_scanned": True, "error": "ocr_not_available"},
                    )

        tracker.complete_stage(ProcessingStage.PARSING, "–ü–∞—Ä—Å–∏–Ω–≥ –∑–∞–≤–µ—Ä—à–µ–Ω")
        status_value = "success" if parsing_result.get("parsed") else "partial"

        # –î–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ AI-–∞–Ω–∞–ª–∏–∑–µ –≤ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
        parsing_result_data = parsing_result.copy()
        if "ai_analysis" in parsing_result:
            # –î–æ–±–∞–≤–ª—è–µ–º –∫—Ä–∞—Ç–∫—É—é —Å–≤–æ–¥–∫—É AI-–∞–Ω–∞–ª–∏–∑–∞
            ai_analysis = parsing_result.get("ai_analysis", {})
            parsing_result_data["ai_analysis_summary"] = {
                "confidence_score": ai_analysis.get("confidence_score", 0.0),
                "is_valid": ai_analysis.get("summary", {}).get("is_valid", True),
                "has_anomalies": ai_analysis.get("summary", {}).get(
                    "has_anomalies", False
                ),
                "is_compliant": ai_analysis.get("summary", {}).get(
                    "is_compliant", True
                ),
                "anomaly_count": ai_analysis.get("anomalies", {}).get(
                    "anomaly_count", 0
                ),
                "efficiency_class": ai_analysis.get("summary", {}).get(
                    "efficiency_class", "N/A"
                ),
            }

        # –ö–†–ò–¢–ò–ß–ï–°–ö–û–ï –ò–ó–ú–ï–ù–ï–ù–ò–ï: –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø —Ä–µ—Å—É—Ä—Å–∞ –ü–û–°–õ–ï –ø–∞—Ä—Å–∏–Ω–≥–∞
        # –Ω–∞ –æ—Å–Ω–æ–≤–µ —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ —Ñ–∞–π–ª–∞
        from utils.resource_classifier import ResourceClassifier

        # –§–æ—Ä–º–∏—Ä—É–µ–º raw_json –¥–ª—è –∞–Ω–∞–ª–∏–∑–∞
        raw_json_for_analysis = {
            "file_type": file_type_label.lower(),
            "parsing": parsing_result_data,
        }

        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø —Ä–µ—Å—É—Ä—Å–∞ —Å –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç–æ–º —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ
        detected_resource_type = ResourceClassifier.classify(
            filename=safe_filename,
            raw_json=raw_json_for_analysis if parsing_result.get("parsed") else None,
            user_provided_type=user_provided_type,
        )

        # –û–±–Ω–æ–≤–ª—è–µ–º resource_type –Ω–∞ –æ—Å–Ω–æ–≤–µ –∞–Ω–∞–ª–∏–∑–∞ —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ
        resource_type = detected_resource_type
        logger.info(
            f"–¢–∏–ø —Ä–µ—Å—É—Ä—Å–∞ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω –ø–æ—Å–ª–µ –ø–∞—Ä—Å–∏–Ω–≥–∞ –¥–ª—è {safe_filename}: {resource_type} "
            f"(–ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–∫–∞–∑–∞–ª: {user_provided_type or '–Ω–µ —É–∫–∞–∑–∞–Ω'})"
        )

        # üß† –ò–ù–¢–ï–õ–õ–ï–ö–¢–£–ê–õ–¨–ù–´–ô –ú–ê–†–®–†–£–¢–ò–ó–ê–¢–û–†: –ê–Ω–∞–ª–∏–∑ —Ñ–∞–π–ª–∞ –∏ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ –æ–ø—Ç–∏–º–∞–ª—å–Ω–æ–≥–æ –ø—É—Ç–∏ –æ–±—Ä–∞–±–æ—Ç–∫–∏
        routing_map: Optional[Dict[str, Any]] = None
        try:
            router = IntelligentRouter()
            raw_json_for_routing = {
                "file_type": file_type_label.lower(),
                "parsing": parsing_result_data if parsing_result.get("parsed") else None,
            }
            routing_map = router.analyze_file(
                file_path=dst,
                filename=safe_filename,
                raw_json=raw_json_for_routing if parsing_result.get("parsed") else None,
                fast_mode=True
            )
            
            # –ï—Å–ª–∏ —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å –Ω–∏–∑–∫–∞—è, –≤—ã–ø–æ–ª–Ω—è–µ–º –≥–ª—É–±–æ–∫–∏–π –∞–Ω–∞–ª–∏–∑
            if routing_map.get("analysis", {}).get("confidence", 0.0) < 0.7:
                logger.info(f"‚ö†Ô∏è –ù–∏–∑–∫–∞—è —É–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å –∞–Ω–∞–ª–∏–∑–∞ ({routing_map.get('analysis', {}).get('confidence', 0.0):.2f}), –≤—ã–ø–æ–ª–Ω—è–µ–º –≥–ª—É–±–æ–∫–∏–π –∞–Ω–∞–ª–∏–∑")
                routing_map = router.analyze_file(
                    file_path=dst,
                    filename=safe_filename,
                    raw_json=raw_json_for_routing if parsing_result.get("parsed") else None,
                    fast_mode=False
                )
            
            logger.info(
                f"üß† Intelligent Router: document_type={routing_map.get('analysis', {}).get('document_type', 'unknown')}, "
                f"resource_type={routing_map.get('analysis', {}).get('resource_type', 'unknown')}, "
                f"data_type={routing_map.get('analysis', {}).get('data_type', 'unknown')}, "
                f"confidence={routing_map.get('analysis', {}).get('confidence', 0.0):.2f}, "
                f"primary_module={routing_map.get('routing', {}).get('primary_module', 'unknown')}"
            )
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º routing_map –≤ parsing_result –¥–ª—è –¥–∞–ª—å–Ω–µ–π—à–µ–≥–æ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è
            if parsing_result:
                parsing_result["routing_map"] = routing_map
            
        except Exception as router_exc:
            logger.warning(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ —Ä–∞–±–æ—Ç–µ Intelligent Router: {router_exc}", exc_info=True)
            # –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É –±–µ–∑ routing_map

        parsing_results_cache[batch_id] = {
            "batch_id": batch_id,
            "filename": safe_filename,
            "file_path": dst,
            "file_type": file_type_label,
            "file_size": file_size,
            "parsing": parsing_result_data,
            "status": status_value,
        }
        if resource_type:
            parsing_results_cache[batch_id]["resource_type"] = resource_type
        logger.info(
            "–ü–∞—Ä—Å–∏–Ω–≥ –∑–∞–≤–µ—Ä—à–µ–Ω –¥–ª—è batch_id=%s, parsed=%s, resource_type=%s",
            batch_id,
            parsing_result.get("parsed"),
            resource_type,
        )
    except Exception as exc:  # pragma: no cover - parse failure path
        parsing_error = str(exc)
        logger.exception("–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–∞—Ä—Å–∏–Ω–≥–µ —Ñ–∞–π–ª–∞ %s", dst)
        tracker.set_error(ProcessingStage.PARSING, f"–û—à–∏–±–∫–∞ –ø–∞—Ä—Å–∏–Ω–≥–∞: {exc}")
        parsing_results_cache[batch_id] = {
            "batch_id": batch_id,
            "filename": safe_filename,
            "file_path": dst,
            "file_type": file_type_label,
            "file_size": file_size,
            "parsing": None,
            "error": parsing_error,
            "status": "error",
        }
        if resource_type:
            parsing_results_cache[batch_id]["resource_type"] = resource_type

    parsing_summary = build_parsing_summary(parsing_result, file_ext)
    parsing_status = (
        "success"
        if parsing_result and parsing_result.get("parsed")
        else ("error" if parsing_error else "pending")
    )

    response_data: Dict[str, Any] = {
        "batch_id": batch_id,
        "saved": os.path.basename(dst),
        "file_type": file_type_label,
        "file_size": file_size,
        "parsing_status": parsing_status,
        "enterprise": {"id": enterprise["id"], "name": enterprise["name"]},
    }
    if resource_type:
        response_data["resource_type"] = resource_type
    if parsing_summary:
        response_data["parsing_summary"] = parsing_summary
    if routing_map:
        response_data["routing_map"] = routing_map

    try:
        summary_payload = dict(parsing_summary or {})
        if resource_type:
            summary_payload["resource_type"] = resource_type
            summary_payload["resource_type_label"] = RESOURCE_LABELS.get(
                resource_type, resource_type
            )

        # –î–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –≥–æ—Ç–æ–≤–æ–º Word-–æ—Ç—á–µ—Ç–µ
        if parsing_result and file_type_enum == FileType.WORD:
            word_data = parsing_result.get("data", {})
            if word_data.get("is_ready_report", False):
                summary_payload["is_ready_report"] = True
                summary_payload["table_count"] = word_data.get("table_count", 0)
                summary_payload["report_type"] = "ready_word_report"

        # –î–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–∏ OCR –¥–ª—è PDF
        if parsing_result and file_type_enum == FileType.PDF:
            # –ò—Å–ø–æ–ª—å–∑—É–µ–º pdf_data, —Ç–∞–∫ –∫–∞–∫ –æ–Ω–∞ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∞ –≤—ã—à–µ
            summary_payload["ocr_used"] = pdf_data.get("ocr_used", False)
            summary_payload["is_scanned"] = pdf_data.get("is_scanned", False)
            summary_payload["ocr_success"] = pdf_data.get("ocr_success", False)

        # –î–æ–±–∞–≤–ª—è–µ–º routing_map –≤ summary_payload –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –≤ –ë–î
        if routing_map:
            summary_payload["routing_map"] = {
                "document_type": routing_map.get("analysis", {}).get("document_type"),
                "resource_type": routing_map.get("analysis", {}).get("resource_type"),
                "data_type": routing_map.get("analysis", {}).get("data_type"),
                "period": routing_map.get("analysis", {}).get("period"),
                "confidence": routing_map.get("analysis", {}).get("confidence"),
                "primary_module": routing_map.get("routing", {}).get("primary_module"),
                "target_tables": routing_map.get("routing", {}).get("target_tables", []),
            }

        if summary_payload:
            response_data["parsing_summary"] = summary_payload

        database.create_upload(
            batch_id=batch_id,
            enterprise_id=enterprise["id"],
            filename=safe_filename,
            file_type=file_type_label,
            file_size=file_size,
            status=parsing_status,
            parsing_summary=summary_payload if summary_payload else None,
            file_hash=file_digest,
            file_mtime=file_mtime,
        )

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —è–≤–ª—è–µ—Ç—Å—è –ª–∏ —Ñ–∞–π–ª –≥–æ—Ç–æ–≤—ã–º Word-–æ—Ç—á–µ—Ç–æ–º
        # –î–ª—è –≥–æ—Ç–æ–≤—ã—Ö –æ—Ç—á–µ—Ç–æ–≤ –ø—Ä–æ–ø—É—Å–∫–∞–µ–º –∞–≥—Ä–µ–≥–∞—Ü–∏—é –∏ —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –ø–∞—Ä—Å–∏–Ω–≥
        is_ready_word_report = False
        if parsing_result and file_type_enum == FileType.WORD:
            is_ready_word_report = parsing_result.get("data", {}).get(
                "is_ready_report", False
            )
            if is_ready_word_report:
                logger.info(
                    f"–û–±–Ω–∞—Ä—É–∂–µ–Ω –≥–æ—Ç–æ–≤—ã–π Word –æ—Ç—á–µ—Ç, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º –∞–≥—Ä–µ–≥–∞—Ü–∏—é –∏ —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –ø–∞—Ä—Å–∏–Ω–≥ (—Ç–∞–±–ª–∏—Ü: {parsing_result.get('data', {}).get('table_count', 0)})"
                )

        aggregation_info: Optional[Dict[str, Any]] = None
        equipment_info: Optional[Dict[str, Any]] = None
        envelope_info: Optional[Dict[str, Any]] = None
        nodes_info: Optional[Dict[str, Any]] = None

        # –ò—Å–ø–æ–ª—å–∑—É–µ–º raw_json –¥–ª—è —É–ª—É—á—à–µ–Ω–Ω–æ–≥–æ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è —Ç–∏–ø–∞ (–Ω—É–∂–Ω–æ –ü–ï–†–ï–î –ø—Ä–æ–≤–µ—Ä–∫–æ–π —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤)
        raw_json_for_check = (
            {"file_type": file_type_label.lower(), "parsing": parsing_result_data}
            if parsing_result.get("parsed")
            else None
        )

        # –î–ª—è –≥–æ—Ç–æ–≤—ã—Ö Word-–æ—Ç—á–µ—Ç–æ–≤ –ø—Ä–æ–ø—É—Å–∫–∞–µ–º –∞–≥—Ä–µ–≥–∞—Ü–∏—é
        # –í–ê–ñ–ù–û: –ù–µ –∞–≥—Ä–µ–≥–∏—Ä—É–µ–º —Ñ–∞–π–ª—ã —É–∑–ª–æ–≤ —É—á—ë—Ç–∞, –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –∏ –¥—Ä—É–≥–∏—Ö —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö —Ç–∏–ø–æ–≤
        is_specialized_file = (
            is_nodes_file(safe_filename, raw_json_for_check)
            or is_equipment_file(safe_filename, raw_json_for_check)
            or is_envelope_file(safe_filename, raw_json_for_check)
        )
        
        if is_specialized_file:
            logger.info(
                f"üìã –§–∞–π–ª {safe_filename} –æ–ø—Ä–µ–¥–µ–ª–µ–Ω –∫–∞–∫ —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π (nodes/equipment/envelope), "
                f"–ø—Ä–æ–ø—É—Å–∫–∞–µ–º –∞–≥—Ä–µ–≥–∞—Ü–∏—é —ç–Ω–µ—Ä–≥–æ—Ä–µ—Å—É—Ä—Å–æ–≤"
            )
        
        if should_aggregate_file(safe_filename) and not is_ready_word_report and not is_specialized_file:
            try:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞
                if tracker.is_cancelled():
                    raise InterruptedError("–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")

                tracker.update_stage(
                    ProcessingStage.AGGREGATION,
                    progress=10,
                    message="–ê–≥—Ä–µ–≥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö...",
                )
                aggregation_data = aggregate_energy_data(dst)

                # –ï—Å–ª–∏ –∞–≥—Ä–µ–≥–∞—Ü–∏—è –∏–∑ —Ñ–∞–π–ª–∞ –Ω–µ —É–¥–∞–ª–∞—Å—å, –ø—Ä–æ–±—É–µ–º –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞—Ç—å –∏–∑ —Ä–∞—Å–ø–∞—Ä—Å–µ–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö (–ë–î)
                # –ù–û —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ —ç—Ç–æ –Ω–µ —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π —Ñ–∞–π–ª
                if (
                    not aggregation_data
                    and parsing_result
                    and parsing_result.get("parsed")
                    and not is_specialized_file
                ):
                    logger.info(
                        "‚ö†Ô∏è [DIAG] –ê–≥—Ä–µ–≥–∞—Ü–∏—è –∏–∑ —Ñ–∞–π–ª–∞ –Ω–µ —É–¥–∞–ª–∞—Å—å, –ø—Ä–æ–±—É—é –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞—Ç—å –∏–∑ —Ä–∞—Å–ø–∞—Ä—Å–µ–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö"
                    )
                    raw_json_for_aggregation = {
                        "batch_id": batch_id,
                        "filename": safe_filename,
                        "file_type": file_type_label.lower(),
                        "parsing": parsing_result_data,
                    }
                    aggregation_data = aggregate_from_db_json(raw_json_for_aggregation)
                    if aggregation_data:
                        logger.info(
                            "‚úÖ [DIAG] –ê–≥—Ä–µ–≥–∞—Ü–∏—è –∏–∑ —Ä–∞—Å–ø–∞—Ä—Å–µ–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö —É—Å–ø–µ—à–Ω–∞"
                        )
                        logger.info(
                            f"üìä [DIAG] –°—Ç—Ä—É–∫—Ç—É—Ä–∞ –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö: resources={list(aggregation_data.get('resources', {}).keys())}"
                        )
                    else:
                        logger.warning(
                            "‚ö†Ô∏è [DIAG] –ê–≥—Ä–µ–≥–∞—Ü–∏—è –∏–∑ —Ä–∞—Å–ø–∞—Ä—Å–µ–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö —Ç–∞–∫–∂–µ –Ω–µ —É–¥–∞–ª–∞—Å—å"
                        )

                if aggregation_data:
                    # –ê–≥—Ä–µ–≥–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ –∫–∞—Ç–µ–≥–æ—Ä–∏—è–º –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è (by_usage)
                    tracker.update_stage(
                        ProcessingStage.AGGREGATION,
                        progress=50,
                        message="–û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–∞—Ç–µ–≥–æ—Ä–∏–π –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è...",
                    )

                    # –ü—ã—Ç–∞–µ–º—Å—è –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –¥–ª—è —Å–≤–µ—Ä–∫–∏
                    equipment_data_for_categories = None
                    equipment_json_path = AGGREGATED_DIR / f"{batch_id}_equipment.json"
                    if equipment_json_path.exists():
                        try:
                            equipment_data_for_categories = json.loads(
                                equipment_json_path.read_text(encoding="utf-8")
                            )
                            logger.info(
                                f"–ó–∞–≥—Ä—É–∂–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –¥–ª—è —Å–≤–µ—Ä–∫–∏ –∫–∞—Ç–µ–≥–æ—Ä–∏–π: {equipment_json_path}"
                            )
                        except Exception as eq_exc:
                            logger.warning(
                                f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –¥–ª—è —Å–≤–µ—Ä–∫–∏: {eq_exc}"
                            )

                    usage_categories = aggregate_usage_categories(
                        dst, equipment_data=equipment_data_for_categories
                    )
                    if usage_categories:
                        # –°–æ—Ö—Ä–∞–Ω—è–µ–º usage_categories.json (–æ–∂–∏–¥–∞–µ—Ç—Å—è –≤–∞–ª–∏–¥–∞—Ç–æ—Ä–æ–º)
                        usage_path = AGGREGATED_DIR / "usage_categories.json"
                        try:
                            usage_path.write_text(
                                json.dumps(
                                    usage_categories, ensure_ascii=False, indent=2
                                ),
                                encoding="utf-8",
                            )
                            logger.info(
                                f"–°–æ—Ö—Ä–∞–Ω–µ–Ω—ã –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è: {usage_path}"
                            )
                        except Exception as usage_save_exc:
                            logger.warning(
                                f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å usage_categories.json: {usage_save_exc}"
                            )

                        # –†–∞—Å–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ –ø–æ –∫–≤–∞—Ä—Ç–∞–ª–∞–º –∏ –¥–æ–±–∞–≤–ª—è–µ–º –≤ aggregation_data
                        logger.info("üìä [DIAG] –î–∞–Ω–Ω—ã–µ –∞–≥—Ä–µ–≥–∞—Ü–∏–∏ –î–û —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è –ø–æ –∫–∞—Ç–µ–≥–æ—Ä–∏—è–º: "
                                    f"keys={list(aggregation_data.get('resources', {}).keys())}")
                        
                        aggregation_data = distribute_categories_by_quarter(
                            aggregation_data, usage_categories
                        )
                        
                        # –ö–ª—é—á–µ–≤–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –ø–æ—Å–ª–µ —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è
                        electricity_data_after = aggregation_data.get("resources", {}).get("electricity")
                        if electricity_data_after:
                            logger.info("‚úÖ [DIAG] –î–∞–Ω–Ω—ã–µ –ø–æ —ç–ª–µ–∫—Ç—Ä–æ—ç–Ω–µ—Ä–≥–∏–∏ –ü–†–ò–°–£–¢–°–¢–í–£–Æ–¢ –ø–æ—Å–ª–µ —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è –ø–æ –∫–∞—Ç–µ–≥–æ—Ä–∏—è–º. "
                                        f"–ö–æ–ª-–≤–æ –∑–∞–ø–∏—Å–µ–π: {len(electricity_data_after)}")
                        else:
                            logger.warning("‚ö†Ô∏è [DIAG] –î–∞–Ω–Ω—ã–µ –ø–æ —ç–ª–µ–∫—Ç—Ä–æ—ç–Ω–µ—Ä–≥–∏–∏ –û–¢–°–£–¢–°–¢–í–£–Æ–¢ –ø–æ—Å–ª–µ —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è –ø–æ –∫–∞—Ç–µ–≥–æ—Ä–∏—è–º.")
                        
                        logger.info("–ö–∞—Ç–µ–≥–æ—Ä–∏–∏ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω—ã –ø–æ –∫–≤–∞—Ä—Ç–∞–ª–∞–º")

                    aggregated_file = write_aggregation_json(
                        batch_id, aggregation_data, AGGREGATED_DIR
                    )
                    log_aggregation_event(
                        batch_id=batch_id,
                        source_file=dst,
                        output_file=aggregated_file,
                        status="success",
                    )
                    
                    # –ò–º–ø–æ—Ä—Ç –≤—Å–µ—Ö —Ä–µ—Å—É—Ä—Å–æ–≤ –≤ –ë–î
                    logger.info("=" * 70)
                    logger.info(f"üöÄ –ù–ê–ß–ê–õ–û –ò–ú–ü–û–†–¢–ê –í –ë–î –¥–ª—è batch_id={batch_id}")
                    logger.info("=" * 70)
                    try:
                        resources = aggregation_data.get("resources", {})
                        logger.info(f"üì¶ –ù–∞–π–¥–µ–Ω–æ —Ä–µ—Å—É—Ä—Å–æ–≤ –¥–ª—è –∏–º–ø–æ—Ä—Ç–∞: {list(resources.keys())}")
                        imported_total = 0
                        
                        # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º –≤—Å–µ –¥–æ—Å—Ç—É–ø–Ω—ã–µ —Ä–µ—Å—É—Ä—Å—ã
                        for resource_type, resource_data in resources.items():
                            if resource_data:
                                logger.info(f"üì• –ò–º–ø–æ—Ä—Ç–∏—Ä—É—é —Ä–µ—Å—É—Ä—Å: {resource_type}")
                                logger.info(f"   –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø–µ—Ä–∏–æ–¥–æ–≤: {len(resource_data)}")
                                
                                imported_records = database.import_resource_to_db(
                                    enterprise_id=enterprise["id"],
                                    batch_id=batch_id,
                                    resource_type=resource_type,
                                    resource_data=resource_data,
                                )
                                imported_total += len(imported_records)
                                logger.info(
                                    f"‚úÖ –ò–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {len(imported_records)} –∑–∞–ø–∏—Å–µ–π {resource_type} –≤ –ë–î –¥–ª—è batch_id={batch_id}"
                                )
                                # –î–µ—Ç–∞–ª—å–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ –∫–∞–∂–¥–æ–π –∑–∞–ø–∏—Å–∏
                                for record in imported_records[:5]:  # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –ø–µ—Ä–≤—ã–µ 5
                                    logger.info(f"   ‚Üí –ó–∞–ø–∏—Å—å: {record.get('resource_type')} / {record.get('period')} (ID: {record.get('id')})")
                                if len(imported_records) > 5:
                                    logger.info(f"   ... –∏ –µ—â–µ {len(imported_records) - 5} –∑–∞–ø–∏—Å–µ–π")
                            else:
                                logger.warning(f"‚ö†Ô∏è –†–µ—Å—É—Ä—Å {resource_type} –ø—É—Å—Ç–æ–π, –ø—Ä–æ–ø—É—Å–∫–∞—é")
                        
                        if imported_total > 0:
                            logger.info("=" * 70)
                            logger.info(
                                f"‚úÖ –ò–ú–ü–û–†–¢ –ó–ê–í–ï–†–®–ï–ù: –í—Å–µ–≥–æ –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {imported_total} –∑–∞–ø–∏—Å–µ–π —Ä–µ—Å—É—Ä—Å–æ–≤ –≤ –ë–î –¥–ª—è batch_id={batch_id}"
                            )
                            logger.info("=" * 70)
                        else:
                            logger.warning(
                                f"‚ö†Ô∏è –î–∞–Ω–Ω—ã–µ —Ä–µ—Å—É—Ä—Å–æ–≤ –æ—Ç—Å—É—Ç—Å—Ç–≤—É—é—Ç –≤ –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –¥–ª—è batch_id={batch_id}"
                            )
                    except Exception as import_exc:
                        logger.error("=" * 70)
                        logger.error(
                            f"‚ùå –û–®–ò–ë–ö–ê –ü–†–ò –ò–ú–ü–û–†–¢–ï –†–ï–°–£–†–°–û–í –í –ë–î –¥–ª—è batch_id={batch_id}: {import_exc}",
                            exc_info=True,
                        )
                        logger.error("=" * 70)
                    
                    aggregation_info = {
                        "output_file": aggregated_file.name,
                        "missing_sheets": aggregation_data.get("missing_sheets", []),
                        "resource_quarters": {
                            resource: len(entries)
                            for resource, entries in aggregation_data[
                                "resources"
                            ].items()
                        },
                    }
                    tracker.complete_stage(
                        ProcessingStage.AGGREGATION, "–ê–≥—Ä–µ–≥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö –∑–∞–≤–µ—Ä—à–µ–Ω–∞"
                    )
                else:
                    log_aggregation_event(
                        batch_id=batch_id,
                        source_file=dst,
                        output_file=None,
                        status="skipped",
                        message="Aggregator returned no data",
                    )
            except InterruptedError:
                # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞
                logger.info("–ê–≥—Ä–µ–≥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
                raise
            except Exception as agg_exc:  # pragma: no cover - aggregation failure path
                logger.exception("Aggregation failed for %s", dst)
                log_aggregation_event(
                    batch_id=batch_id,
                    source_file=dst,
                    output_file=None,
                    status="error",
                    message=str(agg_exc),
                )
        # –î–ª—è –≥–æ—Ç–æ–≤—ã—Ö Word-–æ—Ç—á–µ—Ç–æ–≤ –ø—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –ø–∞—Ä—Å–∏–Ω–≥
        # raw_json_for_check —É–∂–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω –≤—ã—à–µ

        if (
            is_equipment_file(safe_filename, raw_json_for_check)
            and not is_ready_word_report
        ):
            try:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞
                if tracker.is_cancelled():
                    raise InterruptedError("–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")

                tracker.update_stage(
                    ProcessingStage.SPECIALIZED_PARSING,
                    progress=10,
                    message="–ü–∞—Ä—Å–∏–Ω–≥ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è...",
                )
                equipment_data = parse_equipment_workbook(dst)
                if equipment_data:
                    equipment_file = write_equipment_json(
                        batch_id, equipment_data, AGGREGATED_DIR
                    )
                    equipment_info = {
                        "output_file": equipment_file.name,
                        "sections": sum(
                            len(sheet["sections"])
                            for sheet in equipment_data.get("sheets", [])
                        ),
                        "total_items": equipment_data.get("summary", {}).get(
                            "total_items"
                        ),
                        "total_power_kw": equipment_data.get("summary", {}).get(
                            "total_power_kw"
                        ),
                    }
                    tracker.complete_stage(
                        ProcessingStage.SPECIALIZED_PARSING,
                        "–ü–∞—Ä—Å–∏–Ω–≥ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –∑–∞–≤–µ—Ä—à–µ–Ω",
                    )
            except InterruptedError:
                # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞
                logger.info("–ü–∞—Ä—Å–∏–Ω–≥ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –æ—Ç–º–µ–Ω–µ–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
                raise
            except (
                Exception
            ) as equipment_exc:  # pragma: no cover - equipment parsing failure
                logger.exception("Equipment parsing failed for %s", dst)
                equipment_info = {
                    "error": str(equipment_exc),
                }
        # –î–ª—è –≥–æ—Ç–æ–≤—ã—Ö Word-–æ—Ç—á–µ—Ç–æ–≤ –ø—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –ø–∞—Ä—Å–∏–Ω–≥
        # –ü–∞—Ä—Å–∏–Ω–≥ —Ñ–∞–π–ª–∞ —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º
        if (
            is_envelope_file(safe_filename, raw_json_for_check)
            and not is_ready_word_report
        ):
            try:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞
                if tracker.is_cancelled():
                    raise InterruptedError("–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")

                envelope_data = parse_building_envelope(dst)
                if envelope_data:
                    envelope_file = write_envelope_json(
                        batch_id, envelope_data, AGGREGATED_DIR
                    )
                    envelope_info = {
                        "output_file": envelope_file.name,
                        "sections": envelope_data.get("summary", {}).get(
                            "total_sections"
                        ),
                        "total_area_m2": envelope_data.get("summary", {}).get(
                            "total_area_m2"
                        ),
                        "total_heat_loss": envelope_data.get("summary", {}).get(
                            "total_heat_loss"
                        ),
                    }
            except InterruptedError:
                # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞
                logger.info(
                    "–ü–∞—Ä—Å–∏–Ω–≥ —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º –æ—Ç–º–µ–Ω–µ–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
                )
                raise
            except (
                Exception
            ) as envelope_exc:  # pragma: no cover - envelope parsing failure
                logger.exception(
                    "–û—à–∏–±–∫–∞ –ø–∞—Ä—Å–∏–Ω–≥–∞ —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º –¥–ª—è %s", dst
                )
                envelope_info = {
                    "error": str(envelope_exc),
                }
        # –î–ª—è –≥–æ—Ç–æ–≤—ã—Ö Word-–æ—Ç—á–µ—Ç–æ–≤ –ø—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–ø–µ—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –ø–∞—Ä—Å–∏–Ω–≥
        if (
            is_nodes_file(safe_filename, raw_json_for_check)
            and not is_ready_word_report
        ):
            try:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞
                if tracker.is_cancelled():
                    raise InterruptedError("–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")

                nodes_data = parse_nodes_workbook(dst)
                if nodes_data:
                    nodes_file = write_nodes_json(batch_id, nodes_data, AGGREGATED_DIR)
                    nodes_info = {
                        "output_file": nodes_file.name,
                        "total_nodes": nodes_data.get("summary", {}).get("total_nodes"),
                    }
                    
                    # –ò–º–ø–æ—Ä—Ç –¥–∞–Ω–Ω—ã—Ö –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ —É–∑–ª–∞–º —É—á—ë—Ç–∞ –≤ –ë–î
                    try:
                        # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º –¥–∞–Ω–Ω—ã–µ —É–∑–ª–æ–≤ –≤ —Ñ–æ—Ä–º–∞—Ç –¥–ª—è –∏–º–ø–æ—Ä—Ç–∞
                        # –ü—Ä–∏–º–µ—á–∞–Ω–∏–µ: –≤ —Ç–µ–∫—É—â–∏—Ö –¥–∞–Ω–Ω—ã—Ö —É–∑–ª–æ–≤ –Ω–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –ø–µ—Ä–∏–æ–¥–∞—Ö,
                        # –ø–æ—ç—Ç–æ–º—É —Å–æ–∑–¥–∞–µ–º –±–∞–∑–æ–≤—ã–µ –∑–∞–ø–∏—Å–∏. –î–∞–Ω–Ω—ã–µ –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ –ø–µ—Ä–∏–æ–¥–∞–º
                        # –Ω—É–∂–Ω–æ –±—É–¥–µ—Ç –ø–æ–ª—É—á–∞—Ç—å –∏–∑ –∞–∫—Ç–æ–≤ –±–∞–ª–∞–Ω—Å–æ–≤ –∏–ª–∏ –¥—Ä—É–≥–∏—Ö –∏—Å—Ç–æ—á–Ω–∏–∫–æ–≤
                        node_consumption_records = []
                        nodes_list = nodes_data.get("nodes", [])
                        
                        for node in nodes_list:
                            node_name = node.get("name")
                            if not node_name:
                                continue
                            
                            # –ï—Å–ª–∏ –µ—Å—Ç—å –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ —ç–ª–µ–∫—Ç—Ä–æ—ç–Ω–µ—Ä–≥–∏–∏, –ø—ã—Ç–∞–µ–º—Å—è —Å–≤—è–∑–∞—Ç—å
                            # –ü–æ–∫–∞ —Å–æ–∑–¥–∞–µ–º –∑–∞–ø–∏—Å—å —Å –±–∞–∑–æ–≤—ã–º–∏ –¥–∞–Ω–Ω—ã–º–∏ —É–∑–ª–∞
                            # TODO: –î–æ–±–∞–≤–∏—Ç—å –ª–æ–≥–∏–∫—É –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ –ø–µ—Ä–∏–æ–¥–∞–º
                            # –∏–∑ –∞–∫—Ç–æ–≤ –±–∞–ª–∞–Ω—Å–æ–≤ –∏–ª–∏ –¥—Ä—É–≥–∏—Ö –∏—Å—Ç–æ—á–Ω–∏–∫–æ–≤
                            
                            node_consumption_records.append({
                                "node_name": node_name,
                                "period": "unknown",  # –ë—É–¥–µ—Ç –æ–±–Ω–æ–≤–ª–µ–Ω–æ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –¥–∞–Ω–Ω—ã—Ö –ø–æ –ø–µ—Ä–∏–æ–¥–∞–º
                                "active_energy_kwh": node.get("active_energy_p"),
                                "reactive_energy_kvarh": node.get("reactive_energy_q"),
                                "cost_sum": None,  # –ù—É–∂–Ω–æ –ø–æ–ª—É—á–∏—Ç—å –∏–∑ –¥—Ä—É–≥–∏—Ö –∏—Å—Ç–æ—á–Ω–∏–∫–æ–≤
                                "data_json": {
                                    "tt": node.get("tt"),
                                    "coefficient": node.get("coefficient"),
                                    "seal_date": node.get("seal_date"),
                                    "supplier_seal_date": node.get("supplier_seal_date"),
                                    "note": node.get("note"),
                                }
                            })
                        
                        if node_consumption_records:
                            imported_nodes = database.import_node_consumption_to_db(
                                enterprise_id=enterprise["id"],
                                batch_id=batch_id,
                                node_consumption_data=node_consumption_records,
                            )
                            logger.info(
                                f"‚úÖ –ò–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {len(imported_nodes)} –∑–∞–ø–∏—Å–µ–π –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ —É–∑–ª–∞–º —É—á—ë—Ç–∞ "
                                f"–≤ –ë–î –¥–ª—è batch_id={batch_id}"
                            )
                            nodes_info["imported_to_db"] = len(imported_nodes)
                    except Exception as import_exc:
                        logger.warning(
                            f"‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å –¥–∞–Ω–Ω—ã–µ –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ —É–∑–ª–∞–º –≤ –ë–î: {import_exc}"
                        )
                        # –ù–µ –ø—Ä–µ—Ä—ã–≤–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É, –µ—Å–ª–∏ –∏–º–ø–æ—Ä—Ç –Ω–µ —É–¥–∞–ª—Å—è
            except InterruptedError:
                # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞
                logger.info("–ü–∞—Ä—Å–∏–Ω–≥ —É–∑–ª–æ–≤ —É—á–µ—Ç–∞ –æ—Ç–º–µ–Ω–µ–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
                raise
            except Exception as nodes_exc:  # pragma: no cover - nodes parsing failure
                logger.exception("Nodes parsing failed for %s", dst)
                nodes_info = {
                    "error": str(nodes_exc),
                }
        
        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –∞–∫—Ç–æ–≤ –±–∞–ª–∞–Ω—Å–æ–≤ –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ —É–∑–ª–∞–º —É—á—ë—Ç–∞
        # –°–æ–≥–ª–∞—Å–Ω–æ —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏—è–º —ç–∫—Å–ø–µ—Ä—Ç–æ–≤:
        # - Software Engineer: "–ò—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å OCR –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è –∏–∑ PDF –∞–∫—Ç–æ–≤"
        # - ML Engineer: "–ò—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å Gemini Vision (95% confidence)"
        # - QA Engineer: "–¢–µ—Å—Ç–∏—Ä–æ–≤–∞—Ç—å –Ω–∞ –Ω–µ—Å–∫–æ–ª—å–∫–∏—Ö —Ñ–∞–π–ª–∞—Ö —Å–Ω–∞—á–∞–ª–∞"
        logger.info(f"üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ —Ñ–∞–π–ª–∞ –Ω–∞ –∞–∫—Ç –±–∞–ª–∞–Ω—Å–∞: {safe_filename}")
        is_balance_sheet = is_balance_sheet_file(safe_filename, raw_json_for_check)
        if is_balance_sheet:
            logger.info(f"‚úÖ –§–∞–π–ª –æ–ø—Ä–µ–¥–µ–ª–µ–Ω –∫–∞–∫ –∞–∫—Ç –±–∞–ª–∞–Ω—Å–∞: {safe_filename}")
        else:
            logger.info(f"‚ùå –§–∞–π–ª –ù–ï –æ–ø—Ä–µ–¥–µ–ª–µ–Ω –∫–∞–∫ –∞–∫—Ç –±–∞–ª–∞–Ω—Å–∞: {safe_filename} (–ø—Ä–æ–≤–µ—Ä–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞)")
        
        if is_balance_sheet and not is_ready_word_report:
            try:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞
                if tracker.is_cancelled():
                    raise InterruptedError("–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
                
                logger.info(f"üìã –û–±–Ω–∞—Ä—É–∂–µ–Ω –∞–∫—Ç –±–∞–ª–∞–Ω—Å–∞: {safe_filename}")
                
                # –ò–∑–≤–ª–µ–∫–∞–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ —É–∑–ª–∞–º —É—á—ë—Ç–∞ –∏–∑ –∞–∫—Ç–∞ –±–∞–ª–∞–Ω—Å–∞
                # (–º–æ–∂–µ—Ç –±—ã—Ç—å –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏–µ, –ø—Ä–æ–∏–∑–≤–æ–¥—Å—Ç–≤–æ –∏–ª–∏ —Ä–µ–∞–ª–∏–∑–∞—Ü–∏—è —ç–ª–µ–∫—Ç—Ä–æ—ç–Ω–µ—Ä–≥–∏–∏)
                node_consumption_data = extract_node_consumption_from_balance_sheet(
                    file_path=dst,
                    batch_id=batch_id,
                    enterprise_id=enterprise["id"],
                    raw_json=raw_json_for_check
                )
                
                if node_consumption_data:
                    # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
                    data_types = set(record.get("data_type", "consumption") for record in node_consumption_data)
                    data_type_label = {
                        "consumption": "–ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è",
                        "production": "–ø—Ä–æ–∏–∑–≤–æ–¥—Å—Ç–≤–∞",
                        "realization": "—Ä–µ–∞–ª–∏–∑–∞—Ü–∏–∏ (–ø—Ä–æ–∏–∑–≤–µ–¥–µ–Ω–Ω–æ–π –∏ –ø—Ä–æ–¥–∞–Ω–Ω–æ–π)"
                    }
                    type_labels = [data_type_label.get(dt, dt) for dt in data_types]
                    
                    # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –≤ –ë–î
                    imported_nodes = database.import_node_consumption_to_db(
                        enterprise_id=enterprise["id"],
                        batch_id=batch_id,
                        node_consumption_data=node_consumption_data,
                    )
                    
                    logger.info(
                        f"‚úÖ –ò–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {len(imported_nodes)} –∑–∞–ø–∏—Å–µ–π –¥–∞–Ω–Ω—ã—Ö {', '.join(type_labels)} –ø–æ —É–∑–ª–∞–º —É—á—ë—Ç–∞ "
                        f"–∏–∑ –∞–∫—Ç–∞ –±–∞–ª–∞–Ω—Å–∞ {safe_filename} –≤ –ë–î –¥–ª—è batch_id={batch_id}"
                    )
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –≤ –æ—Ç–≤–µ—Ç
                    if "balance_sheet" not in response_data:
                        response_data["balance_sheet"] = {}
                    response_data["balance_sheet"]["nodes_imported"] = len(imported_nodes)
                    response_data["balance_sheet"]["file"] = safe_filename
                else:
                    logger.info(
                        f"‚ÑπÔ∏è –î–∞–Ω–Ω—ã–µ –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—è –ø–æ —É–∑–ª–∞–º –Ω–µ –Ω–∞–π–¥–µ–Ω—ã –≤ –∞–∫—Ç–µ –±–∞–ª–∞–Ω—Å–∞: {safe_filename}"
                    )
                    
            except InterruptedError:
                # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞
                logger.info("–û–±—Ä–∞–±–æ—Ç–∫–∞ –∞–∫—Ç–∞ –±–∞–ª–∞–Ω—Å–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
                raise
            except Exception as balance_exc:
                logger.warning(
                    f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∞–∫—Ç–∞ –±–∞–ª–∞–Ω—Å–∞ {safe_filename}: {balance_exc}"
                )
                # –ù–µ –ø—Ä–µ—Ä—ã–≤–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É, –µ—Å–ª–∏ –æ–±—Ä–∞–±–æ—Ç–∫–∞ –∞–∫—Ç–∞ –±–∞–ª–∞–Ω—Å–∞ –Ω–µ —É–¥–∞–ª–∞—Å—å
        
        if aggregation_info:
            response_data["aggregation"] = aggregation_info
        if equipment_info:
            response_data["equipment"] = equipment_info
        if envelope_info:
            response_data["envelope"] = envelope_info
        if nodes_info:
            response_data["nodes"] = nodes_info

        # –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –±—ã–ª–∞ –ª–∏ –æ—Ç–º–µ–Ω–µ–Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∞ –ø–µ—Ä–µ–¥ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ–º
        if tracker.is_cancelled():
            os.remove(dst)  # –£–¥–∞–ª—è–µ–º —Ñ–∞–π–ª –ø—Ä–∏ –æ—Ç–º–µ–Ω–µ
            raise HTTPException(
                status_code=499, detail="–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
            )

        tracker.update_stage(
            ProcessingStage.SAVING, progress=50, message="–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö..."
        )
        editable_text = build_editable_text(parsing_result)
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –¥–∞–Ω–Ω—ã–µ –µ—Å—Ç—å –≤ –∫—ç—à–µ –ø–µ—Ä–µ–¥ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ–º
        if batch_id not in parsing_results_cache:
            logger.error(f"‚ö†Ô∏è –î–∞–Ω–Ω—ã–µ –ø–∞—Ä—Å–∏–Ω–≥–∞ –æ—Ç—Å—É—Ç—Å—Ç–≤—É—é—Ç –≤ –∫—ç—à–µ –¥–ª—è batch_id={batch_id}")
            # –°–æ–∑–¥–∞–µ–º –º–∏–Ω–∏–º–∞–ª—å–Ω—É—é —Å—Ç—Ä—É–∫—Ç—É—Ä—É –∏–∑ –¥–æ—Å—Ç—É–ø–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
            parsing_results_cache[batch_id] = {
                "batch_id": batch_id,
                "filename": safe_filename,
                "file_path": dst,
                "file_type": file_type_label,
                "file_size": file_size,
                "parsing": parsing_result_data if parsing_result else None,
                "status": parsing_status,
                "error": "–î–∞–Ω–Ω—ã–µ –ø–∞—Ä—Å–∏–Ω–≥–∞ –Ω–µ –±—ã–ª–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤ –∫—ç—à",
            }
        
        try:
            database.save_parsed_content(
                batch_id,
                raw_json=parsing_results_cache[batch_id],
                editable_text=editable_text,
            )
            logger.info(f"‚úÖ –î–∞–Ω–Ω—ã–µ –ø–∞—Ä—Å–∏–Ω–≥–∞ —É—Å–ø–µ—à–Ω–æ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤ –ë–î –¥–ª—è batch_id={batch_id}")
            tracker.complete_stage(ProcessingStage.SAVING, "–î–∞–Ω–Ω—ã–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã")
        except Exception as save_exc:
            logger.error(
                f"‚ùå –û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –ø–∞—Ä—Å–∏–Ω–≥–∞ –≤ –ë–î –¥–ª—è batch_id={batch_id}: {save_exc}",
                exc_info=True
            )
            tracker.set_error(ProcessingStage.SAVING, f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è: {save_exc}")
            # –ù–µ –ø—Ä–µ—Ä—ã–≤–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É, –Ω–æ –ª–æ–≥–∏—Ä—É–µ–º –æ—à–∏–±–∫—É
        history = database.list_uploads_for_enterprise(enterprise["id"])
        response_data["history"] = history

        # –ó–∞–≤–µ—Ä—à–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É
        tracker.complete()

        # –û—á–∏—â–∞–µ–º —Ç—Ä–µ–∫–µ—Ä —á–µ—Ä–µ–∑ 5 –º–∏–Ω—É—Ç (–¥–∞–µ–º –≤—Ä–µ–º—è –¥–ª—è –æ–ø—Ä–æ—Å–∞ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞)
        # –í –ø—Ä–æ–¥–∞–∫—à–µ–Ω–µ –ª—É—á—à–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Ñ–æ–Ω–æ–≤—É—é –∑–∞–¥–∞—á—É
        import threading

        def cleanup_tracker():
            import time

            time.sleep(300)  # 5 –º–∏–Ω—É—Ç
            remove_progress_tracker(batch_id)

        threading.Thread(target=cleanup_tracker, daemon=True).start()

    except InterruptedError:
        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º
        logger.info("–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
        if os.path.exists(dst):
            try:
                os.remove(dst)  # –£–¥–∞–ª—è–µ–º —Ñ–∞–π–ª –ø—Ä–∏ –æ—Ç–º–µ–Ω–µ
            except Exception as e:
                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å —Ñ–∞–π–ª –ø—Ä–∏ –æ—Ç–º–µ–Ω–µ: {e}")
        raise HTTPException(
            status_code=499, detail="–û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º"
        )
    except Exception as exc:  # pragma: no cover - DB failure path
        logger.exception(
            f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –º–µ—Ç–∞–¥–∞–Ω–Ω—ã—Ö –∑–∞–≥—Ä—É–∑–∫–∏: {exc}",
            exc_info=True
        )
        # –õ–æ–≥–∏—Ä—É–µ–º –¥–µ—Ç–∞–ª–∏ –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
        logger.error(f"  batch_id: {batch_id}")
        logger.error(f"  filename: {safe_filename if 'safe_filename' in locals() else 'unknown'}")
        logger.error(f"  enterprise_id: {enterprise['id'] if 'enterprise' in locals() else 'unknown'}")
        logger.error(f"  file_type: {file_type_label if 'file_type_label' in locals() else 'unknown'}")
        
        if batch_id:
            tracker = get_progress_tracker(batch_id)
            if tracker:
                tracker.set_error(ProcessingStage.SAVING, f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è: {exc}")
        
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –±–æ–ª–µ–µ –¥–µ—Ç–∞–ª—å–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –æ–± –æ—à–∏–±–∫–µ
        error_detail = f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –∑–∞–≥—Ä—É–∑–∫–∏: {str(exc)}"
        if len(error_detail) > 200:  # –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º –¥–ª–∏–Ω—É –¥–ª—è –±–µ–∑–æ–ø–∞—Å–Ω–æ—Å—Ç–∏
            error_detail = error_detail[:200] + "..."
        
        raise HTTPException(
            status_code=500, detail=error_detail
        ) from exc

    return response_data


@app.post("/api/generate-passport/{batch_id}")
async def generate_energy_passport(
    batch_id: str,
    template_name: str = Query(
        default="",
        description="–ò–º—è —à–∞–±–ª–æ–Ω–∞ –∏–∑ templates_config (–Ω–∞–ø—Ä–∏–º–µ—Ä, 'new_energy_passport', 'metin', 'default')",
    ),
    skip_readiness_check: str = Query(
        default="false",
        description="–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å –ø—Ä–æ–≤–µ—Ä–∫—É –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ (—Ç–æ–ª—å–∫–æ –¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è)",
    ),
):
    """
    –ì–µ–Ω–µ—Ä–∞—Ü–∏—è —ç–Ω–µ—Ä–≥–µ—Ç–∏—á–µ—Å–∫–æ–≥–æ –ø–∞—Å–ø–æ—Ä—Ç–∞ –ø–æ –ü–ö–ú ‚Ññ690 –∏–∑ –¥–∞–Ω–Ω—ã—Ö –≤ –ë–î.

    Args:
        batch_id: ID –∑–∞–≥—Ä—É–∑–∫–∏
        template_name: –ò–º—è —à–∞–±–ª–æ–Ω–∞ –∏–∑ templates_config (–Ω–∞–ø—Ä–∏–º–µ—Ä, "new_energy_passport", "metin", "default")
                      –ï—Å–ª–∏ –Ω–µ —É–∫–∞–∑–∞–Ω–æ, –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –¥–µ—Ñ–æ–ª—Ç–Ω—ã–π —à–∞–±–ª–æ–Ω
        skip_readiness_check: –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å –ø—Ä–æ–≤–µ—Ä–∫—É –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ (—Ç–æ–ª—å–∫–æ –¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è)

    –ü–æ—Ç–æ–∫:
    1. –ü—Ä–æ–≤–µ—Ä–∫–∞ –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö (–µ—Å–ª–∏ –Ω–µ –ø—Ä–æ–ø—É—â–µ–Ω–∞)
    2. –ü–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –∏–∑ –ë–î (parsed_data.raw_json)
    3. –ê–≥—Ä–µ–≥–∏—Ä–æ–≤–∞—Ç—å –ø–æ–∫–≤–∞—Ä—Ç–∞–ª—å–Ω–æ
    4. –ó–∞–ø–æ–ª–Ω–∏—Ç—å Excel —à–∞–±–ª–æ–Ω
    5. –í–µ—Ä–Ω—É—Ç—å —Ñ–∞–π–ª –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è
    """
    if not HAS_GENERATOR:
        raise HTTPException(
            status_code=503,
            detail="–ì–µ–Ω–µ—Ä–∞—Ç–æ—Ä –ø–∞—Å–ø–æ—Ä—Ç–æ–≤ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ —É—Å—Ç–∞–Ω–æ–≤–∫—É –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–µ–π.",
        )

    # 1. –ü–æ–ª—É—á–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ –ë–î
    upload = database.get_upload_by_batch(batch_id)
    if not upload:
        raise HTTPException(status_code=404, detail=f"–ó–∞–≥—Ä—É–∑–∫–∞ {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")

    # 2. –ü—Ä–æ–≤–µ—Ä–∫–∞ –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö (–µ—Å–ª–∏ –Ω–µ –ø—Ä–æ–ø—É—â–µ–Ω–∞)
    # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º —Å—Ç—Ä–æ–∫—É –≤ bool
    skip_check = skip_readiness_check.lower() in ("true", "1", "yes", "on")
    template_name_final = template_name if template_name else None

    if not skip_check:
        # Feature-flag gate via EXCEL_SEMANTIC_AI_MODE
        excel_ai_mode = get_excel_semantic_mode()
        if excel_ai_mode in ("assist", "strict"):
            # Try to load canonical from stored raw_json or reconstruct
            canonical = None
            raw_json = upload.get("raw_json") or {}
            if isinstance(raw_json, dict):
                canonical_dict = raw_json.get("canonical_source")
                if isinstance(canonical_dict, dict):
                    try:
                        canonical = CanonicalSourceData.parse_obj(canonical_dict)
                    except Exception:
                        canonical = None
            if canonical is None:
                # Attempt reconstruction from file path (if known) or skip
                try:
                    filename = upload.get("filename")
                    if filename:
                        canonical = collect_canonical_from_workbook(filename)
                except Exception:
                    canonical = None
            readiness_result: GenerationReadinessResult = evaluate_generation_readiness(
                canonical
            )
            if readiness_result.overall_status == "blocked":
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Canonical data not ready for generation",
                        "overall_status": readiness_result.overall_status,
                        "missing_required": [
                            rf.__dict__ for rf in readiness_result.missing_required
                        ],
                        "missing_optional": [
                            rf.__dict__ for rf in readiness_result.missing_optional
                        ],
                        "notes": readiness_result.notes,
                        "mode": excel_ai_mode,
                    },
                )

        enterprise_id = upload.get("enterprise_id")
        if enterprise_id:
            readiness = validate_generation_readiness(enterprise_id)

            if not readiness["ready"]:
                logger.warning(
                    f"–ü–æ–ø—ã—Ç–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –ø–∞—Å–ø–æ—Ä—Ç–∞ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id} "
                    f"–ø—Ä–∏ –Ω–µ–≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö: {readiness.get('missing_resources', [])}"
                )
                detail = {
                    "message": "–î–∞–Ω–Ω—ã–µ –Ω–µ –≥–æ—Ç–æ–≤—ã –¥–ª—è –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ —ç–Ω–µ—Ä–≥–µ—Ç–∏—á–µ—Å–∫–æ–≥–æ –ø–∞—Å–ø–æ—Ä—Ç–∞",
                    "missing_resources": readiness.get("missing_resources", []),
                    "missing_files": readiness.get("missing_files", []),
                    "missing_sheet_data": readiness.get("missing_sheet_data", []),
                    "sheet_validation": readiness.get("sheet_validation", {}),
                    "completeness_score": readiness.get("completeness_score", 0.0),
                    "warnings": readiness.get("warnings", []),
                    "progress_percentage": readiness.get("progress_percentage", 0),
                    "available_resources": readiness.get("available_resources", []),
                }
                raise HTTPException(status_code=400, detail=detail)

            logger.info(
                f"–ü—Ä–æ–≤–µ—Ä–∫–∞ –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –ø—Ä–æ–π–¥–µ–Ω–∞ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id}: "
                f"completeness={readiness['completeness_score']:.2f}"
            )

    # 2. –ê–≥—Ä–µ–≥–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ –≤—Å–µ—Ö –∑–∞–≥—Ä—É–∑–æ–∫ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è (–Ω–µ —Ç–æ–ª—å–∫–æ –∏–∑ –æ–¥–Ω–æ–≥–æ batch_id)
    enterprise_id = upload.get("enterprise_id")
    logger.info(
        f"–ê–≥—Ä–µ–≥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id} (batch_id: {batch_id})"
    )

    # –ê–≥—Ä–µ–≥–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ –≤—Å–µ—Ö –∑–∞–≥—Ä—É–∑–æ–∫ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è
    aggregated = None
    from utils.readiness_validator import _get_aggregated_data_for_enterprise

    aggregated = _get_aggregated_data_for_enterprise(enterprise_id)

    # –ï—Å–ª–∏ –Ω–µ –ø–æ–ª—É—á–∏–ª–æ—Å—å –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞—Ç—å –∏–∑ –≤—Å–µ—Ö –∑–∞–≥—Ä—É–∑–æ–∫, –ø—Ä–æ–±—É–µ–º –∏–∑ —Ç–µ–∫—É—â–µ–≥–æ batch_id
    if not aggregated:
        raw_json = upload.get("raw_json")
        if not raw_json:
            raise HTTPException(status_code=400, detail="–î–∞–Ω–Ω—ã–µ –Ω–µ —Ä–∞—Å–ø–∞—Ä—Å–µ–Ω—ã")

        logger.info(f"–ê–≥—Ä–µ–≥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö –∏–∑ batch_id: {batch_id}")
        logger.info(f"–°—Ç—Ä—É–∫—Ç—É—Ä–∞ raw_json: {list(raw_json.keys())}")
        logger.info(f"file_type: {raw_json.get('file_type')}")
        logger.info(f"parsing keys: {list(raw_json.get('parsing', {}).keys())}")

        # –õ–æ–≥–∏—Ä—É–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É parsing.data –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
        parsing_data = raw_json.get("parsing", {}).get("data", {})
        if parsing_data:
            logger.info(f"parsing.data keys: {list(parsing_data.keys())}")
            if "sheets" in parsing_data:
                sheets = parsing_data.get("sheets", [])
                logger.info(f"–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏—Å—Ç–æ–≤: {len(sheets)}")
                for sheet in sheets[:5]:  # –ü–µ—Ä–≤—ã–µ 5 –ª–∏—Å—Ç–æ–≤
                    sheet_name = sheet.get("name", "–ë–µ–∑ –∏–º–µ–Ω–∏")
                    rows_count = len(sheet.get("rows", []))
                    logger.info(f"  –õ–∏—Å—Ç '{sheet_name}': {rows_count} —Å—Ç—Ä–æ–∫")

        aggregated = aggregate_from_db_json(raw_json)

    # –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ–µ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞ –∞–≥—Ä–µ–≥–∞—Ü–∏–∏
    if aggregated:
        logger.info(
            f"–ê–≥—Ä–µ–≥–∞—Ü–∏—è —É—Å–ø–µ—à–Ω–∞. –°—Ç—Ä—É–∫—Ç—É—Ä–∞ aggregated: {list(aggregated.keys())}"
        )
        if "resources" in aggregated:
            resources = aggregated["resources"]
            for resource_type, resource_data in resources.items():
                if (
                    resource_data
                    and isinstance(resource_data, dict)
                    and len(resource_data) > 0
                ):
                    quarters = list(resource_data.keys())
                    logger.info(
                        f"  –†–µ—Å—É—Ä—Å {resource_type}: {len(quarters)} –∫–≤–∞—Ä—Ç–∞–ª–æ–≤ - {quarters[:3]}..."
                    )
                    # –õ–æ–≥–∏—Ä—É–µ–º –¥–µ—Ç–∞–ª–∏ –ø–æ –ø–µ—Ä–≤–æ–º—É –∫–≤–∞—Ä—Ç–∞–ª—É –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
                    if quarters:
                        first_quarter = quarters[0]
                        quarter_data = resource_data[first_quarter]
                        totals = quarter_data.get("quarter_totals", {})
                        logger.info(
                            f"    –ö–≤–∞—Ä—Ç–∞–ª {first_quarter}: totals={list(totals.keys())}, values={list(totals.values())[:3]}"
                        )
                else:
                    logger.info(f"  –†–µ—Å—É—Ä—Å {resource_type}: –ø—É—Å—Ç–æ–π")

        # –í–∞–ª–∏–¥–∞—Ü–∏—è –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –ø–µ—Ä–µ–¥ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏–µ–º —à–∞–±–ª–æ–Ω–∞
        logger.info("–í–∞–ª–∏–¥–∞—Ü–∏—è –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –ø–µ—Ä–µ–¥ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏–µ–º —à–∞–±–ª–æ–Ω–∞...")
        try:
            is_valid, errors, warnings = validate_data_for_template(
                aggregated, raise_on_error=False
            )
            if not is_valid:
                logger.warning(f"–û–±–Ω–∞—Ä—É–∂–µ–Ω—ã –æ—à–∏–±–∫–∏ –≤–∞–ª–∏–¥–∞—Ü–∏–∏ –¥–∞–Ω–Ω—ã—Ö: {errors}")
                # –õ–æ–≥–∏—Ä—É–µ–º –ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–µ–Ω–∏—è –æ—Ç–¥–µ–ª—å–Ω–æ
                if warnings:
                    logger.info(f"–ü—Ä–µ–¥—É–ø—Ä–µ–∂–¥–µ–Ω–∏—è –≤–∞–ª–∏–¥–∞—Ü–∏–∏: {warnings}")
            else:
                logger.info("–í–∞–ª–∏–¥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö –ø—Ä–æ–π–¥–µ–Ω–∞ —É—Å–ø–µ—à–Ω–æ")
                if warnings:
                    logger.info(f"–ü—Ä–µ–¥—É–ø—Ä–µ–∂–¥–µ–Ω–∏—è –≤–∞–ª–∏–¥–∞—Ü–∏–∏: {warnings}")
        except Exception as validation_exc:
            logger.error(
                f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≤–∞–ª–∏–¥–∞—Ü–∏–∏ –¥–∞–Ω–Ω—ã—Ö: {validation_exc}", exc_info=True
            )
            # –ù–µ –ø—Ä–µ—Ä—ã–≤–∞–µ–º –≥–µ–Ω–µ—Ä–∞—Ü–∏—é –∏–∑-–∑–∞ –æ—à–∏–±–æ–∫ –≤–∞–ª–∏–¥–∞—Ü–∏–∏, –Ω–æ –ª–æ–≥–∏—Ä—É–µ–º
    else:
        logger.error(f"–ê–≥—Ä–µ–≥–∞—Ü–∏—è –≤–µ—Ä–Ω—É–ª–∞ None –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id}")
        if "raw_json" in locals():
            logger.error(
                f"raw_json.parsing.parsed: {raw_json.get('parsing', {}).get('parsed')}"
            )
            logger.error(
                f"raw_json.parsing.sheets count: {len(raw_json.get('parsing', {}).get('sheets', []))}"
            )
        raise HTTPException(
            status_code=400,
            detail="–ù–µ —É–¥–∞–ª–æ—Å—å –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞—Ç—å –¥–∞–Ω–Ω—ã–µ. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ —Å—Ç—Ä—É–∫—Ç—É—Ä—É —Ñ–∞–π–ª–∞.",
        )

    # 3. –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –¥–∞–Ω–Ω—ã—Ö –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è
    enterprise_data = {
        "id": upload.get("enterprise_id"),
        "name": upload.get("enterprise_name", "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ–µ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ"),
        "inn": None,  # TODO: –¥–æ–±–∞–≤–∏—Ç—å –≤ –ë–î
        "address": None,
        "director_name": None,
        "industry": None,
        "reporting_year": 2024,
    }

    # 4. –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –ø–∞—Å–ø–æ—Ä—Ç
    output_dir = Path("/tmp/passports")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"{batch_id}_energy_passport.xlsx"

    # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ –ø—É—Ç–∏ –∫ —à–∞–±–ª–æ–Ω—É
    template_path = None
    if template_name_final:
        # –ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ templates_config –¥–ª—è –≤—ã–±–æ—Ä–∞ —à–∞–±–ª–æ–Ω–∞ –ø–æ –∏–º–µ–Ω–∏
        try:
            templates_config_path = PROJECT_ROOT / "templates" / "pcm690"
            import sys

            if str(templates_config_path) not in sys.path:
                sys.path.insert(0, str(templates_config_path))
            from templates_config import get_template_path

            template_path = get_template_path(template_name_final)
            logger.info(
                "‚úÖ –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è —à–∞–±–ª–æ–Ω –ø–æ –∏–º–µ–Ω–∏ '%s': %s",
                template_name_final,
                template_path,
            )
            if not template_path.exists():
                logger.error("‚ùå –§–∞–π–ª —à–∞–±–ª–æ–Ω–∞ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç: %s", template_path)
                template_path = None
        except (ImportError, ValueError, FileNotFoundError) as e:
            logger.warning(
                "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å —à–∞–±–ª–æ–Ω –ø–æ –∏–º–µ–Ω–∏ '%s': %s. –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –¥–µ—Ñ–æ–ª—Ç–Ω—ã–π.",
                template_name_final,
                e,
            )
            logger.exception("–î–µ—Ç–∞–ª–∏ –æ—à–∏–±–∫–∏ –∑–∞–≥—Ä—É–∑–∫–∏ —à–∞–±–ª–æ–Ω–∞:")
            template_path = None

    # –ï—Å–ª–∏ —à–∞–±–ª–æ–Ω –Ω–µ –≤—ã–±—Ä–∞–Ω –ø–æ –∏–º–µ–Ω–∏, –∏—Å–ø–æ–ª—å–∑—É–µ–º –¥–µ—Ñ–æ–ª—Ç–Ω—ã–µ –∫–∞–Ω–¥–∏–¥–∞—Ç—ã
    if not template_path:
        template_candidates = [
            PROJECT_ROOT
            / "data"
            / "source_files"
            / "audit_sinergys"
            / "EnergyPassport_PKM690_filled.xlsx",
            PROJECT_ROOT / "templates" / "pcm690" / "energy_passport_template.xlsx",
        ]
        template_path = next(
            (path for path in template_candidates if path.exists()), None
        )
        if not template_path:
            raise FileNotFoundError("–®–∞–±–ª–æ–Ω —ç–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ –æ–∂–∏–¥–∞–µ–º—ã—Ö –ø—É—Ç—è—Ö")
        logger.info("–ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –¥–µ—Ñ–æ–ª—Ç–Ω—ã–π —à–∞–±–ª–æ–Ω —ç–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç–∞: %s", template_path)

    if HAS_FILLER:
        try:
            # –ö–æ–ø–∏—Ä—É–µ–º —à–∞–±–ª–æ–Ω –≤–æ –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª –¥–ª—è —Ä–∞–±–æ—Ç—ã (–∑–∞—â–∏—Ç–∞ –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω–æ–≥–æ —à–∞–±–ª–æ–Ω–∞)
            temp_template_path = output_dir / f"{batch_id}_passport_template_copy.xlsx"
            shutil.copyfile(template_path, temp_template_path)
            logger.info("–®–∞–±–ª–æ–Ω —Å–∫–æ–ø–∏—Ä–æ–≤–∞–Ω –≤–æ –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª: %s", temp_template_path)

            workbook = load_workbook(temp_template_path, data_only=False)

            resources_data = aggregated.get("resources") or aggregated
            if resources_data is None:
                resources_data = {}
            # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º –≤—Å–µ —Ç–∏–ø—ã —Ä–µ—Å—É—Ä—Å–æ–≤
            for key in (
                "electricity",
                "gas",
                "water",
                "fuel",
                "coal",
                "heat",
                "production",
            ):
                resources_data.setdefault(key, {})

            # –õ–æ–≥–∏—Ä—É–µ–º –¥–æ—Å—Ç—É–ø–Ω—ã–µ —Ä–µ—Å—É—Ä—Å—ã –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
            available_resources = [
                k
                for k, v in resources_data.items()
                if v and isinstance(v, dict) and len(v) > 0
            ]
            logger.info(f"–î–æ—Å—Ç—É–ø–Ω—ã–µ —Ä–µ—Å—É—Ä—Å—ã –¥–ª—è –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è: {available_resources}")
            logger.info(f"–í—Å–µ–≥–æ —Ä–µ—Å—É—Ä—Å–æ–≤ –≤ –¥–∞–Ω–Ω—ã—Ö: {list(resources_data.keys())}")

            # –õ–æ–≥–∏—Ä—É–µ–º –¥–µ—Ç–∞–ª–∏ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Ä–µ—Å—É—Ä—Å–∞
            for resource_type, resource_data in resources_data.items():
                if (
                    resource_data
                    and isinstance(resource_data, dict)
                    and len(resource_data) > 0
                ):
                    logger.info(
                        f"  –†–µ—Å—É—Ä—Å {resource_type}: {len(resource_data)} –∫–≤–∞—Ä—Ç–∞–ª–æ–≤"
                    )
                    for quarter, quarter_data in list(resource_data.items())[
                        :2
                    ]:  # –ü–µ—Ä–≤—ã–µ 2 –∫–≤–∞—Ä—Ç–∞–ª–∞
                        totals = (
                            quarter_data.get("quarter_totals", {})
                            if isinstance(quarter_data, dict)
                            else {}
                        )
                        logger.info(
                            f"    {quarter}: totals keys={list(totals.keys())}, sample values={dict(list(totals.items())[:3])}"
                        )

            # (–û–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ) Canonical payload for nodes/equipment when mode != off
            canonical_payload = None
            excel_ai_mode_runtime = get_excel_semantic_mode()
            if excel_ai_mode_runtime in ("assist", "strict"):
                try:
                    c = None
                    raw_json = upload.get("raw_json") or {}
                    if isinstance(raw_json, dict):
                        cdict = raw_json.get("canonical_source")
                        if isinstance(cdict, dict):
                            try:
                                c = CanonicalSourceData.parse_obj(cdict)
                            except Exception:
                                c = None
                    if c is None and upload.get("filename"):
                        c = collect_canonical_from_workbook(upload.get("filename"))
                    if c:
                        canonical_payload = canonical_to_passport_payload(c)
                        logger.info(
                            "Canonical payload prepared for nodes/equipment (mode=%s)",
                            excel_ai_mode_runtime,
                        )
                except Exception as e:
                    logger.warning("Failed to prepare canonical payload: %s", e)

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–°—Ç—Ä—É–∫—Ç—É—Ä–∞ –ø—Ä 2"
            struktura_sheet_names = [
                "–°—Ç—Ä—É–∫—Ç—É—Ä–∞ –ø—Ä 2",
                "–°—Ç—Ä—É–∫—Ç—É—Ä–∞ –ø—Ä 2 ",
                "Struktura pr2",
                "02_–°—Ç—Ä—É–∫—Ç—É—Ä–∞",
            ]
            struktura_sheet = None
            # –°–Ω–∞—á–∞–ª–∞ –∏—â–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ (—Å —É—á–µ—Ç–æ–º –ø—Ä–æ–±–µ–ª–æ–≤)
            for sheet_name in struktura_sheet_names:
                if sheet_name in workbook.sheetnames:
                    struktura_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏, –∏—â–µ–º —Å —É—á–µ—Ç–æ–º –ø—Ä–æ–±–µ–ª–æ–≤ –≤ –∫–æ–Ω—Ü–µ
            if not struktura_sheet:
                for ws_name in workbook.sheetnames:
                    ws_name_stripped = ws_name.strip()
                    for target_name in struktura_sheet_names:
                        if ws_name_stripped == target_name.strip():
                            struktura_sheet = workbook[ws_name]
                            break
                    if struktura_sheet:
                        break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not struktura_sheet:
                for ws_name in workbook.sheetnames:
                    if "—Å—Ç—Ä—É–∫—Ç—É—Ä–∞" in ws_name.lower() or "struktura" in ws_name.lower():
                        struktura_sheet = workbook[ws_name]
                        break

            # –°–æ–∑–¥–∞–µ–º –ª–∏—Å—Ç, –µ—Å–ª–∏ –µ–≥–æ –Ω–µ—Ç
            if not struktura_sheet:
                logger.info("–õ–∏—Å—Ç '–°—Ç—Ä—É–∫—Ç—É—Ä–∞ –ø—Ä 2' –Ω–µ –Ω–∞–π–¥–µ–Ω, —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π")
                struktura_sheet = workbook.create_sheet(title="02_–°—Ç—Ä—É–∫—Ç—É—Ä–∞")

            if struktura_sheet:
                logger.info(
                    f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{struktura_sheet.title}' —Å –¥–∞–Ω–Ω—ã–º–∏: {len(available_resources)} —Ä–µ—Å—É—Ä—Å–æ–≤"
                )
                fill_struktura_pr2(
                    struktura_sheet,
                    resources_data,
                    loss_active_month=0.0,
                    loss_reactive_month=0.0,
                )
                logger.info(f"–õ–∏—Å—Ç '{struktura_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–£–∑–ª—ã —É—á–µ—Ç–∞"
            nodes_json_path = AGGREGATED_DIR / f"{batch_id}_nodes.json"
            # –ò—â–µ–º nodes JSON –≤–æ –≤—Å–µ—Ö –∑–∞–≥—Ä—É–∑–∫–∞—Ö –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è
            if not nodes_json_path.exists() and enterprise_id:
                uploads = database.list_uploads_for_enterprise(enterprise_id)
                for upload_item in uploads:
                    upload_batch_id = upload_item.get("batch_id")
                    if upload_batch_id:
                        candidate_path = (
                            AGGREGATED_DIR / f"{upload_batch_id}_nodes.json"
                        )
                        if candidate_path.exists():
                            nodes_json_path = candidate_path
                            logger.info(
                                f"–ù–∞–π–¥–µ–Ω nodes JSON –∏–∑ –¥—Ä—É–≥–æ–π –∑–∞–≥—Ä—É–∑–∫–∏: {nodes_json_path}"
                            )
                            break

            if canonical_payload and canonical_payload.get("nodes"):
                nodes_data = canonical_payload["nodes"]
                logger.info(
                    "–ò—Å–ø–æ–ª—å–∑—É—é—Ç—Å—è canonical —É–∑–ª—ã —É—á–µ—Ç–∞ (%d —à—Ç.)", len(nodes_data)
                )
            else:
                if nodes_json_path.exists():
                    logger.info("–ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è nodes JSON: %s", nodes_json_path)
                    nodes_data = load_nodes_from_json(nodes_json_path)
                else:
                    nodes_data = load_default_nodes()
                    logger.info("–ò—Å–ø–æ–ª—å–∑—É—é—Ç—Å—è —É–∑–ª—ã —É—á–µ—Ç–∞ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é")

            nodes_sheet_names = [
                "01_–£–∑–ª—ã —É—á–µ—Ç–∞",
                "–£–∑–µ–ª —É—á–µ—Ç–∞",
                "–£–∑–µ–ª —É—á–µ—Ç–∞ ",
                "–£–∑–ª—ã —É—á–µ—Ç–∞",
                "Nodes",
            ]
            nodes_sheet = None
            # –°–Ω–∞—á–∞–ª–∞ –∏—â–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ (—Å —É—á–µ—Ç–æ–º –ø—Ä–æ–±–µ–ª–æ–≤)
            for sheet_name in nodes_sheet_names:
                if sheet_name in workbook.sheetnames:
                    nodes_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏, –∏—â–µ–º —Å —É—á–µ—Ç–æ–º –ø—Ä–æ–±–µ–ª–æ–≤ –≤ –∫–æ–Ω—Ü–µ
            if not nodes_sheet:
                for ws_name in workbook.sheetnames:
                    ws_name_stripped = ws_name.strip()
                    for target_name in nodes_sheet_names:
                        if ws_name_stripped == target_name.strip():
                            nodes_sheet = workbook[ws_name]
                            break
                    if nodes_sheet:
                        break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not nodes_sheet:
                for ws_name in workbook.sheetnames:
                    if "—É–∑–µ–ª" in ws_name.lower() or "nodes" in ws_name.lower():
                        nodes_sheet = workbook[ws_name]
                        break

            # –°–æ–∑–¥–∞–µ–º –ª–∏—Å—Ç, –µ—Å–ª–∏ –µ–≥–æ –Ω–µ—Ç
            if not nodes_sheet:
                logger.info("–õ–∏—Å—Ç '–£–∑–ª—ã —É—á–µ—Ç–∞' –Ω–µ –Ω–∞–π–¥–µ–Ω, —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π")
                nodes_sheet = workbook.create_sheet(title="01_–£–∑–ª—ã —É—á–µ—Ç–∞")

            if nodes_sheet:
                logger.info(
                    f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{nodes_sheet.title}' —Å {len(nodes_data)} —É–∑–ª–∞–º–∏"
                )
                fill_nodes_sheet(nodes_sheet, nodes_data)
                logger.info(f"–õ–∏—Å—Ç '{nodes_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ"
            equipment_json_path = AGGREGATED_DIR / f"{batch_id}_equipment.json"
            if not equipment_json_path.exists() and enterprise_id:
                uploads = database.list_uploads_for_enterprise(enterprise_id)
                for upload_item in uploads:
                    upload_batch_id = upload_item.get("batch_id")
                    if upload_batch_id:
                        candidate_path = (
                            AGGREGATED_DIR / f"{upload_batch_id}_equipment.json"
                        )
                        if candidate_path.exists():
                            equipment_json_path = candidate_path
                            break

            if canonical_payload and canonical_payload.get("equipment"):
                equipment_data = canonical_payload["equipment"]
                logger.info(
                    "–ò—Å–ø–æ–ª—å–∑—É—é—Ç—Å—è canonical –¥–∞–Ω–Ω—ã–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è (sections=%d)",
                    len(equipment_data.get("sheets", [])),
                )
                try:
                    # –ò—â–µ–º –ª–∏—Å—Ç –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –∏ –∑–∞–ø–æ–ª–Ω—è–µ–º
                    equipment_sheet_names = [
                        "Equipment",
                        "–ê–ù–ê–õ–ò–ó –û–ë–û–†–£–î–û–í–ê–ù–ò–Ø",
                        "–ê–Ω–∞–ª–∏–∑ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è",
                        "–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
                        "03_–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
                        "–æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
                    ]
                    equipment_sheet_found = False
                    for sheet_name in equipment_sheet_names:
                        if sheet_name in workbook.sheetnames:
                            fill_equipment_sheet(
                                workbook, equipment_data, sheet_name=sheet_name
                            )
                            equipment_sheet_found = True
                            break
                    if not equipment_sheet_found:
                        for sheet_name in workbook.sheetnames:
                            sheet_lower = sheet_name.lower()
                            if any(
                                keyword in sheet_lower
                                for keyword in ["equipment", "–æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ", "–∞–Ω–∞–ª–∏–∑"]
                            ):
                                fill_equipment_sheet(
                                    workbook, equipment_data, sheet_name=sheet_name
                                )
                                equipment_sheet_found = True
                                break
                    if not equipment_sheet_found:
                        logger.warning(
                            "–õ–∏—Å—Ç –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω –¥–ª—è canonical payload"
                        )
                except Exception as equipment_exc:
                    logger.exception(
                        f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏–∏ –ª–∏—Å—Ç–∞ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è (canonical): {equipment_exc}"
                    )
            else:
                if equipment_json_path.exists():
                    logger.info("–ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è equipment JSON: %s", equipment_json_path)
                    equipment_data = None
                    try:
                        import json as json_module

                        equipment_data = json_module.loads(
                            equipment_json_path.read_text(encoding="utf-8")
                        )
                        logger.info(
                            f"–ó–∞–≥—Ä—É–∂–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è: sheets={len(equipment_data.get('sheets', []))}, summary={equipment_data.get('summary', {})}"
                        )
                    except Exception as equipment_exc:
                        logger.exception(
                            f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —á—Ç–µ–Ω–∏–∏ JSON –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è: {equipment_exc}"
                        )
                        equipment_data = None
                    if equipment_data:
                        # –ò—â–µ–º –ª–∏—Å—Ç "Equipment" –∏–ª–∏ –¥—Ä—É–≥–∏–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã –Ω–∞–∑–≤–∞–Ω–∏–π (—Å–Ω–∞—á–∞–ª–∞ –∏—â–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ)
                        equipment_sheet_names = [
                            "Equipment",
                            "–ê–ù–ê–õ–ò–ó –û–ë–û–†–£–î–û–í–ê–ù–ò–Ø",
                            "–ê–Ω–∞–ª–∏–∑ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è",
                            "–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
                            "03_–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
                            "–æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",  # –° –º–∞–ª–µ–Ω—å–∫–æ–π –±—É–∫–≤—ã
                            "Sheet1",  # –í–æ–∑–º–æ–∂–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –≤ new_energy_passport
                        ]
                        equipment_sheet_found = False

                        # –°–Ω–∞—á–∞–ª–∞ –ø—ã—Ç–∞–µ–º—Å—è –Ω–∞–π—Ç–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ
                        for sheet_name in equipment_sheet_names:
                            if sheet_name in workbook.sheetnames:
                                logger.info(f"–ù–∞–π–¥–µ–Ω –ª–∏—Å—Ç –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è: '{sheet_name}'")
                                fill_equipment_sheet(
                                    workbook, equipment_data, sheet_name=sheet_name
                                )
                                logger.info(f"–õ–∏—Å—Ç '{sheet_name}' –∑–∞–ø–æ–ª–Ω–µ–Ω")
                                equipment_sheet_found = True
                                break

                        # –ï—Å–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, –∏—â–µ–º —á–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ
                        if not equipment_sheet_found:
                            logger.info(
                                "–¢–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, –∏—â–µ–º —á–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ..."
                            )
                            for sheet_name in workbook.sheetnames:
                                sheet_lower = sheet_name.lower()
                                if any(
                                    keyword in sheet_lower
                                    for keyword in [
                                        "equipment",
                                        "–æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ",
                                        "–∞–Ω–∞–ª–∏–∑",
                                    ]
                                ):
                                    logger.info(
                                        f"–ù–∞–π–¥–µ–Ω –ª–∏—Å—Ç –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É —Å–æ–≤–ø–∞–¥–µ–Ω–∏—é: '{sheet_name}'"
                                    )
                                    fill_equipment_sheet(
                                        workbook, equipment_data, sheet_name=sheet_name
                                    )
                                    logger.info(f"–õ–∏—Å—Ç '{sheet_name}' –∑–∞–ø–æ–ª–Ω–µ–Ω")
                                    equipment_sheet_found = True
                                    break

                        if not equipment_sheet_found:
                            # –°–æ–∑–¥–∞–µ–º –ª–∏—Å—Ç Equipment, –µ—Å–ª–∏ –µ–≥–æ –Ω–µ—Ç
                            logger.info("–õ–∏—Å—Ç –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω, —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π")
                            equipment_sheet = workbook.create_sheet(
                                title="03_–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ"
                            )
                            fill_equipment_sheet(
                                workbook, equipment_data, sheet_name="03_–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ"
                            )
                            equipment_sheet_found = True
                            logger.info("–õ–∏—Å—Ç '03_–û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ' —Å–æ–∑–¥–∞–Ω –∏ –∑–∞–ø–æ–ª–Ω–µ–Ω")
                else:
                    logger.warning(f"Equipment JSON –Ω–µ –Ω–∞–π–¥–µ–Ω: {equipment_json_path}")

            # –ï—Å–ª–∏ –µ—Å—Ç—å –≥–æ–¥–æ–≤—ã–µ –∏—Ç–æ–≥–∏ –∏–∑ Canonical, –¥–æ–±–∞–≤–ª—è–µ–º —Å—Ç—Ä–æ–∫—É 'ANNUAL' –¥–ª—è —Ä–µ—Å—É—Ä—Å–æ–≤ –≤ –±–∞–ª–∞–Ω—Å
            try:
                excel_ai_mode_runtime = get_excel_semantic_mode()
                if (
                    excel_ai_mode_runtime in ("assist", "strict")
                    and canonical_payload
                    and canonical_payload.get("balance", {}).get("annual_totals")
                ):
                    annual_totals = canonical_payload["balance"]["annual_totals"]
                    # –ö–∞—Ä—Ç–∞ –∫–ª—é—á–µ–π –¥–ª—è –∏—Ç–æ–≥–æ–≤ –ø–æ —Ä–µ—Å—É—Ä—Å–∞–º (Balance –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è; –∫–≤–∞—Ä—Ç–∞–ª—ã –Ω–µ –ø—Ä–∏–¥—É–º—ã–≤–∞–µ–º)
                    total_key_map = {
                        "electricity": "active_kwh",
                        "gas": "volume_m3",
                        "water": "volume_m3",
                        "heat": "energy_gcal",
                        "fuel": "volume_ton",
                        "coal": "volume_ton",
                    }
                    for resource_name, annual_value in annual_totals.items():
                        if annual_value is None:
                            continue
                        if resource_name not in resources_data:
                            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–æ–∑–¥–∞–Ω–∏–µ —Å–ª–æ–∂–Ω–æ–π —Å—Ç—Ä—É–∫—Ç—É—Ä—ã —Å –Ω—É–ª—è –¥–ª—è –æ—Ç—Å—É—Ç—Å—Ç–≤—É—é—â–∏—Ö —Ä–µ—Å—É—Ä—Å–æ–≤
                            continue
                        key = total_key_map.get(resource_name, "annual_total")
                        resources_data.setdefault(resource_name, {})
                        # –°–ø–µ—Ü–∏–∞–ª—å–Ω—ã–π –∫–ª—é—á –¥–ª—è –≥–æ–¥–æ–≤–æ–≥–æ –∏—Ç–æ–≥–∞ ‚Äî —Ç–æ–ª—å–∫–æ –≤ —Ü–µ–ª—è—Ö —Å—Ä–∞–≤–Ω–µ–Ω–∏—è legacy vs canonical
                        resources_data[resource_name]["ANNUAL"] = {
                            "year": None,
                            "quarter": None,
                            "quarter_totals": {key: float(annual_value)},
                            "by_usage": {},
                        }
                        # –ï—Å–ª–∏ –µ—Å—Ç—å canonical by_usage –¥–ª—è —Ä–µ—Å—É—Ä—Å–∞ ‚Äî –¥–æ–±–∞–≤–ª—è–µ–º –≤ ANNUAL
                        try:
                            byu = (
                                canonical_payload.get("balance", {}).get("by_usage", {})
                                or {}
                            ).get(resource_name)
                            if isinstance(byu, dict) and byu:
                                resources_data[resource_name]["ANNUAL"]["by_usage"] = (
                                    byu
                                )
                                logger.info(
                                    "–î–æ–±–∞–≤–ª–µ–Ω canonical %s.by_usage –≤ ANNUAL: keys=%s",
                                    resource_name,
                                    list(byu.keys()),
                                )

                                # –ö–†–ò–¢–ò–ß–ù–û: –†–∞—Å–ø—Ä–µ–¥–µ–ª—è–µ–º canonical by_usage –ø–æ –∫–≤–∞—Ä—Ç–∞–ª–∞–º –ø—Ä–æ–ø–æ—Ä—Ü–∏–æ–Ω–∞–ª—å–Ω–æ –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—é
                                if (
                                    resource_name == "electricity"
                                    and resources_data.get("electricity")
                                ):
                                    annual_by_usage = byu
                                    annual_total = float(annual_value)

                                    # –í—ã—á–∏—Å–ª—è–µ–º –æ–±—â–µ–µ –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏–µ –ø–æ –≤—Å–µ–º –∫–≤–∞—Ä—Ç–∞–ª–∞–º
                                    total_quarterly_consumption = 0.0
                                    for quarter_key, quarter_data in resources_data[
                                        "electricity"
                                    ].items():
                                        if quarter_key == "ANNUAL":
                                            continue
                                        if isinstance(quarter_data, dict):
                                            quarter_total = quarter_data.get(
                                                "quarter_totals", {}
                                            ).get("active_kwh", 0)
                                            if quarter_total:
                                                total_quarterly_consumption += float(
                                                    quarter_total
                                                )

                                    # –†–∞—Å–ø—Ä–µ–¥–µ–ª—è–µ–º by_usage –ø–æ –∫–≤–∞—Ä—Ç–∞–ª–∞–º –ø—Ä–æ–ø–æ—Ä—Ü–∏–æ–Ω–∞–ª—å–Ω–æ –ø–æ—Ç—Ä–µ–±–ª–µ–Ω–∏—é
                                    if (
                                        total_quarterly_consumption > 0
                                        and annual_total > 0
                                    ):
                                        for quarter_key, quarter_data in resources_data[
                                            "electricity"
                                        ].items():
                                            if quarter_key == "ANNUAL":
                                                continue
                                            if isinstance(quarter_data, dict):
                                                quarter_total = quarter_data.get(
                                                    "quarter_totals", {}
                                                ).get("active_kwh", 0)
                                                if quarter_total and quarter_total > 0:
                                                    # –ü—Ä–æ–ø–æ—Ä—Ü–∏–æ–Ω–∞–ª—å–Ω–æ–µ —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ
                                                    quarter_ratio = (
                                                        float(quarter_total)
                                                        / total_quarterly_consumption
                                                    )
                                                    quarter_by_usage = {
                                                        category: float(value)
                                                        * quarter_ratio
                                                        for category, value in annual_by_usage.items()
                                                    }
                                                    quarter_data["by_usage"] = (
                                                        quarter_by_usage
                                                    )
                                                    logger.debug(
                                                        "–†–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω canonical by_usage –¥–ª—è %s %s: %s (ratio=%.3f)",
                                                        resource_name,
                                                        quarter_key,
                                                        list(quarter_by_usage.keys()),
                                                        quarter_ratio,
                                                    )

                                        logger.info(
                                            "‚úÖ Canonical by_usage —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω –ø–æ %d –∫–≤–∞—Ä—Ç–∞–ª–∞–º electricity (annual_total=%.2f, quarterly_total=%.2f)",
                                            sum(
                                                1
                                                for k, v in resources_data[
                                                    "electricity"
                                                ].items()
                                                if k != "ANNUAL"
                                                and isinstance(v, dict)
                                                and v.get("by_usage")
                                            ),
                                            annual_total,
                                            total_quarterly_consumption,
                                        )
                                    else:
                                        logger.warning(
                                            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–∏—Ç—å canonical by_usage: annual_total=%.2f, quarterly_total=%.2f",
                                            annual_total,
                                            total_quarterly_consumption,
                                        )
                        except Exception as e:
                            logger.warning(
                                "–û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–æ–±–∞–≤–ª–µ–Ω–∏–∏ canonical by_usage: %s", e
                            )
                        logger.info(
                            "–î–æ–±–∞–≤–ª–µ–Ω –≥–æ–¥–æ–≤–æ–π –∏—Ç–æ–≥ Canonical –¥–ª—è —Ä–µ—Å—É—Ä—Å–∞ %s: %.2f (%s)",
                            resource_name,
                            float(annual_value),
                            key,
                        )
            except Exception as e:
                logger.warning(
                    "–ù–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–∏–º–µ–Ω–∏—Ç—å –≥–æ–¥–æ–≤—ã–µ –∏—Ç–æ–≥–∏ Canonical –¥–ª—è –±–∞–ª–∞–Ω—Å–∞: %s", e
                )

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–†–∞—Å—á–µ—Ç —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º"
            envelope_json_path = AGGREGATED_DIR / f"{batch_id}_envelope.json"
            if not envelope_json_path.exists() and enterprise_id:
                uploads = database.list_uploads_for_enterprise(enterprise_id)
                for upload_item in uploads:
                    upload_batch_id = upload_item.get("batch_id")
                    if upload_batch_id:
                        candidate_path = (
                            AGGREGATED_DIR / f"{upload_batch_id}_envelope.json"
                        )
                        if candidate_path.exists():
                            envelope_json_path = candidate_path
                            break

            if envelope_json_path.exists():
                logger.info(
                    "–ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è JSON —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º: %s",
                    envelope_json_path,
                )
                import json as json_module

                envelope_data = json_module.loads(
                    envelope_json_path.read_text(encoding="utf-8")
                )
                envelope_sheet_names = ["02_–ò—Å—Ö–æ–¥–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ", "–û–≥—Ä–∞–∂–¥–∞—é—â–∏–µ", "Envelope"]
                for sheet_name in envelope_sheet_names:
                    if sheet_name in workbook.sheetnames:
                        fill_building_envelope_sheet(
                            workbook, envelope_data, sheet_name=sheet_name
                        )
                        logger.info(
                            f"–õ–∏—Å—Ç '{sheet_name}' –∑–∞–ø–æ–ª–Ω–µ–Ω –¥–∞–Ω–Ω—ã–º–∏ —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º"
                        )
                        break

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–ë–∞–ª–∞–Ω—Å"
            balans_sheet_names = ["–ë–∞–ª–∞–Ω—Å", "–ë–∞–ª–∞–Ω—Å ", "04_–ë–∞–ª–∞–Ω—Å", "Balance", "Balans"]
            balans_sheet = None
            for sheet_name in balans_sheet_names:
                if sheet_name in workbook.sheetnames:
                    balans_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not balans_sheet:
                for ws_name in workbook.sheetnames:
                    if "–±–∞–ª–∞–Ω—Å" in ws_name.lower() or "balance" in ws_name.lower():
                        balans_sheet = workbook[ws_name]
                        break

            # –°–æ–∑–¥–∞–µ–º –ª–∏—Å—Ç, –µ—Å–ª–∏ –µ–≥–æ –Ω–µ—Ç
            if not balans_sheet:
                logger.info("–õ–∏—Å—Ç '–ë–∞–ª–∞–Ω—Å' –Ω–µ –Ω–∞–π–¥–µ–Ω, —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π")
                balans_sheet = workbook.create_sheet(title="04_–ë–∞–ª–∞–Ω—Å")
                # –ö–†–ò–¢–ò–ß–ï–°–ö–û–ï –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –≤—ã–∑—ã–≤–∞–µ–º fill_balans_sheet –¥–ª—è –Ω–æ–≤–æ–≥–æ –ª–∏—Å—Ç–∞
                logger.info(f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –ª–∏—Å—Ç–∞ '{balans_sheet.title}'")
                fill_balans_sheet(balans_sheet, resources_data)
                logger.info(f"‚úÖ –ù–æ–≤—ã–π –ª–∏—Å—Ç '{balans_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            if balans_sheet:
                # ========== –î–ï–¢–ê–õ–¨–ù–ê–Ø –î–ò–ê–ì–ù–û–°–¢–ò–ö–ê –ü–ï–†–ï–î fill_balans_sheet ==========
                logger.info("=" * 80)
                logger.info(
                    "üîç –î–ò–ê–ì–ù–û–°–¢–ò–ö–ê: –°—Ç—Ä—É–∫—Ç—É—Ä–∞ resources_data –ø–µ—Ä–µ–¥ fill_balans_sheet"
                )
                logger.info("=" * 80)

                # 1. –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–∏–ø –∏ —Å—Ç—Ä—É–∫—Ç—É—Ä—É resources_data
                logger.info(f"üìã –¢–∏–ø resources_data: {type(resources_data)}")
                logger.info(
                    f"üìã –ö–ª—é—á–∏ resources_data: {list(resources_data.keys())[:20]}"
                )

                # 2. –ü—Ä–æ–≤–µ—Ä—è–µ–º –æ–±–∞ —Ñ–æ—Ä–º–∞—Ç–∞ –¥–æ—Å—Ç—É–ø–∞ –∫ electricity
                electricity_direct = resources_data.get("electricity", {})
                electricity_via_resources = (
                    resources_data.get("resources", {}).get("electricity", {})
                    if isinstance(resources_data.get("resources"), dict)
                    else {}
                )

                logger.info(
                    f"üìä electricity (–ø—Ä—è–º–æ–π –¥–æ—Å—Ç—É–ø): type={type(electricity_direct)}, len={len(electricity_direct) if isinstance(electricity_direct, dict) else 'N/A'}"
                )
                logger.info(
                    f"üìä electricity (—á–µ—Ä–µ–∑ resources): type={type(electricity_via_resources)}, len={len(electricity_via_resources) if isinstance(electricity_via_resources, dict) else 'N/A'}"
                )

                # –ò—Å–ø–æ–ª—å–∑—É–µ–º —Ç–æ—Ç, –∫–æ—Ç–æ—Ä—ã–π –Ω–µ –ø—É—Å—Ç–æ–π
                electricity = (
                    electricity_direct
                    if electricity_direct
                    else electricity_via_resources
                )

                if not electricity:
                    logger.error("‚ùå –ö–†–ò–¢–ò–ß–ù–û: electricity –ø—É—Å—Ç –≤ –æ–±–æ–∏—Ö —Ñ–æ—Ä–º–∞—Ç–∞—Ö!")
                    logger.error(
                        f"   resources_data.get('electricity'): {electricity_direct}"
                    )
                    logger.error(
                        f"   resources_data.get('resources', {{}}).get('electricity'): {electricity_via_resources}"
                    )
                    logger.error(
                        "   –ü–æ–ª–Ω–∞—è —Å—Ç—Ä—É–∫—Ç—É—Ä–∞ resources_data (–ø–µ—Ä–≤—ã–µ 3 —É—Ä–æ–≤–Ω—è):"
                    )
                    import json

                    try:
                        # –ë–µ–∑–æ–ø–∞—Å–Ω—ã–π –≤—ã–≤–æ–¥ —Å—Ç—Ä—É–∫—Ç—É—Ä—ã (–±–µ–∑ —Ä–µ–∫—É—Ä—Å–∏–∏)
                        struct_preview = {}
                        for key, value in list(resources_data.items())[:10]:
                            if isinstance(value, dict):
                                struct_preview[key] = {
                                    "type": "dict",
                                    "keys": list(value.keys())[:10],
                                    "sample": {
                                        k: type(v).__name__
                                        for k, v in list(value.items())[:5]
                                    },
                                }
                            else:
                                struct_preview[key] = type(value).__name__
                        logger.error(
                            f"   {json.dumps(struct_preview, indent=2, ensure_ascii=False)}"
                        )
                    except Exception as e:
                        logger.error(f"   –û—à–∏–±–∫–∞ –ø—Ä–∏ –≤—ã–≤–æ–¥–µ —Å—Ç—Ä—É–∫—Ç—É—Ä—ã: {e}")
                else:
                    logger.info(f"‚úÖ electricity –Ω–∞–π–¥–µ–Ω: {len(electricity)} —ç–ª–µ–º–µ–Ω—Ç–æ–≤")
                    logger.info(
                        f"   –ö–ª—é—á–∏ electricity: {list(electricity.keys())[:10]}"
                    )

                # 3. –î–µ—Ç–∞–ª—å–Ω—ã–π –∞–Ω–∞–ª–∏–∑ –∫–≤–∞—Ä—Ç–∞–ª–æ–≤
                quarters_with_by_usage = []
                quarters_without_by_usage = []
                annual_has_by_usage = False

                if electricity:
                    for quarter, quarter_data in electricity.items():
                        if quarter == "ANNUAL":
                            # ANNUAL –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –æ—Ç–¥–µ–ª—å–Ω–æ
                            continue
                        if not isinstance(quarter_data, dict):
                            logger.warning(
                                f"   ‚ö†Ô∏è –ö–≤–∞—Ä—Ç–∞–ª {quarter}: –Ω–µ —è–≤–ª—è–µ—Ç—Å—è dict, type={type(quarter_data)}"
                            )
                            continue
                        by_usage = quarter_data.get("by_usage")
                        quarter_totals = quarter_data.get("quarter_totals", {})
                        active_kwh = quarter_totals.get("active_kwh", 0)

                        if (
                            by_usage
                            and isinstance(by_usage, dict)
                            and len(by_usage) > 0
                        ):
                            quarters_with_by_usage.append(quarter)
                            logger.debug(
                                f"   ‚úÖ –ö–≤–∞—Ä—Ç–∞–ª {quarter}: by_usage –Ω–∞–π–¥–µ–Ω - {list(by_usage.keys())}, active_kwh={active_kwh}"
                            )
                        else:
                            quarters_without_by_usage.append(quarter)
                            logger.debug(
                                f"   ‚ö†Ô∏è –ö–≤–∞—Ä—Ç–∞–ª {quarter}: by_usage –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç, active_kwh={active_kwh}, quarter_totals keys={list(quarter_totals.keys())}"
                            )

                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º ANNUAL (–º–æ–∂–µ—Ç —Å–æ–¥–µ—Ä–∂–∞—Ç—å by_usage –∏–∑ canonical)
                    annual_data = electricity.get("ANNUAL")
                    if annual_data and isinstance(annual_data, dict):
                        annual_by_usage = annual_data.get("by_usage")
                        annual_totals = annual_data.get("quarter_totals", {})
                        if (
                            annual_by_usage
                            and isinstance(annual_by_usage, dict)
                            and len(annual_by_usage) > 0
                        ):
                            annual_has_by_usage = True
                            logger.info(
                                f"   ‚úÖ ANNUAL: by_usage –Ω–∞–π–¥–µ–Ω - {list(annual_by_usage.keys())}, totals={list(annual_totals.keys())}"
                            )
                        else:
                            logger.info(
                                f"   ‚ö†Ô∏è ANNUAL: by_usage –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç, totals={list(annual_totals.keys())}"
                            )

                # 4. –ò—Ç–æ–≥–æ–≤–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
                logger.info("=" * 80)
                logger.info(
                    f"üìä –°–¢–ê–¢–ò–°–¢–ò–ö–ê: –∫–≤–∞—Ä—Ç–∞–ª–æ–≤ —Å by_usage: {len(quarters_with_by_usage)}, –±–µ–∑: {len(quarters_without_by_usage)}"
                )
                if annual_has_by_usage:
                    logger.info(
                        "   ‚úÖ ANNUAL —Å–æ–¥–µ—Ä–∂–∏—Ç by_usage –∏–∑ canonical (–¥–ª—è –≥–æ–¥–æ–≤–æ–≥–æ –∏—Ç–æ–≥–∞)"
                    )

                if quarters_without_by_usage:
                    logger.warning(
                        f"   ‚ö†Ô∏è –ö–≤–∞—Ä—Ç–∞–ª—ã –±–µ–∑ by_usage (–±—É–¥—É—Ç –ø—Ä–æ–ø—É—â–µ–Ω—ã –≤ –ª–∏—Å—Ç–µ –±–∞–ª–∞–Ω—Å–∞): {quarters_without_by_usage}"
                    )
                    logger.warning(
                        "   –≠—Ç–æ –º–æ–∂–µ—Ç –ø—Ä–∏–≤–µ—Å—Ç–∏ –∫ –æ—à–∏–±–∫–µ –≤–∞–ª–∏–¥–∞—Ü–∏–∏. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ, —á—Ç–æ:"
                    )
                    logger.warning("     1. –§–∞–π–ª pererashod.xlsx –∑–∞–≥—Ä—É–∂–µ–Ω –∏ –æ–±—Ä–∞–±–æ—Ç–∞–Ω")
                    logger.warning(
                        "     2. –§—É–Ω–∫—Ü–∏—è aggregate_usage_categories() –≤–µ—Ä–Ω—É–ª–∞ –¥–∞–Ω–Ω—ã–µ"
                    )
                    logger.warning(
                        "     3. –§—É–Ω–∫—Ü–∏—è distribute_categories_by_quarter() —É—Å–ø–µ—à–Ω–æ –≤—ã–ø–æ–ª–Ω–∏–ª–∞—Å—å"
                    )
                    logger.warning(
                        "     4. –ï—Å–ª–∏ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è canonical mode, –ø—Ä–æ–≤–µ—Ä—å—Ç–µ –Ω–∞–ª–∏—á–∏–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è –≤ CanonicalSourceData"
                    )

                if quarters_with_by_usage:
                    logger.info(
                        f"   ‚úÖ –ö–≤–∞—Ä—Ç–∞–ª—ã —Å by_usage (–±—É–¥—É—Ç –∑–∞–ø–æ–ª–Ω–µ–Ω—ã): {quarters_with_by_usage}"
                    )

                # 5. –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–∞—è –¥–∏–∞–≥–Ω–æ—Å—Ç–∏–∫–∞: –ø—Ä–æ–≤–µ—Ä—è–µ–º –∏—Å—Ç–æ—á–Ω–∏–∫ –¥–∞–Ω–Ω—ã—Ö
                if canonical_payload and canonical_payload.get("balance", {}).get(
                    "by_usage", {}
                ).get("electricity"):
                    canonical_by_usage = canonical_payload["balance"]["by_usage"][
                        "electricity"
                    ]
                    logger.info(
                        f"   üìä Canonical by_usage –¥–ª—è electricity –Ω–∞–π–¥–µ–Ω: {list(canonical_by_usage.keys())}"
                    )
                    if not annual_has_by_usage:
                        logger.warning(
                            "   ‚ö†Ô∏è Canonical by_usage —Å—É—â–µ—Å—Ç–≤—É–µ—Ç, –Ω–æ –Ω–µ –ø–æ–ø–∞–ª –≤ ANNUAL. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ –∏–Ω—Ç–µ–≥—Ä–∞—Ü–∏—é."
                        )
                elif excel_ai_mode_runtime in ("assist", "strict"):
                    logger.warning(
                        "   ‚ö†Ô∏è Canonical mode –∞–∫—Ç–∏–≤–µ–Ω, –Ω–æ by_usage –¥–ª—è electricity –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ canonical_payload"
                    )
                    logger.warning("      –í–æ–∑–º–æ–∂–Ω—ã–µ –ø—Ä–∏—á–∏–Ω—ã:")
                    logger.warning(
                        "      - –û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –≤ CanonicalSourceData"
                    )
                    logger.warning("      - annual_totals.electricity –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç")
                    logger.warning(
                        "      - compute_electricity_by_usage –≤–µ—Ä–Ω—É–ª –ø—É—Å—Ç–æ–π —Å–ª–æ–≤–∞—Ä—å"
                    )

                logger.info("=" * 80)

                # 6. –ù–æ—Ä–º–∞–ª–∏–∑—É–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É resources_data –¥–ª—è fill_balans_sheet
                # –£–±–µ–∂–¥–∞–µ–º—Å—è, —á—Ç–æ electricity –¥–æ—Å—Ç—É–ø–µ–Ω –≤ –æ–∂–∏–¥–∞–µ–º–æ–º —Ñ–æ—Ä–º–∞—Ç–µ
                if not electricity_direct and electricity_via_resources:
                    # –ï—Å–ª–∏ –¥–∞–Ω–Ω—ã–µ –µ—Å—Ç—å —Ç–æ–ª—å–∫–æ —á–µ—Ä–µ–∑ resources, –Ω–æ—Ä–º–∞–ª–∏–∑—É–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É
                    logger.info(
                        "üîÑ –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è: –ø–µ—Ä–µ–Ω–æ—Å–∏–º electricity –∏–∑ resources –≤ –∫–æ—Ä–µ–Ω—å resources_data"
                    )
                    resources_data["electricity"] = electricity_via_resources
                    electricity = electricity_via_resources

                # –¢–∞–∫–∂–µ —É–±–µ–∂–¥–∞–µ–º—Å—è, —á—Ç–æ –µ—Å—Ç—å —Å—Ç—Ä—É–∫—Ç—É—Ä–∞ resources
                if "resources" not in resources_data:
                    resources_data["resources"] = {}
                if "electricity" not in resources_data["resources"]:
                    resources_data["resources"]["electricity"] = electricity

                logger.info("‚úÖ –í—ã–∑–æ–≤ fill_balans_sheet —Å –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–º–∏ –¥–∞–Ω–Ω—ã–º–∏")
                logger.info(
                    f"   electricity –¥–æ—Å—Ç—É–ø–µ–Ω: {len(electricity) if electricity else 0} —ç–ª–µ–º–µ–Ω—Ç–æ–≤"
                )
                logger.info("=" * 80)

                fill_balans_sheet(balans_sheet, resources_data)
                logger.info(f"‚úÖ –õ–∏—Å—Ç '{balans_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–î–∏–Ω–∞–º–∏–∫–∞ —Å—Ä"
            dinamika_sheet_names = [
                "–î–∏–Ω–∞–º–∏–∫–∞ —Å—Ä",
                "–î–∏–Ω–∞–º–∏–∫–∞ —Å—Ä ",
                "–î–∏–Ω–∞–º–∏–∫–∞",
                "05_–î–∏–Ω–∞–º–∏–∫–∞",
                "Dynamics",
            ]
            dinamika_sheet = None
            for sheet_name in dinamika_sheet_names:
                if sheet_name in workbook.sheetnames:
                    dinamika_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not dinamika_sheet:
                for ws_name in workbook.sheetnames:
                    if "–¥–∏–Ω–∞–º–∏–∫–∞" in ws_name.lower() or "dynamics" in ws_name.lower():
                        dinamika_sheet = workbook[ws_name]
                        break

            if dinamika_sheet:
                logger.info(f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{dinamika_sheet.title}'")
                fill_dinamika_sheet(dinamika_sheet, resources_data)
                logger.info(f"–õ–∏—Å—Ç '{dinamika_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–º–∞–∑—É—Ç,—É–≥–æ–ª—å 5"
            fuel_dynamics_sheet_names = [
                "–º–∞–∑—É—Ç,—É–≥–æ–ª—å 5",
                "–º–∞–∑—É—Ç,—É–≥–æ–ª—å 5 ",
                "–º–∞–∑—É—Ç,—É–≥–æ–ª—å",
                "06_–ú–∞–∑—É—Ç_–£–≥–æ–ª—å",
                "Fuel Dynamics",
            ]
            fuel_dynamics_sheet = None
            for sheet_name in fuel_dynamics_sheet_names:
                if sheet_name in workbook.sheetnames:
                    fuel_dynamics_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not fuel_dynamics_sheet:
                for ws_name in workbook.sheetnames:
                    if (
                        "–º–∞–∑—É—Ç" in ws_name.lower()
                        or "—É–≥–æ–ª—å" in ws_name.lower()
                        or "fuel" in ws_name.lower()
                    ):
                        fuel_dynamics_sheet = workbook[ws_name]
                        break

            if fuel_dynamics_sheet:
                logger.info(f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{fuel_dynamics_sheet.title}'")
                fill_fuel_dynamics_sheet(fuel_dynamics_sheet, resources_data)
                logger.info(f"–õ–∏—Å—Ç '{fuel_dynamics_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–†–∞—Å—Ö–æ–¥ –Ω–∞ –µ–¥.–ø"
            specific_consumption_sheet_names = [
                "–†–∞—Å—Ö–æ–¥ –Ω–∞ –µ–¥.–ø",
                "–†–∞—Å—Ö–æ–¥  –Ω–∞ –µ–¥.–ø",
                "–†–∞—Å—Ö–æ–¥  –Ω–∞ –µ–¥.–ø ",
                "07_–†–∞—Å—Ö–æ–¥_–Ω–∞_–µ–¥",
                "Specific Consumption",
            ]
            specific_consumption_sheet = None
            for sheet_name in specific_consumption_sheet_names:
                if sheet_name in workbook.sheetnames:
                    specific_consumption_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not specific_consumption_sheet:
                for ws_name in workbook.sheetnames:
                    if "—Ä–∞—Å—Ö–æ–¥" in ws_name.lower() and "–µ–¥" in ws_name.lower():
                        specific_consumption_sheet = workbook[ws_name]
                        break

            if specific_consumption_sheet:
                logger.info(f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{specific_consumption_sheet.title}'")
                fill_specific_consumption_sheet(
                    specific_consumption_sheet, resources_data
                )
                logger.info(f"–õ–∏—Å—Ç '{specific_consumption_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "–ú–µ—Ä–∏–∞–ø—Ä–∏—è—Ç–∏—è 1"
            meropriyatiya_sheet_names = [
                "–ú–µ—Ä–∏–∞–ø—Ä–∏—è—Ç–∏—è 1",
                "–ú–µ—Ä–∏–∞–ø—Ä–∏—è—Ç–∏—è 1 ",
                "–ú–µ—Ä–æ–ø—Ä–∏—è—Ç–∏—è",
                "08_–ú–µ—Ä–æ–ø—Ä–∏—è—Ç–∏—è",
                "Measures",
            ]
            meropriyatiya_sheet = None
            # –°–Ω–∞—á–∞–ª–∞ –∏—â–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ (—Å —É—á–µ—Ç–æ–º –ø—Ä–æ–±–µ–ª–æ–≤)
            for sheet_name in meropriyatiya_sheet_names:
                if sheet_name in workbook.sheetnames:
                    meropriyatiya_sheet = workbook[sheet_name]
                    break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏, –∏—â–µ–º —Å —É—á–µ—Ç–æ–º –ø—Ä–æ–±–µ–ª–æ–≤ –≤ –∫–æ–Ω—Ü–µ
            if not meropriyatiya_sheet:
                for ws_name in workbook.sheetnames:
                    ws_name_stripped = ws_name.strip()
                    for target_name in meropriyatiya_sheet_names:
                        if ws_name_stripped == target_name.strip():
                            meropriyatiya_sheet = workbook[ws_name]
                            break
                    if meropriyatiya_sheet:
                        break
            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É
            if not meropriyatiya_sheet:
                for ws_name in workbook.sheetnames:
                    if (
                        "–º–µ—Ä–æ–ø—Ä–∏—è—Ç–∏—è" in ws_name.lower()
                        or "–º–µ—Ä–∏–∞–ø—Ä–∏—è—Ç–∏—è" in ws_name.lower()
                        or "measures" in ws_name.lower()
                    ):
                        meropriyatiya_sheet = workbook[ws_name]
                        break

            if meropriyatiya_sheet:
                logger.info(f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{meropriyatiya_sheet.title}'")
                fill_meropriyatiya_sheet(meropriyatiya_sheet)
                logger.info(f"–õ–∏—Å—Ç '{meropriyatiya_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω")

            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ª–∏—Å—Ç "Monthly" (–º–µ—Å—è—á–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ)
            monthly_sheet_names = [
                "Monthly",
                "MONTHLY",
                "–ú–µ—Å—è—á–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ",
                "–ú–µ—Å—è—á–Ω—ã–π",
                "–º–µ—Å—è—á–Ω—ã–µ",
            ]
            monthly_sheet = None
            for sheet_name in monthly_sheet_names:
                if sheet_name in workbook.sheetnames:
                    monthly_sheet = workbook[sheet_name]
                    break

            # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –∏—â–µ–º —á–∞—Å—Ç–∏—á–Ω–æ–µ
            if not monthly_sheet:
                for sheet_name in workbook.sheetnames:
                    sheet_lower = sheet_name.lower()
                    if (
                        "monthly" in sheet_lower
                        or "–º–µ—Å—è—Ü" in sheet_lower
                        or "–º–µ—Å—è—á–Ω" in sheet_lower
                    ):
                        monthly_sheet = workbook[sheet_name]
                        logger.info(
                            f"–ù–∞–π–¥–µ–Ω –ª–∏—Å—Ç –º–µ—Å—è—á–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –ø–æ —á–∞—Å—Ç–∏—á–Ω–æ–º—É —Å–æ–≤–ø–∞–¥–µ–Ω–∏—é: '{sheet_name}'"
                        )
                        break

            if monthly_sheet:
                logger.info(f"–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ª–∏—Å—Ç–∞ '{monthly_sheet.title}'")
                try:
                    fill_monthly_sheet(monthly_sheet, resources_data)
                    logger.info(f"–õ–∏—Å—Ç '{monthly_sheet.title}' –∑–∞–ø–æ–ª–Ω–µ–Ω —É—Å–ø–µ—à–Ω–æ")
                except Exception as monthly_exc:
                    logger.error(
                        f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏–∏ –ª–∏—Å—Ç–∞ '{monthly_sheet.title}': {monthly_exc}",
                        exc_info=True,
                    )
            else:
                logger.warning(
                    "–õ–∏—Å—Ç 'Monthly' –Ω–µ –Ω–∞–π–¥–µ–Ω. –î–æ—Å—Ç—É–ø–Ω—ã–µ –ª–∏—Å—Ç—ã: "
                    + ", ".join(workbook.sheetnames[:10])
                )

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –≤—Å–µ –ª–∏ –ª–∏—Å—Ç—ã –∑–∞–ø–æ–ª–Ω–µ–Ω—ã
            filled_sheets = set()
            equipment_sheet_name = (
                None  # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º –ø–µ—Ä–µ–º–µ–Ω–Ω—É—é –¥–ª—è –∏–º–µ–Ω–∏ –ª–∏—Å—Ç–∞ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è
            )

            # –°–æ–±–∏—Ä–∞–µ–º —Å–ø–∏—Å–æ–∫ –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã—Ö –ª–∏—Å—Ç–æ–≤
            if struktura_sheet:
                filled_sheets.add(struktura_sheet.title)
            if nodes_sheet:
                filled_sheets.add(nodes_sheet.title)

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ª–∏—Å—Ç –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è (–ø–µ—Ä–µ–ø—Ä–æ–≤–µ—Ä—è–µ–º, –±—ã–ª –ª–∏ –æ–Ω –∑–∞–ø–æ–ª–Ω–µ–Ω)
            if equipment_json_path.exists():
                for sheet_name in workbook.sheetnames:
                    sheet_lower = sheet_name.lower()
                    if any(
                        keyword in sheet_lower
                        for keyword in ["equipment", "–æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ", "–∞–Ω–∞–ª–∏–∑"]
                    ):
                        filled_sheets.add(sheet_name)
                        equipment_sheet_name = sheet_name
                        break

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ª–∏—Å—Ç —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º
            envelope_sheet_name = None
            if envelope_json_path.exists():
                for sheet_name in workbook.sheetnames:
                    if any(
                        name in sheet_name
                        for name in ["02_–ò—Å—Ö–æ–¥–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ", "–û–≥—Ä–∞–∂–¥–∞—é—â–∏–µ", "Envelope"]
                    ):
                        filled_sheets.add(sheet_name)
                        envelope_sheet_name = sheet_name
                        break

            if balans_sheet:
                filled_sheets.add(balans_sheet.title)
            if dinamika_sheet:
                filled_sheets.add(dinamika_sheet.title)
            if fuel_dynamics_sheet:
                filled_sheets.add(fuel_dynamics_sheet.title)
            if specific_consumption_sheet:
                filled_sheets.add(specific_consumption_sheet.title)
            if meropriyatiya_sheet:
                filled_sheets.add(meropriyatiya_sheet.title)
            if monthly_sheet:
                filled_sheets.add(monthly_sheet.title)

            # –ù–∞—Ö–æ–¥–∏–º –Ω–µ–∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã–µ –ª–∏—Å—Ç—ã
            all_sheets = set(workbook.sheetnames)
            unfilled_sheets = all_sheets - filled_sheets

            if unfilled_sheets:
                logger.warning(
                    f"‚ö†Ô∏è  –ù–∞–π–¥–µ–Ω—ã –Ω–µ–∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã–µ –ª–∏—Å—Ç—ã ({len(unfilled_sheets)}): "
                    f"{', '.join(sorted(unfilled_sheets))}"
                )
                logger.info(
                    f"üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è: "
                    f"–∑–∞–ø–æ–ª–Ω–µ–Ω–æ {len(filled_sheets)}/{len(all_sheets)} –ª–∏—Å—Ç–æ–≤ "
                    f"({round(len(filled_sheets) / len(all_sheets) * 100, 1)}%)"
                )
            else:
                logger.info(
                    f"‚úÖ –í—Å–µ –ª–∏—Å—Ç—ã –∑–∞–ø–æ–ª–Ω–µ–Ω—ã: {len(filled_sheets)}/{len(all_sheets)} –ª–∏—Å—Ç–æ–≤ "
                    f"({round(len(filled_sheets) / len(all_sheets) * 100, 1)}%)"
                )

            # –í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ —Ñ–æ—Ä–º—É–ª —Å –ø–æ–º–æ—â—å—é AI (–µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–Ω–æ)
            if HAS_FORMULA_RESTORER:
                try:
                    logger.info("–ù–∞—á–∏–Ω–∞—é –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ —Ñ–æ—Ä–º—É–ª –≤ –ø–∞—Å–ø–æ—Ä—Ç–µ...")
                    restorer = AIFormulaRestorer()

                    # –ò—â–µ–º –≤—Å–µ —è—á–µ–π–∫–∏ —Å –æ—à–∏–±–∫–∞–º–∏ #REF! –∏ –≤–æ—Å—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –∏—Ö
                    restored_count = 0
                    total_ref_errors = 0

                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.data_type == "f" and cell.value:
                                    formula_str = str(cell.value)
                                    if "#REF!" in formula_str:
                                        total_ref_errors += 1
                                        cell_coord = cell.coordinate

                                        restored_formula = restorer.restore_ref_error(
                                            workbook,
                                            sheet_name,
                                            cell_coord,
                                            formula_str,
                                        )

                                        if restored_formula:
                                            cell.value = restored_formula
                                            restored_count += 1
                                            logger.debug(
                                                f"–í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∞ —Ñ–æ—Ä–º—É–ª–∞ –≤ {sheet_name}!{cell_coord}: "
                                                f"{formula_str} -> {restored_formula}"
                                            )

                    if total_ref_errors > 0:
                        logger.info(
                            f"–í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ —Ñ–æ—Ä–º—É–ª –∑–∞–≤–µ—Ä—à–µ–Ω–æ: "
                            f"–≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–æ {restored_count}/{total_ref_errors} —Ñ–æ—Ä–º—É–ª"
                        )
                    else:
                        logger.info(
                            "–û—à–∏–±–æ–∫ #REF! –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ –Ω–µ —Ç—Ä–µ–±—É–µ—Ç—Å—è"
                        )
                except Exception as restore_exc:
                    logger.warning(
                        f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–∏ —Ñ–æ—Ä–º—É–ª: {restore_exc}. –ü—Ä–æ–¥–æ–ª–∂–∞—é –±–µ–∑ –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏—è."
                    )

            debug_output = AGGREGATED_DIR / f"{batch_id}_filler_passport.xlsx"
            workbook.save(debug_output)
            logger.info("–ü–∞—Å–ø–æ—Ä—Ç —Å–æ—Ö—Ä–∞–Ω—ë–Ω –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏: %s", debug_output)
            workbook.save(output_file)
            logger.info(f"–ü–∞—Å–ø–æ—Ä—Ç —Å–æ—Ö—Ä–∞–Ω—ë–Ω: {output_file}")

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Ñ–∞–π–ª —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –∏ –Ω–µ –ø—É—Å—Ç–æ–π
            if not output_file.exists():
                raise HTTPException(
                    status_code=500,
                    detail=f"–§–∞–π–ª –ø–∞—Å–ø–æ—Ä—Ç–∞ –Ω–µ –±—ã–ª —Å–æ–∑–¥–∞–Ω: {output_file}",
                )
            file_size = output_file.stat().st_size
            if file_size == 0:
                raise HTTPException(
                    status_code=500, detail=f"–§–∞–π–ª –ø–∞—Å–ø–æ—Ä—Ç–∞ –ø—É—Å—Ç–æ–π: {output_file}"
                )
            logger.info(
                f"‚úÖ –ü–∞—Å–ø–æ—Ä—Ç —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω (filler): {output_file} (—Ä–∞–∑–º–µ—Ä: {file_size} –±–∞–π—Ç)"
            )

            return FileResponse(
                path=str(output_file),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"–≠–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç_{enterprise_data['name']}_{batch_id[:8]}.xlsx",
            )
        except Exception as filler_exc:
            logger.exception(
                "–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ fill_energy_passport: %s", filler_exc
            )
            if not HAS_GENERATOR:
                raise HTTPException(
                    status_code=500, detail=f"–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏: {filler_exc}"
                ) from filler_exc

    if not HAS_GENERATOR:
        raise HTTPException(status_code=503, detail="–ì–µ–Ω–µ—Ä–∞—Ç–æ—Ä –ø–∞—Å–ø–æ—Ä—Ç–æ–≤ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω")

    try:
        logger.info(
            f"–ì–µ–Ω–µ—Ä–∞—Ü–∏—è –ø–∞—Å–ø–æ—Ä—Ç–∞ —á–µ—Ä–µ–∑ PKM690ExcelGenerator –¥–ª—è {enterprise_data['name']}"
        )
        generator = PKM690ExcelGenerator(
            enterprise_data=enterprise_data, energy_data=aggregated
        )

        success = generator.create_energy_passport(str(output_file))

        if not success or not output_file.exists():
            raise HTTPException(status_code=500, detail="–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –ø–∞—Å–ø–æ—Ä—Ç–∞")

        logger.info(f"‚úÖ –ü–∞—Å–ø–æ—Ä—Ç —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω: {output_file}")

        return FileResponse(
            path=str(output_file),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"–≠–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç_{enterprise_data['name']}_{batch_id[:8]}.xlsx",
        )

    except Exception as e:
        logger.exception(f"–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –ø–∞—Å–ø–æ—Ä—Ç–∞: {e}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏: {str(e)}")


@app.post("/api/generate-word-report/{batch_id}")
async def generate_word_report(
    batch_id: str,
    skip_readiness_check: bool = Query(
        False, description="–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å –ø—Ä–æ–≤–µ—Ä–∫—É –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö"
    ),
):
    """
    –ì–µ–Ω–µ—Ä–∞—Ü–∏—è Word –æ—Ç—á–µ—Ç–∞ —ç–Ω–µ—Ä–≥–æ–∞—É–¥–∏—Ç–∞ –¥–ª—è batch_id
    """
    if not HAS_WORD_GENERATOR:
        raise HTTPException(
            status_code=503,
            detail="–ì–µ–Ω–µ—Ä–∞—Ç–æ—Ä Word –æ—Ç—á–µ—Ç–æ–≤ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω (python-docx –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω)",
        )

    # 1. –ü–æ–ª—É—á–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É
    upload = database.get_upload_by_batch(batch_id)
    if not upload:
        raise HTTPException(status_code=404, detail=f"–ó–∞–≥—Ä—É–∑–∫–∞ {batch_id} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞")

    # 2. –ü—Ä–æ–≤–µ—Ä–∫–∞ –≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö
    if not skip_readiness_check:
        enterprise_id = upload.get("enterprise_id")
        if enterprise_id:
            readiness = validate_generation_readiness(enterprise_id)
            if not readiness["ready"]:
                logger.warning(
                    f"–ü–æ–ø—ã—Ç–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ Word –æ—Ç—á–µ—Ç–∞ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id} –ø—Ä–∏ –Ω–µ–≥–æ—Ç–æ–≤–Ω–æ—Å—Ç–∏ –¥–∞–Ω–Ω—ã—Ö"
                )
                detail = {
                    "message": "–î–∞–Ω–Ω—ã–µ –Ω–µ –≥–æ—Ç–æ–≤—ã –¥–ª—è –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ Word –æ—Ç—á–µ—Ç–∞",
                    "missing_resources": readiness.get("missing_resources", []),
                    "warnings": readiness.get("warnings", []),
                }
                raise HTTPException(status_code=400, detail=detail)

    # 3. –ê–≥—Ä–µ–≥–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤
    enterprise_id = upload.get("enterprise_id")
    logger.info(
        f"–ì–µ–Ω–µ—Ä–∞—Ü–∏—è Word –æ—Ç—á–µ—Ç–∞ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id} (batch_id: {batch_id})"
    )
    logger.info("üìä –ì–µ–Ω–µ—Ä–∞—Ü–∏—è Word –æ—Ç—á–µ—Ç–∞ –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö —Å –ø—Ä–∏–º–µ–Ω–µ–Ω–∏–µ–º AI-–∞–Ω–∞–ª–∏–∑–∞")

    from utils.readiness_validator import _get_aggregated_data_for_enterprise

    aggregated = _get_aggregated_data_for_enterprise(enterprise_id)

    if not aggregated:
        raw_json = upload.get("raw_json")
        if not raw_json:
            raise HTTPException(status_code=400, detail="–î–∞–Ω–Ω—ã–µ –Ω–µ —Ä–∞—Å–ø–∞—Ä—Å–µ–Ω—ã")
        from utils.energy_aggregator import aggregate_from_db_json

        aggregated = aggregate_from_db_json(raw_json)

    if not aggregated:
        raise HTTPException(status_code=400, detail="–ù–µ —É–¥–∞–ª–æ—Å—å –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞—Ç—å –¥–∞–Ω–Ω—ã–µ")

    # 3.1. –ü—Ä–∏–º–µ–Ω—è–µ–º AI-–∞–Ω–∞–ª–∏–∑ –∫ –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã–º –¥–∞–Ω–Ω—ã–º (–µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–µ–Ω)
    try:
        from ai_energy_analysis import enhanced_energy_analysis

        HAS_AI_ENERGY_ANALYSIS = True
    except ImportError:
        HAS_AI_ENERGY_ANALYSIS = False
        logger.debug(
            "ai_energy_analysis –º–æ–¥—É–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. AI-–∞–Ω–∞–ª–∏–∑ –¥–∞–Ω–Ω—ã—Ö –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω."
        )

    ai_processed_data = None
    if HAS_AI_ENERGY_ANALYSIS:
        try:
            logger.info(
                "ü§ñ –ü—Ä–∏–º–µ–Ω–µ–Ω–∏–µ AI-–∞–Ω–∞–ª–∏–∑–∞ –∫ –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã–º –¥–∞–Ω–Ω—ã–º –¥–ª—è —É–ª—É—á—à–µ–Ω–∏—è –∫–∞—á–µ—Å—Ç–≤–∞ –æ—Ç—á–µ—Ç–∞..."
            )
            # –°–æ–∑–¥–∞–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É –¥–∞–Ω–Ω—ã—Ö –¥–ª—è AI-–∞–Ω–∞–ª–∏–∑–∞
            analysis_input = {
                "data": aggregated,
                "metadata": {
                    "enterprise_id": enterprise_id,
                    "batch_id": batch_id,
                    "source": "aggregated_from_files",
                },
            }
            ai_processed_data = enhanced_energy_analysis(analysis_input)
            if ai_processed_data and not ai_processed_data.get("error"):
                logger.info("‚úÖ AI-–∞–Ω–∞–ª–∏–∑ —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–ª –¥–∞–Ω–Ω—ã–µ")
                # –û–±–æ–≥–∞—â–∞–µ–º –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ AI-–∞–Ω–∞–ª–∏–∑–∞
                # –ò–∑–≤–ª–µ–∫–∞–µ–º –∏–Ω—Å–∞–π—Ç—ã –∏–∑ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ AI-–∞–Ω–∞–ª–∏–∑–∞
                ai_insights = {}
                ai_recommendations = []

                # –§–æ—Ä–º–∏—Ä—É–µ–º –∏–Ω—Å–∞–π—Ç—ã –∏–∑ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏, –∞–Ω–æ–º–∞–ª–∏–π, —ç—Ñ—Ñ–µ–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –∏ —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏—è
                summary = ai_processed_data.get("summary", {})
                if summary:
                    ai_insights["–°—Ç–∞—Ç—É—Å –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏"] = (
                        "–î–∞–Ω–Ω—ã–µ –≤–∞–ª–∏–¥–Ω—ã"
                        if summary.get("is_valid")
                        else "–û–±–Ω–∞—Ä—É–∂–µ–Ω—ã –ø—Ä–æ–±–ª–µ–º—ã —Å –¥–∞–Ω–Ω—ã–º–∏"
                    )
                    ai_insights["–ù–∞–ª–∏—á–∏–µ –∞–Ω–æ–º–∞–ª–∏–π"] = (
                        f"–û–±–Ω–∞—Ä—É–∂–µ–Ω–æ –∞–Ω–æ–º–∞–ª–∏–π: {ai_processed_data.get('anomalies', {}).get('anomaly_count', 0)}"
                        if summary.get("has_anomalies")
                        else "–ê–Ω–æ–º–∞–ª–∏–π –Ω–µ –æ–±–Ω–∞—Ä—É–∂–µ–Ω–æ"
                    )
                    ai_insights["–°–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏–µ –Ω–æ—Ä–º–∞–º"] = (
                        "–°–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É–µ—Ç"
                        if summary.get("is_compliant")
                        else "–¢—Ä–µ–±—É–µ—Ç—Å—è —É–ª—É—á—à–µ–Ω–∏–µ"
                    )
                    efficiency_class = summary.get("efficiency_class", "N/A")
                    if efficiency_class != "N/A":
                        ai_insights["–ö–ª–∞—Å—Å —ç–Ω–µ—Ä–≥–æ—ç—Ñ—Ñ–µ–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏"] = efficiency_class

                # –î–æ–±–∞–≤–ª—è–µ–º —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –∏–∑ –∞–Ω–∞–ª–∏–∑–∞ —ç—Ñ—Ñ–µ–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
                efficiency = ai_processed_data.get("efficiency", {})
                if efficiency and "recommendations" in efficiency:
                    ai_recommendations.extend(efficiency.get("recommendations", []))

                # –î–æ–±–∞–≤–ª—è–µ–º —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –∏–∑ –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏—è
                compliance = ai_processed_data.get("compliance", {})
                if compliance and "recommendations" in compliance:
                    ai_recommendations.extend(compliance.get("recommendations", []))

                # –°–æ—Ö—Ä–∞–Ω—è–µ–º AI-–∏–Ω—Å–∞–π—Ç—ã –∏ —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –≤ –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
                if ai_insights:
                    aggregated["ai_insights"] = ai_insights
                if ai_recommendations:
                    aggregated["ai_recommendations"] = ai_recommendations[
                        :20
                    ]  # –ü–µ—Ä–≤—ã–µ 20 —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–π

                logger.info(
                    f"AI-–∞–Ω–∞–ª–∏–∑ –¥–æ–±–∞–≤–∏–ª {len(ai_insights)} –∏–Ω—Å–∞–π—Ç–æ–≤ –∏ {len(ai_recommendations)} —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–π"
                )
            else:
                error_msg = (
                    ai_processed_data.get("error", "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è –æ—à–∏–±–∫–∞")
                    if ai_processed_data
                    else "–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö"
                )
                logger.info(
                    f"AI-–∞–Ω–∞–ª–∏–∑ –Ω–µ –≤–µ—Ä–Ω—É–ª –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ ({error_msg}), –∏—Å–ø–æ–ª—å–∑—É–µ–º –∏—Å—Ö–æ–¥–Ω—ã–µ –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ"
                )
        except Exception as ai_exc:
            logger.warning(
                f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–∏–º–µ–Ω–µ–Ω–∏–∏ AI-–∞–Ω–∞–ª–∏–∑–∞ (–ø—Ä–æ–¥–æ–ª–∂–∞–µ–º –±–µ–∑ AI): {ai_exc}"
            )
            # –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –≥–µ–Ω–µ—Ä–∞—Ü–∏—é –±–µ–∑ AI-–æ–±–æ–≥–∞—â–µ–Ω–∏—è

    # 4. –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –¥–∞–Ω–Ω—ã—Ö –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è
    enterprise = database.get_enterprise_by_id(enterprise_id) if enterprise_id else None
    enterprise_data = {
        "id": enterprise_id,
        "name": enterprise.get("name", "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ–µ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ")
        if enterprise
        else upload.get("enterprise_name", "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ–µ –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ"),
        "address": enterprise.get("address", "–Ω–µ —É–∫–∞–∑–∞–Ω")
        if enterprise
        else "–Ω–µ —É–∫–∞–∑–∞–Ω",
    }

    # 5. –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ (–µ—Å–ª–∏ –µ—Å—Ç—å)
    equipment_data = None
    nodes_data = None
    envelope_data = None

    # –û–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ
    equipment_json_path = AGGREGATED_DIR / f"{batch_id}_equipment.json"
    if not equipment_json_path.exists() and enterprise_id:
        uploads = database.list_uploads_for_enterprise(enterprise_id)
        for upload_item in uploads:
            upload_batch_id = upload_item.get("batch_id")
            if upload_batch_id:
                candidate_path = AGGREGATED_DIR / f"{upload_batch_id}_equipment.json"
                if candidate_path.exists():
                    equipment_json_path = candidate_path
                    break

    if equipment_json_path.exists():
        try:
            equipment_data = json.loads(equipment_json_path.read_text(encoding="utf-8"))
        except Exception as e:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è: {e}")

    # –£–∑–ª—ã —É—á–µ—Ç–∞
    nodes_json_path = AGGREGATED_DIR / f"{batch_id}_nodes.json"
    if not nodes_json_path.exists() and enterprise_id:
        uploads = database.list_uploads_for_enterprise(enterprise_id)
        for upload_item in uploads:
            upload_batch_id = upload_item.get("batch_id")
            if upload_batch_id:
                candidate_path = AGGREGATED_DIR / f"{upload_batch_id}_nodes.json"
                if candidate_path.exists():
                    nodes_json_path = candidate_path
                    break

    if nodes_json_path.exists():
        try:
            from tools.fill_energy_passport import load_nodes_from_json

            nodes_data = load_nodes_from_json(nodes_json_path)
        except Exception as e:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ —É–∑–ª–æ–≤ —É—á–µ—Ç–∞: {e}")

    # –†–∞—Å—á–µ—Ç —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º
    envelope_json_path = AGGREGATED_DIR / f"{batch_id}_envelope.json"
    if not envelope_json_path.exists() and enterprise_id:
        uploads = database.list_uploads_for_enterprise(enterprise_id)
        for upload_item in uploads:
            upload_batch_id = upload_item.get("batch_id")
            if upload_batch_id:
                candidate_path = AGGREGATED_DIR / f"{upload_batch_id}_envelope.json"
                if candidate_path.exists():
                    envelope_json_path = candidate_path
                    break

    if envelope_json_path.exists():
        try:
            envelope_data = json.loads(envelope_json_path.read_text(encoding="utf-8"))
        except Exception as e:
            logger.warning(
                f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ —Ä–∞—Å—á–µ—Ç–∞ —Ç–µ–ø–ª–æ–ø–æ—Ç–µ—Ä—å –ø–æ –∑–¥–∞–Ω–∏—è–º: {e}"
            )

    # 6. –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º Word –æ—Ç—á–µ—Ç –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö —Å –ø—Ä–∏–º–µ–Ω–µ–Ω–∏–µ–º AI-–∞–Ω–∞–ª–∏–∑–∞
    try:
        # –ì–µ–Ω–µ—Ä–∞—Ü–∏—è Word –æ—Ç—á–µ—Ç–∞ –≤—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è –Ω–∞ –æ—Å–Ω–æ–≤–µ:
        # - –ê–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤ (Excel, PDF –∏ —Ç.–¥.)
        # - –†–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ AI-–∞–Ω–∞–ª–∏–∑–∞ –¥–∞–Ω–Ω—ã—Ö (–µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–µ–Ω)
        # - –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö (–æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏–µ, —É–∑–ª—ã —É—á–µ—Ç–∞, –æ–≥—Ä–∞–∂–¥–∞—é—â–∏–µ –∫–æ–Ω—Å—Ç—Ä—É–∫—Ü–∏–∏)
        logger.info("üìÑ –ì–µ–Ω–µ—Ä–∞—Ü–∏—è Word –æ—Ç—á–µ—Ç–∞ –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö —Å AI-–æ–±–æ–≥–∞—â–µ–Ω–∏–µ–º...")

        generator = WordReportGenerator()

        # –°–æ–∑–¥–∞–µ–º –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª
        output_dir = AGGREGATED_DIR
        output_file = output_dir / f"{batch_id}_report.docx"

        # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –¥–æ–∫—É–º–µ–Ω—Ç –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö –∞–≥—Ä–µ–≥–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
        # (AI-–æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ —É–∂–µ –æ–±–æ–≥–∞—Ç–∏–ª–∏ aggregated, –µ—Å–ª–∏ AI –±—ã–ª –¥–æ—Å—Ç—É–ø–µ–Ω)
        generator.generate_report(
            enterprise_data=enterprise_data,
            aggregated_data=aggregated,  # –í–∫–ª—é—á–∞–µ—Ç AI-–∏–Ω—Å–∞–π—Ç—ã, –µ—Å–ª–∏ AI –ø—Ä–∏–º–µ–Ω—è–ª—Å—è
            equipment_data=equipment_data,
            nodes_data=nodes_data,
            envelope_data=envelope_data,
            output_path=output_file,
        )

        logger.info(f"‚úÖ Word –æ—Ç—á–µ—Ç —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω –∏–∑ –∏—Å—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö: {output_file}")
        if ai_processed_data:
            logger.info("üìä –û—Ç—á–µ—Ç –æ–±–æ–≥–∞—â–µ–Ω —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ AI-–∞–Ω–∞–ª–∏–∑–∞")

        return FileResponse(
            path=str(output_file),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=f"–û—Ç—á–µ—Ç_—ç–Ω–µ—Ä–≥–æ–∞—É–¥–∏—Ç–∞_{enterprise_data['name']}_{batch_id[:8]}.docx",
        )
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ Word –æ—Ç—á–µ—Ç–∞: {exc}")
        raise HTTPException(
            status_code=500, detail=f"–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ Word –æ—Ç—á–µ—Ç–∞: {exc}"
        ) from exc


@app.post("/ingest/files")
async def ingest_files(file: UploadFile = File(...)):
    """API endpoint –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ —Ñ–∞–π–ª–æ–≤ —Å –≤–∞–ª–∏–¥–∞—Ü–∏–µ–π —Å–æ–≥–ª–∞—Å–Ω–æ –¢–ó"""
    is_valid, error_msg = validate_file(file)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    size_valid, size_error = await validate_file_size(file)
    if not size_valid:
        raise HTTPException(status_code=400, detail=size_error)

    batch_id = str(uuid4())
    safe_filename = os.path.basename(file.filename)
    save_path = f"/tmp/ingest_{batch_id}_{safe_filename}"

    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file not allowed")
        with open(save_path, "wb") as output_file:
            output_file.write(content)
    except HTTPException:
        raise
    except PermissionError as exc:
        raise HTTPException(
            status_code=500, detail=f"Permission denied when saving file: {exc}"
        )
    except OSError as exc:  # pragma: no cover - IO errors
        raise HTTPException(status_code=500, detail=f"Failed to save file: {exc}")
    except Exception as exc:  # pragma: no cover - unexpected errors
        raise HTTPException(
            status_code=500, detail=f"Unexpected error saving file: {exc}"
        )

    validate_resp: Dict[str, Any] = {}
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                "http://validate:8002/validate/run", json={"batchId": batch_id}
            )
            resp.raise_for_status()
            validate_resp = resp.json()
    except httpx.TimeoutException:
        validate_resp = {"error": "validate service timeout", "batchId": batch_id}
    except httpx.HTTPStatusError as exc:
        validate_resp = {
            "error": f"validate service returned {exc.response.status_code}",
            "batchId": batch_id,
        }
    except httpx.RequestError as exc:
        validate_resp = {
            "error": f"validate service connection failed: {exc}",
            "batchId": batch_id,
        }
    except Exception as exc:  # pragma: no cover - unexpected
        validate_resp = {"error": f"validate call failed: {exc}", "batchId": batch_id}

    parsing_result = None
    try:
        logger.info("–ù–∞—á–∏–Ω–∞—é –ø–∞—Ä—Å–∏–Ω–≥ —Ñ–∞–π–ª–∞ —á–µ—Ä–µ–∑ API: %s", save_path)
        parsing_result = parse_file(save_path)
        status_value = "success" if parsing_result.get("parsed") else "partial"
        parsing_results_cache[batch_id] = {
            "batch_id": batch_id,
            "filename": safe_filename,
            "file_path": save_path,
            "parsing": parsing_result,
            "status": status_value,
        }
    except Exception as exc:  # pragma: no cover - parse failure path
        logger.exception("–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–∞—Ä—Å–∏–Ω–≥–µ —Ñ–∞–π–ª–∞ %s", save_path)
        parsing_results_cache[batch_id] = {
            "batch_id": batch_id,
            "filename": safe_filename,
            "file_path": save_path,
            "parsing": None,
            "error": str(exc),
            "status": "error",
        }

    parsing_status = (
        "success"
        if parsing_result and parsing_result.get("parsed")
        else ("error" if parsing_result is None else "pending")
    )

    return {
        "batchId": batch_id,
        "filename": safe_filename,
        "validate": validate_resp,
        "parsing_status": parsing_status,
    }


@app.post("/api/normative/upload")
async def upload_normative_document(
    file: UploadFile = File(...),
    title: Optional[str] = Form(None),
    document_type: Optional[str] = Form(None),
):
    """
    –ó–∞–≥—Ä—É–∑–∫–∞ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞ —Å AI-–∞–Ω–∞–ª–∏–∑–æ–º –∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ–º –≤ –ë–î.

    –ü–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç —Ñ–æ—Ä–º–∞—Ç—ã: PDF, Word, Excel.
    –î–æ–∫—É–º–µ–Ω—Ç –ø–∞—Ä—Å–∏—Ç—Å—è, –∞–Ω–∞–ª–∏–∑–∏—Ä—É–µ—Ç—Å—è AI –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è —Ñ–æ—Ä–º—É–ª –∏ –Ω–æ—Ä–º–∞—Ç–∏–≤–æ–≤,
    —Ä–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ—Ö—Ä–∞–Ω—è–µ—Ç—Å—è –≤ –ë–î –¥–ª—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –≤ —Ä–∞—Å—á–µ—Ç–∞—Ö —ç–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç–∞.

    Args:
        file: –§–∞–π–ª –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞
        title: –ù–∞–∑–≤–∞–Ω–∏–µ –¥–æ–∫—É–º–µ–Ω—Ç–∞ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ, –µ—Å–ª–∏ –Ω–µ —É–∫–∞–∑–∞–Ω–æ - –±–µ—Ä–µ—Ç—Å—è –∏–∑ –∏–º–µ–Ω–∏ —Ñ–∞–π–ª–∞)
        document_type: –¢–∏–ø –¥–æ–∫—É–º–µ–Ω—Ç–∞ (PKM690, GOST, SNiP –∏ —Ç.–¥., –æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)

    Returns:
        –†–µ–∑—É–ª—å—Ç–∞—Ç –∏–º–ø–æ—Ä—Ç–∞ —Å –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ–º –∏–∑–≤–ª–µ—á–µ–Ω–Ω—ã—Ö –ø—Ä–∞–≤–∏–ª
    """
    try:
        logger.info(
            f"–ü–æ–ª—É—á–µ–Ω –∑–∞–ø—Ä–æ—Å –Ω–∞ –∑–∞–≥—Ä—É–∑–∫—É –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {file.filename}"
        )

        # –í–∞–ª–∏–¥–∞—Ü–∏—è —Ñ–∞–π–ª–∞
        try:
            is_valid, error_msg = validate_file(file)
            if not is_valid:
                logger.warning(f"–í–∞–ª–∏–¥–∞—Ü–∏—è —Ñ–∞–π–ª–∞ –Ω–µ –ø—Ä–æ–π–¥–µ–Ω–∞: {error_msg}")
                raise HTTPException(status_code=400, detail=error_msg)
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –≤–∞–ª–∏–¥–∞—Ü–∏–∏ —Ñ–∞–π–ª–∞: {e}")
            raise HTTPException(status_code=400, detail=f"–û—à–∏–±–∫–∞ –≤–∞–ª–∏–¥–∞—Ü–∏–∏ —Ñ–∞–π–ª–∞: {e}")

        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞ –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –¥—É–±–ª–∏–∫–∞—Ç–∞
        file_size = None
        try:
            size_valid, size_error = await validate_file_size(file)
            if not size_valid:
                logger.warning(f"–ü—Ä–æ–≤–µ—Ä–∫–∞ —Ä–∞–∑–º–µ—Ä–∞ —Ñ–∞–π–ª–∞ –Ω–µ –ø—Ä–æ–π–¥–µ–Ω–∞: {size_error}")
                raise HTTPException(status_code=400, detail=size_error)
            # –ü–æ—Å–ª–µ –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ä–∞–∑–º–µ—Ä–∞ –ø–æ–ª—É—á–∞–µ–º —Ä–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞
            await file.seek(0)
            content_for_size = await file.read()
            file_size = len(content_for_size)
            await file.seek(0)  # –°–±—Ä–∞—Å—ã–≤–∞–µ–º –ø–æ–∑–∏—Ü–∏—é –¥–ª—è –¥–∞–ª—å–Ω–µ–π—à–µ–≥–æ —á—Ç–µ–Ω–∏—è
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ä–∞–∑–º–µ—Ä–∞ —Ñ–∞–π–ª–∞: {e}")
            raise HTTPException(
                status_code=400, detail=f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ä–∞–∑–º–µ—Ä–∞ —Ñ–∞–π–ª–∞: {e}"
            )

        # –ò–º–ø–æ—Ä—Ç –º–æ–¥—É–ª—è
        try:
            from domain.normative_importer import get_normative_importer

            importer = get_normative_importer()
            if not importer:
                logger.warning("–ò–º–ø–æ—Ä—Ç–µ—Ä –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω")
                raise HTTPException(
                    status_code=503,
                    detail="–ò–º–ø–æ—Ä—Ç–µ—Ä –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ AI.",
                )
        except ImportError as e:
            logger.exception("–ù–µ —É–¥–∞–ª–æ—Å—å –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å normative_importer")
            raise HTTPException(
                status_code=500,
                detail=f"–ú–æ–¥—É–ª—å –∏–º–ø–æ—Ä—Ç–∞ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω: {e}",
            )
        except HTTPException:
            raise
        except Exception as e:
            logger.exception(f"–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –∏–º–ø–æ—Ä—Ç–µ –º–æ–¥—É–ª—è: {e}")
            raise HTTPException(
                status_code=500, detail=f"–û—à–∏–±–∫–∞ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏–∏ –∏–º–ø–æ—Ä—Ç–µ—Ä–∞: {e}"
            )

        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∞–π–ª
        safe_filename = os.path.basename(file.filename)
        normative_dir = Path(INBOX_DIR) / "normative"
        normative_dir.mkdir(parents=True, exist_ok=True)

        file_path = normative_dir / safe_filename

        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –¥—É–±–ª–∏–∫–∞—Ç (–ø–æ –∏–º–µ–Ω–∏ –∏ —Ä–∞–∑–º–µ—Ä—É)
        if file_path.exists() and file_size is not None:
            existing_size = file_path.stat().st_size

            if existing_size == file_size:
                logger.warning(
                    f"–§–∞–π–ª {safe_filename} —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É"
                )
                # –ù–æ –≤—Å—ë —Ä–∞–≤–Ω–æ –¥–µ–ª–∞–µ–º –∏–º–ø–æ—Ä—Ç –∏–∑ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —Ñ–∞–π–ª–∞
                try:
                    result = importer.import_normative_document(
                        file_path=str(file_path),
                        title=title,
                        document_type=document_type,
                    )
                    return {
                        **result,
                        "message": "–§–∞–π–ª —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç, –≤—ã–ø–æ–ª–Ω–µ–Ω –∏–º–ø–æ—Ä—Ç –∏–∑ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ",
                        "file_path": str(file_path),
                    }
                except Exception as e:
                    logger.error(f"–û—à–∏–±–∫–∞ –∏–º–ø–æ—Ä—Ç–∞ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —Ñ–∞–π–ª–∞: {e}")
                    raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –∏–º–ø–æ—Ä—Ç–∞: {e}")

        try:
            with open(file_path, "wb") as output_file:
                content = await file.read()
                output_file.write(content)
        except Exception as exc:
            logger.exception("–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞")
            raise HTTPException(
                status_code=500, detail=f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å —Ñ–∞–π–ª: {exc}"
            )

        # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º —Å AI-–∞–Ω–∞–ª–∏–∑–æ–º
        try:
            logger.info(f"–ù–∞—á–∏–Ω–∞—é –∏–º–ø–æ—Ä—Ç –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {safe_filename}")
            result = importer.import_normative_document(
                file_path=str(file_path),
                title=title or safe_filename,
                document_type=document_type,
            )
            logger.info(
                f"‚úÖ –ò–º–ø–æ—Ä—Ç –∑–∞–≤–µ—Ä—à–µ–Ω: {result.get('rules_extracted', 0)} –ø—Ä–∞–≤–∏–ª –∏–∑–≤–ª–µ—á–µ–Ω–æ"
            )
            return {
                **result,
                "file_path": str(file_path),
                "filename": safe_filename,
            }
        except Exception as exc:
            logger.exception(f"–û—à–∏–±–∫–∞ –∏–º–ø–æ—Ä—Ç–∞ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {exc}")
            raise HTTPException(
                status_code=500, detail=f"–û—à–∏–±–∫–∞ –∏–º–ø–æ—Ä—Ç–∞ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {exc}"
            )

    except HTTPException:
        # –ü–µ—Ä–µ–±—Ä–∞—Å—ã–≤–∞–µ–º HTTPException –∫–∞–∫ –µ—Å—Ç—å
        raise
    except Exception as exc:
        # –õ–æ–≤–∏–º –ª—é–±—ã–µ –¥—Ä—É–≥–∏–µ –Ω–µ–æ–∂–∏–¥–∞–Ω–Ω—ã–µ –æ—à–∏–±–∫–∏
        logger.exception(
            f"–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {exc}"
        )
        raise HTTPException(
            status_code=500, detail=f"–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {exc}"
        )


@app.get("/api/normative/documents")
def list_normative_documents():
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
    try:
        documents = database.list_normative_documents()
        return {"documents": documents, "total": len(documents)}
    except Exception as exc:
        logger.exception("–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Å–ø–∏—Å–∫–∞ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Å–ø–∏—Å–∫–∞: {exc}")


@app.get("/api/normative/rules/{rule_type}")
def get_normative_rules_by_type(rule_type: str):
    """–ü–æ–ª—É—á–∏—Ç—å –ø—Ä–∞–≤–∏–ª–∞ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–Ω–æ–≥–æ —Ç–∏–ø–∞"""
    try:
        rules = database.get_normative_rules_by_type(rule_type)
        return {"rule_type": rule_type, "rules": rules, "total": len(rules)}
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –ø—Ä–∞–≤–∏–ª —Ç–∏–ø–∞ {rule_type}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –ø—Ä–∞–≤–∏–ª: {exc}")


@app.get("/api/normative/rules/for-field/{field_name}")
def get_normative_rules_for_field(
    field_name: str,
    sheet_name: Optional[str] = Query(None),
):
    """–ü–æ–ª—É—á–∏—Ç—å –ø—Ä–∞–≤–∏–ª–∞, —Å–≤—è–∑–∞–Ω–Ω—ã–µ —Å –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã–º –ø–æ–ª–µ–º —ç–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç–∞"""
    try:
        rules = database.get_normative_rules_for_field(field_name, sheet_name)
        return {
            "field_name": field_name,
            "sheet_name": sheet_name,
            "rules": rules,
            "total": len(rules),
        }
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –ø—Ä–∞–≤–∏–ª –¥–ª—è –ø–æ–ª—è {field_name}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –ø—Ä–∞–≤–∏–ª: {exc}")


@app.get("/api/normative/documents/{document_id}")
def get_normative_document_by_id(document_id: int):
    """–ü–æ–ª—É—á–∏—Ç—å –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã–π –¥–æ–∫—É–º–µ–Ω—Ç –ø–æ ID (–≤–∫–ª—é—á–∞—è –º–µ—Ç–∞–¥–∞–Ω–Ω—ã–µ)"""
    try:
        doc = database.get_normative_document(document_id)
        if not doc:
            raise HTTPException(status_code=404, detail=f"–î–æ–∫—É–º–µ–Ω—Ç —Å ID={document_id} –Ω–µ –Ω–∞–π–¥–µ–Ω")
        
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –¥–æ–∫—É–º–µ–Ω—Ç (–±–µ–∑ –ø–æ–ª–Ω–æ–≥–æ —Ç–µ–∫—Å—Ç–∞ –¥–ª—è —ç–∫–æ–Ω–æ–º–∏–∏ —Ç—Ä–∞—Ñ–∏–∫–∞)
        return {
            "document": {
                "id": doc.get("id"),
                "title": doc.get("title"),
                "document_type": doc.get("document_type"),
                "file_path": doc.get("file_path"),
                "file_size": doc.get("file_size"),
                "uploaded_at": doc.get("uploaded_at"),
                "ai_processed": doc.get("ai_processed"),
                "processing_status": doc.get("processing_status"),
            },
            "has_full_text": bool(doc.get("full_text")),
            "has_parsed_data": bool(doc.get("parsed_data_json")),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–∞ {document_id}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–∞: {exc}")


@app.get("/api/normative/documents/{document_id}/text")
def get_normative_document_text(document_id: int):
    """–ü–æ–ª—É—á–∏—Ç—å –ø–æ–ª–Ω—ã–π —Ç–µ–∫—Å—Ç –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞"""
    try:
        doc = database.get_normative_document(document_id)
        if not doc:
            raise HTTPException(status_code=404, detail=f"–î–æ–∫—É–º–µ–Ω—Ç —Å ID={document_id} –Ω–µ –Ω–∞–π–¥–µ–Ω")
        
        full_text = doc.get("full_text")
        if not full_text:
            return {
                "document_id": document_id,
                "title": doc.get("title"),
                "text": None,
                "message": "–ü–æ–ª–Ω—ã–π —Ç–µ–∫—Å—Ç –Ω–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω –¥–ª—è —ç—Ç–æ–≥–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞",
            }
        
        return {
            "document_id": document_id,
            "title": doc.get("title"),
            "text": full_text,
            "text_length": len(full_text),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Ç–µ–∫—Å—Ç–∞ –¥–æ–∫—É–º–µ–Ω—Ç–∞ {document_id}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Ç–µ–∫—Å—Ç–∞: {exc}")


@app.post("/api/normative/validate-field")
def validate_field_against_normative(
    field_name: str,
    actual_value: float,
    sheet_name: Optional[str] = Query(None),
    tolerance_percent: float = Query(10.0),
):
    """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏–µ —Ñ–∞–∫—Ç–∏—á–µ—Å–∫–æ–≥–æ –∑–Ω–∞—á–µ–Ω–∏—è –Ω–æ—Ä–º–∞—Ç–∏–≤—É"""
    try:
        from domain.normative_validator import validate_against_normative
        
        result = validate_against_normative(
            actual_value=actual_value,
            field_name=field_name,
            sheet_name=sheet_name,
            tolerance_percent=tolerance_percent,
        )
        
        return {
            "field_name": field_name,
            "sheet_name": sheet_name,
            "validation": result,
        }
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø–æ–ª—è {field_name}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏: {exc}")


@app.get("/api/normative/statistics")
def get_normative_statistics():
    """–ü–æ–ª—É—á–∏—Ç—å —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ –Ω–æ—Ä–º–∞—Ç–∏–≤–Ω—ã–º –¥–æ–∫—É–º–µ–Ω—Ç–∞–º"""
    try:
        from domain.normative_validator import get_normative_statistics
        
        stats = get_normative_statistics()
        return stats
    except Exception as exc:
        logger.exception("–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏: {exc}")


@app.get("/api/normative/critical-fields/{enterprise_id}")
def check_critical_fields_for_enterprise(enterprise_id: int):
    """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏–µ –ø–æ–ª—è –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è"""
    try:
        from domain.normative_validator import check_critical_fields
        
        result = check_critical_fields(enterprise_id=enterprise_id)
        return result
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –ø–æ–ª–µ–π –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏: {exc}")


@app.get("/api/normative/violations")
def get_normative_violations(
    enterprise_id: Optional[int] = Query(None),
    batch_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
):
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –Ω–∞—Ä—É—à–µ–Ω–∏–π –Ω–æ—Ä–º–∞—Ç–∏–≤–æ–≤"""
    try:
        violations = database.get_normative_violations(
            enterprise_id=enterprise_id,
            batch_id=batch_id,
            status=status,
            limit=limit,
        )
        return {
            "violations": violations,
            "total": len(violations),
            "filters": {
                "enterprise_id": enterprise_id,
                "batch_id": batch_id,
                "status": status,
            },
        }
    except Exception as exc:
        logger.exception("–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –Ω–∞—Ä—É—à–µ–Ω–∏–π –Ω–æ—Ä–º–∞—Ç–∏–≤–æ–≤")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –Ω–∞—Ä—É—à–µ–Ω–∏–π: {exc}")


@app.post("/api/normative/monitor-passport")
def monitor_passport_normatives(
    passport_path: str = Form(...),
    enterprise_id: Optional[int] = Form(None),
    batch_id: Optional[str] = Form(None),
):
    """–ú–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –ø–æ–ª–µ–π –∏–∑ —ç–Ω–µ—Ä–≥–æ–ø–∞—Å–ø–æ—Ä—Ç–∞"""
    try:
        from domain.normative_monitor import monitor_critical_fields_from_passport
        
        result = monitor_critical_fields_from_passport(
            passport_path=passport_path,
            enterprise_id=enterprise_id,
            batch_id=batch_id,
        )
        return result
    except Exception as exc:
        logger.exception(f"–û—à–∏–±–∫–∞ –º–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥–∞ –ø–∞—Å–ø–æ—Ä—Ç–∞ {passport_path}")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –º–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥–∞: {exc}")


@app.get("/api/normative/monitoring-summary")
def get_monitoring_summary(
    enterprise_id: Optional[int] = Query(None),
):
    """–ü–æ–ª—É—á–∏—Ç—å —Å–≤–æ–¥–∫—É –º–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥–∞ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –ø–æ–ª–µ–π"""
    try:
        from domain.normative_monitor import get_monitoring_summary
        
        summary = get_monitoring_summary(enterprise_id=enterprise_id)
        return summary
    except Exception as exc:
        logger.exception("–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Å–≤–æ–¥–∫–∏ –º–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥–∞")
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è —Å–≤–æ–¥–∫–∏: {exc}")


@app.post("/api/enterprises/{enterprise_id}/reaggregate")
def api_reaggregate_enterprise(enterprise_id: int):
    """
    –ü–µ—Ä–µ–∞–≥—Ä–µ–≥–∏—Ä—É–µ—Ç –≤—Å–µ –¥–∞–Ω–Ω—ã–µ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è.
    –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –¥–ª—è –ø—Ä–∏–º–µ–Ω–µ–Ω–∏—è –∏—Å–ø—Ä–∞–≤–ª–µ–Ω–∏–π –∫ —É–∂–µ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–º —Ñ–∞–π–ª–∞–º.
    """
    enterprise = database.get_enterprise_by_id(enterprise_id)
    if not enterprise:
        raise HTTPException(status_code=404, detail="–ü—Ä–µ–¥–ø—Ä–∏—è—Ç–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")

    try:
        from utils.reaggregate_all import reaggregate_enterprise

        logger.info(
            f"üîÑ –ó–∞–ø—Ä–æ—Å –Ω–∞ –ø–µ—Ä–µ–∞–≥—Ä–µ–≥–∞—Ü–∏—é –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id} ({enterprise['name']})"
        )
        stats = reaggregate_enterprise(enterprise_id)
        logger.info(f"‚úÖ –ü–µ—Ä–µ–∞–≥—Ä–µ–≥–∞—Ü–∏—è –∑–∞–≤–µ—Ä—à–µ–Ω–∞: {stats}")
        return {
            "enterprise_id": enterprise_id,
            "enterprise_name": enterprise["name"],
            "status": "success",
            "stats": stats,
        }
    except Exception as e:
        logger.exception(
            f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–µ—Ä–µ–∞–≥—Ä–µ–≥–∞—Ü–∏–∏ –¥–ª—è –ø—Ä–µ–¥–ø—Ä–∏—è—Ç–∏—è {enterprise_id}: {e}"
        )
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–µ—Ä–µ–∞–≥—Ä–µ–≥–∞—Ü–∏–∏: {e}")
