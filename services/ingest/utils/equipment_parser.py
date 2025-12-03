from __future__ import annotations

import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


SECTION_TITLE_PATTERN = re.compile(r"^\s*(\d+)\.?\s*(.*)$")


def is_equipment_file(filename: str, raw_json: Optional[Dict[str, Any]] = None) -> bool:
    """
    Проверяет, является ли файл файлом оборудования.

    Args:
        filename: Имя файла
        raw_json: Распарсенные данные файла (опционально, для анализа содержимого)

    Returns:
        True если файл содержит данные об оборудовании
    """
    # Используем единый классификатор для определения типа
    try:
        from utils.resource_classifier import ResourceClassifier

        detected_type = ResourceClassifier.classify(filename, raw_json)
        return detected_type == "equipment"
    except ImportError:
        # Fallback на проверку имени файла, если классификатор недоступен
        return "oborudovanie" in filename.lower()


def parse_equipment_workbook(
    workbook_path: Union[str, Path],
) -> Optional[Dict[str, Any]]:
    """
    Parse equipment workbook and return structured data grouped by production sections.

    Expected layout (per section):
        Row with title: "1. Цех ..." (column B)
        Header row: "№ | Оборудование | Кол-во (шт) | Мощн-едн (кВт*ч) | Общ-мощн (кВт*ч) | ЧП (VFD) | Категория энергоэффективности (опционально)"
        Data rows with consecutive numbers in column B
        Summary row with "Итого"
    
    Категория энергоэффективности извлекается из колонки 7 (если присутствует) и нормализуется к значениям A-G.
    """
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Equipment workbook not found: %s", path)
        return None

    workbook = load_workbook(path, data_only=True)
    sheets_payload: List[Dict[str, Any]] = []
    summary_total_items = 0
    summary_total_power = 0.0
    summary_vfd_true = 0

    for sheet in workbook.worksheets:
        sections = _parse_sheet(sheet)
        if not sections:
            continue

        sheet_info = {
            "sheet": sheet.title,
            "sections": sections,
        }
        sheets_payload.append(sheet_info)

        for section in sections:
            items = section.get("items", [])
            summary_total_items += len(items)
            for item in items:
                total_power = item.get("total_power_kw")
                if isinstance(total_power, (int, float)):
                    summary_total_power += float(total_power)
                if item.get("vfd") is True:
                    summary_vfd_true += 1

    if not sheets_payload:
        return None

    return {
        "source": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sheets": sheets_payload,
        "summary": {
            "total_sheets": len(sheets_payload),
            "total_sections": sum(len(sheet["sections"]) for sheet in sheets_payload),
            "total_items": summary_total_items,
            "total_power_kw": round(summary_total_power, 2),
            "vfd_with_frequency_drive": summary_vfd_true,
        },
    }


def write_equipment_json(
    batch_id: str,
    equipment_data: Dict[str, Any],
    destination_dir: Union[str, Path],
) -> Path:
    destination = Path(destination_dir)
    destination.mkdir(parents=True, exist_ok=True)
    target_file = destination / f"{batch_id}_equipment.json"
    target_file.write_text(
        json.dumps(equipment_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return target_file


def _parse_sheet(sheet: Worksheet) -> List[Dict[str, Any]]:
    sections: List[Dict[str, Any]] = []
    current_section: Optional[Dict[str, Any]] = None

    for row in sheet.iter_rows(values_only=True):
        cells: Sequence[Any] = list(row)
        if not any(value is not None and str(value).strip() for value in cells):
            # пустая строка: закрываем текущую секцию (если есть незаписанная)
            continue

        label_cell = _get_str(cells, 1)
        name_cell = _get_str(cells, 2)

        if _is_section_title(label_cell, name_cell):
            if current_section:
                sections.append(current_section)
            current_section = {
                "title": label_cell.strip(),
                "items": [],
            }
            title_match = SECTION_TITLE_PATTERN.match(label_cell)
            if title_match:
                current_section["section_number"] = int(title_match.group(1))
                cleaned_name = title_match.group(2).strip()
                if cleaned_name:
                    current_section["name"] = cleaned_name
            continue

        if not current_section:
            # пропускаем строки до первой секции
            continue

        if label_cell and label_cell.strip().lower() == "№":
            # это строка заголовков таблицы — пропускаем
            continue

        if label_cell and label_cell.strip().lower().startswith("итого"):
            totals = {
                "equipment_count": _parse_int(cells, 3),
                "total_power_kw": _parse_float(cells, 5),
                "vfd_count": _parse_int(cells, 6),
            }
            current_section["totals"] = totals
            continue

        if not name_cell:
            # Строка без названия оборудования — пропускаем
            continue

        item = _parse_item_row(cells)
        current_section.setdefault("items", []).append(item)

    if current_section:
        sections.append(current_section)

    return sections


def _parse_item_row(cells: Sequence[Any]) -> Dict[str, Any]:
    item: Dict[str, Any] = {
        "order": _parse_int(cells, 1),
        "name": _get_str(cells, 2),
        "quantity": _parse_float(cells, 3),
        "unit_power_kw": _parse_number_from_cell(cells, 4),
        "total_power_kw": _parse_number_from_cell(cells, 5),
    }

    raw_unit_power = cells[4] if len(cells) > 4 else None
    if isinstance(raw_unit_power, str):
        raw_value = raw_unit_power.strip()
        if raw_value:
            item["unit_power_raw"] = raw_value

    vfd_value = _get_str(cells, 6)
    if vfd_value:
        normalized = vfd_value.replace(" ", "")
        if "+" in normalized:
            item["vfd"] = True
        elif "-" in normalized:
            item["vfd"] = False

    # Категория энергоэффективности (опционально, может быть в колонке 7 или позже)
    energy_efficiency_category = _get_str(cells, 7)
    if energy_efficiency_category:
        # Нормализуем значение: убираем пробелы, приводим к верхнему регистру для категорий A-G
        normalized_category = energy_efficiency_category.strip().upper()
        if normalized_category in ["A", "B", "C", "D", "E", "F", "G"]:
            item["energy_efficiency_category"] = normalized_category
        elif normalized_category:
            # Сохраняем как есть, если не соответствует стандартным категориям
            item["energy_efficiency_category"] = normalized_category

    return item


def _is_section_title(label_cell: Optional[str], name_cell: Optional[str]) -> bool:
    if not label_cell:
        return False
    if name_cell:
        # В строке с названием секции колонка "Оборудование" пуста
        return False
    if SECTION_TITLE_PATTERN.match(label_cell):
        return True
    keywords = ("цех", "склад", "офис", "раздевалка", "душ", "производству")
    lowered = label_cell.lower()
    return any(keyword in lowered for keyword in keywords)


def _get_str(cells: Sequence[Any], index: int) -> Optional[str]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    return str(value)


def _parse_int(cells: Sequence[Any], index: int) -> Optional[int]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return int(round(value))
    try:
        stripped = str(value).strip()
        if not stripped:
            return None
        return int(float(stripped.replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _parse_float(cells: Sequence[Any], index: int) -> Optional[float]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        stripped = str(value).strip()
        if not stripped:
            return None
        return float(stripped.replace(",", "."))
    except (ValueError, TypeError):
        return None


def _parse_number_from_cell(cells: Sequence[Any], index: int) -> Optional[float]:
    if index >= len(cells):
        return None
    value = cells[index]
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        matches = re.findall(r"[-+]?\d*[.,]?\d+", stripped)
        if not matches:
            return None
        first = matches[0].replace(",", ".")
        try:
            return float(first)
        except ValueError:
            return None
    return None
