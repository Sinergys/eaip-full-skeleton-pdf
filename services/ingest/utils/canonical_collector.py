from __future__ import annotations
# cSpell:ignore sheetnames

import logging
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from ai.ai_excel_semantic_parser import (
    AnalyzeSheetInput,
    AnalyzeSheetResult,
    CanonicalSourceData,
    analyze_excel_sheet,
)
from settings.excel_semantic_settings import get_excel_semantic_mode

logger = logging.getLogger(__name__)


def analyze_workbook_sheets(workbook_path: str) -> list[dict]:
    """
    Run analyze_excel_sheet per sheet and return lightweight debug info.
    Does not persist anything; intended for canonical-debug endpoint.
    """
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Workbook not found for sheet analysis: %s", path)
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:  # pragma: no cover
        logger.exception("Failed to load workbook for sheet analysis %s: %s", path, exc)
        return []

    debug_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        header_rows = rows[:3]
        sample_rows = rows[3 : 3 + 16]

        input_payload = AnalyzeSheetInput(
            sheet_name=sheet_name,
            header_rows=header_rows,
            sample_rows=sample_rows,
            language_hints=["ru", "uz"],
            current_mapping_rules={},
        )
        result: AnalyzeSheetResult = analyze_excel_sheet(input_payload)
        debug_rows.append(
            {
                "sheet_name": sheet_name,
                "deterministic_conf": None,  # not separated in current implementation
                "ai_conf": None,  # not separated in current implementation
                "used_ai": result.used_ai,
                "confidence": result.confidence,
                "notes": result.notes,
            }
        )
    return debug_rows


def collect_canonical_from_workbook(
    workbook_path: str,
) -> Optional[CanonicalSourceData]:
    """
    Iterate workbook sheets and collect CanonicalSourceData (deterministic-only for now).
    Returns CanonicalSourceData or None if workbook can't be read.
    """
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Workbook not found for canonical collection: %s", path)
        return None
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl internals
        logger.exception(
            "Failed to load workbook for canonical collection %s: %s", path, exc
        )
        return None

    mode = get_excel_semantic_mode()
    logger.info("Collecting CanonicalSourceData (mode=%s) for %s", mode, path.name)

    global_canonical = CanonicalSourceData()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        header_rows = rows[:3]
        sample_rows = rows[3 : 3 + 16]  # up to ~quarter/year lines

        input_payload = AnalyzeSheetInput(
            sheet_name=sheet_name,
            header_rows=header_rows,
            sample_rows=sample_rows,
            language_hints=["ru", "uz"],
            current_mapping_rules={},
        )

        # Deterministic only in this iteration
        result: AnalyzeSheetResult = analyze_excel_sheet(input_payload)
        # Merge partials
        global_canonical.resources.extend(result.partial.resources or [])
        global_canonical.equipment.extend(result.partial.equipment or [])
        global_canonical.nodes.extend(result.partial.nodes or [])
        global_canonical.envelope.extend(result.partial.envelope or [])
        global_canonical.provenance.update(result.partial.provenance or {})

    return global_canonical
