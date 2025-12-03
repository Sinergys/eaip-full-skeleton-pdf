from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def is_envelope_file(filename: str, raw_json: Optional[Dict[str, Any]] = None) -> bool:
    """
    Проверяет, соответствует ли файл данным расчета теплопотерь по зданиям.

    Args:
        filename: Имя файла
        raw_json: Распарсенные данные файла (опционально, для анализа содержимого)

    Returns:
        True если файл содержит данные расчета теплопотерь по зданиям
    """
    # Используем единый классификатор для определения типа
    try:
        from utils.resource_classifier import ResourceClassifier

        detected_type = ResourceClassifier.classify(filename, raw_json)
        return detected_type == "envelope"
    except ImportError:
        # Fallback на проверку имени файла, если классификатор недоступен
        lowered = filename.lower()
        return (
            "ograjdayuschie" in lowered
            or "ograzhdayush" in lowered
            or "teploprovodnost" in lowered
            or "теплопроводность" in lowered
        )


def parse_building_envelope(
    workbook_path: Union[str, Path],
) -> Optional[Dict[str, Any]]:
    """
    Парсит Excel файл расчета теплопотерь по зданиям в структурированный JSON.

    Ожидаемые колонки (строка 2): Раздел | Конструкция | Материал | Толщина м | λ (Вт/м*°С) | R | Rтр | Площадь м² | ΔT | Q
    Строки разделов начинаются, когда колонка A (Раздел) имеет значение.
    Строки элементов имеют заполненную колонку B; итоговые строки содержат суммы в колонках H и J.
    """
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Файл расчета теплопотерь по зданиям не найден: %s", path)
        return None

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl internal issues
        logger.exception(
            "Не удалось загрузить файл расчета теплопотерь по зданиям %s: %s", path, exc
        )
        return None

    sheet = workbook.active
    sections = _parse_sheet(sheet)
    if not sections:
        logger.warning(
            "Не удалось распарсить разделы из файла расчета теплопотерь по зданиям %s",
            path,
        )
        return None

    return {
        "source": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sections": sections,
        "summary": {
            "total_sections": len(sections),
            "total_area_m2": round(
                sum(
                    section.get("totals", {}).get("area_m2", 0.0) or 0.0
                    for section in sections
                ),
                2,
            ),
            "total_heat_loss": round(
                sum(
                    section.get("totals", {}).get("heat_loss_q", 0.0) or 0.0
                    for section in sections
                ),
                2,
            ),
        },
    }


def write_envelope_json(
    batch_id: str,
    envelope_data: Dict[str, Any],
    destination_dir: Union[str, Path],
) -> Path:
    destination = Path(destination_dir)
    destination.mkdir(parents=True, exist_ok=True)
    target_file = destination / f"{batch_id}_envelope.json"
    target_file.write_text(
        json.dumps(envelope_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return target_file


def _parse_sheet(sheet: Worksheet) -> List[Dict[str, Any]]:
    sections: List[Dict[str, Any]] = []
    current_section: Optional[Dict[str, Any]] = None
    for row in sheet.iter_rows(values_only=True):
        cells: Sequence[Any] = list(row)
        if not _row_has_values(cells):
            continue

        first = _get_str(cells, 0)
        construction = _get_str(cells, 1)

        if first and first.lower() == "раздел":
            # header row
            continue

        if first:
            # new section
            if current_section:
                sections.append(current_section)
            current_section = {
                "section": first.strip(),
                "items": [],
            }
            # if row also contains construction data, fall through to item parsing

        if current_section is None:
            continue

        # Summary row: no construction but total area / heat loss present
        if not construction and (
            _parse_float(cells, 7) is not None or _parse_float(cells, 9) is not None
        ):
            totals = {
                "area_m2": _parse_float(cells, 7),
                "heat_loss_q": _parse_float(cells, 9),
            }
            current_section["totals"] = totals
            continue

        if not construction:
            continue

        item = {
            "construction": construction,
            "material": _get_str(cells, 2),
            "thickness_m": _parse_float(cells, 3),
            "lambda_w_mk": _parse_float(cells, 4),
            "resistance_r": _parse_float(cells, 5),
            "normative_r": _parse_float(cells, 6),
            "area_m2": _parse_float(cells, 7),
            "delta_t": _parse_float(cells, 8),
            "heat_loss_q": _parse_float(cells, 9),
        }
        current_section.setdefault("items", []).append(item)

    if current_section:
        sections.append(current_section)

    return sections


def _row_has_values(cells: Sequence[Any]) -> bool:
    return any(cell is not None and str(cell).strip() for cell in cells)


def _get_str(cells: Sequence[Any], index: int) -> Optional[str]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    value_str = str(value).strip()
    return value_str or None


def _parse_float(cells: Sequence[Any], index: int) -> Optional[float]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).strip().replace(",", ".")
        if not cleaned:
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None
