from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def is_nodes_file(filename: str, raw_json: Optional[Dict[str, Any]] = None) -> bool:
    """
    Проверяет, является ли файл файлом узлов учета.

    Args:
        filename: Имя файла
        raw_json: Распарсенные данные файла (опционально, для анализа содержимого)

    Returns:
        True если файл содержит данные узлов учета
    """
    # Используем единый классификатор для определения типа
    try:
        from utils.resource_classifier import ResourceClassifier

        detected_type = ResourceClassifier.classify(filename, raw_json)
        return detected_type == "nodes"
    except ImportError:
        # Fallback на проверку имени файла, если классификатор недоступен
        lowered = filename.lower()
        return "schetch" in lowered or "узлы учета" in lowered or "uzly" in lowered


def parse_nodes_workbook(workbook_path: Union[str, Path]) -> Optional[Dict[str, Any]]:
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Nodes workbook not found: %s", path)
        return None

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:  # pragma: no cover
        logger.exception("Failed to load workbook %s: %s", path, exc)
        return None

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        logger.warning("Nodes workbook %s is empty", path)
        return None

    tables: List[Dict[str, Any]] = []
    idx = 0
    total_nodes = 0
    total_active = 0.0
    total_reactive = 0.0

    while idx < len(rows):
        if not _has_data(rows[idx]):
            idx += 1
            continue

        header_rows = rows[idx : idx + 3]
        if len(header_rows) < 3:
            break
        idx += 3

        table_nodes: List[Dict[str, Any]] = []
        col_count = max(len(r) for r in header_rows)

        while idx < len(rows) and _has_data(rows[idx]):
            cells: Sequence[Any] = list(rows[idx])
            name = _str_or_none(cells, 0)
            if name:
                node: Dict[str, Any] = {
                    "name": name,
                    "resource": "Электрическая энергия",
                    "active_energy_p": _parse_float(cells, 1),
                    "reactive_energy_q": _parse_float(cells, 2),
                    "tt": _str_or_none(cells, 3),
                    "coefficient": _parse_float(cells, 5),
                    "seal_date": _str_or_none(cells, 6),
                    "supplier_seal_date": _str_or_none(cells, 7),
                    "note": _str_or_none(cells, 8),
                    "row_values": [
                        cells[i] if i < len(cells) else None for i in range(col_count)
                    ],
                }
                table_nodes.append(node)
                total_nodes += 1
                total_active += node.get("active_energy_p") or 0.0
                total_reactive += node.get("reactive_energy_q") or 0.0
            idx += 1

        tables.append(
            {
                "header": header_rows,
                "nodes": table_nodes,
                "columns": col_count,
            }
        )
        idx += 1  # пропускаем пустую строку между таблицами

    if not tables:
        return None

    flattened_nodes: List[Dict[str, Any]] = [
        node for table in tables for node in table["nodes"]
    ]

    return {
        "source": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "tables": tables,
        "nodes": flattened_nodes,
        "summary": {
            "total_nodes": total_nodes,
            "total_active_p": round(total_active, 2),
            "total_reactive_q": round(total_reactive, 2),
        },
    }


def write_nodes_json(
    batch_id: str, nodes_data: Dict[str, Any], destination_dir: Union[str, Path]
) -> Path:
    destination = Path(destination_dir)
    destination.mkdir(parents=True, exist_ok=True)
    target_file = destination / f"{batch_id}_nodes.json"
    target_file.write_text(
        json.dumps(nodes_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return target_file


def load_nodes_from_json(path: Union[str, Path]) -> Dict[str, Any]:
    json_path = Path(path)
    if not json_path.exists():
        raise FileNotFoundError(f"Nodes JSON not found: {json_path}")
    return json.loads(json_path.read_text(encoding="utf-8"))


def _has_data(cells: Sequence[Any]) -> bool:
    return any(cell is not None and str(cell).strip() for cell in cells)


def _str_or_none(cells: Sequence[Any], index: int) -> Optional[str]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    return str(value).strip() or None


def _parse_float(cells: Sequence[Any], index: int) -> Optional[float]:
    if index >= len(cells):
        return None
    value = cells[index]
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).strip().replace(" ", "").replace(",", ".")
        if not cleaned:
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None
