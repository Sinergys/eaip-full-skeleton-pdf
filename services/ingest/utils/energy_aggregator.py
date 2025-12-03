from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, Optional, Union

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Project root (3 levels up from this file: utils/ -> ingest/ -> services/ -> eaip_full_skeleton/)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "source_files" / "audit_sinergys"

MONTH_ALIASES = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}

TARGET_FILENAME_KEYWORDS: Iterable[str] = (
    "потребление энергоресурсов",
    "consumption",
    "energy_resources",
    "pererashod",
    "gaz",
    "газ",
    "расчет газа",
    "отопл",
    "неотпл",
    "voda",
    "otoplenie",
    "kotel",
)


def get_data_file_path(filename: str) -> Path:
    """
    Get path to data file, checking both project data directory and fallback locations.

    Args:
        filename: Name of the file (e.g., 'pererashod.xlsx')

    Returns:
        Path to the file

    Raises:
        FileNotFoundError: If file not found in any location
    """
    # Priority 1: Project data directory
    project_path = DATA_DIR / filename
    if project_path.exists():
        return project_path

    # Priority 2: Old external location (fallback for backwards compatibility)
    external_path = Path(r"C:\Users\DELL\Documents\AUDIT\Audit in Sinergys") / filename
    if external_path.exists():
        logger.warning(
            "Using external data file %s. Consider copying to %s",
            external_path,
            DATA_DIR,
        )
        return external_path

    raise FileNotFoundError(
        f"Data file '{filename}' not found in {DATA_DIR} or fallback locations"
    )


def month_to_quarter(month_number: int) -> int:
    return (month_number - 1) // 3 + 1


def _normalise_month_name(value: Optional[str]) -> Optional[str]:
    if not isinstance(value, str):
        return None
    return value.strip().lower()


def should_aggregate_file(filename: str) -> bool:
    name = filename.lower()
    return any(
        keyword in name for keyword in TARGET_FILENAME_KEYWORDS
    ) and filename.lower().endswith((".xlsx", ".xlsm", ".xls"))


def aggregate_single_resource_file(workbook_path: Union[str, Path]) -> Optional[Dict]:
    """
    Aggregate data from single-resource files (gaz.xlsx, voda.xlsx, kotel.xlsx).

    These files have simpler structure with data in a single sheet, organized by years and months.
    Returns None if the workbook does not exist or cannot be processed.
    """
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Workbook for aggregation not found: %s", path)
        return None

    filename = path.name.lower()

    # Определяем тип ресурса с использованием единого классификатора
    # Сначала пробуем определить по имени файла (быстрая проверка)
    # Если нужно, можно добавить анализ содержимого после загрузки workbook
    try:
        from utils.resource_classifier import ResourceClassifier

        resource_type = ResourceClassifier.classify(filename, None)

        # Проверяем, что определенный тип соответствует ожидаемым для этого метода
        if resource_type not in ("gas", "water", "heating", "boiler"):
            # Если классификатор вернул другой тип, пробуем старую логику как fallback
            if "gaz" in filename:
                resource_type = "gas"
            elif "voda" in filename or "water" in filename:
                resource_type = "water"
            elif "otoplenie" in filename or "heating" in filename:
                resource_type = "heating"
            elif "kotel" in filename or "boiler" in filename:
                resource_type = "boiler"
            else:
                logger.warning("Unknown resource type for file: %s", filename)
                return None
    except ImportError:
        # Fallback на старую логику, если классификатор недоступен
        if "gaz" in filename:
            resource_type = "gas"
        elif "voda" in filename or "water" in filename:
            resource_type = "water"
        elif "otoplenie" in filename or "heating" in filename:
            resource_type = "heating"
        elif "kotel" in filename or "boiler" in filename:
            resource_type = "boiler"
        else:
            logger.warning("Unknown resource type for file: %s", filename)
            return None

    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active  # Use first sheet
    except Exception as exc:
        logger.exception("Failed to load workbook %s: %s", path, exc)
        raise

    result: Dict[str, Dict[str, Dict]] = {resource_type: {}}

    if resource_type == "gas":
        # Parse gas files: может быть разная структура
        # Старая структура (gaz.xlsx):
        #   Row 2: None, 2022, None, 2023, None, 2024, None, ...
        #   Row 3: None, 'сум', 'м3', 'сум', 'м3', 'сум', 'м3', ...
        #   Row 4+: 'Январь', values...
        #
        # Новая структура (Расчет газа для отопл и неотпл.xlsx):
        #   Может быть другая структура - нужно определить динамически

        year_cols = {}

        # Пробуем найти годы в разных строках (1-5)
        header_rows_to_check = [1, 2, 3, 4, 5]
        header_row_idx = None
        header_row = None

        for row_idx in header_rows_to_check:
            try:
                row_data = list(
                    sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
                )[0]
                for col_idx, value in enumerate(row_data):
                    if isinstance(value, int) and value in (2022, 2023, 2024):
                        header_row_idx = row_idx
                        header_row = row_data
                        break
                if header_row:
                    break
            except Exception:
                continue

        if not header_row:
            logger.warning(f"Не найдены годы в файле газа: {path.name}")
            # Пробуем найти годы в любом месте первых 10 строк
            for row_idx in range(1, min(11, sheet.max_row + 1)):
                for col_idx in range(1, min(50, sheet.max_column + 1)):
                    value = sheet.cell(row_idx, col_idx).value
                    if isinstance(value, int) and value in (2022, 2023, 2024):
                        # Пробуем определить структуру: следующая колонка может быть volume_m3
                        year_cols[value] = {
                            "volume_m3": col_idx,  # Текущая колонка или следующая
                            "cost_sum": col_idx - 1
                            if col_idx > 1
                            else None,  # Предыдущая или None
                        }
                        # Проверяем следующую колонку
                        next_val = (
                            sheet.cell(row_idx, col_idx + 1).value
                            if col_idx < sheet.max_column
                            else None
                        )
                        if isinstance(next_val, (int, float)) and next_val > 0:
                            year_cols[value]["volume_m3"] = col_idx + 1
                            year_cols[value]["cost_sum"] = col_idx
        else:
            # Стандартная структура: годы в header_row
            for col_idx, value in enumerate(header_row):
                if isinstance(value, int) and value in (2022, 2023, 2024):
                    # Проверяем следующую строку для определения структуры колонок
                    next_row_idx = header_row_idx + 1
                    if next_row_idx <= sheet.max_row:
                        next_row = list(
                            sheet.iter_rows(
                                min_row=next_row_idx,
                                max_row=next_row_idx,
                                values_only=True,
                            )
                        )[0]
                        # Ищем 'м3', 'сум', 'cost', 'volume' в следующей строке
                        cost_col = None
                        volume_col = None

                        # Проверяем колонки вокруг года
                        for offset in [-1, 0, 1, 2]:
                            check_col = col_idx + offset
                            if 0 <= check_col < len(next_row):
                                cell_val = next_row[check_col]
                                if isinstance(cell_val, str):
                                    cell_lower = cell_val.lower()
                                    if any(
                                        kw in cell_lower
                                        for kw in ["сум", "cost", "стоимость"]
                                    ):
                                        cost_col = check_col
                                    elif any(
                                        kw in cell_lower
                                        for kw in ["м3", "volume", "объем"]
                                    ):
                                        volume_col = check_col

                        # Если не нашли по заголовкам, используем стандартную структуру
                        if volume_col is None:
                            volume_col = col_idx + 1
                        if cost_col is None:
                            cost_col = col_idx

                        year_cols[value] = {
                            "cost_sum": cost_col,
                            "volume_m3": volume_col,
                        }

        if not year_cols:
            logger.error(f"Не удалось определить структуру файла газа: {path.name}")
            return None

        logger.info(f"Найдены годы в файле газа: {list(year_cols.keys())}")

        # Ищем строки с месяцами (начинаем с строки после заголовков)
        start_data_row = (header_row_idx + 2) if header_row_idx else 4

        for row_idx in range(start_data_row, sheet.max_row + 1):
            row = list(
                sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
            )[0]

            # Ищем название месяца в первых колонках
            month_name = None
            for col_idx in range(min(5, len(row))):
                val = row[col_idx]
                if isinstance(val, str):
                    month_norm = _normalise_month_name(val)
                    if month_norm in MONTH_ALIASES:
                        month_name = val
                        break

            if not month_name:
                continue

            month_norm = _normalise_month_name(month_name)
            if month_norm not in MONTH_ALIASES:
                continue

            for year, col_indices in year_cols.items():
                quarter = month_to_quarter(MONTH_ALIASES[month_norm])
                quarter_key = f"{year}-Q{quarter}"
                quarter_entry = result[resource_type].setdefault(
                    quarter_key, {"year": year, "quarter": quarter, "months": []}
                )

                cost_sum = None
                volume_m3 = None

                # Безопасно извлекаем значения
                if col_indices.get("cost_sum") is not None:
                    cost_col = col_indices["cost_sum"]
                    if cost_col < len(row):
                        cost_sum = row[cost_col]

                if col_indices.get("volume_m3") is not None:
                    volume_col = col_indices["volume_m3"]
                    if volume_col < len(row):
                        volume_m3 = row[volume_col]

                # Если volume_m3 не найден, пробуем найти в соседних колонках
                if volume_m3 is None or (
                    isinstance(volume_m3, (int, float)) and volume_m3 == 0
                ):
                    # Ищем числовое значение в колонках вокруг года
                    for offset in [-1, 0, 1, 2]:
                        check_col = (
                            col_indices.get("volume_m3", col_indices.get("cost_sum", 0))
                            + offset
                        )
                        if 0 <= check_col < len(row):
                            val = row[check_col]
                            if isinstance(val, (int, float)) and val > 0:
                                volume_m3 = val
                                break

                quarter_entry["months"].append(
                    {
                        "month": month_name,
                        "values": {
                            "cost_sum": cost_sum,
                            "volume_m3": volume_m3,
                        },
                    }
                )

                logger.debug(
                    f"Добавлен месяц {month_name} для {year} Q{quarter}: "
                    f"volume_m3={volume_m3}, cost_sum={cost_sum}"
                )

    elif resource_type == "water":
        # Parse voda.xlsx: simple structure
        # Row 3 header: None, None, 'Месяцы', 'м3', 'Квартал'
        # Row 4+: None, None, 'Январь', 800, None

        current_year = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Look for year in row
            if row[2] and isinstance(row[2], int) and row[2] in (2022, 2023, 2024):
                current_year = row[2]
                continue

            if not current_year:
                continue

            month_name = row[2] if len(row) > 2 else None
            volume = row[3] if len(row) > 3 else None

            if not isinstance(month_name, str):
                continue

            month_norm = _normalise_month_name(month_name)
            if month_norm not in MONTH_ALIASES:
                continue

            quarter = month_to_quarter(MONTH_ALIASES[month_norm])
            quarter_key = f"{current_year}-Q{quarter}"
            quarter_entry = result[resource_type].setdefault(
                quarter_key, {"year": current_year, "quarter": quarter, "months": []}
            )

            quarter_entry["months"].append(
                {
                    "month": month_name,
                    "values": {
                        "volume_m3": volume,
                    },
                }
            )

    elif resource_type == "heating":
        # Parse otoplenie.xlsx: building data
        # This contains building dimensions and volumes, not time-series consumption
        # We'll store it differently - as building inventory
        buildings = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            building_name = row[0]
            if not building_name or not isinstance(building_name, str):
                continue

            if building_name.lower() in ("общее", "здания"):
                continue

            buildings.append(
                {
                    "name": building_name,
                    "width_m": row[1],
                    "length_m": row[2],
                    "height_m": row[3],
                    "area_m2": row[4],
                    "volume_m3": row[5],
                }
            )

        return {
            "source": str(path),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "resource_type": resource_type,
            "buildings": buildings,
        }

    elif resource_type == "boiler":
        # Parse kotel.xlsx: production norms and actual consumption
        # This has multiple sections - we'll parse production data
        production_data = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=3, max_row=7, values_only=True), 3
        ):
            item_num = row[0]
            if not item_num:
                continue

            production_data.append(
                {
                    "id": item_num,
                    "name": row[1],
                    "norm_tons": row[2],
                    "actual_2022": row[3],
                    "actual_2023": row[4],
                    "actual_2024": row[5],
                }
            )

        return {
            "source": str(path),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "resource_type": resource_type,
            "production": production_data,
        }

    _compute_quarter_totals_single_resource(result, resource_type)

    return {
        "source": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "resources": result,
    }


def aggregate_energy_data(workbook_path: Union[str, Path]) -> Optional[Dict]:
    """
    Aggregate monthly production and resource consumption data into quarterly totals.

    The expected structure mirrors the METIN workbook used in the manual pipeline.
    Returns None if the workbook does not exist or cannot be processed.
    """
    path = Path(workbook_path)
    if not path.exists():
        logger.warning("Workbook for aggregation not found: %s", path)
        return None

    # Check if this is a single-resource file
    filename = path.name.lower()
    if any(keyword in filename for keyword in ("gaz", "voda", "otoplenie", "kotel")):
        return aggregate_single_resource_file(path)

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl internal errors
        logger.exception("Failed to load workbook %s: %s", path, exc)
        raise

    result: Dict[str, Dict[str, Dict]] = {
        "electricity": {},
        "gas": {},
        "water": {},
        "production": {},
    }
    missing_sheets = []

    def aggregate_months(
        target: Dict, year: int, month_name: str, payload: Dict[str, Optional[float]]
    ) -> None:
        month_key = _normalise_month_name(month_name)
        if month_key not in MONTH_ALIASES:
            logger.debug(
                f"⚠️ [DIAG] Месяц '{month_name}' не распознан (month_key={month_key})"
            )
            return
        quarter = month_to_quarter(MONTH_ALIASES[month_key])
        quarter_key = f"{year}-Q{quarter}"
        quarter_entry = target.setdefault(
            quarter_key, {"year": year, "quarter": quarter, "months": []}
        )
        quarter_entry["months"].append({"month": month_name, "values": payload})
        logger.debug(
            f"📝 [DIAG] Добавлен месяц {month_name} ({year}) в квартал {quarter_key}: "
            f"поля={list(payload.keys())}, значения={[(k, v) for k, v in payload.items() if v is not None]}"
        )

    # Production sheet
    if "Килограмм да" in workbook.sheetnames:
        sheet = workbook["Килограмм да"]
        current_year: Optional[int] = None
        for row in sheet.iter_rows(values_only=True):
            first_cell = row[0]
            if isinstance(first_cell, int) and first_cell in (2022, 2023, 2024):
                current_year = first_cell
                continue
            if current_year and isinstance(first_cell, str):
                aggregate_months(
                    result["production"],
                    current_year,
                    first_cell,
                    {
                        "Труба хвс": row[1],
                        "Канал труба": row[2],
                        "Канал фитинг": row[3],
                        "Фит хвс и гвс": row[4],
                        "Теплый пол": row[5],
                        "Жами": row[6],
                    },
                )
    else:
        missing_sheets.append("Килограмм да")

    # Ищем лист с электроэнергией (проверяем альтернативные названия)
    electricity_sheet_name = None
    electricity_sheet_names = [
        "ЭЛЕКТР",
        "электр ",
        "электр",
        "Электроэнергия",
        "Электричество",
        "ТП",
        "Electricity",
        "ЭЛЕКТРИЧЕСТВО",
        "Sheet1",
    ]

    # ПРИМЕЧАНИЕ: pererashod.xlsx содержит только категории использования (by_usage),
    # но НЕ содержит месячных данных по электроэнергии. Sheet1 в pererashod.xlsx содержит описание ПКМ 690.
    # Месячные данные по электроэнергии должны быть в другом файле (например, "electro act react.xlsx").

    # Ищем по названию листа (точное совпадение)
    for sheet_name_var in electricity_sheet_names:
        if sheet_name_var in workbook.sheetnames:
            electricity_sheet_name = sheet_name_var
            logger.info(
                f"✅ [DIAG] Найден лист электроэнергии по точному совпадению: '{sheet_name_var}'"
            )
            break

    # Если не нашли по точному совпадению, ищем по частичному (с учетом пробелов в конце)
    if not electricity_sheet_name:
        for sheet_name in workbook.sheetnames:
            sheet_name_clean = sheet_name.strip()
            sheet_name_lower = sheet_name_clean.lower()
            # Расширенный поиск: включаем листы из файлов "Реализация" (Реал, Баланс)
            if any(
                keyword in sheet_name_lower
                for keyword in [
                    "электр", "electric", "электроэнергия",
                    "реал", "баланс", "реализация"  # Для файлов "Реализация"
                ]
            ):
                electricity_sheet_name = sheet_name
                logger.info(
                    f"✅ [DIAG] Найден лист электроэнергии по частичному совпадению: '{sheet_name}'"
                )
                break

    if electricity_sheet_name:
        sheet = workbook[electricity_sheet_name]
        current_year = None

        # Пытаемся определить структуру листа по заголовкам (первые 3 строки)
        header_rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i < 3:
                header_rows.append(list(row))
            else:
                break

        # Ищем колонки с данными (анализируем заголовки)
        # Структура может быть разной:
        # 1. Стандартная: Месяц | Сумма | кВт·ч | кВАр·ч | ...
        # 2. electro act react.xlsx: Год в строке 2, Месяц | Активная Квт/ч | Реактивная квар/ч | Сумма Квт/ч | ...
        col_cost = None  # Будет определена
        col_active_kwh = None  # Будет определена
        col_reactive_kvarh = None  # Будет определена
        col_active_power = None  # Колонка с расходом активной мощности (кВт)

        # Определяем год из первой строки (может быть в строке 2, колонка A или B)
        year_from_header = None
        for header_row in header_rows:
            for col_idx, cell_value in enumerate(header_row):
                if isinstance(cell_value, int) and cell_value in (2022, 2023, 2024):
                    year_from_header = cell_value
                    logger.info(
                        f"📅 [DIAG] Найден год {year_from_header} в заголовке, колонка {col_idx}"
                    )
                    break
            if year_from_header:
                break

        # Анализируем заголовки для определения колонок
        for header_row in header_rows:
            for col_idx, cell_value in enumerate(header_row):
                if cell_value and isinstance(cell_value, str):
                    cell_lower = str(cell_value).lower()
                    # Ищем колонку с активной энергией (кВт·ч)
                    if (
                        any(
                            keyword in cell_lower
                            for keyword in [
                                "активная квт/ч",
                                "активная квтч",
                                "активная квт·ч",
                                "active kwh",
                                "квт·ч",
                                "квтч",
                                "kwh",
                                "активная энергия",
                            ]
                        )
                        and col_active_kwh is None
                    ):
                        col_active_kwh = col_idx
                        logger.info(
                            f"✅ [DIAG] Найдена колонка active_kwh: {col_idx} ('{cell_value}')"
                        )
                    # Ищем колонку с реактивной энергией (кВАр·ч)
                    elif (
                        any(
                            keyword in cell_lower
                            for keyword in [
                                "реактивная квар/ч",
                                "реактивная кварч",
                                "реактивная квар·ч",
                                "reactive kvarh",
                                "квар·ч",
                                "кварч",
                                "kvarh",
                                "реактивная",
                            ]
                        )
                        and col_reactive_kvarh is None
                    ):
                        col_reactive_kvarh = col_idx
                        logger.info(
                            f"✅ [DIAG] Найдена колонка reactive_kvarh: {col_idx} ('{cell_value}')"
                        )
                    # Ищем колонку со стоимостью
                    elif (
                        any(
                            keyword in cell_lower
                            for keyword in [
                                "сумма квт/ч",
                                "сумма",
                                "стоимость",
                                "cost",
                                "цена",
                                "price",
                                "итого сум",
                            ]
                        )
                        and col_cost is None
                    ):
                        col_cost = col_idx
                        logger.info(
                            f"✅ [DIAG] Найдена колонка cost_sum: {col_idx} ('{cell_value}')"
                        )
                    # Ищем колонку с расходом активной мощности (в кВт, а не кВт·ч)
                    elif (
                        any(
                            keyword in cell_lower
                            for keyword in [
                                "расход активной мощности",
                                "мощность квт",
                                "active power",
                                "power kw",
                                "расход мощности",
                                "активная мощность",
                                "мощность активная",
                            ]
                        )
                        and "квт·ч" not in cell_lower
                        and "квтч" not in cell_lower
                    ):
                        col_active_power = col_idx
                        logger.debug(
                            f"Найдена колонка с расходом активной мощности: колонка {col_idx} ('{cell_value}')"
                        )

        # Устанавливаем значения по умолчанию, если не найдены
        if col_active_kwh is None:
            col_active_kwh = 1  # По умолчанию колонка B
        if col_reactive_kvarh is None:
            col_reactive_kvarh = 2  # По умолчанию колонка C
        if col_cost is None:
            col_cost = 4  # По умолчанию колонка E (Сумма Квт/ч)

        logger.info(
            f"📊 [DIAG] Определены колонки: active_kwh={col_active_kwh}, reactive_kvarh={col_reactive_kvarh}, "
            f"cost={col_cost}, year_from_header={year_from_header}"
        )

        # Парсим данные
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            first_cell = row[0] if len(row) > 0 else None

            # Определяем год: из заголовка или из первой колонки строки
            if isinstance(first_cell, int) and first_cell in (2022, 2023, 2024):
                current_year = first_cell
                logger.debug(
                    f"📅 [DIAG] Найден год {current_year} в строке {row_idx}, колонка A"
                )
                continue

            # Если год был определен из заголовка, используем его
            if year_from_header and current_year is None:
                current_year = year_from_header

            # Парсим строки с месяцами
            if current_year and first_cell and isinstance(first_cell, str):
                month_name = first_cell.strip()
                # Проверяем, что это месяц
                month_norm = _normalise_month_name(month_name)
                if month_norm not in MONTH_ALIASES:
                    continue

                # Извлекаем данные из определенных колонок
                values_dict = {
                    "cost_sum": row[col_cost] if len(row) > col_cost else None,
                    "active_kwh": row[col_active_kwh]
                    if len(row) > col_active_kwh
                    else None,
                    "reactive_kvarh": row[col_reactive_kvarh]
                    if len(row) > col_reactive_kvarh
                    else None,
                }

                # Если есть колонка с расходом активной мощности, добавляем её
                if col_active_power is not None and len(row) > col_active_power:
                    power_value = row[col_active_power]
                    if power_value is not None:
                        values_dict["active_power"] = power_value
                        logger.debug(
                            f"⚡ [DIAG] Извлечен расход активной мощности для {month_name}: {power_value} кВт"
                        )

                logger.debug(
                    f"📝 [DIAG] Парсинг {month_name} ({current_year}): "
                    f"active_kwh={values_dict.get('active_kwh')}, reactive_kvarh={values_dict.get('reactive_kvarh')}, "
                    f"cost_sum={values_dict.get('cost_sum')}"
                )

                aggregate_months(
                    result["electricity"],
                    current_year,
                    month_name,
                    values_dict,
                )
    else:
        missing_sheets.append("ЭЛЕКТР (или альтернативные названия)")

    if "ГАЗ" in workbook.sheetnames:
        sheet = workbook["ГАЗ"]
        current_year = None
        for row in sheet.iter_rows(values_only=True):
            first_cell = row[0]
            if isinstance(first_cell, int) and first_cell in (2022, 2023, 2024):
                current_year = first_cell
                continue
            if current_year and isinstance(first_cell, str):
                aggregate_months(
                    result["gas"],
                    current_year,
                    first_cell,
                    {
                        "cost_sum": row[1],
                        "volume_m3": row[2],
                        "extra": row[3],
                    },
                )
    else:
        missing_sheets.append("ГАЗ")

    if "СУВ" in workbook.sheetnames:
        sheet = workbook["СУВ"]
        current_year = None
        for row in sheet.iter_rows(values_only=True):
            first_cell = row[0]
            if isinstance(first_cell, int) and first_cell in (2022, 2023, 2024):
                current_year = first_cell
                continue
            if current_year and isinstance(first_cell, str):
                aggregate_months(
                    result["water"],
                    current_year,
                    first_cell,
                    {
                        "volume_m3": row[1],
                        "cost_sum": row[2],
                    },
                )
    else:
        missing_sheets.append("СУВ")

    logger.info(
        f"🔍 [DIAG] Перед расчетом квартальных итогов. Ресурсы: {list(result.keys())}"
    )
    for resource_type, resource_data in result.items():
        logger.info(
            f"📊 [DIAG] Ресурс {resource_type}: {len(resource_data)} кварталов, "
            f"кварталы={list(resource_data.keys())}"
        )
        for quarter_key, quarter_data in resource_data.items():
            months_count = len(quarter_data.get("months", []))
            logger.info(
                f"  └─ Квартал {quarter_key}: {months_count} месяцев, "
                f"quarter_totals={'есть' if 'quarter_totals' in quarter_data else 'отсутствует'}"
            )

    _compute_quarter_totals(result)

    logger.info("✅ [DIAG] После расчета квартальных итогов")
    for resource_type, resource_data in result.items():
        for quarter_key, quarter_data in resource_data.items():
            quarter_totals = quarter_data.get("quarter_totals", {})
            logger.info(
                f"📊 [DIAG] Ресурс {resource_type}, квартал {quarter_key}: "
                f"quarter_totals={list(quarter_totals.keys()) if quarter_totals else 'пусто'}"
            )

    return {
        "source": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "resources": result,
        "missing_sheets": missing_sheets,
    }


def _compute_quarter_totals(result: Dict[str, Dict[str, Dict]]) -> None:
    """
    Вычисляет квартальные итоги для всех ресурсов ИЗ МЕСЯЧНЫХ ДАННЫХ.

    КРИТИЧЕСКИ ВАЖНО: Квартальные данные ВСЕГДА рассчитываются из месячных,
    а не ожидаются готовыми в файле. Это гарантирует корректность данных.
    """
    logger.info(
        f"🔍 [DIAG] Начало расчета квартальных итогов. Ресурсы: {list(result.keys())}"
    )

    # Обрабатываем все ресурсы, а не только основные
    resource_keys = result.keys()

    for key in resource_keys:
        if key not in result:
            logger.warning(f"⚠️ [DIAG] Ресурс {key} отсутствует в result")
            continue

        resource_data = result[key]
        logger.info(
            f"📊 [DIAG] Обработка ресурса '{key}': {len(resource_data)} кварталов"
        )

        for quarter_key, quarter_data in result[key].items():
            months = quarter_data.get("months", [])
            logger.debug(
                f"🔍 [DIAG] Квартал {quarter_key} для {key}: "
                f"{len(months)} месяцев, "
                f"year={quarter_data.get('year')}, quarter={quarter_data.get('quarter')}"
            )

            if not months:
                logger.warning(
                    f"⚠️ [DIAG] Квартал {quarter_key} для ресурса {key} не содержит месячных данных. "
                    f"Квартальные итоги не могут быть рассчитаны. "
                    f"Структура quarter_data: {list(quarter_data.keys())}"
                )
                # Создаем пустые quarter_totals, если месячных данных нет
                if "quarter_totals" not in quarter_data:
                    quarter_data["quarter_totals"] = {}
                continue

            # ДИАГНОСТИКА: Показываем структуру месячных данных
            logger.debug(
                f"📋 [DIAG] Месячные данные для {key} {quarter_key}: "
                f"месяцев={len(months)}, "
                f"пример первого месяца: {months[0] if months else 'нет'}"
            )

            if key == "production":
                # Для производства используем фиксированный набор полей
                totals: Dict[str, float] = {
                    "Труба хвс": 0.0,
                    "Канал труба": 0.0,
                    "Канал фитинг": 0.0,
                    "Фит хвс и гвс": 0.0,
                    "Теплый пол": 0.0,
                    "Жами": 0.0,
                }
                for month in months:
                    values = month.get("values", {})
                    for product, value in values.items():
                        if isinstance(value, (int, float)) and value is not None:
                            totals[product] = totals.get(product, 0.0) + float(value)
                quarter_data["quarter_totals"] = totals
                logger.debug(
                    f"Рассчитаны квартальные итоги для {key} {quarter_key}: "
                    f"{len(months)} месяцев, итого: {totals}"
                )
            else:
                # Для всех остальных ресурсов суммируем ВСЕ числовые поля из месячных данных
                totals: Dict[str, float] = {}

                for month_idx, month in enumerate(months):
                    values = month.get("values", {})
                    month_name = month.get("month", f"месяц_{month_idx}")

                    if not values:
                        logger.debug(f"⚠️ [DIAG] Месяц {month_name} не содержит values")
                        continue

                    logger.debug(
                        f"📊 [DIAG] Месяц {month_name}: поля={list(values.keys())}, "
                        f"значения={[(k, v) for k, v in values.items() if isinstance(v, (int, float))]}"
                    )

                    for field, value in values.items():
                        # Пропускаем None и нечисловые значения
                        if value is None:
                            continue
                        if not isinstance(value, (int, float)):
                            continue

                        # Суммируем значения по полям
                        if field not in totals:
                            totals[field] = 0.0
                        totals[field] += float(value)
                        logger.debug(
                            f"➕ [DIAG] Суммирование {field}: {value} -> итого={totals[field]}"
                        )

                # Гарантируем наличие quarter_totals
                if "quarter_totals" not in quarter_data:
                    quarter_data["quarter_totals"] = {}

                # ВАЖНО: Перезаписываем quarter_totals рассчитанными значениями из месячных
                # Это гарантирует, что квартальные данные всегда рассчитываются из месячных
                quarter_data["quarter_totals"] = totals

                logger.info(
                    f"✅ [DIAG] Рассчитаны квартальные итоги для {key} {quarter_key} из {len(months)} месяцев: "
                    f"поля={list(totals.keys())}, значения={[(k, v) for k, v in totals.items()]}"
                )

                # Специальная обработка для электроэнергии: расчет active_kwh из месячных данных мощности
                # Проверяем отсутствие, None или 0 (только если не было рассчитано из месячных)
                active_kwh_value = totals.get("active_kwh")
                if key == "electricity" and (
                    active_kwh_value is None or active_kwh_value == 0
                ):
                    # Ищем данные о расходе активной мощности по месяцам
                    # Возможные названия полей: active_power, power_kw, active_power_kw, расход_активной_мощности
                    power_field_names = [
                        "active_power",
                        "power_kw",
                        "active_power_kw",
                        "расход_активной_мощности",
                        "мощность_квт",
                        "мощность",
                        "active_power_monthly",
                        "power_monthly",
                    ]

                    monthly_power_sum = 0.0
                    power_found = False

                    for month in months:
                        values = month.get("values", {})
                        for power_field in power_field_names:
                            if power_field in values and values[power_field]:
                                try:
                                    power_value = float(values[power_field])
                                    if power_value > 0:
                                        monthly_power_sum += power_value
                                        power_found = True
                                        logger.debug(
                                            f"Найдено значение мощности '{power_field}': {power_value} кВт "
                                            f"для месяца {month.get('month')}"
                                        )
                                        break  # Используем первое найденное поле
                                except (ValueError, TypeError):
                                    continue

                    # Если нашли данные о мощности, вычисляем active_kwh
                    if power_found and monthly_power_sum > 0:
                        try:
                            # Импортируем функции расчета (относительный импорт)
                            import sys
                            from pathlib import Path

                            domain_path = (
                                Path(__file__).resolve().parent.parent / "domain"
                            )
                            if str(domain_path) not in sys.path:
                                sys.path.insert(0, str(domain_path))

                            from energy_passport_calculations import (
                                calculate_quarter_consumption_from_monthly_power,
                            )
                            from energy_units import HOURS_PER_MONTH

                            # Вычисляем среднюю мощность за месяц
                            avg_monthly_power = (
                                monthly_power_sum / len(months) if months else 0.0
                            )

                            # Вычисляем квартальное потребление
                            calculated_active_kwh = (
                                calculate_quarter_consumption_from_monthly_power(
                                    avg_monthly_power, HOURS_PER_MONTH
                                )
                            )

                            if calculated_active_kwh > 0:
                                totals["active_kwh"] = calculated_active_kwh
                                logger.info(
                                    f"Вычислен active_kwh из месячных данных мощности: "
                                    f"{calculated_active_kwh:.2f} кВт·ч "
                                    f"(средняя мощность: {avg_monthly_power:.2f} кВт, месяцев: {len(months)})"
                                )
                        except ImportError as e:
                            logger.warning(
                                f"Не удалось импортировать функции расчета: {e}"
                            )
                        except Exception as e:
                            logger.warning(
                                f"Ошибка при расчете active_kwh из мощности: {e}"
                            )

                # Специальная обработка для электроэнергии: расчет reactive_kvarh из месячных данных реактивной мощности
                # Проверяем отсутствие, None или 0
                reactive_kvarh_value = totals.get("reactive_kvarh")
                if key == "electricity" and (
                    reactive_kvarh_value is None or reactive_kvarh_value == 0
                ):
                    # Ищем данные о расходе реактивной мощности по месяцам
                    reactive_power_field_names = [
                        "reactive_power",
                        "reactive_power_kvar",
                        "reactive_power_kvarh",
                        "расход_реактивной_мощности",
                        "реактивная_мощность_квар",
                        "реактивная_мощность",
                        "reactive_power_monthly",
                        "reactive_monthly",
                    ]

                    monthly_reactive_power_sum = 0.0
                    reactive_power_found = False

                    for month in months:
                        values = month.get("values", {})
                        for reactive_power_field in reactive_power_field_names:
                            if (
                                reactive_power_field in values
                                and values[reactive_power_field]
                            ):
                                try:
                                    reactive_power_value = float(
                                        values[reactive_power_field]
                                    )
                                    if reactive_power_value > 0:
                                        monthly_reactive_power_sum += (
                                            reactive_power_value
                                        )
                                        reactive_power_found = True
                                        logger.debug(
                                            f"Найдено значение реактивной мощности '{reactive_power_field}': {reactive_power_value} кВАр "
                                            f"для месяца {month.get('month')}"
                                        )
                                        break  # Используем первое найденное поле
                                except (ValueError, TypeError):
                                    continue

                    # Если нашли данные о реактивной мощности, вычисляем reactive_kvarh
                    if reactive_power_found and monthly_reactive_power_sum > 0:
                        try:
                            # Импортируем функции расчета (относительный импорт)
                            import sys
                            from pathlib import Path

                            domain_path = (
                                Path(__file__).resolve().parent.parent / "domain"
                            )
                            if str(domain_path) not in sys.path:
                                sys.path.insert(0, str(domain_path))

                            from energy_passport_calculations import (
                                calculate_quarter_reactive_consumption_from_monthly_power,
                            )
                            from energy_units import HOURS_PER_MONTH

                            # Вычисляем среднюю реактивную мощность за месяц
                            avg_monthly_reactive_power = (
                                monthly_reactive_power_sum / len(months)
                                if months
                                else 0.0
                            )

                            # Вычисляем квартальное реактивное потребление
                            calculated_reactive_kvarh = calculate_quarter_reactive_consumption_from_monthly_power(
                                avg_monthly_reactive_power, HOURS_PER_MONTH
                            )

                            if calculated_reactive_kvarh > 0:
                                totals["reactive_kvarh"] = calculated_reactive_kvarh
                                logger.info(
                                    f"Вычислен reactive_kvarh из месячных данных реактивной мощности: "
                                    f"{calculated_reactive_kvarh:.2f} кВАр·ч "
                                    f"(средняя реактивная мощность: {avg_monthly_reactive_power:.2f} кВАр, месяцев: {len(months)})"
                                )
                        except ImportError as e:
                            logger.warning(
                                f"Не удалось импортировать функции расчета реактивной энергии: {e}"
                            )
                        except Exception as e:
                            logger.warning(
                                f"Ошибка при расчете reactive_kvarh из реактивной мощности: {e}"
                            )


def _compute_quarter_totals_single_resource(
    result: Dict[str, Dict[str, Dict]], resource_type: str
) -> None:
    """Compute quarterly totals for single-resource files (gas, water, etc.)"""
    for quarter_data in result[resource_type].values():
        months = quarter_data.get("months", [])
        totals: Dict[str, float] = {}
        for month in months:
            for field, value in month.get("values", {}).items():
                if isinstance(value, (int, float)):
                    totals[field] = totals.get(field, 0.0) + float(value)
        quarter_data["quarter_totals"] = totals


def aggregate_usage_categories(
    workbook_path: Union[str, Path], equipment_data: Optional[Dict] = None
) -> Optional[Dict]:
    """
    Parse usage categories (technological/own/production/household) from pererashod.xlsx.

    Структура файла:
    - 4 таблицы, разделенные 2 пустыми строками
    - Каждая таблица начинается с R1C1 (A1)
    - 4-я таблица содержит данные по категориям использования (строки 26-37)

    Структура 4-й таблицы:
    - Строка 26: заголовки - "Год", 2022, 2023, 2024, ИТОГО (столбцы B-F)
    - Строка 29: "Тех-потери кВтч" → technological (столбцы C-E для годов)
    - Строка 30: "Хоз-бытовые нужды кВтч" → household
    - Строка 31: "Производственные нужды" → production

    Args:
        workbook_path: Путь к файлу pererashod.xlsx
        equipment_data: Данные из oborudovanie.xlsx (основной источник для сверки)

    Returns:
        Dict с годовыми данными по категориям, или None если не применимо.

    Примечание:
        - oborudovanie.xlsx - основной источник данных по категориям использования
        - pererashod.xlsx - вспомогательный источник для сверки и корректировки
        - Используется ИИ для анализа структуры таблицы
    """
    path = Path(workbook_path)
    if not path.exists() or "pererashod" not in path.name.lower():
        return None

    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active

        # Находим 4-ю таблицу (с данными по типу расхода)
        # Таблицы разделены 2 пустыми строками
        tables = []
        current_table_start = None
        empty_rows_count = 0

        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            cells = [cell.value for cell in row[:20]]
            has_data = any(c is not None and str(c).strip() for c in cells)

            if not has_data:
                empty_rows_count += 1
                if empty_rows_count >= 2 and current_table_start:
                    tables.append(
                        {"start_row": current_table_start, "end_row": row_idx - 2}
                    )
                    current_table_start = None
            else:
                empty_rows_count = 0
                if current_table_start is None:
                    current_table_start = row_idx

        if current_table_start:
            tables.append({"start_row": current_table_start, "end_row": sheet.max_row})

        if len(tables) < 4:
            logger.warning(
                f"В файле {path.name} найдено только {len(tables)} таблиц, ожидалось 4"
            )
            return None

        # Работаем с 4-й таблицей
        table_4 = tables[3]  # Индекс 3 для 4-й таблицы
        logger.info(f"4-я таблица: строки {table_4['start_row']}-{table_4['end_row']}")

        # Ищем строку с заголовками годов в 4-й таблице
        # Обычно это первая строка таблицы с "Год" и годами
        header_row = None
        years = []

        for row_idx in range(
            table_4["start_row"], min(table_4["start_row"] + 5, table_4["end_row"] + 1)
        ):
            row = list(
                sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
            )[0]
            row_text = " ".join([str(c) if c else "" for c in row[:10]]).lower()

            # Ищем строку с "год" и годами
            if "год" in row_text or any(str(y) in str(row) for y in [2022, 2023, 2024]):
                # Извлекаем годы из строки (столбцы C-E, индексы 2-4)
                years_candidates = [row[2], row[3], row[4], row[5]]  # Столбцы C-F
                years = [int(y) for y in years_candidates if y and str(y).isdigit()]
                if years:
                    header_row = row_idx
                    logger.info(
                        f"Найдена строка заголовков: {header_row}, годы: {years}"
                    )
                    break

        if not years or not header_row:
            logger.warning(
                f"Не найдены годы в 4-й таблице (строки {table_4['start_row']}-{table_4['end_row']})"
            )
            workbook.close()
            return None

        # Используем ИИ для анализа структуры таблицы
        try:
            from ai.ai_excel_semantic_parser import (
                analyze_excel_sheet,
                AnalyzeSheetInput,
            )

            # Подготавливаем данные для ИИ анализа
            header_rows = []
            sample_rows = []

            # Читаем заголовки (первые 3 строки таблицы)
            for row_idx in range(
                table_4["start_row"],
                min(table_4["start_row"] + 3, table_4["end_row"] + 1),
            ):
                row = list(
                    sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
                )[0]
                header_rows.append([str(c) if c is not None else "" for c in row[:10]])

            # Читаем образцы данных (следующие 10 строк)
            for row_idx in range(
                table_4["start_row"] + 3,
                min(table_4["start_row"] + 13, table_4["end_row"] + 1),
            ):
                row = list(
                    sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
                )[0]
                sample_rows.append([str(c) if c is not None else "" for c in row[:10]])

            # Анализ через ИИ
            ai_input = AnalyzeSheetInput(
                sheet_name=sheet.title,
                header_rows=header_rows,
                sample_rows=sample_rows,
                language_hints=["ru", "uz"],
                current_mapping_rules={
                    "usage_categories": {
                        "technological": ["тех", "технологич", "тех-потер"],
                        "household": ["хоз", "бытов", "хоз-быт"],
                        "production": ["производств", "произв"],
                        "own_needs": ["собствен", "с.н.", "нужды"],
                    }
                },
            )

            ai_result = analyze_excel_sheet(ai_input)
            logger.info(
                f"ИИ анализ таблицы: confidence={ai_result.confidence}, used_ai={ai_result.used_ai}"
            )

        except ImportError:
            logger.warning(
                "Модуль ИИ не доступен, используем детерминистический парсинг"
            )
            ai_result = None
        except Exception as ai_exc:
            logger.warning(
                f"Ошибка ИИ анализа: {ai_exc}, используем детерминистический парсинг"
            )
            ai_result = None

        # Детерминистический парсинг 4-й таблицы
        # Ищем строки с категориями использования
        categories_data = {}
        category_mapping = {}  # Будет заполнен при поиске

        # АЛЬТЕРНАТИВНЫЙ ПОДХОД: Если не нашли таблицу через ИИ, ищем по ключевым словам во всем листе
        # Это для файлов pererashod.xlsx с простой структурой (строки 10-15)
        if not years or not header_row:
            logger.info(
                "Пробуем альтернативный подход: поиск категорий по ключевым словам во всем листе"
            )
            # Ищем строки с категориями по всему листу
            for row_idx in range(
                1, min(sheet.max_row + 1, 50)
            ):  # Проверяем первые 50 строк
                row_label_cell = sheet.cell(row=row_idx, column=1).value  # Столбец A
                if not row_label_cell:
                    continue

                row_label = str(row_label_cell).lower()

                # Определяем категорию по тексту в столбце A
                category_key = None
                if any(kw in row_label for kw in ["тех", "технологич", "тех-потер"]):
                    category_key = "technological"
                elif any(kw in row_label for kw in ["хоз", "бытов", "хоз-быт"]):
                    category_key = "household"
                elif any(kw in row_label for kw in ["производств", "произв"]):
                    category_key = "production"
                elif any(kw in row_label for kw in ["собствен", "с.н.", "нужды"]):
                    category_key = "own_needs"

                if category_key:
                    # Читаем значения по годам из текущей строки
                    row_data = list(
                        sheet.iter_rows(
                            min_row=row_idx, max_row=row_idx, values_only=True
                        )
                    )[0]

                    # Пробуем найти годы в заголовках (строка 1)
                    header_row_data = list(
                        sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                    )[0]
                    year_cols = {}  # {year: col_index}
                    for col_idx, cell_value in enumerate(header_row_data):
                        if isinstance(cell_value, int) and cell_value in (
                            2022,
                            2023,
                            2024,
                        ):
                            year_cols[cell_value] = col_idx

                    if not year_cols:
                        # Если годы не найдены в заголовках, используем стандартную структуру pererashod.xlsx:
                        # A=название, B=норма, C=2022 факт, D=2022 %, E=2023 факт, F=2023 %, G=2024 факт, H=2024 %
                        # Для категорий использования данные в колонках C, E, G (индексы 2, 4, 6)
                        year_cols = {2022: 2, 2023: 4, 2024: 6}
                        logger.debug(
                            f"Используем стандартную структуру pererashod.xlsx: годы={list(year_cols.keys())}, колонки={list(year_cols.values())}"
                        )
                    else:
                        logger.debug(f"Найдены годы в заголовках: {year_cols}")

                    logger.info(
                        f"Строка {row_idx} ({row_label_cell}): категория={category_key}, годы={list(year_cols.keys())}"
                    )

                    for year, col_idx in year_cols.items():
                        if col_idx < len(row_data):
                            year_str = str(year)
                            if year_str not in categories_data:
                                categories_data[year_str] = {}
                            value = row_data[col_idx]
                            if value is not None and isinstance(value, (int, float)):
                                categories_data[year_str][category_key] = float(value)
                                logger.info(
                                    f"  ✅ Год {year}: {category_key} = {value}"
                                )
                            else:
                                logger.debug(
                                    f"  ⚠️ Год {year}: {category_key} = None или не число"
                                )

        # Стандартный подход: ищем строки с категориями после строки заголовков
        if not categories_data and header_row:
            for row_idx in range(header_row + 1, table_4["end_row"] + 1):
                row_label_cell = sheet.cell(row=row_idx, column=2).value  # Столбец B
                if not row_label_cell:
                    continue

                row_label = str(row_label_cell).lower()

                # Определяем категорию по тексту в столбце B
                category_key = None
                if any(kw in row_label for kw in ["тех", "технологич", "тех-потер"]):
                    category_key = "technological"
                elif any(kw in row_label for kw in ["хоз", "бытов", "хоз-быт"]):
                    category_key = "household"
                elif any(kw in row_label for kw in ["производств", "произв"]):
                    category_key = "production"

                if category_key:
                    category_mapping[row_idx] = category_key
                    # Читаем значения по годам (столбцы C-E, индексы 2-4)
                    row_data = list(
                        sheet.iter_rows(
                            min_row=row_idx,
                            max_row=row_idx,
                            min_col=3,
                            max_col=5,
                            values_only=True,
                        )
                    )[0]

                    logger.debug(f"Строка {row_idx} ({row_label_cell}): {row_data}")

                    for year_idx, year in enumerate(years):
                        if year_idx >= len(row_data):
                            continue
                        year_str = str(year)
                        if year_str not in categories_data:
                            categories_data[year_str] = {}
                        value = row_data[year_idx]
                        categories_data[year_str][category_key] = (
                            float(value) if value else 0.0
                        )

        # Add own_needs as 0 (не присутствует в pererashod.xlsx, будет рассчитано из oborudovanie.xlsx)
        for year_data in categories_data.values():
            year_data["own_needs"] = 0.0

        result = {
            "source": str(path),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "years": categories_data,
            "table_4_location": {
                "start_row": table_4["start_row"],
                "end_row": table_4["end_row"],
            },
            "ai_analysis": {
                "used": ai_result.used_ai if ai_result else False,
                "confidence": ai_result.confidence if ai_result else 0.0,
            },
        }

        # Если есть данные из оборудования - добавляем метаданные для сверки
        if equipment_data:
            result["equipment_source_available"] = True
            result["note"] = (
                "Данные из pererashod.xlsx используются для сверки с основными данными из oborudovanie.xlsx"
            )
        else:
            result["equipment_source_available"] = False
            result["note"] = (
                "Основной источник (oborudovanie.xlsx) не найден, используются только данные из pererashod.xlsx"
            )

        workbook.close()
        return result

    except Exception as exc:
        logger.exception("Failed to parse usage categories from %s: %s", path, exc)
        return None


def distribute_categories_by_quarter(
    aggregated_data: Dict, categories_data: Dict
) -> Dict:
    """
    Distribute yearly usage categories into quarterly totals.
    If quarterly electricity data is missing, create it from yearly category totals.
    """
    if not categories_data or "years" not in categories_data:
        return aggregated_data

    years_categories = categories_data["years"]
    
    # Ensure the basic structure exists
    if "resources" not in aggregated_data:
        aggregated_data["resources"] = {}
    if "electricity" not in aggregated_data["resources"]:
        aggregated_data["resources"]["electricity"] = {}
        
    electricity_data = aggregated_data["resources"]["electricity"]

    # Check if electricity data is empty. If so, populate it from yearly categories.
    if not electricity_data and years_categories:
        logger.info("Данные по кварталам для электроэнергии отсутствуют. Создание из годовых категорий...")
        for year_str, year_cats in years_categories.items():
            yearly_total = sum(year_cats.values())
            if yearly_total > 0:
                avg_quarterly_total = yearly_total / 4.0
                avg_quarterly_by_usage = {cat: val / 4.0 for cat, val in year_cats.items()}
                for i in range(1, 5):
                    quarter_key = f"{year_str}-Q{i}"
                    electricity_data[quarter_key] = {
                        "year": int(year_str),
                        "quarter": i,
                        "months": [],
                        "quarter_totals": {"active_kwh": avg_quarterly_total},
                        "by_usage": avg_quarterly_by_usage,
                    }
        logger.info(f"Создано {len(electricity_data)} кварталов для электроэнергии из годовых данных.")

    # Original logic for distributing proportionally if quarterly data already exists
    for quarter_key, quarter_data in electricity_data.items():
        # Skip if by_usage is already populated
        if quarter_data.get("by_usage"):
            continue

        year = quarter_data.get("year")
        if not year or str(year) not in years_categories:
            continue

        year_categories = years_categories[str(year)]
        yearly_total = sum(year_categories.values())

        if yearly_total == 0:
            continue

        # Distribute proportionally to quarter_totals
        quarter_total = quarter_data.get("quarter_totals", {}).get("active_kwh", 0)
        if quarter_total and quarter_total > 0:
            quarter_data["by_usage"] = {
                category: (quarter_total * value / yearly_total)
                for category, value in year_categories.items()
            }

    return aggregated_data


def _aggregate_single_resource_from_db(
    data: Dict, filename: str, parsed_json: Dict
) -> Optional[Dict]:
    """
    Агрегация файла одного ресурса из структуры БД.

    Args:
        data: parsing.data из raw_json
        filename: Имя файла
        parsed_json: Полная структура raw_json

    Returns:
        Агрегированные данные в формате для паспорта
    """
    filename_lower = filename.lower()

    # Определяем тип ресурса по имени файла
    if "gaz" in filename_lower:
        resource_type = "gas"
    elif "voda" in filename_lower or "water" in filename_lower:
        resource_type = "water"
    elif "otoplenie" in filename_lower or "heating" in filename_lower:
        resource_type = "heat"
    elif "kotel" in filename_lower or "boiler" in filename_lower:
        resource_type = "production"
    elif "electricity" in filename_lower or "elektr" in filename_lower:
        resource_type = "electricity"
    else:
        logger.warning(f"Неизвестный тип ресурса для файла: {filename}")
        return None

    # Инициализируем результат
    result: Dict[str, Dict[str, Dict]] = {
        "electricity": {},
        "gas": {},
        "water": {},
        "fuel": {},
        "coal": {},
        "heat": {},
        "production": {},
    }

    # Получаем данные из sheets
    sheets = data.get("sheets", [])
    if not sheets:
        logger.warning(f"Нет листов в файле {filename}")
        return None

    # Используем первый лист (обычно активный)
    sheet = sheets[0]
    rows = sheet.get("rows", [])

    if not rows:
        logger.warning(f"Нет строк данных в файле {filename}")
        return None

    logger.info(
        f"Обработка файла одного ресурса '{filename}': {len(rows)} строк, ресурс: {resource_type}"
    )

    # Агрегация для газа
    if resource_type == "gas":
        # Ищем год в заголовках (обычно строка 1 или 2)
        year_cols = {}
        # Проверяем первые несколько строк на наличие годов
        for row_idx in range(min(3, len(rows))):
            header_row = rows[row_idx]
            for col_idx, value in enumerate(header_row):
                if value and isinstance(value, (int, float)):
                    try:
                        year_val = int(value)
                        if year_val in (2022, 2023, 2024):
                            year = year_val
                            # В gaz.xlsx: год, затем "сум", затем "м3"
                            # Проверяем следующие колонки для определения позиций
                            cost_col = col_idx + 1
                            volume_col = col_idx + 2
                            if len(header_row) > col_idx + 1:
                                year_cols[year] = {
                                    "cost_sum": cost_col,
                                    "volume_m3": volume_col,
                                }
                            logger.info(
                                f"Найден год {year} в колонке {col_idx}, cost={cost_col}, volume={volume_col}"
                            )
                    except (ValueError, TypeError):
                        continue

        # Обрабатываем месяцы (начиная со строки 3)
        current_year = None
        for row_idx in range(3, len(rows)):
            row = rows[row_idx]
            if not row or len(row) == 0:
                continue

            month_name = row[0] if len(row) > 0 else None
            if not isinstance(month_name, str):
                continue

            month_norm = _normalise_month_name(month_name)
            if month_norm not in MONTH_ALIASES:
                continue

            # Обрабатываем данные для каждого года
            for year, col_indices in year_cols.items():
                quarter = month_to_quarter(MONTH_ALIASES[month_norm])
                quarter_key = f"{year}-Q{quarter}"
                quarter_entry = result[resource_type].setdefault(
                    quarter_key, {"year": year, "quarter": quarter, "months": []}
                )

                cost_sum = (
                    row[col_indices["cost_sum"]]
                    if len(row) > col_indices["cost_sum"]
                    else None
                )
                volume_m3 = (
                    row[col_indices["volume_m3"]]
                    if len(row) > col_indices["volume_m3"]
                    else None
                )

                quarter_entry["months"].append(
                    {
                        "month": month_name,
                        "values": {
                            "cost_sum": cost_sum,
                            "volume_m3": volume_m3,
                        },
                    }
                )

    # Агрегация для воды
    elif resource_type == "water":
        current_year = None
        for row in rows:
            if not row or len(row) == 0:
                continue

            # Ищем год
            if (
                len(row) > 2
                and isinstance(row[2], (int, float))
                and int(row[2]) in (2022, 2023, 2024)
            ):
                current_year = int(row[2])
                continue

            if not current_year:
                continue

            month_name = row[2] if len(row) > 2 else None
            volume_m3 = row[3] if len(row) > 3 else None

            if not isinstance(month_name, str):
                continue

            month_norm = _normalise_month_name(month_name)
            if month_norm not in MONTH_ALIASES:
                continue

            quarter = month_to_quarter(MONTH_ALIASES[month_norm])
            quarter_key = f"{current_year}-Q{quarter}"
            quarter_entry = result[resource_type].setdefault(
                quarter_key, {"year": current_year, "quarter": quarter, "months": []}
            )

            quarter_entry["months"].append(
                {
                    "month": month_name,
                    "values": {
                        "volume_m3": volume_m3,
                    },
                }
            )

    # Вычисляем квартальные итоги
    _compute_quarter_totals_single_resource(result, resource_type)

    return {
        "source": "database",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "resources": result,
    }


def aggregate_from_db_json(parsed_json: Dict) -> Optional[Dict]:
    """
    Агрегировать данные напрямую из parsed_data.raw_json (БД).

    Преобразует структуру из БД (sheets → rows) в агрегированный формат
    для генерации энергопаспорта.

    Args:
        parsed_json: Структура из parsed_data.raw_json
            {
                "batch_id": "...",
                "filename": "pererashod.xlsx",
                "file_type": "excel",
                "parsing": {
                    "parsed": True,
                    "sheets": [{"name": "ЭЛЕКТР", "rows": [...]}],
                    ...
                },
                "status": "success"
            }

    Returns:
        Агрегированные данные в формате для паспорта или None
    """
    if not parsed_json:
        logger.warning("Данные отсутствуют")
        return None

    # Реальная структура из БД имеет вложенный объект "parsing"
    parsing = parsed_json.get("parsing", {})
    if not parsing or not parsing.get("parsed"):
        logger.warning("Данные не распарсены или отсутствуют")
        return None

    # Данные лежат в parsing.data (результат parse_excel_file)
    data = parsing.get("data", {})
    if not data:
        logger.warning("Нет данных в parsing.data")
        return None

    file_type = parsing.get("file_type", "")

    # Для Excel файлов работаем со структурой sheets
    if file_type == "excel" and "sheets" in data:
        # Определяем тип файла по имени файла
        filename = parsed_json.get("filename", "").lower()
        is_single_resource_file = any(
            keyword in filename
            for keyword in [
                "gaz.xlsx",
                "voda.xlsx",
                "otoplenie.xlsx",
                "kotel.xlsx",
                "electricity.xlsx",
            ]
        )

        # Если это файл одного ресурса, используем другую логику
        if is_single_resource_file:
            logger.info(f"Определен файл одного ресурса: {filename}")
            return _aggregate_single_resource_from_db(data, filename, parsed_json)

        # Преобразуем в формат workbook для aggregate_energy_data
        # Создаём временную структуру, совместимую с openpyxl

        result: Dict[str, Dict[str, Dict]] = {
            "electricity": {},
            "gas": {},
            "water": {},
            "fuel": {},
            "coal": {},
            "heat": {},
            "production": {},
        }
        missing_sheets = []

        # Создаём mapping: имя листа → данные
        sheets_map = {sheet["name"]: sheet for sheet in data.get("sheets", [])}

        def aggregate_months_from_rows(
            target: Dict,
            rows: list,
            year_col: int = 0,
            month_col: int = 0,
            value_cols: Dict[str, int] = None,
        ) -> None:
            """Агрегация из rows в структуру по кварталам"""
            current_year: Optional[int] = None
            processed_rows = 0
            added_months = 0

            for row_idx, row in enumerate(rows):
                if not row or len(row) == 0:
                    continue

                # Проверяем год
                first_cell = row[year_col] if len(row) > year_col else None
                if isinstance(first_cell, int) and first_cell in (2022, 2023, 2024):
                    current_year = first_cell
                    logger.debug(
                        f"📅 [DIAG] Найден год {current_year} в строке {row_idx}"
                    )
                    continue

                # Проверяем месяц
                if current_year and isinstance(first_cell, str):
                    month_name = first_cell
                    month_key = _normalise_month_name(month_name)
                    if month_key not in MONTH_ALIASES:
                        logger.debug(
                            f"⚠️ [DIAG] Месяц '{month_name}' не распознан (строка {row_idx})"
                        )
                        continue

                    quarter = month_to_quarter(MONTH_ALIASES[month_key])
                    quarter_key = f"{current_year}-Q{quarter}"
                    quarter_entry = target.setdefault(
                        quarter_key,
                        {"year": current_year, "quarter": quarter, "months": []},
                    )

                    # Извлекаем значения из указанных колонок
                    values = {}
                    if value_cols:
                        for field_name, col_idx in value_cols.items():
                            value = row[col_idx] if len(row) > col_idx else None
                            values[field_name] = value
                            if value is not None:
                                logger.debug(
                                    f"📊 [DIAG] Извлечено {field_name}={value} из строки {row_idx}, колонка {col_idx}"
                                )

                    quarter_entry["months"].append(
                        {"month": month_name, "values": values}
                    )
                    added_months += 1
                    processed_rows += 1
                    logger.debug(
                        f"✅ [DIAG] Добавлен месяц {month_name} ({current_year}) в квартал {quarter_key}: "
                        f"поля={list(values.keys())}, значения={[(k, v) for k, v in values.items() if v is not None]}"
                    )

            logger.info(
                f"📊 [DIAG] Обработано строк: {processed_rows}, добавлено месяцев: {added_months}"
            )

        def aggregate_months_from_numbered_rows(
            target: Dict,
            rows: list,
            month_num_col: int = 0,
            value_cols: Dict[str, int] = None,
        ) -> None:
            """Агрегация из rows, где первая колонка - номер месяца (1-12)"""
            current_year: Optional[int] = None
            processed_rows = 0
            added_months = 0
            data_row_count = 0  # Счетчик строк с данными (без заголовков)

            # Определяем год по позиции: строки 1-12 = 2022, 13-24 = 2023, 25-36 = 2024
            # Или ищем год в заголовках/метаданных

            for row_idx, row in enumerate(rows):
                if not row or len(row) == 0:
                    continue

                # Пропускаем заголовки
                first_cell = row[month_num_col] if len(row) > month_num_col else None
                if isinstance(first_cell, str) and any(
                    keyword in first_cell.lower()
                    for keyword in ["оаж", "сум", "киловат", "месяц", "год"]
                ):
                    logger.debug(
                        f"⏭️ [DIAG] Пропущен заголовок в строке {row_idx}: {first_cell}"
                    )
                    continue

                # Проверяем, является ли первая ячейка номером месяца (1-12)
                if isinstance(first_cell, (int, float)) and 1 <= first_cell <= 12:
                    month_num = int(first_cell)
                    data_row_count += 1

                    # Определяем год по позиции строки с данными (без учета заголовков)
                    # Предполагаем: строки данных 1-12 = 2022, 13-24 = 2023, 25-36 = 2024
                    if data_row_count <= 12:
                        current_year = 2022
                    elif data_row_count <= 24:
                        current_year = 2023
                    elif data_row_count <= 36:
                        current_year = 2024
                    else:
                        # Если больше 36 строк, продолжаем с 2024 или определяем по-другому
                        current_year = 2024

                    # Преобразуем номер месяца в название
                    month_names = [
                        "Январь",
                        "Февраль",
                        "Март",
                        "Апрель",
                        "Май",
                        "Июнь",
                        "Июль",
                        "Август",
                        "Сентябрь",
                        "Октябрь",
                        "Ноябрь",
                        "Декабрь",
                    ]
                    month_name = month_names[month_num - 1]

                    quarter = month_to_quarter(month_num)
                    quarter_key = f"{current_year}-Q{quarter}"
                    quarter_entry = target.setdefault(
                        quarter_key,
                        {"year": current_year, "quarter": quarter, "months": []},
                    )

                    # Извлекаем значения из указанных колонок
                    values = {}
                    if value_cols:
                        for field_name, col_idx in value_cols.items():
                            value = row[col_idx] if len(row) > col_idx else None
                            values[field_name] = value
                            if value is not None:
                                logger.debug(
                                    f"📊 [DIAG] Извлечено {field_name}={value} из строки {row_idx}, колонка {col_idx}"
                                )

                    quarter_entry["months"].append(
                        {"month": month_name, "values": values}
                    )
                    added_months += 1
                    processed_rows += 1
                    logger.debug(
                        f"✅ [DIAG] Добавлен месяц {month_name} ({current_year}, номер {month_num}) в квартал {quarter_key}: "
                        f"поля={list(values.keys())}, значения={[(k, v) for k, v in values.items() if v is not None]}"
                    )

            logger.info(
                f"📊 [DIAG] Обработано строк (номерной формат): {processed_rows}, добавлено месяцев: {added_months}"
            )

        # Парсим лист "Килограмм да" (Production)
        if "Килограмм да" in sheets_map:
            sheet = sheets_map["Килограмм да"]
            aggregate_months_from_rows(
                result["production"],
                sheet.get("rows", []),
                year_col=0,
                month_col=0,
                value_cols={
                    "Труба хвс": 1,
                    "Канал труба": 2,
                    "Канал фитинг": 3,
                    "Фит хвс и гвс": 4,
                    "Теплый пол": 5,
                    "Жами": 6,
                },
            )
        else:
            missing_sheets.append("Килограмм да")

        # Парсим лист с электроэнергией (проверяем альтернативные названия)
        electricity_sheet_name = None

        # ПРИМЕЧАНИЕ: pererashod.xlsx содержит только категории использования (by_usage),
        # но НЕ содержит месячных данных по электроэнергии. Sheet1 в pererashod.xlsx содержит описание ПКМ 690.
        # Месячные данные по электроэнергии должны быть в другом файле или листе.

        # Ищем по названию листа (точное совпадение)
        if not electricity_sheet_name:
            electricity_sheet_names = [
                "ЭЛЕКТР",
                "электр ",
                "электр",
                "Электроэнергия",
                "Электричество",
                "ТП",
                "Electricity",
                "ЭЛЕКТРИЧЕСТВО",
                "ЭЛЕКТРО",
                "ЭЛЕКТРОЭНЕРГИЯ",
                "электроэнергия",
                "электричество",
                "Энергоресурсы",
                "энергоресурсы",
                "ELECTRO",
                "ELECTRIC",
                # Листы из файлов "Реализация"
                "Реал 04", "Реал 00", "Реал 01", "Реал 02", "Реал 03",
                "Баланс 04", "Баланс 00", "Баланс 01", "Баланс 02", "Баланс 03",
            ]
            for sheet_name_var in electricity_sheet_names:
                if sheet_name_var in sheets_map:
                    electricity_sheet_name = sheet_name_var
                    logger.info(
                        f"✅ [DIAG] Найден лист электроэнергии по точному совпадению: '{sheet_name_var}'"
                    )
                    break

        # Если не нашли по точному совпадению, ищем по частичному совпадению (с учетом пробелов в конце)
        if not electricity_sheet_name:
            for sheet_name in sheets_map.keys():
                sheet_name_clean = sheet_name.strip()
                sheet_name_lower = sheet_name_clean.lower()
                # Расширенный поиск: включаем листы из файлов "Реализация" (Реал, Баланс)
                if any(
                    keyword in sheet_name_lower
                    for keyword in [
                        "электр", "electric", "тп", "электроэнергия",
                        "реал", "баланс", "реализация"  # Для файлов "Реализация"
                    ]
                ):
                    electricity_sheet_name = sheet_name
                    logger.info(
                        f"✅ [DIAG] Найден лист электроэнергии по частичному совпадению: '{sheet_name}'"
                    )
                    break

        if electricity_sheet_name:
            sheet = sheets_map[electricity_sheet_name]
            rows = sheet.get("rows", [])
            logger.info(
                f"🔍 [DIAG] Найден лист электроэнергии '{electricity_sheet_name}': {len(rows)} строк"
            )
            if rows:
                logger.debug(
                    f"📋 [DIAG] Первые 5 строк листа '{electricity_sheet_name}':"
                )
                for idx, row in enumerate(rows[:5]):
                    row_preview = row[:10] if len(row) > 10 else row
                    logger.debug(f"  Строка {idx}: {row_preview} (длина: {len(row)})")
                    if row:
                        first_cell_type = (
                            type(row[0]).__name__ if len(row) > 0 else "нет"
                        )
                        first_cell_value = repr(row[0]) if len(row) > 0 else "нет"
                        logger.debug(
                            f"    Тип первой ячейки: {first_cell_type}, значение: {first_cell_value}"
                        )

            # Определяем колонки по заголовкам (первые 3 строки)
            col_cost = 1  # По умолчанию
            col_active_kwh = 2  # По умолчанию
            col_reactive_kvarh = 3  # По умолчанию
            col_active_other = 4  # По умолчанию

            # Анализируем заголовки для определения колонок
            header_rows = rows[:3] if len(rows) >= 3 else rows
            for header_row in header_rows:
                if not header_row:
                    continue
                for col_idx, cell_value in enumerate(header_row):
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = str(cell_value).lower()
                        # Ищем колонку с активной энергией (кВт·ч)
                        if any(
                            keyword in cell_lower
                            for keyword in [
                                "квт·ч",
                                "квтч",
                                "kwh",
                                "активная энергия",
                                "актив",
                                "active",
                            ]
                        ):
                            if (
                                "реакт" not in cell_lower
                                and "reactive" not in cell_lower
                            ):
                                col_active_kwh = col_idx
                                logger.info(
                                    f"📊 [DIAG] Найдена колонка active_kwh: {col_idx} ('{cell_value}')"
                                )
                        # Ищем колонку с реактивной энергией (кВАр·ч)
                        elif any(
                            keyword in cell_lower
                            for keyword in [
                                "квар·ч",
                                "кварч",
                                "kvarh",
                                "реактивная энергия",
                                "реактив",
                                "reactive",
                            ]
                        ):
                            col_reactive_kvarh = col_idx
                            logger.info(
                                f"📊 [DIAG] Найдена колонка reactive_kvarh: {col_idx} ('{cell_value}')"
                            )
                        # Ищем колонку со стоимостью
                        elif any(
                            keyword in cell_lower
                            for keyword in ["сум", "стоимость", "cost", "цена", "price"]
                        ):
                            col_cost = col_idx
                            logger.info(
                                f"📊 [DIAG] Найдена колонка cost_sum: {col_idx} ('{cell_value}')"
                            )

            # Проверяем структуру данных: если первая строка - заголовок, а дальше номера месяцев
            # то это альтернативный формат (номер месяца в первой колонке)
            first_data_row = None
            for row in rows:
                if row and len(row) > 0:
                    first_cell = row[0]
                    # Пропускаем заголовки
                    if isinstance(first_cell, str) and any(
                        keyword in first_cell.lower()
                        for keyword in [
                            "оаж",
                            "сум",
                            "киловат",
                            "месяц",
                            "год",
                            "наименование",
                        ]
                    ):
                        continue
                    # Если первая ячейка - число от 1 до 12, это номер месяца
                    if isinstance(first_cell, (int, float)) and 1 <= first_cell <= 12:
                        first_data_row = row
                        break

            if first_data_row:
                # Альтернативный формат: номер месяца в первой колонке
                logger.info(
                    "📊 [DIAG] Обнаружен альтернативный формат: номер месяца в первой колонке"
                )
                aggregate_months_from_numbered_rows(
                    result["electricity"],
                    rows,
                    month_num_col=0,
                    value_cols={
                        "cost_sum": col_cost,
                        "active_kwh": col_active_kwh,
                        "reactive_kvarh": col_reactive_kvarh,
                        "active_other": col_active_other,
                    },
                )
            else:
                # Стандартный формат: год и название месяца
                logger.info(
                    "📊 [DIAG] Используется стандартный формат: год и название месяца"
                )
                aggregate_months_from_rows(
                    result["electricity"],
                    rows,
                    year_col=0,
                    month_col=0,
                    value_cols={
                        "cost_sum": col_cost,
                        "active_kwh": col_active_kwh,
                        "reactive_kvarh": col_reactive_kvarh,
                        "active_other": col_active_other,
                    },
                )

            # Проверяем результат агрегации
            electricity_quarters = result["electricity"]
            logger.info(
                f"📊 [DIAG] После агрегации электроэнергии: {len(electricity_quarters)} кварталов"
            )
            for quarter_key, quarter_data in electricity_quarters.items():
                months_count = len(quarter_data.get("months", []))
                logger.debug(f"  └─ {quarter_key}: {months_count} месяцев")
        else:
            missing_sheets.append("ЭЛЕКТР (или альтернативные названия)")
            logger.warning(
                f"⚠️ [DIAG] Лист с электроэнергией не найден. Доступные листы: {list(sheets_map.keys())}"
            )

        # Парсим лист "ГАЗ"
        if "ГАЗ" in sheets_map:
            sheet = sheets_map["ГАЗ"]
            aggregate_months_from_rows(
                result["gas"],
                sheet.get("rows", []),
                year_col=0,
                month_col=0,
                value_cols={
                    "cost_sum": 1,
                    "volume_m3": 2,
                    "extra": 3,
                },
            )
        else:
            missing_sheets.append("ГАЗ")

        # Парсим лист "СУВ" (вода)
        if "СУВ" in sheets_map:
            sheet = sheets_map["СУВ"]
            aggregate_months_from_rows(
                result["water"],
                sheet.get("rows", []),
                year_col=0,
                month_col=0,
                value_cols={
                    "volume_m3": 1,
                    "cost_sum": 2,
                },
            )
        else:
            missing_sheets.append("СУВ")

        # Вычисляем квартальные итоги
        _compute_quarter_totals(result)

        return {
            "source": "database",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "resources": result,
            "missing_sheets": missing_sheets,
        }

    logger.warning("Неподдерживаемый формат данных для агрегации")
    return None


def write_aggregation_json(
    batch_id: str,
    aggregation_data: Dict,
    destination_dir: Union[str, Path],
) -> Path:
    destination = Path(destination_dir)
    destination.mkdir(parents=True, exist_ok=True)
    target_file = destination / f"{batch_id}_aggregated.json"
    target_file.write_text(
        json.dumps(aggregation_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return target_file
