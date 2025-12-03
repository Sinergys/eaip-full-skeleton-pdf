"""
Модуль для восстановления формул в энергопаспорте с помощью ИИ.
"""

import logging
from typing import Dict, Any, Optional
import re

logger = logging.getLogger(__name__)

# Импорт ИИ парсера (опционально)
try:
    from ai_parser import get_ai_parser

    HAS_AI_PARSER = True
except ImportError:
    HAS_AI_PARSER = False
    logger.warning("ai_parser модуль не найден. ИИ функции недоступны.")


class AIFormulaRestorer:
    """
    Восстановление недостающих формул в энергопаспорте с помощью ИИ.
    """

    def __init__(self):
        """Инициализация восстановителя формул."""
        self.ai_parser = None
        if HAS_AI_PARSER:
            try:
                self.ai_parser = get_ai_parser()
            except Exception as e:
                logger.warning(f"Не удалось инициализировать AI парсер: {e}")

        # Шаблоны формул для восстановления
        self.formula_patterns = self._load_formula_patterns()

    def _load_formula_patterns(self) -> Dict[str, Any]:
        """Загружает шаблоны формул для восстановления."""
        return {
            "cross_sheet_reference": {
                "pattern": r"'([^']+)'!#REF!",
                "description": "Ссылка на другой лист с ошибкой #REF!",
            },
            "sum_formula": {
                "indicators": ["итого", "итог", "сумма", "total", "всего"],
                "description": "Формула суммы",
            },
            "specific_consumption": {
                "pattern": "энергия / производство",
                "description": "Удельный расход",
            },
        }

    def restore_ref_error(
        self, workbook, sheet_name: str, cell_coordinate: str, formula: str
    ) -> Optional[str]:
        """
        Восстанавливает формулу с ошибкой #REF!.

        Args:
            workbook: Рабочая книга openpyxl
            sheet_name: Имя листа с ошибкой
            cell_coordinate: Координаты ячейки (например, "AF13")
            formula: Текущая формула с ошибкой

        Returns:
            Восстановленная формула или None
        """
        # Ищем паттерн #REF! в формуле
        ref_pattern = r"'([^']+)'!#REF!"
        match = re.search(ref_pattern, formula)

        if not match:
            return None

        source_sheet_name = match.group(1).strip()

        # Получаем текущий лист
        current_sheet = workbook[sheet_name]
        current_sheet[cell_coordinate]

        # Анализируем соседние ячейки для определения паттерна
        restored = self._restore_by_pattern(
            workbook, current_sheet, cell_coordinate, source_sheet_name
        )

        if restored:
            logger.info(
                f"Восстановлена формула {cell_coordinate} в листе '{sheet_name}': "
                f"{formula} -> {restored}"
            )
            return restored

        # Если не удалось восстановить по паттерну, используем ИИ
        if self.ai_parser and self.ai_parser.enabled:
            return self._restore_with_ai(
                workbook, sheet_name, cell_coordinate, formula, source_sheet_name
            )

        return None

    def _restore_by_pattern(
        self, workbook, current_sheet, cell_coordinate: str, source_sheet_name: str
    ) -> Optional[str]:
        """
        Восстанавливает формулу по паттерну соседних ячеек.

        Логика:
        - Ищем похожие формулы в соседних ячейках
        - Определяем паттерн (смещение строки/столбца)
        - Применяем паттерн для восстановления
        """
        try:
            from openpyxl.utils import column_index_from_string

            # Парсим координаты (формат: "AF13")
            # Извлекаем букву колонки и номер строки
            import re

            coord_match = re.match(r"([A-Z]+)(\d+)", cell_coordinate)
            if not coord_match:
                logger.warning(f"Неверный формат координат: {cell_coordinate}")
                return None

            col_letter = coord_match.group(1)
            row_num = int(coord_match.group(2))
            column_index_from_string(col_letter)

            # Проверяем, существует ли исходный лист (с учетом возможных пробелов)
            # Ищем точное совпадение или с пробелами
            matching_sheet_name = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.strip() == source_sheet_name.strip():
                    matching_sheet_name = sheet_name
                    break

            if not matching_sheet_name:
                logger.warning(
                    f"Лист '{source_sheet_name}' не найден в рабочей книге. Доступные: {workbook.sheetnames[:5]}"
                )
                return None

            source_sheet = workbook[matching_sheet_name]

            # Ищем похожие формулы в соседних ячейках (вверх/вниз)
            pattern_found = False
            target_cell = None

            # Проверяем ячейки выше (row_num - 1, row_num - 2, row_num - 3)
            for offset in [1, 2, 3]:
                check_row = row_num - offset
                if check_row >= 1:
                    check_cell = current_sheet[f"{col_letter}{check_row}"]
                    if check_cell.data_type == "f" and check_cell.value:
                        formula_str = str(check_cell.value)
                        # Ищем ссылку на исходный лист
                        if (
                            f"'{source_sheet_name}'!" in formula_str
                            or f"'{matching_sheet_name}'!" in formula_str
                        ):
                            # Извлекаем ссылку на ячейку
                            cell_ref_match = re.search(r"!([A-Z]+\d+)", formula_str)
                            if cell_ref_match:
                                ref_cell = cell_ref_match.group(1)
                                # Парсим ссылку на ячейку (формат: "AM14")
                                ref_match = re.match(r"([A-Z]+)(\d+)", ref_cell)
                                if ref_match:
                                    ref_col = ref_match.group(1)
                                    ref_row_idx = int(ref_match.group(2))
                                # Применяем то же смещение
                                target_row = ref_row_idx + offset
                                target_cell = f"{ref_col}{target_row}"

                                # Проверяем, что целевая ячейка существует
                                if target_row <= source_sheet.max_row:
                                    pattern_found = True
                                    break

            # Если не нашли выше, проверяем ниже
            if not pattern_found:
                for offset in [1, 2, 3]:
                    check_row = row_num + offset
                    if check_row <= current_sheet.max_row:
                        check_cell = current_sheet[f"{col_letter}{check_row}"]
                        if check_cell.data_type == "f" and check_cell.value:
                            formula_str = str(check_cell.value)
                            if (
                                f"'{source_sheet_name}'!" in formula_str
                                or f"'{matching_sheet_name}'!" in formula_str
                            ):
                                cell_ref_match = re.search(r"!([A-Z]+\d+)", formula_str)
                                if cell_ref_match:
                                    ref_cell = cell_ref_match.group(1)
                                    # Парсим ссылку на ячейку
                                    ref_match = re.match(r"([A-Z]+)(\d+)", ref_cell)
                                    if ref_match:
                                        ref_col = ref_match.group(1)
                                        ref_row_idx = int(ref_match.group(2))
                                    target_row = ref_row_idx - offset
                                    target_cell = f"{ref_col}{target_row}"

                                    if target_row >= 1:
                                        pattern_found = True
                                        break

            if pattern_found and target_cell:
                restored_formula = f"='{source_sheet_name}'!{target_cell}"
                return restored_formula

        except Exception as e:
            logger.error(f"Ошибка при восстановлении по паттерну: {e}")

        return None

    def _restore_with_ai(
        self,
        workbook,
        sheet_name: str,
        cell_coordinate: str,
        formula: str,
        source_sheet_name: str,
    ) -> Optional[str]:
        """
        Восстанавливает формулу с помощью ИИ.

        TODO: Реализовать использование ИИ для анализа контекста
        и восстановления формулы на основе содержимого листов.
        """
        # Пока возвращаем None - будет реализовано позже
        logger.info(f"ИИ восстановление пока не реализовано для {cell_coordinate}")
        return None


def restore_formulas_in_file(
    file_path: str, output_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    Восстанавливает формулы в файле энергопаспорта.

    Args:
        file_path: Путь к файлу энергопаспорта
        output_path: Путь для сохранения результата (если None, не сохраняет)

    Returns:
        Отчет о восстановленных формулах
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

    workbook = load_workbook(file_path, data_only=False)
    restorer = AIFormulaRestorer()

    restored = {}
    total_found = 0
    total_restored = 0

    # Ищем все ячейки с ошибками #REF!
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_restored = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    formula_str = str(cell.value)
                    if "#REF!" in formula_str:
                        total_found += 1
                        cell_coord = cell.coordinate

                        restored_formula = restorer.restore_ref_error(
                            workbook, sheet_name, cell_coord, formula_str
                        )

                        if restored_formula:
                            # Применяем восстановленную формулу
                            cell.value = restored_formula
                            total_restored += 1

                            sheet_restored.append(
                                {
                                    "cell": cell_coord,
                                    "old_formula": formula_str,
                                    "new_formula": restored_formula,
                                }
                            )

        if sheet_restored:
            restored[sheet_name] = sheet_restored

    # Сохраняем результат, если указан путь
    if output_path:
        workbook.save(output_path)
        logger.info(f"Файл с восстановленными формулами сохранен: {output_path}")

    return {
        "file_path": file_path,
        "output_path": output_path,
        "total_ref_errors_found": total_found,
        "total_restored": total_restored,
        "restored_by_sheet": restored,
    }
