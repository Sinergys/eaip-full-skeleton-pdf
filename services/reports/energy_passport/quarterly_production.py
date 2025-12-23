"""Модуль для агрегации месячных данных по производству продукции в квартальные таблицы."""

from collections import defaultdict
from typing import Dict, List, Any
from openpyxl.worksheet.worksheet import Worksheet
import logging

logger = logging.getLogger(__name__)

# Маппинг месяцев к кварталам
MONTH_TO_QUARTER = {
    "Январь": 1,
    "Февраль": 1,
    "Март": 1,
    "Апрель": 2,
    "Май": 2,
    "Июнь": 2,
    "Июль": 3,
    "Август": 3,
    "Сентябрь": 3,
    "Октябрь": 4,
    "Ноябрь": 4,
    "Декабрь": 4,
}

# Названия продуктов (стандартизированные)
PRODUCT_COLUMNS = [
    "Труба ХВС",
    "Канал труба",  # Канализационные трубы
    "Канал фитинг",  # Канализационные фитинги
    "Фит ХВС и ГВС",  # Фитинги ХВС и ГВС
    "Топ. пол",  # Трубы тёплого пола
]

# Маппинг различных названий продуктов к стандартным
PRODUCT_NAME_MAPPING = {
    "труба хвс": "Труба ХВС",
    "трубы хвс": "Труба ХВС",
    "канализационные трубы": "Канал труба",
    "канал труба": "Канал труба",
    "канализационные фитинги": "Канал фитинг",
    "канал фитинг": "Канал фитинг",
    "фитинги хвс и гвс": "Фит ХВС и ГВС",
    "фит хвс и гвс": "Фит ХВС и ГВС",
    "фитинги хвс (хвс и гвс)": "Фит ХВС и ГВС",
    "трубы тёплого пола": "Топ. пол",
    "тёплый пол": "Топ. пол",
    "теплый пол": "Топ. пол",
    "топ. пол": "Топ. пол",
}


def normalize_product_name(product_name: str) -> str:
    """
    Нормализует название продукта к стандартному виду.

    Args:
        product_name: Исходное название продукта

    Returns:
        Стандартизированное название продукта
    """
    if not product_name:
        return ""

    product_lower = product_name.lower().strip()

    # Прямое совпадение
    if product_lower in PRODUCT_NAME_MAPPING:
        return PRODUCT_NAME_MAPPING[product_lower]

    # Частичное совпадение
    for key, standard_name in PRODUCT_NAME_MAPPING.items():
        if key in product_lower or product_lower in key:
            return standard_name

    # Если не найдено, возвращаем как есть
    return product_name


def aggregate_quarters_by_product(
    monthly_production: Dict[int, Dict[str, Dict[str, float]]], year: int
) -> Dict[int, Dict[str, float]]:
    """
    Агрегирует месячные данные по производству продукции в квартальные итоги.

    Args:
        monthly_production: Словарь вида {year: {month: {product: value}}}
            где month - строки "Январь", "Февраль", ..., "Декабрь"
        year: Год для агрегации

    Returns:
        Словарь вида {quarter: {product: total_value, "ИТОГО": total}}
        где quarter - 1, 2, 3, 4
    """
    result: Dict[int, Dict[str, float]] = defaultdict(lambda: defaultdict(float))

    data_for_year = monthly_production.get(year, {})

    if not data_for_year:
        logger.warning(f"Нет данных по производству для года {year}")
        return {}

    for month, products in data_for_year.items():
        q = MONTH_TO_QUARTER.get(month.strip())
        if q is None:
            logger.warning(f"Неизвестный месяц: {month}")
            continue

        for product_name, value in products.items():
            if value is None:
                continue

            try:
                value_float = float(value)
                if value_float <= 0:
                    continue

                # Нормализуем название продукта
                normalized_product = normalize_product_name(product_name)

                if normalized_product:
                    result[q][normalized_product] += value_float
            except (ValueError, TypeError) as e:
                logger.debug(f"Пропущено значение для {product_name} в {month}: {e}")
                continue

    # Считаем ИТОГО по каждому кварталу
    for q, prod_dict in result.items():
        total = sum(v for v in prod_dict.values())
        prod_dict["ИТОГО"] = total
        logger.debug(f"Квартал {q}: ИТОГО = {total}")

    return dict(result)


def write_quarter_tables_to_sheet(
    ws: Worksheet,
    monthly_production: Dict[int, Dict[str, Dict[str, float]]],
    years: List[int],
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    """
    Записывает квартальные таблицы по производству продукции в лист Excel.

    Args:
        ws: Лист Excel для записи
        monthly_production: Словарь месячных данных {year: {month: {product: value}}}
        years: Список лет для обработки
        start_row: Начальная строка для записи (по умолчанию 1)
        start_col: Начальная колонка для записи (по умолчанию 1)
    """
    row = start_row

    # Заголовок
    ws.cell(row=row, column=start_col, value="Год")
    ws.cell(row=row, column=start_col + 1, value="Квартал")

    for i, prod in enumerate(PRODUCT_COLUMNS, start=start_col + 2):
        ws.cell(row=row, column=i, value=prod)

    ws.cell(row=row, column=start_col + 2 + len(PRODUCT_COLUMNS), value="ИТОГО")
    row += 1

    # Данные по годам и кварталам
    for year in sorted(years):
        q_data = aggregate_quarters_by_product(monthly_production, year)

        if not q_data:
            logger.warning(f"Нет данных для года {year}, пропускаем")
            continue

        for q in sorted(q_data.keys()):
            ws.cell(row=row, column=start_col, value=year)
            ws.cell(row=row, column=start_col + 1, value=f"Кв{q}")

            # Продукты по колонкам
            for i, prod in enumerate(PRODUCT_COLUMNS, start=start_col + 2):
                value = q_data[q].get(prod, 0.0)
                ws.cell(row=row, column=i, value=value)

            # ИТОГО по кварталу
            total_value = q_data[q].get("ИТОГО", 0.0)
            ws.cell(
                row=row,
                column=start_col + 2 + len(PRODUCT_COLUMNS),
                value=total_value,
            )

            logger.debug(
                f"Записана строка: год={year}, квартал={q}, итого={total_value}"
            )
            row += 1

    logger.info(f"Записано {row - start_row - 1} строк квартальных данных")


def extract_monthly_production_from_aggregated(
    aggregated_data: Dict[str, Any],
) -> Dict[int, Dict[str, Dict[str, float]]]:
    """
    Извлекает месячные данные по производству из aggregated_data.

    Ищет данные в структуре:
    - aggregated_data["resources"]["production"][quarter_key]["months"][month]["values"][product]
    - или aggregated_data["production"][quarter_key]["months"][month]["values"][product]

    Args:
        aggregated_data: Агрегированные данные из EAIP

    Returns:
        Словарь вида {year: {month: {product: value}}}
    """
    monthly_production: Dict[int, Dict[str, Dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(float))
    )

    # Пробуем разные пути к данным
    production_data = None

    # Вариант 1: aggregated_data["resources"]["production"]
    resources = aggregated_data.get("resources", {})
    if isinstance(resources, dict):
        production_data = resources.get("production", {})

    # Вариант 2: aggregated_data["production"]
    if not production_data:
        production_data = aggregated_data.get("production", {})

    # Вариант 3: если aggregated_data уже содержит production на верхнем уровне
    if not production_data and "production" in aggregated_data:
        production_data = aggregated_data["production"]

    if not production_data:
        logger.warning("Нет данных по производству в aggregated_data")
        return {}

    logger.debug(f"Найдены данные по производству: {len(production_data)} кварталов")

    # Обрабатываем квартальные данные
    for quarter_key, quarter_data in production_data.items():
        if not isinstance(quarter_data, dict):
            continue

        if "-Q" not in str(quarter_key):
            continue

        try:
            year_str, quarter_str = str(quarter_key).split("-Q")
            year = int(year_str)
        except (ValueError, AttributeError) as e:
            logger.debug(f"Неверный формат квартала: {quarter_key}, ошибка: {e}")
            continue

        months = quarter_data.get("months", [])

        if not months:
            logger.debug(f"Нет месячных данных для квартала {quarter_key}")
            continue

        for month_entry in months:
            if not isinstance(month_entry, dict):
                continue

            month_name = month_entry.get("month", "").strip()
            if not month_name:
                continue

            values = month_entry.get("values", {})
            if not isinstance(values, dict):
                continue

            for product_name, value in values.items():
                if value is None:
                    continue

                # Пропускаем "Жами" (итого)
                if product_name.lower() in ("жами", "итого", "total"):
                    continue

                try:
                    value_float = float(value)
                    if value_float > 0:
                        monthly_production[year][month_name][product_name] = value_float
                        logger.debug(
                            f"Добавлено: {year}, {month_name}, {product_name} = {value_float}"
                        )
                except (ValueError, TypeError) as e:
                    logger.debug(
                        f"Пропущено значение для {product_name} в {month_name}: {e}"
                    )
                    continue

    total_months = sum(len(months) for months in monthly_production.values())
    logger.info(
        f"Извлечено месячных данных: {len(monthly_production)} лет, "
        f"всего месяцев: {total_months}"
    )

    if total_months == 0:
        logger.warning("Не удалось извлечь месячные данные по производству")

    return dict(monthly_production)
