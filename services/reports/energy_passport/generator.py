"""
Главная функция для генерации энергопаспорта.

Единая точка входа для создания полного Excel-энергопаспорта.
"""

from pathlib import Path
from typing import Dict, Any, Optional
import logging
from openpyxl import load_workbook

from .data_collector import collect_energy_passport_data, EnergyPassportData

# Импортируем функции заполнения из tools/fill_energy_passport
import sys
from pathlib import Path as PathLib

# Определяем путь к tools относительно корня проекта
_project_root = PathLib(__file__).parent.parent.parent.parent
_fill_module_path = _project_root / "tools"
if str(_fill_module_path) not in sys.path:
    sys.path.insert(0, str(_fill_module_path))

logger = logging.getLogger(__name__)

try:
    from fill_energy_passport import (
        fill_struktura_pr2,
        fill_electricity_by_product,
        fill_gas_specific_consumption,
        fill_gas_by_usage_categories,
        fill_nodes_sheet,
        fill_equipment_sheet,
        fill_building_envelope_sheet,
        fill_losses_sheet,
    )
except ImportError as e:
    logger.error(f"Не удалось импортировать функции заполнения: {e}")
    raise


def generate_energy_passport(
    enterprise_id: str,
    year: int,
    template_path: Path,
    output_path: Path,
    aggregated_data: Optional[Dict[str, Any]] = None,
    enterprise_data: Optional[Dict[str, Any]] = None,
    building_data: Optional[Dict[str, Any]] = None,
    nodes_json_path: Optional[Path] = None,
    equipment_json_path: Optional[Path] = None,
    envelope_json_path: Optional[Path] = None,
    loss_active_month: float = 0.0,
    loss_reactive_month: float = 0.0,
    transformer_power_kva: float = 0.0,
) -> Path:
    """
    Генерирует полный энергопаспорт в формате Excel.

    Args:
        enterprise_id: ID предприятия (UUID или строка)
        year: Год для генерации паспорта
        template_path: Путь к шаблону Excel
        output_path: Путь для сохранения готового паспорта
        aggregated_data: Агрегированные данные энергопотребления (опционально)
        enterprise_data: Данные предприятия (опционально)
        building_data: Данные о здании (опционально)
        nodes_json_path: Путь к JSON с данными узлов учета (опционально)
        equipment_json_path: Путь к JSON с данными оборудования (опционально)
        envelope_json_path: Путь к JSON с данными ограждающих конструкций (опционально)
        loss_active_month: Потери активной энергии за месяц (кВт·ч)
        loss_reactive_month: Потери реактивной энергии за месяц (кВАр·ч)
        transformer_power_kva: Мощность трансформатора (кВА)

    Returns:
        Path к сгенерированному файлу

    Raises:
        FileNotFoundError: Если шаблон не найден
        ValueError: Если данные некорректны
    """
    logger.info(
        f"Начало генерации энергопаспорта: enterprise_id={enterprise_id}, year={year}"
    )

    # Проверяем наличие шаблона
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    # Загружаем шаблон
    logger.info(f"Загрузка шаблона: {template_path}")
    workbook = load_workbook(template_path, data_only=False)

    # Собираем данные для паспорта
    if aggregated_data is None:
        logger.warning("aggregated_data не предоставлен, используем пустые данные")
        aggregated_data = {"resources": {}}

    passport_data: EnergyPassportData = collect_energy_passport_data(
        enterprise_id=enterprise_id,
        year=year,
        aggregated_data=aggregated_data,
        enterprise_data=enterprise_data,
        building_data=building_data,
    )

    logger.info(
        f"Данные собраны: площадь={passport_data.building_area_m2} м², "
        f"продуктов={len(passport_data.electricity_by_product)}, "
        f"газ={passport_data.gas_data is not None}"
    )

    # Подготавливаем данные для fill_struktura_pr2
    resources_data = aggregated_data.get("resources") or aggregated_data

    # Заполняем лист "Структура пр 2"
    if "Структура пр 2" in workbook.sheetnames:
        fill_struktura_pr2(
            workbook["Структура пр 2"],
            resources_data,
            loss_active_month=loss_active_month,
            loss_reactive_month=loss_reactive_month,
        )
        logger.info("Заполнен лист 'Структура пр 2'")
    else:
        logger.warning("Лист 'Структура пр 2' не найден в шаблоне")

    # Заполняем газ по категориям использования
    if passport_data.gas_data:
        fill_gas_by_usage_categories(
            workbook,
            resources_data,
            sheet_name="Структура пр 2",
        )
        logger.info("Заполнены категории использования газа")

    # Заполняем электроэнергию по видам продукции
    if passport_data.electricity_by_product:
        electricity_data_dict = {}
        for product in passport_data.electricity_by_product:
            electricity_data_dict[product.product_name] = {
                "norm_kw": product.norm_kw,
                "fact_2022_kw": product.fact_2022_kw,
                "fact_2023_kw": product.fact_2023_kw,
                "fact_2024_kw": product.fact_2024_kw,
            }

        fill_electricity_by_product(
            workbook,
            electricity_data_dict,
            sheet_name="Структура пр 2",
        )
        logger.info(
            f"Заполнена электроэнергия по {len(passport_data.electricity_by_product)} видам продукции"
        )

    # Заполняем удельный расход газа
    if passport_data.gas_data and passport_data.gas_data.yearly.get(year):
        gas_year_m3 = passport_data.gas_data.yearly[year]

        fill_gas_specific_consumption(
            workbook,
            gas_year_m3=gas_year_m3,
            building_area_m2=passport_data.building_area_m2,
            production_units=passport_data.total_production_units,
            gas_norm_per_m2=passport_data.gas_norm_per_m2,
            gas_norm_per_unit=passport_data.gas_norm_per_unit,
            sheet_name="Удельный расход газа",
        )
        logger.info(f"Заполнен удельный расход газа: {gas_year_m3:.1f} м³/год")

    # Заполняем узлы учета
    if "01_Узлы учета" in workbook.sheetnames or "Узел учета " in workbook.sheetnames:
        sheet_name = (
            "01_Узлы учета" if "01_Узлы учета" in workbook.sheetnames else "Узел учета "
        )
        if nodes_json_path and nodes_json_path.exists():
            import json

            nodes_data = json.loads(nodes_json_path.read_text(encoding="utf-8"))
        else:
            from fill_energy_passport import load_default_nodes

            nodes_data = load_default_nodes()

        fill_nodes_sheet(workbook[sheet_name], nodes_data)
        logger.info("Заполнены узлы учета")

    # Заполняем оборудование
    if equipment_json_path and equipment_json_path.exists():
        import json

        equipment_data = json.loads(equipment_json_path.read_text(encoding="utf-8"))
        fill_equipment_sheet(workbook, equipment_data, sheet_name="Equipment")
        logger.info("Заполнено оборудование")

    # Заполняем ограждающие конструкции
    if envelope_json_path and envelope_json_path.exists():
        import json

        envelope_data = json.loads(envelope_json_path.read_text(encoding="utf-8"))
        fill_building_envelope_sheet(
            workbook,
            envelope_data,
            sheet_name="02_Исходные данные",
        )
        logger.info("Заполнены ограждающие конструкции")

    # Заполняем потери
    if loss_active_month or loss_reactive_month:
        fill_losses_sheet(
            workbook,
            loss_active_month,
            loss_reactive_month,
            transformer_power_kva=transformer_power_kva,
            hours_per_month=720.0,
        )
        logger.info("Заполнены потери")

    # Заполняем квартальные таблицы по производству продукции
    from .quarterly_production import (
        extract_monthly_production_from_aggregated,
        write_quarter_tables_to_sheet,
    )

    try:
        monthly_production = extract_monthly_production_from_aggregated(aggregated_data)

        if monthly_production:
            # Ищем или создаем лист для квартальных данных
            sheet_name = "Объемы по кварталам"
            if sheet_name in workbook.sheetnames:
                ws_quarters = workbook[sheet_name]
                # Очищаем существующий лист (оставляем заголовки если нужно)
                ws_quarters.delete_rows(1, ws_quarters.max_row)
            else:
                ws_quarters = workbook.create_sheet(sheet_name)

            # Записываем квартальные таблицы
            years = sorted(set(monthly_production.keys()))
            write_quarter_tables_to_sheet(
                ws_quarters,
                monthly_production,
                years=years,
                start_row=1,
                start_col=1,
            )
            logger.info(
                f"Заполнены квартальные таблицы по производству для {len(years)} лет"
            )
        else:
            logger.warning(
                "Нет данных по месячному производству для заполнения квартальных таблиц"
            )
    except Exception as e:
        logger.warning(f"Ошибка при заполнении квартальных таблиц: {e}")
        import traceback

        logger.debug(traceback.format_exc())

    # Сохраняем файл
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    logger.info(f"✅ Энергопаспорт сохранен: {output_path}")

    return output_path
