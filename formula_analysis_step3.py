#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Детальный анализ ошибки #REF! и восстановление формулы"""
import openpyxl
import json
from pathlib import Path

file_path = r"C:\Users\DELL\Downloads\Telegram Desktop\ЭНЕРГО_ПАСПОРТ_23102025.xlsm"

wb = openpyxl.load_workbook(file_path, data_only=False)

print("Analyzing #REF! error in Баланс sheet...")
sheet = wb['Баланс']
cell = sheet['AF13']

print(f"Cell AF13:")
print(f"  Current formula: {cell.value}")
print(f"  Data type: {cell.data_type}")

# Анализируем контекст вокруг ячейки
print("\nContext analysis:")
print(f"  Row 13 values:")
for col in range(26, 34):  # AF = column 32, смотрим несколько ячеек вокруг
    col_letter = openpyxl.utils.get_column_letter(col)
    val = sheet[f'{col_letter}13'].value
    print(f"    {col_letter}13: {val}")

print(f"\n  Column AF values (around row 13):")
for row in range(10, 16):
    val = sheet[f'AF{row}'].value
    data_type = sheet[f'AF{row}'].data_type
    print(f"    AF{row}: {val} (type: {data_type})")

# Ищем похожие формулы на листе
print("\nLooking for similar formulas on Баланс sheet...")
similar_formulas = []
for row in sheet.iter_rows(min_row=1, max_row=76, min_col=1, max_col=59):
    for cell in row:
        if cell.data_type == 'f' and cell.value:
            formula = str(cell.value)
            # Ищем формулы, которые могут быть похожи
            if 'SUM' in formula or '+' in formula or '-' in formula:
                similar_formulas.append({
                    'cell': cell.coordinate,
                    'formula': formula[:100]  # Первые 100 символов
                })

print(f"  Found {len(similar_formulas)} formulas with SUM or +")
if similar_formulas:
    print("  Sample formulas:")
    for sf in similar_formulas[:5]:
        print(f"    {sf['cell']}: {sf['formula'][:80]}")

# Анализ структуры листа "Баланс"
print("\nAnalyzing Баланс sheet structure...")
# Проверяем заголовки
print("  Headers (first 5 rows):")
for row_idx in range(1, 6):
    row_data = []
    for col_idx in range(1, 10):
        val = sheet.cell(row=row_idx, column=col_idx).value
        if val:
            row_data.append(str(val)[:30])
    print(f"    Row {row_idx}: {', '.join(row_data[:5])}")

# Сохраняем результат
result = {
    'ref_error_cell': 'AF13',
    'sheet': 'Баланс',
    'current_formula': str(cell.value),
    'context': {
        'row_13_values': {openpyxl.utils.get_column_letter(col): str(sheet.cell(row=13, column=col).value)[:50] 
                         for col in range(26, 34)},
        'col_AF_values': {f'AF{row}': str(sheet[f'AF{row}'].value)[:50] 
                         for row in range(10, 16)}
    },
    'similar_formulas': similar_formulas[:20],
    'recommendation': 'Need to analyze what AF13 should reference'
}

output_file = Path(__file__).parent / 'formula_analysis_step3_result.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\nStep 3 results saved to: {output_file}")

