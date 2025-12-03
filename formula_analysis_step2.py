#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Детальный анализ ячеек для восстановления формул"""
import openpyxl
import json
from pathlib import Path
from typing import List, Dict, Any

file_path = r"C:\Users\DELL\Downloads\Telegram Desktop\ЭНЕРГО_ПАСПОРТ_23102025.xlsm"

wb = openpyxl.load_workbook(file_path, data_only=False)

def analyze_cells_for_formulas(sheet, sheet_name: str) -> Dict[str, Any]:
    """Анализирует ячейки, которые должны быть формулами"""
    missing_formulas = []
    ref_errors = []
    potential_formulas = []
    
    # Ключевые слова, которые указывают на то, что ячейка должна быть формулой
    formula_indicators = [
        'итого', 'итог', 'сумма', 'sum', 'всего', 'total',
        'удельный', 'расход', 'эффективность', 'кпд', 'потери',
        'соответствует', 'превышает', 'норматив'
    ]
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            coord = cell.coordinate
            
            # Проверяем #REF! ошибки
            if cell.data_type == 'f' and cell.value:
                if '#REF!' in str(cell.value):
                    ref_errors.append({
                        'cell': coord,
                        'row': row_idx,
                        'col': col_idx,
                        'formula': str(cell.value),
                        'sheet': sheet_name
                    })
            
            # Ищем ячейки со значениями, которые могут быть формулами
            # (проверяем соседние ячейки на наличие подсказок)
            if cell.data_type != 'f' and cell.value:
                cell_value_str = str(cell.value).lower()
                
                # Проверяем заголовки строк/столбцов
                header_hint = False
                if row_idx <= 5:  # Первые строки часто содержат заголовки
                    for indicator in formula_indicators:
                        if indicator in cell_value_str:
                            header_hint = True
                            break
                
                # Если в строке есть подсказка о формуле, но сама ячейка не формула
                if header_hint or (isinstance(cell.value, (int, float)) and row_idx > 5):
                    # Проверяем, может ли это быть суммой или расчетом
                    potential_formulas.append({
                        'cell': coord,
                        'row': row_idx,
                        'col': col_idx,
                        'value': cell.value,
                        'sheet': sheet_name,
                        'reason': 'potential_calculation'
                    })
    
    return {
        'missing_formulas': missing_formulas,
        'ref_errors': ref_errors,
        'potential_formulas': potential_formulas[:50]  # Первые 50
    }

print("Analyzing cells for missing formulas...")
analysis = {}

# Анализируем ключевые листы
key_sheets = ['Структура пр 2 ', 'Баланс', 'Динамика ср', 'Расход  на ед.п']
for sheet_name in key_sheets:
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"  Analyzing {sheet_name}...")
        analysis[sheet_name] = analyze_cells_for_formulas(sheet, sheet_name)

# Сохраняем результат
result = {
    'file_path': file_path,
    'analysis': analysis,
    'summary': {
        'sheets_analyzed': len(analysis),
        'total_ref_errors': sum(len(a['ref_errors']) for a in analysis.values()),
        'total_potential_formulas': sum(len(a['potential_formulas']) for a in analysis.values())
    }
}

output_file = Path(__file__).parent / 'formula_analysis_step2_result.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\nStep 2 results saved to: {output_file}")
print(f"Found {result['summary']['total_ref_errors']} #REF! errors")
print(f"Found {result['summary']['total_potential_formulas']} potential formula cells")

