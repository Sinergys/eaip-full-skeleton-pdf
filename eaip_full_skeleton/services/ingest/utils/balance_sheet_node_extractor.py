"""
Модуль для извлечения данных потребления электроэнергии по узлам учёта из актов балансов.
Согласно рекомендациям экспертов:
- Software Engineer: "Использовать OCR для извлечения из PDF актов"
- ML Engineer: "Использовать Gemini Vision (95% confidence)"
- QA Engineer: "Тестировать на нескольких файлах сначала"
"""
import logging
import re
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime

from openpyxl import load_workbook
from .balance_sheet_detector import is_balance_sheet_file, get_balance_sheet_type

logger = logging.getLogger(__name__)

# Попытка импорта OCR модулей (опционально)
try:
    from .ocr_integration import process_pdf_with_ocr
    from .gemini_vision_ocr import extract_with_gemini_vision
    from pdf2image import convert_from_path
    HAS_OCR = True
except ImportError:
    HAS_OCR = False
    logger.warning("OCR модули не найдены. Обработка PDF будет ограничена.")


def extract_node_consumption_from_balance_sheet(
    file_path: str,
    batch_id: str,
    enterprise_id: int,
    raw_json: Optional[Dict[str, Any]] = None
) -> List[Dict[str, Any]]:
    """
    Извлекает данные потребления/производства/реализации по узлам учёта из акта баланса.
    
    Согласно рекомендациям экспертов:
    - Software Engineer: "Использовать OCR для извлечения из PDF актов"
    - ML Engineer: "Использовать Gemini Vision (95% confidence)"
    - QA Engineer: "Валидировать после получения"
    
    Args:
        file_path: Путь к файлу акта баланса
        batch_id: ID загрузки
        enterprise_id: ID предприятия
        raw_json: Распарсенные данные файла (опционально, для Excel/Word)
    
    Returns:
        Список валидированных данных по узлам в формате для импорта в БД
        Каждая запись содержит поле data_type: 'consumption', 'production' или 'realization'
    """
    file_path_obj = Path(file_path)
    if not file_path_obj.exists():
        logger.error(f"Файл не найден: {file_path}")
        return []
    
    file_ext = file_path_obj.suffix.lower()
    filename = file_path_obj.name.lower()
    
    # Определяем тип данных на основе имени файла
    # "Реализация" - это произведенная и реализованная (проданная) электроэнергия
    data_type = "consumption"  # По умолчанию - потребление
    if "реализация" in filename or "реализация нэс" in filename:
        data_type = "realization"
        logger.info(f"📊 Определен тип данных: 'realization' (произведенная и реализованная электроэнергия) для файла {file_path_obj.name}")
    elif "производство" in filename or "production" in filename:
        data_type = "production"
        logger.info(f"📊 Определен тип данных: 'production' (произведенная электроэнергия) для файла {file_path_obj.name}")
    
    # Определяем тип файла и обрабатываем соответственно
    raw_data = []
    if file_ext in [".xlsx", ".xls"]:
        raw_data = _extract_from_excel(file_path, batch_id, enterprise_id, raw_json, data_type)
    elif file_ext == ".pdf":
        raw_data = _extract_from_pdf(file_path, batch_id, enterprise_id, data_type)
    elif file_ext in [".docx", ".doc"]:
        raw_data = _extract_from_word(file_path, batch_id, enterprise_id, raw_json, data_type)
    else:
        logger.warning(f"Неподдерживаемый формат файла: {file_ext}")
        return []
    
    # Валидируем извлеченные данные перед возвратом
    validated_data = validate_node_consumption_data(raw_data, file_path)
    
    # Дедупликация: если один узел встречается несколько раз, объединяем данные
    deduplicated_data = _deduplicate_nodes(validated_data, file_path)
    
    # Логируем статистику извлечения
    if deduplicated_data:
        _log_extraction_statistics(deduplicated_data, file_path)
    
    return deduplicated_data


def _log_extraction_statistics(
    node_data: List[Dict[str, Any]],
    source_file: str
) -> None:
    """
    Логирует статистику извлеченных данных по узлам.
    
    Args:
        node_data: Список извлеченных данных по узлам
        source_file: Путь к исходному файлу
    """
    if not node_data:
        return
    
    total_nodes = len(node_data)
    nodes_with_active = sum(1 for n in node_data if n.get("active_energy_kwh") is not None)
    nodes_with_reactive = sum(1 for n in node_data if n.get("reactive_energy_kvarh") is not None)
    nodes_with_cost = sum(1 for n in node_data if n.get("cost_sum") is not None)
    
    # Группируем по периодам
    periods = {}
    for node in node_data:
        period = node.get("period", "unknown")
        periods[period] = periods.get(period, 0) + 1
    
    # Суммируем значения
    total_active = sum(n.get("active_energy_kwh") or 0 for n in node_data)
    total_reactive = sum(n.get("reactive_energy_kvarh") or 0 for n in node_data)
    total_cost = sum(n.get("cost_sum") or 0 for n in node_data)
    
    logger.info(
        f"📊 Статистика извлечения данных по узлам из {Path(source_file).name}:\n"
        f"  • Всего узлов: {total_nodes}\n"
        f"  • Узлов с активной энергией: {nodes_with_active} ({nodes_with_active/total_nodes*100:.1f}%)\n"
        f"  • Узлов с реактивной энергией: {nodes_with_reactive} ({nodes_with_reactive/total_nodes*100:.1f}%)\n"
        f"  • Узлов со стоимостью: {nodes_with_cost} ({nodes_with_cost/total_nodes*100:.1f}%)\n"
        f"  • Периодов: {len(periods)} ({', '.join(sorted(periods.keys()))})\n"
        f"  • Суммарная активная энергия: {total_active:,.2f} кВт·ч\n"
        f"  • Суммарная реактивная энергия: {total_reactive:,.2f} кВар·ч\n"
        f"  • Суммарная стоимость: {total_cost:,.2f} сум"
    )


def _deduplicate_nodes(
    node_data: List[Dict[str, Any]],
    source_file: str
) -> List[Dict[str, Any]]:
    """
    Удаляет дубликаты узлов, объединяя данные по узлу и периоду.
    
    Если один узел встречается несколько раз с одинаковым периодом,
    выбирается запись с наибольшим количеством данных.
    
    Args:
        node_data: Список данных по узлам
        source_file: Путь к исходному файлу (для логирования)
    
    Returns:
        Список данных без дубликатов
    """
    if not node_data:
        return []
    
    # Группируем по узлу и периоду
    nodes_map = {}  # (node_name, period) -> record
    
    for record in node_data:
        key = (record["node_name"], record["period"])
        
        if key not in nodes_map:
            nodes_map[key] = record
        else:
            # Объединяем данные: выбираем запись с большим количеством данных
            existing = nodes_map[key]
            existing_data_count = sum(
                1 for v in [existing.get("active_energy_kwh"), 
                           existing.get("reactive_energy_kvarh"),
                           existing.get("cost_sum")]
                if v is not None
            )
            new_data_count = sum(
                1 for v in [record.get("active_energy_kwh"),
                           record.get("reactive_energy_kvarh"),
                           record.get("cost_sum")]
                if v is not None
            )
            
            # Если новая запись содержит больше данных, заменяем
            if new_data_count > existing_data_count:
                nodes_map[key] = record
                logger.debug(
                    f"Объединение дубликата узла '{record['node_name']}' "
                    f"период '{record['period']}': выбрана запись с большим количеством данных"
                )
            else:
                # Дополняем существующую запись недостающими данными
                if existing.get("active_energy_kwh") is None and record.get("active_energy_kwh") is not None:
                    existing["active_energy_kwh"] = record["active_energy_kwh"]
                if existing.get("reactive_energy_kvarh") is None and record.get("reactive_energy_kvarh") is not None:
                    existing["reactive_energy_kvarh"] = record["reactive_energy_kvarh"]
                if existing.get("cost_sum") is None and record.get("cost_sum") is not None:
                    existing["cost_sum"] = record["cost_sum"]
    
    deduplicated = list(nodes_map.values())
    
    if len(deduplicated) < len(node_data):
        logger.info(
            f"🔗 Дедупликация узлов из {Path(source_file).name}: "
            f"{len(node_data)} записей → {len(deduplicated)} уникальных узлов"
        )
    
    return deduplicated


def validate_node_consumption_data(
    node_data: List[Dict[str, Any]],
    source_file: str
) -> List[Dict[str, Any]]:
    """
    Валидирует данные потребления по узлам учёта.
    
    Согласно рекомендации QA Engineer: "Валидировать после получения"
    
    Args:
        node_data: Список извлеченных данных по узлам
        source_file: Путь к исходному файлу (для логирования)
    
    Returns:
        Список валидированных данных (невалидные записи удаляются)
    """
    if not node_data:
        return []
    
    validated = []
    errors = []
    warnings = []
    
    for idx, record in enumerate(node_data):
        record_errors = []
        record_warnings = []
        
        # Проверка обязательных полей
        if not record.get("node_name"):
            record_errors.append("Отсутствует название узла учёта")
            continue
        
        node_name = str(record["node_name"]).strip()
        if not node_name or len(node_name) < 2:
            record_errors.append(f"Некорректное название узла: '{node_name}'")
            continue
        
        # Проверка периода
        period = record.get("period", "unknown")
        if period == "unknown":
            record_warnings.append("Период не определен, используется 'unknown'")
        
        # Проверка числовых значений
        active_energy = record.get("active_energy_kwh")
        reactive_energy = record.get("reactive_energy_kvarh")
        cost = record.get("cost_sum")
        
        # Проверка на отрицательные значения
        if active_energy is not None and active_energy < 0:
            record_warnings.append(f"Отрицательное значение активной энергии: {active_energy}")
            active_energy = None  # Устанавливаем в None для некорректных значений
        
        if reactive_energy is not None and reactive_energy < 0:
            record_warnings.append(f"Отрицательное значение реактивной энергии: {reactive_energy}")
            reactive_energy = None
        
        if cost is not None and cost < 0:
            record_warnings.append(f"Отрицательное значение стоимости: {cost}")
            cost = None
        
        # Проверка на нереалистично большие значения (предупреждение, не ошибка)
        if active_energy is not None and active_energy > 100000000:  # 100 млн кВт·ч
            record_warnings.append(f"Подозрительно большое значение активной энергии: {active_energy}")
        
        if reactive_energy is not None and reactive_energy > 100000000:  # 100 млн кВар·ч
            record_warnings.append(f"Подозрительно большое значение реактивной энергии: {reactive_energy}")
        
        # Проверка на наличие хотя бы одного значения потребления
        if active_energy is None and reactive_energy is None and cost is None:
            record_warnings.append("Все значения потребления отсутствуют")
            # Не удаляем запись, но логируем предупреждение
        
        # Если есть ошибки, пропускаем запись
        if record_errors:
            errors.extend([f"Запись {idx + 1} (узел '{node_name}'): {err}" for err in record_errors])
            continue
        
        # Обновляем запись с исправленными значениями
        validated_record = {
            "node_name": node_name,
            "period": period,
            "active_energy_kwh": active_energy,
            "reactive_energy_kvarh": reactive_energy,
            "cost_sum": cost,
            "data_type": record.get("data_type", "consumption"),
            "data_json": record.get("data_json", {})
        }
        
        # Добавляем информацию о валидации в data_json
        if record_warnings:
            validated_record["data_json"]["validation_warnings"] = record_warnings
        
        validated.append(validated_record)
        
        if record_warnings:
            warnings.extend([f"Запись {idx + 1} (узел '{node_name}'): {warn}" for warn in record_warnings])
    
    # Логируем результаты валидации
    if errors:
        logger.warning(
            f"⚠️ Валидация данных по узлам из {Path(source_file).name}: "
            f"найдено {len(errors)} ошибок, {len(errors)} записей пропущено"
        )
        for error in errors[:5]:  # Логируем первые 5 ошибок
            logger.debug(f"  - {error}")
        if len(errors) > 5:
            logger.debug(f"  ... и ещё {len(errors) - 5} ошибок")
    
    if warnings:
        logger.info(
            f"ℹ️ Валидация данных по узлам из {Path(source_file).name}: "
            f"найдено {len(warnings)} предупреждений"
        )
        for warning in warnings[:5]:  # Логируем первые 5 предупреждений
            logger.debug(f"  - {warning}")
        if len(warnings) > 5:
            logger.debug(f"  ... и ещё {len(warnings) - 5} предупреждений")
    
    if validated:
        logger.info(
            f"✅ Валидация завершена: {len(validated)} из {len(node_data)} записей прошли валидацию"
        )
    
    return validated


def _extract_from_excel(
    file_path: str,
    batch_id: str,
    enterprise_id: int,
    raw_json: Optional[Dict[str, Any]] = None,
    data_type: str = "consumption"
) -> List[Dict[str, Any]]:
    """
    Извлекает данные по узлам учёта из Excel файла акта баланса.
    
    Для файлов "Реализация" обрабатывает оба листа:
    - Детальное по потребителям
    - Общее за год
    
    Args:
        file_path: Путь к Excel файлу
        batch_id: ID загрузки
        enterprise_id: ID предприятия
        raw_json: Распарсенные данные файла (опционально)
        data_type: Тип данных ('consumption', 'production', 'realization')
    
    Returns:
        Список данных по узлам (каждая запись содержит поле data_type)
    """
    try:
        filename = Path(file_path).name
        logger.info(f"📊 Извлечение данных по узлам из Excel: {filename} (тип: {data_type})")
        
        workbook = load_workbook(file_path, data_only=True)
        node_consumption_data = []
        
        # Для файлов "Реализация" обрабатываем все листы
        is_realization_file = "реализация" in filename.lower()
        
        logger.info(f"  Найдено листов в файле: {len(workbook.sheetnames)}")
        for sheet_name in workbook.sheetnames:
            logger.debug(f"  Проверка листа: {sheet_name}")
            
            sheet = workbook[sheet_name]
            sheet_name_lower = sheet_name.lower()
            
            # Для файлов "Реализация" проверяем все листы более тщательно
            if is_realization_file:
                # Проверяем, содержит ли лист данные (детальные или общие)
                if _is_realization_sheet(sheet_name_lower, sheet):
                    logger.info(f"  ✅ Найден лист с данными реализации: {sheet_name}")
                    sheet_data = _parse_node_sheet(sheet, sheet_name, data_type)
                    if sheet_data:
                        logger.info(f"    Извлечено {len(sheet_data)} записей из листа {sheet_name}")
                        node_consumption_data.extend(sheet_data)
                    else:
                        logger.warning(f"    Не удалось извлечь данные из листа {sheet_name}")
                else:
                    logger.debug(f"  ⏭️ Лист {sheet_name} пропущен (не содержит данных по узлам)")
            else:
                # Для других файлов используем стандартную проверку
                if not _is_node_sheet(sheet_name_lower, sheet):
                    continue
                
                logger.info(f"  Найден лист с данными по узлам: {sheet_name}")
                # Передаем filename для извлечения периода
                sheet_data = _parse_node_sheet(sheet, sheet_name, data_type, filename=filename)
                node_consumption_data.extend(sheet_data)
        
        workbook.close()
        
        logger.info(
            f"✅ Извлечено {len(node_consumption_data)} записей по узлам (тип: {data_type}) "
            f"из Excel файла {filename} (обработано листов: {len(workbook.sheetnames)})"
        )
        
        return node_consumption_data
        
    except Exception as e:
        logger.error(f"Ошибка извлечения данных из Excel файла {file_path}: {e}", exc_info=True)
        return []


def _extract_from_pdf(
    file_path: str,
    batch_id: str,
    enterprise_id: int,
    data_type: str = "consumption"
) -> List[Dict[str, Any]]:
    """
    Извлекает данные по узлам учёта из PDF файла акта баланса через OCR.
    
    Согласно рекомендациям:
    - Software Engineer: "Использовать OCR для извлечения из PDF актов"
    - ML Engineer: "Использовать Gemini Vision (95% confidence)"
    
    Args:
        file_path: Путь к PDF файлу
        batch_id: ID загрузки
        enterprise_id: ID предприятия
        data_type: Тип данных ('consumption', 'production', 'realization')
    
    Returns:
        Список данных по узлам (каждая запись содержит поле data_type)
    """
    if not HAS_OCR:
        logger.warning("OCR модули не доступны. Невозможно обработать PDF файл.")
        return []
    
    try:
        logger.info(f"🔍 Извлечение данных по узлам из PDF через OCR: {Path(file_path).name} (тип: {data_type})")
        
        # Обрабатываем PDF через OCR
        ocr_result = process_pdf_with_ocr(
            pdf_path=file_path,
            batch_id=batch_id,
            debug_dir=None,
            save_debug=False
        )
        
        if not ocr_result:
            logger.warning(f"Не удалось обработать PDF через OCR: {file_path}")
            return []
        
        # Извлекаем данные по узлам из результатов OCR
        node_consumption_data = _parse_ocr_result_for_nodes(ocr_result, file_path, data_type)
        
        logger.info(
            f"✅ Извлечено {len(node_consumption_data)} записей по узлам (тип: {data_type}) "
            f"из PDF файла {Path(file_path).name} через OCR"
        )
        
        return node_consumption_data
        
    except Exception as e:
        logger.error(f"Ошибка извлечения данных из PDF файла {file_path}: {e}")
        return []


def _extract_from_word(
    file_path: str,
    batch_id: str,
    enterprise_id: int,
    raw_json: Optional[Dict[str, Any]] = None,
    data_type: str = "consumption"
) -> List[Dict[str, Any]]:
    """
    Извлекает данные по узлам учёта из Word файла акта баланса.
    
    Args:
        file_path: Путь к Word файлу
        batch_id: ID загрузки
        enterprise_id: ID предприятия
        raw_json: Распарсенные данные файла (опционально)
        data_type: Тип данных ('consumption', 'production', 'realization')
    
    Returns:
        Список данных по узлам (каждая запись содержит поле data_type)
    """
    try:
        logger.info(f"📄 Извлечение данных по узлам из Word: {Path(file_path).name} (тип: {data_type})")
        
        if not raw_json:
            logger.warning("raw_json не предоставлен для Word файла")
            return []
        
        # Извлекаем данные из таблиц Word
        node_consumption_data = []
        tables = raw_json.get("tables", [])
        
        for table_idx, table in enumerate(tables):
            if _is_node_table(table):
                logger.info(f"  Найдена таблица с данными по узлам: таблица {table_idx + 1}")
                table_data = _parse_node_table_from_word(table, data_type)
                node_consumption_data.extend(table_data)
        
        logger.info(
            f"✅ Извлечено {len(node_consumption_data)} записей по узлам (тип: {data_type}) "
            f"из Word файла {Path(file_path).name}"
        )
        
        return node_consumption_data
        
    except Exception as e:
        logger.error(f"Ошибка извлечения данных из Word файла {file_path}: {e}")
        return []


def _is_node_sheet(sheet_name: str, sheet) -> bool:
    """Проверяет, содержит ли лист данные по узлам учёта."""
    # Проверяем название листа
    node_keywords = ["узел", "тп", "подстанция", "счетчик", "баланс", "акт"]
    if any(keyword in sheet_name for keyword in node_keywords):
        return True
    
    # Проверяем первые строки листа на наличие признаков таблицы узлов
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx > 5:  # Проверяем только первые 5 строк
            break
        
        row_text = " ".join(str(cell).lower() for cell in row if cell)
        if any(keyword in row_text for keyword in node_keywords):
            return True
    
    return False


def _is_realization_sheet(sheet_name: str, sheet) -> bool:
    """
    Проверяет, содержит ли лист данные реализации (детальные по потребителям или общие за год).
    
    Для файлов "Реализация" обрабатывает:
    - Листы с детальными данными по потребителям
    - Листы с общими данными за год
    """
    sheet_name_lower = sheet_name.lower()
    
    # Ключевые слова для листов реализации
    realization_keywords = [
        "потребитель", "потребители", "детальн", "детально",
        "общее", "общий", "год", "годов", "итого", "итог",
        "узел", "тп", "подстанция", "счетчик", "баланс",
        "активная", "реактивная", "энергия", "квт", "квар"
    ]
    
    # Проверяем название листа
    if any(keyword in sheet_name_lower for keyword in realization_keywords):
        return True
    
    # Проверяем содержимое листа более тщательно (первые 20 строк для файлов реализации)
    rows_checked = 0
    has_node_keywords = False
    has_energy_keywords = False
    has_numeric_data = False
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx > 20:  # Проверяем больше строк для файлов реализации
            break
        
        rows_checked += 1
        row_text = " ".join(str(cell).lower() for cell in row if cell)
        
        # Проверяем наличие ключевых слов узлов
        node_keywords = ["узел", "тп", "подстанция", "счетчик", "потребитель", "наименование", "название"]
        if any(keyword in row_text for keyword in node_keywords):
            has_node_keywords = True
        
        # Проверяем наличие ключевых слов энергии
        energy_keywords = ["активная", "реактивная", "квт", "квар", "энергия", "p", "q", "стоимость", "сумма"]
        if any(keyword in row_text for keyword in energy_keywords):
            has_energy_keywords = True
        
        # Проверяем наличие числовых данных (признак таблицы с данными)
        if any(isinstance(cell, (int, float)) and cell > 0 for cell in row if cell is not None):
            has_numeric_data = True
    
    # Лист считается листом реализации, если:
    # 1. Есть ключевые слова узлов И энергии (детальные данные)
    # 2. ИЛИ есть числовые данные И (ключевые слова узлов ИЛИ энергии) (общие данные)
    if (has_node_keywords and has_energy_keywords) or (has_numeric_data and (has_node_keywords or has_energy_keywords)):
        return True
    
    return False


def _parse_node_sheet(sheet, sheet_name: str, data_type: str = "consumption", filename: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Парсит лист Excel и извлекает данные по узлам учёта.
    
    Для файлов "Реализация" обрабатывает:
    - Детальные данные по потребителям
    - Общие данные за год
    
    Args:
        sheet: Лист Excel
        sheet_name: Название листа
        data_type: Тип данных ('consumption', 'production', 'realization')
    
    Returns:
        Список данных по узлам (каждая запись содержит поле data_type)
    """
    node_data = []
    rows = list(sheet.iter_rows(values_only=True))
    
    if not rows:
        logger.debug(f"Лист {sheet_name} пуст")
        return []
    
    # Определяем тип листа (детальный по потребителям или общий за год)
    sheet_name_lower = sheet_name.lower()
    is_detail_sheet = any(keyword in sheet_name_lower for keyword in ["потребитель", "детальн", "детально"])
    is_summary_sheet = any(keyword in sheet_name_lower for keyword in ["общее", "общий", "год", "годов", "итого", "итог"])
    
    # Ищем заголовки таблицы
    header_row_idx = _find_header_row(rows)
    if header_row_idx is None:
        logger.warning(f"Не найдена строка заголовков в листе {sheet_name}")
        return []
    
    headers = rows[header_row_idx]
    
    # Определяем индексы колонок
    # Для листов "Баланс" передаем дополнительные данные для анализа структуры
    rows_for_analysis = rows[header_row_idx + 1:header_row_idx + 6] if len(rows) > header_row_idx + 1 else []
    col_indices = _find_column_indices(headers, sheet_name=sheet_name, rows_data=rows_for_analysis)
    
    if not col_indices.get("node_name"):
        logger.warning(f"Не найдена колонка с названиями узлов в листе {sheet_name}")
        # Для листов "Баланс" пробуем использовать первую колонку как fallback
        if "баланс" in sheet_name.lower():
            logger.info(f"Для листа 'Баланс' используем первую колонку как название узла (fallback)")
            col_indices["node_name"] = 0
        else:
            return []
    
    # Извлекаем период из названия листа или файла
    # Передаем filename для извлечения года, если он не найден в названии листа
    period = _extract_period_from_text(sheet_name, filename=filename) if filename else _extract_period_from_text(sheet_name)
    
    # Парсим строки данных
    for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if not any(cell for cell in row):  # Пропускаем пустые строки
            continue
        
        node_name = _get_cell_value(row, col_indices.get("node_name"))
        if not node_name or not str(node_name).strip():
            continue
        
        # Пропускаем строки с итогами (если это не общий лист)
        node_name_str = str(node_name).strip().lower()
        if not is_summary_sheet and any(keyword in node_name_str for keyword in ["итого", "итог", "всего", "сумма"]):
            continue
        
        # Извлекаем значения
        active_energy = _parse_float_value(_get_cell_value(row, col_indices.get("active_energy")))
        reactive_energy = _parse_float_value(_get_cell_value(row, col_indices.get("reactive_energy")))
        cost = _parse_float_value(_get_cell_value(row, col_indices.get("cost")))
        
        # Определяем тип листа для data_json
        sheet_type = "unknown"
        if is_detail_sheet:
            sheet_type = "detail_by_consumers"
        elif is_summary_sheet:
            sheet_type = "summary_by_year"
        
        # Создаём запись
        node_record = {
            "node_name": str(node_name).strip(),
            "period": period or "unknown",
            "active_energy_kwh": active_energy,
            "reactive_energy_kvarh": reactive_energy,
            "cost_sum": cost,
            "data_type": data_type,
            "data_json": {
                "source_sheet": sheet_name,
                "sheet_type": sheet_type,
                "row_number": row_idx,
            }
        }
        
        node_data.append(node_record)
    
    logger.debug(
        f"Извлечено {len(node_data)} записей из листа {sheet_name} "
        f"(тип листа: {sheet_type if node_data else 'unknown'})"
    )
    
    return node_data


def _find_header_row(rows: List[tuple]) -> Optional[int]:
    """
    Находит строку с заголовками таблицы.
    
    Использует более гибкий алгоритм поиска заголовков:
    1. Ищет строку с ключевыми словами узлов и энергии
    2. Проверяет наличие числовых данных в следующих строках
    3. Учитывает различные форматы таблиц
    """
    node_keywords = ["узел", "тп", "подстанция", "наименование", "название", "счетчик", "счётчик"]
    energy_keywords = ["активная", "реактивная", "квт", "квар", "энергия", "p", "q"]
    
    # Проверяем первые 15 строк (увеличено для учета возможных пустых строк)
    for idx, row in enumerate(rows[:15]):
        if not any(cell for cell in row):  # Пропускаем полностью пустые строки
            continue
        
        row_text = " ".join(str(cell).lower() for cell in row if cell)
        
        # Проверяем наличие ключевых слов узлов и энергии
        has_node_keyword = any(keyword in row_text for keyword in node_keywords)
        has_energy_keyword = any(keyword in row_text for keyword in energy_keywords)
        
        # Если найдены оба типа ключевых слов - это заголовок
        if has_node_keyword and has_energy_keyword:
            # Дополнительная проверка: следующая строка должна содержать данные
            if idx + 1 < len(rows):
                next_row = rows[idx + 1]
                # Проверяем, есть ли в следующей строке хотя бы одно значение
                if any(cell for cell in next_row):
                    return idx
        
        # Альтернативный вариант: если есть только ключевые слова узлов,
        # но в следующих строках есть числовые данные - тоже может быть заголовком
        if has_node_keyword and not has_energy_keyword:
            # Проверяем следующие 2-3 строки на наличие числовых данных
            for check_idx in range(idx + 1, min(idx + 4, len(rows))):
                check_row = rows[check_idx]
                # Проверяем, есть ли числовые значения (не только текст)
                has_numbers = any(
                    isinstance(cell, (int, float)) and cell is not None
                    for cell in check_row
                    if cell is not None
                )
                if has_numbers:
                    return idx
    
    # Если не нашли по ключевым словам, пробуем найти строку с максимальным количеством
    # непустых ячеек в первых строках (часто это заголовок)
    max_cells = 0
    header_candidate = None
    for idx, row in enumerate(rows[:10]):
        non_empty_count = sum(1 for cell in row if cell and str(cell).strip())
        if non_empty_count > max_cells and non_empty_count >= 3:  # Минимум 3 колонки
            max_cells = non_empty_count
            header_candidate = idx
    
    return header_candidate


def _find_column_indices(headers: tuple, sheet_name: str = "", rows_data: Optional[List[tuple]] = None) -> Dict[str, Optional[int]]:
    """
    Находит индексы колонок с нужными данными.
    
    Согласно рекомендациям экспертов:
    - QA Engineer: "Тестировать на нескольких файлах сначала"
    - Software Engineer: "Использовать гибкий алгоритм для разных структур"
    
    Использует гибкий алгоритм поиска с приоритетами и альтернативными вариантами.
    Для листов "Баланс" применяет специальную логику.
    
    Args:
        headers: Кортеж заголовков
        sheet_name: Название листа (для специальной обработки листов "Баланс")
        rows_data: Первые несколько строк данных (для анализа структуры)
    """
    indices = {
        "node_name": None,
        "active_energy": None,
        "reactive_energy": None,
        "cost": None,
    }
    
    # Если headers пустой или слишком короткий, пробуем использовать первую колонку для узлов
    if not headers or len(headers) < 2:
        # В очень простых таблицах первая колонка может быть названием узла
        indices["node_name"] = 0
        return indices
    
    sheet_name_lower = sheet_name.lower() if sheet_name else ""
    is_balance_sheet = "баланс" in sheet_name_lower
    
    # Для листов "Баланс" используем более широкий набор ключевых слов
    node_keywords_standard = [
        "узел", "тп", "подстанция", "наименование", "название",
        "наименование узла", "название узла", "узел учёта", "узел учета"
    ]
    
    node_keywords_balance = node_keywords_standard + [
        "потребитель", "потребители", "наименование потребителя",
        "название потребителя", "объект", "объекты", "наименование объекта"
    ]
    
    node_keywords = node_keywords_balance if is_balance_sheet else node_keywords_standard
    
    for idx, header in enumerate(headers):
        if not header:
            continue
        
        header_lower = str(header).lower().strip()
        
        # Название узла (приоритет 1: точные совпадения, приоритет 2: частичные)
        if not indices["node_name"]:
            # Точные совпадения
            if any(keyword in header_lower for keyword in node_keywords):
                indices["node_name"] = idx
            # Частичные совпадения (если не нашли точных)
            elif idx == 0 and any(keyword in header_lower for keyword in ["наимен", "назван", "узел", "потребит", "объект"]):
                indices["node_name"] = idx
        
        # Активная энергия (расширенный список для листов "Баланс")
        if not indices["active_energy"]:
            active_keywords = [
                "активная", "квт", "квт·ч", "квтч", "active", "p",
                "активная энергия", "активная мощность", "w", "квт.ч"
            ]
            if is_balance_sheet:
                active_keywords.extend([
                    "активная мощность", "активная энергия, квт", "активная, квт",
                    "p, квт", "активная (квт)", "активная квт/ч"
                ])
            if any(keyword in header_lower for keyword in active_keywords):
                indices["active_energy"] = idx
        
        # Реактивная энергия (расширенный список для листов "Баланс")
        if not indices["reactive_energy"]:
            reactive_keywords = [
                "реактивная", "квар", "квар·ч", "кварч", "reactive", "q",
                "реактивная энергия", "реактивная мощность", "var", "квар.ч"
            ]
            if is_balance_sheet:
                reactive_keywords.extend([
                    "реактивная мощность", "реактивная энергия, квар", "реактивная, квар",
                    "q, квар", "реактивная (квар)", "реактивная квар/ч"
                ])
            if any(keyword in header_lower for keyword in reactive_keywords):
                indices["reactive_energy"] = idx
        
        # Стоимость (расширенный список для листов "Баланс")
        if not indices["cost"]:
            cost_keywords = [
                "стоимость", "сум", "cost", "цена", "сумма",
                "стоимость, сум", "стоимость (сум)", "цена, сум"
            ]
            if is_balance_sheet:
                cost_keywords.extend([
                    "стоимость, сум.", "стоимость (сум.)", "сумма, сум",
                    "стоимость в сумах", "цена в сумах", "сумма к оплате"
                ])
            if any(keyword in header_lower for keyword in cost_keywords):
                indices["cost"] = idx
    
    # Если не нашли колонку с названием узла, пробуем использовать первую непустую колонку
    if not indices["node_name"]:
        for idx, header in enumerate(headers):
            if header and str(header).strip():
                # Проверяем, что это не числовое значение (это может быть название узла)
                header_str = str(header).strip()
                if not header_str.replace(".", "").replace(",", "").replace("-", "").isdigit():
                    indices["node_name"] = idx
                    logger.debug(f"Использована первая непустая колонка {idx} как название узла: {header_str}")
                    break
    
    # Для листов "Баланс": если не нашли колонки, пробуем проанализировать данные
    if is_balance_sheet and rows_data and len(rows_data) > 0:
        # Анализируем первые несколько строк данных для определения структуры
        indices = _analyze_balance_sheet_structure(headers, rows_data, indices)
    
    return indices


def _analyze_balance_sheet_structure(
    headers: tuple,
    rows_data: List[tuple],
    current_indices: Dict[str, Optional[int]]
) -> Dict[str, Optional[int]]:
    """
    Анализирует структуру листа "Баланс" для определения колонок.
    
    Согласно рекомендации QA Engineer: "Тестировать на нескольких файлах сначала"
    
    Args:
        headers: Заголовки таблицы
        rows_data: Первые несколько строк данных
        current_indices: Текущие найденные индексы
    
    Returns:
        Обновленные индексы колонок
    """
    indices = current_indices.copy()
    
    # Если уже нашли все колонки, возвращаем как есть
    if all(indices.values()):
        return indices
    
    # Анализируем первые 5 строк данных
    sample_rows = rows_data[:5] if len(rows_data) > 5 else rows_data
    
    if not sample_rows:
        return indices
    
    # Ищем колонку с названиями узлов: первая колонка с текстовыми значениями
    if not indices["node_name"]:
        for col_idx in range(min(len(headers), 10)):  # Проверяем первые 10 колонок
            # Проверяем, что в этой колонке есть текстовые значения (не только числа)
            has_text = False
            has_numbers = False
            
            for row in sample_rows:
                if col_idx < len(row):
                    cell = row[col_idx]
                    if cell is not None:
                        cell_str = str(cell).strip()
                        # Проверяем, что это текст (не только число)
                        if cell_str and not cell_str.replace(".", "").replace(",", "").replace("-", "").isdigit():
                            has_text = True
                        elif isinstance(cell, (int, float)):
                            has_numbers = True
            
            # Если в колонке есть текст и мало чисел - это может быть название узла
            if has_text and not has_numbers:
                indices["node_name"] = col_idx
                logger.debug(f"Найдена колонка с названиями узлов (анализ данных): колонка {col_idx}")
                break
    
    # Ищем колонки с энергией: колонки с числовыми значениями
    if not indices["active_energy"] or not indices["reactive_energy"]:
        numeric_columns = []
        
        for col_idx in range(len(headers)):
            if col_idx == indices.get("node_name"):
                continue  # Пропускаем колонку с названиями
            
            # Проверяем, есть ли в колонке числовые значения
            has_numbers = False
            numeric_count = 0
            
            for row in sample_rows:
                if col_idx < len(row):
                    cell = row[col_idx]
                    if cell is not None and isinstance(cell, (int, float)) and cell > 0:
                        has_numbers = True
                        numeric_count += 1
            
            if has_numbers and numeric_count >= 2:  # Хотя бы 2 числовых значения
                numeric_columns.append((col_idx, numeric_count))
        
        # Сортируем по количеству числовых значений
        numeric_columns.sort(key=lambda x: x[1], reverse=True)
        
        # Первая числовая колонка - активная энергия, вторая - реактивная
        if numeric_columns and not indices["active_energy"]:
            indices["active_energy"] = numeric_columns[0][0]
            logger.debug(f"Найдена колонка с активной энергией (анализ данных): колонка {numeric_columns[0][0]}")
        
        if len(numeric_columns) > 1 and not indices["reactive_energy"]:
            indices["reactive_energy"] = numeric_columns[1][0]
            logger.debug(f"Найдена колонка с реактивной энергией (анализ данных): колонка {numeric_columns[1][0]}")
        
        # Третья числовая колонка может быть стоимостью
        if len(numeric_columns) > 2 and not indices["cost"]:
            indices["cost"] = numeric_columns[2][0]
            logger.debug(f"Найдена колонка со стоимостью (анализ данных): колонка {numeric_columns[2][0]}")
    
    return indices


def _get_cell_value(row: tuple, col_idx: Optional[int]) -> Any:
    """Получает значение ячейки по индексу колонки."""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


def _parse_float_value(value: Any) -> Optional[float]:
    """Парсит значение в float."""
    if value is None:
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    try:
        # Удаляем пробелы и заменяем запятую на точку
        cleaned = str(value).strip().replace(" ", "").replace(",", ".")
        if not cleaned:
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _extract_period_from_text(text: str, filename: Optional[str] = None) -> Optional[str]:
    """
    Извлекает период из текста (название листа/файла/заголовков).
    
    Поддерживает форматы:
    - "2022-Q1", "2022 Q1", "2022 1 квартал"
    - "Январь 2022", "01.2022", "2022-01"
    - "Реал 04" (04 = апрель, год из имени файла)
    - "2022" (год без квартала)
    
    Args:
        text: Текст для извлечения периода (название листа, заголовок и т.д.)
        filename: Имя файла (для извлечения года, если не найден в text)
    """
    if not text:
        # Если текст пуст, пытаемся извлечь год из имени файла
        if filename:
            year_match = re.search(r'\b(20\d{2})\b', filename)
            if year_match:
                return f"{year_match.group(1)}"
        return None
    
    text_lower = text.lower()
    
    # Ищем год (2000-2099)
    year_match = re.search(r'\b(20\d{2})\b', text)
    year = year_match.group(1) if year_match else None
    
    # Если год не найден в тексте, пытаемся извлечь из имени файла
    if not year and filename:
        year_match = re.search(r'\b(20\d{2})\b', filename)
        year = year_match.group(1) if year_match else None
    
    if not year:
        return None
    
    # Ищем квартал (приоритет 1)
    quarter_patterns = [
        r'\bq[1-4]\b',  # Q1, Q2, Q3, Q4
        r'\b([1-4])\s*квартал\b',  # 1 квартал, 2 квартал и т.д.
        r'\b([i]{1,4})\s*квартал\b',  # I квартал, II квартал и т.д.
    ]
    
    for pattern in quarter_patterns:
        quarter_match = re.search(pattern, text_lower)
        if quarter_match:
            quarter_text = quarter_match.group(0).lower()
            if "q1" in quarter_text or "1 квартал" in quarter_text or quarter_text.startswith("i "):
                quarter = "Q1"
            elif "q2" in quarter_text or "2 квартал" in quarter_text or quarter_text.startswith("ii "):
                quarter = "Q2"
            elif "q3" in quarter_text or "3 квартал" in quarter_text or quarter_text.startswith("iii "):
                quarter = "Q3"
            elif "q4" in quarter_text or "4 квартал" in quarter_text or quarter_text.startswith("iv "):
                quarter = "Q4"
            else:
                continue
            
            return f"{year}-{quarter}"
    
    # Ищем месяц (приоритет 2) - определяем квартал по месяцу
    months = {
        "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
        "май": 5, "июнь": 6, "июль": 7, "август": 8,
        "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
        "янв": 1, "фев": 2, "мар": 3, "апр": 4,
        "июн": 6, "июл": 7, "авг": 8, "сен": 9,
        "окт": 10, "ноя": 11, "дек": 12
    }
    
    for month_name, month_num in months.items():
        if month_name in text_lower:
            # Определяем квартал по месяцу
            quarter = (month_num - 1) // 3 + 1
            return f"{year}-Q{quarter}"
    
    # Ищем формат "Реал 04", "Баланс 00" и т.д. (номер месяца)
    # "04" = апрель, "00" может быть общим листом за год
    month_number_match = re.search(r'\b(0[1-9]|1[0-2]|00)\b', text)
    if month_number_match:
        month_str = month_number_match.group(1)
        if month_str == "00":
            # "00" обычно означает общий лист за год
            return f"{year}"
        else:
            month_num = int(month_str)
            # Определяем квартал по месяцу
            quarter = (month_num - 1) // 3 + 1
            return f"{year}-Q{quarter}"
    
    # Ищем формат даты MM.YYYY или YYYY-MM
    date_patterns = [
        r'\b(0?[1-9]|1[0-2])\.(20\d{2})\b',  # 01.2022, 1.2022
        r'\b(20\d{2})-(0?[1-9]|1[0-2])\b',  # 2022-01, 2022-1
    ]
    
    for pattern in date_patterns:
        date_match = re.search(pattern, text)
        if date_match:
            if len(date_match.groups()) == 2:
                if date_match.group(1).isdigit() and len(date_match.group(1)) <= 2:
                    # Формат MM.YYYY
                    month = int(date_match.group(1))
                    year_from_date = date_match.group(2)
                else:
                    # Формат YYYY-MM
                    year_from_date = date_match.group(1)
                    month = int(date_match.group(2))
                
                quarter = (month - 1) // 3 + 1
                return f"{year_from_date}-Q{quarter}"
    
    # Если найден только год, возвращаем с unknown кварталом
    return f"{year}-unknown"


def _is_node_table(table: Dict[str, Any]) -> bool:
    """Проверяет, содержит ли таблица данные по узлам учёта."""
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    
    if not headers and not rows:
        return False
    
    # Проверяем заголовки
    header_text = " ".join(str(h).lower() for h in headers if h)
    node_keywords = ["узел", "тп", "подстанция", "счетчик"]
    energy_keywords = ["активная", "реактивная", "квт", "квар"]
    
    has_node_keyword = any(keyword in header_text for keyword in node_keywords)
    has_energy_keyword = any(keyword in header_text for keyword in energy_keywords)
    
    return has_node_keyword and has_energy_keyword


def _parse_node_table_from_word(table: Dict[str, Any], data_type: str = "consumption") -> List[Dict[str, Any]]:
    """Парсит таблицу Word и извлекает данные по узлам учёта."""
    node_data = []
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    
    if not headers or not rows:
        return []
    
    # Определяем индексы колонок
    col_indices = _find_column_indices(tuple(headers))
    if not col_indices.get("node_name"):
        return []
    
    # Пытаемся извлечь период из заголовков таблицы
    header_text = " ".join(str(h).lower() for h in headers if h)
    period = _extract_period_from_text(header_text)
    
    # Парсим строки данных
    for row in rows:
        if not any(cell for cell in row):
            continue
        
        node_name = _get_cell_value(tuple(row), col_indices.get("node_name"))
        if not node_name or not str(node_name).strip():
            continue
        
        active_energy = _parse_float_value(_get_cell_value(tuple(row), col_indices.get("active_energy")))
        reactive_energy = _parse_float_value(_get_cell_value(tuple(row), col_indices.get("reactive_energy")))
        cost = _parse_float_value(_get_cell_value(tuple(row), col_indices.get("cost")))
        
        # Если период не найден в заголовках, пытаемся извлечь из строки данных
        row_period = period
        if not row_period or row_period == "unknown":
            row_text = " ".join(str(cell).lower() for cell in row if cell)
            row_period = _extract_period_from_text(row_text) or "unknown"
        
        node_record = {
            "node_name": str(node_name).strip(),
            "period": row_period,
            "active_energy_kwh": active_energy,
            "reactive_energy_kvarh": reactive_energy,
            "cost_sum": cost,
            "data_type": data_type,
            "data_json": {}
        }
        
        node_data.append(node_record)
    
    return node_data


def _parse_ocr_result_for_nodes(ocr_result: Dict[str, Any], file_path: str, data_type: str = "consumption") -> List[Dict[str, Any]]:
    """
    Парсит результаты OCR и извлекает данные по узлам учёта.
    
    Согласно рекомендации ML Engineer: "Использовать Gemini Vision (95% confidence)"
    
    Args:
        ocr_result: Результаты OCR
        file_path: Путь к файлу
        data_type: Тип данных ('consumption', 'production', 'realization')
    
    Returns:
        Список данных по узлам (каждая запись содержит поле data_type)
    """
    node_data = []
    
    # Извлекаем таблицы из результатов OCR
    tables = ocr_result.get("tables", [])
    if not tables:
        logger.warning("Таблицы не найдены в результатах OCR")
        return []
    
    # Извлекаем период из имени файла
    period = _extract_period_from_text(Path(file_path).stem)
    
    # Обрабатываем каждую таблицу
    for table in tables:
        if _is_node_table(table):
            table_data = _parse_node_table_from_word(table, data_type)
            # Обновляем период для всех записей из этой таблицы
            for record in table_data:
                if record["period"] == "unknown":
                    record["period"] = period or "unknown"
            node_data.extend(table_data)
    
    return node_data

