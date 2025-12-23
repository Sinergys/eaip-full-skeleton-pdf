"""
Комплексное тестирование энергопаспорта
Проверяет все пункты из STAGE2_TESTING_CHECKLIST.md
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import json

# Добавляем путь к модулям
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.energy_aggregator import aggregate_from_db_json
from database import get_upload_by_batch, list_enterprises

class PassportTester:
    def __init__(self, passport_path: Path):
        self.passport_path = passport_path
        self.workbook = None
        self.results = {
            "file_checks": {},
            "sheet_checks": {},
            "formula_checks": {},
            "data_validation": {},
            "errors": [],
            "warnings": [],
        }
    
    def load_workbook(self) -> bool:
        """Загружает рабочую книгу Excel"""
        try:
            if not self.passport_path.exists():
                self.results["errors"].append(f"Файл не найден: {self.passport_path}")
                return False
            
            self.workbook = openpyxl.load_workbook(self.passport_path, data_only=False)
            self.results["file_checks"]["exists"] = True
            self.results["file_checks"]["size_mb"] = self.passport_path.stat().st_size / (1024 * 1024)
            return True
        except Exception as e:
            self.results["errors"].append(f"Ошибка загрузки файла: {e}")
            return False
    
    def check_file_structure(self):
        """Проверка 1-2: Общие проверки файла и структуры"""
        print("\n" + "=" * 70)
        print("📋 БЛОК 1: ОБЩИЕ ПРОВЕРКИ ФАЙЛА")
        print("=" * 70)
        
        # 1.1 Существование и доступность
        self.results["file_checks"]["extension"] = self.passport_path.suffix == ".xlsx"
        print(f"✅ Расширение .xlsx: {self.results['file_checks']['extension']}")
        
        # 1.2 Структура рабочей книги
        if self.workbook:
            sheet_names = self.workbook.sheetnames
            self.results["file_checks"]["sheet_count"] = len(sheet_names)
            self.results["file_checks"]["sheet_names"] = sheet_names
            
            print(f"✅ Количество листов: {len(sheet_names)}")
            print(f"   Листы: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}")
            
            # Проверка ожидаемых листов (с учетом вариаций названий)
            expected_sheets_variants = {
                "структура": ["Struktura pr2", "Структура пр 2", "Struktura", "Структура"],
                "узлы": ["01_Узлы учета", "Узел учета", "Узлы учета", "01_Узлы учёта"],
                "оборудование": ["Equipment", "Оборудование"],
                "исходные": ["02_Исходные данные", "Исходные данные"],
                "потери": ["08_Потери_электроэнергии", "Потери", "Потери электроэнергии"],
                "баланс": ["04_Баланс", "Баланс"],
                "динамика": ["05_Динамика", "Динамика ср", "Динамика"],
                "мероприятия": ["06_Мероприятия", "Мериаприятия 1", "Мероприятия"]
            }
            
            found_sheets = {}
            for category, variants in expected_sheets_variants.items():
                found = None
                for variant in variants:
                    if variant in sheet_names:
                        found = variant
                        break
                found_sheets[category] = found
                if found:
                    print(f"   ✅ Найден лист '{category}': {found}")
                else:
                    print(f"   ⚠️ Лист '{category}' не найден (ожидались: {', '.join(variants[:2])})")
            
            self.results["file_checks"]["found_sheets"] = found_sheets
    
    def check_sheet_data(self, sheet_name: str) -> Dict[str, Any]:
        """Проверяет наличие данных на листе"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {"exists": False, "has_data": False}
        
        ws = self.workbook[sheet_name]
        has_data = False
        
        # Проверяем первые 100 строк и колонок
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), 
                                min_col=1, max_col=min(20, ws.max_column), values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                has_data = True
                break
        
        return {
            "exists": True,
            "has_data": has_data,
            "max_row": ws.max_row,
            "max_column": ws.max_column
        }
    
    def check_formulas(self, sheet_name: str) -> Dict[str, Any]:
        """Проверяет формулы на листе"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {"error": "Sheet not found"}
        
        ws = self.workbook[sheet_name]
        formula_errors = {
            "ref_errors": [],
            "value_errors": [],
            "div_zero_errors": [],
            "name_errors": [],
            "circular_refs": [],
            "total_formulas": 0
        }
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formula_errors["total_formulas"] += 1
                    formula = str(cell.value)
                    
                    # Проверка на ошибки
                    if cell.value is None:
                        continue
                    
                    # Проверяем значение ячейки (может быть ошибкой)
                    try:
                        if hasattr(cell, 'value') and isinstance(cell.value, str):
                            if '#REF!' in cell.value:
                                formula_errors["ref_errors"].append(f"{cell.coordinate}: {formula}")
                            elif '#VALUE!' in cell.value:
                                formula_errors["value_errors"].append(f"{cell.coordinate}: {formula}")
                            elif '#DIV/0!' in cell.value or '#DIV/0' in cell.value:
                                formula_errors["div_zero_errors"].append(f"{cell.coordinate}: {formula}")
                            elif '#NAME?' in cell.value:
                                formula_errors["name_errors"].append(f"{cell.coordinate}: {formula}")
                    except:
                        pass
        
        return formula_errors
    
    def test_block_1(self):
        """Блок 1: Общие проверки файла"""
        if not self.load_workbook():
            return False
        
        self.check_file_structure()
        return True
    
    def test_block_2(self):
        """Блок 2: Проверка основных листов"""
        print("\n" + "=" * 70)
        print("📊 БЛОК 2: ПРОВЕРКА ОСНОВНЫХ ЛИСТОВ")
        print("=" * 70)
        
        sheets_to_check = [
            "Struktura pr2",
            "01_Узлы учета",
            "Equipment",
            "02_Исходные данные"
        ]
        
        for sheet_name in sheets_to_check:
            print(f"\n📄 Проверка листа: {sheet_name}")
            sheet_info = self.check_sheet_data(sheet_name)
            self.results["sheet_checks"][sheet_name] = sheet_info
            
            if sheet_info.get("exists"):
                print(f"   ✅ Лист существует")
                if sheet_info.get("has_data"):
                    print(f"   ✅ Содержит данные ({sheet_info['max_row']} строк × {sheet_info['max_column']} столбцов)")
                else:
                    print(f"   ⚠️ Лист пуст")
                    self.results["warnings"].append(f"Лист '{sheet_name}' пуст")
            else:
                print(f"   ❌ Лист не найден")
                self.results["warnings"].append(f"Лист '{sheet_name}' отсутствует")
    
    def test_block_3(self):
        """Блок 3: Проверка расчетных листов"""
        print("\n" + "=" * 70)
        print("🧮 БЛОК 3: ПРОВЕРКА РАСЧЕТНЫХ ЛИСТОВ")
        print("=" * 70)
        
        sheets_to_check = [
            "08_Потери_электроэнергии",
            "04_Баланс",
            "05_Динамика",
            "06_Мероприятия"
        ]
        
        for sheet_name in sheets_to_check:
            print(f"\n📄 Проверка листа: {sheet_name}")
            sheet_info = self.check_sheet_data(sheet_name)
            formula_info = self.check_formulas(sheet_name)
            
            self.results["sheet_checks"][sheet_name] = sheet_info
            self.results["formula_checks"][sheet_name] = formula_info
            
            if sheet_info.get("exists"):
                print(f"   ✅ Лист существует")
                if formula_info.get("total_formulas", 0) > 0:
                    print(f"   ✅ Найдено формул: {formula_info['total_formulas']}")
                    
                    errors = []
                    if formula_info.get("ref_errors"):
                        errors.append(f"#REF!: {len(formula_info['ref_errors'])}")
                    if formula_info.get("value_errors"):
                        errors.append(f"#VALUE!: {len(formula_info['value_errors'])}")
                    if formula_info.get("div_zero_errors"):
                        errors.append(f"#DIV/0!: {len(formula_info['div_zero_errors'])}")
                    
                    if errors:
                        print(f"   ⚠️ Ошибки формул: {', '.join(errors)}")
                        for err_type, err_list in formula_info.items():
                            if err_type.endswith("_errors") and err_list:
                                self.results["warnings"].extend([f"{sheet_name}: {err}" for err in err_list[:3]])
                    else:
                        print(f"   ✅ Формулы без ошибок")
            else:
                print(f"   ⚠️ Лист не найден")
    
    def test_block_4(self):
        """Блок 4: Проверка формул и связей"""
        print("\n" + "=" * 70)
        print("🔗 БЛОК 4: ПРОВЕРКА ФОРМУЛ И СВЯЗЕЙ")
        print("=" * 70)
        
        # Проверяем все листы на формулы
        total_formulas = 0
        total_errors = 0
        
        for sheet_name in self.workbook.sheetnames:
            formula_info = self.check_formulas(sheet_name)
            total_formulas += formula_info.get("total_formulas", 0)
            
            errors_count = (
                len(formula_info.get("ref_errors", [])) +
                len(formula_info.get("value_errors", [])) +
                len(formula_info.get("div_zero_errors", [])) +
                len(formula_info.get("name_errors", []))
            )
            total_errors += errors_count
        
        print(f"✅ Всего формул в книге: {total_formulas}")
        print(f"{'✅' if total_errors == 0 else '⚠️'} Ошибок формул: {total_errors}")
        
        self.results["formula_checks"]["total_formulas"] = total_formulas
        self.results["formula_checks"]["total_errors"] = total_errors
        
        # Проверка связей между листами (базовая)
        if "Struktura pr2" in self.workbook.sheetnames and "04_Баланс" in self.workbook.sheetnames:
            print("\n🔗 Проверка связей между листами...")
            ws_struktura = self.workbook["Struktura pr2"]
            links_found = 0
            
            for row in ws_struktura.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        if "'04_Баланс'" in formula or "'Баланс'" in formula or "Баланс!" in formula:
                            links_found += 1
            
            if links_found > 0:
                print(f"   ✅ Найдено ссылок на лист 'Баланс': {links_found}")
            else:
                print(f"   ⚠️ Ссылки на лист 'Баланс' не найдены")
    
    def generate_report(self) -> str:
        """Генерирует итоговый отчет"""
        report = []
        report.append("=" * 70)
        report.append("📊 ИТОГОВЫЙ ОТЧЕТ ТЕСТИРОВАНИЯ ЭНЕРГОПАСПОРТА")
        report.append("=" * 70)
        report.append(f"\nФайл: {self.passport_path.name}")
        report.append(f"Размер: {self.results['file_checks'].get('size_mb', 0):.2f} MB")
        report.append(f"Листов: {self.results['file_checks'].get('sheet_count', 0)}")
        
        # Статистика
        report.append("\n📈 СТАТИСТИКА:")
        report.append(f"   Формул: {self.results['formula_checks'].get('total_formulas', 0)}")
        report.append(f"   Ошибок формул: {self.results['formula_checks'].get('total_errors', 0)}")
        
        # Ошибки
        if self.results["errors"]:
            report.append("\n❌ КРИТИЧЕСКИЕ ОШИБКИ:")
            for err in self.results["errors"]:
                report.append(f"   - {err}")
        
        # Предупреждения
        if self.results["warnings"]:
            report.append(f"\n⚠️ ПРЕДУПРЕЖДЕНИЯ ({len(self.results['warnings'])}):")
            for warn in self.results["warnings"][:10]:
                report.append(f"   - {warn}")
            if len(self.results["warnings"]) > 10:
                report.append(f"   ... и еще {len(self.results['warnings']) - 10} предупреждений")
        
        # Итог
        report.append("\n" + "=" * 70)
        if not self.results["errors"] and not self.results["warnings"]:
            report.append("✅ ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ УСПЕШНО!")
        elif not self.results["errors"]:
            report.append("⚠️ ЕСТЬ ПРЕДУПРЕЖДЕНИЯ, НО КРИТИЧЕСКИХ ОШИБОК НЕТ")
        else:
            report.append("❌ ОБНАРУЖЕНЫ КРИТИЧЕСКИЕ ОШИБКИ")
        report.append("=" * 70)
        
        return "\n".join(report)


def main():
    """Основная функция тестирования"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Комплексное тестирование энергопаспорта")
    parser.add_argument("passport", type=str, help="Путь к файлу энергопаспорта")
    parser.add_argument("--blocks", type=str, nargs="+", 
                       choices=["1", "2", "3", "4", "all"],
                       default=["all"],
                       help="Блоки для тестирования")
    
    args = parser.parse_args()
    
    passport_path = Path(args.passport)
    if not passport_path.exists():
        print(f"❌ Файл не найден: {passport_path}")
        return 1
    
    tester = PassportTester(passport_path)
    
    blocks_to_run = args.blocks if "all" not in args.blocks else ["1", "2", "3", "4"]
    
    print("=" * 70)
    print("🔍 КОМПЛЕКСНОЕ ТЕСТИРОВАНИЕ ЭНЕРГОПАСПОРТА")
    print("=" * 70)
    
    # Блок 1
    if "1" in blocks_to_run:
        if not tester.test_block_1():
            print("\n❌ Блок 1 провален: не удалось загрузить файл")
            return 1
    
    # Блок 2
    if "2" in blocks_to_run:
        tester.test_block_2()
    
    # Блок 3
    if "3" in blocks_to_run:
        tester.test_block_3()
    
    # Блок 4
    if "4" in blocks_to_run:
        tester.test_block_4()
    
    # Итоговый отчет
    print("\n" + tester.generate_report())
    
    return 0 if not tester.results["errors"] else 1


if __name__ == "__main__":
    sys.exit(main())

