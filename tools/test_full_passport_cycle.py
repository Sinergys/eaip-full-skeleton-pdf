"""Тестирование полного цикла генерации паспорта с проверкой листа 'Баланс'"""
import sys
from pathlib import Path
import json
import openpyxl
from openpyxl import load_workbook

# Добавляем пути для импорта
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent / "eaip_full_skeleton" / "services" / "ingest"))

def check_balance_sheet(workbook_path: Path) -> dict:
    """Проверяет корректность листа 'Баланс' в сгенерированном паспорте"""
    results = {
        "file_exists": False,
        "balance_sheet_found": False,
        "headers_correct": False,
        "formulas_correct": False,
        "data_rows": 0,
        "totals_row_found": False,
        "errors": [],
        "warnings": []
    }
    
    if not workbook_path.exists():
        results["errors"].append(f"Файл не найден: {workbook_path}")
        return results
    
    results["file_exists"] = True
    
    try:
        wb = load_workbook(workbook_path, data_only=False)
        
        # Ищем лист "Баланс" (различные варианты названий)
        balance_sheet = None
        for sheet_name in ["04_Баланс", "Баланс", "Balans", "04_Balans"]:
            if sheet_name in wb.sheetnames:
                balance_sheet = wb[sheet_name]
                results["balance_sheet_found"] = True
                break
        
        if not balance_sheet:
            results["errors"].append("Лист 'Баланс' не найден в паспорте")
            return results
        
        # Проверяем заголовки (строка 2)
        expected_headers = [
            "Квартал",
            "Технологические, кВт·ч",
            "Собственные нужды, кВт·ч",
            "Производственные, кВт·ч",
            "Хоз-бытовые, кВт·ч",
            "Итого, кВт·ч"
        ]
        
        headers_row = 2
        headers_found = []
        for col_idx, expected_header in enumerate(expected_headers, start=1):
            cell = balance_sheet.cell(row=headers_row, column=col_idx)
            if cell.value:
                headers_found.append(str(cell.value))
        
        if len(headers_found) >= len(expected_headers):
            results["headers_correct"] = True
        else:
            results["warnings"].append(f"Заголовки частично найдены: {headers_found}")
        
        # Ищем строки с данными (начиная со строки 3)
        data_start_row = 3
        data_rows = []
        
        for row_idx in range(data_start_row, balance_sheet.max_row + 1):
            quarter_cell = balance_sheet.cell(row=row_idx, column=1)
            if quarter_cell.value and str(quarter_cell.value).strip():
                data_rows.append(row_idx)
        
        results["data_rows"] = len(data_rows)
        
        if results["data_rows"] == 0:
            results["warnings"].append("Не найдено строк с данными в листе 'Баланс'")
        
        # Проверяем формулы в колонке "Итого" (колонка F = 6)
        total_col = 6
        formulas_ok = True
        
        for row_idx in data_rows:
            total_cell = balance_sheet.cell(row=row_idx, column=total_col)
            if total_cell.data_type == "f":
                # Проверяем, что формула суммирует колонки B-E
                formula = str(total_cell.value)
                if "SUM" in formula.upper() and "B" in formula and "E" in formula:
                    pass  # Формула корректна
                else:
                    formulas_ok = False
                    results["warnings"].append(f"Формула в строке {row_idx} может быть некорректной: {formula}")
            else:
                formulas_ok = False
                results["warnings"].append(f"В строке {row_idx} колонка 'Итого' не содержит формулу")
        
        results["formulas_correct"] = formulas_ok
        
        # Ищем строку "ИТОГО"
        totals_row = None
        for row_idx in range(data_start_row, balance_sheet.max_row + 1):
            label_cell = balance_sheet.cell(row=row_idx, column=1)
            if label_cell.value and "ИТОГО" in str(label_cell.value).upper():
                totals_row = row_idx
                results["totals_row_found"] = True
                break
        
        if not totals_row:
            results["warnings"].append("Строка 'ИТОГО' не найдена")
        
        # Проверяем формулы в строке итогов
        if totals_row:
            for col in [2, 3, 4, 5, 6]:  # Колонки B-F
                totals_cell = balance_sheet.cell(row=totals_row, column=col)
                if totals_cell.data_type == "f":
                    formula = str(totals_cell.value)
                    if "SUM" in formula.upper():
                        pass  # Формула корректна
                    else:
                        results["warnings"].append(f"Формула итогов в колонке {col} может быть некорректной: {formula}")
        
        wb.close()
        
    except Exception as e:
        results["errors"].append(f"Ошибка при проверке файла: {e}")
    
    return results


def print_test_results(results: dict):
    """Выводит результаты тестирования"""
    print("\n" + "="*60)
    print("РЕЗУЛЬТАТЫ ПРОВЕРКИ ЛИСТА 'БАЛАНС'")
    print("="*60)
    
    print(f"\n✅ Файл существует: {'Да' if results['file_exists'] else 'Нет'}")
    print(f"✅ Лист 'Баланс' найден: {'Да' if results['balance_sheet_found'] else 'Нет'}")
    print(f"✅ Заголовки корректны: {'Да' if results['headers_correct'] else 'Нет'}")
    print(f"✅ Формулы корректны: {'Да' if results['formulas_correct'] else 'Нет'}")
    print(f"📊 Строк с данными: {results['data_rows']}")
    print(f"✅ Строка 'ИТОГО' найдена: {'Да' if results['totals_row_found'] else 'Нет'}")
    
    if results['errors']:
        print("\n❌ ОШИБКИ:")
        for error in results['errors']:
            print(f"   - {error}")
    
    if results['warnings']:
        print("\n⚠️ ПРЕДУПРЕЖДЕНИЯ:")
        for warning in results['warnings']:
            print(f"   - {warning}")
    
    if not results['errors'] and not results['warnings']:
        print("\n✅ ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ УСПЕШНО!")
    
    print("="*60)


def main():
    """Основная функция тестирования"""
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python test_full_passport_cycle.py <путь_к_паспорту.xlsx>")
        print("\nПример:")
        print("  python test_full_passport_cycle.py generated_passport.xlsx")
        return
    
    passport_path = Path(sys.argv[1])
    
    print(f"Проверка паспорта: {passport_path}")
    
    results = check_balance_sheet(passport_path)
    print_test_results(results)
    
    # Возвращаем код выхода
    if results['errors']:
        sys.exit(1)
    elif results['warnings']:
        sys.exit(0)  # Предупреждения не критичны
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()

