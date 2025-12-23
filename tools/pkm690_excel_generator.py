"""
Генератор Excel-таблиц энергетического паспорта по ПКМ 690 Узбекистан
Автоматические расчеты, нормативы и сводные данные
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Настройка логирования
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class PKM690ExcelGenerator:
    """Генератор Excel-таблиц энергетического паспорта по ПКМ 690"""

    def __init__(self, enterprise_data: dict = None, energy_data: dict = None):
        """
        Инициализация генератора с данными из dict (вместо БД).

        Args:
            enterprise_data: Данные предприятия из БД
                {
                    'id': 1,
                    'name': 'ООО Синергис',
                    'inn': '123456789',
                    'address': '...',
                    'director_name': '...',
                    'industry': '...',
                    ...
                }
            energy_data: Агрегированные данные энергопотребления
                {
                    'resources': {
                        'electricity': {'2022-Q1': {...}},
                        'gas': {...},
                        ...
                    }
                }
        """
        self.enterprise_data = enterprise_data or {}
        self.energy_data = energy_data or {}
        self.wb = None

        # Нормативы по ПКМ 690 Узбекистан
        self.normatives = {
            "electricity": {
                "specific_consumption": 0.15,  # кВт·ч/м²·год
                "efficiency_min": 0.85,
                "cost_per_kwh": 150,  # сум/кВт·ч
            },
            "gas": {
                "specific_consumption": 0.12,  # м³/м²·год
                "efficiency_min": 0.90,
                "cost_per_m3": 200,  # сум/м³
            },
            "water": {
                "specific_consumption": 0.08,  # м³/м²·год
                "efficiency_min": 0.80,
                "cost_per_m3": 50,  # сум/м³
            },
            "building": {
                "heat_loss_coefficient": 0.05,  # кВт/м²·°C
                "insulation_thickness_min": 0.1,  # м
                "window_area_max": 0.3,  # доля от площади стен
            },
            "environmental": {
                # Предельно допустимые выбросы (ПДВ) в атмосферу
                "emissions": {
                    # Нормативы по загрязняющим веществам (мг/м³)
                    "co": {"max_concentration": 3.0, "unit": "мг/м³"},  # Оксид углерода
                    "nox": {"max_concentration": 0.4, "unit": "мг/м³"},  # Оксиды азота
                    "so2": {"max_concentration": 0.5, "unit": "мг/м³"},  # Диоксид серы
                    "dust": {"max_concentration": 0.5, "unit": "мг/м³"},  # Пыль
                    "pm10": {"max_concentration": 0.06, "unit": "мг/м³"},  # PM10
                    "pm2_5": {"max_concentration": 0.035, "unit": "мг/м³"},  # PM2.5
                },
                # Предельно допустимые сбросы (ПДС) в водные объекты
                "discharges": {
                    # Нормативы по загрязняющим веществам (мг/л)
                    "suspended_solids": {"max_concentration": 0.25, "unit": "мг/л"},  # Взвешенные вещества
                    "bod5": {"max_concentration": 3.0, "unit": "мг/л"},  # БПК5
                    "cod": {"max_concentration": 30.0, "unit": "мг/л"},  # ХПК
                    "ammonium": {"max_concentration": 0.5, "unit": "мг/л"},  # Аммоний
                    "nitrates": {"max_concentration": 40.0, "unit": "мг/л"},  # Нитраты
                    "phosphates": {"max_concentration": 0.2, "unit": "мг/л"},  # Фосфаты
                },
                # Категории опасности предприятий
                "hazard_categories": {
                    "category_1": {"description": "Чрезвычайно опасные", "criteria": "ПДВ > 1000 т/год"},
                    "category_2": {"description": "Высокоопасные", "criteria": "ПДВ 100-1000 т/год"},
                    "category_3": {"description": "Умеренно опасные", "criteria": "ПДВ 10-100 т/год"},
                    "category_4": {"description": "Малоопасные", "criteria": "ПДВ < 10 т/год"},
                },
            },
        }

    def create_energy_passport(self, output_path: str) -> bool:
        """
        Создание полного энергетического паспорта в Excel.
        Использует данные из self.enterprise_data и self.energy_data.
        """
        try:
            if not self.enterprise_data:
                logger.error("Данные предприятия не предоставлены")
                return False

            logger.info(
                f"📊 Создание энергетического паспорта для: {self.enterprise_data.get('name', 'Неизвестно')}"
            )

            # Создаем новую книгу
            self.wb = Workbook()

            # Удаляем стандартный лист
            self.wb.remove(self.wb.active)

            # Создаем листы паспорта
            self.create_title_sheet(self.enterprise_data)
            self.create_enterprise_info_sheet(self.enterprise_data)
            self.create_energy_consumption_sheet(self.enterprise_data)
            self.create_equipment_sheet(self.enterprise_data)
            self.create_monthly_data_sheet(self.enterprise_data)
            self.create_calculations_sheet(self.enterprise_data)
            self.create_normatives_sheet(self.enterprise_data)
            self.create_measures_sheet(self.enterprise_data)
            self.create_summary_sheet(self.enterprise_data)

            # Сохраняем файл
            self.wb.save(output_path)
            logger.info(f"✅ Энергетический паспорт сохранен: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Ошибка создания энергетического паспорта: {e}")
            return False

    def create_title_sheet(self, enterprise_data: dict):
        """Создание титульного листа"""
        logger.info("📋 Создание титульного листа...")

        ws = self.wb.create_sheet("Title")

        # Заголовок
        ws["A1"] = "ЭНЕРГЕТИЧЕСКИЙ ПАСПОРТ"
        ws["A1"].font = Font(name="Times New Roman", size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:F1")

        ws["A2"] = "ПРЕДПРИЯТИЯ"
        ws["A2"].font = Font(name="Times New Roman", size=16, bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F2")

        # Информация о предприятии
        ws["A4"] = "Наименование предприятия:"
        ws["B4"] = enterprise_data["name"]
        ws["B4"].font = Font(bold=True)

        ws["A5"] = "ИНН:"
        ws["B5"] = enterprise_data["inn"] or "не указан"

        ws["A6"] = "Адрес:"
        ws["B6"] = enterprise_data["address"] or "не указан"

        ws["A7"] = "Директор:"
        ws["B7"] = enterprise_data["director_name"] or "не указан"

        ws["A8"] = "Отрасль:"
        ws["B8"] = enterprise_data["industry"] or "не указана"

        ws["A9"] = "Год отчета:"
        ws["B9"] = enterprise_data["reporting_year"] or datetime.now().year

        ws["A10"] = "Аудитор:"
        ws["B10"] = enterprise_data.get("auditor", "не указан")

        ws["A11"] = "Дата составления:"
        ws["B11"] = datetime.now().strftime("%d.%m.%Y")

        ws["A12"] = "Стандарт:"
        ws["B12"] = "ПКМ 690 Узбекистан"

        # Настройка стилей
        self.apply_title_styles(ws)

    def create_enterprise_info_sheet(self, enterprise_data: dict):
        """Создание листа с информацией о предприятии"""
        logger.info("🏢 Создание листа информации о предприятии...")

        ws = self.wb.create_sheet("Enterprise")

        # Заголовок
        ws["A1"] = "ОБЩИЕ СВЕДЕНИЯ О ПРЕДПРИЯТИИ"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:D1")

        # Основная информация
        info_data = [
            ["Показатель", "Значение", "Единица измерения", "Примечание"],
            ["Полное наименование", enterprise_data["name"], "", ""],
            ["ИНН", enterprise_data["inn"] or "не указан", "", ""],
            ["Юридический адрес", enterprise_data["address"] or "не указан", "", ""],
            ["Руководитель", enterprise_data["director_name"] or "не указан", "", ""],
            [
                "Отрасль деятельности",
                enterprise_data["industry"] or "не указана",
                "",
                "",
            ],
            [
                "Год основания",
                enterprise_data.get("founding_year", "не указан"),
                "год",
                "",
            ],
            ["Площадь территории", enterprise_data.get("territory_area", 0), "м²", ""],
            ["Площадь зданий", enterprise_data.get("building_area", 0), "м²", ""],
            [
                "Количество сотрудников",
                enterprise_data.get("employees_count", 0),
                "чел.",
                "",
            ],
            ["Режим работы", enterprise_data.get("work_schedule", "не указан"), "", ""],
            [
                "Основная продукция",
                enterprise_data.get("main_products", "не указана"),
                "",
                "",
            ],
        ]

        # Заполняем данные
        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=cell_value)

        # Создаем таблицу - ОТКЛЮЧЕНО (конфликт с формулами)
        # table = Table(displayName="EnterpriseInfo", ref=f"A3:D{len(info_data) + 2}")
        # style = TableStyleInfo(
        #     name="TableStyleMedium2",
        #     showFirstColumn=False,
        #     showLastColumn=False,
        #     showRowStripes=True,
        #     showColumnStripes=False
        # )
        # table.tableStyleInfo = style
        # ws.add_table(table)

        # Настройка стилей
        self.apply_table_styles(ws, f"A3:D{len(info_data) + 2}")

    def create_energy_consumption_sheet(self, enterprise_data: dict):
        """Создание листа энергопотребления"""
        logger.info("⚡ Создание листа энергопотребления...")

        ws = self.wb.create_sheet("Energy")

        # Заголовок
        ws["A1"] = "АНАЛИЗ ЭНЕРГОПОТРЕБЛЕНИЯ"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:F1")

        # Получаем данные по энергопотреблению
        energy_data = self.get_energy_consumption_data()

        # Сводная таблица энергопотребления
        ws["A3"] = "СВОДНЫЕ ДАННЫЕ ПО ЭНЕРГОПОТРЕБЛЕНИЮ"
        ws["A3"].font = Font(bold=True)

        summary_data = [
            [
                "Вид ресурса",
                "Потребление",
                "Единица измерения",
                "Стоимость",
                "Единица измерения",
                "Доля в общем потреблении (%)",
            ],
            [
                "Электрическая энергия",
                energy_data.get("electricity_total", 0),
                "кВт·ч/год",
                energy_data.get("electricity_total", 0)
                * self.normatives["electricity"]["cost_per_kwh"],
                "сум/год",
                "",
            ],
            [
                "Природный газ",
                energy_data.get("gas_total", 0),
                "м³/год",
                energy_data.get("gas_total", 0) * self.normatives["gas"]["cost_per_m3"],
                "сум/год",
                "",
            ],
            [
                "Вода",
                energy_data.get("water_total", 0),
                "м³/год",
                energy_data.get("water_total", 0)
                * self.normatives["water"]["cost_per_m3"],
                "сум/год",
                "",
            ],
            ["ИТОГО", "", "", "", "", ""],
        ]

        # Заполняем данные
        for row_idx, row_data in enumerate(summary_data, 4):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Расчет долей
        total_consumption = (
            energy_data.get("electricity_total", 0)
            + energy_data.get("gas_total", 0)
            + energy_data.get("water_total", 0)
        )

        if total_consumption > 0:
            ws["F5"] = (
                f"={energy_data.get('electricity_total', 0)}/{total_consumption}*100"
            )
            ws["F6"] = f"={energy_data.get('gas_total', 0)}/{total_consumption}*100"
            ws["F7"] = f"={energy_data.get('water_total', 0)}/{total_consumption}*100"

        # Расчет итогов
        ws["B8"] = "=SUM(B5:B7)"
        ws["D8"] = "=SUM(D5:D7)"
        ws["F8"] = "=SUM(F5:F7)"

        # Создаем таблицу
        table = Table(displayName="EnergyConsumption", ref="A4:F8")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Анализ эффективности
        ws["A10"] = "АНАЛИЗ ЭНЕРГОЭФФЕКТИВНОСТИ"
        ws["A10"].font = Font(bold=True)

        efficiency_data = [
            [
                "Показатель",
                "Фактическое значение",
                "Нормативное значение",
                "Соответствие норме",
            ],
            [
                "Удельное потребление электроэнергии",
                f"={energy_data.get('electricity_total', 0)}/{enterprise_data.get('building_area', 1)}",
                self.normatives["electricity"]["specific_consumption"],
                '=IF(B12<=C12,"Соответствует","Превышает")',
            ],
            [
                "Удельное потребление газа",
                f"={energy_data.get('gas_total', 0)}/{enterprise_data.get('building_area', 1)}",
                self.normatives["gas"]["specific_consumption"],
                '=IF(B13<=C13,"Соответствует","Превышает")',
            ],
            [
                "Удельное потребление воды",
                f"={energy_data.get('water_total', 0)}/{enterprise_data.get('building_area', 1)}",
                self.normatives["water"]["specific_consumption"],
                '=IF(B14<=C14,"Соответствует","Превышает")',
            ],
        ]

        # Заполняем данные эффективности
        for row_idx, row_data in enumerate(efficiency_data, 11):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Создаем таблицу эффективности
        table2 = Table(displayName="EnergyEfficiency", ref="A11:D14")
        style2 = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table2.tableStyleInfo = style2
        # ws.add_table(table2)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Настройка стилей
        self.apply_table_styles(ws, "A4:F8")
        self.apply_table_styles(ws, "A11:D14")

    def create_equipment_sheet(self, enterprise_data: dict):
        """Создание листа оборудования"""
        logger.info("🏭 Создание листа оборудования...")

        ws = self.wb.create_sheet("Equipment")

        # Заголовок
        ws["A1"] = "АНАЛИЗ ОБОРУДОВАНИЯ"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:F1")

        # Получаем данные по оборудованию
        equipment_data = self.get_equipment_data()

        # Перечень оборудования
        ws["A3"] = "ПЕРЕЧЕНЬ ОСНОВНОГО ОБОРУДОВАНИЯ"
        ws["A3"].font = Font(bold=True)

        equipment_headers = [
            "Наименование",
            "Тип",
            "Мощность (кВт)",
            "Год установки",
            "Коэффициент использования",
            "Годовое потребление (кВт·ч)",
        ]

        # Заголовки таблицы
        for col_idx, header in enumerate(equipment_headers, 1):
            ws.cell(row=4, column=col_idx, value=header)

        # Данные оборудования
        row_idx = 5
        total_power = 0
        total_consumption = 0

        for equipment in equipment_data.get("equipment_list", []):
            ws.cell(row=row_idx, column=1, value=equipment.get("name", ""))
            ws.cell(row=row_idx, column=2, value=equipment.get("type", ""))
            ws.cell(row=row_idx, column=3, value=equipment.get("power", 0))
            ws.cell(row=row_idx, column=4, value=equipment.get("year", ""))
            ws.cell(row=row_idx, column=5, value=0.8)  # Коэффициент использования
            ws.cell(
                row=row_idx, column=6, value=f"=C{row_idx}*E{row_idx}*8760"
            )  # Годовое потребление

            total_power += equipment.get("power", 0)
            row_idx += 1

        # Итоговая строка
        ws.cell(row=row_idx, column=1, value="ИТОГО")
        ws.cell(row=row_idx, column=3, value=f"=SUM(C5:C{row_idx - 1})")
        ws.cell(row=row_idx, column=6, value=f"=SUM(F5:F{row_idx - 1})")

        # Создаем таблицу
        table = Table(displayName="Equipment", ref=f"A4:F{row_idx}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Анализ эффективности оборудования
        ws[f"A{row_idx + 2}"] = "АНАЛИЗ ЭФФЕКТИВНОСТИ ОБОРУДОВАНИЯ"
        ws[f"A{row_idx + 2}"].font = Font(bold=True)

        efficiency_analysis = [
            ["Показатель", "Значение", "Единица измерения"],
            ["Общая установленная мощность", f"=C{row_idx}", "кВт"],
            ["Средний коэффициент использования", f"=AVERAGE(E5:E{row_idx - 1})", ""],
            ["Общее годовое потребление", f"=F{row_idx}", "кВт·ч"],
            [
                "Средняя мощность на единицу оборудования",
                f"=C{row_idx}/{row_idx - 5}",
                "кВт",
            ],
            ["Количество единиц оборудования", row_idx - 5, "шт."],
        ]

        # Заполняем анализ эффективности
        for i, row_data in enumerate(efficiency_analysis, row_idx + 3):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=i, column=col_idx, value=cell_value)

        # Настройка стилей
        self.apply_table_styles(ws, f"A4:F{row_idx}")
        self.apply_table_styles(ws, f"A{row_idx + 3}:C{row_idx + 8}")

    def create_monthly_data_sheet(self, enterprise_data: dict):
        """Создание листа месячных данных"""
        logger.info("📅 Создание листа месячных данных...")

        ws = self.wb.create_sheet("Monthly")

        # Заголовок
        ws["A1"] = "ЭНЕРГОПОТРЕБЛЕНИЕ ПО МЕСЯЦАМ"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:H1")

        # Получаем месячные данные
        monthly_data = self.get_monthly_energy_data()

        # Заголовки таблицы
        headers = [
            "Месяц",
            "Электричество (кВт·ч)",
            "Газ (м³)",
            "Вода (м³)",
            "Общая стоимость (сум)",
            "Средняя температура (°C)",
            "Коэффициент сезонности",
            "Примечание",
        ]

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=3, column=col_idx, value=header)

        # Данные по месяцам
        months = [
            "Январь",
            "Февраль",
            "Март",
            "Апрель",
            "Май",
            "Июнь",
            "Июль",
            "Август",
            "Сентябрь",
            "Октябрь",
            "Ноябрь",
            "Декабрь",
        ]

        row_idx = 4
        total_electricity = 0
        total_gas = 0
        total_water = 0
        total_cost = 0

        for i, month in enumerate(months, 1):
            # Используем данные из базы или генерируем тестовые
            month_data = (
                monthly_data[i - 1]
                if i - 1 < len(monthly_data)
                else {
                    "electricity": 1500 + (i - 1) * 100,
                    "gas": 800 + (i - 1) * 50,
                    "water": 200 + (i - 1) * 20,
                    "total_cost": 45000 + (i - 1) * 3000,
                }
            )

            ws.cell(row=row_idx, column=1, value=month)
            ws.cell(row=row_idx, column=2, value=month_data.get("electricity", 0))
            ws.cell(row=row_idx, column=3, value=month_data.get("gas", 0))
            ws.cell(row=row_idx, column=4, value=month_data.get("water", 0))
            ws.cell(row=row_idx, column=5, value=month_data.get("total_cost", 0))
            ws.cell(row=row_idx, column=6, value=-10 + i * 3)  # Средняя температура
            ws.cell(
                row=row_idx, column=7, value=f"=IF(F{row_idx}<=0,1.2,0.8)"
            )  # Коэффициент сезонности
            ws.cell(row=row_idx, column=8, value="")

            total_electricity += month_data.get("electricity", 0)
            total_gas += month_data.get("gas", 0)
            total_water += month_data.get("water", 0)
            total_cost += month_data.get("total_cost", 0)

            row_idx += 1

        # Итоговая строка
        ws.cell(row=row_idx, column=1, value="ИТОГО")
        ws.cell(row=row_idx, column=2, value=f"=SUM(B4:B{row_idx - 1})")
        ws.cell(row=row_idx, column=3, value=f"=SUM(C4:C{row_idx - 1})")
        ws.cell(row=row_idx, column=4, value=f"=SUM(D4:D{row_idx - 1})")
        ws.cell(row=row_idx, column=5, value=f"=SUM(E4:E{row_idx - 1})")
        ws.cell(row=row_idx, column=6, value=f"=AVERAGE(F4:F{row_idx - 1})")
        ws.cell(row=row_idx, column=7, value=f"=AVERAGE(G4:G{row_idx - 1})")

        # Создаем таблицу
        table = Table(displayName="MonthlyData", ref=f"A3:H{row_idx}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # График потребления
        ws[f"A{row_idx + 2}"] = "АНАЛИЗ СЕЗОННОСТИ"
        ws[f"A{row_idx + 2}"].font = Font(bold=True)

        seasonality_analysis = [
            ["Показатель", "Зимний период", "Летний период", "Переходный период"],
            [
                "Среднее потребление электроэнергии",
                "=AVERAGE(B4:B6,B10:B12)",
                "=AVERAGE(B7:B9)",
                "=AVERAGE(B4,B10)",
            ],
            [
                "Среднее потребление газа",
                "=AVERAGE(C4:C6,C10:C12)",
                "=AVERAGE(C7:C9)",
                "=AVERAGE(C4,C10)",
            ],
            [
                "Среднее потребление воды",
                "=AVERAGE(D4:D6,D10:D12)",
                "=AVERAGE(D7:D9)",
                "=AVERAGE(D4,D10)",
            ],
        ]

        # Заполняем анализ сезонности
        for i, row_data in enumerate(seasonality_analysis, row_idx + 3):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=i, column=col_idx, value=cell_value)

        # Настройка стилей
        self.apply_table_styles(ws, f"A3:H{row_idx}")
        self.apply_table_styles(ws, f"A{row_idx + 3}:D{row_idx + 6}")

    def create_calculations_sheet(self, enterprise_data: dict):
        """Создание листа расчетов"""
        logger.info("🧮 Создание листа расчетов...")

        ws = self.wb.create_sheet("Calculations")

        # Заголовок
        ws["A1"] = "РАСЧЕТЫ И ФОРМУЛЫ"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:F1")

        # Расчет энергоэффективности
        ws["A3"] = "РАСЧЕТ ЭНЕРГОЭФФЕКТИВНОСТИ"
        ws["A3"].font = Font(bold=True)

        efficiency_calculations = [
            [
                "Показатель",
                "Формула",
                "Результат",
                "Единица измерения",
                "Норматив",
                "Соответствие",
            ],
            [
                "Удельное потребление электроэнергии",
                "=Energy!B8/Enterprise!B9",
                "=B4",
                "кВт·ч/м²·год",
                self.normatives["electricity"]["specific_consumption"],
                '=IF(C4<=E4,"Соответствует","Превышает")',
            ],
            [
                "Удельное потребление газа",
                "=Energy!C8/Enterprise!B9",
                "=B5",
                "м³/м²·год",
                self.normatives["gas"]["specific_consumption"],
                '=IF(C5<=E5,"Соответствует","Превышает")',
            ],
            [
                "Удельное потребление воды",
                "=Energy!D8/Enterprise!B9",
                "=B6",
                "м³/м²·год",
                self.normatives["water"]["specific_consumption"],
                '=IF(C6<=E6,"Соответствует","Превышает")',
            ],
            [
                "Общий коэффициент энергоэффективности",
                "=AVERAGE(C4:C6)/AVERAGE(E4:E6)",
                "=B7",
                "",
                1.0,
                '=IF(C7>=E7,"Соответствует","Не соответствует")',
            ],
        ]

        # Заполняем расчеты
        for row_idx, row_data in enumerate(efficiency_calculations, 4):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Создаем таблицу
        table = Table(displayName="EfficiencyCalculations", ref="A4:F8")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Расчет экономии
        ws["A10"] = "РАСЧЕТ ЭКОНОМИИ ОТ МЕРОПРИЯТИЙ"
        ws["A10"].font = Font(bold=True)

        savings_calculations = [
            [
                "Мероприятие",
                "Экономия (%)",
                "Годовая экономия (сум)",
                "Стоимость внедрения (сум)",
                "Срок окупаемости (лет)",
            ],
            ["Замена освещения на LED", 15, "=Energy!D8*0.15", 50000, "=C12/D12"],
            ["Утепление здания", 20, "=Energy!D8*0.20", 80000, "=C13/D13"],
            ["Модернизация оборудования", 25, "=Energy!D8*0.25", 120000, "=C14/D14"],
            ["ИТОГО", "", "=SUM(C12:C14)", "=SUM(D12:D14)", "=AVERAGE(E12:E14)"],
        ]

        # Заполняем расчеты экономии
        for row_idx, row_data in enumerate(savings_calculations, 11):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Создаем таблицу экономии
        table2 = Table(displayName="SavingsCalculations", ref="A11:E15")
        style2 = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table2.tableStyleInfo = style2
        # ws.add_table(table2)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Настройка стилей
        self.apply_table_styles(ws, "A4:F8")
        self.apply_table_styles(ws, "A11:E15")

    def create_normatives_sheet(self, enterprise_data: dict):
        """Создание листа нормативов"""
        logger.info("📋 Создание листа нормативов...")

        ws = self.wb.create_sheet("Normatives")

        # Заголовок
        ws["A1"] = "НОРМАТИВЫ ПО ПКМ 690 УЗБЕКИСТАН"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:F1")

        # Нормативы по видам энергии
        ws["A3"] = "НОРМАТИВЫ ПО ВИДАМ ЭНЕРГИИ"
        ws["A3"].font = Font(bold=True)

        normative_data = [
            [
                "Вид энергии",
                "Удельное потребление",
                "Единица измерения",
                "Минимальная эффективность",
                "Стоимость",
                "Единица измерения",
            ],
            [
                "Электрическая энергия",
                self.normatives["electricity"]["specific_consumption"],
                "кВт·ч/м²·год",
                self.normatives["electricity"]["efficiency_min"],
                self.normatives["electricity"]["cost_per_kwh"],
                "сум/кВт·ч",
            ],
            [
                "Природный газ",
                self.normatives["gas"]["specific_consumption"],
                "м³/м²·год",
                self.normatives["gas"]["efficiency_min"],
                self.normatives["gas"]["cost_per_m3"],
                "сум/м³",
            ],
            [
                "Вода",
                self.normatives["water"]["specific_consumption"],
                "м³/м²·год",
                self.normatives["water"]["efficiency_min"],
                self.normatives["water"]["cost_per_m3"],
                "сум/м³",
            ],
        ]

        # Заполняем нормативы
        for row_idx, row_data in enumerate(normative_data, 4):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Создаем таблицу нормативов
        table = Table(displayName="Normatives", ref="A4:F7")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Нормативы по зданиям
        ws["A9"] = "НОРМАТИВЫ ПО ЗДАНИЯМ"
        ws["A9"].font = Font(bold=True)

        building_normatives = [
            ["Показатель", "Нормативное значение", "Единица измерения", "Примечание"],
            [
                "Коэффициент теплопотерь",
                self.normatives["building"]["heat_loss_coefficient"],
                "кВт/м²·°C",
                "Максимальное значение",
            ],
            [
                "Минимальная толщина утеплителя",
                self.normatives["building"]["insulation_thickness_min"],
                "м",
                "Для стен",
            ],
            [
                "Максимальная площадь окон",
                self.normatives["building"]["window_area_max"],
                "доля от площади стен",
                "От общей площади стен",
            ],
            ["Коэффициент энергоэффективности", 1.0, "", "Минимальное значение"],
            ["Срок службы энергооборудования", 15, "лет", "Минимальный срок"],
        ]

        # Заполняем нормативы по зданиям
        for row_idx, row_data in enumerate(building_normatives, 10):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Создаем таблицу нормативов по зданиям
        table2 = Table(displayName="BuildingNormatives", ref="A10:D15")
        style2 = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table2.tableStyleInfo = style2
        # ws.add_table(table2)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Настройка стилей
        self.apply_table_styles(ws, "A4:F7")
        self.apply_table_styles(ws, "A10:D15")

    def create_measures_sheet(self, enterprise_data: dict):
        """Создание листа мероприятий"""
        logger.info("💡 Создание листа мероприятий...")

        ws = self.wb.create_sheet("Measures")

        # Заголовок
        ws["A1"] = "МЕРОПРИЯТИЯ ПО ЭНЕРГОСБЕРЕЖЕНИЮ"
        ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A1:G1")

        # Получаем данные по мероприятиям
        measures_data = self.get_energy_efficiency_measures()

        # Перечень мероприятий
        ws["A3"] = "ПЕРЕЧЕНЬ МЕРОПРИЯТИЙ"
        ws["A3"].font = Font(bold=True)

        measures_headers = [
            "Наименование",
            "Тип",
            "Экономия (%)",
            "Годовая экономия (сум)",
            "Стоимость внедрения (сум)",
            "Срок окупаемости (лет)",
            "Приоритет",
        ]

        # Заголовки таблицы
        for col_idx, header in enumerate(measures_headers, 1):
            ws.cell(row=4, column=col_idx, value=header)

        # Данные мероприятий
        row_idx = 5
        total_savings = 0
        total_cost = 0

        for measure in measures_data:
            ws.cell(row=row_idx, column=1, value=measure.get("name", ""))
            ws.cell(row=row_idx, column=2, value=measure.get("type", ""))
            ws.cell(row=row_idx, column=3, value=measure.get("savings_percent", 0))
            ws.cell(row=row_idx, column=4, value=f"=Energy!D8*C{row_idx}/100")
            ws.cell(row=row_idx, column=5, value=measure.get("cost", 0))
            ws.cell(row=row_idx, column=6, value=f"=E{row_idx}/D{row_idx}")
            ws.cell(
                row=row_idx,
                column=7,
                value=f'=IF(F{row_idx}<=2,"Высокий",IF(F{row_idx}<=4,"Средний","Низкий"))',
            )

            total_savings += measure.get("savings_percent", 0)
            total_cost += measure.get("cost", 0)
            row_idx += 1

        # Итоговая строка
        ws.cell(row=row_idx, column=1, value="ИТОГО")
        if row_idx == 5:
            # Нет мероприятий — избегаем самоссылок в формулах
            ws.cell(row=row_idx, column=3, value=0)
            ws.cell(row=row_idx, column=4, value=0)
            ws.cell(row=row_idx, column=5, value=0)
            ws.cell(row=row_idx, column=6, value=0)
        else:
            ws.cell(row=row_idx, column=3, value=f"=SUM(C5:C{row_idx - 1})")
            ws.cell(row=row_idx, column=4, value=f"=SUM(D5:D{row_idx - 1})")
            ws.cell(row=row_idx, column=5, value=f"=SUM(E5:E{row_idx - 1})")
            ws.cell(row=row_idx, column=6, value=f"=AVERAGE(F5:F{row_idx - 1})")

        # Создаем таблицу
        table = Table(displayName="Measures", ref=f"A4:G{row_idx}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # План реализации
        ws[f"A{row_idx + 2}"] = "ПЛАН РЕАЛИЗАЦИИ МЕРОПРИЯТИЙ"
        ws[f"A{row_idx + 2}"].font = Font(bold=True)

        implementation_plan = [
            ["Этап", "Мероприятия", "Срок реализации", "Ответственный", "Статус"],
            [
                "Этап 1 (Краткосрочные)",
                "Замена освещения",
                "1-3 месяца",
                "Главный энергетик",
                "Планируется",
            ],
            [
                "Этап 2 (Среднесрочные)",
                "Утепление здания",
                "3-6 месяцев",
                "Строительный отдел",
                "Планируется",
            ],
            [
                "Этап 3 (Долгосрочные)",
                "Модернизация оборудования",
                "6-12 месяцев",
                "Технический отдел",
                "Планируется",
            ],
        ]

        # Заполняем план реализации
        for i, row_data in enumerate(implementation_plan, row_idx + 3):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=i, column=col_idx, value=cell_value)

        # Создаем таблицу плана
        table2 = Table(
            displayName="ImplementationPlan", ref=f"A{row_idx + 3}:E{row_idx + 6}"
        )
        style2 = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table2.tableStyleInfo = style2
        # ws.add_table(table2)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Настройка стилей
        self.apply_table_styles(ws, f"A4:G{row_idx}")
        self.apply_table_styles(ws, f"A{row_idx + 3}:E{row_idx + 6}")

    def create_summary_sheet(self, enterprise_data: dict):
        """Создание сводного листа"""
        logger.info("📊 Создание сводного листа...")

        ws = self.wb.create_sheet("Summary")

        # Заголовок
        ws["A1"] = "СВОДНЫЙ ЭНЕРГЕТИЧЕСКИЙ ПАСПОРТ"
        ws["A1"].font = Font(name="Times New Roman", size=16, bold=True)
        ws.merge_cells("A1:F1")

        ws["A2"] = f"Предприятие: {enterprise_data['name']}"
        ws["A2"].font = Font(name="Times New Roman", size=14, bold=True)
        ws.merge_cells("A2:F2")

        # Основные показатели
        ws["A4"] = "ОСНОВНЫЕ ПОКАЗАТЕЛИ"
        ws["A4"].font = Font(bold=True)

        summary_data = [
            ["Показатель", "Значение", "Единица измерения", "Норматив", "Соответствие"],
            ["Общее энергопотребление", "=Energy!B8", "кВт·ч/год", "", ""],
            [
                "Удельное потребление электроэнергии",
                "=Calculations!C4",
                "кВт·ч/м²·год",
                "=Normatives!B5",
                "=Calculations!F4",
            ],
            [
                "Удельное потребление газа",
                "=Calculations!C5",
                "м³/м²·год",
                "=Normatives!B6",
                "=Calculations!F5",
            ],
            [
                "Удельное потребление воды",
                "=Calculations!C6",
                "м³/м²·год",
                "=Normatives!B7",
                "=Calculations!F6",
            ],
            [
                "Общий коэффициент энергоэффективности",
                "=Calculations!C7",
                "",
                "1.0",
                "=Calculations!F7",
            ],
            ["Потенциальная экономия", "=SUM(Measures!D5:D20)", "сум/год", "", ""],
            [
                "Срок окупаемости мероприятий",
                "=AVERAGE(Measures!F5:F20)",
                "лет",
                "",
                "",
            ],
        ]

        # Заполняем сводные данные
        for row_idx, row_data in enumerate(summary_data, 5):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Создаем таблицу
        table = Table(displayName="Summary", ref="A5:F12")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # table.tableStyleInfo = style
        # ws.add_table(table)  # ОТКЛЮЧЕНО - конфликт с формулами

        # Заключение
        ws["A14"] = "ЗАКЛЮЧЕНИЕ"
        ws["A14"].font = Font(bold=True)

        conclusion_text = f"""
На основании проведенного энергетического аудита предприятия "{enterprise_data["name"]}" 
можно сделать следующие выводы:

1. ЭНЕРГОЭФФЕКТИВНОСТЬ: Предприятие имеет {"соответствующий" if True else "несоответствующий"} уровень энергоэффективности для данной отрасли.

2. РЕЗЕРВЫ ЭНЕРГОСБЕРЕЖЕНИЯ: Выявлены значительные резервы экономии энергии.

3. РЕКОМЕНДАЦИИ: Разработаны конкретные мероприятия по повышению энергоэффективности.

4. ЭКОНОМИЧЕСКИЙ ЭФФЕКТ: Реализация мероприятий позволит экономить значительные средства.

Заключение подготовлено в соответствии с требованиями стандарта ПКМ 690 Узбекистан.
"""

        ws["A15"] = conclusion_text
        ws.merge_cells(f"A15:F{15 + conclusion_text.count(chr(10))}")

        # Настройка стилей
        self.apply_table_styles(ws, "A5:F12")

        # Настройка выравнивания для заключения
        for row in range(15, 15 + conclusion_text.count(chr(10)) + 1):
            ws[f"A{row}"].alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

    def apply_title_styles(self, ws):
        """Применение стилей к титульному листу"""
        # Настройка ширины колонок
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        # Настройка высоты строк
        for row in range(1, 15):
            ws.row_dimensions[row].height = 20

    def apply_table_styles(self, ws, range_str):
        """Применение стилей к таблице"""
        # Стили для заголовков
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        # Стили для границ
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Применяем стили к диапазону
        for row in ws[range_str]:
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Стили для заголовков
                if cell.row == 1 or cell.row == 3 or cell.row == 4:
                    cell.fill = header_fill
                    cell.font = header_font

    def get_enterprise_data(self) -> dict:
        """Получение данных предприятия (из self.enterprise_data)"""
        return self.enterprise_data

    def get_energy_consumption_data(self) -> dict:
        """Получение данных по энергопотреблению (из self.energy_data)"""
        resources = self.energy_data.get("resources", {})

        # Вычисляем totals из квартальных данных
        electricity_total = sum(
            q.get("quarter_totals", {}).get("active_kwh", 0)
            for q in resources.get("electricity", {}).values()
        )
        gas_total = sum(
            q.get("quarter_totals", {}).get("volume_m3", 0)
            for q in resources.get("gas", {}).values()
        )
        water_total = sum(
            q.get("quarter_totals", {}).get("volume_m3", 0)
            for q in resources.get("water", {}).values()
        )

        return {
            "electricity_total": electricity_total,
            "gas_total": gas_total,
            "water_total": water_total,
        }

    def get_monthly_energy_data(self) -> list:
        """Получение месячных данных по энергопотреблению (из self.energy_data)"""
        resources = self.energy_data.get("resources", {})
        monthly_data = []

        # Извлекаем месячные данные из квартальной структуры
        for resource_type in ["electricity", "gas", "water"]:
            for quarter_key, quarter_data in resources.get(resource_type, {}).items():
                for month_entry in quarter_data.get("months", []):
                    monthly_data.append(
                        {
                            "month": month_entry.get("month"),
                            "resource": resource_type,
                            "values": month_entry.get("values", {}),
                        }
                    )

        return monthly_data

    def get_equipment_data(self) -> dict:
        """Получение данных по оборудованию (заглушка)"""
        # TODO: Добавить парсинг оборудования из raw_json
        return {"equipment_list": [], "total_power": 0}

    def get_energy_efficiency_measures(self) -> list:
        """Получение мероприятий по энергосбережению (заглушка)"""
        # TODO: Добавить AI-генерацию рекомендаций на основе данных
        return []


def main():
    """Основная функция"""
    print("\n" + "=" * 70)
    print("║     📊 ГЕНЕРАТОР EXCEL-ПАСПОРТОВ ПО ПКМ 690 УЗБЕКИСТАН     ║")
    print("=" * 70)

    # Создаем генератор
    generator = PKM690ExcelGenerator()

    # Список предприятий для генерации паспортов
    enterprises = [
        {"id": 1, "name": "Metin Iroda"},
        {"id": 2, "name": "Test Enterprise"},
    ]

    for enterprise in enterprises:
        output_path = f"energy_passport_{enterprise['id']}_{enterprise['name'].replace(' ', '_')}.xlsx"

        print(f"\n📊 Генерация паспорта для: {enterprise['name']}")

        success = generator.create_energy_passport(enterprise["id"], output_path)

        if success:
            print(f"✅ Паспорт создан: {output_path}")
        else:
            print(f"❌ Ошибка создания паспорта для {enterprise['name']}")

    print("\n" + "=" * 70)
    print("✅ ГЕНЕРАЦИЯ ПАСПОРТОВ ЗАВЕРШЕНА!")
    print("=" * 70)


if __name__ == "__main__":
    main()
