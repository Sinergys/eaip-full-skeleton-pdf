"""Детальный анализ структуры листа Структура пр 2"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=False)
ws = wb["Структура пр 2"]

print("=" * 80)
print("АНАЛИЗ СТРУКТУРЫ ЛИСТА 'Структура пр 2'")
print("=" * 80)

# Ищем заголовки с годами
print("\n1. Поиск заголовков с годами (строки 1-8):")
for row in range(1, 9):
    for col in range(1, 200):
        val = ws.cell(row, col).value
        if isinstance(val, int) and val in (2022, 2023, 2024):
            col_letter = get_column_letter(col)
            print(f"  Год {val} найден: {col_letter}{row}")

# Анализируем строку 32 (где E32)
print("\n2. Анализ строки 32 (где находится E32):")
print("   Заголовок строки 32:", ws.cell(32, 1).value)
for col in range(1, 200):
    val = ws.cell(32, col).value
    if val is not None and (isinstance(val, (int, float)) or (isinstance(val, str) and len(str(val)) < 50)):
        col_letter = get_column_letter(col)
        cell = ws.cell(32, col)
        data_type = cell.data_type
        print(f"   {col_letter}32: {val} (тип: {data_type})")

# Анализируем строку 9 (где по маппингу должны быть данные)
print("\n3. Анализ строки 9 (ожидаемое расположение по маппингу):")
print("   Заголовок строки 9:", ws.cell(9, 1).value)
for col in range(1, 200):
    val = ws.cell(9, col).value
    if val is not None and isinstance(val, (int, float)):
        col_letter = get_column_letter(col)
        print(f"   {col_letter}9: {val}")

# Ищем паттерн: где находятся данные по газу для 2023
print("\n4. Поиск данных по газу для 2023:")
# Проверяем строки вокруг 32
for row in range(30, 40):
    val_a = ws.cell(row, 1).value
    if val_a and isinstance(val_a, str) and "газ" in val_a.lower():
        print(f"   Строка {row}: {val_a}")
        for col in range(1, 200):
            val = ws.cell(row, col).value
            if val is not None and isinstance(val, (int, float)):
                col_letter = get_column_letter(col)
                print(f"     {col_letter}{row}: {val}")

# Проверяем строку 32 более детально - какие колонки соответствуют каким кварталам
print("\n5. Определение структуры кварталов в строке 32:")
# Ищем паттерн: B32 (2022), E32 (2023 Q1?), Q32 (2023?), AF32 (2024?)
quarters_2023 = []
for col in range(1, 200):
    val = ws.cell(32, col).value
    if val is not None and isinstance(val, (int, float)) and 10 < val < 20:
        col_letter = get_column_letter(col)
        # Проверяем, что это может быть газ (значения в диапазоне 10-20 тыс. м³)
        print(f"   {col_letter}32: {val} (возможно газ 2023)")

wb.close()

