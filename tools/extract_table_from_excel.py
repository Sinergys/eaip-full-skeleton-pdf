"""
Извлечение таблицы данных из Excel файлов (без заголовков документа)
Автоматически находит начало таблицы и извлекает только данные
"""
import sys
from pathlib import Path
import json

try:
    import openpyxl
except ImportError:
    print("Требуется openpyxl: pip install openpyxl")
    sys.exit(1)

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

def find_table_start(sheet, min_cols=5):
    """
    Находит начало таблицы данных в Excel листе
    
    Ищет строку, которая содержит:
    - Заголовки таблицы (№, Наименование, Количество, Цена, и т.д.)
    - Минимум min_cols непустых ячеек
    """
    # Паттерны для поиска заголовков таблицы
    header_patterns = [
        '№', 'N', 'No', 'номер',
        'наименование', 'название', 'товар', 'услуга',
        'количество', 'кол-во', 'ед', 'единица',
        'цена', 'стоимость', 'сумма', 'ндс'
    ]
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_data = [str(cell).strip().lower() if cell else "" for cell in row]
        
        # Пропускаем пустые строки
        if not any(row_data):
            continue
        
        # Проверяем, содержит ли строка паттерны заголовков
        matches = sum(1 for cell in row_data if any(pattern in cell for pattern in header_patterns))
        
        # Если найдено достаточно совпадений и достаточно столбцов
        non_empty_cols = sum(1 for cell in row_data if cell)
        if matches >= 2 and non_empty_cols >= min_cols:
            return row_idx
    
    return None

def extract_table_data(sheet, start_row):
    """Извлекает данные таблицы начиная с указанной строки"""
    rows = []
    headers = []
    
    # Читаем заголовки (первая строка таблицы)
    header_row = list(sheet.iter_rows(values_only=True))[start_row - 1]
    headers = [str(cell).strip() if cell else "" for cell in header_row]
    
    # Читаем данные (начиная со следующей строки)
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True, min_row=start_row + 1), start_row + 1):
        row_data = [str(cell).strip() if cell else "" for cell in row]
        
        # Пропускаем полностью пустые строки
        if not any(row_data):
            continue
        
        # Останавливаемся, если встретили строку "Итого" или похожую
        if any(keyword in " ".join(row_data).lower() for keyword in ['итого', 'total', 'сумма', 'всего']):
            rows.append(row_data)
            break
        
        rows.append(row_data)
    
    return headers, rows

def normalize_table(headers, rows):
    """Нормализует таблицу: убирает пустые столбцы, выравнивает количество столбцов"""
    # Находим максимальное количество столбцов
    max_cols = max(len(headers), max((len(row) for row in rows), default=0))
    
    # Выравниваем заголовки
    headers = headers + [""] * (max_cols - len(headers))
    
    # Выравниваем строки
    normalized_rows = []
    for row in rows:
        normalized_row = row + [""] * (max_cols - len(row))
        normalized_rows.append(normalized_row)
    
    # Убираем полностью пустые столбцы справа
    while max_cols > 0 and not any(headers[max_cols - 1]) and not any(row[max_cols - 1] for row in normalized_rows):
        max_cols -= 1
    
    if max_cols < len(headers):
        headers = headers[:max_cols]
        normalized_rows = [row[:max_cols] for row in normalized_rows]
    
    return headers, normalized_rows

def extract_table_from_excel(file_path: str, output_path: str = None) -> dict:
    """Извлекает таблицу данных из Excel файла"""
    file_path_obj = Path(file_path)
    
    if not file_path_obj.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    workbook = openpyxl.load_workbook(file_path_obj, data_only=True)
    result = {
        "tables": [],
        "text": ""
    }
    
    # Обрабатываем первый лист (обычно там таблица)
    sheet = workbook.active
    
    # Находим начало таблицы
    start_row = find_table_start(sheet, min_cols=5)
    
    if start_row is None:
        print(f"⚠️  Не удалось найти начало таблицы в {file_path_obj.name}")
        print("   Пытаюсь использовать первую строку с данными...")
        # Пробуем использовать первую непустую строку
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if any(cell for cell in row if cell):
                start_row = row_idx
                break
    
    if start_row is None:
        raise ValueError(f"Не удалось найти данные в файле {file_path_obj.name}")
    
    print(f"✅ Найдено начало таблицы: строка {start_row}")
    
    # Извлекаем данные
    headers, rows = extract_table_data(sheet, start_row)
    
    # Нормализуем
    headers, rows = normalize_table(headers, rows)
    
    # Формируем результат
    table_data = {
        "rows": [headers] + rows,
        "headers": headers
    }
    
    result["tables"].append(table_data)
    
    # Сохраняем в JSON
    if output_path is None:
        output_path = file_path_obj.parent / f"{file_path_obj.stem}_table_extracted.json"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Таблица извлечена: {len(rows)} строк данных, {len(headers)} столбцов")
    
    return str(output_path), result

def main():
    """Основная функция"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Извлечение таблицы данных из Excel файлов')
    parser.add_argument('files', nargs='+', help='Пути к Excel файлам')
    parser.add_argument('--output-dir', type=str, help='Директория для сохранения JSON файлов')
    
    args = parser.parse_args()
    
    print("=" * 80)
    print("ИЗВЛЕЧЕНИЕ ТАБЛИЦЫ ИЗ EXCEL ФАЙЛОВ")
    print("=" * 80)
    print()
    
    output_dir = Path(args.output_dir) if args.output_dir else project_root / "reports" / "ocr"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    extracted_files = []
    
    for file_path in args.files:
        print(f"\n📄 Обработка: {file_path}")
        try:
            output_file = output_dir / f"{Path(file_path).stem}_table_extracted.json"
            extracted_path, data = extract_table_from_excel(file_path, str(output_file))
            extracted_files.append(extracted_path)
            
            # Показываем статистику
            if data.get('tables'):
                table = data['tables'][0]
                print(f"   📊 Статистика:")
                print(f"      - Строк (включая заголовок): {len(table.get('rows', []))}")
                print(f"      - Строк данных: {len(table.get('rows', [])) - 1}")
                print(f"      - Столбцов: {len(table.get('headers', []))}")
                print(f"      - Заголовки: {', '.join(table.get('headers', [])[:5])}...")
            
        except Exception as e:
            print(f"❌ Ошибка обработки {file_path}: {e}")
            import traceback
            traceback.print_exc()
    
    if extracted_files:
        print()
        print("=" * 80)
        print("✅ ИЗВЛЕЧЕНИЕ ЗАВЕРШЕНО")
        print("=" * 80)
        print()
        print("📋 Извлеченные файлы:")
        for f in extracted_files:
            print(f"  - {f}")
        print()
        print("📌 Теперь можно запустить сравнение:")
        if len(extracted_files) == 1:
            print(f"  python tools/compare_recognition_results.py --manual \"{extracted_files[0]}\"")
        else:
            print(f"  python tools/compare_recognition_results.py --manual1 \"{extracted_files[0]}\" --manual2 \"{extracted_files[1]}\"")

if __name__ == "__main__":
    main()

