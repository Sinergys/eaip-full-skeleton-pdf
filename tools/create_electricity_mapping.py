"""Создание полного маппинга для таблиц электроэнергии по всем кварталам"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=False)
ws = wb["Структура пр 2"]

print("=" * 80)
print("СОЗДАНИЕ МАППИНГА ДЛЯ ВСЕХ КВАРТАЛОВ")
print("=" * 80)

# Определяем структуру на основе анализа
# Строки для каждого квартала:
quarter_rows = {
    "Q1": 17,
    "Q2": 39,
    "Q3": 61,
    "Q4": 83,
}

# Колонки для каждого квартала (начало таблицы)
# На основе скриншота: A=1 (Q1), P=16 (Q2), AE=31 (Q3), AT=46 (Q4)
quarter_start_cols = {
    "Q1": 1,   # A
    "Q2": 16,  # P
    "Q3": 31,  # AE
    "Q4": 46,  # AT
}

# Структура колонок внутри каждой таблицы:
# A/P/AE/AT = название продукции
# B/Q/AF/AU = норма
# C/R/AG/AV = факт 2022
# D/S/AH/AW = факт 2023
# E/T/AI/AX = факт 2024
# F/U/AJ/AY = перерасход 2022
# G/V/AK/AZ = перерасход 2023
# H/W/AL/BA = перерасход 2024

print("\nМаппинг для всех кварталов:\n")

mapping = {}

for quarter, start_row in quarter_rows.items():
    start_col = quarter_start_cols[quarter]
    print(f"{quarter} (строка {start_row}, начинается с колонки {get_column_letter(start_col)}):")
    
    for year in [2022, 2023, 2024]:
        # Для каждого года определяем смещение колонок
        # Q1: 2022=B-H, 2023=?, 2024=?
        # Q2: 2022=Q-W, 2023=?, 2024=?
        # Q3: 2022=AE-AL, 2023=?, 2024=?
        # Q4: 2022=AT-BA, 2023=?, 2024=?
        
        # Пока используем структуру: каждый квартал занимает 15 колонок
        # Q1: A-O (1-15), Q2: P-AD (16-30), Q3: AE-AS (31-45), Q4: AT-BN (46-60)
        # Внутри каждого квартала: 2022, 2023, 2024 идут подряд
        
        if quarter == "Q1":
            if year == 2022:
                col_offset = 1  # B (норма), C (2022), D (2023), E (2024)
            elif year == 2023:
                col_offset = 9  # J, K, L, M
            else:  # 2024
                col_offset = 13  # N, O, P, Q
        elif quarter == "Q2":
            if year == 2022:
                col_offset = 16  # Q (норма), R (2022), S (2023), T (2024)
            elif year == 2023:
                col_offset = 24  # Y, Z, AA, AB
            else:  # 2024
                col_offset = 28  # AC, AD, AE, AF
        elif quarter == "Q3":
            if year == 2022:
                col_offset = 31  # AE (норма), AF (2022), AG (2023), AH (2024)
            elif year == 2023:
                col_offset = 39  # AM, AN, AO, AP
            else:  # 2024
                col_offset = 43  # AQ, AR, AS, AT
        else:  # Q4
            if year == 2022:
                col_offset = 46  # AT (норма), AU (2022), AV (2023), AW (2024)
            elif year == 2023:
                col_offset = 54  # BC, BD, BE, BF
            else:  # 2024
                col_offset = 58  # BG, BH, BI, BJ
        
        key = f"{year}-{quarter}"
        mapping[key] = {
            "row": start_row,
            "col_norm": col_offset,
            "col_2022": col_offset + 1,
            "col_2023": col_offset + 2,
            "col_2024": col_offset + 3,
            "col_overrun_2022": col_offset + 4,
            "col_overrun_2023": col_offset + 5,
            "col_overrun_2024": col_offset + 6,
        }
        
        print(f"  {year}: норма={get_column_letter(col_offset)}, "
              f"2022={get_column_letter(col_offset+1)}, "
              f"2023={get_column_letter(col_offset+2)}, "
              f"2024={get_column_letter(col_offset+3)}")

print("\n" + "=" * 80)
print("ПРОВЕРКА: ищем фактические колонки в шаблоне")
print("=" * 80)

# Проверяем, есть ли данные в других колонках строки 17
print("\nСтрока 17 - все непустые колонки:")
for col in range(1, 100):
    val = ws.cell(17, col).value
    if val is not None:
        col_letter = get_column_letter(col)
        print(f"  {col_letter}17: {val}")

wb.close()

print("\n" + "=" * 80)
print("РЕКОМЕНДУЕМЫЙ МАППИНГ (на основе структуры шаблона)")
print("=" * 80)
print("""
Структура:
- Q1: строки 17-21, колонки A-H (2022), J-M (2023?), N-Q (2024?)
- Q2: строки 39-43, колонки P-W (2022), Y-AB (2023?), AC-AF (2024?)
- Q3: строки 61-65, колонки AE-AL (2022), AM-AP (2023?), AQ-AT (2024?)
- Q4: строки 83-87, колонки AT-BA (2022), BC-BF (2023?), BG-BJ (2024?)

НО: нужно проверить фактическую структуру в шаблоне!
""")

