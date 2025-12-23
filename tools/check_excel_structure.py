"""
Проверка структуры Excel файлов для правильного чтения
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Требуется openpyxl: pip install openpyxl")
    sys.exit(1)

def check_excel_structure(file_path: str):
    """Проверяет структуру Excel файла"""
    print(f"\n{'='*80}")
    print(f"Проверка файла: {file_path}")
    print(f"{'='*80}\n")
    
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Лист: {sheet_name}")
        print(f"Размер: {sheet.max_row} строк × {sheet.max_column} столбцов\n")
        
        # Показываем первые 10 строк
        print("Первые 10 строк:")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True, max_row=10), 1):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            # Показываем только непустые ячейки
            non_empty = [f"{i}:{val[:50]}" for i, val in enumerate(row_data) if val]
            if non_empty:
                print(f"  Строка {row_idx}: {', '.join(non_empty)}")
        
        print()

if __name__ == "__main__":
    files = [
        r"C:\AUDIT\OBJECTS\Navoiy IES\INBOX\акт выполненых работ май.xlsx",
        r"C:\AUDIT\OBJECTS\Navoiy IES\INBOX\акт выполненых работ май1.xlsx"
    ]
    
    for file_path in files:
        if Path(file_path).exists():
            check_excel_structure(file_path)
        else:
            print(f"Файл не найден: {file_path}")

