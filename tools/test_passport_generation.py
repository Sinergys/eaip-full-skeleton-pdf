"""
Тестовая генерация паспорта с шаблоном new_energy_passport
"""
import sys
from pathlib import Path

# Добавляем пути
sys.path.insert(0, str(Path(__file__).parent.parent))

from templates.pcm690.templates_config import get_template_path
from openpyxl import load_workbook

def test_template_structure():
    """Проверяет структуру шаблона new_energy_passport"""
    
    print("=" * 80)
    print("ТЕСТ СТРУКТУРЫ ШАБЛОНА new_energy_passport")
    print("=" * 80)
    
    try:
        template_path = get_template_path("new_energy_passport")
        print(f"\n✅ Шаблон найден: {template_path}")
        
        wb = load_workbook(template_path, data_only=False)
        print(f"\n📊 Листов в шаблоне: {len(wb.sheetnames)}")
        print("\nСписок листов:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            formulas = sum(1 for row in ws.iter_rows() for cell in row if cell.data_type == 'f' and cell.value)
            data_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
            print(f"  {i}. {sheet_name}")
            print(f"     - Строк: {ws.max_row}, Колонок: {ws.max_column}")
            print(f"     - Формул: {formulas}, Ячеек с данными: {data_cells}")
        
        # Проверка ключевых листов
        print("\n" + "=" * 80)
        print("ПРОВЕРКА КЛЮЧЕВЫХ ЛИСТОВ")
        print("=" * 80)
        
        key_sheets = {
            "Структура": ["Структура пр 2", "Struktura pr2", "02_Структура"],
            "Узлы учета": ["01_Узлы учета", "Узел учета", "Узлы учета", "Nodes"],
            "Баланс": ["04_Баланс", "Баланс", "Balance", "Balans"],
            "Оборудование": ["03_Оборудование", "Equipment", "Оборудование", "Sheet1"],
            "Динамика": ["05_Динамика", "Динамика", "Dinamika sr", "Динамика ср"],
            "Мероприятия": ["06_Мероприятия", "Мероприятия", "Meropriyatiya", "Мериаприятия 1"]
        }
        
        found_sheets = {}
        for category, names in key_sheets.items():
            found = None
            for name in names:
                if name in wb.sheetnames:
                    found = name
                    break
            found_sheets[category] = found
            status = "✅" if found else "❌"
            print(f"{status} {category}: {found if found else 'НЕ НАЙДЕН'}")
        
        wb.close()
        
        return found_sheets
        
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    test_template_structure()

