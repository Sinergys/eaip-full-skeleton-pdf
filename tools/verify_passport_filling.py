"""Проверка заполнения энергопаспорта"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")

print("=" * 80)
print("ПРОВЕРКА ЗАПОЛНЕНИЯ ЭНЕРГОПАСПОРТА")
print("=" * 80)

if not output_path.exists():
    print(f"❌ Файл не найден: {output_path}")
    exit(1)

wb = load_workbook(output_path, data_only=True)
ws = wb["Структура пр 2"]

print(f"✅ Файл загружен: {output_path.name}\n")

# Проверка E32 (2023 Q1 газ)
print("=" * 80)
print("1. ПРОВЕРКА ГАЗА (E32 - 2023 Q1)")
print("=" * 80)
e32_value = ws["E32"].value
print(f"E32 (2023 Q1 газ): {e32_value}")
if e32_value and abs(e32_value - 14.819) < 0.001:
    print("✅ ОШИБКА ИСПРАВЛЕНА! Значение правильное: 14.819 тыс. м³")
else:
    print(f"⚠️ Значение: {e32_value} (ожидается 14.819)")

# Проверка электроэнергии по кварталам
print("\n" + "=" * 80)
print("2. ПРОВЕРКА ЭЛЕКТРОЭНЕРГИИ ПО ВСЕМ КВАРТАЛАМ")
print("=" * 80)

quarters = {
    "Q1": {"rows": [17, 18, 19, 20, 21], "start_col": 1, "name": "Q1 (строки 17-21, колонки B-H)"},
    "Q2": {"rows": [39, 40, 41, 42, 43], "start_col": 16, "name": "Q2 (строки 39-43, колонки Q-W)"},
    "Q3": {"rows": [61, 62, 63, 64, 65], "start_col": 31, "name": "Q3 (строки 61-65, колонки AF-AL)"},
    "Q4": {"rows": [83, 84, 85, 86, 87], "start_col": 46, "name": "Q4 (строки 83-87, колонки AU-BA)"},
}

products = ["труба ХВС", "фитинги ХВС и ГВС", "канализационные трубы", "канализационные фитинги", "тёплый пол"]

total_filled = 0
total_expected = 0

for quarter, info in quarters.items():
    print(f"\n{info['name']}:")
    print("-" * 60)
    
    start_col = info["start_col"]
    rows = info["rows"]
    
    quarter_filled = 0
    quarter_expected = 0
    
    for i, (row, product) in enumerate(zip(rows, products)):
        col_norm = start_col + 1
        col_2022 = start_col + 2
        col_2023 = start_col + 3
        col_2024 = start_col + 4
        
        norm = ws.cell(row, col_norm).value
        fact_2022 = ws.cell(row, col_2022).value
        fact_2023 = ws.cell(row, col_2023).value
        fact_2024 = ws.cell(row, col_2024).value
        
        # Проверяем заполненность
        filled_count = sum(1 for v in [norm, fact_2022, fact_2023, fact_2024] if v is not None and v != 0)
        expected_count = 4
        
        if filled_count > 0:
            status = "✅"
            quarter_filled += filled_count
        else:
            status = "❌"
        
        quarter_expected += expected_count
        
        print(f"  {status} {product} (строка {row}):")
        print(f"     Норма ({get_column_letter(col_norm)}): {norm}")
        print(f"     Факт 2022 ({get_column_letter(col_2022)}): {fact_2022}")
        print(f"     Факт 2023 ({get_column_letter(col_2023)}): {fact_2023}")
        print(f"     Факт 2024 ({get_column_letter(col_2024)}): {fact_2024}")
        print(f"     Заполнено: {filled_count}/{expected_count}")
    
    total_filled += quarter_filled
    total_expected += quarter_expected
    
    print(f"\n  Итого по {quarter}: {quarter_filled}/{quarter_expected} полей заполнено")

print("\n" + "=" * 80)
print("ИТОГОВАЯ СТАТИСТИКА")
print("=" * 80)
print(f"Всего заполнено: {total_filled}/{total_expected} полей ({total_filled*100/total_expected:.1f}%)")

if total_filled == total_expected:
    print("✅ ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ ПОЛНОСТЬЮ!")
elif total_filled > 0:
    print(f"⚠️ Заполнено частично: {total_filled}/{total_expected}")
else:
    print("❌ КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")

# Проверка категорий газа
print("\n" + "=" * 80)
print("3. ПРОВЕРКА КАТЕГОРИЙ ГАЗА")
print("=" * 80)
e34_value = ws["E34"].value
e36_value = ws["E36"].value
print(f"E34 (собственные нужды 2023 Q1): {e34_value}")
print(f"E36 (хоз-быт 2023 Q1): {e36_value}")
if e34_value and e36_value:
    total = (e34_value or 0) + (e36_value or 0)
    print(f"Сумма: {total} (ожидается ~14.819)")
    if abs(total - 14.819) < 1.0:
        print("✅ Категории газа заполнены корректно")

wb.close()

print("\n" + "=" * 80)
print("ПРОВЕРКА ЗАВЕРШЕНА")
print("=" * 80)

