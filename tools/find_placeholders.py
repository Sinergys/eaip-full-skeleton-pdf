"""
Утилита для поиска placeholder'ов в Excel шаблоне энергопаспорта.
Ищет значения в формате {{key}} или других паттернах.
"""

import re
import json
from pathlib import Path
from typing import Dict, List, Set
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import argparse


def find_placeholders_in_cell(cell_value: any) -> Set[str]:
    """Найти все placeholder'ы в значении ячейки."""
    placeholders = set()

    if cell_value is None:
        return placeholders

    cell_str = str(cell_value)

    # Паттерн для {{key}} или {{key.subkey}}
    pattern1 = r"\{\{([^}]+)\}\}"
    matches1 = re.findall(pattern1, cell_str)
    placeholders.update(matches1)

    # Паттерн для [key] или [key.subkey]
    pattern2 = r"\[([^\]]+)\]"
    matches2 = re.findall(pattern2, cell_str)
    placeholders.update(matches2)

    # Паттерн для <key> или <key.subkey>
    pattern3 = r"<([^>]+)>"
    matches3 = re.findall(pattern3, cell_str)
    placeholders.update(matches3)

    # Паттерн для подчеркиваний: ____, ___, __ (плейсхолдеры для заполнения)
    pattern4 = r"_+"
    matches4 = re.findall(pattern4, cell_str)
    if matches4:
        # Создаем описательные имена на основе контекста
        context = cell_str.lower()
        if "год" in context or "year" in context:
            placeholders.add("year")
        if "квартал" in context or "quarter" in context:
            placeholders.add("quarter")
        if "месяц" in context or "month" in context:
            placeholders.add("month")
        if matches4:
            placeholders.add("fill_blank")

    # Паттерн для "20___" или "20__" (год с подчеркиваниями)
    pattern5 = r"20_+"
    if re.search(pattern5, cell_str):
        placeholders.add("year")

    return placeholders


def find_placeholders_in_workbook(template_path: Path) -> Dict[str, List[Dict]]:
    """
    Найти все placeholder'ы в Excel файле.

    Returns:
        Dict с ключами по именам листов, значения - списки найденных placeholder'ов
        с информацией о ячейке (адрес, значение, тип)
    """
    workbook = load_workbook(template_path, data_only=False)
    all_placeholders = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_placeholders = []

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                # Пропускаем объединенные ячейки (они read-only)
                if isinstance(cell, MergedCell):
                    continue

                cell_value = cell.value
                if cell_value is None:
                    continue

                # Проверяем значение ячейки
                placeholders = find_placeholders_in_cell(cell_value)

                # Проверяем формулу, если есть
                if cell.data_type == "f" and cell.value:
                    formula = str(cell.value)
                    formula_placeholders = find_placeholders_in_cell(formula)
                    placeholders.update(formula_placeholders)

                if placeholders:
                    cell_address = cell.coordinate
                    sheet_placeholders.append(
                        {
                            "cell": cell_address,
                            "row": row_idx,
                            "column": col_idx,
                            "value": str(cell_value)[:200],  # Ограничиваем длину
                            "placeholders": sorted(list(placeholders)),
                            "data_type": cell.data_type,
                        }
                    )

        if sheet_placeholders:
            all_placeholders[sheet_name] = sheet_placeholders

    return all_placeholders


def main():
    parser = argparse.ArgumentParser(description="Найти placeholder'ы в Excel шаблоне")
    parser.add_argument("--template", help="Путь к шаблону Excel")
    parser.add_argument("--template-name", help="Имя шаблона из templates_config")
    parser.add_argument("--output", help="Путь для сохранения JSON с результатами")
    parser.add_argument(
        "--format",
        choices=["json", "summary", "detailed"],
        default="summary",
        help="Формат вывода: json, summary, detailed",
    )

    args = parser.parse_args()

    # Определение пути к шаблону
    if args.template_name:
        try:
            import sys
            from pathlib import Path

            templates_config_path = (
                Path(__file__).parent.parent / "templates" / "pcm690"
            )
            if str(templates_config_path) not in sys.path:
                sys.path.insert(0, str(templates_config_path))
            from templates_config import get_template_path

            template_path = get_template_path(args.template_name)
        except ImportError:
            raise ImportError(
                "Не удалось импортировать templates_config. "
                "Убедитесь, что файл templates/pcm690/templates_config.py существует."
            )
    elif args.template:
        template_path = Path(args.template)
    else:
        raise ValueError("Необходимо указать либо --template-name, либо --template")

    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    print(f"Анализ шаблона: {template_path}")
    print("=" * 80)

    placeholders = find_placeholders_in_workbook(template_path)

    # Подсчет уникальных placeholder'ов
    all_unique_placeholders = set()
    total_count = 0
    for sheet_data in placeholders.values():
        for item in sheet_data:
            all_unique_placeholders.update(item["placeholders"])
            total_count += len(item["placeholders"])

    # Вывод результатов
    if args.format == "json":
        output_data = {
            "template_path": str(template_path),
            "total_unique_placeholders": len(all_unique_placeholders),
            "total_occurrences": total_count,
            "unique_placeholders": sorted(list(all_unique_placeholders)),
            "by_sheet": placeholders,
        }
        output_str = json.dumps(output_data, ensure_ascii=False, indent=2)
        print(output_str)

        if args.output:
            Path(args.output).write_text(output_str, encoding="utf-8")
            print(f"\nРезультаты сохранены в: {args.output}")

    elif args.format == "summary":
        print("\n📊 Сводка:")
        print(f"  Всего уникальных placeholder'ов: {len(all_unique_placeholders)}")
        print(f"  Всего вхождений: {total_count}")
        print(f"  Листов с placeholder'ами: {len(placeholders)}")

        print(f"\n📋 Уникальные placeholder'ы ({len(all_unique_placeholders)}):")
        for placeholder in sorted(all_unique_placeholders):
            print(f"  - {placeholder}")

        print("\n📄 По листам:")
        for sheet_name, sheet_data in placeholders.items():
            sheet_placeholders = set()
            for item in sheet_data:
                sheet_placeholders.update(item["placeholders"])
            print(
                f"  {sheet_name}: {len(sheet_data)} ячеек, {len(sheet_placeholders)} уникальных placeholder'ов"
            )

        if args.output:
            output_data = {
                "template_path": str(template_path),
                "total_unique_placeholders": len(all_unique_placeholders),
                "total_occurrences": total_count,
                "unique_placeholders": sorted(list(all_unique_placeholders)),
                "by_sheet": placeholders,
            }
            Path(args.output).write_text(
                json.dumps(output_data, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            print(f"\nРезультаты сохранены в: {args.output}")

    elif args.format == "detailed":
        print("\n📊 Сводка:")
        print(f"  Всего уникальных placeholder'ов: {len(all_unique_placeholders)}")
        print(f"  Всего вхождений: {total_count}")
        print(f"  Листов с placeholder'ами: {len(placeholders)}")

        for sheet_name, sheet_data in placeholders.items():
            print(f"\n📄 Лист: {sheet_name}")
            print(f"  Найдено {len(sheet_data)} ячеек с placeholder'ами")
            for item in sheet_data:
                print(f"    {item['cell']}: {', '.join(item['placeholders'])}")
                if len(item["value"]) < 100:
                    print(f"      Значение: {item['value']}")


if __name__ == "__main__":
    main()
