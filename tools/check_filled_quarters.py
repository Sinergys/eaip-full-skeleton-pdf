"""Проверка заполнения всех кварталов"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

output_path = r"test_output_metin\generated_passport_fixed.xlsx"

wb = load_workbook(output_path, data_only=True)
ws = wb["Структура пр 2"]

print("=" * 80)
print("ПРОВЕРКА ЗАПОЛНЕНИЯ ВСЕХ КВАРТАЛОВ")
print("=" * 80)

# Проверяем все кварталы
quarters = {
    "Q1": {"rows": [17, 18, 19, 20, 21], "start_col": 1},
    "Q2": {"rows": [39, 40, 41, 42, 43], "start_col": 16},
    "Q3": {"rows": [61, 62, 63, 64, 65], "start_col": 31},
    "Q4": {"rows": [83, 84, 85, 86, 87], "start_col": 46},
}

products = ["труба ХВС", "фитинги ХВС и ГВС", "канализационные трубы", "канализационные фитинги", "тёплый пол"]

for quarter, info in quarters.items():
    print(f"\n{'='*60}")
    print(f"КВАРТАЛ {quarter}")
    print(f"{'='*60}")
    
    start_col = info["start_col"]
    rows = info["rows"]
    
    for i, (row, product) in enumerate(zip(rows, products)):
        col_norm = start_col + 1
        col_2022 = start_col + 2
        
        norm = ws.cell(row, col_norm).value
        fact_2022 = ws.cell(row, col_2022).value
        
        status = "✅" if (norm and norm > 0) or (fact_2022 and fact_2022 > 0) else "❌"
        print(f"{status} {product} (строка {row}): норма={norm}, факт 2022={fact_2022}")

wb.close()

