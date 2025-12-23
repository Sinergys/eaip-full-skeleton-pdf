"""Проверка структуры файла газа"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = Path("data/source_files/audit_sinergys/Расчет газа для отопл и неотпл.xlsx")

results = []

results.append("=" * 80)
results.append("АНАЛИЗ ФАЙЛА: Расчет газа для отопл и неотпл.xlsx")
results.append("=" * 80)

if not file_path.exists():
    results.append(f"❌ Файл не найден: {file_path}")
    print('\n'.join(results))
    sys.exit(1)

wb = load_workbook(file_path, data_only=True)

results.append(f"\n✅ Файл найден")
results.append(f"Листы: {', '.join(wb.sheetnames)}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    results.append(f"\n{'='*80}")
    results.append(f"ЛИСТ: {sheet_name}")
    results.append(f"{'='*80}")
    
    # Анализ структуры
    results.append(f"\nРазмер: {ws.max_row} строк, {ws.max_column} колонок")
    
    # Первые 15 строк
    results.append(f"\nПервые 15 строк:")
    for row in range(1, min(16, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(25, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val is not None:
                col_letter = get_column_letter(col)
                if isinstance(val, (int, float)):
                    row_data.append(f"{col_letter}{row}={val}")
                elif isinstance(val, str) and len(str(val)) < 50:
                    row_data.append(f"{col_letter}{row}='{val}'")
        if row_data:
            results.append(f"  R{row}: {', '.join(row_data[:8])}")
    
    # Поиск годов
    results.append(f"\nГоды в файле:")
    years_found = {}
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, min(50, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if isinstance(val, int) and val in (2022, 2023, 2024):
                col_letter = get_column_letter(col)
                if val not in years_found:
                    years_found[val] = []
                years_found[val].append(f"{col_letter}{row}")
    for year, positions in years_found.items():
        results.append(f"  {year}: {', '.join(positions)}")
    
    # Поиск месяцев
    results.append(f"\nМесяцы в файле:")
    months_found = []
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if isinstance(val, str):
                month_names = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                              "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
                for month in month_names:
                    if month in val:
                        col_letter = get_column_letter(col)
                        months_found.append(f"{col_letter}{row}={val}")
                        break
    if months_found:
        results.append(f"  Найдено: {len(months_found)} записей")
        results.append(f"  Примеры: {', '.join(months_found[:10])}")
    else:
        results.append("  Месяцы не найдены")
    
    # Поиск объемов газа (числовые значения)
    results.append(f"\nЧисловые значения (объемы газа):")
    numeric_samples = []
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(30, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if isinstance(val, (int, float)) and 100 < val < 1000000:
                col_letter = get_column_letter(col)
                numeric_samples.append(f"{col_letter}{row}={val}")
                if len(numeric_samples) >= 15:
                    break
        if len(numeric_samples) >= 15:
            break
    if numeric_samples:
        results.append(f"  Примеры: {', '.join(numeric_samples[:10])}")

wb.close()

# Выводим и сохраняем
output = '\n'.join(results)
print(output)

# Сохраняем в файл
output_file = Path("data/analysis/gas_file_analysis.txt")
output_file.parent.mkdir(parents=True, exist_ok=True)
with open(output_file, 'w', encoding='utf-8') as f:
    f.write(output)

print(f"\n✅ Результат сохранен в: {output_file}")

