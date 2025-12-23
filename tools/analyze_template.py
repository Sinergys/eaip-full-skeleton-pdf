"""
Детальный анализ Excel шаблона для понимания структуры и поиска полей для заполнения.
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import argparse
import sys


def analyze_template(template_path: Path, max_rows_per_sheet: int = 50) -> dict:
    """Анализ структуры шаблона."""
    workbook = load_workbook(template_path, data_only=False)
    analysis = {
        "template_path": str(template_path),
        "sheets": {},
        "summary": {
            "total_sheets": len(workbook.sheetnames),
            "sheets_with_data": 0,
            "total_cells_with_data": 0,
            "total_formulas": 0,
            "total_merged_ranges": 0
        }
    }
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "cells_with_data": [],
            "formulas": [],
            "empty_areas": []
        }
        
        # Анализ ячеек
        cells_with_data = 0
        formulas_count = 0
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows_per_sheet), start=1):
            row_data = []
            for col_idx, cell in enumerate(row, start=1):
                if isinstance(cell, MergedCell):
                    continue
                
                cell_info = {
                    "address": cell.coordinate,
                    "row": row_idx,
                    "column": col_idx,
                    "value": None,
                    "data_type": cell.data_type,
                    "is_formula": False,
                    "formula": None
                }
                
                if cell.value is not None:
                    cells_with_data += 1
                    cell_info["value"] = str(cell.value)[:200]  # Ограничиваем длину
                    
                    if cell.data_type == 'f':
                        formulas_count += 1
                        cell_info["is_formula"] = True
                        cell_info["formula"] = str(cell.value)
                        sheet_info["formulas"].append({
                            "cell": cell.coordinate,
                            "formula": str(cell.value)[:200]
                        })
                    
                    row_data.append(cell_info)
            
            if row_data:
                sheet_info["cells_with_data"].extend(row_data)
        
        sheet_info["cells_count"] = cells_with_data
        sheet_info["formulas_count"] = formulas_count
        
        analysis["summary"]["total_cells_with_data"] += cells_with_data
        analysis["summary"]["total_formulas"] += formulas_count
        analysis["summary"]["total_merged_ranges"] += len(sheet_info["merged_ranges"])
        
        if cells_with_data > 0:
            analysis["summary"]["sheets_with_data"] += 1
        
        analysis["sheets"][sheet_name] = sheet_info
    
    return analysis


def main():
    parser = argparse.ArgumentParser(description="Детальный анализ Excel шаблона")
    parser.add_argument("--template", help="Путь к шаблону Excel")
    parser.add_argument("--template-name", help="Имя шаблона из templates_config")
    parser.add_argument("--output", help="Путь для сохранения JSON с результатами")
    parser.add_argument("--max-rows", type=int, default=50, help="Максимальное количество строк для анализа на листе")
    parser.add_argument("--format", choices=["json", "summary"], default="summary",
                       help="Формат вывода")
    
    args = parser.parse_args()
    
    # Определение пути к шаблону
    if args.template_name:
        try:
            templates_config_path = Path(__file__).parent.parent / "templates" / "pcm690"
            if str(templates_config_path) not in sys.path:
                sys.path.insert(0, str(templates_config_path))
            from templates_config import get_template_path
            template_path = get_template_path(args.template_name)
        except ImportError:
            raise ImportError(
                "Не удалось импортировать templates_config. "
                "Убедитесь, что файл templates/pcm690/templates_config.py существует."
            )
    elif args.template:
        template_path = Path(args.template)
    else:
        raise ValueError("Необходимо указать либо --template-name, либо --template")
    
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    print(f"Анализ шаблона: {template_path}")
    print("=" * 80)
    
    analysis = analyze_template(template_path, max_rows_per_sheet=args.max_rows)
    
    if args.format == "summary":
        print("\n📊 Сводка:")
        print(f"  Всего листов: {analysis['summary']['total_sheets']}")
        print(f"  Листов с данными: {analysis['summary']['sheets_with_data']}")
        print(f"  Всего ячеек с данными: {analysis['summary']['total_cells_with_data']}")
        print(f"  Всего формул: {analysis['summary']['total_formulas']}")
        print(f"  Всего объединенных диапазонов: {analysis['summary']['total_merged_ranges']}")
        
        print("\n📄 Детали по листам:")
        for sheet_name, sheet_info in analysis["sheets"].items():
            print(f"\n  {sheet_name}:")
            print(f"    Размер: {sheet_info['max_row']} строк × {sheet_info['max_column']} столбцов")
            print(f"    Ячеек с данными: {sheet_info['cells_count']}")
            print(f"    Формул: {sheet_info['formulas_count']}")
            print(f"    Объединенных диапазонов: {len(sheet_info['merged_ranges'])}")
            
            if sheet_info['cells_with_data']:
                print("    Примеры ячеек (первые 5):")
                for cell in sheet_info['cells_with_data'][:5]:
                    value_preview = cell['value'][:50] if cell['value'] else "None"
                    print(f"      {cell['address']}: {value_preview}")
    
    else:
        output_str = json.dumps(analysis, ensure_ascii=False, indent=2)
        print(output_str)
    
    if args.output:
        output_str = json.dumps(analysis, ensure_ascii=False, indent=2)
        Path(args.output).write_text(output_str, encoding="utf-8")
        print(f"\nРезультаты сохранены в: {args.output}")


if __name__ == "__main__":
    main()

