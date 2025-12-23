"""Проверка расположения E32 в шаблоне"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=False)
ws = wb["Структура пр 2"]

print("Проверка строки 32 (где находится E32):")
for col in range(1, 75):
    cell = ws.cell(32, col)
    val = cell.value
    if val is not None:
        col_letter = get_column_letter(col)
        print(f"  {col_letter}32: {val} (тип: {cell.data_type})")

print("\nПроверка строки 9 (где по маппингу должен быть 2023-Q1):")
for col in range(65, 75):
    cell = ws.cell(9, col)
    val = cell.value
    col_letter = get_column_letter(col)
    print(f"  {col_letter}9: {val} (тип: {cell.data_type})")

print("\nПоиск '2023' в заголовках:")
for row in range(1, 10):
    for col in range(1, 200):
        val = ws.cell(row, col).value
        if val == 2023:
            col_letter = get_column_letter(col)
            print(f"  Найден 2023: {col_letter}{row}")

wb.close()

