"""
Модуль для замены placeholder'ов в новом шаблоне new_energy_passport.xlsx.
Заполняет все поля данными из aggregated JSON.
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import re


def replace_text_placeholders(ws, replacements: Dict[str, Any]) -> None:
    """
    Заменить текстовые placeholder'ы в ячейках листа.

    Args:
        ws: Рабочий лист Excel
        replacements: Словарь {placeholder: value} для замены
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            cell_value = str(cell.value)
            original_value = cell_value

            # Замена placeholder'ов в формате {{key}}
            for placeholder, value in replacements.items():
                pattern = r"\{\{" + re.escape(placeholder) + r"\}\}"
                if re.search(pattern, cell_value):
                    cell_value = re.sub(pattern, str(value), cell_value)

            # Замена подчеркиваний в контексте года/квартала
            if (
                "____ год" in cell_value
                or "20___" in cell_value
                or "20__" in cell_value
            ):
                if "year" in replacements:
                    cell_value = re.sub(r"20_+", str(replacements["year"]), cell_value)
                    cell_value = re.sub(
                        r"_+ год", f" {replacements['year']} год", cell_value
                    )

            if "____ квартал" in cell_value or "___ квартал" in cell_value:
                if "quarter" in replacements:
                    cell_value = re.sub(
                        r"_+ квартал", f" {replacements['quarter']} квартал", cell_value
                    )

            if cell_value != original_value:
                try:
                    cell.value = cell_value
                except (AttributeError, ValueError):
                    pass  # Пропускаем проблемные ячейки


def fill_summary_sheet(
    ws, agg_data: Dict, enterprise_name: Optional[str] = None
) -> None:
    """Заполнить лист Summary (Sheet1) с метаданными."""
    replacements = {
        "enterprise.name": enterprise_name or "ООО Синергис",
        "year": None,
        "quarter": None,
    }

    # Определяем период из данных
    resources = agg_data.get("resources", {})
    if resources:
        # Берем первый доступный квартал для определения года
        for resource_type, resource_data in resources.items():
            if resource_data:
                first_quarter = (
                    sorted(resource_data.keys())[0] if resource_data else None
                )
                if first_quarter:
                    year = first_quarter.split("-")[0] if "-" in first_quarter else None
                    if year:
                        replacements["year"] = year
                        break

    replace_text_placeholders(ws, replacements)


def fill_structure_sheet(ws, agg_data: Dict, quarter: Optional[str] = None) -> None:
    """Заполнить лист 'Структура пр 2' с данными по кварталам."""
    replacements = {}

    # Определяем квартал
    if quarter:
        replacements["quarter"] = quarter.split("-Q")[1] if "-Q" in quarter else quarter
        replacements["year"] = quarter.split("-")[0] if "-" in quarter else None
    else:
        # Берем первый доступный квартал
        resources = agg_data.get("resources", {})
        for resource_type, resource_data in resources.items():
            if resource_data:
                first_quarter = (
                    sorted(resource_data.keys())[0] if resource_data else None
                )
                if first_quarter:
                    replacements["quarter"] = (
                        first_quarter.split("-Q")[1]
                        if "-Q" in first_quarter
                        else first_quarter
                    )
                    replacements["year"] = (
                        first_quarter.split("-")[0] if "-" in first_quarter else None
                    )
                    break

    replace_text_placeholders(ws, replacements)

    # Заполнение таблицы данными
    fill_structure_table(ws, agg_data)


def fill_structure_table(ws, agg_data: Dict) -> None:
    """
    Заполнение таблицы структуры параметров энергетического баланса.

    Структура таблицы:
    - Row 3: B3 - "____ год ____ квартал" (заменяется placeholder'ом)
    - Row 4: Заголовки - C4: "Электроэнергия", E4: "Тепловая энергия", F4: "Газ", N4: "Вода"
    - Row 5: Подзаголовки - C5: "актив Р", D5: "реактив Q"
    - Row 6: Единицы измерения
    - Row 7+: Данные по показателям

    Маппинг строк:
    - Row 9: "Общее потребление по предприятию"
    - Row 10: "для технологических нужд"
    - Row 11: "для собственных нужд"
    - Row 12: "для производственных нужд"
    - Row 13: "для хозяйственно-бытовых нужд"
    """
    resources = agg_data.get("resources", {})
    electricity = resources.get("electricity", {})
    gas = resources.get("gas", {})
    water = resources.get("water", {})

    # Определяем все доступные кварталы
    all_quarters = set()
    for resource_data in [electricity, gas, water]:
        all_quarters.update(resource_data.keys())
    all_quarters = sorted(all_quarters)

    if not all_quarters:
        return  # Нет данных для заполнения

    # Маппинг столбцов для данных
    # C - активная электроэнергия (кВт·ч)
    # D - реактивная электроэнергия (кВАр·ч)
    # F - газ (тыс. м³) - нужно разделить на 1000
    # N - вода (м³)

    # Заполняем данные для каждого квартала
    # В новом шаблоне структура может отличаться, поэтому заполняем основные показатели

    # Row 9: "Общее потребление по предприятию"
    # Для первого квартала заполняем в столбцах C, D, F, N
    if all_quarters:
        first_quarter = all_quarters[0]

        # Электроэнергия активная (C9)
        elec_totals = electricity.get(first_quarter, {}).get("quarter_totals", {})
        active_kwh = elec_totals.get("active_kwh", 0)
        if active_kwh:
            ws.cell(row=9, column=3).value = active_kwh  # C9

        # Электроэнергия реактивная (D9)
        reactive_kvarh = elec_totals.get("reactive_kvarh", 0)
        if reactive_kvarh:
            ws.cell(row=9, column=4).value = reactive_kvarh  # D9

        # Газ (F9) - в тыс. м³
        gas_totals = gas.get(first_quarter, {}).get("quarter_totals", {})
        gas_volume_m3 = gas_totals.get("volume_m3", 0)
        if gas_volume_m3:
            ws.cell(row=9, column=6).value = gas_volume_m3 / 1000.0  # F9 в тыс. м³

        # Вода (N9)
        water_totals = water.get(first_quarter, {}).get("quarter_totals", {})
        water_volume_m3 = water_totals.get("volume_m3", 0)
        if water_volume_m3:
            ws.cell(row=9, column=14).value = water_volume_m3  # N9

    # Заполнение по категориям потребления (если есть данные by_usage)
    # Row 10: "для технологических нужд"
    # Row 11: "для собственных нужд"
    # Row 12: "для производственных нужд"
    # Row 13: "для хозяйственно-бытовых нужд"

    if all_quarters and first_quarter in electricity:
        by_usage = electricity[first_quarter].get("by_usage", {})
        if by_usage:
            # Технологические (C10)
            tech = by_usage.get("technological", 0)
            if tech:
                ws.cell(row=10, column=3).value = tech

            # Собственные нужды (C11)
            own = by_usage.get("own_needs", 0)
            if own:
                ws.cell(row=11, column=3).value = own

            # Производственные (C12)
            prod = by_usage.get("production", 0)
            if prod:
                ws.cell(row=12, column=3).value = prod

            # Хозяйственно-бытовые (C13)
            household = by_usage.get("household", 0)
            if household:
                ws.cell(row=13, column=3).value = household


def fill_balance_sheet(ws, agg_data: Dict) -> None:
    """Заполнить лист 'Баланс' с энергетическим балансом."""
    # Заполнение таблицы баланса
    fill_balance_table(ws, agg_data)


def fill_balance_table(ws, agg_data: Dict) -> None:
    """
    Заполнение таблицы энергетического баланса.

    Структура таблицы:
    - Row 1: A1 - "Энергетический баланс предприятия"
    - Row 4: Заголовки - C4: "Электр энергияси", E4: "Тепловая энергия", F4: "Газ", O4: "Вода"
    - Row 5: Подзаголовки - C5: "актив Р", D5: "реактив Q"
    - Row 7: Единицы измерения
    - Row 10: "Общее потребление по предприятию"
    - Row 11: "– на технологические нужды"
    - Row 12: "– на собственные нужды предприятия"
    - Row 13: "– на производственные нужды"
    - Row 14: "– на хозяйственно-бытовые нужды"

    Маппинг столбцов:
    - C - активная электроэнергия (кВт·ч)
    - D - реактивная электроэнергия (кВАр·ч)
    - E - тепловая энергия (Гкал)
    - F - газ (тыс. м³)
    - O - вода (м³)
    """
    resources = agg_data.get("resources", {})
    electricity = resources.get("electricity", {})
    gas = resources.get("gas", {})
    water = resources.get("water", {})

    # Определяем все доступные кварталы
    all_quarters = set()
    for resource_data in [electricity, gas, water]:
        all_quarters.update(resource_data.keys())
    all_quarters = sorted(all_quarters)

    if not all_quarters:
        return  # Нет данных для заполнения

    # Используем первый квартал для заполнения (или можно сделать для всех кварталов)
    first_quarter = all_quarters[0]

    # Row 10: "Общее потребление по предприятию"
    # Электроэнергия активная (C10)
    elec_totals = electricity.get(first_quarter, {}).get("quarter_totals", {})
    active_kwh = elec_totals.get("active_kwh", 0)
    if active_kwh:
        try:
            cell = ws.cell(row=10, column=3)
            if not isinstance(cell, MergedCell):
                cell.value = active_kwh  # C10
        except (AttributeError, ValueError):
            pass

    # Электроэнергия реактивная (D10)
    reactive_kvarh = elec_totals.get("reactive_kvarh", 0)
    if reactive_kvarh:
        try:
            cell = ws.cell(row=10, column=4)
            if not isinstance(cell, MergedCell):
                cell.value = reactive_kvarh  # D10
        except (AttributeError, ValueError):
            pass

    # Газ (F10) - в тыс. м³
    gas_totals = gas.get(first_quarter, {}).get("quarter_totals", {})
    gas_volume_m3 = gas_totals.get("volume_m3", 0)
    if gas_volume_m3:
        try:
            cell = ws.cell(row=10, column=6)
            if not isinstance(cell, MergedCell):
                cell.value = gas_volume_m3 / 1000.0  # F10 в тыс. м³
        except (AttributeError, ValueError):
            pass

    # Вода (O10)
    water_totals = water.get(first_quarter, {}).get("quarter_totals", {})
    water_volume_m3 = water_totals.get("volume_m3", 0)
    if water_volume_m3:
        try:
            cell = ws.cell(row=10, column=15)
            if not isinstance(cell, MergedCell):
                cell.value = water_volume_m3  # O10
        except (AttributeError, ValueError):
            pass

    # Заполнение по категориям потребления (если есть данные by_usage)
    # Row 11: "– на технологические нужды"
    # Row 12: "– на собственные нужды предприятия"
    # Row 13: "– на производственные нужды"
    # Row 14: "– на хозяйственно-бытовые нужды"

    if first_quarter in electricity:
        by_usage = electricity[first_quarter].get("by_usage", {})
        if by_usage:
            # Технологические (C11)
            tech = by_usage.get("technological", 0)
            if tech:
                try:
                    cell = ws.cell(row=11, column=3)
                    if not isinstance(cell, MergedCell):
                        cell.value = tech
                except (AttributeError, ValueError):
                    pass

            # Собственные нужды (C12)
            own = by_usage.get("own_needs", 0)
            if own:
                try:
                    cell = ws.cell(row=12, column=3)
                    if not isinstance(cell, MergedCell):
                        cell.value = own
                except (AttributeError, ValueError):
                    pass

            # Производственные (C13)
            prod = by_usage.get("production", 0)
            if prod:
                try:
                    cell = ws.cell(row=13, column=3)
                    if not isinstance(cell, MergedCell):
                        cell.value = prod
                except (AttributeError, ValueError):
                    pass

            # Хозяйственно-бытовые (C14)
            household = by_usage.get("household", 0)
            if household:
                try:
                    cell = ws.cell(row=14, column=3)
                    if not isinstance(cell, MergedCell):
                        cell.value = household
                except (AttributeError, ValueError):
                    pass


def fill_dynamics_sheet(ws, agg_data: Dict) -> None:
    """Заполнить лист 'Динамика ср' с динамикой потребления."""
    replacements = {}

    # Определяем год из данных
    resources = agg_data.get("resources", {})
    if resources:
        for resource_type, resource_data in resources.items():
            if resource_data:
                first_quarter = (
                    sorted(resource_data.keys())[0] if resource_data else None
                )
                if first_quarter:
                    year = first_quarter.split("-")[0] if "-" in first_quarter else None
                    if year:
                        replacements["year"] = year
                        break

    replace_text_placeholders(ws, replacements)

    # Заполнение таблицы динамики
    fill_dynamics_table(ws, agg_data)


def fill_dynamics_table(ws, agg_data: Dict) -> None:
    """
    Заполнение таблицы динамики потребления.

    Структура таблицы:
    - Row 1: A1 - "20___ г.Сравнителные показатели динамика расходов" (заменяется placeholder'ом)
    - Row 3: Заголовки ресурсов - C3: "Электроэнергия, кВт·ч", H3: "Тепловая энергия, Гкал", M3: "Природный газ, тыс. м³"
    - Row 4: Кварталы - C4: "I", D4: "II", E4: "III", F4: "IV", G4: "йил" (год)
    - Row 15: "Общее потребление энергоресурсов"

    Маппинг столбцов для электроэнергии:
    - C - I квартал
    - D - II квартал
    - E - III квартал
    - F - IV квартал
    - G - год (итого)

    Маппинг столбцов для тепловой энергии:
    - H - I квартал
    - I - II квартал
    - J - III квартал
    - K - IV квартал
    - L - год (итого)

    Маппинг столбцов для газа:
    - M - I квартал
    - N - II квартал
    - O - III квартал
    - P - IV квартал
    - Q - год (итого)
    """
    resources = agg_data.get("resources", {})
    electricity = resources.get("electricity", {})
    gas = resources.get("gas", {})
    # water = resources.get("water", {})  # Вода не отображается в этой таблице

    # Определяем все доступные кварталы
    all_quarters = set()
    for resource_data in [electricity, gas]:
        all_quarters.update(resource_data.keys())
    all_quarters = sorted(all_quarters)

    if not all_quarters:
        return  # Нет данных для заполнения

    # Группируем кварталы по годам
    quarters_by_year = {}
    for quarter in all_quarters:
        if "-" in quarter:
            year = quarter.split("-")[0]
            quarter_num = quarter.split("-Q")[1] if "-Q" in quarter else None
            if year not in quarters_by_year:
                quarters_by_year[year] = {}
            if quarter_num:
                quarters_by_year[year][quarter_num] = quarter

    # Заполняем данные для каждого года
    # Используем первый доступный год
    if quarters_by_year:
        first_year = sorted(quarters_by_year.keys())[0]
        year_quarters = quarters_by_year[first_year]

        # Row 15: "Общее потребление энергоресурсов"
        # Заполняем квартальные данные для электроэнергии (столбцы C, D, E, F)
        quarter_col_map = {"1": 3, "2": 4, "3": 5, "4": 6}  # C, D, E, F

        total_electricity = 0
        for quarter_num, col in quarter_col_map.items():
            if quarter_num in year_quarters:
                quarter = year_quarters[quarter_num]
                elec_totals = electricity.get(quarter, {}).get("quarter_totals", {})
                active_kwh = elec_totals.get("active_kwh", 0)
                if active_kwh:
                    try:
                        cell = ws.cell(row=15, column=col)
                        if not isinstance(cell, MergedCell):
                            # Конвертируем в тыс. кВт·ч
                            cell.value = active_kwh / 1000.0
                            total_electricity += active_kwh
                    except (AttributeError, ValueError):
                        pass

        # Год (итого) для электроэнергии (G15)
        if total_electricity > 0:
            try:
                cell = ws.cell(row=15, column=7)  # G
                if not isinstance(cell, MergedCell):
                    cell.value = total_electricity / 1000.0  # в тыс. кВт·ч
            except (AttributeError, ValueError):
                pass

        # Заполняем квартальные данные для газа (столбцы M, N, O, P)
        gas_quarter_col_map = {"1": 13, "2": 14, "3": 15, "4": 16}  # M, N, O, P

        total_gas = 0
        for quarter_num, col in gas_quarter_col_map.items():
            if quarter_num in year_quarters:
                quarter = year_quarters[quarter_num]
                gas_totals = gas.get(quarter, {}).get("quarter_totals", {})
                gas_volume_m3 = gas_totals.get("volume_m3", 0)
                if gas_volume_m3:
                    try:
                        cell = ws.cell(row=15, column=col)
                        if not isinstance(cell, MergedCell):
                            # Конвертируем в тыс. м³
                            cell.value = gas_volume_m3 / 1000.0
                            total_gas += gas_volume_m3
                    except (AttributeError, ValueError):
                        pass

        # Год (итого) для газа (Q15)
        if total_gas > 0:
            try:
                cell = ws.cell(row=15, column=17)  # Q
                if not isinstance(cell, MergedCell):
                    cell.value = total_gas / 1000.0  # в тыс. м³
            except (AttributeError, ValueError):
                pass


def fill_measures_sheet(ws, measures_data: Optional[list] = None) -> None:
    """Заполнить лист 'Мериаприятия 1' с мероприятиями."""
    # Эта функция уже реализована в fill_energy_passport.py как fill_meropriyatiya_sheet
    # Здесь можно добавить специфичные для нового шаблона замены
    pass


def fill_fuel_sheet(ws, agg_data: Dict) -> None:
    """Заполнить лист 'мазут,уголь 5' с данными по топливу."""
    replacements = {}

    # Определяем год из данных
    resources = agg_data.get("resources", {})
    if resources:
        for resource_type, resource_data in resources.items():
            if resource_data:
                first_quarter = (
                    sorted(resource_data.keys())[0] if resource_data else None
                )
                if first_quarter:
                    year = first_quarter.split("-")[0] if "-" in first_quarter else None
                    if year:
                        replacements["year"] = year
                        break

    replace_text_placeholders(ws, replacements)

    # TODO: Заполнение таблицы с данными по топливу
    # Если в aggregated данных есть информация о топливе


def fill_specific_consumption_sheet(ws, agg_data: Dict) -> None:
    """Заполнить лист 'Расход на ед.п' с удельным потреблением."""
    replacements = {}

    # Определяем годы из данных (может быть несколько лет)
    resources = agg_data.get("resources", {})
    years = set()
    if resources:
        for resource_type, resource_data in resources.items():
            for quarter in resource_data.keys():
                if "-" in quarter:
                    year = quarter.split("-")[0]
                    if year.isdigit():
                        years.add(year)

    # Заменяем все вхождения года
    if years:
        # Берем первый год для замены всех placeholder'ов
        first_year = sorted(years)[0] if years else None
        if first_year:
            replacements["year"] = first_year

    replace_text_placeholders(ws, replacements)

    # TODO: Заполнение таблицы с удельным потреблением
    # Расчет: потребление электроэнергии / производство


def fill_new_template(
    template_path: Path,
    agg_data: Dict,
    output_path: Path,
    enterprise_name: Optional[str] = None,
    measures_data: Optional[list] = None,
    nodes_data: Optional[list] = None,
) -> None:
    """
    Заполнить новый шаблон new_energy_passport.xlsx данными.

    Args:
        template_path: Путь к шаблону
        agg_data: Агрегированные данные в формате {"resources": {...}}
        output_path: Путь для сохранения заполненного шаблона
        enterprise_name: Название предприятия
        measures_data: Данные о мероприятиях
        nodes_data: Данные по узлам учета
    """
    import sys
    from pathlib import Path

    tools_path = Path(__file__).parent.parent
    if str(tools_path) not in sys.path:
        sys.path.insert(0, str(tools_path))

    from fill_energy_passport import fill_nodes_sheet

    workbook = load_workbook(template_path, data_only=False)

    # Определяем первый доступный квартал для замены placeholder'ов
    resources = agg_data.get("resources", {})
    first_quarter = None
    for resource_type, resource_data in resources.items():
        if resource_data:
            first_quarter = sorted(resource_data.keys())[0] if resource_data else None
            if first_quarter:
                break

    # Заполнение каждого листа
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        if sheet_name == "Sheet1":
            fill_summary_sheet(ws, agg_data, enterprise_name)
        elif "Узел учета" in sheet_name or "узел" in sheet_name.lower():
            if nodes_data:
                fill_nodes_sheet(ws, nodes_data)
        elif "Структура" in sheet_name or "структура" in sheet_name.lower():
            fill_structure_sheet(ws, agg_data, first_quarter)
            # TODO: Добавить заполнение таблицы с данными для нового шаблона
        elif "Баланс" in sheet_name or "баланс" in sheet_name.lower():
            fill_balance_sheet(ws, agg_data)
            # TODO: Добавить заполнение таблицы баланса для нового шаблона
        elif "Динамика" in sheet_name or "динамика" in sheet_name.lower():
            fill_dynamics_sheet(ws, agg_data)
            # TODO: Добавить заполнение таблицы динамики для нового шаблона
        elif "мазут" in sheet_name.lower() or "уголь" in sheet_name.lower():
            fill_fuel_sheet(ws, agg_data)
        elif "Расход" in sheet_name or "расход" in sheet_name.lower():
            fill_specific_consumption_sheet(ws, agg_data)
        elif "Мериаприятия" in sheet_name or "мероприятия" in sheet_name.lower():
            fill_measures_sheet(ws, measures_data)
            # TODO: Добавить заполнение таблицы мероприятий для нового шаблона

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"✅ Заполненный шаблон сохранен: {output_path}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Заполнить новый шаблон энергопаспорта"
    )
    parser.add_argument("--template", help="Путь к шаблону")
    parser.add_argument("--template-name", help="Имя шаблона из templates_config")
    parser.add_argument("--aggregated", required=True, help="Путь к aggregated JSON")
    parser.add_argument("--output", required=True, help="Путь для сохранения")
    parser.add_argument("--enterprise-name", help="Название предприятия")
    parser.add_argument("--measures-json", help="JSON с мероприятиями")

    args = parser.parse_args()

    # Определение пути к шаблону
    if args.template_name:
        import sys

        templates_config_path = Path(__file__).parent.parent / "templates" / "pcm690"
        if str(templates_config_path) not in sys.path:
            sys.path.insert(0, str(templates_config_path))
        from templates_config import get_template_path

        template_path = get_template_path(args.template_name)
    elif args.template:
        template_path = Path(args.template)
    else:
        raise ValueError("Необходимо указать либо --template-name, либо --template")

    # Загрузка данных
    agg_data = json.loads(Path(args.aggregated).read_text(encoding="utf-8"))

    # Нормализация структуры данных (как в fill_energy_passport.py)
    if not isinstance(agg_data, dict) or "resources" not in agg_data:
        normalized_data = {"resources": {}}
        for file_data in agg_data.values() if isinstance(agg_data, dict) else []:
            if "resources" in file_data:
                for resource_type, resource_data in file_data["resources"].items():
                    if resource_type not in normalized_data["resources"]:
                        normalized_data["resources"][resource_type] = {}
                    for quarter, quarter_data in resource_data.items():
                        normalized_data["resources"][resource_type][quarter] = (
                            quarter_data
                        )
        agg_data = normalized_data

    # Загрузка мероприятий
    measures_data = None
    if args.measures_json:
        measures_data = json.loads(Path(args.measures_json).read_text(encoding="utf-8"))
        if isinstance(measures_data, dict) and "measures" in measures_data:
            measures_data = measures_data["measures"]

    fill_new_template(
        template_path=template_path,
        agg_data=agg_data,
        output_path=Path(args.output),
        enterprise_name=args.enterprise_name,
        measures_data=measures_data,
    )
