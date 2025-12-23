"""Показывает структуру эталонного Excel файла в читаемом виде"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def show_excel_structure(file_path: Path, output_file: Path = None):
    """Показывает полную структуру Excel файла"""
    
    if not file_path.exists():
        print(f"❌ Файл не найден: {file_path}")
        return
    
    wb = load_workbook(file_path, data_only=False)
    
    output_lines = []
    output_lines.append("=" * 100)
    output_lines.append(f"СТРУКТУРА ФАЙЛА: {file_path.name}")
    output_lines.append("=" * 100)
    output_lines.append("")
    
    # Список всех листов
    output_lines.append("📋 ВСЕ ЛИСТЫ В ФАЙЛЕ:")
    output_lines.append("-" * 100)
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        output_lines.append(f"  {idx}. {sheet_name} ({ws.max_row} строк × {ws.max_column} колонок)")
    output_lines.append("")
    
    # Детальный анализ листа "Баланс"
    balance_sheet = None
    for name in ["Баланс", "04_Баланс", "Balans", "04_Balans"]:
        if name in wb.sheetnames:
            balance_sheet = wb[name]
            break
    
    if balance_sheet:
        output_lines.append("=" * 100)
        output_lines.append(f"ДЕТАЛЬНАЯ СТРУКТУРА ЛИСТА: '{balance_sheet.title}'")
        output_lines.append("=" * 100)
        output_lines.append("")
        
        # Показываем первые 20 строк полностью
        output_lines.append("📊 СТРУКТУРА ЛИСТА (первые 20 строк):")
        output_lines.append("-" * 100)
        
        for row_idx in range(1, min(21, balance_sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(17, balance_sheet.max_column + 1)):
                cell = balance_sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    if cell.data_type == "f":
                        val = f"={str(cell.value)[:30]}"
                    else:
                        val = str(cell.value)
                        if len(val) > 30:
                            val = val[:27] + "..."
                    row_data.append(f"{get_column_letter(col_idx)}:{val}")
            
            if row_data:
                output_lines.append(f"Строка {row_idx:2d}: {' | '.join(row_data)}")
        
        output_lines.append("")
        
        # Маппинг категорий и ресурсов
        output_lines.append("=" * 100)
        output_lines.append("МАППИНГ КАТЕГОРИЙ ПОТРЕБЛЕНИЯ И РЕСУРСОВ")
        output_lines.append("=" * 100)
        output_lines.append("")
        
        # Категории потребления
        output_lines.append("📌 КАТЕГОРИИ ПОТРЕБЛЕНИЯ (строки):")
        output_lines.append("-" * 100)
        categories = {}
        for row_idx in range(8, 16):
            cell_a = balance_sheet.cell(row=row_idx, column=1)
            if cell_a.value:
                label = str(cell_a.value).strip()
                categories[row_idx] = label
                output_lines.append(f"  Строка {row_idx:2d}: {label}")
        output_lines.append("")
        
        # Ресурсы (колонки)
        output_lines.append("📌 РЕСУРСЫ (колонки):")
        output_lines.append("-" * 100)
        resource_columns = {}
        
        # Анализируем строки 4-7 для определения колонок ресурсов
        for row_idx in range(4, 8):
            for col_idx in range(1, min(17, balance_sheet.max_column + 1)):
                cell = balance_sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    val = str(cell.value).strip()
                    col_letter = get_column_letter(col_idx)
                    
                    # Определяем тип ресурса
                    val_lower = val.lower()
                    resource_type = None
                    
                    if "электр" in val_lower or "энергия" in val_lower:
                        if "актив" in val_lower or ("р" in val_lower and "реактив" not in val_lower):
                            resource_type = "Электроэнергия активная (Р)"
                        elif "реактив" in val_lower or "q" in val_lower:
                            resource_type = "Электроэнергия реактивная (Q)"
                        else:
                            resource_type = "Электроэнергия"
                    elif "тепл" in val_lower or "гкал" in val_lower:
                        resource_type = "Тепловая энергия (Гкал)"
                    elif "газ" in val_lower:
                        resource_type = "Газ (тыс. м³)"
                    elif "мазут" in val_lower:
                        resource_type = "Мазут (тонна)"
                    elif "уголь" in val_lower:
                        resource_type = "Уголь (тонна)"
                    elif "бензин" in val_lower:
                        resource_type = "Бензин (тонна)"
                    elif "дизель" in val_lower:
                        resource_type = "Дизель (тонна)"
                    elif "керосин" in val_lower:
                        resource_type = "Керосин (тонна)"
                    elif "гсм" in val_lower or "горюче" in val_lower:
                        resource_type = "ГСМ (тонна)"
                    elif "вода" in val_lower or "м³" in val_lower:
                        resource_type = "Вода (м³)"
                    
                    if resource_type:
                        if col_letter not in resource_columns:
                            resource_columns[col_letter] = {
                                "type": resource_type,
                                "rows": [],
                                "unit": None
                            }
                        resource_columns[col_letter]["rows"].append(row_idx)
                        # Определяем единицу измерения из строки 7
                        if row_idx == 7:
                            resource_columns[col_letter]["unit"] = val
        
        for col_letter in sorted(resource_columns.keys()):
            info = resource_columns[col_letter]
            unit_str = f" ({info['unit']})" if info['unit'] else ""
            output_lines.append(f"  Колонка {col_letter:3s}: {info['type']}{unit_str}")
            output_lines.append(f"              Заголовок в строках: {', '.join(map(str, info['rows']))}")
        output_lines.append("")
        
        # Пример заполнения данных
        output_lines.append("=" * 100)
        output_lines.append("ПРИМЕР ЗАПОЛНЕНИЯ ДАННЫХ (строки 10-14)")
        output_lines.append("=" * 100)
        output_lines.append("")
        
        # Создаем таблицу
        header_cols = ["Категория"] + sorted(resource_columns.keys())[:8]
        output_lines.append(f"{'Категория':<45} | " + " | ".join(f"{col:>12}" for col in header_cols[1:]))
        output_lines.append("-" * 100)
        
        for row_idx in [10, 11, 12, 13, 14]:
            cell_a = balance_sheet.cell(row=row_idx, column=1)
            if cell_a.value:
                label = str(cell_a.value).strip()
                if len(label) > 43:
                    label = label[:40] + "..."
                
                row_values = [label]
                for col_letter in sorted(resource_columns.keys())[:8]:
                    col_idx = ord(col_letter) - ord('A') + 1
                    cell = balance_sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        if cell.data_type == "f":
                            val = f"={str(cell.value)[:10]}"
                        else:
                            val = str(cell.value)
                            if len(val) > 10:
                                val = val[:7] + "..."
                        row_values.append(val)
                    else:
                        row_values.append("")
                
                output_lines.append(" | ".join(f"{val:<45}" if i == 0 else f"{val:>12}" for i, val in enumerate(row_values)))
        
        output_lines.append("")
    
    # Показываем структуру других листов (кратко)
    output_lines.append("=" * 100)
    output_lines.append("СТРУКТУРА ДРУГИХ ЛИСТОВ (первые 5 строк)")
    output_lines.append("=" * 100)
    output_lines.append("")
    
    for sheet_name in wb.sheetnames:
        if sheet_name == balance_sheet.title if balance_sheet else False:
            continue
        
        ws = wb[sheet_name]
        output_lines.append(f"📄 Лист: {sheet_name}")
        output_lines.append("-" * 100)
        
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    val = str(cell.value)
                    if len(val) > 20:
                        val = val[:17] + "..."
                    row_data.append(f"{get_column_letter(col_idx)}:{val}")
            
            if row_data:
                output_lines.append(f"  Строка {row_idx:2d}: {' | '.join(row_data)}")
        
        output_lines.append("")
    
    wb.close()
    
    # Выводим результат
    result_text = "\n".join(output_lines)
    print(result_text)
    
    # Сохраняем в файл
    if output_file:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(result_text)
        print(f"\n✅ Структура сохранена в файл: {output_file}")
    else:
        # Сохраняем по умолчанию
        default_output = Path("docs") / "reference_excel_structure.txt"
        default_output.parent.mkdir(parents=True, exist_ok=True)
        with open(default_output, "w", encoding="utf-8") as f:
            f.write(result_text)
        print(f"\n✅ Структура сохранена в файл: {default_output}")

if __name__ == "__main__":
    file_path = Path("docs/input_templates/энергопаспорт (3) (10) (2).xlsx")
    output_file = Path("docs/reference_excel_structure.txt")
    show_excel_structure(file_path, output_file)

