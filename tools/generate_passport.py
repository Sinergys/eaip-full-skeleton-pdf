"""
Вспомогательная функция для генерации энергопаспорта с поддержкой выбора шаблона
"""
from pathlib import Path
from typing import Optional, Dict, Any
import logging

logger = logging.getLogger(__name__)


def generate_passport(
    aggregated_data: Dict[str, Any],
    enterprise_data: Dict[str, Any],
    output_path: Path,
    template_name: Optional[str] = None,
    nodes_json_path: Optional[Path] = None,
    equipment_json_path: Optional[Path] = None,
    envelope_json_path: Optional[Path] = None,
    loss_active_month: float = 0.0,
    loss_reactive_month: float = 0.0,
    transformer_power_kva: float = 0.0,
) -> Path:
    """
    Генерация энергопаспорта с поддержкой выбора шаблона
    
    Args:
        aggregated_data: Агрегированные данные энергопотребления
        enterprise_data: Данные предприятия
        output_path: Путь для сохранения сгенерированного паспорта
        template_name: Имя шаблона из templates_config (например, "new_energy_passport", "metin", "default")
        nodes_json_path: Путь к JSON с данными узлов учета (опционально)
        equipment_json_path: Путь к JSON с данными оборудования (опционально)
        envelope_json_path: Путь к JSON с данными ограждающих конструкций (опционально)
        loss_active_month: Потери активной энергии за месяц (кВт·ч)
        loss_reactive_month: Потери реактивной энергии за месяц (кВАр·ч)
        transformer_power_kva: Мощность трансформатора (кВА)
    
    Returns:
        Path к сгенерированному файлу
    
    Raises:
        FileNotFoundError: Если шаблон не найден
        ValueError: Если имя шаблона не существует
    """
    from openpyxl import load_workbook
    import json
    
    # Определение пути к шаблону
    template_path = None
    if template_name:
        try:
            # Добавляем путь к templates_config
            templates_config_dir = Path(__file__).parent.parent / "templates" / "pcm690"
            import sys
            if str(templates_config_dir) not in sys.path:
                sys.path.insert(0, str(templates_config_dir))
            from templates_config import get_template_path
            template_path = get_template_path(template_name)
            logger.info("Используется шаблон по имени '%s': %s", template_name, template_path)
        except (ImportError, ValueError, FileNotFoundError) as e:
            logger.error("Не удалось загрузить шаблон по имени '%s': %s", template_name, e)
            raise
    
    # Если шаблон не выбран по имени, используем дефолтный
    if not template_path:
        default_template = Path(__file__).parent.parent / "templates" / "pcm690" / "energy_passport_template.xlsx"
        if default_template.exists():
            template_path = default_template
            logger.info("Используется дефолтный шаблон: %s", template_path)
        else:
            raise FileNotFoundError(
                f"Шаблон не найден. Укажите template_name или убедитесь, что дефолтный шаблон существует: {default_template}"
            )
    
    # Загрузка шаблона
    workbook = load_workbook(template_path, data_only=False)
    
    # Импорт функций заполнения
    # Используем относительный импорт из того же модуля
    import sys
    fill_energy_passport_path = Path(__file__).parent
    if str(fill_energy_passport_path) not in sys.path:
        sys.path.insert(0, str(fill_energy_passport_path))
    
    from fill_energy_passport import (
        fill_struktura_pr2,
        fill_nodes_sheet,
        load_default_nodes,
        fill_building_envelope_sheet,
        fill_equipment_sheet,
        fill_losses_sheet,
    )
    
    # Подготовка данных ресурсов
    resources_data = aggregated_data.get("resources") or aggregated_data
    if resources_data is None:
        resources_data = {}
    for key in ("electricity", "gas", "water", "production"):
        resources_data.setdefault(key, {})
    
    # Заполнение листа Struktura pr2
    if "Struktura pr2" in workbook.sheetnames:
        fill_struktura_pr2(
            workbook["Struktura pr2"],
            resources_data,
            loss_active_month=loss_active_month,
            loss_reactive_month=loss_reactive_month,
        )
    
    # Заполнение узлов учета
    if nodes_json_path and nodes_json_path.exists():
        from fill_energy_passport import load_nodes_from_json
        nodes_data = load_nodes_from_json(nodes_json_path)
    else:
        nodes_data = load_default_nodes()
    
    if "01_Узлы учета" in workbook.sheetnames:
        fill_nodes_sheet(workbook["01_Узлы учета"], nodes_data)
    
    # Заполнение оборудования
    if equipment_json_path and equipment_json_path.exists():
        equipment_data = json.loads(equipment_json_path.read_text(encoding="utf-8"))
        fill_equipment_sheet(workbook, equipment_data, sheet_name="Equipment")
    
    # Заполнение ограждающих конструкций
    if envelope_json_path and envelope_json_path.exists():
        envelope_data = json.loads(envelope_json_path.read_text(encoding="utf-8"))
        fill_building_envelope_sheet(
            workbook,
            envelope_data,
            sheet_name="02_Исходные данные",
        )
    
    # Заполнение потерь
    if loss_active_month or loss_reactive_month:
        fill_losses_sheet(
            workbook,
            loss_active_month,
            loss_reactive_month,
            transformer_power_kva=transformer_power_kva,
            hours_per_month=720.0,
        )
    
    # Сохранение
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    logger.info("✅ Энергопаспорт сгенерирован: %s", output_path)
    
    return output_path

