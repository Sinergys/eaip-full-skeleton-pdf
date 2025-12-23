"""
Детальный анализ шаблона энергопаспорта для определения необходимых данных.
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict

template_path = Path(r"C:\eaip\data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx")

print("=" * 80)
print("АНАЛИЗ ШАБЛОНА ЭНЕРГОПАСПОРТА")
print("=" * 80)
print(f"Файл: {template_path.name}\n")

if not template_path.exists():
    print(f"❌ Файл не найден: {template_path}")
    exit(1)

wb = load_workbook(template_path, data_only=False)
print(f"✅ Файл загружен")
print(f"📋 Листы: {wb.sheetnames}\n")

# Анализ каждого листа
required_data = {
    "enterprise_info": [],
    "gas_data": [],
    "electricity_data": [],
    "production_data": [],
    "building_data": [],
    "equipment_data": [],
    "nodes_data": [],
    "other": [],
}

# 1. Анализ листа "Структура пр 2"
if "Структура пр 2" in wb.sheetnames:
    ws = wb["Структура пр 2"]
    print("=" * 80)
    print("ЛИСТ 'Структура пр 2'")
    print("=" * 80)
    
    # Ищем строки с данными
    print("\n📊 Строки с данными по ресурсам:")
    for row in range(1, 100):
        val_a = ws.cell(row, 1).value
        if val_a and isinstance(val_a, str):
            val_lower = val_a.lower()
            if any(keyword in val_lower for keyword in [
                "потребление", "электро", "газ", "вода", "тепло",
                "труб", "фитинг", "канал", "собственн", "хозяйствен"
            ]):
                # Показываем несколько колонок
                row_data = []
                for col in range(1, 20):
                    cell_val = ws.cell(row, col).value
                    if cell_val is not None:
                        row_data.append(f"{get_column_letter(col)}{row}={cell_val}")
                if row_data:
                    print(f"  Строка {row}: {val_a}")
                    print(f"    Данные: {', '.join(row_data[:5])}")
    
    # Анализ структуры кварталов
    print("\n📅 Структура кварталов (поиск заголовков годов):")
    for row in range(1, 15):
        for col in range(1, 200):
            val = ws.cell(row, col).value
            if isinstance(val, int) and val in (2022, 2023, 2024):
                print(f"  Год {val} найден: строка {row}, колонка {get_column_letter(col)} ({col})")
    
    # Проверяем ячейку E32 (2023 Q1 газ)
    print(f"\n🔍 Ячейка E32 (2023 Q1 газ):")
    e32_val = ws["E32"].value
    e32_type = ws["E32"].data_type
    print(f"  Значение: {e32_val}")
    print(f"  Тип: {e32_type}")
    if e32_type == 'f':
        print(f"  Формула: {ws['E32'].value}")
    elif e32_type == 'n':
        print(f"  Числовое значение: {e32_val}")
        if e32_val and abs(e32_val - 14.819) > 0.001:
            print(f"  ⚠️ ВНИМАНИЕ: Ожидается 14.819, получено {e32_val}")

# 2. Анализ листа "Объемы продукции"
if "Объемы продукции" in wb.sheetnames:
    ws = wb["Объемы продукции"]
    print("\n" + "=" * 80)
    print("ЛИСТ 'Объемы продукции'")
    print("=" * 80)
    
    print("\n📊 Структура таблицы:")
    for row in range(1, 10):
        row_data = [ws.cell(row, col).value for col in range(1, 8)]
        if any(v for v in row_data):
            print(f"  Строка {row}: {row_data}")

# 3. Анализ листа "Баланс"
if "Баланс" in wb.sheetnames:
    ws = wb["Баланс"]
    print("\n" + "=" * 80)
    print("ЛИСТ 'Баланс'")
    print("=" * 80)
    
    print("\n📊 Поиск структуры баланса:")
    for row in range(1, 50):
        val_a = ws.cell(row, 1).value
        if val_a and isinstance(val_a, str):
            val_lower = val_a.lower()
            if any(keyword in val_lower for keyword in [
                "баланс", "технологическ", "собственн", "производственн", "хозяйствен"
            ]):
                print(f"  Строка {row}: {val_a}")

# 4. Анализ листа "Расход  на ед.п"
if "Расход  на ед.п" in wb.sheetnames:
    ws = wb["Расход  на ед.п"]
    print("\n" + "=" * 80)
    print("ЛИСТ 'Расход  на ед.п'")
    print("=" * 80)
    
    print("\n📊 Структура листа:")
    for row in range(1, 30):
        val_a = ws.cell(row, 1).value
        val_b = ws.cell(row, 2).value
        if val_a or val_b:
            print(f"  Строка {row}: A={val_a}, B={val_b}")

# 5. Анализ листа "паспорт здание"
if "паспорт здание " in wb.sheetnames:
    ws = wb["паспорт здание "]
    print("\n" + "=" * 80)
    print("ЛИСТ 'паспорт здание'")
    print("=" * 80)
    
    print("\n📊 Поиск данных о здании:")
    for row in range(1, 50):
        val_a = ws.cell(row, 1).value
        val_b = ws.cell(row, 2).value
        if val_a and isinstance(val_a, str):
            val_lower = val_a.lower()
            if any(keyword in val_lower for keyword in [
                "площадь", "объем", "здание", "м²", "м3"
            ]):
                print(f"  Строка {row}: {val_a} = {val_b}")

wb.close()

# Формируем итоговый список необходимых данных
print("\n" + "=" * 80)
print("НЕОБХОДИМЫЕ ДАННЫЕ ДЛЯ ГЕНЕРАЦИИ ЭНЕРГОПАСПОРТА")
print("=" * 80)

print("""
1. ДАННЫЕ ПРЕДПРИЯТИЯ:
   - Название предприятия
   - ИНН
   - Адрес
   - Отрасль
   - Директор

2. ДАННЫЕ О ЗДАНИИ:
   - Площадь здания (м²)
   - Объем здания (м³)
   - Данные ограждающих конструкций (опционально)

3. ДАННЫЕ ПО ГАЗУ (помесячно за 2022-2024):
   - Январь, Февраль, Март, ... (м³)
   - Квартальные итоги (рассчитываются из месячных)
   - Годовые итоги (рассчитываются из квартальных)
   - Разбивка: собственные нужды (432 м³/мес), хоз-быт (остаток)

4. ДАННЫЕ ПО ЭЛЕКТРОЭНЕРГИИ:
   - Активная энергия (кВт·ч) по кварталам
   - Реактивная энергия (кВАр·ч) по кварталам
   - По видам продукции (5 видов):
     * Трубы ХВС (норма: 630 кВт)
     * Фитинги ХВС (норма: 2100 кВт)
     * Канализационные трубы (норма: 750 кВт)
     * Канализационные фитинги (норма: 2100 кВт)
     * Трубы тёплого пола (норма: 670 кВт)
   - Фактическое потребление по годам (2022, 2023, 2024)

5. ДАННЫЕ ПО ПРОИЗВОДСТВУ:
   - Объемы продукции по видам (т/шт/м)
   - Помесячные данные по производству
   - Квартальные и годовые итоги

6. ДАННЫЕ ПО ОБОРУДОВАНИЮ (опционально):
   - Перечень основного оборудования
   - Мощность оборудования
   - Год установки
   - Коэффициент использования

7. ДАННЫЕ ПО УЗЛАМ УЧЕТА (опционально):
   - Тип счетчика
   - Серийный номер
   - Место установки
   - Коэффициент учета

8. ДАННЫЕ ПО ПОТЕРЯМ (опционально):
   - Потери активной энергии (кВт·ч/мес)
   - Потери реактивной энергии (кВАр·ч/мес)
   - Мощность трансформатора (кВА)

9. НОРМАТИВЫ:
   - Норматив газа на м² (м³/(м²·год))
   - Норматив газа на условную единицу (м³/усл.ед.)
   - Нормативы по электроэнергии по видам продукции

10. ДРУГИЕ РЕСУРСЫ (опционально):
    - Вода (м³)
    - Тепло (Гкал)
    - Топливо (т)
    - Уголь (т)
""")

print("\n" + "=" * 80)
print("СТРУКТУРА ДАННЫХ ДЛЯ ПЕРЕДАЧИ В generate_energy_passport()")
print("=" * 80)

print("""
aggregated_data = {
    "resources": {
        "gas": {
            "2022-Q1": {
                "year": 2022,
                "quarter": 1,
                "months": [
                    {"month": "Январь", "values": {"volume_m3": 2500.0}},
                    {"month": "Февраль", "values": {"volume_m3": 2400.0}},
                    {"month": "Март", "values": {"volume_m3": 2600.0}},
                ],
                "quarter_totals": {"volume_m3": 7500.0},  # Рассчитывается из месяцев
            },
            # ... другие кварталы
        },
        "electricity": {
            "2022-Q1": {
                "year": 2022,
                "quarter": 1,
                "quarter_totals": {
                    "active_kwh": 100000.0,
                    "reactive_kvarh": 50000.0,
                },
            },
            # ... другие кварталы
        },
        "production": {
            "2022-Q1": {
                "year": 2022,
                "quarter": 1,
                "months": [
                    {"month": "Январь", "values": {
                        "Труба хвс": 50.0,
                        "Канал труба": 30.0,
                        # ... другие виды
                    }},
                    # ... другие месяцы
                ],
                "quarter_totals": {
                    "Труба хвс": 150.0,
                    # ... другие виды
                },
            },
            # ... другие кварталы
        },
    }
}

enterprise_data = {
    "id": "metin-iroda",
    "name": "Метин Ирода",
    "inn": "123456789",
    "address": "...",
    "director_name": "...",
    "industry": "...",
}

building_data = {
    "area_m2": 5000.0,  # Площадь здания в м²
    "volume_m3": 15000.0,  # Объем здания в м³ (опционально)
}
""")

print("\n✅ Анализ завершен")

