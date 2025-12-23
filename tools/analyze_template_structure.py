"""
Скрипт для анализа структуры шаблона энергопаспорта
и поиска ячеек для заполнения данных
"""
from openpyxl import load_workbook
from pathlib import Path

template_path = Path(r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx")

print("=" * 80)
print("АНАЛИЗ СТРУКТУРЫ ШАБЛОНА ЭНЕРГОПАСПОРТА")
print("=" * 80)

wb = load_workbook(template_path, data_only=False)
print(f"\n📋 Листы: {wb.sheetnames}\n")

# Анализ листа "Структура пр 2"
if "Структура пр 2" in wb.sheetnames:
    ws = wb["Структура пр 2"]
    print("=" * 80)
    print("ЛИСТ 'Структура пр 2'")
    print("=" * 80)
    
    # Поиск строки с газом для 2023 Q1 (должна быть строка 32, колонка E)
    print("\n🔍 Поиск данных по газу:")
    print(f"E32 (2023 Q1 газ): значение={ws['E32'].value}, тип={ws['E32'].data_type}")
    if ws['E32'].data_type == 'f':
        print(f"  Формула: {ws['E32'].value}")
    
    # Поиск строк с газом
    print("\n📊 Строки с данными по газу (2022-2024):")
    for row in range(30, 40):
        val_a = ws.cell(row, 1).value
        val_e = ws.cell(row, 5).value
        if val_a and isinstance(val_a, str) and ('газ' in val_a.lower() or 'потребление' in val_a.lower()):
            print(f"  Строка {row}: A={val_a}, E={val_e}")
    
    # Поиск таблицы электроэнергии по видам продукции
    print("\n🔍 Поиск таблицы электроэнергии по видам продукции:")
    for row in range(1, 150):
        val = ws.cell(row, 1).value
        if val and isinstance(val, str):
            val_lower = val.lower()
            if any(keyword in val_lower for keyword in ['труб', 'фитинг', 'канал', 'теплый пол', 'хвс', 'гвс']):
                print(f"  Строка {row}: {val}")
                # Показываем несколько колонок
                row_data = [ws.cell(row, col).value for col in range(1, 10)]
                print(f"    Данные: {row_data[:5]}")

# Анализ листа "Расход  на ед.п"
if "Расход  на ед.п" in wb.sheetnames:
    ws = wb["Расход  на ед.п"]
    print("\n" + "=" * 80)
    print("ЛИСТ 'Расход  на ед.п'")
    print("=" * 80)
    
    print("\n📊 Первые 30 строк:")
    for row in range(1, 31):
        val_a = ws.cell(row, 1).value
        val_b = ws.cell(row, 2).value
        if val_a or val_b:
            print(f"  Строка {row}: A={val_a}, B={val_b}")

# Анализ листа "Объемы продукции"
if "Объемы продукции" in wb.sheetnames:
    ws = wb["Объемы продукции"]
    print("\n" + "=" * 80)
    print("ЛИСТ 'Объемы продукции'")
    print("=" * 80)
    
    print("\n📊 Первые 20 строк:")
    for row in range(1, 21):
        row_data = [ws.cell(row, col).value for col in range(1, 8)]
        if any(v for v in row_data):
            print(f"  Строка {row}: {row_data}")

wb.close()
print("\n" + "=" * 80)
print("✅ Анализ завершен")
print("=" * 80)

