"""Анализ структуры файла 'Расчет газа для отопл и неотпл.xlsx'"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

file_path = Path("data/source_files/audit_sinergys/Расчет газа для отопл и неотпл.xlsx")

print("=" * 80)
print("АНАЛИЗ СТРУКТУРЫ ФАЙЛА ГАЗА")
print("=" * 80)
print(f"Файл: {file_path.name}\n")

if not file_path.exists():
    print(f"❌ Файл не найден: {file_path}")
    exit(1)

wb = load_workbook(file_path, data_only=True)

print(f"Листы в файле: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("=" * 80)
    print(f"ЛИСТ: {sheet_name}")
    print("=" * 80)
    
    # Анализируем первые 20 строк
    print("\nПервые 20 строк:")
    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val is not None:
                col_letter = get_column_letter(col)
                row_data.append(f"{col_letter}{row}={val}")
        if row_data:
            print(f"  Строка {row}: {', '.join(row_data[:10])}")
    
    # Ищем заголовки с годами
    print("\nПоиск заголовков с годами:")
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, min(50, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if isinstance(val, int) and val in (2022, 2023, 2024):
                col_letter = get_column_letter(col)
                print(f"  Год {val} найден: {col_letter}{row}")
    
    # Ищем месяцы
    print("\nПоиск месяцев:")
    months_found = []
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if isinstance(val, str) and any(month in val for month in ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]):
                col_letter = get_column_letter(col)
                months_found.append(f"{col_letter}{row}={val}")
    if months_found:
        print(f"  Найдено: {', '.join(months_found[:10])}")
    
    # Ищем числовые значения (вероятно объемы газа)
    print("\nПоиск числовых значений (объемы газа):")
    numeric_values = []
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(30, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if isinstance(val, (int, float)) and 1000 < val < 100000:
                col_letter = get_column_letter(col)
                numeric_values.append(f"{col_letter}{row}={val}")
                if len(numeric_values) >= 20:
                    break
        if len(numeric_values) >= 20:
            break
    if numeric_values:
        print(f"  Примеры: {', '.join(numeric_values[:10])}")

wb.close()

# Сохраняем результат в файл
output_file = Path("data/analysis/gas_file_structure.txt")
output_file.parent.mkdir(parents=True, exist_ok=True)
print(f"\n✅ Результат сохранен в: {output_file}")

