"""Тест заполнения электроэнергии по всем кварталам"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

output_path = r"test_output_metin\generated_passport_fixed.xlsx"

try:
    wb = load_workbook(output_path, data_only=True)
    ws = wb["Структура пр 2"]
    
    print("Проверка Q1 (строка 17):")
    print(f"  B17 (норма): {ws['B17'].value}")
    print(f"  C17 (2022): {ws['C17'].value}")
    print(f"  D17 (2023): {ws['D17'].value}")
    print(f"  E17 (2024): {ws['E17'].value}")
    
    print("\nПроверка Q2 (строка 39):")
    print(f"  Q39 (норма): {ws['Q39'].value}")
    print(f"  R39 (2022): {ws['R39'].value}")
    print(f"  S39 (2023): {ws['S39'].value}")
    print(f"  T39 (2024): {ws['T39'].value}")
    
    print("\nПроверка Q3 (строка 61):")
    print(f"  AF61 (норма): {ws['AF61'].value}")
    print(f"  AG61 (2022): {ws['AG61'].value}")
    print(f"  AH61 (2023): {ws['AH61'].value}")
    print(f"  AI61 (2024): {ws['AI61'].value}")
    
    print("\nПроверка Q4 (строка 83):")
    print(f"  AU83 (норма): {ws['AU83'].value}")
    print(f"  AV83 (2022): {ws['AV83'].value}")
    print(f"  AW83 (2023): {ws['AW83'].value}")
    print(f"  AX83 (2024): {ws['AX83'].value}")
    
    wb.close()
    print("\n✅ Проверка завершена")
except Exception as e:
    print(f"Ошибка: {e}")
    import traceback
    traceback.print_exc()

