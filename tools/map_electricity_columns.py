"""Определение маппинга колонок для таблиц электроэнергии по всем кварталам"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=True)
ws = wb["Структура пр 2"]

print("=" * 80)
print("МАППИНГ КОЛОНОК ДЛЯ ТАБЛИЦ ЭЛЕКТРОЭНЕРГИИ")
print("=" * 80)

# Анализируем строку 17 (Q1) - где есть заполненные данные
print("\nСтрока 17 (Q1, труба ХВС) - все колонки:")
row17_data = []
for col in range(1, 200):
    val = ws.cell(17, col).value
    if val is not None:
        col_letter = get_column_letter(col)
        row17_data.append((col, col_letter, val))

for col, col_letter, val in row17_data[:30]:
    print(f"  {col_letter}17 ({col}): {val}")

# Определяем структуру: где находятся данные для Q1
# По описанию пользователя: B=норма, C=2022, D=2023, E=2024, F=перерасход 2022, G=перерасход 2023, H=перерасход 2024
print("\nОпределенная структура для Q1 (строка 17):")
print("  B17 (колонка 2) = Норма")
print("  C17 (колонка 3) = Факт 2022")
print("  D17 (колонка 4) = Факт 2023")
print("  E17 (колонка 5) = Факт 2024")
print("  F17 (колонка 6) = Перерасход 2022")
print("  G17 (колонка 7) = Перерасход 2023")
print("  H17 (колонка 8) = Перерасход 2024")

# Ищем повторяющийся паттерн для других кварталов
# Проверяем строки 39 (Q2), 61 (Q3), 83 (Q4)
print("\n" + "=" * 80)
print("ПОИСК ПАТТЕРНА ДЛЯ ДРУГИХ КВАРТАЛОВ")
print("=" * 80)

# Таблицы находятся в строках:
# Q1: 17-21
# Q2: 39-43
# Q3: 61-65
# Q4: 83-87 (нужно проверить)

quarter_rows = {
    "2022-Q1": 17,
    "2022-Q2": 39,
    "2022-Q3": 61,
    "2022-Q4": 83,
    "2023-Q1": 17,  # Та же строка, другие колонки
    "2023-Q2": 39,
    "2023-Q3": 61,
    "2023-Q4": 83,
    "2024-Q1": 17,
    "2024-Q2": 39,
    "2024-Q3": 61,
    "2024-Q4": 83,
}

# Определяем смещение колонок для каждого квартала/года
# Проверяем, есть ли данные в других колонках строки 17
print("\nПроверка строки 17 на наличие данных в других колонках (для других годов):")
for col in range(1, 200):
    val = ws.cell(17, col).value
    if isinstance(val, (int, float)) and 100 < val < 10000:  # Вероятно факт
        col_letter = get_column_letter(col)
        # Проверяем заголовок колонки
        header_val = (
            ws.cell(16, col).value if ws.cell(16, col).value else ws.cell(15, col).value
        )
        print(f"  {col_letter}17: {val} (заголовок: {header_val})")

# Ищем паттерн: каждые ~16 колонок новый блок
print("\nГипотеза: блоки колонок повторяются каждые ~16 колонок")
print("Проверяем колонки для Q1 разных годов:")
blocks = [
    (2, 8, "2022-Q1"),
    (17, 23, "2023-Q1?"),
    (32, 38, "2024-Q1?"),
]

for start_col, end_col, label in blocks:
    print(f"\n{label} (колонки {start_col}-{end_col}):")
    for col in range(start_col, min(end_col + 1, 200)):
        val = ws.cell(17, col).value
        if val is not None:
            col_letter = get_column_letter(col)
            print(f"  {col_letter}17: {val}")

wb.close()
