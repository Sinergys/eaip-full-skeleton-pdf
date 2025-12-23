"""Проверка структуры таблиц электроэнергии в шаблоне"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=False)
ws = wb["Структура пр 2"]

print("=" * 80)
print("ПРОВЕРКА СТРУКТУРЫ ТАБЛИЦ ЭЛЕКТРОЭНЕРГИИ")
print("=" * 80)

# Проверяем строки для каждого квартала
quarters = {
    "Q1": {"rows": [17, 18, 19, 20, 21]},
    "Q2": {"rows": [39, 40, 41, 42, 43]},
    "Q3": {"rows": [61, 62, 63, 64, 65]},
    "Q4": {"rows": [83, 84, 85, 86, 87]},
}

for quarter, info in quarters.items():
    print(f"\n{'='*60}")
    print(f"КВАРТАЛ {quarter}")
    print(f"{'='*60}")
    
    for row in info["rows"]:
        product_name = ws.cell(row, 1).value
        print(f"\nСтрока {row}: {product_name}")
        
        # Проверяем все колонки с данными
        for col in range(1, 100):
            val = ws.cell(row, col).value
            if val is not None and isinstance(val, (int, float)) and val > 0:
                col_letter = get_column_letter(col)
                print(f"  {col_letter}{row}: {val}")

# Проверяем, где находятся таблицы для разных кварталов
print("\n" + "=" * 80)
print("ПОИСК НАЧАЛА ТАБЛИЦ ДЛЯ КАЖДОГО КВАРТАЛА")
print("=" * 80)

# Ищем колонки, где есть "труба ХВС" (начало таблицы)
for quarter, info in quarters.items():
    first_row = info["rows"][0]
    print(f"\n{quarter} (строка {first_row}):")
    
    # Ищем все колонки с "труба ХВС"
    table_starts = []
    for col in range(1, 200):
        val = ws.cell(first_row, col).value
        if val and isinstance(val, str) and "труба" in val.lower():
            col_letter = get_column_letter(col)
            table_starts.append((col, col_letter))
            print(f"  Начало таблицы: {col_letter} ({col})")
    
    # Для каждого начала таблицы проверяем структуру колонок
    for start_col, start_letter in table_starts:
        print(f"\n  Структура таблицы, начинающейся с {start_letter}:")
        # Проверяем колонки: название, норма, факт 2022, факт 2023, факт 2024
        for offset in range(0, 8):
            col = start_col + offset
            col_letter = get_column_letter(col)
            val = ws.cell(first_row, col).value
            if val is not None:
                print(f"    {col_letter}{first_row}: {val}")

wb.close()

