"""
Детальная проверка шаблона энергопаспорта на ошибки.
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict
import re

template_path = Path(r"C:\eaip\data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx")

print("=" * 80)
print("ПРОВЕРКА ШАБЛОНА ЭНЕРГОПАСПОРТА НА ОШИБКИ")
print("=" * 80)
print(f"Файл: {template_path.name}\n")

if not template_path.exists():
    print(f"❌ Файл не найден: {template_path}")
    exit(1)

# Загружаем с вычисленными значениями для проверки ошибок
wb_data = load_workbook(template_path, data_only=True)
# Загружаем с формулами для анализа
wb_formulas = load_workbook(template_path, data_only=False)

print(f"✅ Файл загружен")
print(f"📋 Листы: {wb_data.sheetnames}\n")

# Счетчики ошибок
errors_by_type = defaultdict(list)
warnings = []
info = []

# Проверка каждого листа
for sheet_name in wb_data.sheetnames:
    ws_data = wb_data[sheet_name]
    ws_formulas = wb_formulas[sheet_name]
    
    print("=" * 80)
    print(f"ЛИСТ: '{sheet_name}'")
    print("=" * 80)
    
    sheet_errors = []
    sheet_warnings = []
    
    # Проверяем все ячейки на ошибки Excel
    for row in ws_data.iter_rows():
        for cell in row:
            if cell.value is not None:
                value_str = str(cell.value)
                
                # Проверка на ошибки Excel
                if value_str.startswith('#'):
                    error_info = {
                        'sheet': sheet_name,
                        'cell': cell.coordinate,
                        'error': value_str,
                        'row': cell.row,
                        'col': cell.column_letter,
                    }
                    errors_by_type[value_str].append(error_info)
                    sheet_errors.append(error_info)
                
                # Проверка на необычные значения
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        sheet_warnings.append({
                            'cell': cell.coordinate,
                            'message': f'Отрицательное значение: {cell.value}',
                        })
                    if abs(cell.value) > 1e10:
                        sheet_warnings.append({
                            'cell': cell.coordinate,
                            'message': f'Очень большое значение: {cell.value}',
                        })
    
    # Специальные проверки для ключевых листов
    if sheet_name == "Структура пр 2":
        print("\n🔍 Специальные проверки для 'Структура пр 2':")
        
        # Проверка E32 (2023 Q1 газ)
        e32_data = ws_data["E32"].value
        e32_formula = ws_formulas["E32"]
        
        print(f"\n  Ячейка E32 (2023 Q1 газ):")
        print(f"    Значение: {e32_data}")
        print(f"    Тип данных: {ws_data['E32'].data_type}")
        
        if e32_formula.data_type == 'f':
            print(f"    Формула: {e32_formula.value}")
        elif e32_formula.data_type == 'n':
            print(f"    Числовое значение (не формула)")
            if e32_data and abs(e32_data - 14.819) > 0.001:
                error_info = {
                    'sheet': sheet_name,
                    'cell': 'E32',
                    'row': 32,
                    'error': 'НЕПРАВИЛЬНОЕ_ЗНАЧЕНИЕ',
                    'expected': 14.819,
                    'actual': e32_data,
                    'message': f'Ожидается 14.819, получено {e32_data}',
                }
                errors_by_type['НЕПРАВИЛЬНОЕ_ЗНАЧЕНИЕ'].append(error_info)
                print(f"    ❌ ОШИБКА: Ожидается 14.819, получено {e32_data}")
        
        # Проверка структуры данных по газу
        print(f"\n  Проверка структуры данных по газу:")
        gas_rows = [32, 34, 36]  # 2022
        for row in gas_rows:
            val_a = ws_data.cell(row, 1).value
            val_e = ws_data.cell(row, 5).value
            if val_a:
                print(f"    Строка {row} ({val_a}): E={val_e}")
        
        # Проверка электроэнергии по видам продукции
        print(f"\n  Проверка электроэнергии по видам продукции:")
        product_rows = [17, 18, 19, 20]  # 2022
        for row in product_rows:
            val_a = ws_data.cell(row, 1).value
            if val_a:
                row_data = [ws_data.cell(row, col).value for col in range(2, 9)]
                print(f"    Строка {row} ({val_a}): {row_data[:3]}")
    
    elif sheet_name == "Объемы продукции":
        print("\n🔍 Специальные проверки для 'Объемы продукции':")
        
        # Проверяем таблицу продукции
        for row in range(2, 7):
            product_name = ws_data.cell(row, 2).value
            norm = ws_data.cell(row, 3).value
            fact_2022 = ws_data.cell(row, 4).value
            fact_2023 = ws_data.cell(row, 5).value
            fact_2024 = ws_data.cell(row, 6).value
            
            if product_name:
                print(f"  {product_name}:")
                print(f"    Норма: {norm}, Факт 2022: {fact_2022}, 2023: {fact_2023}, 2024: {fact_2024}")
                
                # Проверка на наличие данных
                if not norm or not fact_2022 or not fact_2023 or not fact_2024:
                    sheet_warnings.append({
                        'cell': f'B{row}',
                        'message': f'Неполные данные для {product_name}',
                    })
    
    elif sheet_name == "Баланс":
        print("\n🔍 Специальные проверки для 'Баланс':")
        
        # Проверяем наличие ошибок в формулах
        error_cells = []
        for row in ws_data.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('#'):
                    error_cells.append(cell.coordinate)
        
        if error_cells:
            print(f"  ⚠️ Найдено {len(error_cells)} ячеек с ошибками")
            for cell_coord in error_cells[:10]:
                cell_formula = ws_formulas[cell_coord]
                if cell_formula.data_type == 'f':
                    print(f"    {cell_coord}: {cell_formula.value}")
        else:
            print(f"  ✅ Ошибок не найдено")
    
    # Выводим ошибки и предупреждения для листа
    if sheet_errors:
        print(f"\n  ❌ Найдено ошибок: {len(sheet_errors)}")
        for err in sheet_errors[:5]:
            print(f"    {err['cell']}: {err['error']}")
        if len(sheet_errors) > 5:
            print(f"    ... и еще {len(sheet_errors) - 5} ошибок")
    
    if sheet_warnings:
        print(f"\n  ⚠️ Найдено предупреждений: {len(sheet_warnings)}")
        for warn in sheet_warnings[:5]:
            print(f"    {warn['cell']}: {warn['message']}")
        if len(sheet_warnings) > 5:
            print(f"    ... и еще {len(sheet_warnings) - 5} предупреждений")

# Итоговая сводка
print("\n" + "=" * 80)
print("ИТОГОВАЯ СВОДКА ОШИБОК")
print("=" * 80)

total_errors = sum(len(errors) for errors in errors_by_type.values())

if total_errors == 0:
    print("\n✅ Ошибок Excel не обнаружено!")
else:
    print(f"\n❌ Найдено ошибок: {total_errors}\n")
    
    for error_type, errors in sorted(errors_by_type.items()):
        print(f"\n{error_type}: {len(errors)} ячеек")
        
        # Показываем примеры
        for err in errors[:10]:
            row_info = f" (строка {err['row']})" if 'row' in err else ""
            print(f"  • {err['sheet']}!{err['cell']}{row_info}")
            if 'message' in err:
                print(f"    {err['message']}")
        
        if len(errors) > 10:
            print(f"  ... и еще {len(errors) - 10} ячеек")
        
        # Анализ по листам
        sheets_affected = defaultdict(int)
        for err in errors:
            sheets_affected[err['sheet']] += 1
        
        if len(sheets_affected) > 1:
            print(f"  Затронуто листов: {len(sheets_affected)}")
            for sheet, count in sorted(sheets_affected.items(), key=lambda x: -x[1]):
                print(f"    - {sheet}: {count} ошибок")

# Специальная проверка формул с ошибками
if total_errors > 0:
    print("\n" + "=" * 80)
    print("АНАЛИЗ ФОРМУЛ С ОШИБКАМИ")
    print("=" * 80)
    
    for error_type, errors in sorted(errors_by_type.items()):
        if error_type.startswith('#'):
            print(f"\n{error_type}:")
            
            # Анализируем первые 5 ошибок
            for err in errors[:5]:
                cell = wb_formulas[err['sheet']][err['cell']]
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    print(f"  {err['sheet']}!{err['cell']}:")
                    print(f"    Формула: {formula[:100]}")
                    
                    # Анализ типа ошибки
                    if '#DIV/0!' in error_type:
                        if '/' in formula:
                            print(f"    ⚠️ Деление на ноль в формуле")
                    elif '#REF!' in error_type:
                        if "'" in formula or "!" in formula:
                            print(f"    ⚠️ Ссылка на несуществующую ячейку/лист")
                    elif '#VALUE!' in error_type:
                        print(f"    ⚠️ Неправильный тип данных в формуле")

# Рекомендации
print("\n" + "=" * 80)
print("РЕКОМЕНДАЦИИ ПО ИСПРАВЛЕНИЮ")
print("=" * 80)

if '#DIV/0!' in errors_by_type:
    print("\n1. ОШИБКИ ДЕЛЕНИЯ НА НОЛЬ (#DIV/0!):")
    print(f"   Найдено: {len(errors_by_type['#DIV/0!'])} ячеек")
    print("   Решение:")
    print("   • Заменить =A1/B1 на =IF(B1=0, 0, A1/B1)")
    print("   • Или использовать =IFERROR(A1/B1, 0)")
    print("   • Проверить исходные данные - возможно, делитель должен быть заполнен")

if '#REF!' in errors_by_type:
    print("\n2. ОШИБКИ ССЫЛОК (#REF!):")
    print(f"   Найдено: {len(errors_by_type['#REF!'])} ячеек")
    print("   Решение:")
    print("   • Проверить ссылки на другие листы - возможно, листы переименованы")
    print("   • Проверить ссылки на ячейки - возможно, строки/колонки были удалены")
    print("   • Исправить имена листов в формулах (убедиться в пробелах в конце)")

if 'НЕПРАВИЛЬНОЕ_ЗНАЧЕНИЕ' in errors_by_type:
    print("\n3. НЕПРАВИЛЬНЫЕ ЗНАЧЕНИЯ:")
    for err in errors_by_type['НЕПРАВИЛЬНОЕ_ЗНАЧЕНИЕ']:
        print(f"   • {err['sheet']}!{err['cell']}: {err['message']}")
        print(f"     Решение: Использовать расчет из помесячных данных при генерации")

wb_data.close()
wb_formulas.close()

print("\n" + "=" * 80)
print("✅ Проверка завершена")
print("=" * 80)

