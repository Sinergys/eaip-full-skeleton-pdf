"""Проверка заполнения таблиц электроэнергии по всем кварталам"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

output_path = r"test_output_metin\generated_passport_fixed.xlsx"

print("=" * 80)
print("ПРОВЕРКА ЗАПОЛНЕНИЯ ТАБЛИЦ ЭЛЕКТРОЭНЕРГИИ ПО ВСЕМ КВАРТАЛАМ")
print("=" * 80)

wb = load_workbook(output_path, data_only=True)
ws = wb["Структура пр 2"]

# Проверяем все кварталы
quarters = {
    "Q1": {"rows": [17, 18, 19, 20, 21], "start_col": 1, "products": ["труба ХВС", "фитинги ХВС и ГВС", "канализационные трубы", "канализационные фитинги", "тёплый пол"]},
    "Q2": {"rows": [39, 40, 41, 42, 43], "start_col": 16, "products": ["труба ХВС", "фитинги ХВС и ГВС", "канализационные трубы", "канализационные фитинги", "тёплый пол"]},
    "Q3": {"rows": [61, 62, 63, 64, 65], "start_col": 31, "products": ["труба ХВС", "фитинги ХВС и ГВС", "канализационные трубы", "канализационные фитинги", "тёплый пол"]},
    "Q4": {"rows": [83, 84, 85, 86, 87], "start_col": 46, "products": ["труба ХВС", "фитинги ХВС и ГВС", "канализационные трубы", "канализационные фитинги", "тёплый пол"]},
}

for quarter, info in quarters.items():
    print(f"\n{'='*80}")
    print(f"КВАРТАЛ {quarter}")
    print(f"{'='*80}")
    
    start_col = info["start_col"]
    rows = info["rows"]
    products = info["products"]
    
    for i, (row, product) in enumerate(zip(rows, products)):
        print(f"\n{product} (строка {row}):")
        
        # Проверяем колонки: норма, факт 2022, факт 2023, факт 2024
        col_norm = start_col + 1
        col_2022 = start_col + 2
        col_2023 = start_col + 3
        col_2024 = start_col + 4
        
        norm = ws.cell(row, col_norm).value
        fact_2022 = ws.cell(row, col_2022).value
        fact_2023 = ws.cell(row, col_2023).value
        fact_2024 = ws.cell(row, col_2024).value
        
        print(f"  Норма ({get_column_letter(col_norm)}): {norm}")
        print(f"  Факт 2022 ({get_column_letter(col_2022)}): {fact_2022}")
        print(f"  Факт 2023 ({get_column_letter(col_2023)}): {fact_2023}")
        print(f"  Факт 2024 ({get_column_letter(col_2024)}): {fact_2024}")
        
        # Проверяем заполненность
        filled = sum(1 for v in [norm, fact_2022, fact_2023, fact_2024] if v is not None and v != 0)
        if filled > 0:
            print(f"  ✅ Заполнено: {filled}/4 полей")
        else:
            print(f"  ❌ Не заполнено")

wb.close()

print("\n" + "=" * 80)
print("ПРОВЕРКА ЗАВЕРШЕНА")
print("=" * 80)

