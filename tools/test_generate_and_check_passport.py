"""
Скрипт для генерации паспорта и проверки результата:
1. Находит batch_id из БД
2. Генерирует паспорт через API или напрямую
3. Проверяет наличие данных и формул
"""
import sys
from pathlib import Path
import httpx
from openpyxl import load_workbook

# Добавляем пути
sys.path.insert(0, str(Path(__file__).parent.parent / "eaip_full_skeleton" / "services" / "ingest"))
import database

def find_latest_batch_id():
    """Находит последний batch_id из БД"""
    try:
        import sqlite3
        from pathlib import Path
        
        # Находим путь к БД (пробуем несколько вариантов)
        possible_paths = [
            Path(__file__).parent.parent / "eaip_full_skeleton" / "services" / "ingest" / "ingest_data.db",
            Path("C:/eaip/eaip_full_skeleton/services/ingest/ingest_data.db"),
            Path("ingest_data.db"),
        ]
        
        db_path = None
        for path in possible_paths:
            if path.exists():
                db_path = path
                break
        
        if not db_path:
            print(f"❌ БД не найдена. Проверял пути:")
            for path in possible_paths:
                print(f"   - {path}")
            return None, None, None
        
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                """
                SELECT batch_id, filename, enterprise_id, status
                FROM uploads
                WHERE status = 'completed'
                ORDER BY created_at DESC
                LIMIT 1
                """
            ).fetchone()
            if row:
                return row["batch_id"], row["filename"], row["enterprise_id"]
        finally:
            conn.close()
    except Exception as e:
        print(f"Ошибка при поиске batch_id: {e}")
        import traceback
        traceback.print_exc()
    return None, None, None

def generate_passport_via_api(batch_id: str, base_url: str = "http://localhost:8001"):
    """Генерирует паспорт через API endpoint"""
    try:
        url = f"{base_url}/api/generate-passport/{batch_id}"
        print(f"Вызов API: {url}")
        
        with httpx.Client(timeout=300.0) as client:
            response = client.post(url)
            response.raise_for_status()
            
            # Сохраняем файл
            output_path = Path("/tmp/passports") / f"{batch_id}_energy_passport.xlsx"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_bytes(response.content)
            print(f"✅ Паспорт сохранен: {output_path}")
            return output_path
    except Exception as e:
        print(f"❌ Ошибка при генерации через API: {e}")
        return None

def check_passport_formulas_and_data(passport_path: Path):
    """Проверяет наличие формул и данных в паспорте"""
    if not passport_path.exists():
        print(f"❌ Файл не найден: {passport_path}")
        return False
    
    print("\n" + "=" * 80)
    print(f"ПРОВЕРКА ПАСПОРТА: {passport_path.name}")
    print("=" * 80)
    
    try:
        wb = load_workbook(passport_path, data_only=False)
        
        print(f"\n📊 Всего листов: {len(wb.sheetnames)}")
        print("\n📋 Список всех листов:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            formulas = sum(1 for row in ws.iter_rows() for cell in row if cell.data_type == 'f' and cell.value)
            data_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
            print(f"  {i}. '{sheet_name}'")
            print(f"     - Строк: {ws.max_row}, Колонок: {ws.max_column}")
            print(f"     - Формул: {formulas}, Ячеек с данными: {data_cells}")
        
        # Проверяем ключевые листы (расширенный список вариантов)
        key_sheets = {
            "Balans": ["04_Баланс", "Баланс", "Balance", "Balans", "Баланс ", "04_Баланс ", "Energy"],
            "Struktura pr2": ["Структура пр 2", "Struktura pr2", "02_Структура", "Структура пр 2 ", "Struktura pr2 ", "02_Структура ", "Energy"],
            "Equipment": ["03_Оборудование", "Equipment", "Оборудование", "Sheet1", "03_Оборудование ", "Equipment "],
        }
        
        results = {}
        
        for sheet_category, sheet_names in key_sheets.items():
            found_sheet = None
            for name in sheet_names:
                # Проверяем точное совпадение и варианты с пробелами
                for sheet_name in wb.sheetnames:
                    if sheet_name.strip() == name.strip() or sheet_name == name:
                        found_sheet = sheet_name
                        break
                if found_sheet:
                    break
            
            if not found_sheet:
                print(f"\n❌ {sheet_category}: лист не найден")
                results[sheet_category] = {"found": False}
                continue
            
            ws = wb[found_sheet]
            
            # Подсчитываем формулы и данные
            formulas = []
            data_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formulas.append((cell.coordinate, str(cell.value)))
                    elif cell.value is not None and cell.value != "":
                        data_cells.append((cell.coordinate, cell.value))
            
            print(f"\n✅ {sheet_category} (лист '{found_sheet}'):")
            print(f"   - Формул: {len(formulas)}")
            print(f"   - Ячеек с данными: {len(data_cells)}")
            
            # Показываем примеры формул
            if formulas:
                print(f"   - Примеры формул:")
                for coord, formula in formulas[:5]:
                    print(f"     {coord}: {formula[:60]}...")
            else:
                print(f"   ⚠️  ВНИМАНИЕ: Формулы не найдены!")
            
            # Показываем примеры данных
            if data_cells:
                print(f"   - Примеры данных:")
                for coord, value in data_cells[:5]:
                    if isinstance(value, (int, float)):
                        print(f"     {coord}: {value}")
                    elif isinstance(value, str) and len(value) < 50:
                        print(f"     {coord}: {value}")
            
            results[sheet_category] = {
                "found": True,
                "formulas_count": len(formulas),
                "data_count": len(data_cells),
                "has_formulas": len(formulas) > 0,
                "has_data": len(data_cells) > 0,
            }
        
        # Специальная проверка для Struktura pr2 - проверяем заполнение ресурсов
        struktura_sheet = None
        for category, names in key_sheets.items():
            if category == "Struktura pr2":
                for name in names:
                    for sheet_name in wb.sheetnames:
                        if sheet_name.strip() == name.strip() or sheet_name == name:
                            struktura_sheet = wb[sheet_name]
                            break
                    if struktura_sheet:
                        break
                break
        
        if struktura_sheet:
            print(f"\n" + "=" * 80)
            print("ПРОВЕРКА ЗАПОЛНЕНИЯ STRUKTURA PR2")
            print("=" * 80)
            
            # Проверяем заполнение данных по кварталам
            # Согласно build_quarter_mapping, данные должны быть в строках 9 для каждого квартала
            quarters_filled = 0
            resources_found = {"electricity": False, "gas": False, "water": False, "heat": False}
            
            # Проверяем ячейки для кварталов 2022-Q1 (колонки 3, 4, 6, 14)
            # Активная энергия: колонка 3, строка 9
            # Реактивная энергия: колонка 4, строка 9
            # Газ: колонка 6, строка 9
            # Вода: колонка 14, строка 9
            
            for quarter_col in [3, 19, 35, 51, 67, 83, 99, 115, 131, 147, 163, 179]:
                row = 9
                active_cell = struktura_sheet.cell(row=row, column=quarter_col)
                reactive_cell = struktura_sheet.cell(row=row, column=quarter_col + 1)
                gas_cell = struktura_sheet.cell(row=row, column=quarter_col + 3)
                water_cell = struktura_sheet.cell(row=row, column=quarter_col + 11)
                
                has_data = False
                if active_cell.value and isinstance(active_cell.value, (int, float)) and active_cell.value > 0:
                    resources_found["electricity"] = True
                    has_data = True
                if reactive_cell.value and isinstance(reactive_cell.value, (int, float)) and reactive_cell.value > 0:
                    has_data = True
                if gas_cell.value and isinstance(gas_cell.value, (int, float)) and gas_cell.value > 0:
                    resources_found["gas"] = True
                    has_data = True
                if water_cell.value and isinstance(water_cell.value, (int, float)) and water_cell.value > 0:
                    resources_found["water"] = True
                    has_data = True
                
                if has_data:
                    quarters_filled += 1
            
            print(f"   - Кварталов с данными: {quarters_filled}/12")
            print(f"   - Ресурсы найдены:")
            for resource, found in resources_found.items():
                status = "✅" if found else "❌"
                print(f"     {status} {resource}")
        
        wb.close()
        
        # Итоговая оценка
        print(f"\n" + "=" * 80)
        print("ИТОГОВАЯ ОЦЕНКА")
        print("=" * 80)
        
        all_ok = True
        for category, result in results.items():
            if not result.get("found"):
                print(f"❌ {category}: лист не найден")
                all_ok = False
            elif not result.get("has_formulas"):
                print(f"⚠️  {category}: формулы не найдены")
                all_ok = False
            elif not result.get("has_data"):
                print(f"⚠️  {category}: данные не найдены")
                all_ok = False
            else:
                print(f"✅ {category}: OK (формул: {result['formulas_count']}, данных: {result['data_count']})")
        
        return all_ok
        
    except Exception as e:
        print(f"\n❌ Ошибка при проверке паспорта: {e}")
        import traceback
        traceback.print_exc()
        return False

def find_existing_passport():
    """Ищет уже сгенерированный паспорт"""
    possible_dirs = [
        Path("/tmp/passports"),
        Path("C:/tmp/passports"),
        Path(__file__).parent.parent / "data" / "source_files" / "audit_sinergys",
    ]
    
    for dir_path in possible_dirs:
        if dir_path.exists():
            passports = sorted(
                dir_path.glob("*_energy_passport.xlsx"),
                key=lambda p: p.stat().st_mtime if p.exists() else 0,
                reverse=True
            )
            if passports:
                return passports[0]
    
    return None

def main():
    print("=" * 80)
    print("ГЕНЕРАЦИЯ И ПРОВЕРКА ЭНЕРГОПАСПОРТА")
    print("=" * 80)
    
    # Если передан путь к файлу как аргумент
    if len(sys.argv) > 1 and Path(sys.argv[1]).exists():
        passport_path = Path(sys.argv[1])
        print(f"\n✅ Используется указанный файл: {passport_path}")
    else:
        # Находим batch_id
        batch_id, filename, enterprise_id = find_latest_batch_id()
        
        if batch_id:
            print(f"\n✅ Найден batch_id: {batch_id}")
            print(f"   Файл: {filename}")
            print(f"   Предприятие ID: {enterprise_id}")
            
            # Пробуем сгенерировать через API
            base_url = sys.argv[1] if len(sys.argv) > 1 else "http://localhost:8001"
            passport_path = generate_passport_via_api(batch_id, base_url)
            
            if not passport_path:
                print("\n⚠️  Генерация через API не удалась. Ищу уже сгенерированный паспорт...")
                passport_path = find_existing_passport()
        else:
            print("\n⚠️  Не найден batch_id в БД. Ищу уже сгенерированный паспорт...")
            passport_path = find_existing_passport()
        
        if not passport_path:
            print("\n❌ Не найден сгенерированный паспорт.")
            print("   Варианты:")
            print("   1. Загрузите файлы через UI и сгенерируйте паспорт")
            print("   2. Укажите путь к файлу: python test_generate_and_check_passport.py <путь_к_файлу>")
            return
        
        print(f"\n✅ Найден паспорт: {passport_path}")
    
    # Проверяем паспорт
    success = check_passport_formulas_and_data(passport_path)
    
    if success:
        print("\n✅ ВСЕ ПРОВЕРКИ ПРОЙДЕНЫ!")
    else:
        print("\n⚠️  Некоторые проверки не пройдены. Проверьте вывод выше.")

if __name__ == "__main__":
    main()

