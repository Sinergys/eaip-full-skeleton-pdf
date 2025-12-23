"""Поиск всех колонок для каждого квартала и года"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=True)
ws = wb["Структура пр 2"]

print("=" * 80)
print("ПОИСК КОЛОНОК ДЛЯ ВСЕХ КВАРТАЛОВ И ГОДОВ")
print("=" * 80)

# Проверяем строку 17 (Q1) - где есть данные
print("\nСтрока 17 (Q1) - проверка всех колонок с числовыми данными:")
for col in range(1, 200):
    val = ws.cell(17, col).value
    if isinstance(val, (int, float)) and val > 10:
        col_letter = get_column_letter(col)
        # Проверяем заголовки выше
        header1 = ws.cell(16, col).value
        header2 = ws.cell(15, col).value
        header3 = ws.cell(14, col).value
        print(f"  {col_letter}17: {val}")
        if header1:
            print(f"    Заголовок строка 16: {header1}")
        if header2:
            print(f"    Заголовок строка 15: {header2}")
        if header3:
            print(f"    Заголовок строка 14: {header3}")

# Проверяем строку 39 (Q2)
print("\n" + "=" * 80)
print("Строка 39 (Q2) - проверка всех колонок:")
for col in range(1, 200):
    val = ws.cell(39, col).value
    if isinstance(val, (int, float)) and val > 10:
        col_letter = get_column_letter(col)
        print(f"  {col_letter}39: {val}")

# Ищем паттерн: проверяем, где в строке 17 находятся повторяющиеся блоки
print("\n" + "=" * 80)
print("АНАЛИЗ ПОВТОРЯЮЩИХСЯ БЛОКОВ")
print("=" * 80)

# Находим все колонки, где в строке 17 есть "труба ХВС" (начало таблицы)
product_columns = []
for col in range(1, 200):
    val = ws.cell(17, col).value
    if val and isinstance(val, str) and "труба" in val.lower():
        col_letter = get_column_letter(col)
        product_columns.append((col, col_letter))
        print(f"Найдена таблица в колонке {col_letter} ({col})")

print(f"\nВсего найдено таблиц: {len(product_columns)}")

# Для каждой таблицы определяем структуру колонок
print("\nСтруктура колонок для каждой таблицы:")
for i, (start_col, start_letter) in enumerate(product_columns):
    print(f"\nТаблица {i + 1} (начинается с {start_letter}):")

    # Проверяем колонки после начала таблицы
    # Структура: A=название, B=норма, C=2022, D=2023, E=2024, F=перерасход 2022, G=перерасход 2023, H=перерасход 2024
    structure = {
        "название": start_col,
        "норма": start_col + 1,
        "факт_2022": start_col + 2,
        "факт_2023": start_col + 3,
        "факт_2024": start_col + 4,
        "перерасход_2022": start_col + 5,
        "перерасход_2023": start_col + 6,
        "перерасход_2024": start_col + 7,
    }

    for key, col_num in structure.items():
        col_letter = get_column_letter(col_num)
        val = ws.cell(17, col_num).value
        print(f"  {col_letter}17 ({key}): {val}")

wb.close()
