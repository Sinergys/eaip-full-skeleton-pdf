"""Детальная проверка структуры Excel файла"""
import openpyxl

file_path = r"C:\AUDIT\OBJECTS\Navoiy IES\INBOX\акт выполненых работ май.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

print("=" * 80)
print("ДЕТАЛЬНАЯ ПРОВЕРКА СТРУКТУРЫ EXCEL")
print("=" * 80)
print(f"Всего строк: {sheet.max_row}, столбцов: {sheet.max_column}\n")

# Показываем строки 15-35
print("Строки 15-35:")
for row_idx in range(15, min(36, sheet.max_row + 1)):
    row = list(sheet.iter_rows(values_only=True))[row_idx - 1]
    row_data = [str(cell)[:30] if cell else "" for cell in row[:10]]
    non_empty = sum(1 for c in row_data if c)
    if non_empty > 0:
        print(f"Строка {row_idx}: {non_empty} непустых ячеек")
        print(f"  Первые 5: {row_data[:5]}")

# Ищем строку с "№" в первом столбце
print("\n" + "=" * 80)
print("Поиск строки с '№' в первом столбце:")
for row_idx in range(1, min(40, sheet.max_row + 1)):
    row = list(sheet.iter_rows(values_only=True))[row_idx - 1]
    if row and str(row[0]).strip() in ['№', 'N', 'No', '1', '2', '3']:
        print(f"Строка {row_idx}: первый столбец = '{row[0]}'")
        print(f"  Непустых ячеек: {sum(1 for c in row if c)}")
        print(f"  Первые 5: {[str(c)[:30] if c else '' for c in row[:5]]}")

