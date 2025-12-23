"""Анализ структуры листа 'Баланс' в существующих паспортах"""
import sys
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def analyze_balance_sheet(file_path: Path):
    """Анализирует структуру листа 'Баланс'"""
    print(f"\n{'='*60}")
    print(f"АНАЛИЗ: {file_path.name}")
    print('='*60)
    
    if not file_path.exists():
        print(f"❌ Файл не найден: {file_path}")
        return
    
    try:
        wb = load_workbook(file_path, data_only=False)
        
        # Ищем лист "Баланс"
        balance_sheet = None
        for sheet_name in ["04_Баланс", "Баланс", "Balans", "04_Balans"]:
            if sheet_name in wb.sheetnames:
                balance_sheet = wb[sheet_name]
                print(f"✅ Найден лист: {sheet_name}")
                break
        
        if not balance_sheet:
            print("❌ Лист 'Баланс' не найден")
            print(f"   Доступные листы: {wb.sheetnames[:10]}")
            wb.close()
            return
        
        # Анализируем структуру
        print(f"\n📊 СТРУКТУРА ЛИСТА:")
        print(f"   Максимальная строка: {balance_sheet.max_row}")
        print(f"   Максимальная колонка: {balance_sheet.max_column}")
        
        # Анализируем первые 10 строк
        print(f"\n📋 ПЕРВЫЕ 10 СТРОК:")
        for row_idx in range(1, min(11, balance_sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(7, balance_sheet.max_column + 1)):
                cell = balance_sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    if cell.data_type == "f":
                        row_data.append(f"={cell.value}")
                    else:
                        row_data.append(str(cell.value)[:30])
                else:
                    row_data.append("")
            print(f"   Строка {row_idx}: {' | '.join(row_data)}")
        
        # Ищем строку с "ИТОГО"
        print(f"\n🔍 ПОИСК 'ИТОГО':")
        for row_idx in range(1, balance_sheet.max_row + 1):
            for col_idx in range(1, min(7, balance_sheet.max_column + 1)):
                cell = balance_sheet.cell(row=row_idx, column=col_idx)
                if cell.value and "ИТОГО" in str(cell.value).upper():
                    print(f"   ✅ Найдено в строке {row_idx}, колонка {col_idx}: {cell.value}")
        
        # Ищем формулы
        print(f"\n🔢 ФОРМУЛЫ:")
        formula_count = 0
        for row_idx in range(1, min(20, balance_sheet.max_row + 1)):
            for col_idx in range(1, min(7, balance_sheet.max_column + 1)):
                cell = balance_sheet.cell(row=row_idx, column=col_idx)
                if cell.data_type == "f":
                    formula_count += 1
                    if formula_count <= 5:  # Показываем первые 5 формул
                        print(f"   Строка {row_idx}, колонка {col_idx}: ={cell.value}")
        print(f"   Всего формул найдено: {formula_count}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")

def main():
    if len(sys.argv) < 2:
        # Тестируем несколько файлов
        test_files = [
            "data/aggregated/EnergyPassport_PKM690_test7.xlsx",
            "data/aggregated/EnergyPassport_PKM690_filled_test.xlsx",
            "EnergyPassport_PKM690_filled_test.xlsx"
        ]
        
        for file_path in test_files:
            path = Path(file_path)
            if path.exists():
                analyze_balance_sheet(path)
                break
    else:
        analyze_balance_sheet(Path(sys.argv[1]))

if __name__ == "__main__":
    main()

