"""Анализ эталонного файла энергопаспорта для понимания правильной структуры"""
import sys
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def analyze_reference_passport(file_path: Path):
    """Анализирует структуру эталонного паспорта"""
    print(f"\n{'='*60}")
    print(f"АНАЛИЗ ЭТАЛОННОГО ПАСПОРТА: {file_path.name}")
    print('='*60)
    
    if not file_path.exists():
        print(f"❌ Файл не найден: {file_path}")
        return
    
    try:
        wb = load_workbook(file_path, data_only=False)
        
        print(f"\n📋 ВСЕ ЛИСТЫ:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
        
        # Ищем лист "Баланс"
        balance_sheet = None
        balance_sheet_name = None
        for sheet_name in ["04_Баланс", "Баланс", "Balans", "04_Balans"]:
            if sheet_name in wb.sheetnames:
                balance_sheet = wb[sheet_name]
                balance_sheet_name = sheet_name
                print(f"\n✅ Найден лист 'Баланс': {sheet_name}")
                break
        
        if not balance_sheet:
            print("\n❌ Лист 'Баланс' не найден в эталонном файле")
            print("   Проверяю все листы на наличие слова 'Баланс'...")
            for sheet_name in wb.sheetnames:
                if "баланс" in sheet_name.lower() or "balans" in sheet_name.lower():
                    print(f"   Найден похожий лист: {sheet_name}")
                    balance_sheet = wb[sheet_name]
                    balance_sheet_name = sheet_name
                    break
        
        if balance_sheet:
            print(f"\n📊 СТРУКТУРА ЛИСТА '{balance_sheet_name}':")
            print(f"   Максимальная строка: {balance_sheet.max_row}")
            print(f"   Максимальная колонка: {balance_sheet.max_column}")
            
            # Анализируем первые 15 строк
            print(f"\n📋 СТРУКТУРА (первые 15 строк):")
            for row_idx in range(1, min(16, balance_sheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(10, balance_sheet.max_column + 1)):
                    cell = balance_sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        if cell.data_type == "f":
                            row_data.append(f"={str(cell.value)[:25]}")
                        else:
                            val = str(cell.value)
                            if len(val) > 30:
                                val = val[:27] + "..."
                            row_data.append(val)
                    else:
                        row_data.append("")
                if any(row_data):  # Показываем только непустые строки
                    print(f"   Строка {row_idx:2d}: {' | '.join(row_data)}")
            
            # Ищем заголовки
            print(f"\n🔍 ПОИСК ЗАГОЛОВКОВ:")
            for row_idx in range(1, min(6, balance_sheet.max_row + 1)):
                for col_idx in range(1, min(10, balance_sheet.max_column + 1)):
                    cell = balance_sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        val = str(cell.value)
                        if any(keyword in val.lower() for keyword in ["квартал", "технологич", "собственн", "производств", "хоз", "итого", "баланс"]):
                            print(f"   Строка {row_idx}, колонка {col_idx}: {val}")
            
            # Ищем строку "ИТОГО"
            print(f"\n🔍 ПОИСК 'ИТОГО':")
            found_totals = False
            for row_idx in range(1, balance_sheet.max_row + 1):
                for col_idx in range(1, min(10, balance_sheet.max_column + 1)):
                    cell = balance_sheet.cell(row=row_idx, column=col_idx)
                    if cell.value and "итого" in str(cell.value).lower():
                        print(f"   ✅ Найдено в строке {row_idx}, колонка {col_idx}: {cell.value}")
                        found_totals = True
                        # Показываем всю строку
                        row_data = []
                        for c in range(1, min(10, balance_sheet.max_column + 1)):
                            c_cell = balance_sheet.cell(row=row_idx, column=c)
                            if c_cell.value:
                                if c_cell.data_type == "f":
                                    row_data.append(f"={str(c_cell.value)[:20]}")
                                else:
                                    row_data.append(str(c_cell.value)[:20])
                        print(f"      Полная строка: {' | '.join(row_data)}")
            if not found_totals:
                print("   ⚠️ Строка 'ИТОГО' не найдена")
            
            # Анализируем формулы
            print(f"\n🔢 ФОРМУЛЫ (первые 10):")
            formula_count = 0
            for row_idx in range(1, min(20, balance_sheet.max_row + 1)):
                for col_idx in range(1, min(10, balance_sheet.max_column + 1)):
                    cell = balance_sheet.cell(row=row_idx, column=col_idx)
                    if cell.data_type == "f":
                        formula_count += 1
                        if formula_count <= 10:
                            formula = str(cell.value)
                            if len(formula) > 50:
                                formula = formula[:47] + "..."
                            print(f"   Строка {row_idx}, колонка {col_idx}: ={formula}")
            print(f"   Всего формул найдено: {formula_count}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

def main():
    file_path = Path("docs/input_templates/энергопаспорт (3) (10) (2).xlsx")
    analyze_reference_passport(file_path)

if __name__ == "__main__":
    main()

