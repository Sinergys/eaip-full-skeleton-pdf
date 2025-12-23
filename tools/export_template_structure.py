"""
Экспорт детальной структуры нового шаблона энергопаспорта.
Создает JSON с информацией о всех листах, ячейках, placeholder'ах и типах данных.
"""

import json
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
import re
import sys


def detect_data_type(cell_value: Any, cell) -> str:
    """Определить тип данных ячейки."""
    if cell_value is None:
        return "empty"

    if cell.data_type == "f":
        return "formula"

    if isinstance(cell_value, (int, float)):
        return "number"

    if isinstance(cell_value, bool):
        return "boolean"

    cell_str = str(cell_value)

    # Проверка на дату
    if re.match(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", cell_str):
        return "date_like"

    # Проверка на placeholder
    if re.search(r"\{\{[^}]+\}\}", cell_str):
        return "placeholder_curly"

    if re.search(r"_+", cell_str) and (
        "год" in cell_str.lower()
        or "quarter" in cell_str.lower()
        or "квартал" in cell_str.lower()
    ):
        return "placeholder_underscore"

    if re.search(r"20_+", cell_str):
        return "placeholder_year"

    # Проверка на заголовок/метку
    if len(cell_str) > 0 and cell_str[0].isupper() and len(cell_str) < 100:
        return "label"

    return "text"


def find_placeholders_in_value(value: str) -> List[str]:
    """Найти все placeholder'ы в значении."""
    placeholders = []

    # {{key}} формат
    matches = re.findall(r"\{\{([^}]+)\}\}", value)
    placeholders.extend(matches)

    # Подчеркивания с контекстом
    if re.search(r"_+", value):
        context = value.lower()
        if "год" in context or "year" in context or re.search(r"20_+", value):
            placeholders.append("year")
        if "квартал" in context or "quarter" in context:
            placeholders.append("quarter")
        if "месяц" in context or "month" in context:
            placeholders.append("month")

    return list(set(placeholders))


def analyze_sheet(ws, max_rows: int = 200, max_cols: int = 20) -> Dict[str, Any]:
    """Анализ структуры листа."""
    sheet_info = {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "cells": [],
        "placeholders": [],
        "formulas": [],
        "data_regions": [],
        "headers": [],
    }

    # Анализ ячеек
    cells_with_data = []
    placeholder_cells = []
    formula_cells = []
    header_cells = []

    # Определение области данных (первые непустые строки и столбцы)
    data_start_row = None
    data_end_row = None
    data_start_col = None
    data_end_col = None

    for row_idx, row in enumerate(
        ws.iter_rows(max_row=max_rows, max_col=max_cols), start=1
    ):
        row_data = []
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell, MergedCell):
                continue

            cell_info = {
                "address": cell.coordinate,
                "row": row_idx,
                "column": col_idx,
                "column_letter": get_column_letter(col_idx),
                "value": None,
                "data_type": cell.data_type,
                "detected_type": None,
                "placeholders": [],
                "is_formula": False,
                "formula": None,
                "number_format": cell.number_format
                if hasattr(cell, "number_format")
                else None,
            }

            if cell.value is not None:
                cell_info["value"] = str(cell.value)[:500]  # Ограничиваем длину
                cell_info["detected_type"] = detect_data_type(cell.value, cell)

                # Поиск placeholder'ов
                placeholders = find_placeholders_in_value(str(cell.value))
                if placeholders:
                    cell_info["placeholders"] = placeholders
                    placeholder_cells.append(cell_info)

                # Формулы
                if cell.data_type == "f":
                    cell_info["is_formula"] = True
                    cell_info["formula"] = str(cell.value)
                    formula_cells.append(cell_info)

                # Заголовки (обычно в первых строках)
                if row_idx <= 5 and cell_info["detected_type"] in ["label", "text"]:
                    header_cells.append(cell_info)

                # Определение области данных
                if data_start_row is None:
                    data_start_row = row_idx
                    data_start_col = col_idx
                data_end_row = row_idx
                data_end_col = max(data_end_col or col_idx, col_idx)

                row_data.append(cell_info)
                cells_with_data.append(cell_info)

        if row_data:
            sheet_info["cells"].append({"row": row_idx, "cells": row_data})

    sheet_info["cells_with_data_count"] = len(cells_with_data)
    sheet_info["placeholder_cells"] = placeholder_cells
    sheet_info["formula_cells"] = formula_cells
    sheet_info["header_cells"] = header_cells

    if data_start_row:
        sheet_info["data_region"] = {
            "start_row": data_start_row,
            "end_row": data_end_row,
            "start_column": data_start_col,
            "end_column": data_end_col,
            "start_address": f"{get_column_letter(data_start_col)}{data_start_row}",
            "end_address": f"{get_column_letter(data_end_col)}{data_end_row}",
        }

    # Уникальные placeholder'ы
    all_placeholders = set()
    for cell in placeholder_cells:
        all_placeholders.update(cell["placeholders"])
    sheet_info["unique_placeholders"] = sorted(list(all_placeholders))

    return sheet_info


def export_template_structure(
    template_path: Path, output_path: Path, max_rows: int = 200
) -> Dict[str, Any]:
    """Экспорт полной структуры шаблона."""
    workbook = load_workbook(template_path, data_only=False)

    structure = {
        "template_path": str(template_path),
        "template_name": template_path.stem,
        "total_sheets": len(workbook.sheetnames),
        "sheets": {},
        "summary": {
            "total_placeholders": 0,
            "total_formulas": 0,
            "total_cells_with_data": 0,
            "unique_placeholders_all": set(),
        },
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"Анализ листа: {sheet_name}...")
        sheet_info = analyze_sheet(ws, max_rows=max_rows)
        structure["sheets"][sheet_name] = sheet_info

        # Обновление сводки
        structure["summary"]["total_placeholders"] += len(
            sheet_info["placeholder_cells"]
        )
        structure["summary"]["total_formulas"] += len(sheet_info["formula_cells"])
        structure["summary"]["total_cells_with_data"] += sheet_info[
            "cells_with_data_count"
        ]
        structure["summary"]["unique_placeholders_all"].update(
            sheet_info["unique_placeholders"]
        )

    # Конвертируем set в list для JSON
    structure["summary"]["unique_placeholders_all"] = sorted(
        list(structure["summary"]["unique_placeholders_all"])
    )

    # Сохранение
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(structure, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    return structure


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Экспорт структуры Excel шаблона")
    parser.add_argument("--template", help="Путь к шаблону")
    parser.add_argument("--template-name", help="Имя шаблона из templates_config")
    parser.add_argument("--output", required=True, help="Путь для сохранения JSON")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=200,
        help="Максимальное количество строк для анализа",
    )

    args = parser.parse_args()

    # Определение пути к шаблону
    if args.template_name:
        templates_config_path = Path(__file__).parent.parent / "templates" / "pcm690"
        if str(templates_config_path) not in sys.path:
            sys.path.insert(0, str(templates_config_path))
        from templates_config import get_template_path

        template_path = get_template_path(args.template_name)
    elif args.template:
        template_path = Path(args.template)
    else:
        raise ValueError("Необходимо указать либо --template-name, либо --template")

    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    print(f"Экспорт структуры шаблона: {template_path}")
    print("=" * 80)

    structure = export_template_structure(
        template_path, Path(args.output), max_rows=args.max_rows
    )

    print(f"\n✅ Структура экспортирована в: {args.output}")
    print("\n📊 Сводка:")
    print(f"  Листов: {structure['total_sheets']}")
    print(f"  Ячеек с данными: {structure['summary']['total_cells_with_data']}")
    print(f"  Placeholder'ов: {structure['summary']['total_placeholders']}")
    print(f"  Формул: {structure['summary']['total_formulas']}")
    print(
        f"  Уникальных placeholder'ов: {len(structure['summary']['unique_placeholders_all'])}"
    )

    print("\n📄 Листы:")
    for sheet_name, sheet_info in structure["sheets"].items():
        print(f"  {sheet_name}:")
        print(f"    Ячеек: {sheet_info['cells_with_data_count']}")
        print(f"    Placeholder'ов: {len(sheet_info['placeholder_cells'])}")
        print(f"    Формул: {len(sheet_info['formula_cells'])}")
        if sheet_info["unique_placeholders"]:
            print(f"    Placeholder'ы: {', '.join(sheet_info['unique_placeholders'])}")


if __name__ == "__main__":
    main()
