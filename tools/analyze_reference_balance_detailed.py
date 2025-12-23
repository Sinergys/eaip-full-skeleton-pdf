"""Детальный анализ структуры листа 'Баланс' в эталонном файле"""
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def analyze_balance_detailed(file_path: Path):
    """Детальный анализ структуры листа 'Баланс'"""
    print(f"\n{'='*70}")
    print(f"ДЕТАЛЬНЫЙ АНАЛИЗ ЛИСТА 'БАЛАНС'")
    print(f"Файл: {file_path.name}")
    print('='*70)
    
    wb = load_workbook(file_path, data_only=False)
    balance_sheet = wb["Баланс"]
    
    print(f"\n📊 РАЗМЕРЫ ЛИСТА:")
    print(f"   Строк: {balance_sheet.max_row}")
    print(f"   Колонок: {balance_sheet.max_column}")
    
    # Анализируем все строки с данными
    print(f"\n📋 ПОЛНАЯ СТРУКТУРА ЛИСТА:")
    print(f"{'Строка':<6} | {'A':<30} | {'B':<20} | {'C':<20} | {'D':<20} | {'E':<20} | {'F':<20}")
    print("-" * 140)
    
    for row_idx in range(1, min(35, balance_sheet.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(7, balance_sheet.max_column + 1)):
            cell = balance_sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                if cell.data_type == "f":
                    val = f"={str(cell.value)[:15]}"
                else:
                    val = str(cell.value)
                    if len(val) > 18:
                        val = val[:15] + "..."
                row_values.append(val)
            else:
                row_values.append("")
        
        if any(row_values):  # Показываем только непустые строки
            print(f"{row_idx:<6} | {row_values[0]:<30} | {row_values[1]:<20} | {row_values[2]:<20} | {row_values[3]:<20} | {row_values[4]:<20} | {row_values[5]:<20}")
    
    # Анализируем структуру категорий потребления
    print(f"\n🔍 КАТЕГОРИИ ПОТРЕБЛЕНИЯ:")
    categories = []
    for row_idx in range(1, balance_sheet.max_row + 1):
        cell_a = balance_sheet.cell(row=row_idx, column=1)
        if cell_a.value:
            val = str(cell_a.value).lower()
            if any(keyword in val for keyword in ["технологич", "собственн", "производств", "хоз", "хозяйствен"]):
                categories.append((row_idx, str(cell_a.value)))
                print(f"   Строка {row_idx}: {cell_a.value}")
    
    # Анализируем колонки с ресурсами
    print(f"\n🔍 КОЛОНКИ С РЕСУРСАМИ:")
    header_row = None
    for row_idx in range(1, 10):
        cell = balance_sheet.cell(row=row_idx, column=2)
        if cell.value and ("электр" in str(cell.value).lower() or "энергия" in str(cell.value).lower()):
            header_row = row_idx
            print(f"   Заголовок ресурсов в строке {row_idx}")
            break
    
    if header_row:
        print(f"\n   Колонки в строке {header_row}:")
        for col_idx in range(1, min(17, balance_sheet.max_column + 1)):
            cell = balance_sheet.cell(row=header_row, column=col_idx)
            if cell.value:
                print(f"      Колонка {col_idx}: {cell.value}")
    
    wb.close()

if __name__ == "__main__":
    file_path = Path("docs/input_templates/энергопаспорт (3) (10) (2).xlsx")
    analyze_balance_detailed(file_path)

