"""
Заполнение существующего паспорта energopasport.xlsx данными из листа AggregatedData
"""

import sys
from pathlib import Path
import json
import logging
from openpyxl import load_workbook
from typing import Dict, Any, Optional

# Добавляем пути для импорта
sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(
    0, str(Path(__file__).parent.parent / "eaip_full_skeleton" / "services" / "ingest")
)

from tools.fill_energy_passport import (
    fill_struktura_pr2,
    fill_nodes_sheet,
    fill_equipment_sheet,
    fill_balans_sheet,
    fill_dinamika_sheet,
    fill_meropriyatiya_sheet,
    load_default_nodes,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

PASSPORT_PATH = (
    Path(__file__).parent.parent
    / "data"
    / "source_files"
    / "audit_sinergys"
    / "energopasport.xlsx"
)


def read_aggregated_data_from_sheet(workbook) -> Optional[Dict[str, Any]]:
    """Читает агрегированные данные из листа AggregatedData и преобразует в формат aggregated"""
    if "AggregatedData" not in workbook.sheetnames:
        logger.error("Лист 'AggregatedData' не найден")
        return None

    ws = workbook["AggregatedData"]
    logger.info(f"Чтение данных из листа 'AggregatedData' ({ws.max_row} строк)")

    # Читаем заголовки из первой строки
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        headers.append(str(cell.value) if cell.value else f"Col{col_idx}")

    logger.info(f"Заголовки: {headers}")

    # Ищем индексы колонок
    col_indices = {}
    for idx, header in enumerate(headers, 1):
        header_lower = str(header).lower()
        if "year" in header_lower or "год" in header_lower:
            col_indices["year"] = idx
        elif "quarter" in header_lower or "квартал" in header_lower:
            col_indices["quarter"] = idx
        elif "resource" in header_lower or "ресурс" in header_lower:
            col_indices["resource"] = idx
        elif "field" in header_lower or "поле" in header_lower:
            col_indices["field"] = idx
        elif "value" in header_lower or "значение" in header_lower:
            col_indices["value"] = idx

    if not all(
        k in col_indices for k in ["year", "quarter", "resource", "field", "value"]
    ):
        logger.error(f"Не найдены все необходимые колонки. Найдено: {col_indices}")
        return None

    logger.info(f"Индексы колонок: {col_indices}")

    # Читаем данные и преобразуем в формат aggregated
    aggregated = {"resources": {}}

    for row_idx in range(2, ws.max_row + 1):
        year = ws.cell(row=row_idx, column=col_indices["year"]).value
        quarter = ws.cell(row=row_idx, column=col_indices["quarter"]).value
        resource = ws.cell(row=row_idx, column=col_indices["resource"]).value
        field = ws.cell(row=row_idx, column=col_indices["field"]).value
        value = ws.cell(row=row_idx, column=col_indices["value"]).value

        if not all([year, quarter, resource, field, value is not None]):
            continue

        # Нормализуем значения
        year = int(year) if isinstance(year, (int, float)) else int(str(year).strip())
        quarter_str = str(quarter).strip().upper()
        if "Q" not in quarter_str:
            # Если квартал как число (1, 2, 3, 4)
            try:
                q_num = int(quarter_str)
                quarter_str = f"{year}-Q{q_num}"
            except:
                quarter_str = f"{year}-{quarter_str}"
        else:
            quarter_str = f"{year}-{quarter_str}"

        resource_str = str(resource).strip().lower()
        field_str = str(field).strip().lower()

        # Нормализуем название ресурса
        resource_map = {
            "electricity": "electricity",
            "электроэнергия": "electricity",
            "электро": "electricity",
            "gas": "gas",
            "газ": "gas",
            "water": "water",
            "вода": "water",
            "heat": "heat",
            "тепло": "heat",
            "отопление": "heat",
        }
        resource_str = resource_map.get(resource_str, resource_str)

        # Преобразуем значение
        try:
            if isinstance(value, (int, float)):
                num_value = float(value)
            else:
                num_value = float(str(value).replace(",", ".").replace(" ", ""))
        except:
            continue

        # Создаем структуру
        if resource_str not in aggregated["resources"]:
            aggregated["resources"][resource_str] = {}

        if quarter_str not in aggregated["resources"][resource_str]:
            aggregated["resources"][resource_str][quarter_str] = {"quarter_totals": {}}

        # Нормализуем название поля
        field_map = {
            "active_kwh": "active_kwh",
            "active": "active_kwh",
            "активная": "active_kwh",
            "reactive_kvarh": "reactive_kvarh",
            "reactive": "reactive_kvarh",
            "реактивная": "reactive_kvarh",
            "volume_m3": "volume_m3",
            "volume": "volume_m3",
            "объем": "volume_m3",
            "energy_gcal": "energy_gcal",
            "energy": "energy_gcal",
            "энергия": "energy_gcal",
        }
        field_str = field_map.get(field_str, field_str)

        aggregated["resources"][resource_str][quarter_str]["quarter_totals"][
            field_str
        ] = num_value

    logger.info(f"Преобразовано данных:")
    for resource, quarters in aggregated["resources"].items():
        logger.info(f"  {resource}: {len(quarters)} кварталов")
        for quarter in list(quarters.keys())[:3]:
            totals = quarters[quarter].get("quarter_totals", {})
            logger.info(f"    {quarter}: {list(totals.keys())}")

    return aggregated


def fill_passport_from_aggregated_data(
    passport_path: Path, output_path: Optional[Path] = None
):
    """Заполняет паспорт данными из листа AggregatedData"""

    if not passport_path.exists():
        logger.error(f"Файл паспорта не найден: {passport_path}")
        return False

    logger.info(f"Открытие паспорта: {passport_path}")

    # Копируем файл для работы
    import shutil

    if output_path is None:
        output_path = passport_path.parent / f"{passport_path.stem}_filled.xlsx"

    shutil.copyfile(passport_path, output_path)
    logger.info(f"Создана копия: {output_path}")

    # Загружаем workbook
    workbook = load_workbook(output_path, data_only=False)

    # Читаем агрегированные данные
    aggregated = read_aggregated_data_from_sheet(workbook)

    if not aggregated:
        logger.error("Не удалось прочитать агрегированные данные")
        # Пытаемся найти JSON файл рядом с паспортом
        json_files = list(passport_path.parent.glob("*aggregated*.json"))
        if json_files:
            json_path = json_files[0]
            logger.info(f"Найден JSON файл: {json_path}")
            aggregated = json.loads(json_path.read_text(encoding="utf-8"))
        else:
            logger.error("Не найден JSON файл с агрегированными данными")
            workbook.close()
            return False

    # Нормализуем структуру данных
    if "resources" not in aggregated:
        logger.warning(
            "Структура данных не содержит 'resources', пытаемся нормализовать"
        )
        aggregated = {"resources": aggregated}

    logger.info(f"Структура данных: {list(aggregated.keys())}")
    if "resources" in aggregated:
        logger.info(f"Ресурсы: {list(aggregated['resources'].keys())}")

    # Преобразуем структуру для fill_struktura_pr2 (она ожидает resources на верхнем уровне)
    agg_data_for_fill = aggregated.get("resources", aggregated)

    # Заполняем листы
    filled_sheets = []

    # 1. Struktura pr2
    if "Struktura pr2" in workbook.sheetnames:
        logger.info("Заполнение листа 'Struktura pr2'")
        try:
            fill_struktura_pr2(
                workbook["Struktura pr2"],
                agg_data_for_fill,
                loss_active_month=0.0,
                loss_reactive_month=0.0,
            )
            filled_sheets.append("Struktura pr2")
            logger.info("✅ Лист 'Struktura pr2' заполнен")
        except Exception as e:
            logger.exception(f"Ошибка заполнения 'Struktura pr2': {e}")

    # 2. Uzel ucheta
    if "Uzel ucheta" in workbook.sheetnames:
        logger.info("Заполнение листа 'Uzel ucheta'")
        try:
            nodes = load_default_nodes()
            # Пытаемся загрузить узлы учета из файла
            nodes_file = passport_path.parent / "schetchiki.xlsx"
            if nodes_file.exists():
                logger.info(f"Парсинг узлов учета из {nodes_file}")
                try:
                    from eaip_full_skeleton.services.ingest.utils.nodes_parser import (
                        parse_nodes_workbook,
                    )

                    parsed_nodes = parse_nodes_workbook(str(nodes_file))
                    if parsed_nodes and parsed_nodes.get("nodes"):
                        nodes = parsed_nodes["nodes"]
                        logger.info(f"Загружено {len(nodes)} узлов учета")
                except Exception as parse_exc:
                    logger.warning(f"Не удалось распарсить узлы учета: {parse_exc}")

            fill_nodes_sheet(workbook["Uzel ucheta"], nodes)
            filled_sheets.append("Uzel ucheta")
            logger.info("✅ Лист 'Uzel ucheta' заполнен")
        except Exception as e:
            logger.exception(f"Ошибка заполнения 'Uzel ucheta': {e}")

    # 3. Sheet1 (Equipment)
    if "Sheet1" in workbook.sheetnames:
        logger.info("Заполнение листа 'Sheet1' (Equipment)")
        try:
            # Ищем данные оборудования
            equipment_data = {"sheets": [], "summary": {}}
            # Пытаемся найти equipment JSON
            equipment_json = list(passport_path.parent.glob("*equipment*.json"))
            if equipment_json:
                equipment_data = json.loads(
                    equipment_json[0].read_text(encoding="utf-8")
                )
                logger.info(f"Загружены данные оборудования из {equipment_json[0]}")
            else:
                # Пытаемся парсить oborudovanie.xlsx напрямую
                equipment_file = passport_path.parent / "oborudovanie.xlsx"
                if equipment_file.exists():
                    logger.info(f"Парсинг оборудования из {equipment_file}")
                    try:
                        from eaip_full_skeleton.services.ingest.utils.equipment_parser import (
                            parse_equipment_workbook,
                        )

                        equipment_data = parse_equipment_workbook(str(equipment_file))
                        logger.info(
                            f"Загружено {len(equipment_data.get('sheets', []))} листов оборудования"
                        )
                    except Exception as parse_exc:
                        logger.warning(
                            f"Не удалось распарсить оборудование: {parse_exc}"
                        )

            fill_equipment_sheet(workbook, equipment_data, sheet_name="Sheet1")
            filled_sheets.append("Sheet1")
            logger.info("✅ Лист 'Sheet1' заполнен")
        except Exception as e:
            logger.exception(f"Ошибка заполнения 'Sheet1': {e}")

    # 4. Balans
    if "Balans" in workbook.sheetnames:
        logger.info("Заполнение листа 'Balans'")
        try:
            # Добавляем by_usage распределение для баланса
            from eaip_full_skeleton.services.ingest.utils.readiness_validator import (
                _create_standard_by_usage_distribution,
            )

            for resource_type, quarters in agg_data_for_fill.items():
                for quarter, quarter_data in quarters.items():
                    if "by_usage" not in quarter_data:
                        quarter_totals = quarter_data.get("quarter_totals", {})
                        total = (
                            quarter_totals.get("active_kwh")
                            or quarter_totals.get("volume_m3")
                            or 0.0
                        )
                        if total > 0:
                            quarter_data["by_usage"] = (
                                _create_standard_by_usage_distribution(total)
                            )

            # fill_balans_sheet ожидает структуру с "resources" на верхнем уровне
            fill_balans_sheet(workbook["Balans"], aggregated)
            filled_sheets.append("Balans")
            logger.info("✅ Лист 'Balans' заполнен")
        except Exception as e:
            logger.exception(f"Ошибка заполнения 'Balans': {e}")

    # 5. Meropriyatiya
    if "Meropriyatiya" in workbook.sheetnames:
        logger.info("Заполнение листа 'Meropriyatiya'")
        try:
            # Ищем данные мероприятий
            measures_data = None
            measures_json = list(passport_path.parent.glob("*measures*.json"))
            if measures_json:
                measures_data = json.loads(measures_json[0].read_text(encoding="utf-8"))
                logger.info(f"Загружены данные мероприятий из {measures_json[0]}")

            fill_meropriyatiya_sheet(workbook["Meropriyatiya"], measures_data)
            filled_sheets.append("Meropriyatiya")
            logger.info("✅ Лист 'Meropriyatiya' заполнен")
        except Exception as e:
            logger.exception(f"Ошибка заполнения 'Meropriyatiya': {e}")

    # Сохраняем
    workbook.save(output_path)
    workbook.close()

    logger.info(f"\n{'=' * 80}")
    logger.info(f"✅ ПАСПОРТ ЗАПОЛНЕН: {output_path}")
    logger.info(f"Заполнено листов: {len(filled_sheets)}/{len(workbook.sheetnames)}")
    logger.info(f"Листы: {', '.join(filled_sheets)}")
    logger.info(f"{'=' * 80}")

    return True


if __name__ == "__main__":
    fill_passport_from_aggregated_data(PASSPORT_PATH)
