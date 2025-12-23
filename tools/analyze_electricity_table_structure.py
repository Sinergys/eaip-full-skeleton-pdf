"""Детальный анализ структуры таблиц электроэнергии по кварталам"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

template_path = r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"

wb = load_workbook(template_path, data_only=False)
ws = wb["Структура пр 2"]

print("=" * 80)
print("АНАЛИЗ СТРУКТУРЫ ТАБЛИЦ ЭЛЕКТРОЭНЕРГИИ ПО КВАРТАЛАМ")
print("=" * 80)

# Анализируем каждую таблицу
tables = [
    (17, "Q1 2022"),
    (39, "Q2 2022"),
    (61, "Q3 2022"),
    (83, "Q4 2022"),
]

for start_row, quarter_name in tables:
    print(f"\n{'='*80}")
    print(f"ТАБЛИЦА: {quarter_name} (строки {start_row}-{start_row+4})")
    print(f"{'='*80}")
    
    # Ищем заголовок таблицы
    print("\nЗаголовки:")
    for header_row in range(start_row-3, start_row):
        row_data = []
        for col in range(1, 30):
            val = ws.cell(header_row, col).value
            if val is not None and isinstance(val, (str, int, float)):
                col_letter = get_column_letter(col)
                row_data.append(f"{col_letter}{header_row}={val}")
        if row_data:
            print(f"  Строка {header_row}: {', '.join(row_data[:10])}")
    
    # Анализируем данные по продукции
    print(f"\nДанные по продукции (строка {start_row} - труба ХВС):")
    for col in range(1, 30):
        val = ws.cell(start_row, col).value
        if val is not None:
            col_letter = get_column_letter(col)
            cell = ws.cell(start_row, col)
            data_type = cell.data_type
            print(f"  {col_letter}{start_row}: {val} (тип: {data_type})")
    
    # Ищем паттерн колонок: норма, факт 2022, факт 2023, факт 2024, перерасходы
    print(f"\nОпределение структуры колонок:")
    # Проверяем строку 17 (где есть данные)
    if start_row == 17:
        print("  Найдены данные в колонках:")
        for col in range(2, 9):
            val = ws.cell(start_row, col).value
            if val is not None:
                col_letter = get_column_letter(col)
                # Определяем назначение колонки
                if col == 2:
                    print(f"    {col_letter} = Норма")
                elif col == 3:
                    print(f"    {col_letter} = Факт 2022")
                elif col == 4:
                    print(f"    {col_letter} = Факт 2023")
                elif col == 5:
                    print(f"    {col_letter} = Факт 2024")
                elif col == 6:
                    print(f"    {col_letter} = Перерасход 2022")
                elif col == 7:
                    print(f"    {col_letter} = Перерасход 2023")
                elif col == 8:
                    print(f"    {col_letter} = Перерасход 2024")

# Ищем повторяющийся паттерн колонок для других кварталов
print("\n" + "=" * 80)
print("ПОИСК ПАТТЕРНА КОЛОНОК ДЛЯ ДРУГИХ КВАРТАЛОВ")
print("=" * 80)

# Проверяем, есть ли повторяющийся паттерн каждые ~16 колонок
print("\nПроверка колонок для Q2 (строка 39):")
for col in range(1, 50):
    val = ws.cell(39, col).value
    if val is not None:
        col_letter = get_column_letter(col)
        print(f"  {col_letter}39: {val}")

# Ищем заголовки с годами или кварталами
print("\nПоиск заголовков с годами/кварталами вокруг строки 39:")
for row in range(36, 40):
    for col in range(1, 50):
        val = ws.cell(row, col).value
        if val and (isinstance(val, int) and val in (2022, 2023, 2024)) or (isinstance(val, str) and ("Q" in str(val) or "квартал" in str(val).lower())):
            col_letter = get_column_letter(col)
            print(f"  {col_letter}{row}: {val}")

wb.close()

