"""Полный анализ структуры эталонного файла для адаптации fill_balans_sheet"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def analyze_complete(file_path: Path):
    """Полный анализ структуры листа 'Баланс'"""
    print(f"\n{'='*80}")
    print(f"ПОЛНЫЙ АНАЛИЗ ЭТАЛОННОГО ФАЙЛА")
    print(f"Файл: {file_path.name}")
    print('='*80)
    
    wb = load_workbook(file_path, data_only=False)
    ws = wb["Баланс"]
    
    print(f"\n📊 РАЗМЕРЫ: {ws.max_row} строк × {ws.max_column} колонок\n")
    
    # Анализ структуры заголовков (строки 1-7)
    print("=" * 80)
    print("СТРУКТУРА ЗАГОЛОВКОВ (строки 1-7):")
    print("=" * 80)
    for row_idx in range(1, 8):
        row_data = []
        for col_idx in range(1, min(17, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                val = str(cell.value)
                if len(val) > 25:
                    val = val[:22] + "..."
                row_data.append(f"{get_column_letter(col_idx)}{row_idx}:{val}")
        if row_data:
            print(f"Строка {row_idx:2d}: {' | '.join(row_data)}")
    
    # Анализ категорий потребления (строки 8-15)
    print("\n" + "=" * 80)
    print("КАТЕГОРИИ ПОТРЕБЛЕНИЯ (строки 8-15):")
    print("=" * 80)
    categories_map = {}
    for row_idx in range(8, 16):
        cell_a = ws.cell(row=row_idx, column=1)
        if cell_a.value:
            label = str(cell_a.value).strip()
            categories_map[row_idx] = label
            # Показываем значения в колонках для этой категории
            row_values = []
            for col_idx in range(2, min(17, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    if cell.data_type == "f":
                        row_values.append(f"{get_column_letter(col_idx)}:={str(cell.value)[:20]}")
                    else:
                        val = str(cell.value)
                        if len(val) > 15:
                            val = val[:12] + "..."
                        row_values.append(f"{get_column_letter(col_idx)}:{val}")
            values_str = " | ".join(row_values[:5]) if row_values else "нет данных"
            print(f"Строка {row_idx:2d}: {label[:50]:<50} | Значения: {values_str}")
    
    # Определяем колонки ресурсов
    print("\n" + "=" * 80)
    print("КОЛОНКИ РЕСУРСОВ:")
    print("=" * 80)
    resource_columns = {}
    # Ищем в строках 4-7 заголовки ресурсов
    for row_idx in range(4, 8):
        for col_idx in range(1, min(17, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                val = str(cell.value).lower()
                col_letter = get_column_letter(col_idx)
                if "электр" in val or "энергия" in val:
                    if "актив" in val or "р" in val:
                        resource_columns[col_letter] = {"type": "electricity_active", "row": row_idx}
                    elif "реактив" in val or "q" in val:
                        resource_columns[col_letter] = {"type": "electricity_reactive", "row": row_idx}
                    else:
                        resource_columns[col_letter] = {"type": "electricity", "row": row_idx}
                elif "тепл" in val or "гкал" in val:
                    resource_columns[col_letter] = {"type": "heat", "row": row_idx}
                elif "газ" in val:
                    resource_columns[col_letter] = {"type": "gas", "row": row_idx}
                elif "мазут" in val:
                    resource_columns[col_letter] = {"type": "fuel", "row": row_idx}
                elif "уголь" in val:
                    resource_columns[col_letter] = {"type": "coal", "row": row_idx}
    
    for col_letter, info in sorted(resource_columns.items()):
        print(f"Колонка {col_letter:3s}: {info['type']:25s} (заголовок в строке {info['row']})")
    
    # Показываем пример заполнения для категорий
    print("\n" + "=" * 80)
    print("ПРИМЕР ЗАПОЛНЕНИЯ ДАННЫХ (строки 10-14):")
    print("=" * 80)
    print(f"{'Категория':<40} | ", end="")
    for col_letter in sorted(resource_columns.keys())[:5]:
        print(f"{col_letter:>10} | ", end="")
    print()
    print("-" * 80)
    
    for row_idx in [10, 11, 12, 13, 14]:
        cell_a = ws.cell(row=row_idx, column=1)
        if cell_a.value:
            label = str(cell_a.value).strip()
            if len(label) > 38:
                label = label[:35] + "..."
            print(f"{label:<40} | ", end="")
            for col_letter in sorted(resource_columns.keys())[:5]:
                col_idx = ord(col_letter) - ord('A') + 1
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    if cell.data_type == "f":
                        val = f"={str(cell.value)[:8]}"
                    else:
                        val = str(cell.value)
                        if len(val) > 8:
                            val = val[:5] + "..."
                    print(f"{val:>10} | ", end="")
                else:
                    print(f"{'':>10} | ", end="")
            print()
    
    wb.close()
    
    # Вывод рекомендаций
    print("\n" + "=" * 80)
    print("РЕКОМЕНДАЦИИ ПО АДАПТАЦИИ:")
    print("=" * 80)
    print("""
1. СТРУКТУРА ЛИСТА:
   - Строка 1: "Энергетический баланс предприятия"
   - Строка 4: "Структура параметров баланса"
   - Строки 5-7: Заголовки ресурсов (электроэнергия актив/реактив, тепло, газ, мазут, уголь)
   - Строка 10: "Общее потребление по предприятию"
   - Строка 11: "– на технологические нужды"
   - Строка 12: "– на собственные нужды предприятия"
   - Строка 13: "– на производственные нужды"
   - Строка 14: "– на хозяйственно-бытовые нужды"

2. ЗАПОЛНЕНИЕ ДАННЫХ:
   - Категории потребления в строках (11-14)
   - Ресурсы в колонках (определены в строках 4-7)
   - Значения заполняются на пересечении категории и ресурса

3. ИЗМЕНЕНИЯ В fill_balans_sheet():
   - НЕ создавать структуру по кварталам
   - Заполнять строки 11-14 для каждой категории
   - Заполнять колонки ресурсов (электроэнергия активная/реактивная, тепло, газ и т.д.)
   - Использовать данные из agg_data по ресурсам и категориям потребления
    """)

if __name__ == "__main__":
    file_path = Path("docs/input_templates/энергопаспорт (3) (10) (2).xlsx")
    analyze_complete(file_path)

