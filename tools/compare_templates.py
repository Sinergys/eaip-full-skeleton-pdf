"""
Скрипт для сравнения шаблонов энергопаспорта
Сравнивает new_energy_passport.xlsx с эталонным energy_passport_template.xlsx
"""

from pathlib import Path
from openpyxl import load_workbook

# Пути к шаблонам
TEMPLATES_DIR = Path(__file__).parent.parent / "templates" / "pcm690"
REFERENCE_TEMPLATE = TEMPLATES_DIR / "energy_passport_template.xlsx"
NEW_TEMPLATE = TEMPLATES_DIR / "new_energy_passport.xlsx"


def compare_templates():
    """Сравнивает два шаблона и выводит различия"""

    print("=" * 80)
    print("СРАВНЕНИЕ ШАБЛОНОВ ЭНЕРГОПАСПОРТА")
    print("=" * 80)

    # Проверка существования файлов
    if not REFERENCE_TEMPLATE.exists():
        print(f"❌ Эталонный шаблон не найден: {REFERENCE_TEMPLATE}")
        return

    if not NEW_TEMPLATE.exists():
        print(f"❌ Новый шаблон не найден: {NEW_TEMPLATE}")
        return

    print(f"\n📄 Эталонный шаблон: {REFERENCE_TEMPLATE.name}")
    print(f"📄 Новый шаблон: {NEW_TEMPLATE.name}\n")

    # Загружаем оба шаблона
    try:
        ref_wb = load_workbook(REFERENCE_TEMPLATE, data_only=False)
        new_wb = load_workbook(NEW_TEMPLATE, data_only=False)
    except Exception as e:
        print(f"❌ Ошибка при загрузке шаблонов: {e}")
        return

    # Сравнение листов
    ref_sheets = set(ref_wb.sheetnames)
    new_sheets = set(new_wb.sheetnames)

    print("=" * 80)
    print("СРАВНЕНИЕ ЛИСТОВ")
    print("=" * 80)

    print(f"\n📊 Эталонный шаблон: {len(ref_sheets)} листов")
    for sheet in sorted(ref_sheets):
        ws = ref_wb[sheet]
        formulas_count = sum(
            1
            for row in ws.iter_rows()
            for cell in row
            if cell.data_type == "f" and cell.value
        )
        print(
            f"  - {sheet} ({ws.max_row} строк, {ws.max_column} колонок, {formulas_count} формул)"
        )

    print(f"\n📊 Новый шаблон: {len(new_sheets)} листов")
    for sheet in sorted(new_sheets):
        ws = new_wb[sheet]
        formulas_count = sum(
            1
            for row in ws.iter_rows()
            for cell in row
            if cell.data_type == "f" and cell.value
        )
        print(
            f"  - {sheet} ({ws.max_row} строк, {ws.max_column} колонок, {formulas_count} формул)"
        )

    # Ключевые листы, которые должны быть
    required_sheets = [
        "02_Структура",
        "Struktura pr2",
        "01_Узлы учета",
        "Узлы учета",
        "Nodes",
        "03_Оборудование",
        "Equipment",
        "Оборудование",
        "04_Баланс",
        "Баланс",
        "Balans",
        "05_Динамика",
        "Динамика",
        "Dinamika sr",
        "06_Мероприятия",
        "Мероприятия",
        "Meropriyatiya",
    ]

    print("\n" + "=" * 80)
    print("ПРОВЕРКА КЛЮЧЕВЫХ ЛИСТОВ")
    print("=" * 80)

    ref_has_required = {}
    new_has_required = {}

    for req_name in required_sheets:
        ref_has_required[req_name] = req_name in ref_sheets
        new_has_required[req_name] = req_name in new_sheets

    print("\nЭталонный шаблон:")
    for req_name in required_sheets:
        status = "✅" if ref_has_required[req_name] else "❌"
        print(f"  {status} {req_name}")

    print("\nНовый шаблон:")
    for req_name in required_sheets:
        status = "✅" if new_has_required[req_name] else "❌"
        print(f"  {status} {req_name}")

    # Листы, которые есть в эталоне, но отсутствуют в новом
    missing_in_new = ref_sheets - new_sheets
    extra_in_new = new_sheets - ref_sheets

    print("\n" + "=" * 80)
    print("РАЗЛИЧИЯ")
    print("=" * 80)

    if missing_in_new:
        print(
            f"\n❌ Листы, которые есть в эталоне, но отсутствуют в новом ({len(missing_in_new)}):"
        )
        for sheet in sorted(missing_in_new):
            print(f"  - {sheet}")
    else:
        print("\n✅ Все листы эталона присутствуют в новом шаблоне")

    if extra_in_new:
        print(f"\n➕ Листы, которые есть только в новом шаблоне ({len(extra_in_new)}):")
        for sheet in sorted(extra_in_new):
            print(f"  - {sheet}")

    # Итоговый вывод
    print("\n" + "=" * 80)
    print("ВЫВОД")
    print("=" * 80)

    if len(missing_in_new) > 0:
        print("\n⚠️  ВНИМАНИЕ: Новый шаблон НЕПОЛНЫЙ!")
        print("   Отсутствуют важные листы из эталонного шаблона.")
        print("   Рекомендация: использовать эталонный шаблон или дополнить новый.")
    elif len(new_sheets) < len(ref_sheets):
        print("\n⚠️  Новый шаблон содержит меньше листов, чем эталонный.")
    else:
        print("\n✅ Новый шаблон содержит все необходимые листы.")

    ref_wb.close()
    new_wb.close()


if __name__ == "__main__":
    compare_templates()
