"""
Извлечение данных по узлам из raw_json без файла на диске.
"""
import sys
import json
from pathlib import Path
from typing import Dict, Any, List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent / "eaip_full_skeleton" / "services" / "ingest"))

from utils.balance_sheet_node_extractor import _parse_node_sheet, _extract_period_from_text
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def create_sheet_from_data(sheet_data: Dict[str, Any]) -> Worksheet:
    """Создает объект Worksheet из данных листа."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_data.get('name', 'Sheet1')
    
    # Добавляем строки
    rows = sheet_data.get('rows', [])
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    return ws


def extract_nodes_from_raw_json(
    raw_json: Dict[str, Any],
    batch_id: str,
    enterprise_id: int,
    filename: str = ""
) -> List[Dict[str, Any]]:
    """
    Извлекает данные по узлам из raw_json без файла на диске.
    
    Args:
        raw_json: Распарсенные данные файла
        batch_id: ID загрузки
        enterprise_id: ID предприятия
        filename: Имя файла (для определения типа данных)
    
    Returns:
        Список данных по узлам
    """
    # Определяем тип данных
    filename_lower = filename.lower() if filename else ""
    data_type = "consumption"
    if "реализация" in filename_lower or "реализация нэс" in filename_lower:
        data_type = "realization"
    elif "производство" in filename_lower or "production" in filename_lower:
        data_type = "production"
    
    # Извлекаем листы из raw_json
    sheets_data = None
    if 'parsing' in raw_json and 'data' in raw_json['parsing']:
        if 'sheets' in raw_json['parsing']['data']:
            sheets_data = raw_json['parsing']['data']['sheets']
    elif 'data' in raw_json and 'sheets' in raw_json['data']:
        sheets_data = raw_json['data']['sheets']
    
    if not sheets_data:
        print(f"   ❌ Листы не найдены в raw_json")
        return []
    
    print(f"   ✅ Найдено {len(sheets_data)} листов")
    
    # Обрабатываем каждый лист
    all_node_data = []
    
    for sheet_data in sheets_data:
        sheet_name = sheet_data.get('name', 'Без имени')
        rows = sheet_data.get('rows', [])
        
        if not rows:
            continue
        
        print(f"   📊 Обработка листа: {sheet_name} ({len(rows)} строк)")
        
        # Создаем объект Worksheet из данных
        try:
            ws = create_sheet_from_data(sheet_data)
            
            # Парсим лист
            # Сначала извлекаем период с учетом filename
            period = _extract_period_from_text(sheet_name, filename=filename) if filename else None
            
            # Парсим лист (пока без filename, так как период уже извлечен)
            node_data = _parse_node_sheet(ws, sheet_name, data_type)
            
            # Обновляем период в записях, если он был извлечен
            if period and node_data:
                for record in node_data:
                    if record.get('period') == 'unknown' or not record.get('period'):
                        record['period'] = period
            
            if node_data:
                print(f"      ✅ Извлечено {len(node_data)} записей")
                all_node_data.extend(node_data)
            else:
                print(f"      ⚠️ Данные не извлечены")
        except Exception as e:
            print(f"      ❌ Ошибка обработки листа: {e}")
            import traceback
            traceback.print_exc()
    
    return all_node_data


if __name__ == "__main__":
    # Тест
    import sqlite3
    from pathlib import Path
    
    DB_PATH = Path(__file__).parent.parent / "eaip_full_skeleton" / "services" / "ingest" / "ingest_data.db"
    
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    
    cursor = conn.execute(
        """
        SELECT u.batch_id, u.filename, pd.raw_json
        FROM uploads u
        JOIN parsed_data pd ON u.id = pd.upload_id
        WHERE u.filename LIKE '%Реализация 2022%'
        ORDER BY u.created_at DESC
        LIMIT 1
        """
    )
    
    row = cursor.fetchone()
    conn.close()
    
    if row:
        raw_json = json.loads(row['raw_json'])
        node_data = extract_nodes_from_raw_json(
            raw_json=raw_json,
            batch_id=row['batch_id'],
            enterprise_id=1,
            filename=row['filename']
        )
        
        print(f"\n✅ Всего извлечено: {len(node_data)} записей")
        for i, record in enumerate(node_data[:5], 1):
            print(f"   {i}. {record['node_name']}: {record.get('active_energy_kwh', 'N/A')} кВт·ч")

