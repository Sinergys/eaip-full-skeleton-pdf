"""
Скрипт для проверки незаполненных листов в шаблоне.
Показывает, какие листы не имеют функций заполнения.
"""
import sys
from pathlib import Path
from typing import List, Dict, Set
from openpyxl import load_workbook

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Импортируем все функции заполнения
try:
    from tools.fill_energy_passport import (
        fill_struktura_pr2,
        fill_nodes_sheet,
        fill_building_envelope_sheet,
        fill_equipment_sheet,
        fill_balans_sheet,
        fill_dinamika_sheet,
        fill_meropriyatiya_sheet,
        fill_fuel_dynamics_sheet,
        fill_specific_consumption_sheet,
        fill_monthly_sheet,
    )
except ImportError as e:
    print(f"Ошибка импорта: {e}")
    sys.exit(1)


def get_filled_sheets() -> Dict[str, List[str]]:
    """Возвращает словарь: функция -> список названий листов, которые она заполняет"""
    return {
        "fill_struktura_pr2": [
            "Структура пр 2", "Struktura pr2", "02_Структура", "Структура пр 2 "
        ],
        "fill_nodes_sheet": [
            "01_Узлы учета", "Узел учета", "Узлы учета", "Nodes", "Uzel ucheta", "Узел учета "
        ],
        "fill_building_envelope_sheet": [
            "02_Исходные данные", "Ограждающие", "Envelope", "Ограждающие конструкции"
        ],
        "fill_equipment_sheet": [
            "Equipment", "АНАЛИЗ ОБОРУДОВАНИЯ", "Анализ оборудования", "Оборудование", 
            "03_Оборудование", "оборудование"
        ],
        "fill_balans_sheet": [
            "Баланс", "04_Баланс", "Balance"
        ],
        "fill_dinamika_sheet": [
            "Динамика ср", "Динамика", "05_Динамика", "Dynamics", "Dinamika sr"
        ],
        "fill_fuel_dynamics_sheet": [
            "мазут,уголь 5", "мазут,уголь", "06_Мазут_Уголь", "Fuel Dynamics", 
            "мазут уголь 5", "Мазут,уголь 5"
        ],
        "fill_specific_consumption_sheet": [
            "Расход на ед.п", "Расход  на ед.п", "07_Расход_на_ед", "Specific Consumption",
            "Расход на единиц", "Расход на ед"
        ],
        "fill_meropriyatiya_sheet": [
            "Мериаприятия 1", "Мероприятия", "08_Мероприятия", "Measures",
            "Мериаприятия 1 ", "Мероприятия 1"
        ],
        "fill_monthly_sheet": [
            "Monthly", "MONTHLY", "Месячные данные", "Месячный", "месячные"
        ],
    }


def check_template_sheets(template_path: Path) -> Dict[str, any]:
    """Проверяет, какие листы в шаблоне заполняются, а какие нет"""
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    all_sheets = workbook.sheetnames
    
    filled_sheets_map = get_filled_sheets()
    
    # Создаем обратный маппинг: название листа -> функция
    sheet_to_function: Dict[str, str] = {}
    covered_sheets: Set[str] = set()
    
    for func_name, sheet_names in filled_sheets_map.items():
        for sheet_name in sheet_names:
            sheet_to_function[sheet_name] = func_name
            # Добавляем также варианты с разными регистрами
            sheet_to_function[sheet_name.lower()] = func_name
            sheet_to_function[sheet_name.upper()] = func_name
    
    # Проверяем каждый лист шаблона
    covered = {}
    uncovered = []
    
    for sheet_name in all_sheets:
        found = False
        
        # Точное совпадение
        if sheet_name in sheet_to_function:
            func_name = sheet_to_function[sheet_name]
            if func_name not in covered:
                covered[func_name] = []
            covered[func_name].append(sheet_name)
            covered_sheets.add(sheet_name)
            found = True
        
        # Частичное совпадение (ищем по ключевым словам)
        if not found:
            sheet_lower = sheet_name.lower()
            for func_name, sheet_names in filled_sheets_map.items():
                for known_sheet in sheet_names:
                    known_lower = known_sheet.lower()
                    # Проверяем, содержит ли название листа ключевые слова
                    if known_lower in sheet_lower or any(
                        word in sheet_lower 
                        for word in known_lower.split() 
                        if len(word) > 3
                    ):
                        if func_name not in covered:
                            covered[func_name] = []
                        covered[func_name].append(sheet_name)
                        covered_sheets.add(sheet_name)
                        found = True
                        break
                if found:
                    break
        
        if not found:
            uncovered.append(sheet_name)
    
    workbook.close()
    
    return {
        "template_path": str(template_path),
        "total_sheets": len(all_sheets),
        "all_sheets": all_sheets,
        "covered": covered,
        "uncovered": uncovered,
        "coverage_percentage": round((len(covered_sheets) / len(all_sheets) * 100) if all_sheets else 0, 2)
    }


def print_report(report: Dict) -> None:
    """Выводит отчет в консоль"""
    print("=" * 80)
    print("📋 ОТЧЕТ О ПОКРЫТИИ ЛИСТОВ ШАБЛОНА ФУНКЦИЯМИ ЗАПОЛНЕНИЯ")
    print("=" * 80)
    print(f"Шаблон: {report['template_path']}")
    print(f"Всего листов: {report['total_sheets']}")
    print(f"Покрытие: {report['coverage_percentage']}%")
    print()
    
    # Заполненные листы
    print("✅ ЛИСТЫ С ФУНКЦИЯМИ ЗАПОЛНЕНИЯ:")
    print("-" * 80)
    for func_name, sheets in report['covered'].items():
        print(f"  {func_name}:")
        for sheet in sheets:
            print(f"    - {sheet}")
    
    print()
    
    # Незаполненные листы
    if report['uncovered']:
        print(f"❌ ЛИСТЫ БЕЗ ФУНКЦИЙ ЗАПОЛНЕНИЯ ({len(report['uncovered'])}):")
        print("-" * 80)
        for sheet in report['uncovered']:
            print(f"  - {sheet}")
        print()
        print("⚠️  ТРЕБУЕТСЯ СОЗДАТЬ ФУНКЦИИ ЗАПОЛНЕНИЯ ДЛЯ ЭТИХ ЛИСТОВ!")
    else:
        print("✅ ВСЕ ЛИСТЫ ИМЕЮТ ФУНКЦИИ ЗАПОЛНЕНИЯ!")
    
    print("=" * 80)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Проверка покрытия листов шаблона функциями заполнения")
    parser.add_argument("--template", required=True, help="Путь к шаблону Excel")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    
    try:
        report = check_template_sheets(template_path)
        print_report(report)
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        sys.exit(1)

