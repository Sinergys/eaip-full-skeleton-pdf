"""Поиск всех таблиц электроэнергии по видам продукции на листе Структура пр 2"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

template_path = Path(
    r"data\source_files\audit_sinergys\Энергопаспорт Метин Ирода_с_объемами_и_месяцами.xlsx"
)

print("=" * 80)
print("ПОИСК ТАБЛИЦ ЭЛЕКТРОЭНЕРГИИ ПО ВИДАМ ПРОДУКЦИИ")
print("=" * 80)

wb = load_workbook(template_path, data_only=False)
ws = wb["Структура пр 2"]

# Ищем все строки с названиями продукции
print("\nПоиск строк с названиями продукции:\n")
product_keywords = ["труба", "фитинг", "канал", "тёплый", "теплый"]

found_tables = []

for row in range(1, 200):
    val_a = ws.cell(row, 1).value
    if val_a and isinstance(val_a, str):
        val_lower = val_a.lower()
        if any(keyword in val_lower for keyword in product_keywords):
            # Нашли строку с продукцией, проверяем структуру
            row_data = []
            for col in range(1, 20):
                val = ws.cell(row, col).value
                if val is not None:
                    col_letter = get_column_letter(col)
                    row_data.append(f"{col_letter}{row}={val}")

            # Проверяем, есть ли заголовок таблицы выше
            header_row = None
            for check_row in range(max(1, row - 5), row):
                header_val = ws.cell(check_row, 1).value
                if header_val and isinstance(header_val, str):
                    if (
                        "продукция" in header_val.lower()
                        or "норма" in header_val.lower()
                        or "перерасход" in header_val.lower()
                    ):
                        header_row = check_row
                        break

            # Определяем квартал по колонкам
            # Ищем колонки с годами (2022, 2023, 2024)
            years_found = {}
            for col in range(1, 200):
                val = ws.cell(row - 2, col).value if row > 2 else None
                if isinstance(val, int) and val in (2022, 2023, 2024):
                    years_found[val] = col

            print(f"Строка {row}: {val_a}")
            if header_row:
                print(f"  Заголовок таблицы: строка {header_row}")
            if years_found:
                print(f"  Годы найдены: {years_found}")
            print(f"  Данные: {', '.join(row_data[:8])}")
            print()

            found_tables.append(
                {
                    "row": row,
                    "product": val_a,
                    "header_row": header_row,
                    "years": years_found,
                }
            )

# Группируем по кварталам
print("=" * 80)
print("ГРУППИРОВКА ПО КВАРТАЛАМ")
print("=" * 80)

# Ищем паттерн: таблицы повторяются каждые ~22 строки (для каждого квартала)
tables_by_quarter = {}
current_quarter = None
current_year = None

for i, table in enumerate(found_tables):
    row = table["row"]

    # Определяем квартал по позиции строки
    # Примерно: строки 17-21 = Q1, 39-43 = Q2, 61-65 = Q3, 83-87 = Q4
    if 15 <= row <= 25:
        quarter = "Q1"
    elif 37 <= row <= 47:
        quarter = "Q2"
    elif 59 <= row <= 69:
        quarter = "Q3"
    elif 81 <= row <= 91:
        quarter = "Q4"
    else:
        quarter = f"Unknown_{row}"

    # Определяем год по колонкам с данными
    # Ищем колонки с числовыми значениями в этой строке
    year = None
    for col in range(2, 20):
        val = ws.cell(row, col).value
        if isinstance(val, (int, float)) and val > 100:  # Вероятно факт
            # Проверяем заголовок колонки
            header_val = ws.cell(row - 2, col).value if row > 2 else None
            if isinstance(header_val, int) and header_val in (2022, 2023, 2024):
                year = header_val
                break

    if not year:
        # Пытаемся определить по позиции колонки
        # Колонки 2-8 обычно 2022, 9-15 = 2023, 16-22 = 2024
        if 2 <= col <= 8:
            year = 2022
        elif 9 <= col <= 15:
            year = 2023
        elif 16 <= col <= 22:
            year = 2024

    quarter_key = f"{year}-{quarter}" if year else quarter
    if quarter_key not in tables_by_quarter:
        tables_by_quarter[quarter_key] = []
    tables_by_quarter[quarter_key].append(table)

print("\nНайденные таблицы по кварталам:")
for quarter_key, tables in sorted(tables_by_quarter.items()):
    print(f"\n{quarter_key}:")
    for table in tables:
        print(f"  Строка {table['row']}: {table['product']}")

# Проверяем структуру первой таблицы детально
if found_tables:
    first_table = found_tables[0]
    row = first_table["row"]
    print("\n" + "=" * 80)
    print(f"ДЕТАЛЬНАЯ СТРУКТУРА ПЕРВОЙ ТАБЛИЦЫ (строка {row})")
    print("=" * 80)

    # Показываем заголовки
    print("\nЗаголовки таблицы:")
    for header_row in range(max(1, row - 3), row):
        row_data = []
        for col in range(1, 10):
            val = ws.cell(header_row, col).value
            if val is not None:
                col_letter = get_column_letter(col)
                row_data.append(f"{col_letter}{header_row}={val}")
        if row_data:
            print(f"  Строка {header_row}: {', '.join(row_data)}")

    # Показываем данные по продукции
    print("\nДанные по продукции:")
    for product_row in range(row, row + 5):
        row_data = []
        for col in range(1, 10):
            val = ws.cell(product_row, col).value
            if val is not None:
                col_letter = get_column_letter(col)
                row_data.append(f"{col_letter}{product_row}={val}")
        if row_data:
            print(f"  Строка {product_row}: {', '.join(row_data)}")

wb.close()
