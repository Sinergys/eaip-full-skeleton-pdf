"""
Улучшенное извлечение таблицы данных из Excel файлов
Более точный поиск начала таблицы и извлечение всех строк данных
"""
import sys
from pathlib import Path
import json

try:
    import openpyxl
except ImportError:
    print("Требуется openpyxl: pip install openpyxl")
    sys.exit(1)

# Добавляем путь к проекту
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

def find_table_start_improved(sheet, min_cols=5):
    """
    Улучшенный поиск начала таблицы данных
    
    Ищет строку с заголовками таблицы, которая содержит:
    - Паттерны заголовков (№, Наименование, Количество, Цена, и т.д.)
    - Минимум min_cols непустых ячеек
    - Следующая строка содержит числовые данные
    """
    # Паттерны для поиска заголовков таблицы
    header_patterns = [
        '№', 'N', 'No', 'номер',
        'наименование', 'название', 'товар', 'услуга',
        'количество', 'кол-во', 'ед', 'единица',
        'цена', 'стоимость', 'сумма', 'ндс'
    ]
    
    rows_list = list(sheet.iter_rows(values_only=True))
    
    for row_idx, row in enumerate(rows_list, 1):
        row_data = [str(cell).strip().lower() if cell else "" for cell in row]
        
        # Пропускаем пустые строки
        if not any(row_data):
            continue
        
        # Проверяем, содержит ли строка паттерны заголовков
        matches = sum(1 for cell in row_data if any(pattern in cell for pattern in header_patterns))
        non_empty_cols = sum(1 for cell in row_data if cell)
        
        # Если найдено достаточно совпадений и достаточно столбцов
        if matches >= 2 and non_empty_cols >= min_cols:
            # Проверяем, что следующая строка содержит данные (не пустая)
            if row_idx < len(rows_list):
                next_row = rows_list[row_idx]
                next_row_data = [str(cell).strip() if cell else "" for cell in next_row]
                if any(next_row_data):
                    return row_idx
    
    return None

def extract_table_data_improved(sheet, start_row):
    """Извлекает данные таблицы начиная с указанной строки"""
    rows = []
    headers = []
    
    rows_list = list(sheet.iter_rows(values_only=True))
    
    # Читаем заголовки (может быть 1 или 2 строки)
    header_row1 = rows_list[start_row - 1]
    headers1 = [str(cell).strip() if cell else "" for cell in header_row1]
    
    # Проверяем, есть ли вторая строка заголовков
    if start_row < len(rows_list):
        header_row2 = rows_list[start_row]
        headers2 = [str(cell).strip() if cell else "" for cell in header_row2]
        
        # Если вторая строка содержит данные (не пустая и не начинается с числа)
        if any(headers2) and not str(headers2[0]).strip().isdigit():
            # Объединяем заголовки
            max_cols = max(len(headers1), len(headers2))
            headers = []
            for i in range(max_cols):
                h1 = headers1[i] if i < len(headers1) else ""
                h2 = headers2[i] if i < len(headers2) else ""
                combined = f"{h1} {h2}".strip()
                headers.append(combined)
            data_start_row = start_row + 2  # Пропускаем обе строки заголовков + возможную пустую
        else:
            headers = headers1
            data_start_row = start_row + 1
    else:
        headers = headers1
        data_start_row = start_row + 1
    
    # Читаем данные (начиная с data_start_row)
    for row_idx in range(data_start_row - 1, len(rows_list)):
        row = rows_list[row_idx]
        row_data = [str(cell).strip() if cell else "" for cell in row]
        
        # Пропускаем полностью пустые строки
        if not any(row_data):
            # Если уже есть данные, проверяем следующие строки
            if rows:
                # Пропускаем одну пустую строку, но если следующая тоже пустая - останавливаемся
                if row_idx + 1 < len(rows_list):
                    next_row = rows_list[row_idx + 1]
                    next_row_data = [str(cell).strip() if cell else "" for cell in next_row]
                    if not any(next_row_data):
                        break
                else:
                    break
            continue
        
        # Останавливаемся, если встретили строку "Итого" или похожую
        row_text = " ".join(row_data).lower()
        if any(keyword in row_text for keyword in ['итого', 'total', 'сумма', 'всего', 'стороны претензий', 'стоимость принятой']):
            rows.append(row_data)
            break
        
        # Проверяем, похожа ли строка на данные таблицы
        first_cell = str(row_data[0]).strip() if row_data else ""
        
        # Если первая ячейка - число (номер строки) или пустая, но есть данные в других столбцах
        is_data_row = False
        if first_cell.isdigit() or (not first_cell and any(row_data[1:5])):  # Первые 5 столбцов после номера
            is_data_row = True
        elif first_cell in ['№', 'N', 'No'] and rows:  # Если это заголовок после данных - конец
            break
        elif not first_cell and rows and not any(row_data[1:3]):  # Пустая строка после данных
            break
        
        if is_data_row or (rows and any(row_data[1:5])):  # Если есть данные в столбцах 2-5
            rows.append(row_data)
    
    return headers, rows

def normalize_table(headers, rows):
    """Нормализует таблицу: убирает пустые столбцы, выравнивает количество столбцов"""
    # Находим максимальное количество столбцов
    max_cols = max(len(headers), max((len(row) for row in rows), default=0))
    
    # Выравниваем заголовки
    headers = headers + [""] * (max_cols - len(headers))
    
    # Выравниваем строки
    normalized_rows = []
    for row in rows:
        normalized_row = row + [""] * (max_cols - len(row))
        normalized_rows.append(normalized_row)
    
    # Убираем полностью пустые столбцы справа
    while max_cols > 0:
        col_has_data = any(headers[max_cols - 1]) or any(row[max_cols - 1] for row in normalized_rows)
        if col_has_data:
            break
        max_cols -= 1
    
    if max_cols < len(headers):
        headers = headers[:max_cols]
        normalized_rows = [row[:max_cols] for row in normalized_rows]
    
    return headers, normalized_rows

def extract_table_from_excel_improved(file_path: str, output_path: str = None) -> tuple:
    """Извлекает таблицу данных из Excel файла (улучшенная версия)"""
    file_path_obj = Path(file_path)
    
    if not file_path_obj.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    workbook = openpyxl.load_workbook(file_path_obj, data_only=True)
    result = {
        "tables": [],
        "text": ""
    }
    
    # Обрабатываем первый лист (обычно там таблица)
    sheet = workbook.active
    
    # Находим начало таблицы
    start_row = find_table_start_improved(sheet, min_cols=5)
    
    if start_row is None:
        # Пробуем найти строку с "№" в первом столбце
        rows_list = list(sheet.iter_rows(values_only=True))
        for row_idx, row in enumerate(rows_list, 1):
            if row and str(row[0]).strip() in ['№', 'N', 'No', '1', '2', '3']:
                # Проверяем, что в строке достаточно столбцов
                non_empty = sum(1 for cell in row if cell)
                if non_empty >= 5:
                    start_row = row_idx - 1  # Предыдущая строка - заголовки
                    if start_row < 1:
                        start_row = row_idx
                    break
    
    if start_row is None:
        raise ValueError(f"Не удалось найти данные в файле {file_path_obj.name}")
    
    print(f"✅ Найдено начало таблицы: строка {start_row}")
    
    # Извлекаем данные
    headers, rows = extract_table_data_improved(sheet, start_row)
    
    # Нормализуем
    headers, rows = normalize_table(headers, rows)
    
    # Формируем результат
    table_data = {
        "rows": [headers] + rows,
        "headers": headers
    }
    
    result["tables"].append(table_data)
    
    # Сохраняем в JSON
    if output_path is None:
        output_path = file_path_obj.parent / f"{file_path_obj.stem}_table_extracted.json"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Таблица извлечена: {len(rows)} строк данных, {len(headers)} столбцов")
    
    return str(output_path), result

def main():
    """Основная функция"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Улучшенное извлечение таблицы данных из Excel файлов')
    parser.add_argument('files', nargs='+', help='Пути к Excel файлам')
    parser.add_argument('--output-dir', type=str, help='Директория для сохранения JSON файлов')
    
    args = parser.parse_args()
    
    print("=" * 80)
    print("УЛУЧШЕННОЕ ИЗВЛЕЧЕНИЕ ТАБЛИЦЫ ИЗ EXCEL ФАЙЛОВ")
    print("=" * 80)
    print()
    
    output_dir = Path(args.output_dir) if args.output_dir else project_root / "reports" / "ocr"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    extracted_files = []
    
    for file_path in args.files:
        print(f"\n📄 Обработка: {file_path}")
        try:
            output_file = output_dir / f"{Path(file_path).stem}_table_extracted.json"
            extracted_path, data = extract_table_from_excel_improved(file_path, str(output_file))
            extracted_files.append(extracted_path)
            
            # Показываем статистику
            if data.get('tables'):
                table = data['tables'][0]
                print(f"   📊 Статистика:")
                print(f"      - Строк (включая заголовок): {len(table.get('rows', []))}")
                print(f"      - Строк данных: {len(table.get('rows', [])) - 1}")
                print(f"      - Столбцов: {len(table.get('headers', []))}")
                headers_preview = ', '.join([h[:20] for h in table.get('headers', [])[:5]])
                print(f"      - Заголовки: {headers_preview}...")
            
        except Exception as e:
            print(f"❌ Ошибка обработки {file_path}: {e}")
            import traceback
            traceback.print_exc()
    
    if extracted_files:
        print()
        print("=" * 80)
        print("✅ ИЗВЛЕЧЕНИЕ ЗАВЕРШЕНО")
        print("=" * 80)
        print()
        print("📋 Извлеченные файлы:")
        for f in extracted_files:
            print(f"  - {f}")
        print()
        print("📌 Теперь можно запустить сравнение:")
        if len(extracted_files) == 1:
            print(f"  python tools/compare_recognition_results.py --manual \"{extracted_files[0]}\"")
        else:
            print(f"  python tools/compare_recognition_results.py --manual1 \"{extracted_files[0]}\" --manual2 \"{extracted_files[1]}\"")

if __name__ == "__main__":
    main()

