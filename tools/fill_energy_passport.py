"""
Utility script to populate the EnergyPassport_PKM690 template with aggregated
energy data, meter nodes, and transformer losses;
configurable through command-line arguments.
"""

from __future__ import annotations
from openpyxl.cell.cell import MergedCell

import argparse
import json
import os
import logging
import shutil
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

# Импорт централизованных формул и единиц измерения
import sys

_calc_module_path = (
    Path(__file__).parent.parent
    / "eaip_full_skeleton"
    / "services"
    / "ingest"
    / "domain"
)
if str(_calc_module_path) not in sys.path:
    sys.path.insert(0, str(_calc_module_path))

try:
    from energy_passport_calculations import (
        calculate_quarter_losses,
        calculate_loss_percentage,
        calculate_specific_consumption,
        calculate_equipment_usage_coefficient,
        calculate_annual_consumption_from_power,
        calculate_average_power_per_unit,
        calculate_balance_total,
        calculate_equipment_used_power,
        extract_quarter_data,
        extract_equipment_data,
        QuarterData,
        EquipmentData,
    )
    from energy_units import (
        to_kwh,
        to_mwh,
        to_kvarh,
        to_m3,
        to_ton,
        to_gcal,
        HOURS_PER_YEAR,
        HOURS_PER_QUARTER,
        MONTHS_PER_QUARTER,
    )

    HAS_CALCULATIONS = True
except ImportError as e:
    import logging

    logging.warning(
        f"Не удалось импортировать модули расчётов: {e}. Используются встроенные формулы."
    )
    HAS_CALCULATIONS = False

# Настройка логирования
logger = logging.getLogger(__name__)

DEFAULT_NODES_HEADER: List[List[Any]] = [
    [
        "Пункты учёта",
        "Вид учёта мощности",
        None,
        "Место установки",
        None,
        "Коэффициент учёта",
        "Дата пломбировки",
        "Дата пломбировки поставщика",
        "Примечание",
    ],
    [None, "P", "Q", "TT", "КT", None, None, None, None],
    [None, "кВт·ч", "кВАр·ч", None, None, None, None, None, None],
]

QuarterMapping = Dict[str, Tuple[int, int, int, int, int]]


def build_quarter_mapping() -> QuarterMapping:
    """
    Маппинг кварталов на позиции в листе Struktura pr2.
    Формат: (row, col_active, col_reactive, col_gas, col_water)
    """
    mapping = {}
    # 2022
    mapping["2022-Q1"] = (9, 3, 4, 6, 14)
    mapping["2022-Q2"] = (9, 19, 20, 22, 30)
    mapping["2022-Q3"] = (9, 35, 36, 38, 46)
    mapping["2022-Q4"] = (9, 51, 52, 54, 62)
    # 2023
    mapping["2023-Q1"] = (9, 67, 68, 70, 78)
    mapping["2023-Q2"] = (9, 83, 84, 86, 94)
    mapping["2023-Q3"] = (9, 99, 100, 102, 110)
    mapping["2023-Q4"] = (9, 115, 116, 118, 126)
    # 2024
    mapping["2024-Q1"] = (9, 131, 132, 134, 142)
    mapping["2024-Q2"] = (9, 147, 148, 150, 158)
    mapping["2024-Q3"] = (9, 163, 164, 166, 174)
    mapping["2024-Q4"] = (9, 179, 180, 182, 190)
    return mapping


def quarter_loss_totals(loss_month: float) -> Dict[str, float]:
    """
    Расчёт потерь за квартал из месячных потерь.

    Использует централизованную формулу из energy_passport_calculations.
    """
    if HAS_CALCULATIONS:
        quarter_loss = calculate_quarter_losses(loss_month)
        return {quarter: quarter_loss for quarter in build_quarter_mapping().keys()}
    else:
        # Fallback на старую логику
        multiplier = 3
        return {
            quarter: loss_month * multiplier
            for quarter in build_quarter_mapping().keys()
        }


def fill_struktura_pr2(
    ws,
    agg_data: Dict[str, Dict],
    loss_active_month: float = 0.0,
    loss_reactive_month: float = 0.0,
) -> None:
    import logging

    logger = logging.getLogger(__name__)
    debug_mode = os.getenv("INGEST_PASSPORT_DEBUG", "false").lower() == "true"

    quarter_map = build_quarter_mapping()
    loss_active = quarter_loss_totals(loss_active_month) if loss_active_month else {}
    loss_reactive = (
        quarter_loss_totals(loss_reactive_month) if loss_reactive_month else {}
    )

    for quarter, (
        row,
        col_active,
        col_reactive,
        col_gas,
        col_water,
    ) in quarter_map.items():
        elec_totals = (
            agg_data.get("electricity", {}).get(quarter, {}).get("quarter_totals", {})
        )
        gas_totals = agg_data.get("gas", {}).get(quarter, {}).get("quarter_totals", {})
        water_totals = (
            agg_data.get("water", {}).get(quarter, {}).get("quarter_totals", {})
        )

        active_kwh = elec_totals.get("active_kwh", 0) or 0

        # Если active_kwh отсутствует, пытаемся вычислить из месячных данных мощности
        if active_kwh == 0:
            elec_quarter_data = agg_data.get("electricity", {}).get(quarter, {})
            months = elec_quarter_data.get("months", [])

            if months:
                # Ищем данные о расходе активной мощности по месяцам
                power_field_names = [
                    "active_power",
                    "power_kw",
                    "active_power_kw",
                    "расход_активной_мощности",
                    "мощность_квт",
                    "мощность",
                    "active_power_monthly",
                    "power_monthly",
                ]

                monthly_power_sum = 0.0
                power_found = False

                for month in months:
                    values = month.get("values", {})
                    for power_field in power_field_names:
                        if power_field in values and values[power_field]:
                            try:
                                power_value = float(values[power_field])
                                if power_value > 0:
                                    monthly_power_sum += power_value
                                    power_found = True
                                    break
                            except (ValueError, TypeError):
                                continue

                # Если нашли данные о мощности, вычисляем active_kwh
                if power_found and monthly_power_sum > 0:
                    try:
                        avg_monthly_power = monthly_power_sum / len(months)
                        # Используем функцию расчета из модуля расчетов
                        if HAS_CALCULATIONS:
                            from energy_passport_calculations import (
                                calculate_quarter_consumption_from_monthly_power,
                            )
                            from energy_units import HOURS_PER_MONTH

                            calculated_active_kwh = (
                                calculate_quarter_consumption_from_monthly_power(
                                    avg_monthly_power, HOURS_PER_MONTH
                                )
                            )
                        else:
                            # Fallback: простая формула
                            calculated_active_kwh = (
                                avg_monthly_power * 720.0 * 3.0
                            )  # мощность * часы_в_месяце * 3

                        if calculated_active_kwh > 0:
                            active_kwh = calculated_active_kwh
                            logger.info(
                                f"Вычислен active_kwh из месячных данных мощности для квартала {quarter}: "
                                f"{calculated_active_kwh:.2f} кВт·ч "
                                f"(средняя мощность: {avg_monthly_power:.2f} кВт)"
                            )
                    except Exception as e:
                        logger.warning(
                            f"Ошибка при расчете active_kwh из мощности для квартала {quarter}: {e}"
                        )

        reactive_kvarh = elec_totals.get("reactive_kvarh", 0) or 0

        # ВАЖНО: Гарантируем правильный расчет газа из помесячных данных
        # Сначала пытаемся получить из месячных данных
        gas_quarter_data = agg_data.get("gas", {}).get(quarter, {})
        gas_months = gas_quarter_data.get("months", [])
        gas_m3_from_months = 0.0

        if gas_months:
            # Суммируем помесячные данные по газу
            for month in gas_months:
                values = month.get("values", {})
                # Ищем поле volume_m3 или другие варианты
                month_gas = values.get("volume_m3", 0) or values.get("gas_m3", 0) or 0
                if month_gas:
                    try:
                        gas_m3_from_months += float(month_gas)
                    except (ValueError, TypeError):
                        pass

        # Используем данные из месячных, если они есть, иначе из quarter_totals
        if gas_m3_from_months > 0:
            gas_m3 = gas_m3_from_months
            logger.debug(
                f"Газ для квартала {quarter} рассчитан из месячных данных: {gas_m3:.3f} м³"
            )
        else:
            gas_m3 = gas_totals.get("volume_m3", 0) or 0
            if gas_m3 > 0:
                logger.debug(
                    f"Газ для квартала {quarter} взят из quarter_totals: {gas_m3:.3f} м³"
                )

        water_m3 = water_totals.get("volume_m3", 0) or 0

        # Логирование недостающих данных
        if active_kwh == 0:
            logger.warning(
                f"Лист 'Struktura pr2', квартал {quarter}: отсутствуют данные по активной электроэнергии (ячейка {row},{col_active})"
            )
        if reactive_kvarh == 0:
            logger.warning(
                f"Лист 'Struktura pr2', квартал {quarter}: отсутствуют данные по реактивной электроэнергии (ячейка {row},{col_reactive})"
            )
        if gas_m3 == 0:
            logger.warning(
                f"Лист 'Struktura pr2', квартал {quarter}: отсутствуют данные по газу (ячейка {row},{col_gas})"
            )
        if water_m3 == 0:
            logger.info(
                f"Лист 'Struktura pr2', квартал {quarter}: отсутствуют данные по воде (ячейка {row},{col_water}) - опционально"
            )

        # Получаем данные по теплу (heat), если есть
        heat_totals = (
            agg_data.get("heat", {}).get(quarter, {}).get("quarter_totals", {})
        )
        heat_gcal = heat_totals.get("volume_gcal", 0) or heat_totals.get("gcal", 0) or 0

        if debug_mode:
            logger.debug(
                f"Заполнение 'Struktura pr2', {quarter}: активная={active_kwh}, реактивная={reactive_kvarh}, газ={gas_m3}, вода={water_m3}, тепло={heat_gcal}"
            )

        # Используем безопасную запись с сохранением формул
        _safe_set_cell_value(ws, row, col_active, active_kwh)
        _safe_set_cell_value(ws, row, col_reactive, reactive_kvarh)
        # ВАЖНО: Газ в шаблоне должен быть в тысячах м³
        gas_m3_thousands = gas_m3 / 1000.0 if gas_m3 > 0 else 0.0
        _safe_set_cell_value(ws, row, col_gas, gas_m3_thousands)
        _safe_set_cell_value(ws, row, col_water, water_m3)

        # Заполняем тепло, если есть колонка для него (обычно после воды)
        # Ищем колонку для тепла: обычно это col_water + 1 или col_water + 8
        # Проверяем несколько возможных позиций
        if heat_gcal > 0:
            # Пробуем найти колонку для тепла (обычно после воды)
            for heat_col_offset in [1, 8, 9]:
                heat_col = col_water + heat_col_offset
                cell = ws.cell(row=row, column=heat_col)
                # Проверяем, не является ли это ячейкой с формулой
                if cell.data_type != "f" or not cell.value:
                    _safe_set_cell_value(ws, row, heat_col, heat_gcal)
                    logger.debug(
                        f"Заполнено тепло для квартала {quarter} в колонку {heat_col}: {heat_gcal} Гкал"
                    )
                    break

        if loss_active_month:
            _safe_set_cell_value(ws, 11, col_active, loss_active.get(quarter, 0.0))
        if loss_reactive_month:
            _safe_set_cell_value(ws, 11, col_reactive, loss_reactive.get(quarter, 0.0))

    # ДОПОЛНИТЕЛЬНО: Заполняем строку 32 "Общее потребление по предприятию" для газа
    # E32 (2023 Q1) должна быть исправлена с 14.0819 на 14.819
    # Структура: строка 32, колонки для газа: E=5 (2023 Q1), T=20 (2023 Q2?), AI=35 (2023 Q3?), AX=50 (2023 Q4?)
    # Определяем маппинг для строки 32 на основе анализа шаблона
    gas_row32_mapping = {
        "2022-Q1": (32, 6),  # F32
        "2022-Q2": (32, 22),  # V32
        "2022-Q3": (32, 38),  # AL32
        "2022-Q4": (32, 54),  # BB32
        "2023-Q1": (32, 5),  # E32 - ИСПРАВЛЯЕМ ЗДЕСЬ!
        "2023-Q2": (32, 20),  # T32
        "2023-Q3": (32, 35),  # AI32
        "2023-Q4": (32, 50),  # AX32
        "2024-Q1": (32, 66),  # BR32
        "2024-Q2": (32, 82),  # CD32
        "2024-Q3": (32, 98),  # CV32
        "2024-Q4": (32, 114),  # DJ32
    }

    for quarter, (row32, col32_gas) in gas_row32_mapping.items():
        gas_quarter_data = agg_data.get("gas", {}).get(quarter, {})
        gas_months = gas_quarter_data.get("months", [])
        gas_m3_from_months = 0.0

        if gas_months:
            # Суммируем помесячные данные по газу (избегаем дубликатов)
            seen_months = set()
            for month in gas_months:
                month_name = month.get("month", "").strip()
                values = month.get("values", {})
                month_gas = values.get("volume_m3", 0) or values.get("gas_m3", 0) or 0

                if not month_gas or month_gas is None:
                    continue

                month_key = f"{quarter}-{month_name}"
                if month_key in seen_months:
                    continue

                try:
                    month_gas = float(month_gas)
                    if month_gas > 0:
                        gas_m3_from_months += month_gas
                        seen_months.add(month_key)
                except (ValueError, TypeError):
                    pass

        if gas_m3_from_months > 0:
            # Записываем в строку 32 в тысячах м³
            gas_m3_thousands = gas_m3_from_months / 1000.0
            _safe_set_cell_value(
                ws, row32, col32_gas, gas_m3_thousands, preserve_formula=False
            )
            logger.info(
                f"Заполнена строка 32 для {quarter}: {gas_m3_thousands:.3f} тыс. м³ (ячейка {get_column_letter(col32_gas)}{row32})"
            )


def fill_losses_sheet(
    workbook,
    loss_active_month: float,
    loss_reactive_month: float,
    transformer_power_kva: float,
    hours_per_month: float,
) -> None:
    sheet_name = "08_Потери_электроэнергии"
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = workbook.create_sheet(title=sheet_name)

    ws["A1"] = "Расчёт потерь трансформатора"
    ws["A3"] = "Мощность трансформатора, kVA"
    ws["B3"] = transformer_power_kva
    ws["A4"] = "Потери активной энергии за месяц, кВт·ч"
    ws["B4"] = loss_active_month
    ws["A5"] = "Потери реактивной энергии за месяц, кВАр·ч"
    ws["B5"] = loss_reactive_month
    ws["A6"] = "Количество часов в месяце"
    ws["B6"] = hours_per_month
    # Используем централизованные формулы
    if HAS_CALCULATIONS:
        loss_active_pct = calculate_loss_percentage(
            loss_active_month, transformer_power_kva, hours_per_month
        )
        loss_reactive_pct = calculate_loss_percentage(
            loss_reactive_month, transformer_power_kva, hours_per_month
        )
        loss_active_quarter = calculate_quarter_losses(loss_active_month)
        loss_reactive_quarter = calculate_quarter_losses(loss_reactive_month)
    else:
        # Fallback на старую логику
        loss_active_pct = (
            (loss_active_month / (transformer_power_kva * 0.9 * hours_per_month) * 100)
            if transformer_power_kva > 0 and hours_per_month > 0
            else 0.0
        )
        loss_reactive_pct = (
            (
                loss_reactive_month
                / (transformer_power_kva * 0.9 * hours_per_month)
                * 100
            )
            if transformer_power_kva > 0 and hours_per_month > 0
            else 0.0
        )
        loss_active_quarter = loss_active_month * 3
        loss_reactive_quarter = loss_reactive_month * 3

    ws["A8"] = "Процент потерь активной энергии"
    ws["B8"] = loss_active_pct
    ws["A9"] = "Процент потерь реактивной энергии"
    ws["B9"] = loss_reactive_pct
    ws["A11"] = "Потери активной энергии за квартал, кВт·ч"
    ws["B11"] = loss_active_quarter
    ws["A12"] = "Потери реактивной энергии за квартал, кВАр·ч"
    ws["B12"] = loss_reactive_quarter


def fill_nodes_sheet(ws, nodes_data: Iterable[Dict[str, Any]] | Dict[str, Any]) -> None:
    ws = _reset_sheet(ws)
    ws.protection = SheetProtection(sheet=False, password=None)

    if isinstance(nodes_data, dict) and "tables" in nodes_data:
        tables = nodes_data.get("tables", [])
    else:
        nodes_list = (
            list(nodes_data)
            if not isinstance(nodes_data, dict)
            else nodes_data.get("nodes", [])
        )
        tables = [
            {
                "header": DEFAULT_NODES_HEADER,
                "columns": len(DEFAULT_NODES_HEADER[0]),
                "nodes": _build_default_rows(nodes_list),
            }
        ]

    max_columns = 0
    current_row = 1
    for table_index, table in enumerate(tables):
        if table_index > 0:
            current_row += 1
        header_rows = table.get("header") or DEFAULT_NODES_HEADER
        nodes_rows = table.get("nodes", [])
        columns = table.get("columns") or len(header_rows[0])
        max_columns = max(max_columns, columns)
        current_row = _write_nodes_table(
            ws, header_rows, nodes_rows, start_row=current_row, column_count=columns
        )

    if max_columns:
        _auto_fit_columns(ws, num_columns=max_columns)


def load_nodes_from_json(path: Path) -> List[Dict[str, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("nodes", [])


def load_default_nodes() -> List[Dict[str, Any]]:
    return [
        {
            "name": "Трёхфазный счётчик",
            "resource": "Электрическая энергия",
            "meter_type": "EX518",
            "serial_number": "124200393818",
            "location": "ТП-1 (ПС «Каракамиш» 35/6 кВ)",
            "kI": 240,
            "k": 240,
            "details": "год пломбировки: 2024; дата: 12.2024",
        },
        {
            "name": "Трёхфазный счётчик",
            "resource": "Электрическая энергия",
            "meter_type": "EX518",
            "serial_number": "124200507111",
            "location": "ТП-2 (ПС «Каракамиш» 35/6 кВ)",
            "kI": 240,
            "k": 240,
            "details": "год пломбировки: 2024; дата: 12.2024",
        },
        {
            "name": "Трёхфазный счётчик",
            "resource": "Электрическая энергия",
            "meter_type": "EX518",
            "serial_number": "124200573458",
            "location": "ТП-3 (ПС «Каракамиш» 35/6 кВ)",
            "kI": 300,
            "k": 300,
            "details": "год пломбировки: 2024; дата: 12.2024",
        },
    ]


def fill_building_envelope_sheet(
    workbook, envelope_data: Dict, sheet_name: str
) -> None:
    """
    Заполняет лист расчета теплопотерь по зданиям данными из envelope_data.

    Args:
        workbook: Рабочая книга Excel
        envelope_data: Данные расчета теплопотерь по зданиям (sections, summary)
        sheet_name: Имя листа для заполнения
    """
    sections = envelope_data.get("sections") or []
    if not sections:
        return

    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        ws.protection = SheetProtection(sheet=False, password=None)
    else:
        ws = workbook.create_sheet(title=sheet_name)
        ws.protection = SheetProtection(sheet=False, password=None)

    headers = [
        "Раздел",
        "Конструкция",
        "Материал",
        "Толщина, м",
        "λ, Вт/(м·°C)",
        "R, м²·°C/Вт",
        "Rтр, м²·°C/Вт",
        "Площадь, м²",
        "ΔT, °C",
        "Теплопотери Q",
    ]
    for col_idx, header in enumerate(headers, start=1):
        safe_cell_write(ws.cell(row=1, column=col_idx), header)

    current_row = 2
    for section in sections:
        section_name = section.get("section") or "Раздел"
        items = section.get("items") or []
        totals = section.get("totals") or {}

        for idx, item in enumerate(items):
            safe_cell_write(ws.cell(row=current_row, column=1), section_name if idx == 0 else "")
            safe_cell_write(ws.cell(row=current_row, column=2), item.get("construction"))
            safe_cell_write(ws.cell(row=current_row, column=3), item.get("material"))
            safe_cell_write(ws.cell(row=current_row, column=4), _format_float(
                item.get("thickness_m")
            ))
            safe_cell_write(ws.cell(row=current_row, column=5), _format_float(
                item.get("lambda_w_mk")
            ))
            safe_cell_write(ws.cell(row=current_row, column=6), _format_float(
                item.get("resistance_r")
            ))
            safe_cell_write(ws.cell(row=current_row, column=7), _format_float(
                item.get("normative_r")
            ))
            safe_cell_write(ws.cell(row=current_row, column=8), _format_float(
                item.get("area_m2")
            ))
            safe_cell_write(ws.cell(row=current_row, column=9), _format_float(
                item.get("delta_t")
            ))
            safe_cell_write(ws.cell(row=current_row, column=10), _format_float(
                item.get("heat_loss_q")
            ))
            current_row += 1

        if totals:
            safe_cell_write(ws.cell(row=current_row, column=2), "Итого")
            safe_cell_write(ws.cell(row=current_row, column=8), _format_float(
                totals.get("area_m2")
            ))
            safe_cell_write(ws.cell(row=current_row, column=10), _format_float(
                totals.get("heat_loss_q")
            ))
            current_row += 1

        current_row += 1  # пустая строка между секциями

    summary = envelope_data.get("summary") or {}
    if summary:
        safe_cell_write(ws.cell(row=current_row, column=1), "Сводные показатели")
        safe_cell_write(ws.cell(row=current_row + 1, column=1), "Количество участков")
        safe_cell_write(ws.cell(row=current_row + 1, column=2), summary.get("total_sections"))
        safe_cell_write(ws.cell(row=current_row + 2, column=1), "Суммарная площадь, м²")
        safe_cell_write(ws.cell(row=current_row + 2, column=2), _format_float(
            summary.get("total_area_m2")
        ))
        safe_cell_write(ws.cell(row=current_row + 3, column=1), "Суммарные теплопотери Q")
        safe_cell_write(ws.cell(row=current_row + 3, column=2), _format_float(
            summary.get("total_heat_loss")
        ))

    _auto_fit_columns(ws, num_columns=len(headers))


def _fix_circular_references(ws) -> None:
    """
    Исправляет циклические ссылки в формулах листа.

    Находит формулы, которые ссылаются на сами себя (например, =СУММ(F4:F5) в ячейке F4)
    и исправляет их, исключая текущую ячейку из диапазона.
    """
    import logging

    logger = logging.getLogger(__name__)
    import re

    fixed_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value:  # Формула
                formula = str(cell.value)
                cell_coord = cell.coordinate  # Например, "F4"

                # Проверяем, содержит ли формула ссылку на текущую ячейку
                # Паттерн для поиска диапазонов типа F4:F5, F:F, 4:5 и т.д.
                patterns = [
                    r"([A-Z]+)(\d+):([A-Z]+)(\d+)",  # F4:F5
                    r"([A-Z]+)(\d+):([A-Z]+)(\d+)",  # Повтор для других форматов
                ]

                # Извлекаем номер строки и букву колонки из координаты ячейки
                col_letter = "".join([c for c in cell_coord if c.isalpha()])
                row_num = int("".join([c for c in cell_coord if c.isdigit()]))

                # Проверяем, есть ли в формуле ссылка на текущую ячейку
                if col_letter and row_num:
                    # Ищем все вхождения координаты текущей ячейки в формуле
                    cell_ref_pattern = f"{col_letter}{row_num}"
                    if cell_ref_pattern in formula:
                        # Это циклическая ссылка - нужно исправить
                        logger.warning(
                            f"Найдена циклическая ссылка в ячейке {cell_coord}: {formula}"
                        )

                        # Пытаемся исправить диапазон, исключая текущую ячейку
                        # Например, =СУММ(F4:F5) в F4 -> =СУММ(F5:F6) или =F5
                        # Простое решение: заменяем ссылку на текущую ячейку на пустую строку или 0
                        # Или заменяем диапазон, исключая текущую ячейку

                        # Если формула типа =СУММ(F4:F5) в ячейке F4, заменяем на =СУММ(F5:F5) или =F5
                        if "СУММ" in formula or "SUM" in formula.upper():
                            # Находим диапазон в формуле
                            range_match = re.search(r"([A-Z]+\d+):([A-Z]+\d+)", formula)
                            if range_match:
                                start_ref = range_match.group(1)
                                end_ref = range_match.group(2)

                                # Если начало диапазона - это текущая ячейка
                                if start_ref == cell_coord:
                                    # Заменяем на диапазон, начинающийся со следующей строки
                                    start_col = "".join(
                                        [c for c in start_ref if c.isalpha()]
                                    )
                                    start_row = int(
                                        "".join([c for c in start_ref if c.isdigit()])
                                    )
                                    new_start = f"{start_col}{start_row + 1}"
                                    new_formula = formula.replace(
                                        f"{start_ref}:{end_ref}",
                                        f"{new_start}:{end_ref}",
                                    )
                                    cell.value = new_formula
                                    fixed_count += 1
                                    logger.info(
                                        f"Исправлена циклическая ссылка в {cell_coord}: {formula} -> {new_formula}"
                                    )
                                # Если конец диапазона - это текущая ячейка
                                elif end_ref == cell_coord:
                                    # Заменяем на диапазон, заканчивающийся предыдущей строкой
                                    end_col = "".join(
                                        [c for c in end_ref if c.isalpha()]
                                    )
                                    end_row = int(
                                        "".join([c for c in end_ref if c.isdigit()])
                                    )
                                    new_end = f"{end_col}{end_row - 1}"
                                    new_formula = formula.replace(
                                        f"{start_ref}:{end_ref}",
                                        f"{start_ref}:{new_end}",
                                    )
                                    cell.value = new_formula
                                    fixed_count += 1
                                    logger.info(
                                        f"Исправлена циклическая ссылка в {cell_coord}: {formula} -> {new_formula}"
                                    )

    if fixed_count > 0:
        logger.info(f"Исправлено {fixed_count} циклических ссылок в листе {ws.title}")


def fill_equipment_sheet(workbook, equipment_data: Dict, sheet_name: str) -> None:
    """Заполняет лист анализа оборудования"""
    import logging

    logger = logging.getLogger(__name__)

    sheets_data = equipment_data.get("sheets") or []
    sections: List[Dict] = []
    for sheet in sheets_data:
        sheet_sections = sheet.get("sections") or []
        for section in sheet_sections:
            section_copy = dict(section)
            section_copy["sheet_name"] = sheet.get("sheet")
            sections.append(section_copy)

    if sheet_name not in workbook.sheetnames:
        logger.warning(
            f"Лист '{sheet_name}' не найден в шаблоне. Доступные листы: {workbook.sheetnames[:5]}"
        )
        return

    ws = workbook[sheet_name]
    ws.protection = SheetProtection(sheet=False, password=None)

    # Исправляем циклические ссылки в формулах перед заполнением
    _fix_circular_references(ws)

    logger.info(
        f"Обработка листа '{sheet_name}', max_row={ws.max_row}, sections={len(sections)}"
    )

    # Получаем summary для использования в анализе эффективности
    summary = equipment_data.get("summary") or {}
    logger.info(f"Summary данных: {summary}")
    logger.info(
        f"Данные оборудования: sheets={len(sheets_data)}, sections={len(sections)}, summary_items={summary.get('total_items', 0)}, summary_power={summary.get('total_power_kw', 0.0)}"
    )

    # Находим или создаем заголовок "ПЕРЕЧЕНЬ ОСНОВНОГО ОБОРУДОВАНИЯ"
    equipment_list_start_row = None
    efficiency_start_row = None

    # Ищем заголовки в существующем листе (проверяем больше строк и колонок)
    max_check_rows = min(100, ws.max_row + 1) if ws.max_row else 100
    for row_idx in range(1, max_check_rows):
        # Проверяем несколько колонок для поиска заголовков
        for col_idx in range(1, 10):  # Увеличиваем диапазон поиска
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                if cell_value and isinstance(cell_value, str):
                    cell_upper = cell_value.upper().strip()
                    if not equipment_list_start_row:
                        # Ищем различные варианты заголовка
                        if (
                            ("ПЕРЕЧЕНЬ" in cell_upper and "ОБОРУДОВАНИЯ" in cell_upper)
                            or ("LIST" in cell_upper and "EQUIPMENT" in cell_upper)
                            or ("ПЕРЕЧЕНЬ" in cell_upper and "ОСНОВНОГО" in cell_upper)
                        ):
                            equipment_list_start_row = row_idx
                            logger.info(
                                f"✅ Найден заголовок перечня оборудования в строке {row_idx}, колонка {col_idx}: '{cell_value}'"
                            )
                    if not efficiency_start_row:
                        if (
                            "АНАЛИЗ" in cell_upper and "ЭФФЕКТИВНОСТИ" in cell_upper
                        ) or ("ANALYSIS" in cell_upper and "EFFICIENCY" in cell_upper):
                            efficiency_start_row = row_idx
                            logger.info(
                                f"✅ Найден заголовок анализа эффективности в строке {row_idx}, колонка {col_idx}: '{cell_value}'"
                            )
            except Exception as e:
                logger.debug(f"Ошибка при проверке ячейки {row_idx},{col_idx}: {e}")
                continue

    # Если не нашли заголовки, но есть данные, создаем структуру с начала
    if not equipment_list_start_row:
        if sections or summary:
            logger.info(
                "⚠️ Заголовок 'ПЕРЕЧЕНЬ ОСНОВНОГО ОБОРУДОВАНИЯ' не найден, создаем структуру с начала листа"
            )
            equipment_list_start_row = 1
        else:
            logger.warning("⚠️ Нет данных оборудования (sections и summary пустые)")

    # Заполняем перечень оборудования
    total_installed_power = 0.0
    total_used_power = 0.0
    total_items_count = 0

    # Определяем структуру таблицы (глобально для функции)
    has_year_column = False
    has_coefficient_column = False
    has_consumption_column = False

    # Если есть данные из sections, заполняем перечень
    if equipment_list_start_row and sections:
        # Заполняем перечень оборудования
        header_row = equipment_list_start_row + 1
        data_start_row = header_row + 1

        # Определяем структуру: либо таблица в строках ниже заголовка, либо создаем новую
        # Проверяем наличие заголовков таблицы
        has_table_headers = False
        for col_idx in range(1, 10):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value and isinstance(cell.value, str) and len(cell.value) > 3:
                has_table_headers = True
                break

        # Определяем структуру заголовков таблицы (может быть 5 или 6 колонок)
        # Проверяем существующие заголовки для определения структуры
        existing_headers = []
        for col_idx in range(1, 7):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                existing_headers.append(cell.value.strip())

        # Определяем, какая структура таблицы используется
        # Вариант 1: Наименование | Тип | Мощность (кВт) | Год установки | Коэффициент использования | Годовое потребление (кВт·ч)
        # Вариант 2: Наименование | Тип | Мощность (кВт) установленная | Мощность (кВт) используемая | Мощность (кВт) потребленная
        has_year_column = any(
            "год" in h.lower() or "year" in h.lower() for h in existing_headers
        )
        has_coefficient_column = any(
            "коэффициент" in h.lower() or "coefficient" in h.lower()
            for h in existing_headers
        )
        has_consumption_column = any(
            "потребление" in h.lower() or "consumption" in h.lower()
            for h in existing_headers
        )

        if not has_table_headers:
            # Создаем заголовки таблицы в зависимости от структуры
            if has_year_column or has_coefficient_column or has_consumption_column:
                # Структура с годом, коэффициентом и потреблением
                headers = [
                    "Наименование",
                    "Тип",
                    "Мощность (кВт)",
                    "Год установки",
                    "Коэффициент использования",
                    "Годовое потребление (кВт·ч)",
                ]
            else:
                # Старая структура
                headers = [
                    "Наименование",
                    "Тип",
                    "Мощность (кВт) установленная",
                    "Мощность (кВт) используемая",
                    "Мощность (кВт) потребленная",
                ]
            for col_idx, header in enumerate(headers, start=1):
                safe_cell_write(ws.cell(row=header_row, column=col_idx), header)

        # Заполняем данные оборудования
        current_row = data_start_row

        for section in sections:
            items = section.get("items") or []
            for item in items:
                name = item.get("name") or ""
                equipment_type = item.get("type") or ""
                unit_power = item.get("unit_power_kw") or 0.0
                quantity = item.get("quantity") or 0
                total_power_item = item.get("total_power_kw") or (
                    unit_power * quantity if quantity else 0.0
                )
                year = item.get("year") or item.get("installation_year") or ""
                usage_factor = (
                    item.get("utilization_factor") or item.get("usage_factor") or 0.8
                )

                if name:  # Заполняем только если есть название
                    # Используем глобальную функцию _safe_set_cell_value для сохранения формул
                    _safe_set_cell_value(ws, current_row, 1, name)
                    _safe_set_cell_value(ws, current_row, 2, equipment_type)

                    # Определяем структуру колонок
                    if (
                        has_year_column
                        or has_coefficient_column
                        or has_consumption_column
                    ):
                        # Структура с годом, коэффициентом и потреблением
                        _safe_set_cell_value(
                            ws, current_row, 3, _format_float(total_power_item)
                        )  # Мощность
                        _safe_set_cell_value(
                            ws, current_row, 4, str(year) if year else ""
                        )  # Год установки

                        # Коэффициент использования для единицы оборудования
                        if total_power_item > 0:
                            if HAS_CALCULATIONS:
                                used_power_item = calculate_equipment_used_power(
                                    total_power_item, usage_factor=usage_factor
                                )
                            else:
                                used_power_item = total_power_item * float(usage_factor)
                            item_coefficient = (
                                min(1.0, used_power_item / total_power_item)
                                if total_power_item > 0
                                else 0.0
                            )
                        else:
                            item_coefficient = 0.0
                            used_power_item = 0.0

                        # Для коэффициента и потребления проверяем формулы - если есть формула, не перезаписываем
                        cell_coeff = ws.cell(row=current_row, column=5)
                        if not (cell_coeff.data_type == "f" and cell_coeff.value):
                            _safe_set_cell_value(
                                ws,
                                current_row,
                                5,
                                _format_float(item_coefficient),
                                preserve_formula=False,
                            )  # Коэффициент использования

                        # Годовое потребление
                        if HAS_CALCULATIONS:
                            annual_consumption = (
                                calculate_annual_consumption_from_power(used_power_item)
                            )
                        else:
                            annual_consumption = (
                                used_power_item * 8760 if used_power_item > 0 else 0.0
                            )
                        cell_consumption = ws.cell(row=current_row, column=6)
                        if not (
                            cell_consumption.data_type == "f" and cell_consumption.value
                        ):
                            _safe_set_cell_value(
                                ws,
                                current_row,
                                6,
                                _format_float(annual_consumption),
                                preserve_formula=False,
                            )  # Годовое потребление
                    else:
                        # Старая структура
                        _safe_set_cell_value(
                            ws, current_row, 3, _format_float(total_power_item)
                        )  # Установленная
                        # Используем централизованную формулу для используемой мощности
                        if HAS_CALCULATIONS:
                            used_power = calculate_equipment_used_power(
                                total_power_item, usage_factor=usage_factor
                            )
                        else:
                            used_power = total_power_item * float(usage_factor)

                        # Проверяем формулы перед записью
                        cell_used = ws.cell(row=current_row, column=4)
                        if not (cell_used.data_type == "f" and cell_used.value):
                            _safe_set_cell_value(
                                ws,
                                current_row,
                                4,
                                _format_float(used_power),
                                preserve_formula=False,
                            )  # Используемая
                        cell_consumed = ws.cell(row=current_row, column=5)
                        if not (cell_consumed.data_type == "f" and cell_consumed.value):
                            _safe_set_cell_value(
                                ws,
                                current_row,
                                5,
                                _format_float(used_power),
                                preserve_formula=False,
                            )  # Потребленная

                    total_installed_power += float(total_power_item)
                    if (
                        has_year_column
                        or has_coefficient_column
                        or has_consumption_column
                    ):
                        if HAS_CALCULATIONS:
                            used_power = calculate_equipment_used_power(
                                total_power_item, usage_factor=usage_factor
                            )
                        else:
                            used_power = total_power_item * float(usage_factor)
                    total_used_power += float(used_power)
                    total_items_count += int(quantity) if quantity else 1
                    current_row += 1

        # Заполняем итоговую строку
        if total_items_count > 0:
            # Ищем строку "ИТОГО" или создаем
            itogo_row = current_row
            # Проверяем, есть ли уже строка "ИТОГО"
            for check_row in range(data_start_row, current_row):
                cell_value = ws.cell(row=check_row, column=1).value
                if (
                    cell_value
                    and isinstance(cell_value, str)
                    and "ИТОГО" in cell_value.upper()
                ):
                    itogo_row = check_row
                    break

            # Устанавливаем "ИТОГО" только если ячейка пустая или не содержит формулу
            cell_itogo = ws.cell(row=itogo_row, column=1)
            if not (cell_itogo.data_type == "f" and cell_itogo.value):
                cell_itogo.value = "ИТОГО"

            # Определяем структуру колонок для итоговой строки
            # ВАЖНО: В итоговой строке могут быть формулы (например, =SUM(...)), их нужно сохранить!
            if has_year_column or has_coefficient_column or has_consumption_column:
                # Структура с годом, коэффициентом и потреблением
                # Проверяем каждую ячейку на наличие формулы
                cell_power = ws.cell(row=itogo_row, column=3)
                if not (cell_power.data_type == "f" and cell_power.value):
                    cell_power.value = _format_float(total_installed_power)  # Мощность

                cell_year = ws.cell(row=itogo_row, column=4)
                if not (cell_year.data_type == "f" and cell_year.value):
                    cell_year.value = ""  # Год установки (пусто для итога)

                # Общий коэффициент использования - проверяем формулу
                cell_coeff = ws.cell(row=itogo_row, column=5)
                if not (cell_coeff.data_type == "f" and cell_coeff.value):
                    if total_installed_power > 0:
                        overall_coefficient = min(
                            1.0, total_used_power / total_installed_power
                        )
                    else:
                        overall_coefficient = 0.0
                    cell_coeff.value = _format_float(
                        overall_coefficient
                    )  # Коэффициент использования

                # Общее годовое потребление - проверяем формулу
                cell_consumption = ws.cell(row=itogo_row, column=6)
                if not (cell_consumption.data_type == "f" and cell_consumption.value):
                    if HAS_CALCULATIONS:
                        total_annual_consumption = (
                            calculate_annual_consumption_from_power(total_used_power)
                        )
                    else:
                        total_annual_consumption = (
                            total_used_power * 8760 if total_used_power > 0 else 0.0
                        )
                    cell_consumption.value = _format_float(
                        total_annual_consumption
                    )  # Годовое потребление
            else:
                # Старая структура - проверяем формулы
                cell_power = ws.cell(row=itogo_row, column=3)
                if not (cell_power.data_type == "f" and cell_power.value):
                    cell_power.value = _format_float(total_installed_power)

                cell_used = ws.cell(row=itogo_row, column=4)
                if not (cell_used.data_type == "f" and cell_used.value):
                    cell_used.value = _format_float(total_used_power)

                cell_consumed = ws.cell(row=itogo_row, column=5)
                if not (cell_consumed.data_type == "f" and cell_consumed.value):
                    cell_consumed.value = _format_float(total_used_power)

            if not efficiency_start_row:
                efficiency_start_row = itogo_row + 3
    # Если sections пусто, но есть summary, используем данные из summary
    if not sections and summary:
        logger.info("⚠️ Sections пусто, используем данные из summary")
        total_items_count = summary.get("total_items", 0)
        total_power_kw = summary.get("total_power_kw", 0.0)
        total_installed_power = total_power_kw
        # Используем централизованную формулу
        if HAS_CALCULATIONS:
            total_used_power = calculate_equipment_used_power(
                total_power_kw, usage_factor=0.8
            )
        else:
            total_used_power = total_power_kw * 0.8

        logger.info(
            f"📊 Данные из summary: items={total_items_count}, power={total_power_kw}, used_power={total_used_power}"
        )

        # Устанавливаем equipment_list_start_row, если он не найден
        if not equipment_list_start_row:
            equipment_list_start_row = 1
            logger.info(
                "🔧 Установлен equipment_list_start_row=1 для заполнения из summary"
            )

        # Заполняем заголовки и итоговую строку, если есть данные
        if equipment_list_start_row and (total_items_count > 0 or total_power_kw > 0):
            header_row = equipment_list_start_row + 1

            # Проверяем существующие заголовки для определения структуры
            existing_headers_summary = []
            for col_idx in range(1, 7):
                cell = ws.cell(row=header_row, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    existing_headers_summary.append(cell.value.strip())

            has_year_col_summary = any(
                "год" in h.lower() or "year" in h.lower()
                for h in existing_headers_summary
            )
            has_coeff_col_summary = any(
                "коэффициент" in h.lower() or "coefficient" in h.lower()
                for h in existing_headers_summary
            )
            has_cons_col_summary = any(
                "потребление" in h.lower() or "consumption" in h.lower()
                for h in existing_headers_summary
            )

            # Создаем заголовки только если их нет
            has_headers = len(existing_headers_summary) > 0
            if not has_headers:
                if (
                    has_year_col_summary
                    or has_coeff_col_summary
                    or has_cons_col_summary
                ):
                    # Структура с годом, коэффициентом и потреблением
                    headers = [
                        "Наименование",
                        "Тип",
                        "Мощность (кВт)",
                        "Год установки",
                        "Коэффициент использования",
                        "Годовое потребление (кВт·ч)",
                    ]
                else:
                    # Старая структура
                    headers = [
                        "Наименование",
                        "Тип",
                        "Мощность (кВт) установленная",
                        "Мощность (кВт) используемая",
                        "Мощность (кВт) потребленная",
                    ]
                for col_idx, header in enumerate(headers, start=1):
                    safe_cell_write(ws.cell(row=header_row, column=col_idx), header)

            # Заполняем итоговую строку
            itogo_row = header_row + 1
            # Устанавливаем "ИТОГО" только если ячейка пустая или не содержит формулу
            cell_itogo = ws.cell(row=itogo_row, column=1)
            if not (cell_itogo.data_type == "f" and cell_itogo.value):
                cell_itogo.value = "ИТОГО"

            # Определяем структуру колонок для итоговой строки (проверяем заголовки)
            header_row_check = (
                equipment_list_start_row + 1 if equipment_list_start_row else header_row
            )
            existing_headers_check = []
            for col_idx in range(1, 7):
                cell = ws.cell(row=header_row_check, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    existing_headers_check.append(cell.value.strip())

            has_year_col = any(
                "год" in h.lower() or "year" in h.lower()
                for h in existing_headers_check
            )
            has_coeff_col = any(
                "коэффициент" in h.lower() or "coefficient" in h.lower()
                for h in existing_headers_check
            )
            has_cons_col = any(
                "потребление" in h.lower() or "consumption" in h.lower()
                for h in existing_headers_check
            )

            if has_year_col or has_coeff_col or has_cons_col:
                # Структура с годом, коэффициентом и потреблением
                # Проверяем формулы перед записью
                cell_power = ws.cell(row=itogo_row, column=3)
                if not (cell_power.data_type == "f" and cell_power.value):
                    cell_power.value = _format_float(total_installed_power)  # Мощность

                cell_year = ws.cell(row=itogo_row, column=4)
                if not (cell_year.data_type == "f" and cell_year.value):
                    cell_year.value = ""  # Год установки (пусто для итога)

                # Общий коэффициент использования - проверяем формулу
                cell_coeff = ws.cell(row=itogo_row, column=5)
                if not (cell_coeff.data_type == "f" and cell_coeff.value):
                    if total_installed_power > 0:
                        overall_coefficient = min(
                            1.0, total_used_power / total_installed_power
                        )
                    else:
                        overall_coefficient = 0.0
                    cell_coeff.value = _format_float(
                        overall_coefficient
                    )  # Коэффициент использования

                # Общее годовое потребление - проверяем формулу
                cell_consumption = ws.cell(row=itogo_row, column=6)
                if not (cell_consumption.data_type == "f" and cell_consumption.value):
                    if HAS_CALCULATIONS:
                        total_annual_consumption = (
                            calculate_annual_consumption_from_power(total_used_power)
                        )
                    else:
                        total_annual_consumption = (
                            total_used_power * 8760 if total_used_power > 0 else 0.0
                        )
                    cell_consumption.value = _format_float(
                        total_annual_consumption
                    )  # Годовое потребление
            else:
                # Старая структура - проверяем формулы
                cell_power = ws.cell(row=itogo_row, column=3)
                if not (cell_power.data_type == "f" and cell_power.value):
                    cell_power.value = _format_float(total_installed_power)

                cell_used = ws.cell(row=itogo_row, column=4)
                if not (cell_used.data_type == "f" and cell_used.value):
                    cell_used.value = _format_float(total_used_power)

                cell_consumed = ws.cell(row=itogo_row, column=5)
                if not (cell_consumed.data_type == "f" and cell_consumed.value):
                    cell_consumed.value = _format_float(total_used_power)

            if not efficiency_start_row:
                efficiency_start_row = itogo_row + 3

    # Используем данные из summary для дополнения, если они более полные
    # ИЛИ если sections пустые, но summary есть
    if summary:
        summary_items = summary.get("total_items", 0)
        summary_power = summary.get("total_power_kw", 0.0)

        # Если sections пустые, используем summary как основной источник
        if not sections and (summary_items > 0 or summary_power > 0):
            logger.info(
                f"📊 Используем summary как основной источник данных: items={summary_items}, power={summary_power}"
            )
            total_items_count = (
                summary_items
                if summary_items > total_items_count
                else total_items_count
            )
            if summary_power > total_installed_power:
                total_installed_power = summary_power
                # Используем централизованную формулу
                if HAS_CALCULATIONS:
                    total_used_power = calculate_equipment_used_power(
                        summary_power, usage_factor=0.8
                    )
                else:
                    total_used_power = summary_power * 0.8
                logger.info(
                    f"✅ Обновлена мощность оборудования из summary: {total_installed_power} кВт, used={total_used_power} кВт"
                )
        else:
            # Дополняем существующие данные
            if summary_items > total_items_count:
                total_items_count = summary_items
                logger.info(
                    f"Обновлено количество единиц оборудования из summary: {total_items_count}"
                )

            if summary_power > total_installed_power:
                total_installed_power = summary_power
                # Используем централизованную формулу
                if HAS_CALCULATIONS:
                    total_used_power = calculate_equipment_used_power(
                        summary_power, usage_factor=0.8
                    )
                else:
                    total_used_power = summary_power * 0.8
                logger.info(
                    f"Обновлена мощность оборудования из summary: {total_installed_power} кВт"
                )

    # Заполняем анализ эффективности оборудования (ВСЕГДА, если есть данные)
    logger.info(
        f"📊 Перед заполнением анализа эффективности: total_items={total_items_count}, total_power={total_installed_power}, total_used={total_used_power}"
    )

    if not efficiency_start_row:
        # Ищем заголовок "АНАЛИЗ ЭФФЕКТИВНОСТИ" в большем диапазоне
        max_check_rows = min(100, ws.max_row + 1) if ws.max_row else 100
        for row_idx in range(1, max_check_rows):
            for col_idx in range(1, 6):
                try:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value and isinstance(cell_value, str):
                        cell_upper = cell_value.upper()
                        if "АНАЛИЗ" in cell_upper and "ЭФФЕКТИВНОСТИ" in cell_upper:
                            efficiency_start_row = row_idx
                            logger.info(
                                f"Найден заголовок анализа эффективности в строке {row_idx}, колонка {col_idx}"
                            )
                            break
                except Exception:
                    continue
            if efficiency_start_row:
                break

    # Если заголовок не найден, но есть данные, создаем структуру
    if not efficiency_start_row and (
        total_items_count > 0 or total_installed_power > 0
    ):
        # Размещаем после перечня оборудования или с начала листа
        if equipment_list_start_row:
            efficiency_start_row = equipment_list_start_row + 20  # После перечня
        else:
            efficiency_start_row = 15  # Начинаем с строки 15
        logger.info(
            f"Заголовок анализа эффективности не найден, создаем в строке {efficiency_start_row}"
        )

    # Заполняем анализ эффективности, если есть данные или структура найдена
    if efficiency_start_row:
        logger.info(f"Заполнение анализа эффективности с строки {efficiency_start_row}")
        # Ищем существующую таблицу анализа или создаем
        analysis_row = efficiency_start_row + 1

        # Проверяем наличие заголовков таблицы
        has_analysis_headers = False
        for col_idx in range(1, 4):
            try:
                cell = ws.cell(row=analysis_row, column=col_idx)
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and "Показатель" in cell.value
                ):
                    has_analysis_headers = True
                    logger.info(
                        f"Найдены заголовки анализа эффективности в строке {analysis_row}"
                    )
                    break
            except Exception:
                continue

        if not has_analysis_headers:
            # Создаем заголовки если их нет
            logger.info(
                f"Создание заголовков анализа эффективности в строке {analysis_row}"
            )
            safe_cell_write(ws.cell(row=analysis_row, column=1), "Показатель")
            safe_cell_write(ws.cell(row=analysis_row, column=2), "Значение")
            safe_cell_write(ws.cell(row=analysis_row, column=3), "Единица измерения")
            analysis_row += 1
        else:
            analysis_row += 1

        # Определяем названия показателей для поиска существующих строк или создания новых
        # Ключи - наши внутренние названия, значения - варианты названий в шаблоне
        indicator_mapping = {
            "Установленная мощность": {
                "search_names": [
                    "Общая установленная мощность",
                    "Установленная мощность",
                    "Установленная",
                    "Общая",
                    "МОЩНОСТЬ",
                    "POWER",
                ],
                "default_name": "Общая установленная мощность",
                "value": total_installed_power,
                "unit": "кВт",
            },
            "Коэффициент использования": {
                "search_names": [
                    "Средний коэффициент использования",
                    "Коэффициент использования",
                    "коэффициент",
                    "КОЭФФИЦИЕНТ",
                    "COEFFICIENT",
                ],
                "default_name": "Средний коэффициент использования",
                "value": 0.0,  # Будет вычислен отдельно
                "unit": "",
            },
            "Годовое потребление": {
                "search_names": [
                    "Общее годовое потребление",
                    "Годовое потребление",
                    "Годовое",
                    "ПОТРЕБЛЕНИЕ",
                    "CONSUMPTION",
                ],
                "default_name": "Общее годовое потребление",
                "value": calculate_annual_consumption_from_power(total_used_power)
                if HAS_CALCULATIONS
                else (total_used_power * 8760 if total_used_power > 0 else 0.0),
                "unit": "кВт·ч",
            },
            "Мощность на единицу": {
                "search_names": [
                    "Средняя мощность на единицу оборудования",
                    "Мощность на единицу",
                    "на единицу",
                    "на едини",
                ],
                "default_name": "Средняя мощность на единицу оборудования",
                "value": calculate_average_power_per_unit(
                    total_installed_power, total_items_count
                )
                if HAS_CALCULATIONS
                else (
                    total_installed_power / total_items_count
                    if total_items_count > 0
                    else 0.0
                ),
                "unit": "кВт",
            },
            "Количество единиц": {
                "search_names": [
                    "Количество единиц оборудования",
                    "единиц оборудования",
                    "Количество единиц",
                    "единиц об",
                ],
                "default_name": "Количество единиц оборудования",
                "value": total_items_count,
                "unit": "шт.",
            },
        }

        # Создаем список для заполнения (сохраняем порядок)
        efficiency_data_order = [
            "Установленная мощность",
            "Коэффициент использования",
            "Годовое потребление",
            "Мощность на единицу",
            "Количество единиц",
        ]

        logger.info(
            f"Заполнение {len(efficiency_data_order)} показателей эффективности. total_items={total_items_count}, total_power={total_installed_power}"
        )

        for idx, indicator_key in enumerate(efficiency_data_order):
            if indicator_key not in indicator_mapping:
                continue

            indicator_info = indicator_mapping[indicator_key]
            default_row = analysis_row + idx

            # Пытаемся найти существующую строку с показателем
            found_row = None
            search_names = indicator_info["search_names"]

            # Ищем в широком диапазоне (до 30 строк после заголовков)
            search_end = min(analysis_row + 30, ws.max_row + 1 if ws.max_row else 100)
            for check_row in range(analysis_row, search_end):
                try:
                    cell_value = ws.cell(row=check_row, column=1).value
                    if cell_value and isinstance(cell_value, str):
                        cell_upper = cell_value.upper().strip()
                        for search_name in search_names:
                            search_upper = search_name.upper().strip()
                            # Проверяем частичное совпадение (слова из поискового названия есть в ячейке)
                            if search_upper in cell_upper or any(
                                word in cell_upper
                                for word in search_upper.split()
                                if len(word) > 3
                            ):
                                found_row = check_row
                                logger.info(
                                    f"Найдена существующая строка для '{indicator_key}' ('{cell_value}') в строке {found_row}"
                                )
                                break
                        if found_row:
                            break
                except Exception as e:
                    logger.debug(f"Ошибка при поиске строки {check_row}: {e}")
                    continue

            if not found_row:
                found_row = default_row
                logger.info(
                    f"Строка для '{indicator_key}' не найдена, используем строку {found_row}"
                )

            # Заполняем показатель
            try:
                # Сохраняем существующее название или устанавливаем стандартное
                existing_name = ws.cell(row=found_row, column=1).value
                if (
                    not existing_name
                    or not isinstance(existing_name, str)
                    or existing_name.strip() == ""
                ):
                    ws.cell(row=found_row, column=1).value = indicator_info[
                        "default_name"
                    ]
                else:
                    logger.debug(
                        f"Используем существующее название показателя: '{existing_name}' в строке {found_row}"
                    )

                # Заполняем значение
                # ВАЖНО: Проверяем, есть ли формула в ячейке значения - если есть, не перезаписываем!
                cell_value = ws.cell(row=found_row, column=2)
                has_formula = cell_value.data_type == "f" and cell_value.value

                if has_formula:
                    logger.info(
                        f"⚠️ Пропущена ячейка {cell_value.coordinate} с формулой: {cell_value.value}"
                    )
                    # Не перезаписываем формулу, но обновляем единицу измерения если нужно
                    unit = indicator_info["unit"]
                    if unit:
                        existing_unit = ws.cell(row=found_row, column=3).value
                        if not existing_unit or existing_unit == "":
                            safe_cell_write(ws.cell(row=found_row, column=3), unit)
                    continue

                if indicator_key == "Коэффициент использования":
                    # Используем централизованную формулу
                    if HAS_CALCULATIONS:
                        if total_installed_power > 0:
                            coefficient = calculate_equipment_usage_coefficient(
                                used_power_kw=total_used_power,
                                installed_power_kw=total_installed_power,
                            )
                        else:
                            coefficient = 0.0
                            logger.warning(
                                f"Коэффициент использования = 0 из-за нулевой установленной мощности (строка {found_row})"
                            )
                    else:
                        # Fallback на старую логику с проверкой деления на ноль
                        if total_installed_power > 0:
                            coefficient = min(
                                1.0, total_used_power / total_installed_power
                            )
                        else:
                            coefficient = 0.0
                            logger.warning(
                                f"Коэффициент использования = 0 из-за нулевой установленной мощности (строка {found_row})"
                            )

                    # Записываем значение, избегая ошибки #ДЕЛ/0!
                    if coefficient > 0:
                        cell_value.value = _format_float(coefficient)
                    else:
                        cell_value.value = 0.0
                    logger.info(
                        f"Заполнен коэффициент использования: {coefficient} (строка {found_row}, power={total_installed_power}, used={total_used_power})"
                    )
                else:
                    value = indicator_info["value"]
                    formatted_value = (
                        _format_float(value)
                        if isinstance(value, (int, float))
                        else value
                    )
                    cell_value.value = formatted_value
                    logger.info(
                        f"Заполнен показатель '{indicator_key}': {formatted_value} (строка {found_row})"
                    )

                # Заполняем единицу измерения
                unit = indicator_info["unit"]
                if unit:
                    existing_unit = ws.cell(row=found_row, column=3).value
                    if not existing_unit or existing_unit == "":
                        safe_cell_write(ws.cell(row=found_row, column=3), unit)
                        logger.debug(
                            f"Установлена единица измерения '{unit}' для '{indicator_key}' в строке {found_row}"
                        )
            except Exception as e:
                logger.error(
                    f"Ошибка при заполнении показателя '{indicator_key}' в строке {found_row}: {e}",
                    exc_info=True,
                )
    else:
        logger.warning(
            "⚠️ Не найдена структура для анализа эффективности и нет данных для создания"
        )
        # Даже если структура не найдена, но есть данные, создаем показатели
        if total_items_count > 0 or total_installed_power > 0:
            logger.info(
                "🔧 Создаем структуру анализа эффективности, так как есть данные"
            )
            # Определяем, где разместить таблицу
            if equipment_list_start_row:
                efficiency_start_row = equipment_list_start_row + 20
            else:
                efficiency_start_row = 15

            analysis_row = efficiency_start_row + 1
            # Создаем заголовки
            safe_cell_write(ws.cell(row=analysis_row, column=1), "Показатель")
            safe_cell_write(ws.cell(row=analysis_row, column=2), "Значение")
            safe_cell_write(ws.cell(row=analysis_row, column=3), "Единица измерения")
            analysis_row += 1

            # Заполняем показатели напрямую
            indicators = [
                ("Общая установленная мощность", total_installed_power, "кВт"),
                (
                    "Средний коэффициент использования",
                    (
                        min(1.0, total_used_power / total_installed_power)
                        if total_installed_power > 0
                        else 0.0
                    ),
                    "",
                ),
                (
                    "Общее годовое потребление",
                    (
                        calculate_annual_consumption_from_power(total_used_power)
                        if HAS_CALCULATIONS
                        else (total_used_power * 8760 if total_used_power > 0 else 0.0)
                    ),
                    "кВт·ч",
                ),
                (
                    "Средняя мощность на единицу оборудования",
                    (
                        total_installed_power / total_items_count
                        if total_items_count > 0
                        else 0.0
                    ),
                    "кВт",
                ),
                ("Количество единиц оборудования", total_items_count, "шт."),
            ]

            for idx, (name, value, unit) in enumerate(indicators):
                row = analysis_row + idx
                safe_cell_write(ws.cell(row=row, column=1), name)
                if isinstance(value, (int, float)):
                    safe_cell_write(ws.cell(row=row, column=2), _format_float(value))
                else:
                    safe_cell_write(ws.cell(row=row, column=2), value)
                safe_cell_write(ws.cell(row=row, column=3), unit)
                logger.info(f"✅ Заполнен показатель '{name}': {value} {unit}")

    _auto_fit_columns(ws, num_columns=9)


def fill_balans_sheet(ws, agg_data: Dict, usage_data: Optional[Dict] = None) -> None:
    """
    Fill energy balance sheet with consumption by usage categories.

    Expected structure:
    - Row 1: Title "Энергетический баланс"
    - Row 2: Headers (Квартал, Технологические, Собственные нужды, Производственные, Хоз-бытовые, Итого)
    - Rows 3+: Data rows
    """
    ws.protection = SheetProtection(sheet=False, password=None)

    # НЕ очищаем лист полностью - сохраняем формулы!
    # Очищаем только ячейки без формул в области данных (начиная со строки 3)
    # Это позволяет сохранить формулы в шаблоне
    logger.info(
        "Сохранение формул в листе 'Баланс' - очищаем только ячейки данных без формул"
    )

    # Headers
    # Проверяем, не является ли ячейка объединенной перед записью
    from openpyxl.cell.cell import MergedCell


def safe_cell_write(cell, value):
    """
    Безопасная запись в ячейку Excel с поддержкой merged cells.
    
    Args:
        cell: ячейка Excel
        value: значение для записи
    
    Returns:
        bool: True если запись успешна, False если ячейка объединенная
    """
    try:
        # Проверяем, является ли ячейка объединенной
        if isinstance(cell, MergedCell):
            logger.debug(
                f"Пропущена объединенная ячейка {cell.coordinate} при записи значения: {value}"
            )
            return False
        
        # Проверяем, не находится ли ячейка в объединенном диапазоне
        if hasattr(cell, 'coordinate') and cell.coordinate in cell.parent.merged_cells:
            # Находим parent cell для объединенного диапазона
            for merged_range in cell.parent.merged_cells.ranges:
                if cell.coordinate in str(merged_range):
                    # Получаем первую ячейку диапазона (parent cell)
                    start_coord = merged_range.coord.split(':')[0]
                    parent_cell = cell.parent[start_coord]
                    parent_cell.value = value
                    logger.debug(
                        f"Записано значение {value} в parent cell {start_coord} "
                        f"объединенного диапазона {merged_range.coord}"
                    )
                    return True
            return False
        else:
            # Обычная ячейка - записываем напрямую
            cell.value = value
            return True
            
    except AttributeError as e:
        logger.debug(f"Не удалось записать значение в {cell.coordinate}: {e}")
        return False
    except Exception as e:
        logger.warning(f"Неожиданная ошибка при записи в {cell.coordinate}: {e}")
        return False

    header_cell = ws.cell(row=1, column=1)
    if not isinstance(header_cell, MergedCell):
        try:
            header_cell.value = "Энергетический баланс по категориям потребления"
        except AttributeError:
            logger.debug(
                f"Не удалось записать заголовок в ячейку A1 (возможно, объединенная)"
            )

    headers = [
        "Квартал",
        "Технологические, кВт·ч",
        "Собственные нужды, кВт·ч",
        "Производственные, кВт·ч",
        "Хоз-бытовые, кВт·ч",
        "Итого, кВт·ч",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        # Проверяем, не является ли ячейка частью объединенных ячеек
        # Если ячейка в объединенном диапазоне, пропускаем запись
        if isinstance(cell, MergedCell):
            logger.debug(
                f"Пропущена объединенная ячейка {cell.coordinate} при записи заголовка"
            )
            continue
        try:
            # Пробуем записать значение
            cell.value = header
        except AttributeError as e:
            # Если ячейка объединенная (MergedCell), пропускаем
            logger.debug(f"Не удалось записать заголовок в {cell.coordinate}: {e}")
            continue

    debug_mode = os.getenv("INGEST_PASSPORT_DEBUG", "false").lower() == "true"

    # Поддержка двух форматов данных:
    # 1. agg_data = {"resources": {"electricity": {...}}}
    # 2. agg_data = {"electricity": {...}} (прямой доступ)
    resources = agg_data.get("resources", {})
    if not resources:
        # Пробуем прямой доступ к electricity
        resources = agg_data

    electricity = resources.get("electricity", {})

    if not electricity:
        logger.warning(
            "Лист 'Баланс': отсутствуют данные по электроэнергии (resources.electricity пуст)"
        )
        logger.debug(f"Структура agg_data: keys={list(agg_data.keys())[:10]}")
        if "resources" in agg_data:
            logger.debug(
                f"Структура resources: keys={list(agg_data['resources'].keys())[:10]}"
            )
        # Если данных нет, не создаем формулы - это нормально
        return

    # Collect all quarters and sort
    all_quarters = sorted(electricity.keys())

    if not all_quarters:
        logger.warning(
            "Лист 'Баланс': не найдено ни одного квартала данных по электроэнергии"
        )
        return

    # Константы для колонок
    COL_QUARTER = 1  # A
    COL_TECH = 2  # B - Технологические
    COL_OWN = 3  # C - Собственные нужды
    COL_PROD = 4  # D - Производственные
    COL_HOUSE = 5  # E - Хоз-бытовые
    COL_TOTAL = 6  # F - Итого

    # Первая строка данных (после заголовков)
    data_start_row = 3
    current_row = data_start_row

    for quarter in all_quarters:
        quarter_data = electricity.get(quarter, {})
        by_usage = quarter_data.get("by_usage", {})

        if not by_usage:
            logger.warning(
                f"Лист 'Баланс', квартал {quarter}: отсутствуют данные по категориям потребления (by_usage)"
            )
            # Fallback: try to get from usage_data if provided
            if usage_data and "years" in usage_data:
                year = quarter_data.get("year")
                if year and str(year) in usage_data["years"]:
                    year_cats = usage_data["years"][str(year)]
                    quarter_total = quarter_data.get("quarter_totals", {}).get(
                        "active_kwh", 0
                    )
                    yearly_total = sum(year_cats.values())
                    if yearly_total > 0:
                        # Используем централизованную функцию распределения
                        if HAS_CALCULATIONS:
                            from energy_passport_calculations import (
                                distribute_quarter_by_usage_categories,
                            )

                            by_usage = distribute_quarter_by_usage_categories(
                                quarter_total_kwh=quarter_total,
                                yearly_categories=year_cats,
                            )
                        else:
                            # Fallback на старую логику
                            by_usage = {
                                cat: (quarter_total * val / yearly_total)
                                if yearly_total
                                else 0.0
                                for cat, val in year_cats.items()
                            }

        # Fill row: quarter, tech, own, prod, house
        # Используем безопасную запись с сохранением формул (но не для колонки F - там будет формула)
        _safe_set_cell_value(ws, current_row, COL_QUARTER, quarter)
        _safe_set_cell_value(
            ws, current_row, COL_TECH, _format_float(by_usage.get("technological", 0))
        )
        _safe_set_cell_value(
            ws, current_row, COL_OWN, _format_float(by_usage.get("own_needs", 0))
        )
        _safe_set_cell_value(
            ws, current_row, COL_PROD, _format_float(by_usage.get("production", 0))
        )
        _safe_set_cell_value(
            ws, current_row, COL_HOUSE, _format_float(by_usage.get("household", 0))
        )

        # Колонка F (Итого) - ВСЕГДА создаем формулу программно
        # Используем централизованную функцию для валидации перед созданием формулы
        tech_value = by_usage.get("technological", 0)
        own_value = by_usage.get("own_needs", 0)
        prod_value = by_usage.get("production", 0)
        house_value = by_usage.get("household", 0)
        
        # Валидация через централизованную функцию (если доступна)
        if HAS_CALCULATIONS:
            try:
                calculated_total = calculate_balance_total(
                    technological=tech_value,
                    own_needs=own_value,
                    production=prod_value,
                    household=house_value,
                )
                if debug_mode:
                    logger.debug(
                        f"Валидация баланса для {quarter}: рассчитано={calculated_total}, "
                        f"тех={tech_value}, собств={own_value}, произв={prod_value}, хоз={house_value}"
                    )
            except Exception as calc_exc:
                logger.warning(
                    f"Ошибка при валидации баланса через calculate_balance_total: {calc_exc}"
                )
        
        # Создаём формулу Excel
        total_cell = ws.cell(row=current_row, column=COL_TOTAL)
        if not isinstance(total_cell, MergedCell):
            col_b_letter = get_column_letter(COL_TECH)
            col_e_letter = get_column_letter(COL_HOUSE)
            total_cell.value = (
                f"=SUM({col_b_letter}{current_row}:{col_e_letter}{current_row})"
            )
            logger.debug(
                f"Создана формула в {total_cell.coordinate}: {total_cell.value}"
            )

        current_row += 1

    # data_end_row - последняя строка с данными
    data_end_row = current_row - 1

    # Проверяем, что есть данные для обработки
    if data_end_row >= data_start_row:
        logger.info(
            f"Создание формул для листа 'Баланс': строки данных {data_start_row}-{data_end_row}"
        )

        # 1. Убеждаемся, что все строки данных имеют формулы в колонке F
        # (на случай, если что-то пропустили в цикле выше)
        for r in range(data_start_row, data_end_row + 1):
            total_cell = ws.cell(row=r, column=COL_TOTAL)
            if not isinstance(total_cell, MergedCell):
                # Если формула отсутствует или это значение, создаем/обновляем формулу
                if total_cell.data_type != "f" or not total_cell.value:
                    col_b_letter = get_column_letter(COL_TECH)
                    col_e_letter = get_column_letter(COL_HOUSE)
                    total_cell.value = f"=SUM({col_b_letter}{r}:{col_e_letter}{r})"
                    logger.debug(
                        f"Обновлена формула в {total_cell.coordinate}: {total_cell.value}"
                    )

        # 2. Добавляем строку итогов после данных
        totals_row = data_end_row + 1

        # Записываем подпись "ИТОГО" в колонку A (если не MergedCell)
        totals_label_cell = ws.cell(row=totals_row, column=COL_QUARTER)
        if not isinstance(totals_label_cell, MergedCell):
            totals_label_cell.value = "ИТОГО"

        # Создаем формулы суммирования по столбцам
        for col in [COL_TECH, COL_OWN, COL_PROD, COL_HOUSE, COL_TOTAL]:
            col_letter = get_column_letter(col)
            totals_cell = ws.cell(row=totals_row, column=col)

            if not isinstance(totals_cell, MergedCell):
                # Суммируем значения из соответствующей колонки для всех строк данных
                totals_cell.value = (
                    f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                )
                logger.debug(
                    f"Создана формула итогов в {totals_cell.coordinate}: {totals_cell.value}"
                )

        logger.info(
            f"✅ Создана строка итогов в строке {totals_row} с формулами суммирования"
        )
    else:
        logger.warning("Лист 'Баланс': нет данных для создания формул")

    _auto_fit_columns(ws, num_columns=len(headers))



def safe_cell_write_assign(cell, value):
    """
    Безопасная запись значения в ячейку для сложных присваиваний.
    Используется когда значение вычисляется на нескольких строках.
    """
    return safe_cell_write(cell, value)

def fill_dinamika_sheet(ws, agg_data: Dict) -> None:
    """
    Fill dynamics sheet with quarterly consumption and specific consumption indicators.

    Expected structure:
    - Row 1: Headers (Год, Квартал, Электроэнергия, Газ, Вода, Производство, Удельный расход)
    - Rows 2+: Data rows
    """
    ws.protection = SheetProtection(sheet=False, password=None)

    # Clear sheet if it exists (skip merged cells)
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value = None
            except AttributeError:
                # MergedCell objects are read-only, skip them
                pass

    # Headers

    headers = [
        "Год",
        "Квартал",
        "Электроэнергия, кВт·ч",
        "Газ, м³",
        "Вода, м³",
        "Производство, кг",
        "Удельный расход, кВт·ч/кг",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if isinstance(cell, MergedCell):
            logger.debug(
                f"Пропущена объединенная ячейка {cell.coordinate} при записи заголовка динамики"
            )
            continue
        try:
            cell.value = header
        except AttributeError:
            logger.debug(f"Не удалось записать заголовок в {cell.coordinate}")
            continue

    resources = agg_data.get("resources", {})
    electricity = resources.get("electricity", {})
    gas = resources.get("gas", {})
    water = resources.get("water", {})
    production = resources.get("production", {})

    # Collect all quarters and sort
    all_quarters = sorted(
        set(
            list(electricity.keys())
            + list(gas.keys())
            + list(water.keys())
            + list(production.keys())
        )
    )

    current_row = 2
    for quarter in all_quarters:
        year = quarter.split("-")[0] if "-" in quarter else None
        quarter_num = quarter.split("-Q")[1] if "-Q" in quarter else None

        elec_totals = electricity.get(quarter, {}).get("quarter_totals", {})
        gas_totals = gas.get(quarter, {}).get("quarter_totals", {})
        water_totals = water.get(quarter, {}).get("quarter_totals", {})
        prod_totals = production.get(quarter, {}).get("quarter_totals", {})

        active_kwh = elec_totals.get("active_kwh", 0) or 0
        volume_m3_gas = gas_totals.get("volume_m3", 0) or 0
        volume_m3_water = water_totals.get("volume_m3", 0) or 0

        # Sum all production values
        production_kg = 0
        if prod_totals:
            production_kg = sum(
                v for v in prod_totals.values() if isinstance(v, (int, float))
            )

        safe_cell_write_assign(ws.cell(row=current_row, column=1), (
            int(year) if year and year.isdigit() else None
        ))
        safe_cell_write_assign(ws.cell(row=current_row, column=2), (
            f"Q{quarter_num}" if quarter_num else quarter
        ))
        safe_cell_write(ws.cell(row=current_row, column=3), _format_float(active_kwh))
        safe_cell_write(ws.cell(row=current_row, column=4), _format_float(volume_m3_gas))
        safe_cell_write(ws.cell(row=current_row, column=5), _format_float(volume_m3_water))
        safe_cell_write(ws.cell(row=current_row, column=6), _format_float(production_kg))

        # Используем централизованную формулу для удельного расхода
        if HAS_CALCULATIONS:
            specific = calculate_specific_consumption(
                energy_kwh=active_kwh, production_kg=production_kg, default_on_zero=0.0
            )
            specific_cell = ws.cell(row=current_row, column=7)
            safe_cell_write(specific_cell, _format_float(specific))
        else:
            # Fallback на формулу Excel
            if production_kg > 0:
                col_elec = get_column_letter(3)
                col_prod = get_column_letter(6)
                specific_cell = ws.cell(
                    row=current_row, column=7
                )
                safe_cell_write(specific_cell, f"=IF({col_prod}{current_row}>0,{col_elec}{current_row}/{col_prod}{current_row},0)")
            else:
                specific_cell = ws.cell(row=current_row, column=7)
                safe_cell_write(specific_cell, 0)
                specific = 0.0

        # Проверка соответствия нормативу (если значение вычислено)
        if HAS_CALCULATIONS and specific > 0:
            try:
                # Импортируем модуль проверки нормативов
                _ingest_path = (
                    Path(__file__).parent.parent
                    / "eaip_full_skeleton"
                    / "services"
                    / "ingest"
                    / "domain"
                )
                if str(_ingest_path) not in sys.path:
                    sys.path.insert(0, str(_ingest_path))
                
                from normative_integration import validate_and_log_critical_field
                
                # Проверяем удельный расход на соответствие нормативу
                validation_result = validate_and_log_critical_field(
                    field_name="Удельный расход",
                    actual_value=specific,
                    sheet_name="Динамика ср",
                    cell=f"G{current_row}",
                )
                if validation_result:
                    logger.debug(f"Валидация удельного расхода для {quarter}: {validation_result}")
            except ImportError:
                # Модуль нормативов не доступен - пропускаем проверку
                pass
            except Exception as e:
                logger.warning(f"Ошибка при проверке норматива для {quarter}: {e}")

        current_row += 1

    _auto_fit_columns(ws, num_columns=len(headers))


# Вспомогательные функции
def _safe_set_cell_value(ws, row: int, col: int, value, preserve_formula: bool = True) -> None:
    """Безопасная запись значения в ячейку с сохранением формул."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    if preserve_formula and cell.data_type == 'f' and cell.value:
        return
    try:
        cell.value = value
    except AttributeError:
        pass


def _format_float(value, decimals: int = 3) -> float:
    """Форматирование числа с заданным количеством знаков после запятой."""
    if value is None:
        return 0.0
    try:
        return round(float(value), decimals)
    except (ValueError, TypeError):
        return 0.0


def _reset_sheet(ws):
    """Очистка листа с сохранением структуры."""
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value = None
            except AttributeError:
                pass
    return ws


def _build_default_rows(nodes_list: List[Dict]) -> List[List[Any]]:
    """Построение строк данных для узлов учета."""
    rows = []
    for node in nodes_list:
        row = [
            node.get('name', ''),
            node.get('active_power', ''),
            node.get('reactive_power', ''),
            node.get('location', ''),
            node.get('coefficient', ''),
            node.get('k', ''),
            node.get('seal_date', ''),
            node.get('supplier_seal_date', ''),
            node.get('notes', ''),
        ]
        rows.append(row)
    return rows


def _write_nodes_table(ws, header_rows, nodes_rows, start_row: int, column_count: int) -> int:
    """Запись таблицы узлов учета."""
    current_row = start_row
    for header_row in header_rows:
        for col_idx, value in enumerate(header_row, start=1):
            if col_idx <= column_count:
                safe_cell_write(ws.cell(row=current_row, column=col_idx), value)
        current_row += 1
    for data_row in nodes_rows:
        for col_idx, value in enumerate(data_row, start=1):
            if col_idx <= column_count:
                safe_cell_write(ws.cell(row=current_row, column=col_idx), value)
        current_row += 1
    return current_row


def _auto_fit_columns(ws, num_columns: int = 10) -> None:
    """Автоподбор ширины колонок."""
    for col_idx in range(1, num_columns + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def fill_fuel_dynamics_sheet(ws, agg_data: Dict) -> None:
    """
    Fill fuel dynamics sheet (мазут, уголь, нефтепродукты).
    Placeholder - реализовать при необходимости.
    """
    logger.info("fill_fuel_dynamics_sheet: placeholder - не реализовано")
    pass


def fill_specific_consumption_sheet(ws, agg_data: Dict) -> None:
    """
    Fill specific consumption sheet (расход на единицу продукции).
    Placeholder - реализовать при необходимости.
    """
    logger.info("fill_specific_consumption_sheet: placeholder - не реализовано")
    pass


def fill_meropriyatiya_sheet(ws, measures_data: Optional[List[Dict]] = None) -> None:
    """
    Fill measures sheet (мероприятия по энергосбережению).
    Placeholder - реализовать при необходимости.
    """
    logger.info("fill_meropriyatiya_sheet: placeholder - не реализовано")
    pass


def fill_monthly_sheet(ws, agg_data: Dict, resource_type: str = "electricity") -> None:
    """
    Fill monthly consumption sheet.
    Placeholder - реализовать при необходимости.
    """
    logger.info(f"fill_monthly_sheet: placeholder для {resource_type} - не реализовано")
    pass


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Fill energy passport template")
    parser.add_argument("--template", required=True, help="Path to template file")
    parser.add_argument("--output", required=True, help="Path to output file")
    parser.add_argument("--data", required=True, help="Path to JSON data file")
    args = parser.parse_args()
    print(f"Template: {args.template}")
    print(f"Output: {args.output}")
    print(f"Data: {args.data}")
