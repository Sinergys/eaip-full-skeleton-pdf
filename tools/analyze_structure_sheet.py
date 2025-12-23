"""Анализ структуры листа 'Структура пр 2' для понимания расположения данных."""
from openpyxl import load_workbook
from pathlib import Path
import sys

# Добавляем путь к templates_config
templates_config_path = Path(__file__).parent.parent / "templates" / "pcm690"
if str(templates_config_path) not in sys.path:
    sys.path.insert(0, str(templates_config_path))
from templates_config import get_template_path

template_path = get_template_path("new_energy_passport")
wb = load_workbook(template_path, data_only=False)

sheet_name = "Структура пр 2 "
if sheet_name not in wb.sheetnames:
    # Попробуем найти похожий лист
    for name in wb.sheetnames:
        if "структура" in name.lower():
            sheet_name = name
            break

ws = wb[sheet_name]

print(f"Анализ листа: {sheet_name}")
print("=" * 80)
print(f"Размер: {ws.max_row} строк × {ws.max_column} столбцов")
print("\nПервые 20 строк с данными:")

for row_idx in range(1, min(21, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(16, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            value = str(cell.value)[:50]
            row_data.append(f"{cell.coordinate}:{value}")
    
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data)}")

print("\n\nПоиск заголовков таблицы:")
# Ищем строки с заголовками (обычно содержат "Электроэнергия", "Газ", "Вода" и т.д.)
keywords = ["Электроэнергия", "Газ", "Вода", "Тепловая", "Структура", "параметров"]
for row_idx in range(1, min(15, ws.max_row + 1)):
    row_text = " ".join([str(ws.cell(row=row_idx, column=col).value or "") 
                         for col in range(1, min(16, ws.max_column + 1))])
    if any(keyword in row_text for keyword in keywords):
        print(f"\nRow {row_idx}:")
        for col_idx in range(1, min(16, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                print(f"  {cell.coordinate}: {str(cell.value)[:80]}")

