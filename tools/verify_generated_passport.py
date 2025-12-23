"""Проверка сгенерированного энергопаспорта"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")

print("=" * 80)
print("ПРОВЕРКА СГЕНЕРИРОВАННОГО ЭНЕРГОПАСПОРТА")
print("=" * 80)
print(f"Файл: {output_path.name}\n")

if not output_path.exists():
    print(f"❌ Файл не найден: {output_path}")
    exit(1)

wb = load_workbook(output_path, data_only=True)
ws = wb["Структура пр 2"]

print("✅ Файл загружен\n")

# Проверка E32 (2023 Q1 газ)
print("=" * 80)
print("ПРОВЕРКА ИСПРАВЛЕНИЯ E32")
print("=" * 80)
e32_value = ws["E32"].value
print(f"E32 (2023 Q1 газ): {e32_value}")
if e32_value and abs(e32_value - 14.819) < 0.001:
    print("✅ ОШИБКА ИСПРАВЛЕНА! Значение правильное: 14.819 тыс. м³")
else:
    print(f"⚠️ Значение: {e32_value} (ожидается 14.819)")

# Проверка категорий газа для 2023 Q1
print(f"\nE34 (собственные нужды 2023 Q1): {ws['E34'].value}")
print(f"E36 (хоз-быт 2023 Q1): {ws['E36'].value}")

# Проверка электроэнергии по продукции
print("\n" + "=" * 80)
print("ПРОВЕРКА ЭЛЕКТРОЭНЕРГИИ ПО ВИДАМ ПРОДУКЦИИ")
print("=" * 80)
product_rows = {
    17: "Трубы ХВС",
    18: "Фитинги ХВС",
    19: "Канализационные трубы",
    20: "Канализационные фитинги",
}

for row, product_name in product_rows.items():
    norm = ws.cell(row, 2).value
    fact_2022 = ws.cell(row, 3).value
    fact_2023 = ws.cell(row, 4).value
    fact_2024 = ws.cell(row, 5).value
    print(f"\n{product_name} (строка {row}):")
    print(f"  Норма: {norm}")
    print(f"  Факт 2022: {fact_2022}")
    print(f"  Факт 2023: {fact_2023}")
    print(f"  Факт 2024: {fact_2024}")

# Проверка удельного расхода газа
print("\n" + "=" * 80)
print("ПРОВЕРКА УДЕЛЬНОГО РАСХОДА ГАЗА")
print("=" * 80)
if "Удельный расход газа" in wb.sheetnames:
    ws_gas = wb["Удельный расход газа"]
    fact_per_m2 = ws_gas["C5"].value
    print(f"Удельный расход на м²: {fact_per_m2} м³/(м²·год)")
    if fact_per_m2:
        print("✅ Рассчитан")
else:
    print("⚠️ Лист 'Удельный расход газа' не найден")

# Проверка других кварталов газа в строке 32
print("\n" + "=" * 80)
print("ПРОВЕРКА ГАЗА ПО КВАРТАЛАМ (строка 32)")
print("=" * 80)
gas_quarters = {
    "2022-Q1": "F32",
    "2022-Q2": "V32",
    "2023-Q1": "E32",
    "2023-Q2": "T32",
    "2023-Q3": "AI32",
    "2023-Q4": "AX32",
    "2024-Q1": "BR32",
}

for quarter, cell_coord in gas_quarters.items():
    value = ws[cell_coord].value
    if value:
        print(f"{quarter}: {cell_coord} = {value} тыс. м³")

wb.close()

print("\n" + "=" * 80)
print("✅ ПРОВЕРКА ЗАВЕРШЕНА")
print("=" * 80)

