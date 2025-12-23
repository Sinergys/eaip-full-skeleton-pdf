"""
Проверка сгенерированного паспорта
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

def check_passport(passport_path: Path):
    """Проверяет структуру сгенерированного паспорта"""
    
    if not passport_path.exists():
        print(f"❌ Файл не найден: {passport_path}")
        return
    
    print("=" * 80)
    print(f"ПРОВЕРКА ПАСПОРТА: {passport_path.name}")
    print("=" * 80)
    
    try:
        wb = load_workbook(passport_path, data_only=False)
        
        print(f"\n📊 Всего листов: {len(wb.sheetnames)}")
        print("\n📋 Список листов:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            formulas = sum(1 for row in ws.iter_rows() for cell in row if cell.data_type == 'f' and cell.value)
            data_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
            print(f"  {i}. '{sheet_name}'")
            print(f"     - Строк: {ws.max_row}, Колонок: {ws.max_column}")
            print(f"     - Формул: {formulas}, Ячеек с данными: {data_cells}")
        
        # Проверка ключевых листов
        print("\n" + "=" * 80)
        print("ПРОВЕРКА КЛЮЧЕВЫХ ЛИСТОВ")
        print("=" * 80)
        
        key_sheets = {
            "Структура": ["Структура пр 2", "Struktura pr2", "02_Структура"],
            "Узлы учета": ["01_Узлы учета", "Узел учета", "Узлы учета", "Nodes"],
            "Баланс": ["04_Баланс", "Баланс", "Balance"],
            "Оборудование": ["03_Оборудование", "Equipment", "Оборудование", "Sheet1"],
            "Динамика": ["05_Динамика", "Динамика", "Dinamika sr", "Динамика ср"],
            "Мероприятия": ["06_Мероприятия", "Мероприятия", "Meropriyatiya", "Мериаприятия 1"]
        }
        
        found_count = 0
        for category, names in key_sheets.items():
            found = None
            for name in names:
                # Проверяем точное совпадение и варианты с пробелами
                for sheet_name in wb.sheetnames:
                    if sheet_name.strip() == name.strip() or sheet_name == name:
                        found = sheet_name
                        break
                if found:
                    break
            status = "✅" if found else "❌"
            print(f"{status} {category}: {found if found else 'НЕ НАЙДЕН'}")
            if found:
                found_count += 1
        
        print(f"\n📊 Найдено ключевых листов: {found_count}/{len(key_sheets)}")
        
        wb.close()
        
        return found_count == len(key_sheets)
        
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) > 1:
        passport_path = Path(sys.argv[1])
    else:
        # Ищем последний сгенерированный паспорт
        output_dir = Path("/tmp/passports")
        if not output_dir.exists():
            print("❌ Директория /tmp/passports не найдена")
            sys.exit(1)
        
        passports = sorted(output_dir.glob("*_energy_passport.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not passports:
            print("❌ Не найдено сгенерированных паспортов")
            sys.exit(1)
        
        passport_path = passports[0]
        print(f"Используется последний паспорт: {passport_path.name}\n")
    
    check_passport(passport_path)

