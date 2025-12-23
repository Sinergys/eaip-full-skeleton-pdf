"""
Проверка всех доступных шаблонов
"""
from pathlib import Path
from openpyxl import load_workbook

TEMPLATES_DIR = Path(__file__).parent.parent / "templates" / "pcm690"

templates_to_check = [
    "energy_passport_template.xlsx",
    "new_energy_passport.xlsx",
    "template_metin.xlsx"
]

print("=" * 80)
print("ПРОВЕРКА ВСЕХ ШАБЛОНОВ")
print("=" * 80)

for template_name in templates_to_check:
    template_path = TEMPLATES_DIR / template_name
    if not template_path.exists():
        print(f"\n❌ {template_name} - не найден")
        continue
    
    try:
        wb = load_workbook(template_path, data_only=False)
        print(f"\n📄 {template_name}")
        print(f"   Листов: {len(wb.sheetnames)}")
        print(f"   Названия листов:")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            formulas = sum(1 for row in ws.iter_rows() for cell in row if cell.data_type == 'f' and cell.value)
            print(f"     - {sheet_name} ({ws.max_row} строк, {formulas} формул)")
        wb.close()
    except Exception as e:
        print(f"\n❌ {template_name} - ошибка: {e}")

