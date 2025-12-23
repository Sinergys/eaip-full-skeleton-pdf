"""
Исправление ссылок на листы в формулах
Проблема: формулы ссылаются на 'Узел учета' и 'Структура пр 2', 
а реальные листы называются 'Узел учета ' и 'Структура пр 2 ' (с пробелами в конце)
"""
from openpyxl import load_workbook
from pathlib import Path
import re
import shutil
from collections import defaultdict

template_path = Path(r"C:\eaip\data\source_files\audit_sinergys\ЭНЕРГО_ПАСПОРТ_ 1.xlsm")
output_path = Path(r"C:\eaip\data\source_files\audit_sinergys\ЭНЕРГО_ПАСПОРТ_ 1_FIXED.xlsm")

if not template_path.exists():
    print(f"❌ Файл не найден: {template_path}")
    exit(1)

print(f"📋 Исправление ссылок в формулах")
print("=" * 80)
print(f"📂 Исходный файл: {template_path.name}")
print(f"📂 Выходной файл: {output_path.name}\n")

# Копируем файл
shutil.copy2(template_path, output_path)
print(f"✅ Файл скопирован\n")

wb = load_workbook(output_path, keep_vba=True)  # Сохраняем макросы для .xlsm
sheet_names = wb.sheetnames

# Маппинг неправильных имен на правильные
SHEET_NAME_FIXES = {
    'Узел учета': 'Узел учета ',
    'Структура пр 2': 'Структура пр 2 ',
}

print(f"📊 Всего листов: {len(sheet_names)}")
print(f"   Реальные листы: {sheet_names}\n")

total_fixes = 0
fixes_by_sheet = defaultdict(int)

# Проверяем все листы (кроме Sheet1)
for sheet_name in sheet_names:
    if sheet_name == "Sheet1":
        continue
    
    ws = wb[sheet_name]
    sheet_fixes = 0
    
    print(f"📄 Лист: '{sheet_name}'")
    
    # Проверяем все ячейки с формулами
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Формула
                formula_str = str(cell.value)
                original_formula = formula_str
                
                # Исправляем ссылки на листы
                for wrong_name, correct_name in SHEET_NAME_FIXES.items():
                    # Ищем ссылки вида 'Узел учета'!A1
                    pattern = f"'{wrong_name}'!"
                    replacement = f"'{correct_name}'!"
                    
                    if pattern in formula_str:
                        formula_str = formula_str.replace(pattern, replacement)
                        sheet_fixes += 1
                
                # Если формула изменилась, обновляем ячейку
                if formula_str != original_formula:
                    cell.value = formula_str
    
    if sheet_fixes > 0:
        print(f"   ✅ Исправлено формул: {sheet_fixes}")
        fixes_by_sheet[sheet_name] = sheet_fixes
        total_fixes += sheet_fixes
    else:
        print(f"   ✅ Изменений не требуется")

# Сохраняем файл
wb.save(output_path)
wb.close()

print(f"\n{'='*80}")
print(f"📋 ИТОГОВЫЙ ОТЧЕТ")
print(f"{'='*80}\n")

print(f"✅ Всего исправлено формул: {total_fixes}")
print(f"\n📊 По листам:")
for sheet_name, count in sorted(fixes_by_sheet.items(), key=lambda x: x[1], reverse=True):
    print(f"   {sheet_name}: {count} формул")

print(f"\n💾 Исправленный файл сохранен: {output_path}")
print(f"✅ Готово!")

