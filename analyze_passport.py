#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Анализ энергопаспорта для восстановления формул"""
import openpyxl
import json
from pathlib import Path

file_path = r"C:\Users\DELL\Downloads\Telegram Desktop\ЭНЕРГО_ПАСПОРТ_23102025.xlsm"

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path, data_only=False)

sheets_info = {}
for sheet in wb.worksheets[:10]:
    formulas_count = 0
    values_count = 0
    ref_errors = []
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formulas_count += 1
                if cell.value and '#REF!' in str(cell.value):
                    ref_errors.append(cell.coordinate)
            elif cell.value and cell.data_type != 'f':
                values_count += 1
    
    sheets_info[sheet.title] = {
        'max_row': sheet.max_row,
        'max_column': sheet.max_column,
        'formulas': formulas_count,
        'values': values_count,
        'ref_errors': len(ref_errors),
        'ref_error_cells': ref_errors[:10]  # Первые 10
    }

print("\nSheets analysis:")
for name, info in sheets_info.items():
    print(f"  {name}:")
    print(f"    Rows: {info['max_row']}, Cols: {info['max_column']}")
    print(f"    Formulas: {info['formulas']}, Values: {info['values']}")
    if info['ref_errors'] > 0:
        print(f"    #REF! errors: {info['ref_errors']}")
        print(f"    Error cells: {', '.join(info['ref_error_cells'])}")

# Сохраняем результат
result = {
    'file_path': file_path,
    'sheets_analysis': sheets_info,
    'total_sheets': len(sheets_info),
    'total_formulas': sum(s['formulas'] for s in sheets_info.values()),
    'total_ref_errors': sum(s['ref_errors'] for s in sheets_info.values())
}

output_file = Path(__file__).parent / 'formula_analysis_step1_result.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\nResults saved to: {output_file}")

