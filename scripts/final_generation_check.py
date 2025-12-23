"""Финальная генерация и проверка"""
import sys
from pathlib import Path

# Добавляем пути
sys.path.insert(0, str(Path(__file__).parent))

print("=" * 80)
print("ГЕНЕРАЦИЯ И ПРОВЕРКА ЭНЕРГОПАСПОРТА")
print("=" * 80)

# Импортируем и запускаем генерацию
try:
    from tools.generate_metin_passport_full import *
    import importlib
    importlib.reload(sys.modules.get('tools.generate_metin_passport_full', None))
    
    # Запускаем через exec
    with open('tools/generate_metin_passport_full.py', 'r', encoding='utf-8') as f:
        code = f.read()
    exec(code)
    
except Exception as e:
    print(f"Ошибка: {e}")
    import traceback
    traceback.print_exc()

# Проверяем результат
print("\n" + "=" * 80)
print("ПРОВЕРКА РЕЗУЛЬТАТА")
print("=" * 80)

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")

if output_path.exists():
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(output_path, data_only=True)
    ws = wb["Структура пр 2"]
    
    print(f"\n✅ Файл создан: {output_path}")
    
    # Проверка E32
    e32 = ws["E32"].value
    print(f"\nE32 (2023 Q1 газ): {e32}")
    
    # Проверка всех кварталов
    quarters_info = [
        ("Q1", 17, 1, "B-H"),
        ("Q2", 39, 16, "Q-W"),
        ("Q3", 61, 31, "AF-AL"),
        ("Q4", 83, 46, "AU-BA"),
    ]
    
    total_filled = 0
    total_expected = 0
    
    for quarter, start_row, start_col, col_range in quarters_info:
        print(f"\n{quarter} (строка {start_row}, колонки {col_range}):")
        quarter_filled = 0
        
        for row in range(start_row, start_row + 5):
            for offset in [1, 2, 3, 4]:  # норма, 2022, 2023, 2024
                col = start_col + offset
                val = ws.cell(row, col).value
                total_expected += 1
                if val is not None and val != 0:
                    total_filled += 1
                    quarter_filled += 1
        
        print(f"  Заполнено: {quarter_filled}/20 полей")
    
    print(f"\n{'='*80}")
    print(f"ИТОГО: {total_filled}/{total_expected} полей заполнено ({total_filled*100/total_expected:.1f}%)")
    
    if total_filled == total_expected:
        print("✅ ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ ПОЛНОСТЬЮ!")
    elif total_filled > 0:
        print(f"⚠️ Заполнено частично")
    else:
        print("❌ КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")
    
    wb.close()
else:
    print(f"❌ Файл не найден: {output_path}")

print("\n" + "=" * 80)

