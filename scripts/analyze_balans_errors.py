"""
Детальный анализ ошибок на листе "Баланс"
Проверка ячеек-источников проблемы
"""
from openpyxl import load_workbook
from pathlib import Path

template_path = Path(r"C:\Users\DELL\Documents\AUDIT\METIN\Энерг паспорт\ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm")

print(f"📋 ДЕТАЛЬНЫЙ АНАЛИЗ ОШИБОК НА ЛИСТЕ 'Баланс'")
print("=" * 80)

# Загружаем с вычисленными значениями
wb_values = load_workbook(template_path, data_only=True)
ws_balans_values = wb_values["Баланс"]

# Загружаем с формулами
wb_formulas = load_workbook(template_path, data_only=False)
ws_balans_formulas = wb_formulas["Баланс"]

print("\n🔍 АНАЛИЗ ЦЕПОЧКИ ОШИБОК\n")

# Ошибки на строке 9
error_cells = ['C9', 'V9', 'AR9', 'AU9', 'BB9', 'BE9']

print("📌 Ячейки с ошибками #VALUE! на строке 9:\n")

for cell_coord in error_cells:
    cell_value = ws_balans_values[cell_coord]
    cell_formula = ws_balans_formulas[cell_coord]
    
    print(f"   {cell_coord}:")
    print(f"      Ошибка: {cell_value.value}")
    if cell_formula.data_type == 'f':
        print(f"      Формула: {cell_formula.value}")
    print()

# Анализ цепочки зависимостей
print("=" * 80)
print("🔗 АНАЛИЗ ЦЕПОЧКИ ЗАВИСИМОСТЕЙ\n")

print("V9 = U9+T9+S9+R9")
print("Проверяем ячейки R9, S9, T9, U9:\n")

source_cells = ['R9', 'S9', 'T9', 'U9']
for cell_coord in source_cells:
    cell_value = ws_balans_values[cell_coord]
    cell_formula = ws_balans_formulas[cell_coord]
    
    value_str = str(cell_value.value) if cell_value.value is not None else "пусто"
    value_type = type(cell_value.value).__name__
    
    print(f"   {cell_coord}:")
    print(f"      Значение: {value_str[:50]}")
    print(f"      Тип: {value_type}")
    
    if cell_formula.data_type == 'f':
        print(f"      Формула: {cell_formula.value}")
    
    # Проверка на ошибку
    if value_str.startswith('#'):
        print(f"      ⚠️ ОШИБКА: {value_str}")
    elif value_type == 'str' and not value_str.replace('.', '').replace('-', '').replace('E', '').replace('+', '').isdigit():
        print(f"      ⚠️ ТЕКСТОВОЕ ЗНАЧЕНИЕ (может вызывать #VALUE!)")
    print()

# Проверяем дополнительные ячейки
print("=" * 80)
print("🔍 ПРОВЕРКА ДОПОЛНИТЕЛЬНЫХ ЯЧЕЕК\n")

additional_cells = ['AW9', 'AZ9']
for cell_coord in additional_cells:
    cell_value = ws_balans_values[cell_coord]
    cell_formula = ws_balans_formulas[cell_coord]
    
    value_str = str(cell_value.value) if cell_value.value is not None else "пусто"
    value_type = type(cell_value.value).__name__
    
    print(f"   {cell_coord}:")
    print(f"      Значение: {value_str[:50]}")
    print(f"      Тип: {value_type}")
    
    if cell_formula.data_type == 'f':
        print(f"      Формула: {cell_formula.value}")
    
    if value_str.startswith('#'):
        print(f"      ⚠️ ОШИБКА: {value_str}")
    print()

# Проверяем контекст строки 9
print("=" * 80)
print("📊 КОНТЕКСТ СТРОКИ 9 (первые 30 колонок)\n")

header_row = 8  # Предполагаем, что заголовки в строке 8
print("Заголовки (строка 8):")
for col in range(1, min(31, ws_balans_values.max_column + 1)):
    header_cell = ws_balans_values.cell(8, col)
    if header_cell.value:
        col_letter = ws_balans_values.cell(8, col).column_letter
        print(f"   {col_letter}8: {str(header_cell.value)[:30]}")

print("\nЗначения строки 9 (первые 30 колонок):")
for col in range(1, min(31, ws_balans_values.max_column + 1)):
    cell = ws_balans_values.cell(9, col)
    if cell.value is not None:
        col_letter = cell.column_letter
        value_str = str(cell.value)
        if value_str.startswith('#'):
            print(f"   {col_letter}9: ❌ {value_str}")
        else:
            print(f"   {col_letter}9: {value_str[:30]}")

# Рекомендации
print("\n" + "=" * 80)
print("💡 РЕШЕНИЕ ПРОБЛЕМЫ")
print("=" * 80)

print("\n📌 ПРОБЛЕМА:")
print("   Ошибка #VALUE! возникает, когда формула пытается выполнить")
print("   математическую операцию с текстовым значением или ошибкой.")
print("\n   Цепочка ошибок:")
print("   1. V9 = U9+T9+S9+R9")
print("   2. Если R9, S9, T9 или U9 содержат текст/ошибку → V9 = #VALUE!")
print("   3. C9 = V9 → C9 = #VALUE!")
print("   4. AR9 = S9/1000*0.123 → если S9 текст → AR9 = #VALUE!")
print("   5. AU9 = V9/1000*0.123 → если V9 = #VALUE! → AU9 = #VALUE!")
print("   6. BB9 = AR9+AW9 → если AR9 = #VALUE! → BB9 = #VALUE!")
print("   7. BE9 = AU9+AZ9 → если AU9 = #VALUE! → BE9 = #VALUE!")

print("\n🔧 РЕШЕНИЕ:")
print("   1. Проверить ячейки R9, S9, T9, U9:")
print("      • Если там должны быть числа - убедиться, что нет текста")
print("      • Если там формулы - проверить, что они возвращают числа")
print("      • Если там пусто - заменить на 0 или использовать IF")
print("\n   2. Исправить формулы:")
print("      • V9: =IFERROR(U9+T9+S9+R9, 0)")
print("      • Или: =SUM(IFERROR(U9,0), IFERROR(T9,0), IFERROR(S9,0), IFERROR(R9,0))")
print("      • AR9: =IFERROR(S9/1000*0.123, 0)")
print("      • AU9: =IFERROR(V9/1000*0.123, 0)")
print("      • BB9: =IFERROR(AR9+AW9, 0)")
print("      • BE9: =IFERROR(AU9+AZ9, 0)")

print("\n   3. Альтернативное решение:")
print("      • Проверить источник данных для строки 9")
print("      • Убедиться, что данные заполнены корректно")
print("      • Если строка 9 должна быть пустой - формулы должны обрабатывать это")

wb_values.close()
wb_formulas.close()

print("\n✅ Анализ завершен")

