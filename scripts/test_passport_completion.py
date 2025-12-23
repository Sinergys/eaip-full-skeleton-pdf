"""
Тест заполнения энергопаспорта - проверка процента заполнения листов и ячеек.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def count_filled_cells(worksheet) -> Tuple[int, int]:
    """
    Подсчёт заполненных и пустых ячеек на листе.
    
    Returns:
        (filled_count, total_count)
    """
    filled = 0
    total = 0
    
    for row in worksheet.iter_rows():
        for cell in row:
            total += 1
            if cell.value is not None:
                # Проверяем, что это не пустая строка и не только пробелы
                value_str = str(cell.value).strip()
                if value_str and value_str not in ("", "None"):
                    filled += 1
    
    return filled, total


def analyze_sheet(worksheet) -> Dict:
    """Анализ заполнения одного листа."""
    filled, total = count_filled_cells(worksheet)
    percentage = (filled / total * 100) if total > 0 else 0.0
    
    # Подсчёт формул
    formula_count = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":  # formula
                formula_count += 1
    
    return {
        "name": worksheet.title,
        "filled_cells": filled,
        "total_cells": total,
        "percentage": round(percentage, 2),
        "formulas": formula_count,
        "rows": worksheet.max_row,
        "columns": worksheet.max_column,
    }


def analyze_passport(passport_path: Path) -> Dict:
    """Анализ заполнения всего энергопаспорта."""
    if not passport_path.exists():
        return {"error": f"Файл не найден: {passport_path}"}
    
    workbook = load_workbook(passport_path, data_only=False)
    
    sheets_analysis = []
    total_filled = 0
    total_cells = 0
    total_formulas = 0
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        analysis = analyze_sheet(sheet)
        sheets_analysis.append(analysis)
        total_filled += analysis["filled_cells"]
        total_cells += analysis["total_cells"]
        total_formulas += analysis["formulas"]
    
    overall_percentage = (total_filled / total_cells * 100) if total_cells > 0 else 0.0
    
    return {
        "file": str(passport_path),
        "sheets_count": len(workbook.sheetnames),
        "overall": {
            "filled_cells": total_filled,
            "total_cells": total_cells,
            "percentage": round(overall_percentage, 2),
            "formulas": total_formulas,
        },
        "sheets": sheets_analysis,
    }


def find_template() -> Path | None:
    """Поиск шаблона энергопаспорта."""
    possible_paths = [
        PROJECT_ROOT / "templates" / "pcm690" / "energy_passport_template.xlsx",
        PROJECT_ROOT / "data" / "source_files" / "metin" / "EnergyPassport_PKM690_filled.xlsx",
        Path(r"C:\Users\DELL\Documents\AUDIT\METIN\EnergyPassport_PKM690_Template_v1.1.2.xlsx"),
    ]
    
    for path in possible_paths:
        if path.exists():
            return path
    return None


def find_aggregated_data() -> Path | None:
    """Поиск агрегированных данных."""
    possible_paths = [
        PROJECT_ROOT / "data" / "aggregated" / "aggregated_full_resources_2022_2024.json",
        PROJECT_ROOT / "data" / "source_files" / "metin" / "aggregated_energy_2022_2024.json",
        Path(r"C:\Users\DELL\Documents\AUDIT\METIN\aggregated_energy_2022_2024.json"),
    ]
    
    for path in possible_paths:
        if path.exists():
            return path
    return None


def find_additional_data() -> Dict[str, Path]:
    """Поиск дополнительных JSON файлов."""
    data_dir = PROJECT_ROOT / "data" / "aggregated"
    result = {}
    
    if data_dir.exists():
        # Equipment
        equipment_path = data_dir / "oborudovanie_equipment.json"
        if equipment_path.exists():
            result["equipment"] = equipment_path
        
        # Envelope
        envelope_path = data_dir / "ograjdayuschie_envelope.json"
        if envelope_path.exists():
            result["envelope"] = envelope_path
        
        # Nodes
        nodes_path = data_dir / "test_nodes.xlsx"
        if nodes_path.exists():
            result["nodes"] = nodes_path
    
    return result


def generate_passport(template_path: Path, aggregated_path: Path, output_path: Path, additional: Dict[str, Path]) -> bool:
    """Генерация энергопаспорта через fill_energy_passport.py."""
    script_path = PROJECT_ROOT / "tools" / "fill_energy_passport.py"
    
    if not script_path.exists():
        print(f"❌ Скрипт не найден: {script_path}")
        return False
    
    cmd = [
        sys.executable,
        str(script_path),
        "--template", str(template_path),
        "--aggregated", str(aggregated_path),
        "--output", str(output_path),
        "--loss-active-month", "3200",
        "--loss-reactive-month", "13600",
        "--transformer-power", "630",
    ]
    
    if "equipment" in additional:
        cmd.extend(["--equipment-json", str(additional["equipment"])])
    
    if "envelope" in additional:
        cmd.extend(["--envelope-json", str(additional["envelope"])])
    
    print("🔄 Генерация энергопаспорта...")
    print(f"   Команда: {' '.join(cmd)}")
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
        if result.returncode == 0:
            print(f"✅ Энергопаспорт сгенерирован: {output_path}")
            return True
        else:
            print("❌ Ошибка генерации:")
            print(result.stderr)
            return False
    except Exception as e:
        print(f"❌ Исключение при генерации: {e}")
        return False


def print_report(analysis: Dict):
    """Вывод отчёта о заполнении."""
    if "error" in analysis:
        print(f"❌ {analysis['error']}")
        return
    
    print("\n" + "=" * 80)
    print("📊 ОТЧЁТ О ЗАПОЛНЕНИИ ЭНЕРГОПАСПОРТА")
    print("=" * 80)
    print(f"\n📁 Файл: {analysis['file']}")
    print(f"📋 Количество листов: {analysis['sheets_count']}")
    
    overall = analysis["overall"]
    print("\n📈 ОБЩАЯ СТАТИСТИКА:")
    print(f"   Заполнено ячеек: {overall['filled_cells']:,} из {overall['total_cells']:,}")
    print(f"   Процент заполнения: {overall['percentage']:.2f}%")
    print(f"   Формул в файле: {overall['formulas']:,}")
    
    print("\n📑 ДЕТАЛИЗАЦИЯ ПО ЛИСТАМ:")
    print(f"{'Лист':<30} {'Заполнено':<15} {'Всего':<12} {'%':<8} {'Формулы':<10}")
    print("-" * 80)
    
    for sheet in sorted(analysis["sheets"], key=lambda x: x["percentage"], reverse=True):
        name = sheet["name"][:28] + ".." if len(sheet["name"]) > 30 else sheet["name"]
        filled = f"{sheet['filled_cells']:,}"
        total = f"{sheet['total_cells']:,}"
        pct = f"{sheet['percentage']:.2f}%"
        formulas = f"{sheet['formulas']:,}"
        print(f"{name:<30} {filled:<15} {total:<12} {pct:<8} {formulas:<10}")
    
    # Топ-5 самых заполненных листов
    top_sheets = sorted(analysis["sheets"], key=lambda x: x["percentage"], reverse=True)[:5]
    print("\n🏆 ТОП-5 САМЫХ ЗАПОЛНЕННЫХ ЛИСТОВ:")
    for idx, sheet in enumerate(top_sheets, 1):
        print(f"   {idx}. {sheet['name']}: {sheet['percentage']:.2f}% ({sheet['filled_cells']:,} ячеек)")
    
    # Листы с низким заполнением (< 10%)
    low_sheets = [s for s in analysis["sheets"] if s["percentage"] < 10.0]
    if low_sheets:
        print("\n⚠️  ЛИСТЫ С НИЗКИМ ЗАПОЛНЕНИЕМ (< 10%):")
        for sheet in sorted(low_sheets, key=lambda x: x["percentage"]):
            print(f"   - {sheet['name']}: {sheet['percentage']:.2f}% ({sheet['filled_cells']:,} ячеек)")


def main():
    """Основная функция теста."""
    print("🧪 ТЕСТ ЗАПОЛНЕНИЯ ЭНЕРГОПАСПОРТА")
    print("=" * 80)
    
    # Поиск файлов
    template_path = find_template()
    if not template_path:
        print("❌ Шаблон энергопаспорта не найден!")
        print("   Искал в:")
        for path in [
            PROJECT_ROOT / "templates" / "pcm690",
            PROJECT_ROOT / "data" / "source_files" / "metin",
        ]:
            print(f"   - {path}")
        return 1
    
    print(f"✅ Шаблон найден: {template_path}")
    
    aggregated_path = find_aggregated_data()
    if not aggregated_path:
        print("❌ Агрегированные данные не найдены!")
        return 1
    
    print(f"✅ Агрегированные данные найдены: {aggregated_path}")
    
    additional = find_additional_data()
    if additional:
        print(f"✅ Дополнительные данные найдены: {', '.join(additional.keys())}")
    
    # Генерация энергопаспорта
    output_path = PROJECT_ROOT / "data" / "aggregated" / "test_passport_completion.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not generate_passport(template_path, aggregated_path, output_path, additional):
        return 1
    
    # Анализ заполнения
    print("\n🔍 Анализ заполнения...")
    analysis = analyze_passport(output_path)
    
    # Вывод отчёта
    print_report(analysis)
    
    # Сохранение JSON отчёта
    report_path = PROJECT_ROOT / "data" / "aggregated" / "passport_completion_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
    print(f"\n💾 JSON отчёт сохранён: {report_path}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

