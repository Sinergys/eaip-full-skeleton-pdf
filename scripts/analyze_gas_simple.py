from pathlib import Path
from openpyxl import load_workbook

p = Path("data/source_files/audit_sinergys/Расчет газа для отопл и неотпл.xlsx")

output_file = Path("gas_analysis_result.txt")

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("АНАЛИЗ ФАЙЛА ГАЗА\n")
    f.write("=" * 80 + "\n\n")
    
    f.write(f"File: {p.name}\n")
    f.write(f"Exists: {p.exists()}\n\n")
    
    if p.exists():
        wb = load_workbook(p, data_only=True)
        f.write(f"Sheets: {', '.join(wb.sheetnames)}\n\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"{'='*80}\n")
            f.write(f"Sheet: {sheet_name}\n")
            f.write(f"{'='*80}\n")
            f.write(f"Rows: {ws.max_row}, Cols: {ws.max_column}\n\n")
            
            # First 15 rows
            f.write("First 15 rows:\n")
            for row in range(1, min(16, ws.max_row + 1)):
                vals = []
                for col in range(1, min(20, ws.max_column + 1)):
                    v = ws.cell(row, col).value
                    if v is not None:
                        s = str(v)
                        if len(s) > 40:
                            s = s[:40] + "..."
                        vals.append(f"C{col}={s}")
                if vals:
                    f.write(f"  Row {row}: {' | '.join(vals[:6])}\n")
            
            # Find years
            f.write("\nYears found:\n")
            for row in range(1, min(15, ws.max_row + 1)):
                for col in range(1, min(50, ws.max_column + 1)):
                    v = ws.cell(row, col).value
                    if isinstance(v, int) and v in (2022, 2023, 2024):
                        f.write(f"  Year {v} at row {row}, col {col}\n")
        
        wb.close()
    
    f.write("\n" + "=" * 80 + "\n")
    f.write("ANALYSIS COMPLETE\n")
    f.write("=" * 80 + "\n")

print(f"Analysis saved to: {output_file}")

