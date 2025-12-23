"""
Проверка исправления формулы P28
"""
from openpyxl import load_workbook
from pathlib import Path

template_path = Path(r"C:\Users\DELL\Documents\AUDIT\METIN\Энерг паспорт\ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm")

print(f"📋 ПРОВЕРКА ИСПРАВЛЕНИЯ ФОРМУЛЫ P28")
print("=" * 80)

# Загружаем с формулами
wb = load_workbook(template_path, data_only=False)
ws = wb["Динамика ср"]

p28_cell = ws['P28']
p28_formula = p28_cell.value

print(f"\n✅ Формула в P28:")
print(f"   {p28_formula}\n")

# Проверяем, что формула содержит защиту
if 'IF(P27=0' in str(p28_formula) or 'ЕСЛИ(P27=0' in str(p28_formula):
    print("✅ Формула содержит защиту от деления на ноль")
else:
    print("⚠️ Формула не содержит защиту от деления на ноль")

# Проверяем значения P26 и P27
p26_cell = ws['P26']
p27_cell = ws['P27']

print(f"\n📊 Значения ячеек-источников:")
print(f"   P26 формула: {p26_cell.value if p26_cell.data_type == 'f' else p26_cell.value}")
print(f"   P27 формула: {p27_cell.value if p27_cell.data_type == 'f' else p27_cell.value}")

# Загружаем с вычисленными значениями для проверки ошибок
wb_values = load_workbook(template_path, data_only=True)
ws_values = wb_values["Динамика ср"]

p28_value = ws_values['P28'].value
p26_value = ws_values['P26'].value
p27_value = ws_values['P27'].value

print(f"\n📊 Вычисленные значения:")
print(f"   P26: {p26_value}")
print(f"   P27: {p27_value}")
print(f"   P28: {p28_value}")

if p28_value and str(p28_value).startswith('#'):
    print(f"\n❌ ОШИБКА: P28 все еще содержит ошибку {p28_value}")
elif p28_value == "-" or p28_value == "0%" or (p28_value and not str(p28_value).startswith('#')):
    print(f"\n✅ УСПЕХ: P28 вычисляется корректно (значение: {p28_value})")
else:
    print(f"\n⚠️ P28 пусто или None (возможно, нужно пересчитать в Excel)")

# Проверяем все ошибки на листе
print(f"\n{'='*80}")
print("🔍 ПРОВЕРКА ВСЕХ ОШИБОК НА ЛИСТЕ 'Динамика ср'")
print(f"{'='*80}\n")

errors_found = []
for row in ws_values.iter_rows():
    for cell in row:
        if cell.value is not None:
            value_str = str(cell.value)
            if value_str.startswith('#'):
                errors_found.append({
                    'cell': cell.coordinate,
                    'error': value_str
                })

if errors_found:
    print(f"❌ Найдено ошибок: {len(errors_found)}\n")
    for error in errors_found[:10]:
        print(f"   {error['cell']}: {error['error']}")
    if len(errors_found) > 10:
        print(f"   ... и еще {len(errors_found) - 10} ошибок")
else:
    print("✅ Ошибок на листе не найдено!")

wb.close()
wb_values.close()

print(f"\n✅ Проверка завершена")

