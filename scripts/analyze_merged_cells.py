#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Анализ объединенных ячеек и структуры разделов в oborudovanie.xlsx"""

from openpyxl import load_workbook
from pathlib import Path
import re

file_path = Path("data/source_files/audit_sinergys/oborudovanie.xlsx")

wb = load_workbook(file_path, data_only=True)

# Работаем только с Лист1
if "Лист1" not in wb.sheetnames:
    print("Лист1 не найден!")
    exit(1)

ws = wb["Лист1"]

print("=" * 80)
print("АНАЛИЗ ОБЪЕДИНЕННЫХ ЯЧЕЕК И СТРУКТУРЫ РАЗДЕЛОВ")
print(f"ЛИСТ: Лист1 ({ws.max_row} строк, {ws.max_column} столбцов)")
print("=" * 80 + "\n")

# Получаем все объединенные ячейки
merged_ranges = list(ws.merged_cells.ranges)
print(f"Найдено объединенных диапазонов: {len(merged_ranges)}\n")

# Анализируем структуру
sections = []
current_section = None
SECTION_PATTERN = re.compile(r"^\s*(\d+)\.?\s*(.*)$")

for row_idx, row in enumerate(ws.iter_rows(), 1):
    cells = [cell for cell in row[:10]]
    cell_values = [cell.value for cell in cells]
    
    # Пропускаем полностью пустые строки
    if not any(v is not None and str(v).strip() for v in cell_values):
        continue
    
    col_b_cell = cells[1] if len(cells) > 1 else None
    col_c_cell = cells[2] if len(cells) > 2 else None
    
    col_b = str(col_b_cell.value).strip() if col_b_cell and col_b_cell.value else ""
    col_c = str(col_c_cell.value).strip() if col_c_cell and col_c_cell.value else ""
    
    # Проверяем, является ли ячейка частью объединенного диапазона
    is_merged = False
    merged_range = None
    for mr in merged_ranges:
        if col_b_cell and col_b_cell.coordinate in mr:
            is_merged = True
            merged_range = mr
            break
    
    # ЗАГОЛОВОК СЕКЦИИ: в столбце B есть текст, столбец C пуст
    is_section = False
    section_name = None
    
    if col_b and not col_c:
        match = SECTION_PATTERN.match(col_b)
        if match:
            is_section = True
            section_name = col_b
        elif any(keyword in col_b.lower() for keyword in 
                 ["цех", "склад", "офис", "раздевалка", "душ", "производству", 
                  "помещение", "котельная", "подстанция", "тп"]):
            is_section = True
            section_name = col_b
    
    if is_section:
        if current_section:
            sections.append(current_section)
        
        # Проверяем объединенные ячейки для этой секции
        section_merged_info = None
        if is_merged and merged_range:
            section_merged_info = {
                "range": str(merged_range),
                "start_row": merged_range.min_row,
                "end_row": merged_range.max_row,
                "start_col": merged_range.min_col,
                "end_col": merged_range.max_col,
                "value": col_b
            }
        
        current_section = {
            "name": section_name,
            "start_row": row_idx,
            "header_row": None,
            "items": [],
            "totals": None,
            "merged_cell": section_merged_info
        }
        
        print(f"\n{'='*70}")
        print(f"СЕКЦИЯ [{row_idx:3d}]: {section_name}")
        if section_merged_info:
            print(f"  📌 ОБЪЕДИНЕННАЯ ЯЧЕЙКА: {section_merged_info['range']}")
            print(f"     Диапазон: строка {section_merged_info['start_row']}-{section_merged_info['end_row']}, "
                  f"столбцы {section_merged_info['start_col']}-{section_merged_info['end_col']}")
        else:
            print("  ⚠️  НЕ объединенная ячейка")
        print("=" * 70)
        continue
    
    # Если мы в секции, обрабатываем строки
    if current_section:
        # Заголовок таблицы
        if col_b and col_b.strip().lower() in ["№", "n", "no", "номер"]:
            current_section["header_row"] = row_idx
            print(f"  [Строка {row_idx:3d}] 📋 Заголовок таблицы")
            continue
        
        # Итоговая строка
        if col_b and "итого" in col_b.lower():
            # Проверяем, является ли "Итого" объединенной ячейкой
            totals_merged = None
            for mr in merged_ranges:
                if col_b_cell and col_b_cell.coordinate in mr:
                    totals_merged = {
                        "range": str(mr),
                        "start_row": mr.min_row,
                        "end_row": mr.max_row,
                        "start_col": mr.min_col,
                        "end_col": mr.max_col
                    }
                    break
            
            current_section["totals"] = {
                "row": row_idx,
                "count": cell_values[3] if len(cell_values) > 3 else None,
                "power": cell_values[5] if len(cell_values) > 5 else None,
                "vfd": cell_values[6] if len(cell_values) > 6 else None,
                "merged_cell": totals_merged
            }
            
            print(f"  [Строка {row_idx:3d}] 📊 ИТОГО: Кол-во={cell_values[3]}, Мощность={cell_values[5]}, ЧП={cell_values[6]}")
            if totals_merged:
                print(f"      📌 ОБЪЕДИНЕННАЯ ЯЧЕЙКА 'Итого': {totals_merged['range']}")
            continue
        
        # Строка с оборудованием
        if col_c:
            item = {
                "row": row_idx,
                "order": col_b,
                "name": col_c,
                "quantity": cell_values[3] if len(cell_values) > 3 else None,
                "unit_power": cell_values[4] if len(cell_values) > 4 else None,
                "total_power": cell_values[5] if len(cell_values) > 5 else None,
                "vfd": cell_values[6] if len(cell_values) > 6 else None
            }
            current_section["items"].append(item)
            # Показываем только первые 2 элемента для краткости
            if len(current_section["items"]) <= 2:
                print(f"  [Строка {row_idx:3d}] ⚙️  №{col_b}: {col_c[:60]}")
            elif len(current_section["items"]) == 3:
                print(f"  [Строка {row_idx:3d}] ⚙️  ... (еще {len(current_section['items']) - 2} единиц оборудования) ...")

# Закрываем последнюю секцию
if current_section:
    sections.append(current_section)

# Итоговый отчет
print("\n\n" + "=" * 80)
print("ИТОГОВЫЙ ОТЧЕТ")
print("=" * 80 + "\n")

print(f"Найдено секций: {len(sections)}\n")

for idx, sec in enumerate(sections, 1):
    print(f"{idx}. {sec['name']}")
    print(f"   Строки: {sec['start_row']}-{sec.get('totals', {}).get('row', '?')}")
    print(f"   Оборудования: {len(sec['items'])} единиц")
    
    if sec.get('merged_cell'):
        mc = sec['merged_cell']
        print(f"   📌 Заголовок секции в объединенной ячейке: {mc['range']}")
        print(f"      Значение: '{mc['value']}'")
        print(f"      Занимает строки: {mc['start_row']}-{mc['end_row']}")
    
    if sec.get('totals'):
        totals = sec['totals']
        print(f"   📊 Итого: {totals['count']} шт., {totals['power']} кВт, ЧП: {totals['vfd']}")
        if totals.get('merged_cell'):
            mc = totals['merged_cell']
            print(f"   📌 'Итого' в объединенной ячейке: {mc['range']}")
    
    print()

# Статистика по объединенным ячейкам
print("\n" + "=" * 80)
print("СТАТИСТИКА ПО ОБЪЕДИНЕННЫМ ЯЧЕЙКАМ")
print("=" * 80 + "\n")

section_headers_merged = sum(1 for sec in sections if sec.get('merged_cell'))
totals_merged = sum(1 for sec in sections if sec.get('totals', {}).get('merged_cell'))

print(f"Заголовков секций в объединенных ячейках: {section_headers_merged} из {len(sections)}")
print(f"Строк 'Итого' в объединенных ячейках: {totals_merged} из {len([s for s in sections if s.get('totals')])}")

wb.close()

