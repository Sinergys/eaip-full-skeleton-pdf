import sys
from pathlib import Path
from openpyxl import load_workbook

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")
report_path = Path("test_output_metin/check_report.txt")

# Перенаправляем вывод в файл
with open(report_path, 'w', encoding='utf-8') as f:
    sys.stdout = f
    
    print("=" * 80)
    print("ПРОВЕРКА ЗАПОЛНЕНИЯ")
    print("=" * 80)

if not output_path.exists():
    print(f"Файл не найден: {output_path}")
    sys.exit(1)

wb = load_workbook(output_path, data_only=True)
ws = wb["Структура пр 2"]

print(f"Файл загружен: {output_path.name}\n")

# E32
e32 = ws["E32"].value
print(f"E32 (2023 Q1 газ): {e32}")

# Q1
print("\nQ1 (строка 17):")
print(f"  B17 (норма): {ws['B17'].value}")
print(f"  C17 (2022): {ws['C17'].value}")
print(f"  D17 (2023): {ws['D17'].value}")
print(f"  E17 (2024): {ws['E17'].value}")

# Q2
print("\nQ2 (строка 39):")
print(f"  Q39 (норма): {ws['Q39'].value}")
print(f"  R39 (2022): {ws['R39'].value}")
print(f"  S39 (2023): {ws['S39'].value}")
print(f"  T39 (2024): {ws['T39'].value}")

# Q3
print("\nQ3 (строка 61):")
print(f"  AF61 (норма): {ws['AF61'].value}")
print(f"  AG61 (2022): {ws['AG61'].value}")
print(f"  AH61 (2023): {ws['AH61'].value}")
print(f"  AI61 (2024): {ws['AI61'].value}")

# Q4
print("\nQ4 (строка 83):")
print(f"  AU83 (норма): {ws['AU83'].value}")
print(f"  AV83 (2022): {ws['AV83'].value}")
print(f"  AW83 (2023): {ws['AW83'].value}")
print(f"  AX83 (2024): {ws['AX83'].value}")

# Подсчет
filled = 0
total = 0

quarters = [
    (17, 1),   # Q1
    (39, 16),  # Q2
    (61, 31),  # Q3
    (83, 46),  # Q4
]

for start_row, start_col in quarters:
    for row in range(start_row, start_row + 5):
        for offset in [1, 2, 3, 4]:
            col = start_col + offset
            val = ws.cell(row, col).value
            total += 1
            if val is not None and val != 0:
                filled += 1

print(f"\n{'='*80}")
print(f"ИТОГО: {filled}/{total} полей заполнено ({filled*100/total:.1f}%)")
if filled == total:
    print("ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ!")
elif filled > 0:
    print("Заполнено частично")
else:
    print("КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")

wb.close()

# Восстанавливаем stdout
sys.stdout = sys.__stdout__
print(f"Отчет сохранен в: {report_path}")

