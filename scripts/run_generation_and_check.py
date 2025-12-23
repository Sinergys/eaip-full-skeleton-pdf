"""Запуск генерации и проверка результата"""
import subprocess
import sys
from pathlib import Path

print("=" * 80)
print("ЗАПУСК ГЕНЕРАЦИИ ЭНЕРГОПАСПОРТА")
print("=" * 80)

# Запускаем генерацию
result = subprocess.run(
    [sys.executable, "tools/generate_metin_passport_full.py"],
    capture_output=True,
    text=True,
    encoding='utf-8',
    errors='replace'
)

print("Вывод генерации:")
print(result.stdout)
if result.stderr:
    print("Ошибки:")
    print(result.stderr)

print("\n" + "=" * 80)
print("ПРОВЕРКА РЕЗУЛЬТАТА")
print("=" * 80)

# Проверяем результат
output_path = Path("test_output_metin/generated_passport_fixed.xlsx")
if output_path.exists():
    print(f"✅ Файл создан: {output_path}")
    print(f"   Размер: {output_path.stat().st_size} байт")
    
    # Проверяем содержимое
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path, data_only=True)
        ws = wb["Структура пр 2"]
        
        print("\nПроверка заполнения:")
        
        # E32
        e32 = ws["E32"].value
        print(f"  E32 (2023 Q1 газ): {e32}")
        
        # Q1
        b17 = ws["B17"].value
        c17 = ws["C17"].value
        print(f"  Q1 - B17 (норма): {b17}, C17 (2022): {c17}")
        
        # Q2
        q39 = ws["Q39"].value
        r39 = ws["R39"].value
        print(f"  Q2 - Q39 (норма): {q39}, R39 (2022): {r39}")
        
        # Q3
        af61 = ws["AF61"].value
        ag61 = ws["AG61"].value
        print(f"  Q3 - AF61 (норма): {af61}, AG61 (2022): {ag61}")
        
        # Q4
        au83 = ws["AU83"].value
        av83 = ws["AV83"].value
        print(f"  Q4 - AU83 (норма): {au83}, AV83 (2022): {av83}")
        
        wb.close()
        
        # Подсчет заполненных полей
        filled = 0
        total = 0
        
        quarters = {
            "Q1": {"rows": [17, 18, 19, 20, 21], "start_col": 1},
            "Q2": {"rows": [39, 40, 41, 42, 43], "start_col": 16},
            "Q3": {"rows": [61, 62, 63, 64, 65], "start_col": 31},
            "Q4": {"rows": [83, 84, 85, 86, 87], "start_col": 46},
        }
        
        for quarter, info in quarters.items():
            for row in info["rows"]:
                for offset in [1, 2, 3, 4]:  # норма, 2022, 2023, 2024
                    col = info["start_col"] + offset
                    val = ws.cell(row, col).value
                    total += 1
                    if val is not None and val != 0:
                        filled += 1
        
        print(f"\nИтого заполнено: {filled}/{total} полей ({filled*100/total:.1f}%)")
        
        if filled == total:
            print("✅ ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ!")
        elif filled > 0:
            print(f"⚠️ Заполнено частично")
        else:
            print("❌ КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")
            
    except Exception as e:
        print(f"Ошибка при проверке: {e}")
        import traceback
        traceback.print_exc()
else:
    print(f"❌ Файл не создан: {output_path}")

print("\n" + "=" * 80)

