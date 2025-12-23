"""
Детальный анализ ссылок в формулах
"""
from openpyxl import load_workbook
from pathlib import Path
import re

template_path = Path(r"C:\eaip\data\source_files\audit_sinergys\ЭНЕРГО_ПАСПОРТ_ 1.xlsm")

wb = load_workbook(template_path, data_only=False)
sheet_names = wb.sheetnames

print("📋 Анализ ссылок в формулах")
print("=" * 80)
print(f"Реальные листы: {sheet_names}\n")

# Проверяем лист "Баланс" (там больше всего формул)
ws = wb["Баланс"]
print(f"📄 Лист: 'Баланс'\n")

# Находим первые 10 формул со ссылками на другие листы
formulas_with_refs = []
for row in ws.iter_rows():
    for cell in row:
        if cell.data_type == 'f':
            formula_str = str(cell.value)
            # Ищем ссылки на другие листы
            if "'" in formula_str or "!" in formula_str:
                formulas_with_refs.append({
                    'cell': cell.coordinate,
                    'formula': formula_str
                })
                if len(formulas_with_refs) >= 10:
                    break
    if len(formulas_with_refs) >= 10:
        break

print("Примеры формул со ссылками на другие листы:")
for i, f in enumerate(formulas_with_refs, 1):
    print(f"\n{i}. {f['cell']}:")
    print(f"   {f['formula'][:150]}")

wb.close()

