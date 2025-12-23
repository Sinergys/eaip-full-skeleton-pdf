"""
Скрипт для анализа эталонного шаблона энергопаспорта.
Извлекает структуру листов, формулы, именованные диапазоны и создает документацию.
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = PROJECT_ROOT / "data" / "source_files" / "audit_sinergys" / "EnergyPassport_PKM690_filled.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "docs"
OUTPUT_DIR.mkdir(exist_ok=True)


def extract_formulas(worksheet) -> List[Dict[str, Any]]:
    """Извлекает все формулы из листа"""
    formulas = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:  # f = formula
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'value': cell.value if hasattr(cell, 'value') else None,
                    'row': cell.row,
                    'column': cell.column,
                })
    return formulas


def extract_named_ranges(workbook) -> List[Dict[str, Any]]:
    """Извлекает именованные диапазоны из книги"""
    named_ranges = []
    for name, definition in workbook.defined_names.items():
        named_ranges.append({
            'name': name,
            'definition': str(definition),
            'scope': definition.scope if hasattr(definition, 'scope') else 'workbook',
        })
    return named_ranges


def analyze_sheet_structure(worksheet) -> Dict[str, Any]:
    """Анализирует структуру листа"""
    structure = {
        'name': worksheet.title,
        'max_row': worksheet.max_row,
        'max_column': worksheet.max_column,
        'merged_cells': [str(mc) for mc in worksheet.merged_cells.ranges],
        'formulas': extract_formulas(worksheet),
        'data_regions': [],
    }
    
    # Определяем регионы с данными (непустые ячейки)
    data_cells = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                data_cells.append({
                    'cell': cell.coordinate,
                    'value': str(cell.value)[:100],  # Ограничиваем длину
                    'data_type': cell.data_type,
                })
    
    structure['data_cells_count'] = len(data_cells)
    structure['sample_data_cells'] = data_cells[:50]  # Первые 50 ячеек для примера
    
    return structure


def create_field_mapping(worksheet) -> Dict[str, str]:
    """Создает карту соответствия ячеек для заполнения"""
    mapping = {}
    
    # Ищем ячейки с метками (обычно в столбце A)
    for row in worksheet.iter_rows(max_col=5):
        label_cell = row[0]  # Первая колонка - метка
        value_cell = row[1] if len(row) > 1 else None  # Вторая колонка - значение
        
        if label_cell.value and isinstance(label_cell.value, str):
            label = label_cell.value.strip()
            if value_cell:
                mapping[label] = value_cell.coordinate
    
    return mapping


def analyze_template():
    """Главная функция анализа шаблона"""
    print(f"Анализ шаблона: {TEMPLATE_PATH}")
    
    if not TEMPLATE_PATH.exists():
        print(f"ОШИБКА: Файл не найден: {TEMPLATE_PATH}")
        return None
    
    try:
        workbook = load_workbook(TEMPLATE_PATH, data_only=False)  # data_only=False для формул
    except Exception as e:
        print(f"ОШИБКА при загрузке файла: {e}")
        return None
    
    analysis = {
        'template_path': str(TEMPLATE_PATH),
        'sheets': [],
        'named_ranges': extract_named_ranges(workbook),
        'total_sheets': len(workbook.sheetnames),
    }
    
    print(f"Найдено листов: {len(workbook.sheetnames)}")
    print(f"Именованных диапазонов: {len(analysis['named_ranges'])}")
    
    # Анализируем каждый лист
    for sheet_name in workbook.sheetnames:
        print(f"  Анализ листа: {sheet_name}")
        worksheet = workbook[sheet_name]
        sheet_analysis = analyze_sheet_structure(worksheet)
        sheet_analysis['field_mapping'] = create_field_mapping(worksheet)
        analysis['sheets'].append(sheet_analysis)
        
        print(f"    - Строк: {sheet_analysis['max_row']}, Колонок: {sheet_analysis['max_column']}")
        print(f"    - Формул: {len(sheet_analysis['formulas'])}")
        print(f"    - Объединенных ячеек: {len(sheet_analysis['merged_cells'])}")
    
    # Сохраняем результаты
    output_file = OUTPUT_DIR / "template_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
    
    print(f"\nРезультаты сохранены в: {output_file}")
    
    # Создаем текстовый отчет
    report_file = OUTPUT_DIR / "template_analysis_report.md"
    create_text_report(analysis, report_file)
    
    workbook.close()
    return analysis


def create_text_report(analysis: Dict, output_file: Path):
    """Создает текстовый отчет в формате Markdown"""
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("# Анализ эталонного шаблона энергопаспорта\n\n")
        f.write(f"**Файл:** `{analysis['template_path']}`\n\n")
        f.write(f"**Всего листов:** {analysis['total_sheets']}\n")
        f.write(f"**Именованных диапазонов:** {len(analysis['named_ranges'])}\n\n")
        
        f.write("## Именованные диапазоны\n\n")
        if analysis['named_ranges']:
            f.write("| Имя | Определение | Область |\n")
            f.write("|-----|-------------|----------|\n")
            for nr in analysis['named_ranges']:
                f.write(f"| {nr['name']} | {nr['definition']} | {nr['scope']} |\n")
        else:
            f.write("Именованные диапазоны не найдены.\n")
        f.write("\n")
        
        f.write("## Структура листов\n\n")
        for sheet in analysis['sheets']:
            f.write(f"### {sheet['name']}\n\n")
            f.write(f"- **Размер:** {sheet['max_row']} строк × {sheet['max_column']} колонок\n")
            f.write(f"- **Формул:** {len(sheet['formulas'])}\n")
            f.write(f"- **Объединенных ячеек:** {len(sheet['merged_cells'])}\n")
            f.write(f"- **Ячеек с данными:** {sheet['data_cells_count']}\n\n")
            
            if sheet['formulas']:
                f.write("#### Формулы:\n\n")
                f.write("| Ячейка | Формула |\n")
                f.write("|--------|----------|\n")
                for formula in sheet['formulas'][:20]:  # Первые 20 формул
                    formula_text = formula['formula'].replace('|', '\\|')
                    f.write(f"| {formula['cell']} | `{formula_text}` |\n")
                if len(sheet['formulas']) > 20:
                    f.write(f"\n*... и еще {len(sheet['formulas']) - 20} формул*\n")
                f.write("\n")
            
            if sheet['field_mapping']:
                f.write("#### Карта полей (метка → ячейка):\n\n")
                f.write("| Метка | Ячейка |\n")
                f.write("|-------|--------|\n")
                for label, cell in list(sheet['field_mapping'].items())[:20]:
                    f.write(f"| {label} | {cell} |\n")
                if len(sheet['field_mapping']) > 20:
                    f.write(f"\n*... и еще {len(sheet['field_mapping']) - 20} полей*\n")
                f.write("\n")
    
    print(f"Текстовый отчет сохранен в: {output_file}")


if __name__ == "__main__":
    analyze_template()

