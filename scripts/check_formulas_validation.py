"""
Детальная проверка правильности формул в шаблоне ЭНЕРГО_ПАСПОРТ_ 1.xlsm
Проверяет синтаксис, ссылки, логику и выявляет потенциальные проблемы
"""
from openpyxl import load_workbook
from pathlib import Path
import re
from collections import defaultdict

template_path = Path(r"C:\eaip\data\source_files\audit_sinergys\ЭНЕРГО_ПАСПОРТ_ 1.xlsm")

if not template_path.exists():
    print(f"❌ Файл не найден: {template_path}")
    exit(1)

print(f"📋 Проверка формул в шаблоне: {template_path.name}")
print("=" * 80)

wb = load_workbook(template_path, data_only=False)
sheet_names = wb.sheetnames
print(f"✅ Файл загружен")
print(f"📊 Всего листов: {len(sheet_names)}")
print(f"   Листы: {sheet_names}\n")

# Собираем все замечания
all_remarks = []
formula_statistics = defaultdict(int)

# Паттерны для проверки формул
PATTERNS = {
    'sum': r'=SUM\([^)]+\)',
    'if': r'=IF\([^)]+\)',
    'vlookup': r'=VLOOKUP\([^)]+\)',
    'reference': r"'[^']+'![A-Z]+\d+",  # Ссылки на другие листы
    'circular': r'=.*\$\w+\d+.*\$\w+\d+',  # Потенциальные циклические ссылки
}

def check_formula_syntax(formula_str, cell_coord, sheet_name):
    """Проверка синтаксиса формулы"""
    remarks = []
    
    # Проверка на незакрытые скобки
    open_brackets = formula_str.count('(')
    close_brackets = formula_str.count(')')
    if open_brackets != close_brackets:
        remarks.append({
            'type': 'syntax_error',
            'severity': 'critical',
            'message': f"Несбалансированные скобки: открывающих {open_brackets}, закрывающих {close_brackets}",
            'cell': cell_coord,
            'formula': formula_str[:100]
        })
    
    # Проверка на пустые аргументы функций
    if re.search(r'\([,\s]+\)|\([,\s]+,', formula_str):
        remarks.append({
            'type': 'syntax_error',
            'severity': 'medium',
            'message': "Возможные пустые аргументы в функции",
            'cell': cell_coord,
            'formula': formula_str[:100]
        })
    
    # Проверка на двойные операторы
    if re.search(r'[+\-*/]{2,}', formula_str):
        remarks.append({
            'type': 'syntax_error',
            'severity': 'high',
            'message': "Двойные операторы в формуле",
            'cell': cell_coord,
            'formula': formula_str[:100]
        })
    
    return remarks

def check_formula_references(formula_str, cell_coord, sheet_name, all_sheets):
    """Проверка ссылок на другие листы и ячейки"""
    remarks = []
    
    # Проверка ссылок на другие листы
    sheet_refs = re.findall(r"'([^']+)'!", formula_str)
    for ref_sheet in sheet_refs:
        # Проверяем точное совпадение и совпадение с учетом пробелов
        ref_sheet_exact = ref_sheet  # Точное имя из формулы
        ref_sheet_clean = ref_sheet.strip()  # Без пробелов
        
        # Проверяем точное совпадение
        if ref_sheet_exact not in all_sheets:
            # Проверяем совпадение без пробелов
            found = False
            for sheet_name in all_sheets:
                if sheet_name.strip() == ref_sheet_clean:
                    found = True
                    break
            
            if not found:
                remarks.append({
                    'type': 'broken_reference',
                    'severity': 'critical',
                    'message': f"Ссылка на несуществующий лист: '{ref_sheet_exact}' (без пробелов: '{ref_sheet_clean}')",
                    'cell': cell_coord,
                    'formula': formula_str[:100],
                    'referenced_sheet': ref_sheet_exact
                })
    
    # Проверка ссылок на ячейки (базовая валидация)
    cell_refs = re.findall(r'[A-Z]+\d+', formula_str)
    for ref in cell_refs:
        # Проверка на некорректные ссылки (например, A0, ZZZ999999)
        if re.match(r'[A-Z]+0$', ref):
            remarks.append({
                'type': 'invalid_reference',
                'severity': 'high',
                'message': f"Некорректная ссылка на ячейку: {ref} (строка 0 не существует)",
                'cell': cell_coord,
                'formula': formula_str[:100]
            })
    
    return remarks

def check_formula_logic(formula_str, cell_coord, sheet_name):
    """Проверка логики формул"""
    remarks = []
    
    # Проверка на деление на ноль (потенциальное)
    if re.search(r'/[A-Z]+\d+[^)]*\)|/[A-Z]+\d+\s*$', formula_str):
        remarks.append({
            'type': 'logic_warning',
            'severity': 'low',
            'message': "Потенциальное деление на ноль (нет проверки делителя)",
            'cell': cell_coord,
            'formula': formula_str[:100]
        })
    
    # Проверка на SUM с одним аргументом
    sum_matches = re.finditer(r'=SUM\(([^)]+)\)', formula_str)
    for match in sum_matches:
        args = match.group(1)
        if ',' not in args and ':' not in args:
            remarks.append({
                'type': 'logic_warning',
                'severity': 'low',
                'message': "SUM с одним аргументом (можно заменить на прямое значение)",
                'cell': cell_coord,
                'formula': formula_str[:100]
            })
    
    return remarks

def analyze_formula_patterns(formula_str):
    """Анализ паттернов в формулах"""
    patterns_found = {}
    for pattern_name, pattern in PATTERNS.items():
        if re.search(pattern, formula_str, re.IGNORECASE):
            patterns_found[pattern_name] = True
    return patterns_found

# Проверяем все листы (кроме Sheet1)
for sheet_name in sheet_names:
    if sheet_name == "Sheet1":
        continue
    
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"📄 Лист: '{sheet_name}'")
    print(f"{'='*80}")
    
    sheet_remarks = []
    formulas_count = 0
    formulas_by_type = defaultdict(int)
    
    # Проверяем все ячейки с формулами
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Формула
                formulas_count += 1
                formula_str = str(cell.value)
                cell_coord = cell.coordinate
                
                # Анализ паттернов
                patterns = analyze_formula_patterns(formula_str)
                for pattern_name in patterns:
                    formulas_by_type[pattern_name] += 1
                
                # Проверка синтаксиса
                syntax_remarks = check_formula_syntax(formula_str, cell_coord, sheet_name)
                sheet_remarks.extend(syntax_remarks)
                
                # Проверка ссылок
                ref_remarks = check_formula_references(formula_str, cell_coord, sheet_name, sheet_names)
                sheet_remarks.extend(ref_remarks)
                
                # Проверка логики
                logic_remarks = check_formula_logic(formula_str, cell_coord, sheet_name)
                sheet_remarks.extend(logic_remarks)
    
    # Выводим статистику
    print(f"📊 Статистика формул:")
    print(f"   Всего формул: {formulas_count}")
    if formulas_by_type:
        print(f"   По типам:")
        for pattern_name, count in sorted(formulas_by_type.items()):
            print(f"      {pattern_name}: {count}")
    
    # Выводим замечания
    if sheet_remarks:
        print(f"\n⚠️ Замечания ({len(sheet_remarks)}):")
        
        # Группируем по типу
        by_type = defaultdict(list)
        for remark in sheet_remarks:
            by_type[remark['type']].append(remark)
        
        for remark_type, remarks_list in sorted(by_type.items()):
            severity_counts = defaultdict(int)
            for r in remarks_list:
                severity_counts[r['severity']] += 1
            
            print(f"\n   📌 {remark_type} ({len(remarks_list)}):")
            for severity in ['critical', 'high', 'medium', 'low']:
                if severity_counts[severity]:
                    print(f"      {severity.upper()}: {severity_counts[severity]}")
            
            # Показываем примеры (первые 3)
            for remark in remarks_list[:3]:
                print(f"      • {remark['cell']}: {remark['message']}")
                if len(remarks_list) > 3 and remark == remarks_list[2]:
                    print(f"      ... и еще {len(remarks_list) - 3} замечаний")
        
        all_remarks.extend(sheet_remarks)
    else:
        print(f"\n✅ Критических замечаний не найдено")
    
    formula_statistics[sheet_name] = {
        'total': formulas_count,
        'remarks': len(sheet_remarks)
    }

wb.close()

# Итоговый отчет
print(f"\n\n{'='*80}")
print("📋 ИТОГОВЫЙ ОТЧЕТ ПО ФОРМУЛАМ")
print(f"{'='*80}\n")

print("📊 Сводка по листам:")
print(f"{'Лист':<30} {'Формул':<10} {'Замечаний':<12} {'Статус':<10}")
print("-" * 65)
for sheet_name, stats in formula_statistics.items():
    status = "⚠️" if stats['remarks'] > 0 else "✅"
    print(f"{sheet_name:<30} {stats['total']:<10} {stats['remarks']:<12} {status:<10}")

# Группировка замечаний по критичности
if all_remarks:
    print(f"\n⚠️ ВСЕГО ЗАМЕЧАНИЙ: {len(all_remarks)}")
    
    by_severity = defaultdict(list)
    for remark in all_remarks:
        by_severity[remark['severity']].append(remark)
    
    for severity in ['critical', 'high', 'medium', 'low']:
        if by_severity[severity]:
            print(f"\n🔴 {severity.upper()} ({len(by_severity[severity])}):")
            for remark in by_severity[severity][:10]:  # Первые 10
                print(f"   • {remark['sheet'] if 'sheet' in remark else 'N/A'}!{remark['cell']}: {remark['message']}")
            if len(by_severity[severity]) > 10:
                print(f"   ... и еще {len(by_severity[severity]) - 10} замечаний")
    
    # Группировка по типу
    by_type = defaultdict(list)
    for remark in all_remarks:
        by_type[remark['type']].append(remark)
    
    print(f"\n📌 Распределение по типам:")
    for remark_type, remarks_list in sorted(by_type.items()):
        print(f"   {remark_type}: {len(remarks_list)}")
else:
    print(f"\n✅ Критических замечаний не найдено!")

print(f"\n✅ Проверка завершена")

