"""
Параметризованный интеграционный тест для всех эталонных объектов.

Проверяет корректность заполнения энергопаспорта на основе эталонных данных
и сравнивает результаты с ожидаемыми значениями для всех reference_enterprise_N.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Dict, Any, List, Tuple
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Список всех эталонных объектов для тестирования
REFERENCE_ENTERPRISES = [
    "reference_enterprise_1",
    "reference_enterprise_2_heat_intensive",
    "reference_enterprise_3_electric_intensive",
    "reference_enterprise_4_services",
]


def load_reference_data(enterprise_name: str) -> Dict[str, Any]:
    """
    Загружает эталонные данные для указанного предприятия.
    
    Args:
        enterprise_name: Имя эталонного объекта (например, "reference_enterprise_1")
    
    Returns:
        Словарь с эталонными данными
    """
    reference_path = PROJECT_ROOT / "data" / "fixtures" / f"{enterprise_name}.json"
    if not reference_path.exists():
        raise FileNotFoundError(f"Эталонный файл не найден: {reference_path}")
    
    return json.loads(reference_path.read_text(encoding="utf-8"))


def prepare_test_data(reference: Dict[str, Any], output_dir: Path) -> Dict[str, Path]:
    """
    Подготавливает тестовые данные из эталонного объекта.
    
    Returns:
        Словарь с путями к подготовленным файлам
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Создаём агрегированный JSON
    aggregated_path = output_dir / "reference_aggregated.json"
    aggregated_data = {
        "resources": reference["input_data"]["aggregated_resources"]["resources"]
    }
    aggregated_path.write_text(
        json.dumps(aggregated_data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    
    # Создаём equipment JSON
    equipment_path = output_dir / "reference_equipment.json"
    equipment_path.write_text(
        json.dumps(reference["input_data"]["equipment"], ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    
    # Создаём envelope JSON
    envelope_path = output_dir / "reference_envelope.json"
    envelope_path.write_text(
        json.dumps(reference["input_data"]["envelope"], ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    
    # Создаём nodes JSON
    nodes_path = output_dir / "reference_nodes.json"
    nodes_path.write_text(
        json.dumps(reference["input_data"]["nodes"], ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    
    return {
        "aggregated": aggregated_path,
        "equipment": equipment_path,
        "envelope": envelope_path,
        "nodes": nodes_path,
    }


def generate_passport(
    template_path: Path,
    test_data: Dict[str, Path],
    output_path: Path,
    losses: Dict[str, float]
) -> bool:
    """Генерирует паспорт через fill_energy_passport.py."""
    script_path = PROJECT_ROOT / "tools" / "fill_energy_passport.py"
    
    if not script_path.exists():
        print(f"❌ Скрипт не найден: {script_path}")
        return False
    
    cmd = [
        sys.executable,
        str(script_path),
        "--template", str(template_path),
        "--aggregated", str(test_data["aggregated"]),
        "--output", str(output_path),
        "--equipment-json", str(test_data["equipment"]),
        "--envelope-json", str(test_data["envelope"]),
        "--nodes-json", str(test_data["nodes"]),
        "--loss-active-month", str(losses["loss_active_month"]),
        "--loss-reactive-month", str(losses["loss_reactive_month"]),
        "--transformer-power", str(losses["transformer_power_kva"]),
    ]
    
    import subprocess
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
        if result.returncode == 0:
            print(f"✅ Паспорт сгенерирован: {output_path}")
            return True
        else:
            print("❌ Ошибка генерации:")
            print(result.stderr)
            return False
    except Exception as e:
        print(f"❌ Исключение при генерации: {e}")
        return False


def read_cell_value(workbook, sheet_name: str, row: int, column: int) -> Any:
    """Читает значение ячейки из листа."""
    if sheet_name not in workbook.sheetnames:
        return None
    
    ws = workbook[sheet_name]
    try:
        cell = ws.cell(row=row, column=column)
        return cell.value
    except Exception:
        return None


def compare_values(actual: Any, expected: Any, tolerance: float = 0.01) -> Tuple[bool, str]:
    """
    Сравнивает значения с учётом погрешности для чисел.
    
    Returns:
        (is_match, error_message)
    """
    # Обработка None
    if actual is None and expected is None:
        return True, ""
    if actual is None:
        return False, f"Ожидалось {expected}, получено None"
    if expected is None:
        return False, f"Ожидалось None, получено {actual}"
    
    # Обработка чисел
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        diff = abs(actual - expected)
        if diff <= tolerance or (expected > 0 and diff / abs(expected) <= tolerance):
            return True, ""
        return False, f"Ожидалось {expected}, получено {actual}, разница {diff}"
    
    # Обработка строк
    if isinstance(expected, str) and isinstance(actual, str):
        if actual.strip() == expected.strip():
            return True, ""
        return False, f"Ожидалось '{expected}', получено '{actual}'"
    
    # Обработка других типов
    if actual == expected:
        return True, ""
    return False, f"Ожидалось {expected}, получено {actual}"


def validate_sheet(
    workbook,
    sheet_name: str,
    expected_values: Dict[str, Any],
    cell_coordinates: Dict[str, Dict[str, Dict[str, int]]]
) -> List[str]:
    """
    Валидирует заполнение листа.
    
    Returns:
        Список ошибок (пустой, если всё корректно)
    """
    errors = []
    
    if sheet_name not in workbook.sheetnames:
        errors.append(f"Лист '{sheet_name}' отсутствует в паспорте")
        return errors
    
    # Проверяем значения по координатам ячеек
    if sheet_name in cell_coordinates:
        coords = cell_coordinates[sheet_name]
        for key, coord in coords.items():
            row = coord.get("row")
            col = coord.get("column")
            if row and col:
                actual = read_cell_value(workbook, sheet_name, row, col)
                # Извлекаем ожидаемое значение из expected_values
                expected = _get_nested_value(expected_values, key)
                
                if expected is not None:
                    is_match, error_msg = compare_values(actual, expected)
                    if not is_match:
                        errors.append(
                            f"Лист '{sheet_name}', ячейка {_cell_name(row, col)} ({key}): {error_msg}"
                        )
    
    # Проверяем агрегированные значения
    if sheet_name in expected_values:
        sheet_expected = expected_values[sheet_name]
        if isinstance(sheet_expected, dict):
            # Проверяем квартальные значения
            for quarter, quarter_data in sheet_expected.items():
                if isinstance(quarter_data, dict):
                    # Здесь можно добавить более детальную проверку
                    pass
    
    return errors


def _get_nested_value(data: Dict[str, Any], key: str) -> Any:
    """Извлекает значение по ключу вида '2022-Q1_active'."""
    parts = key.split("_")
    if len(parts) >= 2:
        quarter = parts[0]
        field = "_".join(parts[1:])
        
        # Ищем в структуре данных
        for sheet_data in data.values():
            if isinstance(sheet_data, dict) and quarter in sheet_data:
                quarter_data = sheet_data[quarter]
                if isinstance(quarter_data, dict) and field in quarter_data:
                    return quarter_data[field]
    
    return None


def _cell_name(row: int, column: int) -> str:
    """Преобразует координаты в имя ячейки (например, A1)."""
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(column)}{row}"


def run_single_test(enterprise_name: str) -> Tuple[bool, List[str]]:
    """
    Запускает тест для одного эталонного объекта.
    
    Args:
        enterprise_name: Имя эталонного объекта
    
    Returns:
        (success, list_of_errors)
    """
    print(f"\n🧪 ТЕСТ: {enterprise_name}")
    print("=" * 80)
    
    # Загружаем эталонные данные
    try:
        reference = load_reference_data(enterprise_name)
        print(f"✅ Эталонные данные загружены: {reference.get('enterprise_name', enterprise_name)}")
    except Exception as e:
        print(f"❌ Ошибка загрузки эталонных данных: {e}")
        return False, [f"Ошибка загрузки данных: {e}"]
    
    # Находим шаблон
    template_candidates = [
        PROJECT_ROOT / "templates" / "pcm690" / "new_energy_passport.xlsx",
        PROJECT_ROOT / "templates" / "pcm690" / "energy_passport_template.xlsx",
    ]
    template_path = next((p for p in template_candidates if p.exists()), None)
    if not template_path:
        error_msg = "Шаблон не найден"
        print(f"❌ {error_msg}")
        return False, [error_msg]
    print(f"✅ Шаблон найден: {template_path}")
    
    # Подготавливаем тестовые данные
    test_output_dir = PROJECT_ROOT / "test_output" / enterprise_name
    test_data = prepare_test_data(reference, test_output_dir)
    print("✅ Тестовые данные подготовлены")
    
    # Генерируем паспорт
    output_path = test_output_dir / f"{enterprise_name}_passport.xlsx"
    losses = reference["input_data"]["losses"]
    
    if not generate_passport(template_path, test_data, output_path, losses):
        return False, ["Ошибка генерации паспорта"]
    
    # Загружаем сгенерированный паспорт
    try:
        workbook = load_workbook(output_path, data_only=False)
        print("✅ Паспорт загружен для проверки")
    except Exception as e:
        error_msg = f"Ошибка загрузки паспорта: {e}"
        print(f"❌ {error_msg}")
        return False, [error_msg]
    
    # Валидируем листы
    expected_results = reference.get("expected_results", {})
    cell_coordinates = expected_results.get("cell_coordinates", {})
    sheet_values = expected_results.get("sheet_values", {})
    
    all_errors = []
    
    print("\n🔍 Проверка заполнения листов:")
    for sheet_name in sheet_values.keys():
        errors = validate_sheet(workbook, sheet_name, sheet_values, cell_coordinates)
        if errors:
            print(f"  ❌ {sheet_name}: {len(errors)} ошибок")
            all_errors.extend([f"{enterprise_name}:{sheet_name}: {e}" for e in errors])
        else:
            print(f"  ✅ {sheet_name}: OK")
    
    # Выводим результаты
    if all_errors:
        print(f"\n❌ ТЕСТ НЕ ПРОЙДЕН: найдено {len(all_errors)} ошибок")
        return False, all_errors
    else:
        print("\n✅ ТЕСТ ПРОЙДЕН: все проверки успешны")
        return True, []


def run_test() -> int:
    """
    Запускает тесты для всех эталонных объектов.
    
    Returns:
        Код возврата: 0 если все тесты прошли, 1 если есть ошибки
    """
    print("🧪 ПАРАМЕТРИЗОВАННЫЙ ТЕСТ ВСЕХ ЭТАЛОННЫХ ОБЪЕКТОВ")
    print("=" * 80)
    print(f"Тестируем {len(REFERENCE_ENTERPRISES)} эталонных объектов:")
    for i, name in enumerate(REFERENCE_ENTERPRISES, 1):
        print(f"  {i}. {name}")
    
    results = []
    all_errors = []
    
    # Запускаем тесты для каждого объекта
    for enterprise_name in REFERENCE_ENTERPRISES:
        success, errors = run_single_test(enterprise_name)
        results.append({
            "enterprise": enterprise_name,
            "success": success,
            "errors": errors
        })
        all_errors.extend(errors)
    
    # Итоговый отчёт
    print("\n" + "=" * 80)
    print("📊 ИТОГОВЫЙ ОТЧЁТ")
    print("=" * 80)
    
    passed = sum(1 for r in results if r["success"])
    failed = len(results) - passed
    
    print(f"\n✅ Успешно: {passed}/{len(results)}")
    print(f"❌ Провалено: {failed}/{len(results)}")
    
    # Детальная таблица результатов
    print("\n📋 Детализация:")
    print(f"{'Предприятие':<40} {'Статус':<10} {'Ошибок':<10}")
    print("-" * 80)
    
    for result in results:
        status = "✅ PASS" if result["success"] else "❌ FAIL"
        error_count = len(result["errors"])
        enterprise_short = result["enterprise"].replace("reference_enterprise_", "ref_")
        print(f"{enterprise_short:<40} {status:<10} {error_count:<10}")
    
    # Выводим ошибки
    if all_errors:
        print(f"\n❌ ОБНАРУЖЕНЫ ОШИБКИ ({len(all_errors)}):")
        for i, error in enumerate(all_errors[:30], 1):  # Показываем первые 30
            print(f"  {i}. {error}")
        if len(all_errors) > 30:
            print(f"  ... и ещё {len(all_errors) - 30} ошибок")
        
        # Сохраняем отчёт в файл
        report_path = PROJECT_ROOT / "test_output" / "test_report.json"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump({
                "total_tests": len(results),
                "passed": passed,
                "failed": failed,
                "results": results,
                "all_errors": all_errors
            }, f, ensure_ascii=False, indent=2)
        print(f"\n💾 Полный отчёт сохранён: {report_path}")
        
        return 1
    else:
        print("\n✅ ВСЕ ТЕСТЫ ПРОЙДЕНЫ УСПЕШНО!")
        return 0


if __name__ == "__main__":
    sys.exit(run_test())

