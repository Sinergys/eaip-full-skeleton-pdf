#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Диагностика проблемы с electricity.by_usage - проверка всех этапов"""

import json
from pathlib import Path
from openpyxl import load_workbook

print("=" * 80)
print("ДИАГНОСТИКА: Проверка всех этапов обработки by_usage")
print("=" * 80 + "\n")

# ============================================================================
# ПРОВЕРКА 1: Загружен ли файл pererashod.xlsx?
# ============================================================================
print("=" * 80)
print("ПРОВЕРКА 1: Загружен ли файл pererashod.xlsx?")
print("=" * 80 + "\n")

pererashod_paths = [
    Path("data/source_files/audit_sinergys/pererashod.xlsx"),
    Path("eaip_full_skeleton/infra/data/inbox") / "pererashod.xlsx",
]

pererashod_found = False
for path in pererashod_paths:
    if path.exists():
        print(f"✅ Файл найден: {path.absolute()}")
        pererashod_found = True
        
        # Проверяем структуру файла
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            sheet = wb.active
            print(f"   Лист: {sheet.title}")
            print(f"   Размер: {sheet.max_row} строк, {sheet.max_column} столбцов")
            
            # Проверяем наличие ключевых слов
            keywords_found = []
            for row_idx in range(1, min(40, sheet.max_row + 1)):
                for col_idx in range(1, min(20, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_text = str(cell.value).lower()
                        if any(kw in cell_text for kw in ["тех-потер", "хоз-быт", "производств"]):
                            keywords_found.append(f"{cell.column_letter}{row_idx}: {cell.value}")
            
            if keywords_found:
                print(f"   ✅ Найдены ключевые слова категорий:")
                for kw in keywords_found[:5]:
                    print(f"      {kw}")
            else:
                print(f"   ⚠️ Ключевые слова категорий не найдены")
            
            wb.close()
        except Exception as e:
            print(f"   ❌ Ошибка при чтении файла: {e}")
        break

if not pererashod_found:
    print("❌ Файл pererashod.xlsx НЕ найден в ожидаемых местах")
    print("   Ожидаемые пути:")
    for path in pererashod_paths:
        print(f"      {path.absolute()}")

print()

# ============================================================================
# ПРОВЕРКА 2: Создается ли файл usage_categories.json?
# ============================================================================
print("=" * 80)
print("ПРОВЕРКА 2: Создается ли файл usage_categories.json?")
print("=" * 80 + "\n")

usage_categories_paths = [
    Path("data/aggregated/usage_categories.json"),
    Path("eaip_full_skeleton/infra/data/aggregated/usage_categories.json"),
]

usage_categories_found = False
for path in usage_categories_paths:
    if path.exists():
        print(f"✅ Файл найден: {path.absolute()}")
        usage_categories_found = True
        
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            print(f"   Структура: {list(data.keys())}")
            
            if "years" in data:
                years = data["years"]
                print(f"   ✅ Данные по годам найдены: {list(years.keys())}")
                for year, categories in years.items():
                    print(f"      {year}: {categories}")
            else:
                print(f"   ❌ Ключ 'years' отсутствует в данных")
            
            if "source" in data:
                print(f"   Источник: {data['source']}")
            
            if "ai_analysis" in data:
                ai_info = data["ai_analysis"]
                print(f"   ИИ анализ: использован={ai_info.get('used')}, confidence={ai_info.get('confidence')}")
            
        except Exception as e:
            print(f"   ❌ Ошибка при чтении файла: {e}")
            import traceback
            traceback.print_exc()
        break

if not usage_categories_found:
    print("❌ Файл usage_categories.json НЕ найден")
    print("   Это означает, что функция aggregate_usage_categories() не вернула данные")
    print("   Ожидаемые пути:")
    for path in usage_categories_paths:
        print(f"      {path.absolute()}")

print()

# ============================================================================
# ПРОВЕРКА 3: Вызывается ли distribute_categories_by_quarter()?
# ============================================================================
print("=" * 80)
print("ПРОВЕРКА 3: Вызывается ли distribute_categories_by_quarter()?")
print("=" * 80 + "\n")

# Проверяем код main.py
main_py = Path("eaip_full_skeleton/services/ingest/main.py")
if main_py.exists():
    content = main_py.read_text(encoding="utf-8")
    
    if "distribute_categories_by_quarter" in content:
        print(f"✅ Функция найдена в коде: {main_py}")
        
        # Ищем вызов функции
        if "distribute_categories_by_quarter(" in content:
            print("   ✅ Вызов функции найден в коде")
            
            # Показываем контекст вызова
            lines = content.split("\n")
            for i, line in enumerate(lines):
                if "distribute_categories_by_quarter(" in line:
                    print(f"\n   Контекст вызова (строка {i+1}):")
                    for j in range(max(0, i-2), min(len(lines), i+3)):
                        marker = ">>>" if j == i else "   "
                        print(f"   {marker} {j+1:4d}: {lines[j]}")
                    break
        else:
            print("   ❌ Вызов функции НЕ найден в коде")
    else:
        print(f"❌ Функция distribute_categories_by_quarter не найдена в {main_py}")
else:
    print(f"❌ Файл {main_py} не найден")

print()

# ============================================================================
# ПРОВЕРКА 4: Есть ли данные by_usage в aggregated_data перед генерацией паспорта?
# ============================================================================
print("=" * 80)
print("ПРОВЕРКА 4: Есть ли данные by_usage в aggregated_data?")
print("=" * 80 + "\n")

# Ищем aggregated JSON файлы
aggregated_dirs = [
    Path("data/aggregated"),
    Path("eaip_full_skeleton/infra/data/aggregated"),
]

aggregated_files = []
for dir_path in aggregated_dirs:
    if dir_path.exists():
        aggregated_files.extend(list(dir_path.glob("*_aggregated.json")))

if not aggregated_files:
    print("❌ Aggregated JSON файлы не найдены")
    print("   Это означает, что агрегация данных не выполнялась или не сохранилась")
    print("   Ожидаемые директории:")
    for dir_path in aggregated_dirs:
        print(f"      {dir_path.absolute()}")
else:
    # Анализируем последний файл
    latest_file = max(aggregated_files, key=lambda p: p.stat().st_mtime)
    print(f"Анализируем файл: {latest_file.name}\n")
    
    try:
        data = json.loads(latest_file.read_text(encoding="utf-8"))
        
        resources = data.get("resources", {})
        electricity = resources.get("electricity", {})
        
        if not electricity:
            print("❌ ПРОБЛЕМА: resources.electricity отсутствует в aggregated_data")
        else:
            print(f"✅ resources.electricity найден")
            print(f"   Кварталов: {len(electricity)}\n")
            
            # Проверяем наличие by_usage в каждом квартале
            quarters_with_by_usage = []
            quarters_without_by_usage = []
            
            for quarter, quarter_data in electricity.items():
                by_usage = quarter_data.get("by_usage")
                if by_usage:
                    quarters_with_by_usage.append(quarter)
                    categories = list(by_usage.keys())
                    values = {k: v for k, v in by_usage.items()}
                    print(f"✅ {quarter}: by_usage найден")
                    print(f"   Категории: {categories}")
                    print(f"   Значения: {values}\n")
                else:
                    quarters_without_by_usage.append(quarter)
                    available_keys = list(quarter_data.keys())
                    print(f"❌ {quarter}: by_usage ОТСУТСТВУЕТ")
                    print(f"   Доступные ключи: {available_keys}\n")
            
            print("=" * 80)
            print("ИТОГИ ПРОВЕРКИ 4:")
            print("=" * 80)
            print(f"✅ Кварталов с by_usage: {len(quarters_with_by_usage)}")
            if quarters_with_by_usage:
                print(f"   {', '.join(quarters_with_by_usage)}")
            
            print(f"❌ Кварталов без by_usage: {len(quarters_without_by_usage)}")
            if quarters_without_by_usage:
                print(f"   {', '.join(quarters_without_by_usage)}")
                print("\n   ⚠️ ЭТО ПРОБЛЕМА! Данные by_usage отсутствуют в aggregated_data")
                print("   Валидатор выдаст ошибку при генерации паспорта")
    
    except Exception as e:
        print(f"❌ Ошибка при анализе файла: {e}")
        import traceback
        traceback.print_exc()

print()

# ============================================================================
# ИТОГОВАЯ СВОДКА
# ============================================================================
print("=" * 80)
print("ИТОГОВАЯ СВОДКА")
print("=" * 80 + "\n")

checks = [
    ("Файл pererashod.xlsx загружен", pererashod_found),
    ("Файл usage_categories.json создан", usage_categories_found),
    ("Функция distribute_categories_by_quarter() вызывается", "distribute_categories_by_quarter(" in content if main_py.exists() else False),
    ("Данные by_usage в aggregated_data", len(quarters_with_by_usage) > 0 if 'quarters_with_by_usage' in locals() else False),
]

for check_name, result in checks:
    status = "✅" if result else "❌"
    print(f"{status} {check_name}")

print("\n" + "=" * 80)
print("РЕКОМЕНДАЦИИ")
print("=" * 80 + "\n")

if not pererashod_found:
    print("1. Загрузите файл pererashod.xlsx в систему")
if not usage_categories_found:
    print("2. Проверьте логи функции aggregate_usage_categories()")
    print("   Убедитесь, что файл pererashod.xlsx правильно парсится")
if 'quarters_without_by_usage' in locals() and quarters_without_by_usage:
    print("3. Проверьте, что функция distribute_categories_by_quarter() успешно выполняется")
    print("   Убедитесь, что usage_categories не None перед вызовом")
if 'quarters_with_by_usage' in locals() and not quarters_with_by_usage:
    print("4. КРИТИЧНО: Данные by_usage отсутствуют во всех кварталах")
    print("   Это приведет к ошибке валидации при генерации паспорта")

