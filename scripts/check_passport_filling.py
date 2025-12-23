"""Проверка заполнения энергопаспорта"""
from openpyxl import load_workbook
from pathlib import Path

passport_path = Path("data/aggregated/test_passport_with_electricity.xlsx")
wb = load_workbook(passport_path, data_only=True)

print("=" * 80)
print("ПРОВЕРКА ЗАПОЛНЕНИЯ ЭНЕРГОПАСПОРТА")
print("=" * 80)

# Проверка листа Баланс
if "04_Баланс" in wb.sheetnames:
    ws = wb["04_Баланс"]
    print("\n📊 ЛИСТ '04_Баланс':")
    print("-" * 80)
    for i, row in enumerate(ws.iter_rows(max_row=15), 1):
        values = [str(cell.value)[:30] if cell.value is not None else "None" for cell in row[:6]]
        print(f"Row {i:2d}: {values}")
else:
    print("\n❌ Лист '04_Баланс' не найден!")
    print(f"Доступные листы: {wb.sheetnames}")

# Проверка листа Динамика
if "05_Динамика" in wb.sheetnames:
    ws = wb["05_Динамика"]
    print("\n📈 ЛИСТ '05_Динамика':")
    print("-" * 80)
    for i, row in enumerate(ws.iter_rows(max_row=15), 1):
        values = [str(cell.value)[:30] if cell.value is not None else "None" for cell in row[:7]]
        print(f"Row {i:2d}: {values}")

# Проверка листа Electricity
if "Electricity" in wb.sheetnames:
    ws = wb["Electricity"]
    print("\n⚡ ЛИСТ 'Electricity' (первые 5 строк):")
    print("-" * 80)
    for i, row in enumerate(ws.iter_rows(max_row=5), 1):
        values = [str(cell.value)[:20] if cell.value is not None else "None" for cell in row[:5]]
        print(f"Row {i:2d}: {values}")

print("\n" + "=" * 80)

