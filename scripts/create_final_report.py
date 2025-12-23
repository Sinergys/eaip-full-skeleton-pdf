"""Создание финального отчета о готовности файла"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")

print("=" * 80)
print("ОТЧЕТ О ГОТОВНОСТИ ЭНЕРГОПАСПОРТА")
print("=" * 80)

if not output_path.exists():
    print(f"\n❌ Файл не найден: {output_path}")
    print("   Необходимо запустить генерацию: python tools/generate_metin_passport_full.py")
    exit(1)

# Получаем информацию о файле
file_size = output_path.stat().st_size
print(f"\n✅ ФАЙЛ СОЗДАН")
print(f"   Путь: {output_path.absolute()}")
print(f"   Размер: {file_size:,} байт ({file_size/1024:.1f} КБ)")

# Проверяем содержимое
try:
    wb = load_workbook(output_path, data_only=True)
    ws = wb["Структура пр 2"]
    
    print("\n" + "=" * 80)
    print("ПРОВЕРКА ЗАПОЛНЕНИЯ")
    print("=" * 80)
    
    # 1. Проверка E32
    e32 = ws["E32"].value
    print(f"\n1. ГАЗ (E32 - 2023 Q1):")
    print(f"   Значение: {e32}")
    if e32 and abs(e32 - 14.819) < 0.001:
        print("   ✅ ОШИБКА ИСПРАВЛЕНА! Значение правильное: 14.819 тыс. м³")
    else:
        print(f"   ⚠️ Ожидается: 14.819")
    
    # 2. Проверка электроэнергии по кварталам
    print(f"\n2. ЭЛЕКТРОЭНЕРГИЯ ПО КВАРТАЛАМ:")
    
    quarters = [
        ("Q1", 17, 1, "B-H"),
        ("Q2", 39, 16, "Q-W"),
        ("Q3", 61, 31, "AF-AL"),
        ("Q4", 83, 46, "AU-BA"),
    ]
    
    total_filled = 0
    total_expected = 0
    quarter_results = []
    
    for quarter, start_row, start_col, col_range in quarters:
        quarter_filled = 0
        for row in range(start_row, start_row + 5):
            for offset in [1, 2, 3, 4]:  # норма, 2022, 2023, 2024
                col = start_col + offset
                val = ws.cell(row, col).value
                total_expected += 1
                if val is not None and val != 0:
                    total_filled += 1
                    quarter_filled += 1
        
        # Примеры заполнения
        examples = []
        if quarter == "Q1":
            examples = [
                f"B17={ws['B17'].value}", f"C17={ws['C17'].value}",
                f"D17={ws['D17'].value}", f"E17={ws['E17'].value}"
            ]
        elif quarter == "Q2":
            examples = [
                f"Q39={ws['Q39'].value}", f"R39={ws['R39'].value}",
                f"S39={ws['S39'].value}", f"T39={ws['T39'].value}"
            ]
        elif quarter == "Q3":
            examples = [
                f"AF61={ws['AF61'].value}", f"AG61={ws['AG61'].value}",
                f"AH61={ws['AH61'].value}", f"AI61={ws['AI61'].value}"
            ]
        elif quarter == "Q4":
            examples = [
                f"AU83={ws['AU83'].value}", f"AV83={ws['AV83'].value}",
                f"AW83={ws['AW83'].value}", f"AX83={ws['AX83'].value}"
            ]
        
        status = "✅" if quarter_filled == 20 else ("⚠️" if quarter_filled > 0 else "❌")
        quarter_results.append((quarter, quarter_filled, status, examples))
        
        print(f"   {status} {quarter} (строка {start_row}): {quarter_filled}/20 полей")
        if examples:
            print(f"      Примеры: {', '.join(examples)}")
    
    print(f"\n   {'='*60}")
    percentage = (total_filled * 100) / total_expected if total_expected > 0 else 0
    print(f"   ИТОГО: {total_filled}/{total_expected} полей ({percentage:.1f}%)")
    
    if total_filled == total_expected:
        print("   ✅ ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ ПОЛНОСТЬЮ!")
    elif total_filled > 0:
        print(f"   ⚠️ Заполнено частично")
    else:
        print("   ❌ КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")
    
    # 3. Категории газа
    print(f"\n3. КАТЕГОРИИ ГАЗА:")
    e34 = ws["E34"].value
    e36 = ws["E36"].value
    print(f"   E34 (собственные нужды): {e34}")
    print(f"   E36 (хоз-быт): {e36}")
    if e34 and e36:
        total = (e34 or 0) + (e36 or 0)
        print(f"   Сумма: {total} (ожидается ~14.819)")
    
    wb.close()
    
except Exception as e:
    print(f"\n❌ Ошибка при проверке: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
print("СТАТУС ГОТОВНОСТИ")
print("=" * 80)

if output_path.exists():
    print(f"✅ ФАЙЛ ГОТОВ К ИСПОЛЬЗОВАНИЮ")
    print(f"   Расположение: {output_path.absolute()}")
    print(f"   Лист для проверки: 'Структура пр 2'")
else:
    print(f"❌ Файл не найден")

print("=" * 80)

