#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Поиск Excel файлов с листом '04_Баланс'"""

from openpyxl import load_workbook
from pathlib import Path

files_to_check = [
    "templates/pcm690/new_energy_passport.xlsx",
    "templates/pcm690/template_metin.xlsx",
    "templates/pcm690/energy_passport_template.xlsx",
    "data/source_files/audit_sinergys/energopasport.xlsx",
    "data/source_files/audit_sinergys/Новая папка/EnergyPassport_PKM690_filled.xlsx",
]

print("Поиск файлов с листом '04_Баланс' или 'Баланс':\n")

for file_path in files_to_check:
    path = Path(file_path)
    if not path.exists():
        print(f"❌ {file_path} - не найден")
        continue
    
    try:
        wb = load_workbook(path, read_only=True)
        sheetnames = wb.sheetnames
        
        # Ищем листы с "04_Баланс", "Баланс", "Balans"
        balance_sheets = [s for s in sheetnames if "04_Баланс" in s or "Баланс" in s or "Balans" in s]
        
        if balance_sheets:
            print(f"✅ {path.absolute()}")
            print(f"   Листы: {balance_sheets}")
            print()
        else:
            print(f"   {file_path} - нет листа '04_Баланс'")
            print(f"   Доступные листы: {sheetnames[:5]}...")
            print()
        
        wb.close()
    except Exception as e:
        print(f"❌ {file_path} - ошибка: {e}")
        print()

