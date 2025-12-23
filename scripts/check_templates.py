"""Проверка структуры шаблонов энергопаспорта"""

from openpyxl import load_workbook
from pathlib import Path

templates = [
    Path(r"C:\eaip\templates\pcm690\energy_passport_template.xlsx"),
    Path(r"C:\eaip\templates\pcm690\unified_energy_audit_template.xlsx"),
    Path(r"C:\eaip\data\source_files\metin\EnergyPassport_PKM690_filled.xlsx"),
]

print("=" * 80)
print("СТРУКТУРА ШАБЛОНОВ ЭНЕРГОПАСПОРТА")
print("=" * 80)

for template_path in templates:
    if template_path.exists():
        try:
            wb = load_workbook(template_path, data_only=False)
            print(f"\n📄 {template_path.name}")
            print(f"   Путь: {template_path}")
            print(f"   Листов: {len(wb.sheetnames)}\n")
            for i, name in enumerate(wb.sheetnames, 1):
                ws = wb[name]
                print(f"   {i}. {name:<40} (строк: {ws.max_row}, колонок: {ws.max_column})")
        except Exception as e:
            print(f"\n❌ Ошибка при чтении {template_path.name}: {e}")
    else:
        print(f"\n⚠️  Файл не найден: {template_path}")

print("\n" + "=" * 80)

