"""
Проверка ошибки #ДЕЛ/0! на листе "Динамика ср" в ячейке P28
"""
from openpyxl import load_workbook
from pathlib import Path

template_path = Path(r"C:\Users\DELL\Documents\AUDIT\METIN\Энерг паспорт\ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm")

print(f"📋 ПРОВЕРКА ОШИБКИ НА ЛИСТЕ 'Динамика ср'")
print("=" * 80)

# Загружаем с вычисленными значениями
wb_values = load_workbook(template_path, data_only=True)
ws_dinamika_values = wb_values["Динамика ср"]

# Загружаем с формулами
wb_formulas = load_workbook(template_path, data_only=False)
ws_dinamika_formulas = wb_formulas["Динамика ср"]

print("\n🔍 АНАЛИЗ ЯЧЕЙКИ P28\n")

# Проверяем P28
p28_value = ws_dinamika_values['P28']
p28_formula = ws_dinamika_formulas['P28']

print(f"P28:")
print(f"   Значение: {p28_value.value}")
print(f"   Формула: {p28_formula.value if p28_formula.data_type == 'f' else 'нет формулы'}")

# Проверяем ячейки-источники
print(f"\n📊 Ячейки-источники формулы:\n")

p26_value = ws_dinamika_values['P26']
p26_formula = ws_dinamika_formulas['P26']
p27_value = ws_dinamika_values['P27']
p27_formula = ws_dinamika_formulas['P27']

print(f"P26:")
print(f"   Значение: {p26_value.value}")
print(f"   Тип: {type(p26_value.value).__name__}")
if p26_formula.data_type == 'f':
    print(f"   Формула: {p26_formula.value}")

print(f"\nP27:")
print(f"   Значение: {p27_value.value}")
print(f"   Тип: {type(p27_value.value).__name__}")
if p27_formula.data_type == 'f':
    print(f"   Формула: {p27_formula.value}")

# Анализ проблемы
print(f"\n{'='*80}")
print("🔍 АНАЛИЗ ПРОБЛЕМЫ")
print(f"{'='*80}\n")

if p27_value.value == 0 or p27_value.value is None:
    print("❌ ПРОБЛЕМА НАЙДЕНА:")
    if p27_value.value == 0:
        print("   P27 = 0 → деление на ноль")
    else:
        print("   P27 = пусто/None → деление на ноль")
    print(f"\n   Формула: =ТЕКСТ((P26-P27)/P27*100; \"0%\")")
    print(f"   При P27 = {p27_value.value} → (P26-P27)/P27 = деление на ноль")
else:
    print("⚠️ P27 не равно 0, но ошибка все равно возникает")
    print("   Возможно, проблема в другом месте формулы")

# Проверяем контекст (строки 26, 27, 28)
print(f"\n{'='*80}")
print("📊 КОНТЕКСТ (строки 26-28, колонки O-P-R)")
print(f"{'='*80}\n")

for row in [26, 27, 28]:
    print(f"Строка {row}:")
    for col_letter in ['O', 'P', 'R']:
        cell_value = ws_dinamika_values[f'{col_letter}{row}']
        cell_formula = ws_dinamika_formulas[f'{col_letter}{row}']
        
        value_str = str(cell_value.value) if cell_value.value is not None else "пусто"
        if value_str.startswith('#'):
            value_str = f"❌ {value_str}"
        
        formula_str = ""
        if cell_formula.data_type == 'f':
            formula_str = f" = {cell_formula.value}"
        
        print(f"   {col_letter}{row}: {value_str}{formula_str}")
    print()

# Решение
print(f"{'='*80}")
print("💡 РЕШЕНИЕ")
print(f"{'='*80}\n")

print("📌 ПРОБЛЕМА:")
print("   Формула =ТЕКСТ((P26-P27)/P27*100; \"0%\")")
print("   Выполняет деление на P27, что вызывает #ДЕЛ/0! если P27 = 0\n")

print("🔧 РЕШЕНИЕ 1 (рекомендуется):")
print("   Заменить формулу на защищенную версию:")
print("   =ЕСЛИ(P27=0; \"-\"; ТЕКСТ((P26-P27)/P27*100; \"0%\"))")
print("   Или:")
print("   =ЕСЛИ(P27=0; \"0%\"; ТЕКСТ((P26-P27)/P27*100; \"0%\"))\n")

print("🔧 РЕШЕНИЕ 2 (альтернативное):")
print("   Использовать IFERROR:")
print("   =ЕСЛИОШИБКА(ТЕКСТ((P26-P27)/P27*100; \"0%\"); \"-\")\n")

print("🔧 РЕШЕНИЕ 3 (если нужно показывать 0% при делении на ноль):")
print("   =ЕСЛИ(P27=0; \"0%\"; ТЕКСТ((P26-P27)/P27*100; \"0%\"))\n")

# Проверяем, есть ли похожие формулы на листе
print(f"{'='*80}")
print("🔍 ПОИСК ПОХОЖИХ ФОРМУЛ НА ЛИСТЕ")
print(f"{'='*80}\n")

similar_formulas = []
for row in ws_dinamika_formulas.iter_rows():
    for cell in row:
        if cell.data_type == 'f':
            formula_str = str(cell.value)
            # Ищем формулы с делением и ТЕКСТ
            if 'ТЕКСТ' in formula_str and '/' in formula_str and cell.coordinate != 'P28':
                # Проверяем, есть ли деление на ячейку в той же строке
                row_num = cell.row
                if f'/P{row_num-1}' in formula_str or f'/P{row_num-2}' in formula_str:
                    similar_formulas.append({
                        'cell': cell.coordinate,
                        'formula': formula_str
                    })

if similar_formulas:
    print(f"Найдено похожих формул: {len(similar_formulas)}\n")
    for item in similar_formulas[:10]:  # Первые 10
        print(f"   {item['cell']}: {item['formula'][:100]}")
    if len(similar_formulas) > 10:
        print(f"   ... и еще {len(similar_formulas) - 10} формул")
else:
    print("Похожих формул не найдено")

wb_values.close()
wb_formulas.close()

print(f"\n✅ Анализ завершен")

