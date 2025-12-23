"""
Валидация данных после каждого этапа конвертации
Проверяет целостность данных при конвертации между форматами
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# JSON Schema для валидации данных ПКМ №690
PCM690_SCHEMA = {
    "type": "object",
    "required": ["resources"],
    "properties": {
        "resources": {
            "type": "object",
            "properties": {
                "electricity": {
                    "type": "object",
                    "patternProperties": {
                        "^\\d{4}-Q[1-4]$": {
                            "type": "object",
                            "required": ["year", "quarter", "quarter_totals"],
                            "properties": {
                                "year": {"type": "integer", "minimum": 2020, "maximum": 2030},
                                "quarter": {"type": "integer", "minimum": 1, "maximum": 4},
                                "quarter_totals": {
                                    "type": "object",
                                    "properties": {
                                        "active_kwh": {"type": "number", "minimum": 0},
                                        "reactive_kvarh": {"type": "number", "minimum": 0},
                                        "cost_sum": {"type": "number", "minimum": 0}
                                    }
                                }
                            }
                        }
                    }
                },
                "gas": {
                    "type": "object",
                    "patternProperties": {
                        "^\\d{4}-Q[1-4]$": {
                            "type": "object",
                            "required": ["year", "quarter", "quarter_totals"],
                            "properties": {
                                "year": {"type": "integer"},
                                "quarter": {"type": "integer"},
                                "quarter_totals": {
                                    "type": "object",
                                    "properties": {
                                        "volume_m3": {"type": "number", "minimum": 0},
                                        "cost_sum": {"type": "number", "minimum": 0}
                                    }
                                }
                            }
                        }
                    }
                }
            }
        }
    }
}

def validate_table_structure(table: Dict[str, Any]) -> Dict[str, Any]:
    """Валидация структуры таблицы после извлечения из PDF"""
    errors = []
    warnings = []
    
    rows = table.get("rows", [])
    if not rows:
        errors.append("Таблица пуста")
        return {"valid": False, "errors": errors, "warnings": warnings}
    
    # Проверка на одинаковое количество колонок
    col_counts = [len(row) for row in rows]
    if len(set(col_counts)) > 1:
        warnings.append(f"Неравномерное количество колонок: {min(col_counts)}-{max(col_counts)}")
    
    # Проверка на пустые строки
    empty_rows = sum(1 for row in rows if not any(cell and str(cell).strip() for cell in row))
    if empty_rows > len(rows) * 0.5:
        warnings.append(f"Много пустых строк: {empty_rows}/{len(rows)}")
    
    # Проверка на наличие заголовков
    if len(rows) > 0:
        header_row = rows[0]
        if not any(cell and str(cell).strip() for cell in header_row):
            warnings.append("Заголовок таблицы пуст или отсутствует")
    
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "row_count": len(rows),
        "col_count": max(col_counts) if col_counts else 0
    }

def validate_excel_formulas(workbook_path: Path) -> Dict[str, Any]:
    """Валидация формул в Excel файле"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(workbook_path, data_only=False)
        errors = []
        warnings = []
        formulas_found = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Формула
                        formulas_found += 1
                        # Проверка на ошибки в формуле
                        if cell.value and '#REF!' in str(cell.value):
                            errors.append(f"Ошибка формулы в {sheet_name}!{cell.coordinate}: {cell.value}")
                        elif cell.value and '#VALUE!' in str(cell.value):
                            errors.append(f"Ошибка значения в {sheet_name}!{cell.coordinate}: {cell.value}")
        
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "formulas_count": formulas_found,
            "sheets_count": len(wb.sheetnames)
        }
    except ImportError:
        return {"valid": False, "error": "openpyxl не установлен"}
    except Exception as e:
        return {"valid": False, "error": str(e)}

def validate_json_schema(data: Dict[str, Any], schema: Dict[str, Any] = None) -> Dict[str, Any]:
    """Валидация JSON данных по схеме"""
    if schema is None:
        schema = PCM690_SCHEMA
    
    errors = []
    warnings = []
    
    # Простая валидация (без полной реализации JSON Schema)
    # Проверяем обязательные поля
    if "resources" not in data:
        errors.append("Отсутствует обязательное поле 'resources'")
        return {"valid": False, "errors": errors, "warnings": warnings}
    
    resources = data["resources"]
    
    # Проверка структуры ресурсов
    for resource_type, resource_data in resources.items():
        if not isinstance(resource_data, dict):
            errors.append(f"Ресурс '{resource_type}' должен быть объектом")
            continue
        
        # Проверка кварталов
        for quarter_key, quarter_data in resource_data.items():
            if not isinstance(quarter_data, dict):
                errors.append(f"Квартал '{quarter_key}' в '{resource_type}' должен быть объектом")
                continue
            
            # Проверка обязательных полей
            if "year" not in quarter_data:
                errors.append(f"Отсутствует поле 'year' в {resource_type}/{quarter_key}")
            if "quarter" not in quarter_data:
                errors.append(f"Отсутствует поле 'quarter' в {resource_type}/{quarter_key}")
            if "quarter_totals" not in quarter_data:
                errors.append(f"Отсутствует поле 'quarter_totals' в {resource_type}/{quarter_key}")
            
            # Проверка значений
            if "quarter_totals" in quarter_data:
                totals = quarter_data["quarter_totals"]
                for key, value in totals.items():
                    if isinstance(value, (int, float)) and value < 0:
                        warnings.append(f"Отрицательное значение в {resource_type}/{quarter_key}/{key}: {value}")
    
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "resources_count": len(resources),
        "quarters_count": sum(len(r) for r in resources.values())
    }

def auto_fix_common_errors(data: Dict[str, Any]) -> Dict[str, Any]:
    """Автоматическое исправление частых ошибок распознавания"""
    fixes_applied = []
    
    # Исправление чисел с точками вместо запятых
    def fix_number_format(value):
        if isinstance(value, str):
            # Заменяем запятые на точки в числах
            if ',' in value and value.replace(',', '').replace('.', '').isdigit():
                fixed = value.replace(',', '.')
                fixes_applied.append(f"Исправлен формат числа: {value} -> {fixed}")
                return float(fixed)
        return value
    
    # Рекурсивное исправление
    def fix_recursive(obj):
        if isinstance(obj, dict):
            return {k: fix_recursive(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [fix_recursive(item) for item in obj]
        elif isinstance(obj, str):
            return fix_number_format(obj)
        else:
            return obj
    
    fixed_data = fix_recursive(data)
    
    return {
        "fixed_data": fixed_data,
        "fixes_applied": fixes_applied,
        "fixes_count": len(fixes_applied)
    }

def validate_conversion_pipeline(
    pdf_path: Optional[Path] = None,
    excel_path: Optional[Path] = None,
    json_path: Optional[Path] = None
) -> Dict[str, Any]:
    """Валидация всего пайплайна конвертации"""
    results = {
        "validation_date": datetime.now().isoformat(),
        "stages": {}
    }
    
    # Этап 1: PDF → Таблицы
    if pdf_path and pdf_path.exists():
        print(f"📄 Валидация извлечения таблиц из PDF: {pdf_path.name}")
        # Здесь должна быть валидация таблиц из PDF
        # Пока пропускаем, так как нужен результат extract_tables_from_pdf
        results["stages"]["pdf_to_tables"] = {
            "status": "skipped",
            "message": "Требуется результат extract_tables_from_pdf"
        }
    
    # Этап 2: Таблицы → Excel
    if excel_path and excel_path.exists():
        print(f"📊 Валидация Excel файла: {excel_path.name}")
        excel_validation = validate_excel_formulas(excel_path)
        results["stages"]["excel_validation"] = excel_validation
    
    # Этап 3: Excel → JSON
    if json_path and json_path.exists():
        print(f"📋 Валидация JSON файла: {json_path.name}")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                json_data = json.load(f)
            
            json_validation = validate_json_schema(json_data)
            results["stages"]["json_validation"] = json_validation
            
            # Авто-исправление ошибок
            if json_validation["errors"]:
                print("🔧 Применение авто-исправлений...")
                fixes = auto_fix_common_errors(json_data)
                results["stages"]["auto_fixes"] = fixes
                
                # Повторная валидация после исправлений
                if fixes["fixes_count"] > 0:
                    fixed_validation = validate_json_schema(fixes["fixed_data"])
                    results["stages"]["json_validation_after_fixes"] = fixed_validation
                    
        except Exception as e:
            results["stages"]["json_validation"] = {
                "valid": False,
                "error": str(e)
            }
    
    # Общая оценка
    all_valid = all(
        stage.get("valid", False) 
        for stage in results["stages"].values() 
        if isinstance(stage, dict) and "valid" in stage
    )
    results["overall_valid"] = all_valid
    
    return results

def main():
    """Основная функция"""
    print("=" * 80)
    print("ВАЛИДАЦИЯ ДАННЫХ ПОСЛЕ КОНВЕРТАЦИИ")
    print("=" * 80)
    
    # Ищем файлы для валидации
    aggregated_dir = PROJECT_ROOT / "data" / "aggregated"
    json_files = list(aggregated_dir.glob("*.json"))
    
    if not json_files:
        print("❌ JSON файлы не найдены для валидации")
        return
    
    # Валидируем каждый JSON файл
    all_results = []
    for json_file in json_files[:3]:  # Ограничиваем 3 файлами
        print(f"\n{'='*80}")
        print(f"Валидация: {json_file.name}")
        print(f"{'='*80}")
        
        result = validate_conversion_pipeline(json_path=json_file)
        all_results.append({
            "file": str(json_file),
            "file_name": json_file.name,
            "validation": result
        })
        
        # Вывод результатов
        if "json_validation" in result["stages"]:
            validation = result["stages"]["json_validation"]
            if validation.get("valid"):
                print("✅ Валидация пройдена")
            else:
                print("❌ Ошибки валидации:")
                for error in validation.get("errors", []):
                    print(f"   - {error}")
            
            if validation.get("warnings"):
                print("⚠ Предупреждения:")
                for warning in validation["warnings"]:
                    print(f"   - {warning}")
    
    # Сохраняем результаты
    output_file = PROJECT_ROOT / "data" / "aggregated" / "validation_results.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Результаты сохранены в: {output_file}")
    
    # Сводная статистика
    valid_count = sum(1 for r in all_results if r["validation"]["overall_valid"])
    print(f"\n{'='*80}")
    print("СВОДНАЯ СТАТИСТИКА")
    print(f"{'='*80}")
    print(f"✅ Валидных файлов: {valid_count}/{len(all_results)}")
    print(f"❌ Файлов с ошибками: {len(all_results) - valid_count}/{len(all_results)}")

if __name__ == "__main__":
    main()

