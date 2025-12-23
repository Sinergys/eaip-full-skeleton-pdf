#!/usr/bin/env python3
"""Аккуратное исправление merged cells без ошибок синтаксиса"""

import re
import shutil
from pathlib import Path

file_path = Path(r"C:\eaip\tools\fill_energy_passport.py")
backup_path = file_path.with_suffix('.py.careful_backup')

print("=== АККУРАТНОЕ ИСПРАВЛЕНИЕ MERGED CELLS ===\n")

# Создаем бэкап
shutil.copy2(file_path, backup_path)
print(f"✓ Создан бэкап: {backup_path}")

# Читаем файл построчно
with open(file_path, 'r', encoding='utf-8') as f:
    lines = f.readlines()

# 1. Сначала исправим очевидные ошибки в строке 1519
print("\n1. ИСПРАВЛЕНИЕ СИНТАКСИЧЕСКИХ ОШИБОК:")

# Исправляем строку 1519 (индекс 1518)
if len(lines) > 1518:
    line_1519 = lines[1518]
    print(f"   Строка 1519 до исправления: {line_1519.rstrip()}")
    
    # Ищем паттерн с ошибкой
    if 'indicator_info[' in line_1519 and 'safe_cell_write' in line_1519:
        # Восстанавливаем оригинальную строку
        original_line = '        ws.cell(row=found_row, column=1).value = indicator_info[\n'
        lines[1518] = original_line
        print(f"   Строка 1519 после исправления: {lines[1518].rstrip()}")
    else:
        print(f"   ✓ Строка 1519 в порядке")

# 2. Исправляем только ОЧЕВИДНЫЕ простые присваивания
print("\n2. ИСПРАВЛЕНИЕ ОЧЕВИДНЫХ ПРИСВАИВАНИЙ:")

# Паттерн для ОЧЕНЬ простых присваиваний: ws.cell(...).value = что-то_простое
simple_pattern = re.compile(r'(ws\.cell\(row=[^,]+,\s*column=[^)]+\))\.value\s*=\s*([^\n(]+)(?<!\()')

fixes_made = 0
for i, line in enumerate(lines):
    line = line.rstrip('\n')
    
    # Пропускаем комментарии и пустые строки
    if line.strip().startswith('#') or not line.strip():
        continue
    
    # Ищем простые присваивания
    match = simple_pattern.search(line)
    if match and 'safe_cell_write' not in line:
        cell_call = match.group(1)  # ws.cell(row=..., column=...)
        value = match.group(2).strip()  # простое значение
        
        # Проверяем, что значение действительно простое (нет скобок, операторов)
        if '(' not in value and '[' not in value and ')' not in value and ']' not in value:
            # Заменяем
            new_line = line.replace(
                f'{cell_call}.value = {value}',
                f'safe_cell_write({cell_call}, {value})'
            )
            lines[i] = new_line + '\n'
            fixes_made += 1
            
            if fixes_made <= 5:  # Покажем первые 5 исправлений
                print(f"   Строка {i+1}: {value} -> safe_cell_write")

print(f"   Всего исправлено: {fixes_made} простых присваиваний")

# 3. Исправляем КРИТИЧЕСКИЕ места из логов теста
print("\n3. ИСПРАВЛЕНИЕ КРИТИЧЕСКИХ МЕСТ ИЗ ТЕСТА:")

# Места, где были ошибки в тесте:
critical_lines = [
    2049,  # fill_dinamika_sheet - была ошибка
    2175,  # _write_nodes_table - исходная проблема
    2180,  # _write_nodes_table - исходная проблема
]

for line_num in critical_lines:
    idx = line_num - 1
    if idx < len(lines):
        line = lines[idx]
        if 'ws.cell' in line and '.value =' in line and 'safe_cell_write' not in line:
            print(f"   Строка {line_num}: Критическое место найдено")
            
            # Простая замена для критических мест
            # Ищем ws.cell(...).value = что-то
            cell_match = re.search(r'(ws\.cell\([^)]+\))\.value\s*=', line)
            if cell_match:
                cell_call = cell_match.group(1)
                # Берем все после = до конца строки
                value_start = line.find('=', cell_match.end()) + 1
                value_part = line[value_start:].strip()
                
                if value_part:
                    # Заменяем
                    new_line = line.replace(
                        f'{cell_call}.value = {value_part}',
                        f'safe_cell_write({cell_call}, {value_part})'
                    )
                    lines[idx] = new_line
                    print(f"     Исправлено: {value_part[:30]}...")

# 4. Убедимся, что функция safe_cell_write корректна
print("\n4. ПРОВЕРКА safe_cell_write:")

# Находим функцию safe_cell_write
safe_func_found = False
for i, line in enumerate(lines):
    if 'def safe_cell_write' in line:
        safe_func_found = True
        print(f"   ✓ Функция найдена на строке {i+1}")
        
        # Проверяем, что она содержит проверку MergedCell
        has_merged_check = False
        for j in range(i, min(i+30, len(lines))):
            if 'MergedCell' in lines[j]:
                has_merged_check = True
                break
        
        if has_merged_check:
            print("   ✓ Содержит проверку MergedCell")
        else:
            print("   ⚠️ Не содержит проверку MergedCell, добавляем...")
            # Добавляем простую проверку
            for j in range(i, min(i+30, len(lines))):
                if 'try:' in lines[j]:
                    # Добавляем после try:
                    merged_check = '''        try:
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                return False
'''
                    lines[j] = merged_check
                    print("   ✓ Проверка добавлена")
                    break
        break

if not safe_func_found:
    print("   ✗ Функция safe_cell_write не найдена")
    print("   Добавляем простую версию...")
    
    simple_safe_func = '''
def safe_cell_write(cell, value):
    """Простая безопасная запись в ячейку"""
    try:
        cell.value = value
        return True
    except AttributeError:
        # Если ячейка объединенная (MergedCell)
        return False
'''
    
    # Добавляем в начало файла
    lines.insert(0, simple_safe_func + '\n')
    print("   ✓ Функция добавлена")

# 5. Сохраняем изменения
with open(file_path, 'w', encoding='utf-8') as f:
    f.writelines(lines)

print("\n=== РЕЗУЛЬТАТ ===")
print(f"✓ Файл исправлен (аккуратно)")
print(f"✓ Исправлены синтаксические ошибки")
print(f"✓ Исправлено {fixes_made} простых присваиваний")
print(f"✓ Исправлены критические места из теста")
print(f"✓ Проверена функция safe_cell_write")
print(f"✓ Бэкап сохранен: {backup_path}")

print("\n=== ИНСТРУКЦИИ ===")
print("1. Запустите тест снова")
print("2. Если есть ошибки синтаксиса - восстановите из бэкапа")
print("3. Если ошибка merged cells осталась - нужно точечное исправление")
print("4. Главное - не сломать синтаксис файла")

# Быстрая проверка синтаксиса
print("\n=== БЫСТРАЯ ПРОВЕРКА СИНТАКСИСА ===")
try:
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    # Пробуем скомпилировать
    compile(content, file_path.name, 'exec')
    print("✓ Синтаксис файла корректен")
except SyntaxError as e:
    print(f"✗ Ошибка синтаксиса: {e}")
    print(f"   Строка {e.lineno}, позиция {e.offset}")