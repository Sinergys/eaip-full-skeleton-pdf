#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Правильный анализ структуры файла oborudovanie.xlsx"""

from openpyxl import load_workbook
from pathlib import Path
import re

file_path = Path("data/source_files/audit_sinergys/oborudovanie.xlsx")

wb = load_workbook(file_path, data_only=True)

SECTION_PATTERN = re.compile(r"^\s*(\d+)\.?\s*(.*)$")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"ЛИСТ: {sheet_name} ({ws.max_row} строк, {ws.max_column} столбцов)")
    print(f"{'='*80}\n")
    
    sections = []
    current_section = None
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        cells = [cell.value for cell in row[:10]]
        
        # Пропускаем полностью пустые строки
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        
        col_b = str(cells[1]).strip() if cells[1] else ""
        col_c = str(cells[2]).strip() if cells[2] else ""
        
        # ЗАГОЛОВОК СЕКЦИИ: в столбце B есть текст (название цеха/помещения),
        # а в столбце C ПУСТО (нет названия оборудования)
        is_section = False
        section_name = None
        
        if col_b and not col_c:  # Ключевое условие: B заполнен, C пуст
            # Проверяем паттерн "1. Цех ..." или "2. Помещение ..."
            match = SECTION_PATTERN.match(col_b)
            if match:
                is_section = True
                section_name = col_b
            # Или просто название цеха/помещения без номера
            elif any(keyword in col_b.lower() for keyword in 
                     ["цех", "склад", "офис", "раздевалка", "душ", "производству", 
                      "помещение", "котельная", "подстанция", "тп"]):
                is_section = True
                section_name = col_b
        
        if is_section:
            # Сохраняем предыдущую секцию
            if current_section:
                sections.append(current_section)
            
            # Начинаем новую секцию
            current_section = {
                "name": section_name,
                "start_row": row_idx,
                "header_row": None,
                "items": [],
                "totals": None
            }
            print(f"  [Строка {row_idx:3d}] 🏭 СЕКЦИЯ: {section_name}")
            continue
        
        # Если мы в секции, обрабатываем строки
        if current_section:
            # Заголовок таблицы
            if col_b and col_b.strip().lower() in ["№", "n", "no", "номер"]:
                current_section["header_row"] = row_idx
                print(f"  [Строка {row_idx:3d}] 📋 Заголовок таблицы")
                continue
            
            # Итоговая строка
            if col_b and "итого" in col_b.lower():
                current_section["totals"] = {
                    "row": row_idx,
                    "count": cells[3],
                    "power": cells[5],
                    "vfd": cells[6]
                }
                print(f"  [Строка {row_idx:3d}] 📊 ИТОГО: Кол-во={cells[3]}, Мощность={cells[5]}, ЧП={cells[6]}")
                continue
            
            # Строка с оборудованием (есть название в столбце C)
            if col_c:
                item = {
                    "row": row_idx,
                    "order": col_b,
                    "name": col_c,
                    "quantity": cells[3],
                    "unit_power": cells[4],
                    "total_power": cells[5],
                    "vfd": cells[6]
                }
                current_section["items"].append(item)
                # Показываем только первые 3 и последние 3 элемента для краткости
                if len(current_section["items"]) <= 3 or len(current_section["items"]) > len(current_section["items"]) - 3:
                    print(f"  [Строка {row_idx:3d}] ⚙️  №{col_b}: {col_c[:60]}")
                elif len(current_section["items"]) == 4:
                    print(f"  [Строка {row_idx:3d}] ⚙️  ... (пропущено {len(current_section['items']) - 6} строк) ...")
    
    # Закрываем последнюю секцию
    if current_section:
        sections.append(current_section)
    
    print(f"\n📊 ИТОГО: Найдено секций: {len(sections)}")
    for idx, sec in enumerate(sections, 1):
        print(f"   {idx}. {sec['name']}")
        totals = sec.get('totals')
        end_row = totals.get('row') if totals else '?'
        print(f"      Строки: {sec['start_row']}-{end_row}")
        print(f"      Оборудования: {len(sec['items'])} единиц")
        if totals:
            print(f"      Итого: {totals['count']} шт., {totals['power']} кВт, ЧП: {totals['vfd']}")
        print()

wb.close()

