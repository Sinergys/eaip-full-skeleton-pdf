"""
Проверка шаблона ЭНЕРГО_ПАСПОРТ.xlsx
Проверяет все листы (кроме Sheet1) на наличие формул и значений
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

template_path = Path(r"C:\eaip\data\source_files\audit_sinergys\ЭНЕРГО_ПАСПОРТ_ 1.xlsm")

if not template_path.exists():
    print(f"❌ Файл не найден: {template_path}")
    exit(1)

print(f"📋 Проверка шаблона: {template_path.name}")
print("=" * 80)

wb = load_workbook(template_path, data_only=False)
print(f"✅ Файл загружен")
print(f"📊 Всего листов: {len(wb.sheetnames)}")
print(f"   Листы: {wb.sheetnames}\n")

# Собираем статистику по всем листам (кроме Sheet1)
report = []
problems = []

for sheet_name in wb.sheetnames:
    if sheet_name == "Sheet1":
        continue
    
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"📄 Лист: '{sheet_name}'")
    print(f"{'='*80}")
    
    # Статистика листа
    total_cells = 0
    empty_cells = 0
    value_cells = 0
    formula_cells = 0
    merged_cells = 0
    error_cells = []
    formula_examples = []
    value_examples = []
    
    # Проверяем объединенные ячейки
    merged_ranges = list(ws.merged_cells.ranges)
    merged_cells = len(merged_ranges)
    
    # Проверяем все ячейки с данными
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                empty_cells += 1
            else:
                total_cells += 1
                
                # Проверяем тип ячейки
                if cell.data_type == 'f':  # Формула
                    formula_cells += 1
                    if len(formula_examples) < 5:
                        formula_examples.append({
                            'cell': cell.coordinate,
                            'formula': str(cell.value),
                            'row': cell.row,
                            'col': cell.column
                        })
                elif cell.data_type == 'e':  # Ошибка
                    error_cells.append({
                        'cell': cell.coordinate,
                        'error': str(cell.value),
                        'row': cell.row,
                        'col': cell.column
                    })
                else:  # Значение
                    value_cells += 1
                    if len(value_examples) < 5:
                        value_examples.append({
                            'cell': cell.coordinate,
                            'value': cell.value,
                            'row': cell.row,
                            'col': cell.column
                        })
    
    # Выводим статистику
    print(f"📊 Статистика:")
    print(f"   Всего ячеек с данными: {total_cells}")
    print(f"   Пустых ячеек: {empty_cells}")
    print(f"   Ячеек со значениями: {value_cells}")
    print(f"   Ячеек с формулами: {formula_cells}")
    print(f"   Объединенных областей: {merged_cells}")
    
    if error_cells:
        print(f"   ⚠️ Ячеек с ошибками: {len(error_cells)}")
        problems.append({
            'sheet': sheet_name,
            'type': 'errors',
            'count': len(error_cells),
            'details': error_cells[:10]  # Первые 10 ошибок
        })
    
    # Примеры формул
    if formula_examples:
        print(f"\n📐 Примеры формул (первые {len(formula_examples)}):")
        for ex in formula_examples:
            print(f"   {ex['cell']}: {ex['formula'][:80]}")
    
    # Примеры значений
    if value_examples:
        print(f"\n📊 Примеры значений (первые {len(value_examples)}):")
        for ex in value_examples:
            value_str = str(ex['value'])[:50]
            print(f"   {ex['cell']}: {value_str}")
    
    # Проверка на проблемы
    sheet_problems = []
    
    # Проблема 1: Нет формул вообще
    if formula_cells == 0 and total_cells > 0:
        sheet_problems.append("❌ Нет формул на листе (только значения)")
        problems.append({
            'sheet': sheet_name,
            'type': 'no_formulas',
            'severity': 'high'
        })
    
    # Проблема 2: Много ошибок
    if len(error_cells) > 0:
        sheet_problems.append(f"⚠️ Обнаружены ячейки с ошибками: {len(error_cells)}")
    
    # Проблема 3: Нет данных
    if total_cells == 0:
        sheet_problems.append("⚠️ Лист пустой (нет данных)")
        problems.append({
            'sheet': sheet_name,
            'type': 'empty',
            'severity': 'medium'
        })
    
    # Проблема 4: Только значения, нет формул (для листов, где должны быть формулы)
    formula_expected_sheets = ['Баланс', 'Balans', 'Структура', 'Struktura', 'Динамика', 'Расход']
    if any(keyword.lower() in sheet_name.lower() for keyword in formula_expected_sheets):
        if formula_cells == 0 and value_cells > 0:
            sheet_problems.append("❌ КРИТИЧНО: На листе должны быть формулы, но их нет!")
            problems.append({
                'sheet': sheet_name,
                'type': 'missing_formulas',
                'severity': 'critical'
            })
    
    if sheet_problems:
        print(f"\n⚠️ Проблемы на листе '{sheet_name}':")
        for problem in sheet_problems:
            print(f"   {problem}")
    
    # Сохраняем в отчет
    report.append({
        'sheet': sheet_name,
        'total_cells': total_cells,
        'value_cells': value_cells,
        'formula_cells': formula_cells,
        'error_cells': len(error_cells),
        'merged_cells': merged_cells,
        'has_problems': len(sheet_problems) > 0
    })

wb.close()

# Итоговый отчет
print(f"\n\n{'='*80}")
print("📋 ИТОГОВЫЙ ОТЧЕТ")
print(f"{'='*80}\n")

print("📊 Сводка по листам:")
print(f"{'Лист':<30} {'Ячеек':<10} {'Значений':<10} {'Формул':<10} {'Ошибок':<10} {'Проблемы':<10}")
print("-" * 80)
for item in report:
    problems_mark = "⚠️" if item['has_problems'] else "✅"
    print(f"{item['sheet']:<30} {item['total_cells']:<10} {item['value_cells']:<10} "
          f"{item['formula_cells']:<10} {item['error_cells']:<10} {problems_mark:<10}")

# Критические проблемы
critical_problems = [p for p in problems if p.get('severity') == 'critical']
if critical_problems:
    print(f"\n❌ КРИТИЧЕСКИЕ ПРОБЛЕМЫ ({len(critical_problems)}):")
    for p in critical_problems:
        print(f"   Лист '{p['sheet']}': {p['type']}")

# Высокоприоритетные проблемы
high_problems = [p for p in problems if p.get('severity') == 'high']
if high_problems:
    print(f"\n⚠️ ВЫСОКИЕ ПРОБЛЕМЫ ({len(high_problems)}):")
    for p in high_problems:
        print(f"   Лист '{p['sheet']}': {p['type']}")

# Ошибки в ячейках
error_problems = [p for p in problems if p.get('type') == 'errors']
if error_problems:
    print(f"\n🔴 ЯЧЕЙКИ С ОШИБКАМИ:")
    for p in error_problems:
        print(f"   Лист '{p['sheet']}': {p['count']} ячеек с ошибками")
        for err in p['details'][:5]:  # Первые 5 ошибок
            print(f"      {err['cell']}: {err['error']}")

print(f"\n✅ Проверка завершена")

