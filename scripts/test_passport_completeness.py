"""
Тест заполненности энергопаспорта ПКМ №690

Проверяет процент заполненности всех листов после генерации.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def count_filled_cells(ws: Worksheet, start_row: int = 1, end_row: int = None, start_col: int = 1, end_col: int = None) -> Tuple[int, int]:
    """
    Подсчитывает заполненные и пустые ячейки на листе.
    
    Returns:
        (filled_count, total_count)
    """
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    
    filled = 0
    total = 0
    
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            total += 1
            if cell.value is not None:
                # Проверяем, что это не пустая строка и не только пробелы
                value_str = str(cell.value).strip()
                if value_str and value_str not in ("", "None"):
                    filled += 1
    
    return filled, total


def analyze_sheet(ws: Worksheet, sheet_name: str) -> Dict:
    """Анализирует заполненность одного листа."""
    filled, total = count_filled_cells(ws)
    percentage = (filled / total * 100) if total > 0 else 0.0
    
    # Подсчет ячеек с формулами
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # formula
                formula_count += 1
    
    return {
        "sheet_name": sheet_name,
        "filled_cells": filled,
        "total_cells": total,
        "percentage": round(percentage, 2),
        "formulas": formula_count,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
    }


def test_passport_completeness(
    template_path: Path,
    aggregated_path: Path,
    output_path: Path,
    nodes_json: Path = None,
    envelope_json: Path = None,
    equipment_json: Path = None,
    usage_json: Path = None,
    measures_json: Path = None,
    loss_active: float = 3200.0,
    loss_reactive: float = 13600.0,
    transformer_power: float = 630.0,
) -> Dict:
    """
    Генерирует энергопаспорт и анализирует его заполненность.
    """
    print("=" * 70)
    print("ТЕСТ ЗАПОЛНЕННОСТИ ЭНЕРГОПАСПОРТА ПКМ №690")
    print("=" * 70)
    print()
    
    # Шаг 1: Генерация паспорта
    print("Шаг 1: Генерация энергопаспорта...")
    cmd = [
        sys.executable,
        str(Path(__file__).parent.parent / "tools" / "fill_energy_passport.py"),
        "--template", str(template_path),
        "--aggregated", str(aggregated_path),
        "--output", str(output_path),
        "--loss-active-month", str(loss_active),
        "--loss-reactive-month", str(loss_reactive),
        "--transformer-power", str(transformer_power),
    ]
    
    if nodes_json and nodes_json.exists():
        cmd.extend(["--nodes-json", str(nodes_json)])
    if envelope_json and envelope_json.exists():
        cmd.extend(["--envelope-json", str(envelope_json)])
    if equipment_json and equipment_json.exists():
        cmd.extend(["--equipment-json", str(equipment_json)])
    if usage_json and usage_json.exists():
        cmd.extend(["--usage-json", str(usage_json)])
    if measures_json and measures_json.exists():
        cmd.extend(["--measures-json", str(measures_json)])
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        print("✅ Паспорт успешно сгенерирован")
        print(f"   Файл: {output_path}")
    except subprocess.CalledProcessError as e:
        print(f"❌ Ошибка генерации: {e}")
        print(f"   stdout: {e.stdout}")
        print(f"   stderr: {e.stderr}")
        return {"error": str(e)}
    
    print()
    
    # Шаг 2: Анализ заполненности
    print("Шаг 2: Анализ заполненности...")
    
    if not output_path.exists():
        print(f"❌ Файл не найден: {output_path}")
        return {"error": "Output file not found"}
    
    workbook = load_workbook(output_path, data_only=False)
    
    sheet_results = []
    total_filled = 0
    total_cells = 0
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        analysis = analyze_sheet(ws, sheet_name)
        sheet_results.append(analysis)
        total_filled += analysis["filled_cells"]
        total_cells += analysis["total_cells"]
    
    overall_percentage = (total_filled / total_cells * 100) if total_cells > 0 else 0.0
    
    # Сортируем по проценту заполненности (по убыванию)
    sheet_results.sort(key=lambda x: x["percentage"], reverse=True)
    
    # Вывод результатов
    print()
    print("=" * 70)
    print("РЕЗУЛЬТАТЫ АНАЛИЗА")
    print("=" * 70)
    print()
    print(f"📊 ОБЩАЯ ЗАПОЛНЕННОСТЬ: {overall_percentage:.2f}%")
    print(f"   Заполнено ячеек: {total_filled:,} из {total_cells:,}")
    print()
    print("📋 ДЕТАЛИЗАЦИЯ ПО ЛИСТАМ:")
    print()
    print(f"{'Лист':<30} {'Заполнено':<12} {'Всего':<10} {'%':<8} {'Формулы':<10}")
    print("-" * 70)
    
    for result in sheet_results:
        print(
            f"{result['sheet_name']:<30} "
            f"{result['filled_cells']:<12,} "
            f"{result['total_cells']:<10,} "
            f"{result['percentage']:<8.2f}% "
            f"{result['formulas']:<10,}"
        )
    
    print()
    print("=" * 70)
    print("ТОП-5 САМЫХ ЗАПОЛНЕННЫХ ЛИСТОВ:")
    print("=" * 70)
    for i, result in enumerate(sheet_results[:5], 1):
        print(f"{i}. {result['sheet_name']}: {result['percentage']:.2f}% ({result['filled_cells']:,} ячеек)")
    
    print()
    print("=" * 70)
    print("ТОП-5 МЕНЕЕ ЗАПОЛНЕННЫХ ЛИСТОВ:")
    print("=" * 70)
    for i, result in enumerate(reversed(sheet_results[-5:]), 1):
        print(f"{i}. {result['sheet_name']}: {result['percentage']:.2f}% ({result['filled_cells']:,} ячеек)")
    
    print()
    
    # Проверка ключевых листов
    key_sheets = ["Struktura pr2", "01_Узлы учета", "Balans", "Dinamika sr", "Meropriyatiya", "Equipment", "02_Исходные данные"]
    print("=" * 70)
    print("ПРОВЕРКА КЛЮЧЕВЫХ ЛИСТОВ:")
    print("=" * 70)
    for key_sheet in key_sheets:
        found = False
        for result in sheet_results:
            if key_sheet in result["sheet_name"] or result["sheet_name"] in key_sheet:
                found = True
                status = "✅" if result["percentage"] > 50 else "⚠️" if result["percentage"] > 0 else "❌"
                print(f"{status} {result['sheet_name']}: {result['percentage']:.2f}%")
                break
        if not found:
            print(f"❌ {key_sheet}: не найден")
    
    print()
    
    return {
        "overall_percentage": round(overall_percentage, 2),
        "total_filled": total_filled,
        "total_cells": total_cells,
        "sheets": sheet_results,
        "output_file": str(output_path),
    }


def main():
    """Основная функция для запуска теста."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Тест заполненности энергопаспорта")
    parser.add_argument("--template", required=True, help="Путь к шаблону EnergyPassport_PKM690_Template.xlsx")
    parser.add_argument("--aggregated", required=True, help="Путь к aggregated_energy_*.json")
    parser.add_argument("--output", required=True, help="Путь для сохранения заполненного паспорта")
    parser.add_argument("--nodes-json", help="Опционально: JSON с узлами учёта")
    parser.add_argument("--envelope-json", help="Опционально: JSON с ограждающими конструкциями")
    parser.add_argument("--equipment-json", help="Опционально: JSON с оборудованием")
    parser.add_argument("--usage-json", help="Опционально: JSON с категориями потребления")
    parser.add_argument("--measures-json", help="Опционально: JSON с мероприятиями")
    parser.add_argument("--loss-active", type=float, default=3200.0, help="Потери активной энергии за месяц, кВт·ч")
    parser.add_argument("--loss-reactive", type=float, default=13600.0, help="Потери реактивной энергии за месяц, кВАр·ч")
    parser.add_argument("--transformer-power", type=float, default=630.0, help="Мощность трансформатора, кВА")
    
    args = parser.parse_args()
    
    template_path = Path(args.template)
    aggregated_path = Path(args.aggregated)
    output_path = Path(args.output)
    
    if not template_path.exists():
        print(f"❌ Шаблон не найден: {template_path}")
        return 1
    
    if not aggregated_path.exists():
        print(f"❌ Файл агрегированных данных не найден: {aggregated_path}")
        return 1
    
    result = test_passport_completeness(
        template_path=template_path,
        aggregated_path=aggregated_path,
        output_path=output_path,
        nodes_json=Path(args.nodes_json) if args.nodes_json else None,
        envelope_json=Path(args.envelope_json) if args.envelope_json else None,
        equipment_json=Path(args.equipment_json) if args.equipment_json else None,
        usage_json=Path(args.usage_json) if args.usage_json else None,
        measures_json=Path(args.measures_json) if args.measures_json else None,
        loss_active=args.loss_active,
        loss_reactive=args.loss_reactive,
        transformer_power=args.transformer_power,
    )
    
    if "error" in result:
        return 1
    
    # Сохранение результатов в JSON
    report_path = output_path.parent / f"{output_path.stem}_completeness_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"📄 Отчёт сохранён: {report_path}")
    print()
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

