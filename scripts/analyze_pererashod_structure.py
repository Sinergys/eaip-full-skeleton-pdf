#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Анализ структуры файла pererashod.xlsx - поиск таблиц с данными по категориям использования"""

from openpyxl import load_workbook
from pathlib import Path
import re

file_path = Path("data/source_files/audit_sinergys/pererashod.xlsx")

if not file_path.exists():
    print(f"Файл не найден: {file_path}")
    exit(1)

wb = load_workbook(file_path, data_only=True)

print("=" * 80)
print("АНАЛИЗ СТРУКТУРЫ ФАЙЛА pererashod.xlsx")
print("=" * 80 + "\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"ЛИСТ: {sheet_name} ({ws.max_row} строк, {ws.max_column} столбцов)")
    print(f"{'='*80}\n")
    
    # Ищем таблицы, начинающиеся с R1C1 (A1)
    # Таблицы разделены 2 пустыми строками
    tables = []
    current_table_start = None
    current_table_end = None
    empty_rows_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        cells = [cell.value for cell in row[:20]]  # Первые 20 столбцов
        
        # Проверяем, есть ли данные в строке
        has_data = any(c is not None and str(c).strip() for c in cells)
        
        if not has_data:
            empty_rows_count += 1
            # Если накопилось 2 пустые строки и была таблица - закрываем её
            if empty_rows_count >= 2 and current_table_start:
                current_table_end = row_idx - 2
                tables.append({
                    "number": len(tables) + 1,
                    "start_row": current_table_start,
                    "end_row": current_table_end,
                    "rows_count": current_table_end - current_table_start + 1
                })
                current_table_start = None
                current_table_end = None
        else:
            empty_rows_count = 0
            # Если это начало новой таблицы (первая строка с данными после 2+ пустых строк)
            if current_table_start is None:
                current_table_start = row_idx
            current_table_end = row_idx
    
    # Закрываем последнюю таблицу
    if current_table_start:
        tables.append({
            "number": len(tables) + 1,
            "start_row": current_table_start,
            "end_row": current_table_end or ws.max_row,
            "rows_count": (current_table_end or ws.max_row) - current_table_start + 1
        })
    
    print(f"Найдено таблиц: {len(tables)}\n")
    
    # Анализируем каждую таблицу
    for table in tables:
        print(f"{'='*70}")
        print(f"ТАБЛИЦА {table['number']} (строки {table['start_row']}-{table['end_row']}, {table['rows_count']} строк)")
        print(f"{'='*70}")
        
        # Показываем первые 15 строк таблицы
        print("\nПервые 15 строк таблицы:")
        for i, row in enumerate(ws.iter_rows(min_row=table['start_row'], max_row=min(table['start_row'] + 14, table['end_row'])), 1):
            cells = [cell.value for cell in row[:20]]
            # Показываем только непустые ячейки
            non_empty = [str(c) if c is not None else "" for c in cells]
            non_empty_str = " | ".join([f"{i:2d}:{val[:30]}" for i, val in enumerate(non_empty, 1) if val.strip()])
            if non_empty_str:
                print(f"  Строка {table['start_row'] + i - 1:3d}: {non_empty_str}")
        
        # Ищем ключевые слова для определения типа таблицы
        keywords_found = []
        for row_idx in range(table['start_row'], min(table['start_row'] + 20, table['end_row'] + 1)):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
            row_text = " ".join([str(cell.value) if cell.value else "" for cell in row[:20]]).lower()
            
            # Ключевые слова для категорий использования
            if any(kw in row_text for kw in ["тех", "технологич", "тех-потер"]):
                keywords_found.append(f"Строка {row_idx}: технологические/тех-потери")
            if any(kw in row_text for kw in ["хоз", "бытов", "хоз-быт"]):
                keywords_found.append(f"Строка {row_idx}: хоз-бытовые")
            if any(kw in row_text for kw in ["производств", "произв"]):
                keywords_found.append(f"Строка {row_idx}: производственные")
            if any(kw in row_text for kw in ["собствен", "с.н.", "нужды"]):
                keywords_found.append(f"Строка {row_idx}: собственные нужды")
            if any(kw in row_text for kw in ["год", "2022", "2023", "2024"]):
                keywords_found.append(f"Строка {row_idx}: годы/периоды")
            if any(kw in row_text for kw in ["итого", "сумма", "всего"]):
                keywords_found.append(f"Строка {row_idx}: итоги")
        
        if keywords_found:
            print(f"\n🔍 Найдены ключевые слова:")
            for kw in keywords_found[:10]:  # Показываем первые 10
                print(f"   {kw}")
        
        # Проверяем, является ли это 4-й таблицей (с данными по типу расхода)
        if table['number'] == 4:
            print(f"\n⭐ ЭТО 4-Я ТАБЛИЦА - содержит данные по типу расхода!")
            print(f"   Нужно использовать ИИ для анализа структуры")
        
        print()

wb.close()

