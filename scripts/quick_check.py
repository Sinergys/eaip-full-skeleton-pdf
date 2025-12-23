# Простая проверка файла
import sys
from pathlib import Path

# Сохраняем вывод в файл
output_file = Path("test_output_metin/check_result.txt")
with open(output_file, 'w', encoding='utf-8') as f:
    sys.stdout = f
    
    try:
        from openpyxl import load_workbook
        
        p = Path("test_output_metin/generated_passport_fixed.xlsx")
        
        if not p.exists():
            print("ERROR: File not found")
            sys.exit(1)
        
        print(f"OK: File exists, size: {p.stat().st_size} bytes")
    
        wb = load_workbook(p, data_only=True)
        ws = wb["Структура пр 2"]
        
        # Проверяем ключевые ячейки
        checks = {
            "E32": ws["E32"].value,
            "B17": ws["B17"].value,
            "C17": ws["C17"].value,
            "Q39": ws["Q39"].value,
            "R39": ws["R39"].value,
            "AF61": ws["AF61"].value,
            "AG61": ws["AG61"].value,
            "AU83": ws["AU83"].value,
            "AV83": ws["AV83"].value,
        }
        
        print("\nValues:")
        for cell, val in checks.items():
            print(f"  {cell}: {val}")
        
        # Подсчет заполненных
        filled = 0
        total = 0
        
        quarters = [(17, 1), (39, 16), (61, 31), (83, 46)]
        for start_row, start_col in quarters:
            for row in range(start_row, start_row + 5):
                for offset in [1, 2, 3, 4]:
                    col = start_col + offset
                    val = ws.cell(row, col).value
                    total += 1
                    if val is not None and val != 0:
                        filled += 1
        
        print(f"\nFilled: {filled}/{total} ({filled*100/total:.1f}%)")
        
        if filled == total:
            print("SUCCESS: All quarters filled!")
        elif filled > 0:
            print("PARTIAL: Some quarters filled")
        else:
            print("ERROR: No quarters filled")
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# Восстанавливаем stdout
sys.stdout = sys.__stdout__
print(f"Report saved to: {output_file}")

