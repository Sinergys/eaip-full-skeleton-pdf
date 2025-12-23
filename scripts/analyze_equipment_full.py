#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Полный анализ структуры файла oborudovanie.xlsx"""

from openpyxl import load_workbook
from pathlib import Path
import re

file_path = Path("data/source_files/audit_sinergys/oborudovanie.xlsx")

wb = load_workbook(file_path, data_only=True)

SECTION_PATTERN = re.compile(r"^\s*(\d+)\.?\s*(.*)$")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"ЛИСТ: {sheet_name} ({ws.max_row} строк, {ws.max_column} столбцов)")
    print(f"{'='*80}\n")
    
    sections = []
    current_section = None
    section_start_row = None
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        cells = [cell.value for cell in row[:10]]
        
        # Пропускаем полностью пустые строки
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        
        col_b = str(cells[1]).strip() if cells[1] else ""
        col_c = str(cells[2]).strip() if cells[2] else ""
        
        # Проверяем, является ли это заголовком секции
        is_section = False
        section_name = None
        
        if col_b:
            # Паттерн "1. Цех ..." или "2. Помещение ..."
            match = SECTION_PATTERN.match(col_b)
            if match:
                is_section = True
                section_name = col_b
            # Или просто название цеха/помещения без номера
            elif not col_c and any(keyword in col_b.lower() for keyword in 
                                   ["цех", "склад", "офис", "раздевалка", "душ", "производству", 
                                    "помещение", "котельная", "подстанция", "тп"]):
                is_section = True
                section_name = col_b
        
        if is_section:
            if current_section:
                sections.append({
                    "name": current_section["name"],
                    "start_row": current_section["start_row"],
                    "end_row": row_idx - 1,
                    "items_count": current_section["items_count"]
                })
            current_section = {
                "name": section_name,
                "start_row": row_idx,
                "items_count": 0
            }
            print(f"  [Строка {row_idx:3d}] 🏭 СЕКЦИЯ: {section_name}")
            continue
        
        # Если мы в секции, проверяем строки
        if current_section:
            # Заголовок таблицы
            if col_b and col_b.strip().lower() in ["№", "n", "no", "номер"]:
                print(f"  [Строка {row_idx:3d}] 📋 Заголовок таблицы")
                continue
            
            # Итоговая строка
            if col_b and "итого" in col_b.lower():
                print(f"  [Строка {row_idx:3d}] 📊 ИТОГО: Кол-во={cells[3]}, Мощность={cells[5]}, ЧП={cells[6]}")
                continue
            
            # Строка с оборудованием
            if col_c:
                current_section["items_count"] += 1
                print(f"  [Строка {row_idx:3d}] ⚙️  №{col_b}: {col_c[:50]}...")
    
    # Закрываем последнюю секцию
    if current_section:
        sections.append({
            "name": current_section["name"],
            "start_row": current_section["start_row"],
            "end_row": ws.max_row,
            "items_count": current_section["items_count"]
        })
    
    print(f"\n📊 ИТОГО: Найдено секций: {len(sections)}")
    for idx, sec in enumerate(sections, 1):
        print(f"   {idx}. {sec['name']} (строки {sec['start_row']}-{sec['end_row']}, {sec['items_count']} единиц оборудования)")

wb.close()

