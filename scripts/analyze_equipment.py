#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Анализ структуры файла oborudovanie.xlsx"""

from openpyxl import load_workbook
from pathlib import Path
import json

file_path = Path("data/source_files/audit_sinergys/oborudovanie.xlsx")

if not file_path.exists():
    print(f"Файл не найден: {file_path}")
    exit(1)

wb = load_workbook(file_path, data_only=True)
print(f"Листы в файле: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=" * 80)
    print(f"ЛИСТ: {sheet_name}")
    print(f"Размер: {ws.max_row} строк, {ws.max_column} столбцов")
    print(f"=" * 80)
    
    # Анализируем первые 50 строк для понимания структуры
    print("\nСТРУКТУРА ФАЙЛА (первые 50 строк):\n")
    
    sections_found = []
    current_section = None
    in_header = False
    
    for i, row in enumerate(list(ws.iter_rows())[:50], 1):
        cells = [cell.value for cell in row[:10]]  # Первые 10 столбцов
        
        # Пропускаем полностью пустые строки
        if not any(c is not None and str(c).strip() for c in cells):
            if current_section:
                print(f"  [Пустая строка после секции '{current_section}']")
            continue
        
        # Получаем значения из столбцов B (индекс 1) и C (индекс 2)
        col_b = str(cells[1]).strip() if cells[1] else ""
        col_c = str(cells[2]).strip() if cells[2] else ""
        
        # Проверяем, является ли это заголовком секции
        is_section = False
        if col_b and not col_c:
            # Паттерн: "1. Цех ..." или просто название цеха/помещения
            if any(keyword in col_b.lower() for keyword in ["цех", "склад", "офис", "раздевалка", "душ", "производству", "помещение"]):
                is_section = True
            elif col_b and col_b[0].isdigit() and "." in col_b:
                is_section = True
        
        if is_section:
            current_section = col_b
            sections_found.append(current_section)
            print(f"\n{'='*60}")
            print(f"СЕКЦИЯ [{i}]: {col_b}")
            print(f"{'='*60}")
            in_header = False
            continue
        
        # Проверяем, является ли это заголовком таблицы
        if col_b and col_b.strip().lower() in ["№", "n", "no"]:
            print(f"  [Строка {i}] ЗАГОЛОВОК ТАБЛИЦЫ: {[str(c) if c else '' for c in cells[:7]]}")
            in_header = True
            continue
        
        # Проверяем, является ли это итоговой строкой
        if col_b and "итого" in col_b.lower():
            print(f"  [Строка {i}] ИТОГО: {[str(c) if c else '' for c in cells[:7]]}")
            continue
        
        # Если есть название оборудования (col_c), это строка с оборудованием
        if col_c and current_section:
            print(f"  [Строка {i}] ОБОРУДОВАНИЕ: №={col_b}, Название={col_c}, "
                  f"Кол-во={cells[3]}, Мощн-едн={cells[4]}, Общ-мощн={cells[5]}, ЧП={cells[6]}")
        elif current_section:
            print(f"  [Строка {i}] ДРУГАЯ СТРОКА: {[str(c) if c else '' for c in cells[:7]]}")
    
    print(f"\n\nНАЙДЕНО СЕКЦИЙ: {len(sections_found)}")
    for idx, sec in enumerate(sections_found, 1):
        print(f"  {idx}. {sec}")
    
    print("\n" + "="*80 + "\n")

wb.close()

