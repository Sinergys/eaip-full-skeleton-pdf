"""
Проверка файла с вычисленными значениями для обнаружения ошибок Excel
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

template_path = Path(r"C:\Users\DELL\Documents\AUDIT\METIN\Энерг паспорт\ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm")

print(f"📋 ПРОВЕРКА ФАЙЛА С ВЫЧИСЛЕННЫМИ ЗНАЧЕНИЯМИ")
print("=" * 80)
print(f"📂 Файл: {template_path.name}\n")

# Загружаем с вычисленными значениями
wb = load_workbook(template_path, data_only=True)
sheet_names = wb.sheetnames

print(f"✅ Файл загружен (data_only=True)")
print(f"📊 Всего листов: {len(sheet_names)}\n")

# Специальная проверка листа "Баланс"
if "Баланс" not in sheet_names:
    print("❌ Лист 'Баланс' не найден!")
    exit(1)

ws_balans = wb["Баланс"]
print(f"{'='*80}")
print(f"🔍 ДЕТАЛЬНАЯ ПРОВЕРКА ЛИСТА 'Баланс'")
print(f"{'='*80}\n")

# Ищем все ячейки с ошибками
errors_found = []
error_types = defaultdict(list)

print("🔍 Поиск ячеек с ошибками Excel...\n")

for row in ws_balans.iter_rows():
    for cell in row:
        if cell.value is not None:
            value_str = str(cell.value)
            
            # Проверяем на ошибки Excel
            if value_str.startswith('#'):
                error_info = {
                    'cell': cell.coordinate,
                    'error': value_str,
                    'row': cell.row,
                    'col': cell.column,
                    'col_letter': cell.column_letter
                }
                errors_found.append(error_info)
                error_types[value_str].append(error_info)

# Выводим результаты
if errors_found:
    print(f"❌ НАЙДЕНО ОШИБОК: {len(errors_found)}\n")
    
    print(f"📌 Ошибки по типам:")
    for error_type, errors_list in sorted(error_types.items()):
        print(f"\n   {error_type}: {len(errors_list)} ячеек")
        
        # Показываем примеры
        print(f"   Примеры ячеек:")
        for error in errors_list[:10]:
            print(f"      • {error['cell']} (строка {error['row']}, колонка {error['col_letter']})")
        if len(errors_list) > 10:
            print(f"      ... и еще {len(errors_list) - 10} ячеек")
        
        # Анализ паттернов
        rows_with_errors = set(e['row'] for e in errors_list)
        cols_with_errors = set(e['col_letter'] for e in errors_list)
        print(f"   Затронуто строк: {len(rows_with_errors)}, колонок: {len(cols_with_errors)}")
        
        # Показываем диапазоны ошибок
        if rows_with_errors:
            min_row = min(rows_with_errors)
            max_row = max(rows_with_errors)
            print(f"   Диапазон строк: {min_row} - {max_row}")
    
    # Анализ формул, вызывающих ошибки
    print(f"\n{'='*80}")
    print(f"📐 АНАЛИЗ ФОРМУЛ, ВЫЗЫВАЮЩИХ ОШИБКИ")
    print(f"{'='*80}\n")
    
    # Загружаем файл снова, но с формулами
    wb_formulas = load_workbook(template_path, data_only=False)
    ws_balans_formulas = wb_formulas["Баланс"]
    
    print("Формулы в ячейках с ошибками:\n")
    for error in errors_found[:20]:  # Первые 20
        cell = ws_balans_formulas[error['cell']]
        if cell.data_type == 'f':
            formula = str(cell.value)
            print(f"   {error['cell']} ({error['error']}):")
            print(f"      {formula[:200]}")
            
            # Анализ формулы
            if '#DIV/0!' in error['error']:
                # Ищем деление
                if '/' in formula:
                    print(f"      ⚠️ Обнаружено деление - возможно, делитель равен нулю")
            elif '#REF!' in error['error']:
                # Ищем ссылки
                if "'" in formula or "!" in formula:
                    print(f"      ⚠️ Обнаружены ссылки - возможно, ссылка на несуществующую ячейку/лист")
            print()
    
    wb_formulas.close()
    
    # Рекомендации по исправлению
    print(f"\n{'='*80}")
    print(f"💡 РЕКОМЕНДАЦИИ ПО ИСПРАВЛЕНИЮ")
    print(f"{'='*80}\n")
    
    if '#DIV/0!' in error_types:
        print("1. ❌ ОШИБКИ ДЕЛЕНИЯ НА НОЛЬ (#DIV/0!):")
        print(f"   Найдено: {len(error_types['#DIV/0!'])} ячеек")
        print("   Решение:")
        print("   • Заменить формулы вида =A1/B1 на =IF(B1=0, 0, A1/B1)")
        print("   • Или использовать =IFERROR(A1/B1, 0)")
        print("   • Проверить исходные данные - возможно, делитель должен быть заполнен\n")
    
    if '#REF!' in error_types:
        print("2. ❌ ОШИБКИ ССЫЛОК (#REF!):")
        print(f"   Найдено: {len(error_types['#REF!'])} ячеек")
        print("   Решение:")
        print("   • Проверить ссылки на другие листы - возможно, листы переименованы")
        print("   • Проверить ссылки на ячейки - возможно, строки/колонки были удалены")
        print("   • Исправить имена листов в формулах (убедиться в пробелах в конце)\n")
    
    if '#VALUE!' in error_types:
        print("3. ❌ ОШИБКИ ЗНАЧЕНИЙ (#VALUE!):")
        print(f"   Найдено: {len(error_types['#VALUE!'])} ячеек")
        print("   Решение:")
        print("   • Проверить типы данных в аргументах функций")
        print("   • Убедиться, что текстовые значения не используются в числовых операциях")
        print("   • Использовать функции преобразования (VALUE, TEXT)\n")
    
    if '#N/A' in error_types:
        print("4. ⚠️ ЗНАЧЕНИЯ НЕДОСТУПНЫ (#N/A):")
        print(f"   Найдено: {len(error_types['#N/A'])} ячеек")
        print("   Решение:")
        print("   • Проверить функции поиска (VLOOKUP, HLOOKUP, MATCH)")
        print("   • Убедиться, что искомые значения существуют")
        print("   • Использовать IFERROR для обработки отсутствующих значений\n")
    
else:
    print("✅ Ошибок Excel не обнаружено!")
    print("   Все формулы вычисляются корректно")

wb.close()

print(f"\n✅ Проверка завершена")

