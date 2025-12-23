"""
Полная проверка файла ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm
Специальная проверка листа "Баланс" на ошибки
Роль: Senior Excel Analyst / Formula Debugger
"""
from openpyxl import load_workbook
from pathlib import Path
import re
from collections import defaultdict

template_path = Path(r"C:\Users\DELL\Documents\AUDIT\METIN\Энерг паспорт\ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm")

if not template_path.exists():
    print(f"❌ Файл не найден: {template_path}")
    exit(1)

print(f"📋 ПОЛНАЯ ПРОВЕРКА ФАЙЛА")
print("=" * 80)
print(f"📂 Файл: {template_path.name}")
print("=" * 80)

wb = load_workbook(template_path, data_only=False)
sheet_names = wb.sheetnames

print(f"\n✅ Файл загружен")
print(f"📊 Всего листов: {len(sheet_names)}")
print(f"   Листы: {sheet_names}\n")

# Собираем все данные
all_remarks = []
file_statistics = {}
balans_errors = []
balans_warnings = []

def check_cell_errors(cell):
    """Проверка ячейки на ошибки Excel"""
    errors = []
    
    if cell.data_type == 'e':  # Ошибка
        error_value = str(cell.value)
        error_types = {
            '#NULL!': 'Пересечение двух областей, которые не пересекаются',
            '#DIV/0!': 'Деление на ноль',
            '#VALUE!': 'Неверный тип аргумента',
            '#REF!': 'Недействительная ссылка на ячейку',
            '#NAME?': 'Неверное имя',
            '#NUM!': 'Неверное число',
            '#N/A': 'Значение недоступно',
            '#SPILL!': 'Переполнение массива',
            '#CALC!': 'Ошибка вычисления',
            '#UNKNOWN!': 'Неизвестная ошибка'
        }
        
        error_desc = error_types.get(error_value, 'Неизвестная ошибка')
        errors.append({
            'cell': cell.coordinate,
            'error': error_value,
            'description': error_desc,
            'row': cell.row,
            'col': cell.column
        })
    
    return errors

def check_formula_errors(formula_str, cell_coord, sheet_name, all_sheets):
    """Детальная проверка формулы на ошибки"""
    errors = []
    warnings = []
    
    # Проверка ссылок на другие листы
    sheet_refs = re.findall(r"'([^']+)'!", formula_str)
    for ref_sheet in sheet_refs:
        ref_sheet_exact = ref_sheet
        ref_sheet_clean = ref_sheet.strip()
        
        if ref_sheet_exact not in all_sheets:
            found = False
            for sheet_name_check in all_sheets:
                if sheet_name_check.strip() == ref_sheet_clean:
                    found = True
                    break
            
            if not found:
                errors.append({
                    'type': 'broken_reference',
                    'cell': cell_coord,
                    'message': f"Ссылка на несуществующий лист: '{ref_sheet_exact}'",
                    'formula': formula_str[:150]
                })
    
    # Проверка на циклические ссылки (базовая)
    if re.search(r'=\$?[A-Z]+\$?\d+.*\$?[A-Z]+\$?\d+', formula_str):
        # Простая проверка на самоссылку
        cell_col = re.match(r'([A-Z]+)', cell_coord)
        if cell_col:
            col_letter = cell_col.group(1)
            if f'{col_letter}' in formula_str and re.search(rf'{col_letter}\d+', formula_str):
                row_num = re.search(r'\d+', cell_coord)
                if row_num:
                    row = row_num.group(0)
                    if f'{col_letter}{row}' in formula_str:
                        warnings.append({
                            'type': 'potential_circular',
                            'cell': cell_coord,
                            'message': 'Возможная циклическая ссылка (самоссылка)',
                            'formula': formula_str[:150]
                        })
    
    # Проверка синтаксиса
    open_brackets = formula_str.count('(')
    close_brackets = formula_str.count(')')
    if open_brackets != close_brackets:
        errors.append({
            'type': 'syntax_error',
            'cell': cell_coord,
            'message': f"Несбалансированные скобки: открывающих {open_brackets}, закрывающих {close_brackets}",
            'formula': formula_str[:150]
        })
    
    return errors, warnings

# Проверяем все листы
for sheet_name in sheet_names:
    if sheet_name == "Sheet1":
        continue
    
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"📄 Лист: '{sheet_name}'")
    print(f"{'='*80}")
    
    sheet_errors = []
    sheet_warnings = []
    total_cells = 0
    error_cells = 0
    formula_cells = 0
    value_cells = 0
    
    # Проверяем все ячейки
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                total_cells += 1
                
                # Проверка на ошибки Excel
                cell_errors = check_cell_errors(cell)
                if cell_errors:
                    error_cells += len(cell_errors)
                    sheet_errors.extend(cell_errors)
                    if sheet_name == "Баланс":
                        balans_errors.extend(cell_errors)
                
                # Проверка формул
                if cell.data_type == 'f':
                    formula_cells += 1
                    formula_str = str(cell.value)
                    formula_errors, formula_warnings = check_formula_errors(
                        formula_str, cell.coordinate, sheet_name, sheet_names
                    )
                    sheet_errors.extend(formula_errors)
                    sheet_warnings.extend(formula_warnings)
                    if sheet_name == "Баланс":
                        balans_errors.extend(formula_errors)
                        balans_warnings.extend(formula_warnings)
                else:
                    value_cells += 1
    
    # Статистика листа
    print(f"📊 Статистика:")
    print(f"   Ячеек с данными: {total_cells}")
    print(f"   Ячеек со значениями: {value_cells}")
    print(f"   Ячеек с формулами: {formula_cells}")
    print(f"   Ячеек с ошибками: {error_cells}")
    
    # Выводим ошибки
    if sheet_errors:
        print(f"\n❌ ОШИБКИ ({len(sheet_errors)}):")
        for error in sheet_errors[:10]:  # Первые 10
            print(f"   • {error['cell']}: {error.get('error', error.get('message', 'Ошибка'))}")
            if 'description' in error:
                print(f"     {error['description']}")
        if len(sheet_errors) > 10:
            print(f"   ... и еще {len(sheet_errors) - 10} ошибок")
    
    # Выводим предупреждения
    if sheet_warnings:
        print(f"\n⚠️ ПРЕДУПРЕЖДЕНИЯ ({len(sheet_warnings)}):")
        for warning in sheet_warnings[:5]:  # Первые 5
            print(f"   • {warning['cell']}: {warning['message']}")
        if len(sheet_warnings) > 5:
            print(f"   ... и еще {len(sheet_warnings) - 5} предупреждений")
    
    if not sheet_errors and not sheet_warnings:
        print(f"\n✅ Ошибок не найдено")
    
    file_statistics[sheet_name] = {
        'total_cells': total_cells,
        'value_cells': value_cells,
        'formula_cells': formula_cells,
        'error_cells': error_cells,
        'errors': len(sheet_errors),
        'warnings': len(sheet_warnings)
    }
    
    all_remarks.extend(sheet_errors)
    all_remarks.extend(sheet_warnings)

# Специальная проверка листа "Баланс"
if "Баланс" in sheet_names:
    print(f"\n\n{'='*80}")
    print("🔍 ДЕТАЛЬНАЯ ПРОВЕРКА ЛИСТА 'Баланс'")
    print(f"{'='*80}")
    
    ws_balans = wb["Баланс"]
    
    # Группируем ошибки по типам
    errors_by_type = defaultdict(list)
    for error in balans_errors:
        error_type = error.get('error', error.get('type', 'unknown'))
        errors_by_type[error_type].append(error)
    
    print(f"\n📊 Статистика ошибок на листе 'Баланс':")
    print(f"   Всего ошибок: {len(balans_errors)}")
    print(f"   Всего предупреждений: {len(balans_warnings)}")
    
    if errors_by_type:
        print(f"\n📌 Ошибки по типам:")
        for error_type, errors_list in sorted(errors_by_type.items()):
            print(f"   {error_type}: {len(errors_list)}")
            # Показываем примеры
            for error in errors_list[:5]:
                print(f"      • {error['cell']}: {error.get('description', error.get('message', ''))}")
            if len(errors_list) > 5:
                print(f"      ... и еще {len(errors_list) - 5} ошибок")
    
    # Анализ формул на листе "Баланс"
    print(f"\n📐 Анализ формул:")
    formula_patterns = defaultdict(int)
    for row in ws_balans.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula_str = str(cell.value)
                # Анализируем паттерны
                if 'SUM' in formula_str:
                    formula_patterns['SUM'] += 1
                if 'IF' in formula_str:
                    formula_patterns['IF'] += 1
                if "'" in formula_str:  # Ссылки на другие листы
                    formula_patterns['cross_sheet_references'] += 1
                if '!' in formula_str:
                    formula_patterns['sheet_references'] += 1
    
    if formula_patterns:
        print(f"   Использование функций:")
        for pattern, count in sorted(formula_patterns.items()):
            print(f"      {pattern}: {count}")

wb.close()

# ИТОГОВЫЙ ОТЧЕТ
print(f"\n\n{'='*80}")
print("📋 ИТОГОВЫЙ ОТЧЕТ ПО ФАЙЛУ")
print(f"{'='*80}\n")

print("📊 Сводка по листам:")
print(f"{'Лист':<30} {'Ячеек':<10} {'Формул':<10} {'Ошибок':<10} {'Предупр.':<12} {'Статус':<10}")
print("-" * 85)
for sheet_name, stats in sorted(file_statistics.items()):
    status = "❌" if stats['errors'] > 0 else "⚠️" if stats['warnings'] > 0 else "✅"
    print(f"{sheet_name:<30} {stats['total_cells']:<10} {stats['formula_cells']:<10} "
          f"{stats['errors']:<10} {stats['warnings']:<12} {status:<10}")

# Сводка по ошибкам
total_errors = sum(s['errors'] for s in file_statistics.values())
total_warnings = sum(s['warnings'] for s in file_statistics.values())

print(f"\n📌 ОБЩАЯ СТАТИСТИКА:")
print(f"   Всего ошибок в файле: {total_errors}")
print(f"   Всего предупреждений: {total_warnings}")
print(f"   Ошибок на листе 'Баланс': {len(balans_errors)}")
print(f"   Предупреждений на листе 'Баланс': {len(balans_warnings)}")

# Рекомендации
print(f"\n💡 РЕКОМЕНДАЦИИ:")
if balans_errors:
    print(f"   1. ❌ КРИТИЧНО: На листе 'Баланс' найдено {len(balans_errors)} ошибок")
    print(f"      Необходимо исправить перед использованием файла")
    
    # Группируем по типам для рекомендаций
    error_types = defaultdict(int)
    for error in balans_errors:
        error_type = error.get('error', error.get('type', 'unknown'))
        error_types[error_type] += 1
    
    print(f"\n   Типы ошибок на листе 'Баланс':")
    for error_type, count in sorted(error_types.items(), key=lambda x: x[1], reverse=True):
        print(f"      • {error_type}: {count} ошибок")
        
        # Специфичные рекомендации
        if error_type == '#DIV/0!':
            print(f"        → Решение: Добавить проверку делителя (IF или IFERROR)")
        elif error_type == '#REF!':
            print(f"        → Решение: Проверить ссылки на удаленные ячейки/листы")
        elif error_type == '#VALUE!':
            print(f"        → Решение: Проверить типы данных в аргументах функций")
        elif error_type == 'broken_reference':
            print(f"        → Решение: Исправить имена листов в ссылках")
else:
    print(f"   ✅ Критических ошибок на листе 'Баланс' не найдено")

if balans_warnings:
    print(f"\n   2. ⚠️ Найдено {len(balans_warnings)} предупреждений на листе 'Баланс'")
    print(f"      Рекомендуется проверить")

print(f"\n✅ Проверка завершена")

