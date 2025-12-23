"""Анализ незаполненных данных в энергопаспорте"""
from openpyxl import load_workbook
from pathlib import Path
import json

passport_path = Path("data/aggregated/test_passport_with_electricity.xlsx")
wb = load_workbook(passport_path, data_only=True)

print("=" * 80)
print("АНАЛИЗ НЕЗАПОЛНЕННЫХ ДАННЫХ В ЭНЕРГОПАСПОРТЕ")
print("=" * 80)

# Проверка каждого листа
sheets_to_check = {
    "Summary": "Сводная информация",
    "Electricity": "Электроэнергия",
    "Gas": "Газ",
    "Analytics": "Аналитика",
    "02_Исходные данные": "Исходные данные",
    "Equipment": "Оборудование",
    "04_Баланс": "Энергобаланс",
    "05_Динамика": "Динамика",
    "08_Потери_электроэнергии": "Потери",
}

missing_items = []

for sheet_name, description in sheets_to_check.items():
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    print(f"\n📋 ЛИСТ: {sheet_name} ({description})")
    print("-" * 80)
    
    empty_cells = []
    placeholder_cells = []
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            value = cell.value
            
            # Проверка на пустые ячейки
            if value is None or (isinstance(value, str) and value.strip() == ""):
                empty_cells.append((row_idx, col_idx))
            
            # Проверка на placeholder'ы
            elif isinstance(value, str) and ("{{" in value or "}}" in value):
                placeholder_cells.append((row_idx, col_idx, value))
    
    if empty_cells:
        print(f"  ⚠️  Пустые ячейки: {len(empty_cells)}")
        print(f"     Примеры: {empty_cells[:5]}")
        missing_items.append({
            "sheet": sheet_name,
            "type": "empty",
            "count": len(empty_cells),
            "examples": empty_cells[:10]
        })
    
    if placeholder_cells:
        print(f"  ⚠️  Placeholder'ы: {len(placeholder_cells)}")
        for row, col, val in placeholder_cells[:3]:
            print(f"     Row {row}, Col {col}: {val[:50]}")
        missing_items.append({
            "sheet": sheet_name,
            "type": "placeholder",
            "count": len(placeholder_cells),
            "examples": [(r, c, v[:50]) for r, c, v in placeholder_cells[:5]]
        })

# Проверка конкретных проблемных листов
print("\n" + "=" * 80)
print("ДЕТАЛЬНЫЙ АНАЛИЗ ПРОБЛЕМНЫХ ЛИСТОВ")
print("=" * 80)

# 1. Electricity - проверка placeholder'ов
if "Electricity" in wb.sheetnames:
    ws = wb["Electricity"]
    print("\n1. ЛИСТ 'Electricity':")
    print("   Первые 10 строк:")
    for i, row in enumerate(ws.iter_rows(max_row=10), 1):
        values = [str(cell.value)[:30] if cell.value else "EMPTY" for cell in row[:5]]
        print(f"   Row {i}: {values}")

# 2. Summary - что не заполнено
if "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    print("\n2. ЛИСТ 'Summary':")
    print("   Все строки:")
    for i, row in enumerate(ws.iter_rows(), 1):
        values = [str(cell.value) if cell.value else "EMPTY" for cell in row[:3]]
        print(f"   Row {i}: {values}")

# 3. Analytics - что не заполнено
if "Analytics" in wb.sheetnames:
    ws = wb["Analytics"]
    print("\n3. ЛИСТ 'Analytics':")
    print("   Все строки:")
    for i, row in enumerate(ws.iter_rows(), 1):
        values = [str(cell.value) if cell.value else "EMPTY" for cell in row[:3]]
        print(f"   Row {i}: {values}")

# 4. 04_Баланс - проверка формул итогов
if "04_Баланс" in wb.sheetnames:
    ws = wb["04_Баланс"]
    print("\n4. ЛИСТ '04_Баланс':")
    print("   Проверка формул итогов (колонка 6):")
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=15), 1):
        quarter = row[0].value
        total_cell = row[5]  # Колонка "Итого"
        if total_cell.value is None or (isinstance(total_cell.value, str) and "=" not in str(total_cell.value)):
            print(f"   Row {i+2} ({quarter}): ИТОГО = {total_cell.value} (ОТСУТСТВУЕТ ФОРМУЛА)")

# 5. 05_Динамика - проверка производства
if "05_Динамика" in wb.sheetnames:
    ws = wb["05_Динамика"]
    print("\n5. ЛИСТ '05_Динамика':")
    print("   Проверка данных по производству (колонка 6):")
    production_empty = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=14), 1):
        quarter = row[1].value
        production = row[5].value  # Колонка "Производство, кг"
        if production is None or production == 0:
            production_empty.append((i+1, quarter, production))
    if production_empty:
        print(f"   ⚠️  Производство = 0 или пусто в {len(production_empty)} строках:")
        for row, q, val in production_empty[:5]:
            print(f"      Row {row} ({q}): {val}")

# 6. Проверка наличия данных в JSON
print("\n" + "=" * 80)
print("ПРОВЕРКА НАЛИЧИЯ ДАННЫХ В ИСТОЧНИКАХ")
print("=" * 80)

agg_path = Path("data/aggregated/aggregated_full_resources_2022_2024.json")
if agg_path.exists():
    with open(agg_path, "r", encoding="utf-8") as f:
        agg_data = json.load(f)
    
    print("\n📊 Данные в aggregated_full_resources_2022_2024.json:")
    print(f"   Ключи: {list(agg_data.keys())}")
    
    # Проверка production
    production_found = False
    for file_key, file_data in agg_data.items():
        if "resources" in file_data and "production" in file_data["resources"]:
            production_found = True
            prod_data = file_data["resources"]["production"]
            print(f"   ✅ Production найдено в {file_key}: {len(prod_data)} кварталов")
            break
    
    if not production_found:
        print("   ❌ Production НЕ найдено в aggregated JSON")

# Сохранение отчета
report = {
    "missing_items": missing_items,
    "summary": {
        "total_sheets": len(wb.sheetnames),
        "checked_sheets": len(sheets_to_check)
    }
}

report_path = Path("data/aggregated/missing_data_analysis.json")
with open(report_path, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=2)

print(f"\n💾 Отчет сохранен: {report_path}")
print("\n" + "=" * 80)

