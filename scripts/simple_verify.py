from pathlib import Path
from openpyxl import load_workbook

p = Path("test_output_metin/generated_passport_fixed.xlsx")

if not p.exists():
    print("ERROR: File not found")
    exit(1)

print(f"OK: File exists, size: {p.stat().st_size} bytes")

wb = load_workbook(p, data_only=True)
ws = wb["Структура пр 2"]

# Проверяем ключевые ячейки
print("\nKey values:")
print(f"  E32: {ws['E32'].value}")
print(f"  B17: {ws['B17'].value}")
print(f"  C17: {ws['C17'].value}")
print(f"  Q39: {ws['Q39'].value}")
print(f"  R39: {ws['R39'].value}")
print(f"  AF61: {ws['AF61'].value}")
print(f"  AG61: {ws['AG61'].value}")
print(f"  AU83: {ws['AU83'].value}")
print(f"  AV83: {ws['AV83'].value}")

# Подсчет заполненных
filled = 0
total = 0

quarters = [(17, 1), (39, 16), (61, 31), (83, 46)]
for start_row, start_col in quarters:
    for row in range(start_row, start_row + 5):
        for offset in [1, 2, 3, 4]:
            col = start_col + offset
            val = ws.cell(row, col).value
            total += 1
            if val is not None and val != 0:
                filled += 1

print(f"\nFilled: {filled}/{total} ({filled*100/total:.1f}%)")

if filled == total:
    print("SUCCESS: All quarters filled!")
elif filled > 0:
    print("PARTIAL: Some quarters filled")
else:
    print("ERROR: No quarters filled")

wb.close()

