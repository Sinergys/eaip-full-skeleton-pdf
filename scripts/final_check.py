"""Финальная проверка заполнения"""
from pathlib import Path
from openpyxl import load_workbook

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")
report_path = Path("test_output_metin/check_report.txt")

results = []

results.append("=" * 80)
results.append("ПРОВЕРКА ЗАПОЛНЕНИЯ ЭНЕРГОПАСПОРТА")
results.append("=" * 80)

if not output_path.exists():
    results.append(f"❌ Файл не найден: {output_path}")
    results.append("\nСначала запустите генерацию:")
    results.append("python tools/generate_metin_passport_full.py")
else:
    wb = load_workbook(output_path, data_only=True)
    ws = wb["Структура пр 2"]
    
    results.append(f"✅ Файл загружен: {output_path.name}\n")
    
    # E32
    e32 = ws["E32"].value
    results.append(f"E32 (2023 Q1 газ): {e32}")
    if e32 and abs(e32 - 14.819) < 0.001:
        results.append("✅ ОШИБКА ИСПРАВЛЕНА!")
    
    # Проверка кварталов
    quarters = [
        ("Q1", 17, 1, "B-H"),
        ("Q2", 39, 16, "Q-W"),
        ("Q3", 61, 31, "AF-AL"),
        ("Q4", 83, 46, "AU-BA"),
    ]
    
    total_filled = 0
    total_expected = 0
    
    for quarter, start_row, start_col, col_range in quarters:
        results.append(f"\n{quarter} (строка {start_row}, колонки {col_range}):")
        quarter_filled = 0
        
        for row in range(start_row, start_row + 5):
            for offset in [1, 2, 3, 4]:
                col = start_col + offset
                val = ws.cell(row, col).value
                total_expected += 1
                if val is not None and val != 0:
                    total_filled += 1
                    quarter_filled += 1
        
        results.append(f"  Заполнено: {quarter_filled}/20 полей")
    
    results.append(f"\n{'='*80}")
    results.append(f"ИТОГО: {total_filled}/{total_expected} полей заполнено ({total_filled*100/total_expected:.1f}%)")
    
    if total_filled == total_expected:
        results.append("✅ ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ ПОЛНОСТЬЮ!")
    elif total_filled > 0:
        results.append(f"⚠️ Заполнено частично")
    else:
        results.append("❌ КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")
    
    wb.close()

# Сохраняем отчет
with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(results))

# Выводим на экран
print('\n'.join(results))
print(f"\nОтчет также сохранен в: {report_path}")

