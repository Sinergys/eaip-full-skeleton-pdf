"""Генерация паспорта и создание отчета"""
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook

print("=" * 80)
print("ГЕНЕРАЦИЯ ЭНЕРГОПАСПОРТА")
print("=" * 80)

# Запускаем генерацию
print("\n1. Запуск генерации...")
result = subprocess.run(
    [sys.executable, "tools/generate_metin_passport_full.py"],
    capture_output=True,
    text=True,
    encoding='utf-8',
    errors='replace',
    cwd=Path.cwd()
)

if result.stdout:
    print(result.stdout)
if result.stderr:
    print("Ошибки:", result.stderr)

# Проверяем результат
print("\n" + "=" * 80)
print("ПРОВЕРКА РЕЗУЛЬТАТА")
print("=" * 80)

output_path = Path("test_output_metin/generated_passport_fixed.xlsx")

if output_path.exists():
    print(f"\n✅ Файл создан: {output_path}")
    print(f"   Размер: {output_path.stat().st_size:,} байт")
    
    # Проверяем содержимое
    try:
        wb = load_workbook(output_path, data_only=True)
        ws = wb["Структура пр 2"]
        
        print("\n2. Проверка заполнения:")
        
        # E32
        e32 = ws["E32"].value
        print(f"\n   E32 (2023 Q1 газ): {e32}")
        if e32 and abs(e32 - 14.819) < 0.001:
            print("   ✅ ОШИБКА ИСПРАВЛЕНА!")
        
        # Проверка всех кварталов
        quarters = [
            ("Q1", 17, 1, "B-H"),
            ("Q2", 39, 16, "Q-W"),
            ("Q3", 61, 31, "AF-AL"),
            ("Q4", 83, 46, "AU-BA"),
        ]
        
        total_filled = 0
        total_expected = 0
        
        for quarter, start_row, start_col, col_range in quarters:
            quarter_filled = 0
            for row in range(start_row, start_row + 5):
                for offset in [1, 2, 3, 4]:  # норма, 2022, 2023, 2024
                    col = start_col + offset
                    val = ws.cell(row, col).value
                    total_expected += 1
                    if val is not None and val != 0:
                        total_filled += 1
                        quarter_filled += 1
            
            print(f"\n   {quarter} (строка {start_row}, колонки {col_range}):")
            print(f"      Заполнено: {quarter_filled}/20 полей")
            
            # Показываем примеры
            if quarter == "Q1":
                print(f"      Пример - B17 (норма): {ws['B17'].value}, C17 (2022): {ws['C17'].value}")
            elif quarter == "Q2":
                print(f"      Пример - Q39 (норма): {ws['Q39'].value}, R39 (2022): {ws['R39'].value}")
            elif quarter == "Q3":
                print(f"      Пример - AF61 (норма): {ws['AF61'].value}, AG61 (2022): {ws['AG61'].value}")
            elif quarter == "Q4":
                print(f"      Пример - AU83 (норма): {ws['AU83'].value}, AV83 (2022): {ws['AV83'].value}")
        
        print(f"\n   {'='*60}")
        print(f"   ИТОГО: {total_filled}/{total_expected} полей заполнено ({total_filled*100/total_expected:.1f}%)")
        
        if total_filled == total_expected:
            print("   ✅ ВСЕ КВАРТАЛЫ ЗАПОЛНЕНЫ ПОЛНОСТЬЮ!")
        elif total_filled > 0:
            print(f"   ⚠️ Заполнено частично")
        else:
            print("   ❌ КВАРТАЛЫ НЕ ЗАПОЛНЕНЫ!")
        
        wb.close()
        
    except Exception as e:
        print(f"   ❌ Ошибка при проверке: {e}")
        import traceback
        traceback.print_exc()
else:
    print(f"\n❌ Файл не создан: {output_path}")
    print("   Проверьте ошибки выше")

print("\n" + "=" * 80)
print("ГОТОВНОСТЬ ФАЙЛА")
print("=" * 80)

if output_path.exists():
    print(f"✅ Файл готов: {output_path.absolute()}")
    print(f"   Путь: {output_path}")
else:
    print(f"❌ Файл не создан")

print("=" * 80)

