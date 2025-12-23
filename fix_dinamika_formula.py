"""
Исправление формулы в P28 на листе "Динамика ср"
Защита от деления на ноль
"""
from openpyxl import load_workbook
from pathlib import Path
import shutil

template_path = Path(r"C:\Users\DELL\Documents\AUDIT\METIN\Энерг паспорт\ЭНЕРГО_ПАСПОРТ_Серёга_18112025.xlsm")
backup_path = template_path.parent / f"{template_path.stem}_backup{template_path.suffix}"

print(f"📋 ИСПРАВЛЕНИЕ ФОРМУЛЫ В P28")
print("=" * 80)
print(f"📂 Файл: {template_path.name}")

# Создаем резервную копию
if not backup_path.exists():
    shutil.copy2(template_path, backup_path)
    print(f"✅ Создана резервная копия: {backup_path.name}\n")
else:
    print(f"ℹ️ Резервная копия уже существует\n")

# Загружаем файл
wb = load_workbook(template_path, keep_vba=True)  # Сохраняем макросы для .xlsm
ws = wb["Динамика ср"]

print("📄 Лист: 'Динамика ср'\n")

# Проверяем текущую формулу
p28_cell = ws['P28']
current_formula = p28_cell.value if p28_cell.data_type == 'f' else None

print(f"Текущая формула в P28:")
print(f"   {current_formula}\n")

# Исправляем формулу
# Вариант: =ЕСЛИ(P27=0; "-"; ТЕКСТ((P26-P27)/P27*100; "0%"))
# В английской версии Excel: =IF(P27=0, "-", TEXT((P26-P27)/P27*100, "0%"))

# Проверяем, какая версия Excel (русская или английская)
# Попробуем оба варианта
new_formula_ru = '=ЕСЛИ(P27=0; "-"; ТЕКСТ((P26-P27)/P27*100; "0%"))'
new_formula_en = '=IF(P27=0, "-", TEXT((P26-P27)/P27*100, "0%"))'

# Определяем, какая версия используется
if 'ТЕКСТ' in str(current_formula) or 'ЕСЛИ' in str(current_formula):
    new_formula = new_formula_ru
    print("Используется русская версия Excel")
else:
    new_formula = new_formula_en
    print("Используется английская версия Excel")

print(f"\nНовая формула:")
print(f"   {new_formula}\n")

# Применяем исправление
ws['P28'].value = new_formula

# Сохраняем файл
wb.save(template_path)
wb.close()

print("✅ Формула исправлена!")
print(f"✅ Файл сохранен: {template_path.name}")
print(f"\n💡 Резервная копия: {backup_path.name}")

